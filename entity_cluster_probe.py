from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from statistics import median

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TOKEN_RE = re.compile(
    r"<[^>]*>|\{[^{}]+\}|[\u3400-\u9fff\u3040-\u30ff\uac00-\ud7af]|[A-Za-zÀ-ÖØ-öø-ÿ]+(?:[-'][A-Za-zÀ-ÖØ-öø-ÿ]+)*|\d+(?:[./:-]\d+)*|[^\w\s]",
    re.UNICODE,
)
ALPHA_RE = re.compile(r"[A-Za-zÀ-ÖØ-öø-ÿ\u3400-\u9fff\u3040-\u30ff\uac00-\ud7af]")
CJK_RE = re.compile(r"^[\u3400-\u9fff\u3040-\u30ff\uac00-\ud7af]$")
NUMBER_RE = re.compile(r"^\d+(?:[./:-]\d+)*$")
PLACEHOLDER_RE = re.compile(r"^\{[^{}]+\}$")
STOPWORD_RE = re.compile(
    r"^(?:a|an|and|are|as|at|be|by|for|from|in|into|is|of|on|or|the|to|with)$",
    re.IGNORECASE,
)
NO_SPACE_BEFORE = set(".,!?;:%)]}。，、！？；：）」》】")
NO_SPACE_AFTER = set("([{「《【")


@dataclass(frozen=True)
class EntityOccurrence:
    row_number: int
    source: str
    target: str
    entity: str
    entity_token_count: int


@dataclass(frozen=True)
class EntityCluster:
    source_pattern: str
    coverage_count: int
    unique_source_count: int
    unique_entity_count: int
    entity_values: tuple[str, ...]
    confidence: float
    risk: str
    sample_sources: tuple[str, ...]
    sample_targets: tuple[str, ...]
    row_numbers: tuple[int, ...]


def find_entity_clusters(
    rows: list[tuple[str, str]],
    *,
    min_group_size: int = 3,
    max_entity_tokens: int = 4,
    min_literal_tokens: int = 3,
    top: int = 200,
) -> list[EntityCluster]:
    grouped: dict[str, dict[tuple[int, str], EntityOccurrence]] = defaultdict(dict)

    for index, (source, target) in enumerate(rows, start=2):
        tokens = _tokenize(source)
        if len(tokens) < min_literal_tokens + 1:
            continue
        for start in range(len(tokens)):
            max_end = min(len(tokens), start + max_entity_tokens)
            for end in range(start + 1, max_end + 1):
                entity_tokens = tokens[start:end]
                if not _is_entity_candidate(entity_tokens):
                    continue
                literal_tokens = tokens[:start] + tokens[end:]
                if _literal_word_count(literal_tokens) < min_literal_tokens:
                    continue
                pattern = _detokenize(tokens[:start] + ["{entity1}"] + tokens[end:])
                entity = _detokenize(entity_tokens)
                grouped[pattern][(index, entity)] = EntityOccurrence(
                    row_number=index,
                    source=source,
                    target=target,
                    entity=entity,
                    entity_token_count=len(entity_tokens),
                )

    if len(rows) <= 1000 and any(_contains_cjk(source) for source, _ in rows):
        for pattern, occurrences in _pairwise_occurrence_groups(
            rows,
            max_entity_tokens=max_entity_tokens,
            min_literal_tokens=min_literal_tokens,
        ).items():
            grouped[pattern].update(occurrences)

    clusters = [
        _build_cluster(pattern, tuple(occurrences.values()))
        for pattern, occurrences in grouped.items()
        if _has_enough_signal(occurrences.values(), min_group_size)
    ]
    clusters.sort(key=_cluster_sort_key)
    clusters = _remove_fragmented_child_clusters(clusters)
    return _dedupe_clusters(clusters, top)


def generate_entity_cluster_workbook(
    input_path: str | Path,
    output_path: str | Path,
    *,
    source_col: str | int = "en",
    target_col: str | int = "fr",
    min_group_size: int = 3,
    max_entity_tokens: int = 4,
    min_literal_tokens: int = 3,
    top: int = 200,
) -> dict[str, int | str]:
    input_path = Path(input_path)
    output_path = Path(output_path)
    rows = _read_rows(input_path, source_col, target_col)
    clusters = find_entity_clusters(
        rows,
        min_group_size=min_group_size,
        max_entity_tokens=max_entity_tokens,
        min_literal_tokens=min_literal_tokens,
        top=top,
    )
    _write_workbook(output_path, input_path, clusters)
    return {
        "row_count": len(rows),
        "cluster_count": len(clusters),
        "covered_rows": len({row for cluster in clusters for row in cluster.row_numbers}),
        "output_path": str(output_path),
    }


def _tokenize(text: str) -> list[str]:
    return TOKEN_RE.findall(str(text or ""))


def _detokenize(tokens: list[str] | tuple[str, ...]) -> str:
    result = ""
    previous = ""
    for token in tokens:
        if not result:
            result = token
        elif _joins_without_space(previous, token):
            result += token
        else:
            result += " " + token
        previous = token
    return result


def _joins_without_space(previous: str, token: str) -> bool:
    if token in NO_SPACE_BEFORE or previous in NO_SPACE_AFTER:
        return True
    if token.startswith("<") or previous.startswith("<"):
        return True
    if CJK_RE.match(previous) or CJK_RE.match(token):
        return True
    if PLACEHOLDER_RE.match(previous) and CJK_RE.match(token):
        return True
    if CJK_RE.match(previous) and PLACEHOLDER_RE.match(token):
        return True
    return False


def _is_entity_candidate(tokens: list[str]) -> bool:
    if not tokens:
        return False
    if any(PLACEHOLDER_RE.match(token) or NUMBER_RE.match(token) for token in tokens):
        return False
    words = [token for token in tokens if ALPHA_RE.search(token)]
    if not words:
        return False
    if all(STOPWORD_RE.match(word) for word in words):
        return False
    return True


def _literal_word_count(tokens: list[str]) -> int:
    return sum(1 for token in tokens if ALPHA_RE.search(token) or PLACEHOLDER_RE.match(token))


def _contains_cjk(text: str) -> bool:
    return bool(re.search(r"[\u3400-\u9fff\u3040-\u30ff\uac00-\ud7af]", text or ""))


def _pairwise_occurrence_groups(
    rows: list[tuple[str, str]],
    *,
    max_entity_tokens: int,
    min_literal_tokens: int,
) -> dict[str, dict[tuple[int, str], EntityOccurrence]]:
    indexed_rows = [
        (index, source, target)
        for index, (source, target) in enumerate(rows, start=2)
        if source and _contains_cjk(source)
    ]
    grouped: dict[str, dict[tuple[int, str], EntityOccurrence]] = defaultdict(dict)
    for left_index, (left_row, left_source, left_target) in enumerate(indexed_rows):
        for right_row, right_source, right_target in indexed_rows[left_index + 1 :]:
            generalized = _generalize_pair(
                left_source,
                right_source,
                max_entity_tokens=max_entity_tokens,
                min_literal_tokens=min_literal_tokens,
            )
            if generalized is None:
                continue
            pattern, left_entities, right_entities = generalized
            left_entity = " / ".join(left_entities)
            right_entity = " / ".join(right_entities)
            grouped[pattern][(left_row, left_entity)] = EntityOccurrence(
                row_number=left_row,
                source=left_source,
                target=left_target,
                entity=left_entity,
                entity_token_count=sum(len(_tokenize(entity)) for entity in left_entities),
            )
            grouped[pattern][(right_row, right_entity)] = EntityOccurrence(
                row_number=right_row,
                source=right_source,
                target=right_target,
                entity=right_entity,
                entity_token_count=sum(len(_tokenize(entity)) for entity in right_entities),
            )
    return grouped


def _generalize_pair(
    left_source: str,
    right_source: str,
    *,
    max_entity_tokens: int,
    min_literal_tokens: int,
) -> tuple[str, tuple[str, ...], tuple[str, ...]] | None:
    left_tokens = _tokenize(left_source)
    right_tokens = _tokenize(right_source)
    matcher = SequenceMatcher(None, left_tokens, right_tokens, autojunk=False)
    pattern_tokens: list[str] = []
    left_entities: list[str] = []
    right_entities: list[str] = []

    for tag, left_start, left_end, right_start, right_end in matcher.get_opcodes():
        if tag == "equal":
            pattern_tokens.extend(left_tokens[left_start:left_end])
            continue

        left_span = left_tokens[left_start:left_end]
        right_span = right_tokens[right_start:right_end]
        if not left_span or not right_span:
            return None
        if len(left_span) > max_entity_tokens or len(right_span) > max_entity_tokens:
            return None
        if not _is_entity_candidate(left_span) or not _is_entity_candidate(right_span):
            return None
        if len(left_entities) >= 2:
            return None

        left_entities.append(_detokenize(left_span))
        right_entities.append(_detokenize(right_span))
        pattern_tokens.append("{" + f"entity{len(left_entities)}" + "}")

    if not 1 <= len(left_entities) <= 2:
        return None
    if _literal_word_count([token for token in pattern_tokens if not PLACEHOLDER_RE.match(token)]) < min_literal_tokens:
        return None

    return _detokenize(pattern_tokens), tuple(left_entities), tuple(right_entities)


def _has_enough_signal(
    occurrences: object,
    min_group_size: int,
) -> bool:
    occurrence_list = list(occurrences)
    unique_sources = {occurrence.source for occurrence in occurrence_list}
    unique_entities = {occurrence.entity for occurrence in occurrence_list}
    return len(unique_sources) >= min_group_size and len(unique_entities) >= min_group_size


def _build_cluster(
    pattern: str, occurrences: tuple[EntityOccurrence, ...]
) -> EntityCluster:
    sorted_occurrences = tuple(
        sorted(occurrences, key=lambda item: (item.source, item.entity, item.row_number))
    )
    entity_values = tuple(sorted({occurrence.entity for occurrence in sorted_occurrences}))
    source_values = tuple(dict.fromkeys(occurrence.source for occurrence in sorted_occurrences))
    target_values = tuple(dict.fromkeys(occurrence.target for occurrence in sorted_occurrences))
    target_similarity = _average_similarity(target_values[:20])
    literal_ratio = _literal_ratio(pattern, sorted_occurrences)
    length_penalty = max(0.0, (median(occ.entity_token_count for occ in sorted_occurrences) - 1) * 0.05)
    confidence = max(
        0.0,
        min(1.0, (literal_ratio * 0.72) + (target_similarity * 0.23) + 0.08 - length_penalty),
    )
    risks: list[str] = []
    if target_similarity < 0.55:
        risks.append("target side varies")
    if median(occ.entity_token_count for occ in sorted_occurrences) > 2:
        risks.append("long entity span")
    if pattern.startswith("{entity1}") or pattern.endswith("{entity1}"):
        risks.append("edge slot")

    return EntityCluster(
        source_pattern=pattern,
        coverage_count=len(sorted_occurrences),
        unique_source_count=len(source_values),
        unique_entity_count=len(entity_values),
        entity_values=entity_values,
        confidence=round(confidence, 3),
        risk="; ".join(risks),
        sample_sources=source_values[:10],
        sample_targets=target_values[:10],
        row_numbers=tuple(occurrence.row_number for occurrence in sorted_occurrences),
    )


def _literal_ratio(pattern: str, occurrences: tuple[EntityOccurrence, ...]) -> float:
    pattern_tokens = _tokenize(pattern)
    literal_words = _literal_word_count([token for token in pattern_tokens if token != "{entity1}"])
    entity_words = median(occurrence.entity_token_count for occurrence in occurrences)
    return literal_words / max(1.0, literal_words + entity_words)


def _average_similarity(values: tuple[str, ...]) -> float:
    if len(values) < 2:
        return 1.0
    normalized = [_normalize_for_similarity(value) for value in values if value]
    if len(normalized) < 2:
        return 1.0
    pairs = []
    for index, left in enumerate(normalized):
        for right in normalized[index + 1 :]:
            pairs.append(SequenceMatcher(None, left, right).ratio())
            if len(pairs) >= 80:
                return sum(pairs) / len(pairs)
    return sum(pairs) / len(pairs)


def _normalize_for_similarity(text: str) -> str:
    lowered = str(text or "").lower()
    lowered = re.sub(r"\{[^{}]+\}", "{var}", lowered)
    lowered = re.sub(r"\d+(?:[./:-]\d+)*", "{num}", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def _cluster_sort_key(cluster: EntityCluster) -> tuple[object, ...]:
    literal_tokens = _literal_word_count(
        [token for token in _tokenize(cluster.source_pattern) if token != "{entity1}"]
    )
    median_entity_len = median(len(_tokenize(value)) for value in cluster.entity_values)
    compression_score = cluster.coverage_count * cluster.confidence
    return (
        -compression_score,
        -cluster.coverage_count,
        -cluster.confidence,
        -literal_tokens,
        median_entity_len,
        cluster.source_pattern,
    )


def _dedupe_clusters(clusters: list[EntityCluster], top: int) -> list[EntityCluster]:
    selected: list[EntityCluster] = []
    selected_row_sets: list[set[int]] = []
    for cluster in clusters:
        row_set = set(cluster.row_numbers)
        if any(
            _overlaps_existing_cluster(cluster, row_set, existing, existing_rows)
            for existing, existing_rows in zip(selected, selected_row_sets)
        ):
            continue
        selected.append(cluster)
        selected_row_sets.append(row_set)
        if len(selected) >= top:
            break
    return selected


def _overlaps_existing_cluster(
    cluster: EntityCluster,
    row_set: set[int],
    existing: EntityCluster,
    existing_rows: set[int],
) -> bool:
    if _jaccard(row_set, existing_rows) >= 0.85:
        return True
    if row_set < existing_rows:
        return True
    return False


def _remove_fragmented_child_clusters(
    clusters: list[EntityCluster],
) -> list[EntityCluster]:
    kept: list[EntityCluster] = []
    for cluster in clusters:
        if any(
            _absorbs_fragmented_child(broader, cluster)
            for broader in clusters
            if broader is not cluster
        ):
            continue
        kept.append(cluster)
    return kept


def _absorbs_fragmented_child(
    broader: EntityCluster, narrower: EntityCluster
) -> bool:
    if broader.coverage_count <= narrower.coverage_count:
        return False
    narrower_rows = set(narrower.row_numbers)
    broader_rows = set(broader.row_numbers)
    if not narrower_rows:
        return False
    if len(narrower_rows & broader_rows) / len(narrower_rows) < 0.8:
        return False
    return _pattern_absorbs_fragmented_child(
        broader.source_pattern,
        narrower.source_pattern,
    )


def _pattern_absorbs_fragmented_child(
    broader_pattern: str, narrower_pattern: str
) -> bool:
    broader_tokens = _tokenize(broader_pattern)
    narrower_tokens = _tokenize(narrower_pattern)
    if len(broader_tokens) >= len(narrower_tokens):
        return False

    deleted: set[int] = set()
    broader_index = 0
    for narrow_index, token in enumerate(narrower_tokens):
        if (
            broader_index < len(broader_tokens)
            and token == broader_tokens[broader_index]
        ):
            broader_index += 1
        else:
            deleted.add(narrow_index)

    if broader_index != len(broader_tokens) or not deleted:
        return False

    sorted_deleted = sorted(deleted)
    block_start = sorted_deleted[0]
    previous = sorted_deleted[0]
    for index in sorted_deleted[1:] + [None]:
        if index is not None and index == previous + 1:
            previous = index
            continue
        if not _is_deletable_entity_boundary_fragment(
            narrower_tokens,
            block_start,
            previous,
        ):
            return False
        if index is not None:
            block_start = previous = index
    return True


def _is_deletable_entity_boundary_fragment(
    tokens: list[str], start: int, end: int
) -> bool:
    if any(not CJK_RE.match(token) for token in tokens[start : end + 1]):
        return False
    left = tokens[start - 1] if start > 0 else ""
    right = tokens[end + 1] if end + 1 < len(tokens) else ""
    return _is_entity_placeholder(left) or _is_entity_placeholder(right)


def _is_entity_placeholder(token: str) -> bool:
    return bool(re.fullmatch(r"\{entity\d+\}", token))


def _jaccard(left: set[int], right: set[int]) -> float:
    if not left or not right:
        return 0.0
    return len(left & right) / len(left | right)


def _read_rows(
    input_path: Path, source_col: str | int, target_col: str | int
) -> list[tuple[str, str]]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    source_index = _resolve_column(ws, source_col)
    target_index = _resolve_column(ws, target_col)
    rows: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        source = _cell_value(row, source_index)
        target = _cell_value(row, target_index)
        if source is None or str(source).strip() == "":
            continue
        rows.append((str(source).strip(), "" if target is None else str(target).strip()))
    return rows


def _resolve_column(ws, col: str | int) -> int:
    if isinstance(col, int):
        return col
    if str(col).isdigit():
        return int(str(col))
    wanted = str(col).strip()
    wanted_lower = wanted.lower()
    for cell in ws[1]:
        value = "" if cell.value is None else str(cell.value).strip()
        if value == wanted or value.lower() == wanted_lower:
            return cell.column
    raise ValueError(f"Column {col!r} not found in header row")


def _cell_value(row: tuple[object, ...], one_based_index: int) -> object | None:
    zero_based = one_based_index - 1
    return row[zero_based] if zero_based < len(row) else None


def _write_workbook(
    output_path: Path, input_path: Path, clusters: list[EntityCluster]
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    summary.append(["metric", "value"])
    summary.append(["input_file", str(input_path)])
    summary.append(["entity_cluster_count", len(clusters)])
    summary.append(["covered_rows", len({row for cluster in clusters for row in cluster.row_numbers})])
    summary.append(["review_note", "Experimental one-entity-slot clusters. Review manually before TM use."])

    ws = wb.create_sheet("entity_clusters")
    ws.append(
        [
            "cluster_id",
            "source_pattern",
            "coverage_count",
            "unique_source_count",
            "unique_entity_count",
            "confidence",
            "risk",
            "entity_values",
            "sample_sources",
            "sample_targets",
            "row_numbers",
        ]
    )
    for index, cluster in enumerate(clusters, start=1):
        ws.append(
            [
                f"E{index:04d}",
                cluster.source_pattern,
                cluster.coverage_count,
                cluster.unique_source_count,
                cluster.unique_entity_count,
                cluster.confidence,
                cluster.risk or None,
                "\n".join(cluster.entity_values[:30]),
                "\n".join(cluster.sample_sources),
                "\n".join(cluster.sample_targets),
                ",".join(str(row) for row in cluster.row_numbers),
            ]
        )

    for sheet in wb.worksheets:
        _style_sheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max(
            10,
            min(
                70,
                max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                + 2,
            ),
        )
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Probe one-entity-slot reusable sentence structures in completed TM."
    )
    parser.add_argument("input", help="Completed TM workbook")
    parser.add_argument("-o", "--output", help="Review workbook output path")
    parser.add_argument("--source-col", default="en")
    parser.add_argument("--target-col", default="fr")
    parser.add_argument("--min-group-size", type=int, default=3)
    parser.add_argument("--max-entity-tokens", type=int, default=4)
    parser.add_argument("--min-literal-tokens", type=int, default=3)
    parser.add_argument("--top", type=int, default=200)
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    output_path = (
        Path(args.output)
        if args.output
        else input_path.with_name(f"{input_path.stem}_entity_clusters.xlsx")
    )
    stats = generate_entity_cluster_workbook(
        input_path,
        output_path,
        source_col=args.source_col,
        target_col=args.target_col,
        min_group_size=args.min_group_size,
        max_entity_tokens=args.max_entity_tokens,
        min_literal_tokens=args.min_literal_tokens,
        top=args.top,
    )
    print(f"Rows read: {stats['row_count']}")
    print(f"Entity clusters: {stats['cluster_count']}")
    print(f"Covered rows: {stats['covered_rows']}")
    print(f"Output: {stats['output_path']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
