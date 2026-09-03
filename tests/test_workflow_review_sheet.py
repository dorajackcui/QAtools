from __future__ import annotations

import unittest

from openpyxl import Workbook

from tools.excel_output import PROBLEM_BASE_HEADERS
from tools.workflow.review_sheet import collect_review_rows


class ReviewSheetOrderingTests(unittest.TestCase):
    def test_consistency_groups_come_first_without_splitting_multiple_issues(self) -> None:
        workbook = Workbook()
        try:
            tag_sheet = workbook.active
            tag_sheet.title = "Tag 问题"
            tag_sheet.append(PROBLEM_BASE_HEADERS)
            tag_sheet.append([5, "Other", "Other target", "Tag 缺失"])
            tag_sheet.append([20, "Save", "Save A", "Tag 缺失"])
            tag_sheet.append([2, "Plain", "Plain target", "Tag 缺失"])

            source_sheet = workbook.create_sheet("Source 问题")
            source_sheet.append(PROBLEM_BASE_HEADERS + ("同组行号",))
            for row, source, target, grouped_rows in (
                (90, "Save", "Save B", "20、50、80、90"),
                (60, "Open", "Open B", "40、60"),
                (20, "Save", "Save A", "20、50、80、90"),
                (40, "Open", "Open A", "40、60"),
                (50, "Save", "Save B", "20、50、80、90"),
                (80, "Save", "Save A", "20、50、80、90"),
            ):
                source_sheet.append([row, source, target, "译文不一致", grouped_rows])

            target_sheet = workbook.create_sheet("Target 问题")
            target_sheet.append(PROBLEM_BASE_HEADERS + ("同组行号",))
            for row, source, target, grouped_rows in (
                (100, "Extra", "Save A", "20、80、100"),
                (80, "Save", "Save A", "20、80、100"),
                (70, "Zulu", "Shared Z", "10、70"),
                (20, "Save", "Save A", "20、80、100"),
                (30, "Beta", "Shared A", "30、110"),
                (10, "Alpha", "Shared Z", "10、70"),
                (110, "Gamma", "Shared A", "30、110"),
            ):
                target_sheet.append([row, source, target, "原文不一致", grouped_rows])

            checks = (
                ("Tag 检查", tag_sheet.title),
                ("同 Target 不同 Source", target_sheet.title),
                ("同 Source 不同 Target", source_sheet.title),
            )
            rows = collect_review_rows(workbook, checks)
            self.assertEqual(
                [row[0] for row in rows],
                [20, 80, 50, 90, 40, 60, 10, 70, 30, 110, 100, 2, 5],
            )
            self.assertEqual(len({row[0] for row in rows}), len(rows))
            self.assertEqual(
                rows[0],
                (
                    20,
                    "Save",
                    "Save A",
                    None,
                    "【Tag 检查】Tag 缺失；"
                    "【同 Target 不同 Source】原文不一致（同组行号：20、80、100）；"
                    "【同 Source 不同 Target】译文不一致（同组行号：20、50、80、90）",
                    "Tag 检查；同 Target 不同 Source；同 Source 不同 Target",
                ),
            )
            self.assertEqual(
                [row[:4] for row in collect_review_rows(workbook, reversed(checks))],
                [row[:4] for row in rows],
            )
        finally:
            workbook.close()

    def test_ordering_keeps_exact_source_groups_and_blank_target_variants(self) -> None:
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.append(PROBLEM_BASE_HEADERS)
            for row, source, target in (
                (12, "Same ", "Y"),
                (4, "Same", None),
                (10, "Same", "X"),
                (6, "Same ", "X"),
                (14, "Same", None),
                (8, "same", "X"),
                (16, "same", "Y"),
            ):
                sheet.append([row, source, target, "译文不一致"])

            rows = collect_review_rows(
                workbook, (("同 Source 不同 Target", sheet.title),)
            )
            self.assertEqual([row[0] for row in rows], [4, 14, 10, 6, 12, 8, 16])
            self.assertEqual([row[2] for row in rows[:3]], ["", "", "X"])
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
