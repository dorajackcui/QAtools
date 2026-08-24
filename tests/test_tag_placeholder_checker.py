from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from tools.tag_placeholder_checker.check_tags_and_placeholders import (
    extract_tokens,
    load_angle_patterns_from_file,
    process_excel,
)


class ExtractTokensTests(unittest.TestCase):
    def test_extract_tokens_supports_mixed_token_types_in_text_order(self) -> None:
        text = r"前缀</text>{name}\n中间<color=red>后缀"
        self.assertEqual(
            extract_tokens(text, token_types=("angle", "brace", "newline")),
            ["</text>", "{name}", r"\n", "<color=red>"],
        )

    def test_extract_tokens_requires_at_least_one_type(self) -> None:
        with self.assertRaisesRegex(ValueError, "请至少选择一种检查类型"):
            extract_tokens("任意文本", token_types=())

    def test_extract_tokens_treats_all_angle_brackets_as_tags_by_default(self) -> None:
        text = (
            "保留 <apple>、</text>、<br/>、<i>、"
            "<img src='itemsmall_%s'/>、<size={c}>、<a href='https://example.com'> 和 <color=red>"
        )
        self.assertEqual(
            extract_tokens(text, token_types=("angle",)),
            [
                "<apple>",
                "</text>",
                "<br/>",
                "<i>",
                "<img src='itemsmall_%s'/>",
                "<size={c}>",
                "<a href='https://example.com'>",
                "<color=red>",
            ],
        )

    def test_extract_tokens_treats_memoq_tags_as_dedicated_tokens(self) -> None:
        text = "{1}{2>Glace du Néant<3} 和 {name}"

        self.assertEqual(
            extract_tokens(text, token_types=("memoq", "brace")),
            ["{1}", "{2>", "<3}", "{name}"],
        )
        self.assertEqual(extract_tokens(text, token_types=("brace",)), ["{name}"])

    def test_extract_tokens_supports_square_color_tags(self) -> None:
        text = "保留 [color=red]、[color = #fff] 和 [/color]，忽略 [stage1]"

        self.assertEqual(
            extract_tokens(text, token_types=("square_color",)),
            ["[color=red]", "[color = #fff]", "[/color]"],
        )

    def test_extract_tokens_ignores_spaced_comparison_expressions(self) -> None:
        self.assertEqual(
            extract_tokens("Value < 10 and count > 0", token_types=("angle",)),
            [],
        )

    def test_extract_tokens_respects_quotes_inside_angle_tags(self) -> None:
        self.assertEqual(
            extract_tokens(
                '<a title="1 > 0">text</a>',
                token_types=("angle",),
            ),
            ['<a title="1 > 0">', "</a>"],
        )

    def test_extract_tokens_preserves_double_brace_placeholders(self) -> None:
        self.assertEqual(
            extract_tokens("{{name}} and {count}", token_types=("brace",)),
            ["{{name}}", "{count}"],
        )
        self.assertEqual(
            extract_tokens("{{1}}", token_types=("brace", "memoq")),
            ["{{1}}"],
        )

    def test_angle_config_requires_a_json_object(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            config_path = Path(tmp_dir) / "angle-tags.json"
            config_path.write_text('["tag"]', encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "需要 JSON 对象"):
                load_angle_patterns_from_file(config_path)

    def test_empty_angle_config_uses_default_angle_matching(self) -> None:
        self.assertEqual(
            extract_tokens(
                "保留 <b> 和 </b>",
                token_types=("angle",),
                angle_config_file="",
            ),
            ["<b>", "</b>"],
        )


class ProcessExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["A2"] = "保留 <color=red>{name}</text>"
        worksheet["B2"] = "保留 <color=red>{name}</text>"
        worksheet["A3"] = "缺少占位 {name}"
        worksheet["B3"] = "缺少占位"
        worksheet["A4"] = "多出tag <color=red>"
        worksheet["B4"] = "多出tag <color=red></text>"
        worksheet["A5"] = "数量不一致 {count} {count}"
        worksheet["B5"] = "数量不一致 {count}"
        worksheet["A6"] = r"缺少换行 mark \n"
        worksheet["B6"] = "缺少换行 mark"
        worksheet["A7"] = "普通尖括号 <apple>"
        worksheet["B7"] = "普通尖括号"
        worksheet["Z1000"] = "unrelated tail"
        worksheet["A1001"].number_format = "@"
        workbook.save(path)

    def test_process_excel_writes_problem_and_summary_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            expected_output_path = Path(tmp_dir) / "tag_check_input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                token_types=("angle", "brace", "newline", "memoq"),
            )

            self.assertEqual(summary.worksheet_title, "Data")
            self.assertEqual(summary.output_path, expected_output_path.resolve())
            self.assertEqual(summary.total_rows_checked, 6)
            self.assertEqual(summary.rows_with_selected_tokens, 6)
            self.assertEqual(summary.angle_rows, 3)
            self.assertEqual(summary.square_color_rows, 0)
            self.assertEqual(summary.brace_rows, 3)
            self.assertEqual(summary.newline_rows, 1)
            self.assertEqual(summary.memoq_rows, 0)
            self.assertEqual(summary.problem_rows, 5)
            self.assertEqual(summary.problem_count, 5)
            self.assertEqual(summary.selected_token_types, ("angle", "brace", "newline", "memoq"))

            original_workbook = load_workbook(input_path)
            self.assertEqual(original_workbook.sheetnames, ["Data"])
            self.assertEqual(original_workbook["Data"]["A2"].value, "保留 <color=red>{name}</text>")

            workbook = load_workbook(summary.output_path)
            self.assertIn("标签占位问题", workbook.sheetnames)
            self.assertIn("检查汇总", workbook.sheetnames)

            problem_sheet = workbook["标签占位问题"]
            self.assertEqual(
                [problem_sheet.cell(1, column).value for column in range(1, 6)],
                ["行号", "source原文", "target原文", "问题描述", "问题类型"],
            )
            self.assertEqual(problem_sheet["A2"].value, 3)
            self.assertEqual(problem_sheet["B2"].value, "缺少占位 {name}")
            self.assertEqual(problem_sheet["C2"].value, "缺少占位")
            self.assertIn("target缺少={name}", str(problem_sheet["D2"].value))
            self.assertEqual(problem_sheet["E2"].value, "花括号placeholder不一致")

            self.assertEqual(problem_sheet["A3"].value, 4)
            self.assertEqual(problem_sheet["E3"].value, "尖括号tag不一致")
            self.assertIn("target多出=</text>", str(problem_sheet["D3"].value))

            self.assertEqual(problem_sheet["A4"].value, 5)
            self.assertEqual(problem_sheet["E4"].value, "花括号placeholder不一致")
            self.assertIn("target缺少={count}", str(problem_sheet["D4"].value))

            self.assertEqual(problem_sheet["A5"].value, 6)
            self.assertEqual(problem_sheet["E5"].value, r"\n mark不一致")
            self.assertIn(r"target缺少=\n", str(problem_sheet["D5"].value))

            self.assertEqual(problem_sheet["A6"].value, 7)
            self.assertEqual(problem_sheet["E6"].value, "尖括号tag不一致")
            self.assertIn("target缺少=<apple>", str(problem_sheet["D6"].value))
            self.assertEqual(problem_sheet["A2"].hyperlink.location, "'Data'!B3")
            self.assertEqual(problem_sheet["A6"].hyperlink.location, "'Data'!B7")
            self.assertIsNone(problem_sheet["A2"].hyperlink.target)
            self.assertEqual(problem_sheet.freeze_panes, "A2")

            summary_sheet = workbook["检查汇总"]
            summary_values = {
                summary_sheet.cell(row_index, 1).value: summary_sheet.cell(row_index, 2).value
                for row_index in range(2, 16)
            }
            self.assertEqual(summary_values["检查工作表"], "Data")
            self.assertEqual(summary_values["source列"], "A")
            self.assertEqual(summary_values["target列"], "B")
            self.assertEqual(summary_values["开始行"], 2)
            self.assertEqual(summary_values["检查类型"], r"尖括号tag、花括号placeholder、\n mark、memoQ tag")
            self.assertEqual(summary_values["总行数"], 6)
            self.assertEqual(summary_values["命中检查类型行数"], 6)
            self.assertEqual(summary_values["含尖括号tag行数"], 3)
            self.assertEqual(summary_values["含方括号color tag行数"], 0)
            self.assertEqual(summary_values["含花括号placeholder行数"], 3)
            self.assertEqual(summary_values[r"含\n mark行数"], 1)
            self.assertEqual(summary_values["含memoQ tag行数"], 0)
            self.assertEqual(summary_values["问题行数"], 5)
            self.assertEqual(summary_values["问题条数"], 5)

    def test_process_excel_supports_single_selected_type(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                token_types=("angle",),
            )

            self.assertEqual(summary.rows_with_selected_tokens, 3)
            self.assertEqual(summary.angle_rows, 3)
            self.assertEqual(summary.square_color_rows, 0)
            self.assertEqual(summary.brace_rows, 0)
            self.assertEqual(summary.newline_rows, 0)
            self.assertEqual(summary.memoq_rows, 0)
            self.assertEqual(summary.problem_rows, 2)
            self.assertEqual(summary.problem_count, 2)

            workbook = load_workbook(summary.output_path)
            problem_sheet = workbook["标签占位问题"]
            self.assertEqual(problem_sheet.max_row, 3)
            self.assertEqual(problem_sheet["E2"].value, "尖括号tag不一致")

    def test_process_excel_supports_newline_selected_type(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                token_types=("newline",),
            )

            self.assertEqual(summary.rows_with_selected_tokens, 1)
            self.assertEqual(summary.angle_rows, 0)
            self.assertEqual(summary.square_color_rows, 0)
            self.assertEqual(summary.brace_rows, 0)
            self.assertEqual(summary.newline_rows, 1)
            self.assertEqual(summary.memoq_rows, 0)
            self.assertEqual(summary.problem_rows, 1)
            self.assertEqual(summary.problem_count, 1)

            workbook = load_workbook(summary.output_path)
            problem_sheet = workbook["标签占位问题"]
            self.assertEqual(problem_sheet.max_row, 2)
            self.assertEqual(problem_sheet["E2"].value, r"\n mark不一致")

    def test_process_excel_reports_memoq_tag_mismatches_separately(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "{1}{2>Glace du Néant<3}"
            worksheet["B2"] = "{1}{2>Glace du Néant<4}"
            workbook.save(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                token_types=("memoq", "brace"),
            )

            self.assertEqual(summary.rows_with_selected_tokens, 1)
            self.assertEqual(summary.brace_rows, 0)
            self.assertEqual(summary.square_color_rows, 0)
            self.assertEqual(summary.memoq_rows, 1)
            self.assertEqual(summary.problem_rows, 1)
            self.assertEqual(summary.problem_count, 1)

            workbook = load_workbook(summary.output_path)
            problem_sheet = workbook["标签占位问题"]
            self.assertEqual(problem_sheet["E2"].value, "memoQ tag不一致")
            self.assertIn("target缺少=<3}", str(problem_sheet["D2"].value))
            self.assertIn("target多出=<4}", str(problem_sheet["D2"].value))

    def test_process_excel_reports_repeated_numeric_placeholder_count_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet.append(["source", "target"])
            worksheet.append(
                [
                    "<link=8>{1}</link> 和 <link=9>{1}</link>",
                    "<link=8>{1}</link> et <link=9></link>",
                ]
            )
            worksheet.append(
                [
                    "<link=8>{1}</link> 和 <link=9>{1}</link>",
                    "<link=8>{1}</link>, <link=9>{1}</link> et {1}",
                ]
            )
            workbook.save(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                token_types=("angle", "memoq"),
            )

            self.assertEqual(summary.problem_rows, 2)
            self.assertEqual(summary.problem_count, 2)
            problem_sheet = load_workbook(summary.output_path)["标签占位问题"]
            self.assertEqual(problem_sheet["E2"].value, "memoQ tag不一致")
            self.assertIn("source={1} x2", problem_sheet["D2"].value)
            self.assertIn("target={1}", problem_sheet["D2"].value)
            self.assertIn("target缺少={1}", problem_sheet["D2"].value)
            self.assertEqual(problem_sheet["E3"].value, "memoQ tag不一致")
            self.assertIn("source={1} x2", problem_sheet["D3"].value)
            self.assertIn("target={1} x3", problem_sheet["D3"].value)
            self.assertIn("target多出={1}", problem_sheet["D3"].value)

    def test_process_excel_reports_square_color_tag_mismatches(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "颜色 [color=red]文本[/color]"
            worksheet["B2"] = "颜色 [color=red]文本"
            workbook.save(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                token_types=("square_color",),
            )

            self.assertEqual(summary.rows_with_selected_tokens, 1)
            self.assertEqual(summary.square_color_rows, 1)
            self.assertEqual(summary.problem_rows, 1)
            self.assertEqual(summary.problem_count, 1)

            workbook = load_workbook(summary.output_path)
            problem_sheet = workbook["标签占位问题"]
            self.assertEqual(problem_sheet["E2"].value, "方括号color tag不一致")
            self.assertIn("target缺少=[/color]", str(problem_sheet["D2"].value))

    def test_process_excel_ignores_comparisons_and_checks_full_quoted_tags(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet.append(["source", "target"])
            worksheet.append(
                ["Value < 10 and count > 0", "Valeur < 10 et nombre > 0"]
            )
            worksheet.append(
                [
                    '<a title="1 > 0">text</a>',
                    '<a title="1 > 999">texte</a>',
                ]
            )
            workbook.save(input_path)

            summary = process_excel(
                input_path,
                "A",
                "B",
                token_types=("angle",),
            )

            self.assertEqual(summary.problem_rows, 1)
            self.assertEqual(summary.problem_count, 1)
            problem_sheet = load_workbook(summary.output_path)["标签占位问题"]
            self.assertEqual(problem_sheet["A2"].value, 3)
            self.assertIn('<a title="1 > 0">', problem_sheet["D2"].value)
            self.assertIn('<a title="1 > 999">', problem_sheet["D2"].value)

    def test_process_excel_reports_changed_angle_tag_hierarchy(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet.append(["source", "target"])
            worksheet.append(
                ["<b>A</b><i>B</i>", "<b>A<i>B</i></b>"]
            )
            worksheet.append(
                ["<b>Bold</b><i>Italic</i>", "<i>斜体</i><b>粗体</b>"]
            )
            workbook.save(input_path)

            summary = process_excel(
                input_path,
                "A",
                "B",
                token_types=("angle",),
            )

            self.assertEqual(summary.problem_rows, 1)
            self.assertEqual(summary.problem_count, 1)
            problem_sheet = load_workbook(summary.output_path)["标签占位问题"]
            self.assertEqual(problem_sheet["A2"].value, 2)
            self.assertEqual(problem_sheet["E2"].value, "尖括号tag结构不一致")
            self.assertIn("嵌套或闭合结构不同", problem_sheet["D2"].value)

    def test_process_excel_reports_single_vs_double_brace_placeholder(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet.append(["source", "target"])
            worksheet.append(["Hello {{name}}", "Bonjour {name}"])
            workbook.save(input_path)

            summary = process_excel(
                input_path,
                "A",
                "B",
                token_types=("brace",),
            )

            self.assertEqual(summary.problem_rows, 1)
            self.assertEqual(summary.problem_count, 1)
            problem_sheet = load_workbook(summary.output_path)["标签占位问题"]
            self.assertIn("target缺少={{name}}", problem_sheet["D2"].value)
            self.assertIn("target多出={name}", problem_sheet["D2"].value)

    def test_problem_sheet_merges_multiple_issue_types_from_the_same_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "<b>{name}"
            worksheet["B2"] = ""
            workbook.save(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                token_types=("angle", "brace"),
            )

            self.assertEqual(summary.problem_rows, 1)
            self.assertEqual(summary.problem_count, 2)
            problem_sheet = load_workbook(summary.output_path)["标签占位问题"]
            self.assertEqual(problem_sheet.max_row, 2)
            self.assertIn("尖括号tag不一致", problem_sheet["E2"].value)
            self.assertIn("花括号placeholder不一致", problem_sheet["E2"].value)
            self.assertIn("<b>", problem_sheet["D2"].value)
            self.assertIn("{name}", problem_sheet["D2"].value)

    def test_process_excel_rejects_same_source_and_target_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            with self.assertRaisesRegex(ValueError, "不能相同"):
                process_excel(input_path, "A", "a", token_types=("angle",))

    def test_process_excel_compiles_angle_filters_once_per_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            with patch(
                "tools.tag_placeholder_checker.check_tags_and_placeholders.compile_angle_patterns",
                return_value=(),
            ) as compile_mock:
                process_excel(input_path, "A", "B", token_types=("angle",))

            compile_mock.assert_called_once_with(None, None)

    def test_process_excel_rebuilds_reserved_output_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet.append(["source", "target"])
            workbook.create_sheet("标签占位问题")["A1"] = "旧问题"
            workbook.create_sheet("检查汇总")["A1"] = "旧汇总"
            workbook.save(input_path)

            summary = process_excel(
                input_path,
                "A",
                "B",
                token_types=("angle",),
            )

            output_workbook = load_workbook(summary.output_path)
            self.assertEqual(output_workbook["标签占位问题"]["A1"].value, "行号")
            self.assertEqual(output_workbook["检查汇总"]["A1"].value, "统计项")


if __name__ == "__main__":
    unittest.main()
