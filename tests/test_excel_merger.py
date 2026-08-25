from __future__ import annotations

from collections import OrderedDict
from datetime import datetime
import io
from pathlib import Path
from types import SimpleNamespace
import tempfile
import unittest
from unittest.mock import Mock, patch
from xml.etree import ElementTree as ET
import zipfile

from openpyxl import Workbook, load_workbook

from tools.excel_merger.merge_active_sheets import (
    EXCEL_MAX_COLUMNS,
    EXCEL_MAX_ROWS,
    build_default_output_path,
    get_cell_value,
    merge_active_sheets,
    read_shared_strings,
    write_output_xlsx,
)
from tools.excel_merger.merge_active_sheets_gui import MergeActiveSheetsApp


class ExcelMergerTests(unittest.TestCase):
    def write_workbook(
        self,
        path: Path,
        rows: list[list[object]],
        *,
        active_sheet_title: str = "Active",
    ) -> None:
        workbook = Workbook()
        ignored_sheet = workbook.active
        ignored_sheet.title = "Ignored"
        ignored_sheet.append(["must", "not", "appear"])
        active_sheet = workbook.create_sheet(active_sheet_title)
        for row in rows:
            active_sheet.append(row)
        workbook.active = workbook.sheetnames.index(active_sheet_title)
        workbook.save(path)
        workbook.close()

    def test_merges_active_sheets_recursively_and_keeps_only_first_header(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_folder = root / "input"
            nested_folder = input_folder / "nested"
            nested_folder.mkdir(parents=True)
            self.write_workbook(
                input_folder / "a.xlsx",
                [["Key", "Value"], ["alpha", 1]],
            )
            self.write_workbook(
                nested_folder / "b.xlsm",
                [["Key", "Value"], ["beta", 2]],
                active_sheet_title="Current",
            )
            (input_folder / "legacy.xls").write_bytes(b"legacy")
            (input_folder / "~$temporary.xlsx").write_bytes(b"temporary")
            output_path = root / "merged.xlsx"
            progress: list[tuple[int, int]] = []

            summary = merge_active_sheets(
                input_folder,
                output_path=output_path,
                progress_callback=lambda completed, total: progress.append(
                    (completed, total)
                ),
            )

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                rows = list(workbook["MergedData"].values)
            finally:
                workbook.close()

            self.assertEqual(
                rows,
                [
                    ("SourceFile", "Key", "Value"),
                    ("a.xlsx", "alpha", "1"),
                    ("b.xlsm", "beta", "2"),
                ],
            )
            self.assertNotIn("must", str(rows))
            self.assertEqual(summary.supported_file_count, 2)
            self.assertEqual(summary.skipped_file_count, 1)
            self.assertEqual(summary.failed_file_count, 0)
            self.assertEqual(summary.merged_row_count, 3)
            self.assertIsNone(summary.error_log_path)
            self.assertEqual(progress, [(1, 2), (2, 2)])

    def test_keep_all_headers_preserves_each_first_row(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_folder = root / "input"
            input_folder.mkdir()
            self.write_workbook(
                input_folder / "a.xlsx",
                [["Key"], ["alpha"]],
            )
            self.write_workbook(
                input_folder / "b.xlsx",
                [["Key"], ["beta"]],
            )
            output_path = root / "merged.xlsx"

            merge_active_sheets(
                input_folder,
                output_path=output_path,
                keep_all_headers=True,
            )

            workbook = load_workbook(output_path, read_only=True)
            try:
                rows = list(workbook["MergedData"].values)
            finally:
                workbook.close()
            self.assertEqual(
                rows,
                [
                    ("SourceFile", "Key"),
                    ("a.xlsx", "alpha"),
                    ("SourceFile", "Key"),
                    ("b.xlsx", "beta"),
                ],
            )

    def test_unreadable_workbook_is_logged_while_valid_files_are_merged(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_folder = root / "input"
            input_folder.mkdir()
            self.write_workbook(
                input_folder / "good.xlsx",
                [["Key"], ["A&B <value>"]],
            )
            (input_folder / "broken.xlsx").write_bytes(b"not a workbook")

            summary = merge_active_sheets(
                input_folder,
                output_path=root / "merged.xlsx",
            )

            self.assertEqual(summary.supported_file_count, 2)
            self.assertEqual(summary.failed_file_count, 1)
            self.assertIsNotNone(summary.error_log_path)
            assert summary.error_log_path is not None
            error_log = summary.error_log_path.read_text(encoding="utf-8")
            self.assertIn("broken.xlsx", error_log)

            workbook = load_workbook(summary.output_path, read_only=True)
            try:
                rows = list(workbook["MergedData"].values)
            finally:
                workbook.close()
            self.assertEqual(rows[-1], ("good.xlsx", "A&B <value>"))

    def test_default_output_path_matches_source_tool_naming(self) -> None:
        folder = Path("workbooks")

        output_path = build_default_output_path(
            folder,
            now=datetime(2026, 8, 25, 12, 34, 56),
        )

        self.assertEqual(
            output_path.name,
            "workbooks_merged_active_sheet_20260825_123456.xlsx",
        )
        self.assertEqual(output_path.parent, folder.resolve().parent)

    def test_shared_and_inline_strings_preserve_text_node_spaces(self) -> None:
        workbook = io.BytesIO()
        with zipfile.ZipFile(workbook, "w") as zip_file:
            zip_file.writestr(
                "xl/sharedStrings.xml",
                """<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <si><t>Key</t></si>
  <si><t xml:space="preserve"> Key </t></si>
  <si><r><t>Hello</t></r><r><t>World</t></r></si>
</sst>""",
            )

        workbook.seek(0)
        with zipfile.ZipFile(workbook, "r") as zip_file:
            values = read_shared_strings(zip_file)
        self.assertEqual(values, ["Key", " Key ", "HelloWorld"])

        cell = ET.fromstring(
            """<c xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" t="inlineStr">
  <is><t xml:space="preserve"> Value </t></is>
</c>"""
        )
        self.assertEqual(get_cell_value(cell, []), " Value ")

    def test_rejects_a_folder_without_supported_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)
            (folder / "legacy.xls").write_bytes(b"legacy")

            with self.assertRaisesRegex(
                ValueError,
                r"没有支持的 \.xlsx/\.xlsm 文件",
            ):
                merge_active_sheets(folder)

    def test_gui_starts_merge_with_the_previewed_output_path(self) -> None:
        app = MergeActiveSheetsApp.__new__(MergeActiveSheetsApp)
        app.input_folder_var = SimpleNamespace(get=lambda: r"D:\input")
        app.keep_all_headers_var = SimpleNamespace(get=lambda: True)
        app.output_preview_var = SimpleNamespace(set=Mock())
        app._start_merge_worker = Mock()
        expected_output = Path(r"D:\input_merged_active_sheet.xlsx")

        with patch(
            "tools.excel_merger.merge_active_sheets_gui.build_default_output_path",
            return_value=expected_output,
        ):
            app.run_merge()

        app.output_preview_var.set.assert_called_once_with(
            f"输出文件：{expected_output}"
        )
        app._start_merge_worker.assert_called_once_with(
            {
                "folder_path": r"D:\input",
                "output_path": expected_output,
                "keep_all_headers": True,
            }
        )

    def test_rejects_output_path_that_would_overwrite_an_input(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_folder = Path(temp_dir)
            source_path = input_folder / "source.xlsx"
            self.write_workbook(source_path, [["Key"], ["original"]])
            original_bytes = source_path.read_bytes()

            with self.assertRaisesRegex(
                ValueError,
                "输出文件不能覆盖待合并的源工作簿",
            ):
                merge_active_sheets(
                    input_folder,
                    output_path=source_path,
                )

            self.assertEqual(source_path.read_bytes(), original_bytes)

    def test_rejects_output_wider_than_the_excel_column_limit(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "too_wide.xlsx"
            rows = [[""] * (EXCEL_MAX_COLUMNS + 1)]

            with self.assertRaisesRegex(ValueError, "超过 Excel 上限"):
                write_output_xlsx(output_path, rows, OrderedDict())

            self.assertFalse(output_path.exists())

    def test_rejects_output_taller_than_the_excel_row_limit(self) -> None:
        class OversizedRows:
            def __len__(self) -> int:
                return EXCEL_MAX_ROWS + 1

            def __iter__(self):
                raise AssertionError("行数超限后不应继续遍历数据")

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "too_tall.xlsx"

            with self.assertRaisesRegex(ValueError, "超过 Excel 上限"):
                write_output_xlsx(
                    output_path,
                    OversizedRows(),  # type: ignore[arg-type]
                    OrderedDict(),
                )

            self.assertFalse(output_path.exists())


if __name__ == "__main__":
    unittest.main()
