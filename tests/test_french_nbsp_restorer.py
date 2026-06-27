from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.french_nbsp_restorer.restore_french_nbsp import process_excel, restore_french_nbsp


NBSP = "\u00a0"


class RestoreFrenchNbspTextTests(unittest.TestCase):
    def test_restores_nbsp_before_french_double_punctuation(self) -> None:
        self.assertEqual(
            restore_french_nbsp("Bonjour! Attention ; vraiment : oui ?"),
            f"Bonjour{NBSP}! Attention{NBSP}; vraiment{NBSP}: oui{NBSP}?",
        )

    def test_restores_nbsp_before_percent_sign(self) -> None:
        self.assertEqual(
            restore_french_nbsp("Progression 10% ou 20 %."),
            f"Progression 10{NBSP}% ou 20{NBSP}%.",
        )

    def test_restores_nbsp_inside_french_guillemets(self) -> None:
        self.assertEqual(
            restore_french_nbsp('Il dit « Bonjour ! »'),
            f"Il dit «{NBSP}Bonjour{NBSP}!{NBSP}»",
        )

    def test_keeps_url_and_time_colons_unchanged(self) -> None:
        self.assertEqual(
            restore_french_nbsp("Voir https://example.com à 12:30."),
            "Voir https://example.com à 12:30.",
        )

    def test_keeps_url_query_punctuation_unchanged(self) -> None:
        self.assertEqual(
            restore_french_nbsp("Voir https://example.com/search?q=oui!"),
            "Voir https://example.com/search?q=oui!",
        )

    def test_keeps_url_percent_encoding_unchanged(self) -> None:
        self.assertEqual(
            restore_french_nbsp("Voir https://example.com/search?q=10%25."),
            "Voir https://example.com/search?q=10%25.",
        )

    def test_returns_non_string_values_unchanged(self) -> None:
        self.assertIsNone(restore_french_nbsp(None))
        self.assertEqual(restore_french_nbsp(42), 42)


class RestoreFrenchNbspExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["C1"] = "fixed target"
        worksheet["B2"] = "Bonjour !"
        worksheet["B3"] = "Pas besoin."
        worksheet["B4"] = None
        worksheet["C2"] = "old fixed"
        worksheet["C3"] = "old passthrough"
        worksheet["C4"] = "old none"
        workbook.save(path)

    def test_process_excel_overwrites_target_column_in_output_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_file=input_path,
                target_column="B",
                sheet="Data",
                start_row=2,
                output_file=output_path,
            )

            self.assertEqual(summary.output_path, output_path.resolve())
            self.assertEqual(summary.worksheet_title, "Data")
            self.assertEqual(summary.target_column, "B")
            self.assertIsNone(summary.result_column)
            self.assertEqual(summary.processed_count, 3)
            self.assertEqual(summary.changed_count, 1)

            original_workbook = load_workbook(input_path)
            self.assertEqual(original_workbook["Data"]["B2"].value, "Bonjour !")

            output_workbook = load_workbook(output_path)
            output_sheet = output_workbook["Data"]
            self.assertEqual(output_sheet["B2"].value, f"Bonjour{NBSP}!")
            self.assertEqual(output_sheet["B3"].value, "Pas besoin.")
            self.assertIsNone(output_sheet["B4"].value)

    def test_process_excel_writes_complete_fixed_result_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_file=input_path,
                target_column="B",
                result_column="C",
                sheet="Data",
                start_row=2,
                output_file=output_path,
            )

            self.assertEqual(summary.target_column, "B")
            self.assertEqual(summary.result_column, "C")
            self.assertEqual(summary.processed_count, 3)
            self.assertEqual(summary.changed_count, 1)

            output_workbook = load_workbook(output_path)
            output_sheet = output_workbook["Data"]
            self.assertEqual(output_sheet["B2"].value, "Bonjour !")
            self.assertEqual(output_sheet["C2"].value, f"Bonjour{NBSP}!")
            self.assertEqual(output_sheet["C3"].value, "Pas besoin.")
            self.assertIsNone(output_sheet["C4"].value)


if __name__ == "__main__":
    unittest.main()
