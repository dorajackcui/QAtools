import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


class StandardWorkflowScriptTests(unittest.TestCase):
    def test_run_standard_workflow_returns_pair_and_coverage_stats(self):
        from scripts.run_standard_workflow import run_standard_workflow

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_path = tmp_path / "TM.xlsx"
            source_path = tmp_path / "for_test.xlsx"
            _write_tm(tm_path)
            _write_source(source_path)

            result = run_standard_workflow(
                source_workbook=source_path,
                tm_workbook=tm_path,
                source_col="source",
                target_col="target",
            )

            self.assertEqual(result.tm_stats["tm_pair_count"], 2)
            self.assertEqual(result.tm_stats["template_pair_count"], 1)
            self.assertEqual(result.tm_stats["segment_pair_count"], 1)
            self.assertEqual(result.extract_stats["row_count"], 3)
            self.assertEqual(result.extract_stats["translation_unit_count"], 3)
            self.assertEqual(result.extract_stats["prefilled_translation_unit_count"], 2)
            self.assertEqual(result.extract_stats["autofilled_count"], 2)
            self.assertEqual(result.rates["tm_unit_hit_rate"], "66.67%")
            self.assertEqual(result.rates["tm_row_hit_rate"], "66.67%")
            self.assertEqual(result.filled_result["non_empty_targets"], 2)
            self.assertEqual(result.filled_result["residual_token_rows"], 0)

            self.assertTrue(result.tm_pairs_workbook.exists())
            self.assertTrue(result.prefill_pack.exists())
            self.assertTrue(result.translator_todo.exists())
            self.assertTrue(result.filled_workbook.exists())

            filled = load_workbook(result.filled_workbook, data_only=True)
            try:
                rows = list(filled.active.iter_rows(values_only=True))
            finally:
                filled.close()

        self.assertEqual(rows[1][1], "Bonjour")
        self.assertEqual(rows[2][1], "Pack VIP30")
        self.assertIsNone(rows[3][1])


def _write_tm(path: Path) -> None:
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.append(["source", "target"])
        worksheet.append(["Hello", "Bonjour"])
        worksheet.append(["VIP10 Pack", "Pack VIP10"])
        worksheet.append(["VIP20 Pack", "Pack VIP20"])
        workbook.save(path)
    finally:
        workbook.close()


def _write_source(path: Path) -> None:
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.append(["source", "target"])
        worksheet.append(["Hello", None])
        worksheet.append(["VIP30 Pack", None])
        worksheet.append(["New text", None])
        workbook.save(path)
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
