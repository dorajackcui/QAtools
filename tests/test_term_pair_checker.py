from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from tools.term_pair_checker import extract_terms_from_excel as term_pair_module
from tools.term_pair_checker.extract_terms_from_excel import (
    extract_terms,
    process_excel,
)
class ExtractTermsTests(unittest.TestCase):
    def test_term_mark_helpers_are_available_from_focused_module(self) -> None:
        from tools.term_pair_checker.term_marks import (
            extract_terms as extract_terms_from_marks,
            strip_supported_marks,
        )

        text = "建立 [Alpha] 和 【Beta】"

        self.assertEqual(
            extract_terms_from_marks(text, mark_styles=("[]", "【】")),
            ["[Alpha]", "【Beta】"],
        )
        self.assertEqual(
            strip_supported_marks(text, mark_styles=("[]", "【】")),
            "建立 Alpha 和 Beta",
        )

    def test_workbook_output_helpers_are_available_from_focused_module(self) -> None:
        from tools.term_pair_checker.workbook_output import build_row_problem_summaries

        summaries = build_row_problem_summaries(
            [
                term_pair_module.ProblemEntry(
                    row_index=7,
                    problem_source_term="Alpha",
                    expected_target_term="ALPHA_OK",
                    term_source="本批次新增",
                    description="target缺少预期术语",
                    source_snapshot="source",
                    target_snapshot="target",
                ),
                term_pair_module.ProblemEntry(
                    row_index=7,
                    problem_source_term="Beta",
                    expected_target_term="BETA_OK",
                    term_source="本批次新增",
                    description="target术语不匹配：实际术语 - WRONG",
                    source_snapshot="source",
                    target_snapshot="target",
                ),
            ]
        )

        self.assertEqual(
            summaries,
            {
                7: (
                    "Alpha -> ALPHA_OK：target缺少预期术语；"
                    "Beta -> BETA_OK：target术语不匹配：实际术语 - WRONG"
                )
            },
        )

    def test_extract_terms_supports_selected_term_marks_in_text_order(self) -> None:
        text = "前缀[方括号]中间<尖括号tag>后缀【书名号】"
        self.assertEqual(
            extract_terms(text, mark_styles=("[]", "【】")),
            ["[方括号]", "【书名号】"],
        )

    def test_extract_terms_keeps_fullwidth_square_bracket_compatibility(self) -> None:
        self.assertEqual(extract_terms("这里有［全角方括号］", mark_styles=("[]",)), ["［全角方括号］"])

    def test_extract_terms_defaults_to_book_title_and_square_marks(self) -> None:
        self.assertEqual(extract_terms("建立【梨】和[苹果]"), ["【梨】", "[苹果]"])

    def test_extract_terms_requires_at_least_one_mark_style(self) -> None:
        with self.assertRaisesRegex(ValueError, "请至少选择一种 mark 类型"):
            extract_terms("任意文本", mark_styles=())

    def test_extract_terms_ignores_square_color_tags(self) -> None:
        text = "样式 [color=red]文字[/color]，真术语 [苹果]、[color] 和 【梨】"

        self.assertEqual(extract_terms(text, mark_styles=("[]", "【】")), ["[苹果]", "[color]", "【梨】"])

    def test_extract_terms_rejects_angle_brackets_as_term_mark(self) -> None:
        with self.assertRaisesRegex(ValueError, "不支持的 mark 类型"):
            extract_terms("真术语 <苹果>", mark_styles=("<>",))

    def test_extract_terms_ignores_numbered_tag_boundaries_between_marked_terms(self) -> None:
        text = (
            "3. You can obtain corresponding faction crystals by challenging the "
            "{1>[Will] Tower<2}, {3>[Order] Tower<4}, {5>[Origin] Tower<6}, "
            "and {7>Divinity [Tower]<8} in Dragonspiral."
        )

        self.assertEqual(
            extract_terms(text, mark_styles=("[]",)),
            ["[Will]", "[Order]", "[Origin]", "[Tower]"],
        )

    def test_extract_terms_ignores_short_and_symbol_only_pseudo_terms(self) -> None:
        text = "<a> [b] 【Z】 [123] 【+10%】 [火] <A1> [HP]"

        self.assertEqual(
            extract_terms(text, mark_styles=("[]", "【】")),
            ["[火]", "[HP]"],
        )


class ProcessExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["A2"] = "第一行 [Alpha] 和 【Beta】"
        worksheet["B2"] = "第一行 [阿尔法] 和 【贝塔】"
        worksheet["A3"] = "第二行【Gamma】"
        worksheet["B3"] = "第二行【伽马】"
        worksheet["A4"] = "第三行复用 【Beta】"
        worksheet["B4"] = "第三行复用 【错误贝塔】"
        worksheet["A5"] = "第四行 [Alpha] 加【Gamma】"
        worksheet["B5"] = "第四行只有 [阿尔法]"
        workbook.save(path)

    def test_process_excel_supports_multiple_term_mark_types(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            expected_output_path = Path(tmp_dir) / "term_pair_check_input.xlsx"

            worksheet_title, source_col, target_col, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]", "【】"),
            )

            self.assertEqual(worksheet_title, "Data")
            self.assertEqual(source_col, "A")
            self.assertEqual(target_col, "B")
            self.assertEqual(saved_path, expected_output_path.resolve())
            self.assertEqual(term_count, 3)
            self.assertEqual(problem_count, 3)

            original_workbook = load_workbook(input_path)
            self.assertEqual(original_workbook.sheetnames, ["Data"])
            self.assertEqual(original_workbook["Data"]["A2"].value, "第一行 [Alpha] 和 【Beta】")

            workbook = load_workbook(saved_path)
            term_sheet = workbook["术语表"]
            self.assertEqual(term_sheet["A2"].value, "[Alpha]")
            self.assertEqual(term_sheet["B2"].value, "[阿尔法]")
            self.assertEqual(term_sheet["A3"].value, "【Beta】")
            self.assertEqual(term_sheet["B3"].value, "【贝塔】")
            self.assertEqual(term_sheet["A4"].value, "【Gamma】")
            self.assertEqual(term_sheet["B4"].value, "【伽马】")
            self.assertEqual(term_sheet["C2"].value, "Alpha")
            self.assertEqual(term_sheet["D2"].value, "阿尔法")
            self.assertEqual(term_sheet["C3"].value, "Beta")
            self.assertEqual(term_sheet["D3"].value, "贝塔")
            self.assertEqual(term_sheet["C4"].value, "Gamma")
            self.assertEqual(term_sheet["D4"].value, "伽马")

            problem_sheet = workbook["问题列"]
            self.assertEqual(problem_sheet["A2"].value, 5)
            self.assertEqual(problem_sheet["B2"].value, "Alpha、Gamma")
            self.assertEqual(problem_sheet["C2"].value, "阿尔法、伽马")
            self.assertEqual(problem_sheet["D2"].value, "本批次新增")
            self.assertIn("source/target术语数量不一致", str(problem_sheet["E2"].value))
            self.assertEqual(problem_sheet["F2"].value, "第四行 [Alpha] 加【Gamma】")
            self.assertEqual(problem_sheet["G2"].value, "第四行只有 [阿尔法]")
            self.assertEqual(problem_sheet["A3"].value, 4)
            self.assertEqual(problem_sheet["B3"].value, "Beta")
            self.assertEqual(problem_sheet["C3"].value, "贝塔")
            self.assertEqual(problem_sheet["D3"].value, "本批次新增")
            self.assertEqual(problem_sheet["E3"].value, "target术语不匹配：实际术语 - 错误贝塔")
            self.assertEqual(problem_sheet["F3"].value, "第三行复用 【Beta】")
            self.assertEqual(problem_sheet["G3"].value, "第三行复用 【错误贝塔】")
            self.assertEqual(problem_sheet["A4"].value, 5)
            self.assertEqual(problem_sheet["B4"].value, "Gamma")
            self.assertEqual(problem_sheet["C4"].value, "伽马")
            self.assertEqual(problem_sheet["D4"].value, "本批次新增")
            self.assertEqual(problem_sheet["E4"].value, "target缺少预期术语")
            self.assertEqual(problem_sheet["F4"].value, "第四行 [Alpha] 加【Gamma】")
            self.assertEqual(problem_sheet["G4"].value, "第四行只有 [阿尔法]")

    def test_process_excel_dedupes_same_plain_term_with_different_marks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "第一行 [Alpha]"
            worksheet["B2"] = "第一行 [阿尔法]"
            worksheet["A3"] = "第二行 <Alpha>"
            worksheet["B3"] = "第二行 <阿尔法>"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]", "【】"),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            term_sheet = result_workbook["术语表"]
            self.assertEqual(term_sheet["A2"].value, "[Alpha]")
            self.assertEqual(term_sheet["B2"].value, "[阿尔法]")
            self.assertEqual(term_sheet["C2"].value, "Alpha")
            self.assertEqual(term_sheet["D2"].value, "阿尔法")
            self.assertEqual(term_sheet.max_row, 2)

    def test_process_excel_retroactively_checks_unmarked_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "第三行先出现苹果"
            worksheet["B2"] = "第三行先出现banana"
            worksheet["A3"] = "第十一行才标记出 [苹果]"
            worksheet["B3"] = "第十一行才标记出 [apple]"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 1)

            result_workbook = load_workbook(saved_path)
            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet["A2"].value, 2)
            self.assertEqual(problem_sheet["B2"].value, "苹果")
            self.assertEqual(problem_sheet["C2"].value, "apple")
            self.assertEqual(problem_sheet["D2"].value, "本批次新增")
            self.assertEqual(problem_sheet["E2"].value, "target缺少预期术语")
            self.assertEqual(problem_sheet["F2"].value, "第三行先出现苹果")
            self.assertEqual(problem_sheet["G2"].value, "第三行先出现banana")

    def test_process_excel_records_count_mismatch_and_missing_target_separately(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "建立 [Alpha]"
            worksheet["B2"] = "建立 [ALPHA_OK]"
            worksheet["A3"] = "建立 [Beta]"
            worksheet["B3"] = "建立 [BETA_OK]"
            worksheet["A4"] = "复用 [Alpha] and [Beta]"
            worksheet["B4"] = "复用 [ALPHA_OK] only"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 2)
            self.assertEqual(problem_count, 2)

            result_workbook = load_workbook(saved_path)
            data_sheet = result_workbook["Data"]
            self.assertEqual(data_sheet["A1"].value, "source")
            self.assertEqual(data_sheet["B1"].value, "target")
            self.assertEqual(data_sheet["C1"].value, "术语QA问题")
            self.assertIsNone(data_sheet["C2"].value)
            self.assertIsNone(data_sheet["C3"].value)
            self.assertEqual(
                data_sheet["C4"].value,
                (
                    "Alpha、Beta -> ALPHA_OK、BETA_OK：source/target术语数量不一致："
                    "2（预期数量）- 1（实际数量）；Beta -> BETA_OK：target缺少预期术语"
                ),
            )

            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet["A2"].value, 4)
            self.assertEqual(problem_sheet["B2"].value, "Alpha、Beta")
            self.assertIn("source/target术语数量不一致", str(problem_sheet["E2"].value))
            self.assertEqual(problem_sheet["A3"].value, 4)
            self.assertEqual(problem_sheet["B3"].value, "Beta")
            self.assertEqual(problem_sheet["C3"].value, "BETA_OK")
            self.assertEqual(problem_sheet["E3"].value, "target缺少预期术语")

    def test_process_excel_resolves_count_mismatch_for_marked_term_when_other_row_term_is_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Use [Attack]"
            worksheet["B2"] = "Use [ATK]"
            worksheet["A3"] = "Target [Battle_Target_Enemy]"
            worksheet["B3"] = "Target [Lumi adverse actuel]"
            worksheet["A4"] = "Attack [Battle_Target_Enemy]"
            worksheet["B4"] = "Attaque le Lumi adverse actuel."
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 2)
            self.assertEqual(problem_count, 1)

            result_workbook = load_workbook(saved_path)
            data_sheet = result_workbook["Data"]
            self.assertIn("Attack", data_sheet["C4"].value)
            self.assertNotIn("Battle_Target_Enemy", data_sheet["C4"].value)

            problem_sheet = result_workbook[term_pair_module.PROBLEM_SHEET_NAME]
            self.assertEqual(problem_sheet.max_row, 2)
            self.assertEqual(problem_sheet["A2"].value, 4)
            self.assertEqual(problem_sheet["B2"].value, "Attack")
            self.assertEqual(problem_sheet["C2"].value, "ATK")
            self.assertIn("target", problem_sheet["E2"].value)

    def test_process_excel_preserves_mark_boundaries_next_to_digits(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Target [Battle_Target_EnemyOnGround]"
            worksheet["B2"] = "Target [tous les Lumi du camp adverse en combat]"
            worksheet["A3"] = "Restore [Battle_Target_EnemyOnGround]20% Max HP"
            worksheet["B3"] = "Restaure 20 % des PV max de tous les Lumi du camp adverse en combat."
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            data_sheet = result_workbook["Data"]
            self.assertIsNone(data_sheet["C3"].value)

    def test_process_excel_records_mismatched_marked_term_and_missing_target_separately(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "建立 [Alpha]"
            worksheet["B2"] = "建立 [ALPHA_OK]"
            worksheet["A3"] = "建立 [Beta]"
            worksheet["B3"] = "建立 [BETA_OK]"
            worksheet["A4"] = "复用 [Alpha] and Beta"
            worksheet["B4"] = "复用 [WRONG] and untranslated"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 2)
            self.assertEqual(problem_count, 2)

            result_workbook = load_workbook(saved_path)
            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet["A2"].value, 4)
            self.assertEqual(problem_sheet["B2"].value, "Alpha")
            self.assertEqual(problem_sheet["C2"].value, "ALPHA_OK")
            self.assertEqual(problem_sheet["E2"].value, "target术语不匹配：实际术语 - WRONG")
            self.assertEqual(problem_sheet["A3"].value, 4)
            self.assertEqual(problem_sheet["B3"].value, "Beta")
            self.assertEqual(problem_sheet["C3"].value, "BETA_OK")
            self.assertEqual(problem_sheet["E3"].value, "target缺少预期术语")

    def test_process_excel_dedupes_repeated_same_row_term_problem(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "建立 [Sunlight]"
            worksheet["B2"] = "建立 [Rayon Soleil]"
            worksheet["A3"] = "触发 [Sunlight] 后转移 [Sunlight]"
            worksheet["B3"] = "触发 [Lumiere solaire] 后转移 [Lumiere solaire]"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 1)

            result_workbook = load_workbook(saved_path)
            data_sheet = result_workbook["Data"]
            expected_problem = (
                "Sunlight -> Rayon Soleil：target术语不匹配：实际术语 - Lumiere solaire"
            )
            self.assertEqual(data_sheet["C3"].value, expected_problem)

            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet.max_row, 2)
            self.assertEqual(problem_sheet["A2"].value, 3)
            self.assertEqual(problem_sheet["B2"].value, "Sunlight")
            self.assertEqual(problem_sheet["C2"].value, "Rayon Soleil")
            self.assertEqual(problem_sheet["E2"].value, "target术语不匹配：实际术语 - Lumiere solaire")

    def test_process_excel_treats_marked_target_as_aligned_for_unmarked_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "前面未标记的苹果"
            worksheet["B2"] = "前面已写成 [apple]"
            worksheet["A3"] = "后面用 [苹果] 建立术语"
            worksheet["B3"] = "后面用 [apple] 建立术语"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet.max_row, 1)

    def test_process_excel_ignores_extra_target_mark_when_source_terms_are_aligned(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "定义 [Basic CTRL]"
            worksheet["B2"] = "定义 [CTRL de base]"
            worksheet["A3"] = "Receives a [Basic CTRL] effect and enters HP Lock."
            worksheet["B3"] = "Subit [CTRL de base] et entre en [Verrouillage des PV]."
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            self.assertIsNone(result_workbook["Data"]["C3"].value)
            self.assertEqual(result_workbook["问题列"].max_row, 1)

    def test_resolved_count_mismatch_still_checks_other_source_terms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "定义 [Basic CTRL]"
            worksheet["B2"] = "定义 [CTRL de base]"
            worksheet["A3"] = "定义 [Poison]"
            worksheet["B3"] = "定义 [Poison]"
            worksheet["A4"] = "Receives [Basic CTRL] and Poison."
            worksheet["B4"] = "Subit [CTRL de base] et [Marque supplémentaire]."
            workbook.save(input_path)

            _, _, _, saved_path, _, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(problem_count, 1)

            result_workbook = load_workbook(saved_path)
            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet["A2"].value, 4)
            self.assertEqual(problem_sheet["B2"].value, "Poison")
            self.assertEqual(problem_sheet["E2"].value, "target缺少预期术语")

    def test_process_excel_treats_simple_s_plural_source_and_target_as_aligned(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "定义 [shard]"
            worksheet["B2"] = "定义 [éclat]"
            worksheet["A3"] = "Collect shards."
            worksheet["B3"] = "Collectez des éclats."
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            self.assertEqual(result_workbook["问题列"].max_row, 1)

    def test_process_excel_skips_target_only_simple_s_plural_variant(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "定义 [shard]"
            worksheet["B2"] = "定义 [éclat]"
            worksheet["A3"] = "Collect one shard."
            worksheet["B3"] = "Collectez des éclats."
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet.max_row, 1)

    def test_process_excel_skips_source_only_simple_s_plural_variant(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "定义 [shard]"
            worksheet["B2"] = "定义 [éclat]"
            worksheet["A3"] = "Collect shards."
            worksheet["B3"] = "Collectez des fragments."
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet.max_row, 1)

    def test_process_excel_skips_plural_signature_variants(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "定义 [EXP Candy]"
            worksheet["B2"] = "定义 [Bonbon Exp.]"
            worksheet["A3"] = "定义 [Gacha Coin]"
            worksheet["B3"] = "定义 [Pièce de gacha]"
            worksheet["A4"] = "Open EXP Candies."
            worksheet["B4"] = "Ouvrir Bonbons Exp."
            worksheet["A5"] = "Use Gacha Coins."
            worksheet["B5"] = "Utiliser des Pièces de gacha."
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 2)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet.max_row, 1)

    def test_process_excel_uses_hybrid_boundary_for_retroactive_checks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "account setup"
            worksheet["B2"] = "account setup"
            worksheet["A3"] = "定义 [ACC]"
            worksheet["B3"] = "定义 [ACC]"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            self.assertEqual(result_workbook["问题列"].max_row, 1)

    def test_process_excel_rejects_empty_marks_without_history_tb(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            with self.assertRaisesRegex(ValueError, "必须提供历史 TB"):
                process_excel(
                    input_file=input_path,
                    source_column="A",
                    target_column="B",
                    start_row=2,
                    mark_styles=(),
                )

    def test_process_excel_ignores_angle_tags_as_term_marks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "</> <color=red> <outline color=blue>"
            worksheet["B2"] = "</> <color=red> <outline color=blue>"
            worksheet["A3"] = "真实术语 [苹果]"
            worksheet["B3"] = "真实术语 [apple]"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            term_sheet = result_workbook["术语表"]
            self.assertEqual(term_sheet["A2"].value, "[苹果]")
            self.assertEqual(term_sheet["B2"].value, "[apple]")
            self.assertEqual(term_sheet["C2"].value, "苹果")
            self.assertEqual(term_sheet["D2"].value, "apple")
            self.assertEqual(term_sheet.max_row, 2)

    def test_process_excel_ignores_rt_style_tags_around_expected_target_term(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            history_path = Path(tmp_dir) / "history.xlsx"

            history_workbook = Workbook()
            history_sheet = history_workbook.active
            history_sheet.title = "TB"
            history_sheet["A1"] = "source"
            history_sheet["B1"] = "target"
            history_sheet["A2"] = "榴月"
            history_sheet["B2"] = "Mois de la Grenade"
            history_workbook.save(history_path)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "种植<rt style=txt_color_54>榴月</rt>"
            worksheet["B2"] = "Planter <rt style=txt_color_54>Mois de la Grenade</rt>"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
                history_tb_file=history_path,
                history_sheet="TB",
            )

            self.assertEqual(term_count, 1)
            self.assertEqual(problem_count, 0)

            result_workbook = load_workbook(saved_path)
            data_sheet = result_workbook["Data"]
            self.assertIsNone(data_sheet["C2"].value)
            self.assertEqual(result_workbook["问题列"].max_row, 1)

    def test_process_excel_prefers_history_tb_target_for_existing_plain_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            history_path = Path(tmp_dir) / "history.xlsx"

            history_workbook = Workbook()
            history_sheet = history_workbook.active
            history_sheet.title = "术语表"
            history_sheet["A1"] = "source术语"
            history_sheet["B1"] = "target术语"
            history_sheet["C1"] = "source术语（无mark）"
            history_sheet["D1"] = "target术语（无mark）"
            history_sheet["A2"] = "[Apple]"
            history_sheet["B2"] = "[历史苹果]"
            history_sheet["C2"] = "Apple"
            history_sheet["D2"] = "错误CD译法"
            history_workbook.save(history_path)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "本批次标记 [Apple]"
            worksheet["B2"] = "本批次标记 [临时苹果]"
            worksheet["A3"] = "本批次新词 [Banana]"
            worksheet["B3"] = "本批次新词 [香蕉]"
            worksheet["A4"] = "未标记复用 Apple"
            worksheet["B4"] = "未标记复用 临时苹果"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
                history_tb_file=history_path,
            )

            self.assertEqual(term_count, 2)
            self.assertEqual(problem_count, 2)

            result_workbook = load_workbook(saved_path)
            term_sheet = result_workbook["术语表"]
            self.assertEqual(term_sheet["E1"].value, "术语来源")
            self.assertEqual(term_sheet["A2"].value, "Apple")
            self.assertEqual(term_sheet["B2"].value, "历史苹果")
            self.assertEqual(term_sheet["C2"].value, "Apple")
            self.assertEqual(term_sheet["D2"].value, "历史苹果")
            self.assertEqual(term_sheet["E2"].value, "历史TB")
            self.assertEqual(term_sheet["A3"].value, "[Banana]")
            self.assertEqual(term_sheet["B3"].value, "[香蕉]")
            self.assertEqual(term_sheet["C3"].value, "Banana")
            self.assertEqual(term_sheet["D3"].value, "香蕉")
            self.assertEqual(term_sheet["E3"].value, "本批次新增")

            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet["A2"].value, 2)
            self.assertEqual(problem_sheet["B2"].value, "Apple")
            self.assertEqual(problem_sheet["C2"].value, "历史苹果")
            self.assertEqual(problem_sheet["D2"].value, "历史TB")
            self.assertEqual(problem_sheet["E2"].value, "target缺少预期术语")
            self.assertEqual(problem_sheet["A3"].value, 4)
            self.assertEqual(problem_sheet["B3"].value, "Apple")
            self.assertEqual(problem_sheet["C3"].value, "历史苹果")
            self.assertEqual(problem_sheet["D3"].value, "历史TB")

    def test_process_excel_checks_full_history_tb_plus_new_batch_terms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            history_path = Path(tmp_dir) / "history.xlsx"

            history_workbook = Workbook()
            history_sheet = history_workbook.active
            history_sheet.title = "TB"
            history_sheet["A1"] = "source"
            history_sheet["B1"] = "target"
            history_sheet["A2"] = "Apple"
            history_sheet["B2"] = "Pomme"
            history_sheet["A3"] = "Orange"
            history_sheet["B3"] = "Orange historique"
            history_workbook.save(history_path)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "本轮新增 [Banana]"
            worksheet["B2"] = "本轮新增 [Banane]"
            worksheet["A3"] = "未标记复用 Apple"
            worksheet["B3"] = "未标记缺少历史译法"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
                history_tb_file=history_path,
                history_sheet="TB",
            )

            self.assertEqual(term_count, 2)
            self.assertEqual(problem_count, 1)

            result_workbook = load_workbook(saved_path)
            term_sheet = result_workbook["术语表"]
            self.assertEqual(term_sheet["A2"].value, "Apple")
            self.assertEqual(term_sheet["B2"].value, "Pomme")
            self.assertEqual(term_sheet["E2"].value, "历史TB")
            self.assertEqual(term_sheet["A3"].value, "[Banana]")
            self.assertEqual(term_sheet["B3"].value, "[Banane]")
            self.assertEqual(term_sheet["E3"].value, "本批次新增")

            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet["A2"].value, 3)
            self.assertEqual(problem_sheet["B2"].value, "Apple")
            self.assertEqual(problem_sheet["C2"].value, "Pomme")

    def test_process_excel_can_check_history_tb_without_term_marks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            history_path = Path(tmp_dir) / "history.xlsx"

            history_workbook = Workbook()
            history_sheet = history_workbook.active
            history_sheet.title = "TB"
            history_sheet["A1"] = "source"
            history_sheet["B1"] = "target"
            history_sheet["A2"] = "Apple"
            history_sheet["B2"] = "Pomme"
            history_sheet["A3"] = "Banana"
            history_sheet["B3"] = "Banane"
            history_workbook.save(history_path)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Use Apple"
            worksheet["B2"] = "Utiliser Pomme"
            worksheet["A3"] = "Use Banana"
            worksheet["B3"] = "Utiliser le mauvais terme"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=(),
                history_tb_file=history_path,
                history_sheet="TB",
            )

            self.assertEqual(term_count, 2)
            self.assertEqual(problem_count, 1)

            result_workbook = load_workbook(saved_path)
            term_sheet = result_workbook["术语表"]
            self.assertEqual(term_sheet["A2"].value, "Apple")
            self.assertEqual(term_sheet["A3"].value, "Banana")
            self.assertEqual(term_sheet["E2"].value, "历史TB")
            self.assertEqual(term_sheet["E3"].value, "历史TB")

            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet.max_row, 2)
            self.assertEqual(problem_sheet["A2"].value, 3)
            self.assertEqual(problem_sheet["B2"].value, "Banana")
            self.assertEqual(problem_sheet["C2"].value, "Banane")

    def test_problem_sheet_includes_term_source_and_sorts_history_first(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            history_path = Path(tmp_dir) / "history.xlsx"

            history_workbook = Workbook()
            history_sheet = history_workbook.active
            history_sheet.title = "TB"
            history_sheet["A1"] = "source"
            history_sheet["B1"] = "target"
            history_sheet["A2"] = "Zebra"
            history_sheet["B2"] = "Zebre"
            history_workbook.save(history_path)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "建立新增 [Apple]"
            worksheet["B2"] = "建立新增 [Pomme]"
            worksheet["A3"] = "复用新增 Apple"
            worksheet["B3"] = "复用新增错误"
            worksheet["A4"] = "复用历史 Zebra"
            worksheet["B4"] = "复用历史错误"
            workbook.save(input_path)

            _, _, _, saved_path, term_count, problem_count = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                start_row=2,
                mark_styles=("[]",),
                history_tb_file=history_path,
                history_sheet="TB",
            )

            self.assertEqual(term_count, 2)
            self.assertEqual(problem_count, 2)

            result_workbook = load_workbook(saved_path)
            problem_sheet = result_workbook["问题列"]
            self.assertEqual(problem_sheet["C1"].value, "预期target术语")
            self.assertEqual(problem_sheet["D1"].value, "术语来源")
            self.assertEqual(problem_sheet["E1"].value, "问题简述")
            self.assertEqual(problem_sheet["F1"].value, "source原文")
            self.assertEqual(problem_sheet["G1"].value, "target原文")
            self.assertEqual(problem_sheet["B2"].value, "Zebra")
            self.assertEqual(problem_sheet["C2"].value, "Zebre")
            self.assertEqual(problem_sheet["D2"].value, "历史TB")
            self.assertEqual(problem_sheet["B3"].value, "Apple")
            self.assertEqual(problem_sheet["C3"].value, "Pomme")
            self.assertEqual(problem_sheet["D3"].value, "本批次新增")

    def test_history_tb_loading_stops_after_consecutive_empty_tail_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Apple"
            worksheet["B2"] = "Pomme"
            worksheet["A6"] = "Late"
            worksheet["B6"] = "Tardif"
            workbook.save(history_path)

            with patch.object(term_pair_module, "HISTORY_EMPTY_ROW_STOP_THRESHOLD", 2, create=True):
                history_mapping = term_pair_module.load_history_tb_mapping(history_path)

            self.assertIn("apple", history_mapping)
            self.assertNotIn("late", history_mapping)

    def test_history_tb_start_row_keeps_row_one_header_detection(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "source术语"
            worksheet["B1"] = "target术语"
            worksheet["C1"] = "备注"
            worksheet["A2"] = "Skipped"
            worksheet["B2"] = "Ignore"
            worksheet["C2"] = "not a header"
            worksheet["A3"] = "Apple"
            worksheet["B3"] = "Pomme"
            worksheet["C3"] = "use me"
            workbook.save(history_path)

            history_mapping = term_pair_module.load_history_tb_mapping(
                history_path,
                start_row=3,
            )

            self.assertNotIn("skipped", history_mapping)
            self.assertEqual(history_mapping["apple"].target_plain_text, "Pomme")

    def test_history_tb_rejects_ambiguous_duplicate_source_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "source"
            worksheet["B1"] = "source"
            worksheet["C1"] = "target"
            worksheet["A2"] = "Wrong Source"
            worksheet["B2"] = "Apple"
            worksheet["C2"] = "Pomme"
            workbook.save(history_path)

            with self.assertRaisesRegex(ValueError, "缺少 source/target 列"):
                term_pair_module.load_history_tb_mapping(history_path)


if __name__ == "__main__":
    unittest.main()
