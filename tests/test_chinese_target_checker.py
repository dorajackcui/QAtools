from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from tools.chinese_target_checker.check_chinese_target import (
    PROBLEM_SHEET_NAME,
    contains_chinese,
    extract_chinese_characters,
    process_excel,
)


class ChineseTargetTextTests(unittest.TestCase):
    def test_contains_chinese_detects_cjk_ideographs_and_punctuation(self) -> None:
        self.assertTrue(contains_chinese("Keep 中文 text"))
        self.assertTrue(contains_chinese("Only punctuation ，。！？"))
        self.assertTrue(contains_chinese("Fullwidth marks 【】（）《》“”"))
        self.assertTrue(contains_chinese("Typography marks “”‘’·"))
        self.assertFalse(contains_chinese("Allowed punctuation —…"))
        self.assertFalse(contains_chinese(""))
        self.assertFalse(contains_chinese(None))
        self.assertFalse(contains_chinese(42))
        self.assertFalse(contains_chinese("ASCII punctuation ()[]!?"))
        self.assertFalse(contains_chinese("Fullwidth alnum ＡＢ１２"))

    def test_extract_chinese_characters_preserves_match_order(self) -> None:
        self.assertEqual(extract_chinese_characters("A中B，C（D）“E”…"), "中，（）“”")
        self.assertEqual(extract_chinese_characters("—…"), "")
        self.assertEqual(extract_chinese_characters("No Chinese"), "")


class ChineseTargetExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["C1"] = "existing data"
        worksheet["A2"] = "hello"
        worksheet["B2"] = "Bonjour"
        worksheet["C2"] = "keep"
        worksheet["A3"] = "hello again"
        worksheet["B3"] = "包含中文 target"
        worksheet["C3"] = "keep too"
        worksheet["A4"] = "empty target"
        worksheet["B4"] = None
        workbook.save(path)

    def test_process_excel_writes_unified_problem_sheet_without_changing_data(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
            )

            self.assertEqual(
                summary.output_path,
                input_path.with_name("target_chinese_check_input.xlsx").resolve(),
            )
            self.assertEqual(summary.source_column, "A")
            self.assertEqual(summary.target_column, "B")
            self.assertEqual(summary.processed_count, 3)
            self.assertEqual(summary.matched_count, 1)

            workbook = load_workbook(summary.output_path)
            problem_sheet = workbook[PROBLEM_SHEET_NAME]
            self.assertEqual(
                [problem_sheet.cell(1, column).value for column in range(1, 6)],
                ["行号", "source原文", "target原文", "问题描述", "命中字符"],
            )
            self.assertEqual(problem_sheet["A2"].value, 3)
            self.assertEqual(problem_sheet["B2"].value, "hello again")
            self.assertEqual(problem_sheet["C2"].value, "包含中文 target")
            self.assertIn("包含中文", problem_sheet["D2"].value)
            self.assertEqual(problem_sheet["E2"].value, "包含中文")
            self.assertEqual(problem_sheet["A2"].hyperlink.location, "'Data'!B3")
            self.assertIsNone(problem_sheet["A2"].hyperlink.target)
            with ZipFile(summary.output_path) as archive:
                worksheet_xml = "\n".join(
                    archive.read(name).decode("utf-8")
                    for name in archive.namelist()
                    if name.startswith("xl/worksheets/sheet")
                    and name.endswith(".xml")
                )
                self.assertIn('location="\'Data\'!B3"', worksheet_xml)
                self.assertFalse(
                    any(
                        name.startswith("xl/worksheets/_rels/")
                        for name in archive.namelist()
                    )
                )
            self.assertEqual(problem_sheet.freeze_panes, "A2")
            self.assertEqual(problem_sheet.auto_filter.ref, "A1:E2")
            self.assertEqual(workbook["Data"]["C1"].value, "existing data")
            self.assertEqual(workbook["Data"]["C3"].value, "keep too")
            self.assertEqual(load_workbook(input_path).sheetnames, ["Data"])

    def test_process_excel_rebuilds_current_and_removes_legacy_problem_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"
            self.create_workbook(input_path)
            workbook = load_workbook(input_path)
            workbook.create_sheet(PROBLEM_SHEET_NAME)["A1"] = "stale"
            workbook.create_sheet("中文检查问题")["A1"] = "legacy"
            workbook.save(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                output_file=output_path,
            )

            output_workbook = load_workbook(summary.output_path)
            self.assertEqual(output_workbook[PROBLEM_SHEET_NAME]["A1"].value, "行号")
            self.assertNotIn("中文检查问题", output_workbook.sheetnames)

    def test_process_excel_rejects_invalid_start_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            with self.assertRaisesRegex(ValueError, "开始行必须大于等于 1"):
                process_excel(
                    input_file=input_path,
                    source_column="A",
                    target_column="B",
                    start_row=0,
                )


if __name__ == "__main__":
    unittest.main()
