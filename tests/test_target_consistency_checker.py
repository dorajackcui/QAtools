from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.target_consistency_checker.check_target_consistency import (
    PROBLEM_SHEET_NAME,
    process_excel,
)


class TargetConsistencyExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["source", "target"])
        worksheet.append(["Save", "保存"])
        worksheet.append(["Store", "保存"])
        worksheet.append(["Same", "相同"])
        worksheet.append(["Same", "相同"])
        worksheet.append(["Blank target", None])
        worksheet.append(["Whitespace target", "   "])
        worksheet.append([None, "共用"])
        worksheet.append(["Non-empty source", "共用"])
        worksheet.append(["Trailing-space target", "保存 "])
        worksheet["Z1000"] = "unrelated tail"
        worksheet["A1001"].number_format = "@"
        workbook.save(path)
        workbook.close()

    def test_process_excel_lists_each_row_in_inconsistent_target_groups(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_path,
                source_column="a",
                target_column="b",
                sheet="Data",
            )

            self.assertEqual(summary.total_rows_checked, 9)
            self.assertEqual(summary.non_empty_target_rows, 7)
            self.assertEqual(summary.repeated_target_count, 3)
            self.assertEqual(summary.inconsistent_target_count, 2)
            self.assertEqual(summary.problem_rows, 4)
            self.assertEqual(
                summary.output_path,
                input_path.with_name("target_consistency_check_input.xlsx").resolve(),
            )

            workbook = load_workbook(summary.output_path)
            try:
                problem_sheet = workbook[PROBLEM_SHEET_NAME]
                self.assertEqual(
                    [problem_sheet.cell(1, column).value for column in range(1, 7)],
                    [
                        "行号",
                        "source原文",
                        "target原文",
                        "问题描述",
                        "source版本数",
                        "同组行号",
                    ],
                )
                self.assertEqual(
                    [problem_sheet[f"A{row}"].value for row in range(2, 6)],
                    [2, 3, 8, 9],
                )
                self.assertIn("2 个不同 source", problem_sheet["D2"].value)
                self.assertEqual(problem_sheet["E2"].value, 2)
                self.assertEqual(problem_sheet["F2"].value, "2、3")
                self.assertIsNone(problem_sheet["B4"].value)
                self.assertEqual(problem_sheet["C4"].value, "共用")
                self.assertEqual(problem_sheet["A2"].hyperlink.location, "'Data'!B2")
            finally:
                workbook.close()

    def test_process_excel_skips_empty_targets_and_rebuilds_stale_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            workbook = load_workbook(input_path)
            workbook.create_sheet(PROBLEM_SHEET_NAME)["A1"] = "stale"
            workbook.save(input_path)
            workbook.close()

            summary = process_excel(input_path, "A", "B", sheet="Data")

            workbook = load_workbook(summary.output_path)
            try:
                problem_sheet = workbook[PROBLEM_SHEET_NAME]
                self.assertEqual(problem_sheet["A1"].value, "行号")
                self.assertNotIn(6, [cell.value for cell in problem_sheet["A"]])
                self.assertNotIn(7, [cell.value for cell in problem_sheet["A"]])
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
