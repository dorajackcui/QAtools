import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from phraseloom.workflow import (
    fill_target_column_workbook,
    generate_tm_pairs,
    generate_workbook,
)


TESTFILES = Path("testfiles")


def _unit_header_index(headers, column):
    aliases = {
        "source_unit": ("source_unit", "source"),
        "target_unit": ("target_unit", "target"),
    }
    for candidate in aliases.get(column, (column,)):
        if candidate in headers:
            return headers.index(candidate)
    raise ValueError(f"{column!r} is not in list")


def create_tag_testfiles(tag_tm, tag_source):
    tm = Workbook()
    try:
        tm_ws = tm.active
        tm_ws.append(["source", "target"])
        tm_ws.append(
            ['<color=#fff>VIP10 Pack</color>', '<color=#fff>Pack VIP10 FR</color>']
        )
        tm_ws.append(
            ['<color=#fff>VIP20 Pack</color>', '<color=#fff>Pack VIP20 FR</color>']
        )
        tm_ws.append(["Login failed", "Login failed FR"])
        tm.save(tag_tm)
    finally:
        tm.close()

    source = Workbook()
    try:
        source_ws = source.active
        source_ws.append(["source", "target"])
        source_ws.append(['<color=#fff>VIP30 Pack</color>', None])
        source_ws.append(['<color=#fff>VIP40 Pack</color>', None])
        source_ws.append(['<img src="coin.png"/>', None])
        source_ws.append(["Brand new line", None])
        source.save(tag_source)
    finally:
        source.close()


def sheet_rows_by_header(workbook_path, sheet_name):
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        return [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        workbook.close()


class TagWorkflowTestfilesTests(unittest.TestCase):
    def setUp(self):
        TESTFILES.mkdir(exist_ok=True)
        self.tmp = tempfile.TemporaryDirectory(dir=TESTFILES)
        self.work_dir = Path(self.tmp.name)
        self.tag_tm = self.work_dir / "tag_tm.xlsx"
        self.tag_source = self.work_dir / "tag_source.xlsx"
        create_tag_testfiles(self.tag_tm, self.tag_source)

    def tearDown(self):
        self.tmp.cleanup()

    def test_tm_extract_and_fill_use_testfiles_with_tags(self):
        tm_pairs = self.work_dir / "tag_tm_l10n" / "tag_tm_reusable_units.xlsx"
        source_pack = self.work_dir / "tag_source_l10n" / "tag_source_pack.xlsx"
        standalone_todo = (
            self.work_dir / "tag_source_l10n" / "tag_source_translator_todo.xlsx"
        )
        filled_workbook = self.work_dir / "tag_source_l10n" / "tag_source_filled.xlsx"

        tm_stats = generate_tm_pairs(
            self.tag_tm,
            tm_pairs,
            source_col="source",
            target_col="target",
            min_group_size=2,
        )

        self.assertEqual(tm_stats["template_pair_count"], 1)

        stats = generate_workbook(
            self.tag_source,
            source_pack,
            source_col="source",
            target_col="target",
            tm_workbook=tm_pairs,
            min_group_size=2,
            use_existing_targets=False,
        )

        self.assertEqual(stats["prefilled_translation_unit_count"], 2)
        self.assertEqual(stats["untranslated_translation_unit_count"], 1)

        todo_workbook = load_workbook(standalone_todo)
        try:
            todo = todo_workbook["to_translate"]
            todo_headers = [cell.value for cell in todo[1]]
            source_idx = _unit_header_index(todo_headers, "source_unit") + 1
            target_idx = _unit_header_index(todo_headers, "target_unit") + 1
            todo_sources = [
                row[source_idx - 1].value for row in todo.iter_rows(min_row=2)
            ]
            self.assertEqual(todo_sources, ["Brand new line"])
            todo.cell(row=2, column=target_idx).value = "Brand new line FR"
            todo_workbook.save(standalone_todo)
        finally:
            todo_workbook.close()

        fill_stats = fill_target_column_workbook(
            self.tag_source,
            filled_workbook,
            source_col="source",
            target_col="target",
            template_workbook=standalone_todo,
            min_group_size=2,
        )

        self.assertEqual(fill_stats["autofilled_count"], 4)

        filled = load_workbook(filled_workbook, data_only=True)
        try:
            rows = list(filled.active.iter_rows(values_only=True))
            self.assertEqual(rows[1][1], '<color=#fff>Pack VIP30 FR</color>')
            self.assertEqual(rows[2][1], '<color=#fff>Pack VIP40 FR</color>')
            self.assertEqual(rows[3][1], '<img src="coin.png"/>')
            self.assertEqual(rows[4][1], "Brand new line FR")
        finally:
            filled.close()

    def test_fill_writes_target_when_protected_token_mismatch_warning_exists(self):
        pack = self.work_dir / "tag_source_l10n" / "tag_source_pack.xlsx"
        report = self.work_dir / "tag_source_l10n" / "tag_source_report.xlsx"

        generate_workbook(
            self.tag_source,
            pack,
            source_col="source",
            target_col="target",
            min_group_size=2,
            use_existing_targets=False,
        )

        pack_workbook = load_workbook(pack)
        try:
            units = pack_workbook["translation_units"]
            headers = [cell.value for cell in units[1]]
            source_idx = _unit_header_index(headers, "source_unit") + 1
            target_idx = _unit_header_index(headers, "target_unit") + 1
            for row in units.iter_rows(min_row=2):
                source_unit = row[source_idx - 1].value
                if source_unit == "{1>VIP{num1} Pack<2}":
                    row[target_idx - 1].value = "Pack VIP{num1} FR"
                elif source_unit == "Brand new line":
                    row[target_idx - 1].value = "Brand new line FR"
            pack_workbook.save(pack)
        finally:
            pack_workbook.close()

        generate_workbook(
            self.tag_source,
            report,
            source_col="source",
            target_col="target",
            template_workbook=pack,
            min_group_size=2,
            use_existing_targets=False,
        )

        rows = sheet_rows_by_header(report, "source_map")
        self.assertEqual(rows[0]["auto_target"], "Pack VIP30 FR")
        self.assertEqual(rows[1]["auto_target"], "Pack VIP40 FR")
        self.assertTrue(
            any("protected_token_mismatch" in str(row["warning"] or "") for row in rows)
        )


if __name__ == "__main__":
    unittest.main()
