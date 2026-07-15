from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.line_break_checker.check_line_breaks import (
    PROBLEM_SHEET_NAME,
    count_line_breaks,
    process_excel,
)


class LineBreakCountTests(unittest.TestCase):
    def test_count_line_breaks_handles_excel_newline_variants(self) -> None:
        self.assertEqual(count_line_breaks("one\ntwo\nthree"), 2)
        self.assertEqual(count_line_breaks("one\r\ntwo\r\nthree"), 2)
        self.assertEqual(count_line_breaks("one\rtwo"), 1)
        self.assertEqual(count_line_breaks(r"one\ntwo"), 0)
        self.assertEqual(count_line_breaks(None), 0)
        self.assertEqual(count_line_breaks(42), 0)


class LineBreakExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["A2"] = "第一行\n第二行"
        worksheet["B2"] = "First\nSecond"
        worksheet["A3"] = "第一行\n第二行\n第三行"
        worksheet["B3"] = "First\nSecond"
        worksheet["A4"] = "第一行\n第二行"
        worksheet["B4"] = None
        worksheet["A5"] = r"literal\nmark"
        worksheet["B5"] = "no real break"
        worksheet["Z1000"] = "unrelated tail"
        worksheet["A1001"].number_format = "@"
        workbook.save(path)

    def test_process_excel_writes_only_mismatched_rows_to_problem_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="a",
                target_column="b",
                sheet="Data",
                start_row=2,
            )

            self.assertEqual(
                summary.output_path,
                input_path.with_name("line_break_check_input.xlsx").resolve(),
            )
            self.assertEqual(summary.source_column, "A")
            self.assertEqual(summary.target_column, "B")
            self.assertEqual(summary.total_rows_checked, 4)
            self.assertEqual(summary.problem_rows, 2)

            workbook = load_workbook(summary.output_path)
            problem_sheet = workbook[PROBLEM_SHEET_NAME]
            self.assertEqual(problem_sheet.max_row, 3)
            self.assertEqual(
                [problem_sheet.cell(1, column).value for column in range(1, 8)],
                [
                    "行号",
                    "source原文",
                    "target原文",
                    "问题描述",
                    "source换行数",
                    "target换行数",
                    "数量差",
                ],
            )
            self.assertEqual(problem_sheet["A2"].value, 3)
            self.assertEqual(problem_sheet["B2"].value, "第一行\n第二行\n第三行")
            self.assertEqual(problem_sheet["C2"].value, "First\nSecond")
            self.assertIn("换行数量不一致", problem_sheet["D2"].value)
            self.assertEqual(problem_sheet["E2"].value, 2)
            self.assertEqual(problem_sheet["F2"].value, 1)
            self.assertEqual(problem_sheet["G2"].value, -1)
            self.assertEqual(problem_sheet["A3"].value, 4)
            self.assertEqual(problem_sheet["E3"].value, 1)
            self.assertEqual(problem_sheet["F3"].value, 0)
            self.assertIsNone(problem_sheet["C3"].value)
            self.assertEqual(problem_sheet["A2"].hyperlink.location, "'Data'!B3")
            self.assertEqual(problem_sheet["A3"].hyperlink.location, "'Data'!B4")
            self.assertIsNone(problem_sheet["A2"].hyperlink.target)
            self.assertEqual(problem_sheet.freeze_panes, "A2")
            self.assertEqual(problem_sheet.auto_filter.ref, "A1:G3")
            self.assertFalse(input_path.with_name("input.xlsx").resolve() == summary.output_path)
            self.assertNotIn(PROBLEM_SHEET_NAME, load_workbook(input_path).sheetnames)

    def test_process_excel_rebuilds_stale_problem_sheet_and_supports_custom_output(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "result.xlsx"
            self.create_workbook(input_path)
            workbook = load_workbook(input_path)
            stale_sheet = workbook.create_sheet(PROBLEM_SHEET_NAME)
            stale_sheet["A1"] = "stale"
            workbook.save(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                output_file=output_path,
            )

            self.assertEqual(summary.output_path, output_path.resolve())
            problem_sheet = load_workbook(output_path)[PROBLEM_SHEET_NAME]
            self.assertEqual(problem_sheet["A1"].value, "行号")

    def test_process_excel_rejects_invalid_start_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            with self.assertRaisesRegex(ValueError, "开始行必须大于等于 1"):
                process_excel(input_path, "A", "B", start_row=0)

    def test_process_excel_rejects_same_source_and_target_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            with self.assertRaisesRegex(ValueError, "不能相同"):
                process_excel(input_path, "A", "a")


if __name__ == "__main__":
    unittest.main()
