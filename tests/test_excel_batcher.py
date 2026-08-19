from __future__ import annotations

import json
from pathlib import Path
import tempfile
import unittest
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from tools.excel_batcher.excel_batcher import (
    MANIFEST_FILE_NAME,
    restore_batches,
    split_workbook,
)


class ExcelBatcherTests(unittest.TestCase):
    def replace_zip_entry(
        self,
        archive_path: Path,
        entry_name: str,
        replacement: bytes,
    ) -> None:
        with ZipFile(archive_path) as archive:
            entries = [
                (item, replacement if item.filename == entry_name else archive.read(item))
                for item in archive.infolist()
            ]
        with ZipFile(archive_path, "w", compression=ZIP_DEFLATED) as archive:
            for item, data in entries:
                archive.writestr(item, data)

    def create_workbook(self, path: Path, *, data_rows: int = 5) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Strings"
        worksheet.append(["key", "source", "target"])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:C{data_rows + 1}"
        worksheet.column_dimensions["B"].width = 32

        for index in range(1, data_rows + 1):
            worksheet.append(
                [
                    f"key-{index}",
                    f"Source {index}",
                    f"=B{index + 1}",
                ]
            )
            worksheet.cell(index + 1, 2).fill = PatternFill(
                fill_type="solid",
                fgColor="FFF2CC",
            )
            worksheet.row_dimensions[index + 1].height = 18 + index

        table = Table(displayName="StringsTable", ref=f"A1:C{data_rows + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
        )
        worksheet.add_table(table)

        notes = workbook.create_sheet("Notes")
        notes["A1"] = "This sheet must survive restoration."
        workbook.active = 0
        workbook.save(path)
        workbook.close()

    def test_split_creates_ordered_batches_with_repeated_header_and_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            output_dir = Path(tmp_dir) / "batches"
            self.create_workbook(source_path, data_rows=5)

            summary = split_workbook(
                source_path,
                sheet="Strings",
                batch_size=2,
                header_rows=1,
                output_dir=output_dir,
            )

            self.assertEqual(summary.batch_count, 3)
            self.assertEqual(summary.data_row_count, 5)
            self.assertEqual(
                [path.name for path in summary.batch_files],
                [
                    "project_batch_001_of_003.xlsx",
                    "project_batch_002_of_003.xlsx",
                    "project_batch_003_of_003.xlsx",
                ],
            )
            manifest = json.loads(
                (output_dir / MANIFEST_FILE_NAME).read_text(encoding="utf-8")
            )
            self.assertEqual(manifest["worksheet"], "Strings")
            self.assertEqual(manifest["data_row_count"], 5)
            self.assertEqual(
                [batch["source_start_row"] for batch in manifest["batches"]],
                [2, 4, 6],
            )

            expected_keys = (["key-1", "key-2"], ["key-3", "key-4"], ["key-5"])
            for batch_path, batch_keys in zip(
                summary.batch_files,
                expected_keys,
                strict=True,
            ):
                workbook = load_workbook(batch_path, data_only=False)
                try:
                    worksheet = workbook["Strings"]
                    self.assertEqual(workbook.sheetnames, ["Strings"])
                    self.assertEqual(worksheet["A1"].value, "key")
                    self.assertEqual(
                        [
                            worksheet.cell(row_index, 1).value
                            for row_index in range(2, 2 + len(batch_keys))
                        ],
                        batch_keys,
                    )
                    self.assertEqual(worksheet.freeze_panes, "A2")
                    self.assertEqual(
                        worksheet["B2"].fill.fgColor.rgb,
                        "00FFF2CC",
                    )
                finally:
                    workbook.close()

    def test_restore_round_trip_preserves_original_workbook_content_and_formatting(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            self.create_workbook(source_path, data_rows=5)
            summary = split_workbook(source_path, batch_size=2)
            restored_path = Path(tmp_dir) / "roundtrip.xlsx"

            restore_summary = restore_batches(
                summary.output_dir,
                output_file=restored_path,
            )

            self.assertEqual(restore_summary.batch_count, 3)
            self.assertEqual(restore_summary.restored_row_count, 5)
            original = load_workbook(source_path, data_only=False)
            restored = load_workbook(restored_path, data_only=False)
            try:
                self.assertEqual(restored.sheetnames, original.sheetnames)
                self.assertEqual(restored.active.title, original.active.title)
                self.assertEqual(restored["Notes"]["A1"].value, original["Notes"]["A1"].value)
                for row_index in range(1, 7):
                    for column_index in range(1, 4):
                        original_cell = original["Strings"].cell(row_index, column_index)
                        restored_cell = restored["Strings"].cell(row_index, column_index)
                        self.assertEqual(restored_cell.value, original_cell.value)
                        self.assertEqual(restored_cell.style_id, original_cell.style_id)
                self.assertEqual(restored["Strings"].column_dimensions["B"].width, 32)
                self.assertEqual(restored["Strings"].row_dimensions[5].height, 22)
            finally:
                original.close()
                restored.close()

    def test_restore_applies_batch_edits_to_their_original_rows_and_new_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            self.create_workbook(source_path, data_rows=5)
            summary = split_workbook(source_path, batch_size=2)

            first_batch = load_workbook(summary.batch_files[0])
            first_batch["Strings"]["C2"] = "Translated 1"
            first_batch["Strings"]["C3"] = None
            first_batch["Strings"]["D1"] = "review"
            first_batch["Strings"]["D2"] = "pass"
            first_batch.save(summary.batch_files[0])
            first_batch.close()

            second_batch = load_workbook(summary.batch_files[1])
            second_batch["Strings"]["C3"] = "Translated 4"
            second_batch["Strings"]["D1"] = "review"
            second_batch["Strings"]["D3"] = "check"
            second_batch.save(summary.batch_files[1])
            second_batch.close()

            restored_path = Path(tmp_dir) / "edited.xlsx"
            restore_batches(summary.manifest_path, output_file=restored_path)

            restored = load_workbook(restored_path, data_only=False)
            try:
                worksheet = restored["Strings"]
                self.assertEqual(worksheet["C2"].value, "Translated 1")
                self.assertIsNone(worksheet["C3"].value)
                self.assertEqual(worksheet["C5"].value, "Translated 4")
                self.assertEqual(worksheet["D1"].value, "review")
                self.assertEqual(worksheet["D2"].value, "pass")
                self.assertEqual(worksheet["D5"].value, "check")
            finally:
                restored.close()

    def test_restore_reads_shared_strings_from_excel_saved_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            self.create_workbook(source_path, data_rows=2)
            summary = split_workbook(source_path, batch_size=2)
            batch_path = summary.batch_files[0]

            with ZipFile(batch_path) as archive:
                entries = {
                    item.filename: (item, archive.read(item))
                    for item in archive.infolist()
                }
            worksheet_name = "xl/worksheets/sheet1.xml"
            worksheet_root = ElementTree.fromstring(entries[worksheet_name][1])
            namespace = worksheet_root.tag[1:].split("}", 1)[0]
            cell = next(
                node
                for node in worksheet_root.iter(f"{{{namespace}}}c")
                if node.attrib.get("r") == "C2"
            )
            for child in list(cell):
                cell.remove(child)
            cell.attrib["t"] = "s"
            ElementTree.SubElement(cell, f"{{{namespace}}}v").text = "0"
            entries[worksheet_name] = (
                entries[worksheet_name][0],
                ElementTree.tostring(
                    worksheet_root,
                    encoding="utf-8",
                    xml_declaration=True,
                ),
            )
            shared_strings = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<sst xmlns="http://schemas.openxmlformats.org/'
                'spreadsheetml/2006/main" count="1" uniqueCount="1">'
                '<si><t>Bonjour partagé</t></si></sst>'
            ).encode("utf-8")
            with ZipFile(batch_path, "w", compression=ZIP_DEFLATED) as archive:
                for _, (item, data) in entries.items():
                    archive.writestr(item, data)
                archive.writestr("xl/sharedStrings.xml", shared_strings)

            restored_path = Path(tmp_dir) / "shared-restored.xlsx"
            restore_batches(summary.output_dir, output_file=restored_path)

            restored = load_workbook(restored_path)
            try:
                self.assertEqual(
                    restored["Strings"]["C2"].value,
                    "Bonjour partagé",
                )
            finally:
                restored.close()

    def test_restore_expands_reused_shared_formula_groups_per_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            self.create_workbook(source_path, data_rows=4)
            summary = split_workbook(source_path, batch_size=2)

            for batch_index, batch_path in enumerate(summary.batch_files):
                with ZipFile(batch_path) as archive:
                    worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
                root = ElementTree.fromstring(worksheet_xml)
                namespace = root.tag[1:].split("}", 1)[0]
                cells = {
                    node.attrib.get("r"): node
                    for node in root.iter(f"{{{namespace}}}c")
                }
                master_formula = next(
                    child
                    for child in cells["C2"]
                    if child.tag == f"{{{namespace}}}f"
                )
                master_formula.attrib.update(
                    {"t": "shared", "si": "0", "ref": "C2:C3"}
                )
                master_formula.text = f"B{2 + batch_index * 2}"
                dependent_formula = next(
                    child
                    for child in cells["C3"]
                    if child.tag == f"{{{namespace}}}f"
                )
                dependent_formula.attrib.update({"t": "shared", "si": "0"})
                dependent_formula.attrib.pop("ref", None)
                dependent_formula.text = None
                self.replace_zip_entry(
                    batch_path,
                    "xl/worksheets/sheet1.xml",
                    ElementTree.tostring(
                        root,
                        encoding="utf-8",
                        xml_declaration=True,
                    ),
                )

            restored_path = Path(tmp_dir) / "shared-formulas.xlsx"
            restore_batches(summary.output_dir, output_file=restored_path)

            restored = load_workbook(restored_path, data_only=False)
            try:
                self.assertEqual(
                    [restored["Strings"].cell(row, 3).value for row in range(2, 6)],
                    ["=B2", "=B3", "=B4", "=B5"],
                )
            finally:
                restored.close()
            with ZipFile(restored_path) as archive:
                restored_root = ElementTree.fromstring(
                    archive.read("xl/worksheets/sheet1.xml")
                )
            formulas = [
                node
                for node in restored_root.iter()
                if node.tag.rsplit("}", 1)[-1] == "f"
            ]
            self.assertTrue(formulas)
            self.assertTrue(all(node.attrib.get("t") != "shared" for node in formulas))

    def test_restore_rejects_array_formula_batches(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            self.create_workbook(source_path, data_rows=2)
            summary = split_workbook(source_path, batch_size=2)
            batch_path = summary.batch_files[0]
            with ZipFile(batch_path) as archive:
                worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
            root = ElementTree.fromstring(worksheet_xml)
            namespace = root.tag[1:].split("}", 1)[0]
            formula = next(
                child
                for cell in root.iter(f"{{{namespace}}}c")
                if cell.attrib.get("r") == "C2"
                for child in cell
                if child.tag == f"{{{namespace}}}f"
            )
            formula.attrib.update({"t": "array", "ref": "C2:C3"})
            self.replace_zip_entry(
                batch_path,
                "xl/worksheets/sheet1.xml",
                ElementTree.tostring(
                    root,
                    encoding="utf-8",
                    xml_declaration=True,
                ),
            )

            with self.assertRaisesRegex(ValueError, "数组或数据表公式"):
                restore_batches(summary.output_dir)

    def test_restore_preserves_ooxml_namespace_prefixes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "namespaces.xlsx"
            self.create_workbook(source_path, data_rows=2)
            with ZipFile(source_path) as archive:
                worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
            namespace_attributes = (
                b' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
                b' xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"'
                b' xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"'
                b' mc:Ignorable="x14ac"'
            )
            worksheet_xml = worksheet_xml.replace(
                b"<worksheet",
                b"<worksheet" + namespace_attributes,
                1,
            )
            alternate_content = (
                b'<mc:AlternateContent><mc:Choice Requires="x14">'
                b'<x14:conditionalFormattings/></mc:Choice>'
                b'<mc:Fallback/></mc:AlternateContent>'
            )
            worksheet_xml = worksheet_xml.replace(
                b"</worksheet>",
                alternate_content + b"</worksheet>",
                1,
            )
            self.replace_zip_entry(
                source_path,
                "xl/worksheets/sheet1.xml",
                worksheet_xml,
            )

            summary = split_workbook(source_path, batch_size=2)
            restored_path = Path(tmp_dir) / "namespace-restored.xlsx"
            restore_batches(summary.output_dir, output_file=restored_path)

            with ZipFile(restored_path) as archive:
                restored_xml = archive.read("xl/worksheets/sheet1.xml")
            self.assertIn(b'xmlns:mc="', restored_xml)
            self.assertIn(b'xmlns:x14="', restored_xml)
            self.assertIn(b'xmlns:x14ac="', restored_xml)
            self.assertIn(b'mc:Ignorable="x14ac"', restored_xml)
            self.assertIn(b'mc:Choice Requires="x14"', restored_xml)
            ElementTree.fromstring(restored_xml)

    def test_split_rejects_nonempty_output_directory(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            output_dir = Path(tmp_dir) / "batches"
            output_dir.mkdir()
            (output_dir / "keep.txt").write_text("keep", encoding="utf-8")
            self.create_workbook(source_path)

            with self.assertRaisesRegex(ValueError, "必须为空"):
                split_workbook(source_path, output_dir=output_dir)

            self.assertEqual((output_dir / "keep.txt").read_text(encoding="utf-8"), "keep")

    def test_restore_rejects_modified_source_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            self.create_workbook(source_path)
            summary = split_workbook(source_path, batch_size=2)
            manifest = json.loads(summary.manifest_path.read_text(encoding="utf-8"))
            snapshot_path = summary.output_dir / manifest["source_snapshot"]
            snapshot_path.write_bytes(snapshot_path.read_bytes() + b"tampered")

            with self.assertRaisesRegex(ValueError, "复原模板已被修改"):
                restore_batches(summary.output_dir)

    def test_split_validates_sheet_sizes_and_header_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = Path(tmp_dir) / "project.xlsx"
            self.create_workbook(source_path, data_rows=2)

            with self.assertRaisesRegex(ValueError, "行数必须大于 0"):
                split_workbook(source_path, batch_size=0)
            with self.assertRaisesRegex(ValueError, "表头行数不能小于 0"):
                split_workbook(source_path, header_rows=-1)
            with self.assertRaisesRegex(ValueError, "工作表不存在"):
                split_workbook(source_path, sheet="Missing")
            with self.assertRaisesRegex(ValueError, "没有可拆分"):
                split_workbook(source_path, header_rows=3)


if __name__ == "__main__":
    unittest.main()
