from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.source_consistency_checker.check_source_consistency import (
    PROBLEM_SHEET_NAME,
    process_excel,
)


class SourceConsistencyExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["A2"] = "Hello"
        worksheet["B2"] = "Bonjour"
        worksheet["A3"] = "Hello"
        worksheet["B3"] = "Salut"
        worksheet["A4"] = "Same"
        worksheet["B4"] = "Identique"
        worksheet["A5"] = "Same"
        worksheet["B5"] = "Identique"
        worksheet["A6"] = None
        worksheet["B6"] = "Ignored blank source"
        worksheet["A7"] = "   "
        worksheet["B7"] = "Ignored whitespace source"
        worksheet["A8"] = "Hello "
        worksheet["B8"] = "Trailing-space source is distinct"
        worksheet["A9"] = "Empty target"
        worksheet["B9"] = None
        worksheet["A10"] = "Empty target"
        worksheet["B10"] = "Translation"
        workbook.save(path)

    def test_process_excel_lists_every_row_in_inconsistent_source_groups(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_file=input_path,
                source_column="a",
                target_column="b",
                sheet="Data",
                start_row=2,
            )

            self.assertEqual(
                summary.output_path,
                input_path.with_name("source_consistency_check_input.xlsx").resolve(),
            )
            self.assertEqual(summary.total_rows_checked, 9)
            self.assertEqual(summary.non_empty_source_rows, 7)
            self.assertEqual(summary.repeated_source_count, 3)
            self.assertEqual(summary.inconsistent_source_count, 2)
            self.assertEqual(summary.problem_rows, 4)

            workbook = load_workbook(summary.output_path)
            problem_sheet = workbook[PROBLEM_SHEET_NAME]
            self.assertEqual(
                [problem_sheet.cell(1, column).value for column in range(1, 6)],
                ["source文本", "target版本数", "行号", "target文本", "同组行号"],
            )
            self.assertEqual(problem_sheet.max_row, 5)
            self.assertEqual(
                [problem_sheet[f"C{row}"].value for row in range(2, 6)],
                [2, 3, 9, 10],
            )
            self.assertEqual(problem_sheet["A2"].value, "Hello")
            self.assertEqual(problem_sheet["B2"].value, 2)
            self.assertEqual(problem_sheet["E2"].value, "2、3")
            self.assertEqual(problem_sheet["A4"].value, "Empty target")
            self.assertIsNone(problem_sheet["D4"].value)
            self.assertEqual(problem_sheet["E4"].value, "9、10")
            self.assertNotIn(PROBLEM_SHEET_NAME, load_workbook(input_path).sheetnames)

    def test_process_excel_rebuilds_stale_problem_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            workbook = load_workbook(input_path)
            stale_sheet = workbook.create_sheet(PROBLEM_SHEET_NAME)
            stale_sheet["A1"] = "stale"
            workbook.save(input_path)

            summary = process_excel(input_path, "A", "B", sheet="Data")

            problem_sheet = load_workbook(summary.output_path)[PROBLEM_SHEET_NAME]
            self.assertEqual(problem_sheet["A1"].value, "source文本")

    def test_process_excel_rejects_invalid_start_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)
            with self.assertRaisesRegex(ValueError, "开始行必须大于等于 1"):
                process_excel(input_path, "A", "B", start_row=0)


if __name__ == "__main__":
    unittest.main()
