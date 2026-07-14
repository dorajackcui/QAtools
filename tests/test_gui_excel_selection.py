from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook

from tools.chinese_target_checker.check_chinese_target_gui import ChineseTargetCheckerApp
from tools.excel_line_splitter.split_excel_lines_gui import SplitExcelLinesApp
from tools.french_nbsp_restorer.restore_french_nbsp_gui import FrenchNbspRestorerApp
from tools.line_break_checker.check_line_breaks_gui import LineBreakCheckerApp
from tools.source_consistency_checker.check_source_consistency_gui import SourceConsistencyCheckerApp
from tools.tag_placeholder_checker.check_tags_and_placeholders_gui import TagPlaceholderCheckerApp
from tools.term_pair_checker.extract_terms_gui import ExtractTermsApp
from tools.workflow.workflow_gui import WorkflowRunnerApp
from tools.xbench_report_transformer.transform_xbench_report_gui import XbenchReportTransformerApp


class FakeVar:
    def __init__(self, value: str = "") -> None:
        self.value = value

    def get(self) -> str:
        return self.value

    def set(self, value: str) -> None:
        self.value = value


class FakeBoolVar:
    def __init__(self, value: bool = False) -> None:
        self.value = value

    def get(self) -> bool:
        return self.value

    def set(self, value: bool) -> None:
        self.value = value


class FakeCombobox(dict):
    def __init__(self) -> None:
        super().__init__()
        self["values"] = ()


class FakeWidget:
    def __init__(self) -> None:
        self.state = "normal"
        self.visible = True

    def configure(self, **kwargs) -> None:
        if "state" in kwargs:
            self.state = kwargs["state"]

    def grid(self) -> None:
        self.visible = True

    def grid_remove(self) -> None:
        self.visible = False


class GuiSheetSelectionTests(unittest.TestCase):
    def create_term_pair_workbook(self, path: Path) -> None:
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Data"
        data_sheet["A1"] = "source"
        data_sheet["B1"] = "target"

        alternate_sheet = workbook.create_sheet("Alternate")
        alternate_sheet["C1"] = " SOURCE "
        alternate_sheet["E1"] = " target "

        missing_header_sheet = workbook.create_sheet("NoHeader")
        missing_header_sheet["A1"] = "foo"
        missing_header_sheet["B1"] = "bar"

        workbook.active = 1
        workbook.save(path)

    def create_history_tb_workbook(self, path: Path) -> None:
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Raw"
        data_sheet["A1"] = "source"
        data_sheet["B1"] = "target"

        term_sheet = workbook.create_sheet("术语表")
        term_sheet["A1"] = "source术语"
        term_sheet["B1"] = "target术语"
        term_sheet["C1"] = "source术语（无mark）"
        term_sheet["D1"] = "target术语（无mark）"

        workbook.active = 0
        workbook.save(path)

    def create_splitter_workbook(self, path: Path) -> None:
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Split"
        workbook.create_sheet("Other")
        workbook.active = 0
        workbook.save(path)

    def create_xbench_report_workbook(self, path: Path) -> None:
        workbook = Workbook()
        report_sheet = workbook.active
        report_sheet.title = "Xbench QA"
        report_sheet["C4"] = "Source"
        report_sheet["D4"] = "Target"
        report_sheet["E4"] = "Comments"
        report_sheet["F4"] = "Metadata"
        workbook.create_sheet("Other")
        workbook.active = 0
        workbook.save(path)

    def create_tag_checker_workbook(self, path: Path) -> None:
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Tags"
        data_sheet["D1"] = "source"
        data_sheet["F1"] = "target"

        archive_sheet = workbook.create_sheet("Archive")
        archive_sheet["A1"] = "source"
        archive_sheet["B1"] = "target"

        workbook.active = 0
        workbook.save(path)

    def build_extract_terms_app(self, input_path: Path) -> ExtractTermsApp:
        app = ExtractTermsApp.__new__(ExtractTermsApp)
        app.input_file_var = FakeVar(str(input_path))
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_var = FakeVar("")
        app.sheet_combobox = FakeCombobox()
        return app

    def build_extract_terms_app_with_history(self, history_path: Path) -> ExtractTermsApp:
        app = ExtractTermsApp.__new__(ExtractTermsApp)
        app.history_tb_file_var = FakeVar(str(history_path))
        app.history_sheet_var = FakeVar("")
        app.history_source_column_var = FakeVar("")
        app.history_target_column_var = FakeVar("")
        app.history_sheet_combobox = FakeCombobox()
        return app

    def build_splitter_app(self, input_path: Path) -> SplitExcelLinesApp:
        app = SplitExcelLinesApp.__new__(SplitExcelLinesApp)
        app.input_file_var = FakeVar(str(input_path))
        app.output_file_var = FakeVar("")
        app.sheet_var = FakeVar("")
        app.sheet_combobox = FakeCombobox()
        app.output_preview_var = FakeVar("")
        return app

    def build_tag_checker_app(self, input_path: Path) -> TagPlaceholderCheckerApp:
        app = TagPlaceholderCheckerApp.__new__(TagPlaceholderCheckerApp)
        app.input_file_var = FakeVar(str(input_path))
        app.output_file_var = FakeVar("")
        app.sheet_var = FakeVar("")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_combobox = FakeCombobox()
        return app

    def build_line_break_checker_app(self, input_path: Path) -> LineBreakCheckerApp:
        app = LineBreakCheckerApp.__new__(LineBreakCheckerApp)
        app.input_file_var = FakeVar(str(input_path))
        app.sheet_var = FakeVar("")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_combobox = FakeCombobox()
        app.output_preview_var = FakeVar("")
        return app

    def build_source_consistency_checker_app(
        self, input_path: Path
    ) -> SourceConsistencyCheckerApp:
        app = SourceConsistencyCheckerApp.__new__(SourceConsistencyCheckerApp)
        app.input_file_var = FakeVar(str(input_path))
        app.sheet_var = FakeVar("")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_combobox = FakeCombobox()
        return app

    def build_french_nbsp_restorer_app(self, input_path: Path) -> FrenchNbspRestorerApp:
        app = FrenchNbspRestorerApp.__new__(FrenchNbspRestorerApp)
        app.input_file_var = FakeVar(str(input_path))
        app.output_file_var = FakeVar("")
        app.sheet_var = FakeVar("")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_combobox = FakeCombobox()
        return app

    def build_chinese_target_checker_app(self, input_path: Path) -> ChineseTargetCheckerApp:
        app = ChineseTargetCheckerApp.__new__(ChineseTargetCheckerApp)
        app.input_file_var = FakeVar(str(input_path))
        app.output_file_var = FakeVar("")
        app.sheet_var = FakeVar("")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_combobox = FakeCombobox()
        app.output_preview_var = FakeVar("")
        return app

    def build_workflow_app(self, input_path: Path) -> WorkflowRunnerApp:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        app.input_file_var = FakeVar(str(input_path))
        app.output_file_var = FakeVar("")
        app.sheet_var = FakeVar("")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_combobox = FakeCombobox()
        app.output_preview_var = FakeVar("")
        return app

    def build_workflow_app_with_history(self, history_path: Path) -> WorkflowRunnerApp:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        app.term_history_tb_file_var = FakeVar(str(history_path))
        app.term_history_sheet_var = FakeVar("")
        app.term_history_source_column_var = FakeVar("")
        app.term_history_target_column_var = FakeVar("")
        app.term_history_sheet_combobox = FakeCombobox()
        return app

    def build_xbench_report_transformer_app(self, input_path: Path) -> XbenchReportTransformerApp:
        app = XbenchReportTransformerApp.__new__(XbenchReportTransformerApp)
        app.input_file_var = FakeVar(str(input_path))
        app.sheet_var = FakeVar("")
        app.sheet_combobox = FakeCombobox()
        return app

    def test_term_pair_refresh_populates_sheet_choices_and_detects_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "terms.xlsx"
            self.create_term_pair_workbook(workbook_path)
            app = self.build_extract_terms_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Data", "Alternate", "NoHeader"))
            self.assertEqual(app.sheet_var.get(), "Alternate")
            self.assertEqual(app.source_column_var.get(), "C")
            self.assertEqual(app.target_column_var.get(), "E")

    def test_term_pair_switching_sheet_redetects_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "terms.xlsx"
            self.create_term_pair_workbook(workbook_path)
            app = self.build_extract_terms_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)
            app.sheet_var.set("Data")
            app.handle_sheet_selected(show_error=False)

            self.assertEqual(app.source_column_var.get(), "A")
            self.assertEqual(app.target_column_var.get(), "B")

    def test_term_pair_missing_detection_keeps_manual_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "terms.xlsx"
            self.create_term_pair_workbook(workbook_path)
            app = self.build_extract_terms_app(workbook_path)
            app.source_column_var.set("X")
            app.target_column_var.set("Y")

            app.sheet_var.set("NoHeader")
            app.handle_sheet_selected(show_error=False)

            self.assertEqual(app.source_column_var.get(), "X")
            self.assertEqual(app.target_column_var.get(), "Y")

    def test_term_pair_history_tb_defaults_to_term_sheet_and_detects_source_target_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            self.create_history_tb_workbook(history_path)
            app = self.build_extract_terms_app_with_history(history_path)

            app.refresh_history_sheet_choices(show_error=False)

            self.assertEqual(app.history_sheet_combobox["values"], ("Raw", "术语表"))
            self.assertEqual(app.history_sheet_var.get(), "术语表")
            self.assertEqual(app.history_source_column_var.get(), "A")
            self.assertEqual(app.history_target_column_var.get(), "B")

    def test_term_pair_history_details_expand_after_selecting_tb(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            self.create_history_tb_workbook(history_path)
            app = ExtractTermsApp.__new__(ExtractTermsApp)
            app.history_tb_file_var = FakeVar("")
            app.history_sheet_var = FakeVar("")
            app.history_source_column_var = FakeVar("")
            app.history_target_column_var = FakeVar("")
            app.history_start_row_var = FakeVar("2")
            app.history_sheet_combobox = FakeCombobox()
            app.history_details_expanded = False
            app.history_details_button_text_var = FakeVar("展开详情")
            app.history_details_button = FakeWidget()
            app.history_details_frame = FakeWidget()

            with patch(
                "tools.term_pair_checker.extract_terms_gui.filedialog.askopenfilename",
                return_value=str(history_path),
            ):
                app.choose_history_tb_file()

            self.assertEqual(app.history_tb_file_var.get(), str(history_path))
            self.assertTrue(app.history_details_expanded)
            self.assertTrue(app.history_details_frame.visible)
            self.assertEqual(app.history_details_button_text_var.get(), "收起详情")
            self.assertEqual(app.history_sheet_var.get(), "术语表")

    def test_term_pair_output_preview_uses_automatic_output_name(self) -> None:
        app = ExtractTermsApp.__new__(ExtractTermsApp)
        app.input_file_var = FakeVar("D:/project/input.xlsx")
        app.output_preview_var = FakeVar("")

        app.update_output_preview()

        self.assertTrue(
            app.output_preview_var.get().endswith("term_pair_check_input.xlsx")
        )

    def test_simple_tool_pages_preview_their_automatic_output_names(self) -> None:
        cases = (
            (SplitExcelLinesApp, "split_lines_input.xlsx"),
            (TagPlaceholderCheckerApp, "tag_check_input.xlsx"),
            (LineBreakCheckerApp, "line_break_check_input.xlsx"),
            (
                SourceConsistencyCheckerApp,
                "source_consistency_check_input.xlsx",
            ),
            (FrenchNbspRestorerApp, "french_nbsp_restore_input.xlsx"),
            (ChineseTargetCheckerApp, "target_chinese_check_input.xlsx"),
            (XbenchReportTransformerApp, "xbench_transform_input.xlsx"),
        )

        for app_type, expected_name in cases:
            with self.subTest(app=app_type.__name__):
                app = app_type.__new__(app_type)
                app.input_file_var = FakeVar("D:/project/input.xlsx")
                app.output_preview_var = FakeVar("")

                app.update_output_preview()

                self.assertTrue(app.output_preview_var.get().endswith(expected_name))

    def test_splitter_refresh_populates_sheet_choices(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "split.xlsx"
            self.create_splitter_workbook(workbook_path)
            app = self.build_splitter_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Split", "Other"))
            self.assertEqual(app.sheet_var.get(), "Split")

    def test_tag_checker_refresh_populates_sheet_choices_and_detects_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "tags.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_tag_checker_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Tags", "Archive"))
            self.assertEqual(app.sheet_var.get(), "Tags")
            self.assertEqual(app.source_column_var.get(), "D")
            self.assertEqual(app.target_column_var.get(), "F")

    def test_tag_checker_switching_sheet_redetects_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "tags.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_tag_checker_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)
            app.sheet_var.set("Archive")
            app.handle_sheet_selected(show_error=False)

            self.assertEqual(app.source_column_var.get(), "A")
            self.assertEqual(app.target_column_var.get(), "B")

    def test_line_break_checker_detects_source_and_target_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "line_breaks.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_line_break_checker_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Tags", "Archive"))
            self.assertEqual(app.sheet_var.get(), "Tags")
            self.assertEqual(app.source_column_var.get(), "D")
            self.assertEqual(app.target_column_var.get(), "F")

    def test_source_consistency_checker_detects_source_and_target_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "source_consistency.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_source_consistency_checker_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Tags", "Archive"))
            self.assertEqual(app.sheet_var.get(), "Tags")
            self.assertEqual(app.source_column_var.get(), "D")
            self.assertEqual(app.target_column_var.get(), "F")

    def test_french_nbsp_restorer_refresh_populates_sheet_choices_and_detects_target_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "nbsp.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_french_nbsp_restorer_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Tags", "Archive"))
            self.assertEqual(app.sheet_var.get(), "Tags")
            self.assertEqual(app.target_column_var.get(), "F")

    def test_chinese_target_checker_refresh_detects_source_and_target_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "chinese.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_chinese_target_checker_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Tags", "Archive"))
            self.assertEqual(app.sheet_var.get(), "Tags")
            self.assertEqual(app.source_column_var.get(), "D")
            self.assertEqual(app.target_column_var.get(), "F")

    def test_chinese_target_checker_choose_input_updates_output_preview(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "chinese.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_chinese_target_checker_app(Path(""))

            with patch(
                "tools.chinese_target_checker.check_chinese_target_gui.filedialog.askopenfilename",
                return_value=str(workbook_path),
            ):
                app.choose_input_file()

            self.assertEqual(app.input_file_var.get(), str(workbook_path))
            self.assertEqual(app.sheet_var.get(), "Tags")
            self.assertTrue(
                app.output_preview_var.get().endswith(
                    "target_chinese_check_chinese.xlsx"
                )
            )

    def test_chinese_target_checker_has_no_problem_sheet_toggle(self) -> None:
        app = ChineseTargetCheckerApp.__new__(ChineseTargetCheckerApp)
        with (
            patch.object(ChineseTargetCheckerApp, "_build_ui", lambda self: None),
            patch("tools.chinese_target_checker.check_chinese_target_gui.ttk.Frame.__init__", lambda self, master=None, padding=None: None),
            patch("tools.chinese_target_checker.check_chinese_target_gui.tk.StringVar", FakeVar),
            patch("tools.chinese_target_checker.check_chinese_target_gui.tk.BooleanVar", FakeBoolVar),
        ):
            ChineseTargetCheckerApp.__init__(app, object())

        self.assertFalse(hasattr(app, "problem_sheet_var"))
        self.assertFalse(hasattr(app, "result_column_var"))

    def test_french_nbsp_restorer_switching_sheet_redetects_target_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "nbsp.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_french_nbsp_restorer_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)
            app.sheet_var.set("Archive")
            app.handle_sheet_selected(show_error=False)

            self.assertEqual(app.target_column_var.get(), "B")

    def test_workflow_refresh_populates_sheet_choices_and_detects_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "workflow.xlsx"
            self.create_tag_checker_workbook(workbook_path)
            app = self.build_workflow_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Tags", "Archive"))
            self.assertEqual(app.sheet_var.get(), "Tags")
            self.assertEqual(app.source_column_var.get(), "D")
            self.assertEqual(app.target_column_var.get(), "F")

    def test_workflow_history_tb_defaults_to_term_sheet_and_detects_source_target_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            self.create_history_tb_workbook(history_path)
            app = self.build_workflow_app_with_history(history_path)

            app.refresh_term_history_sheet_choices(show_error=False)

            self.assertEqual(app.term_history_sheet_combobox["values"], ("Raw", "术语表"))
            self.assertEqual(app.term_history_sheet_var.get(), "术语表")
            self.assertEqual(app.term_history_source_column_var.get(), "A")
            self.assertEqual(app.term_history_target_column_var.get(), "B")

    def test_xbench_report_transformer_refresh_populates_sheet_choices(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "xbench.xlsx"
            self.create_xbench_report_workbook(workbook_path)
            app = self.build_xbench_report_transformer_app(workbook_path)

            app.refresh_sheet_choices(show_error=False)

            self.assertEqual(app.sheet_combobox["values"], ("Xbench QA", "Other"))
            self.assertEqual(app.sheet_var.get(), "Xbench QA")

    def test_xbench_report_transformer_run_passes_selected_sheet_to_processor(self) -> None:
        app = XbenchReportTransformerApp.__new__(XbenchReportTransformerApp)
        app.input_file_var = FakeVar("/tmp/input.xlsx")
        app.sheet_var = FakeVar("Xbench QA")
        summary = SimpleNamespace(
            output_path=Path("/tmp/xbench_transform_input.xlsx"),
            worksheet_title="Xbench QA",
            detail_count=249,
            grouped_count=221,
        )

        with (
            patch(
                "tools.xbench_report_transformer.transform_xbench_report_gui.process_excel",
                return_value=summary,
            ) as process_excel_mock,
            patch("tools.xbench_report_transformer.transform_xbench_report_gui.messagebox.showinfo"),
        ):
            app.run_transform()

        process_excel_mock.assert_called_once_with(
            input_file="/tmp/input.xlsx",
            sheet="Xbench QA",
            output_file=None,
        )

    def test_term_pair_ignores_invalid_history_start_without_history_file(self) -> None:
        app = ExtractTermsApp.__new__(ExtractTermsApp)
        app.input_file_var = FakeVar("/tmp/input.xlsx")
        app.output_file_var = FakeVar("/tmp/output.xlsx")
        app.history_tb_file_var = FakeVar("")
        app.history_sheet_var = FakeVar("术语表")
        app.history_source_column_var = FakeVar("C")
        app.history_target_column_var = FakeVar("D")
        app.history_start_row_var = FakeVar("not-an-int")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_var = FakeVar("Data")
        app.start_row_var = FakeVar("2")
        app.mark_style_vars = {
            "【】": FakeBoolVar(True),
            "[]": FakeBoolVar(True),
        }

        with (
            patch(
                "tools.term_pair_checker.extract_terms_gui.process_excel",
                return_value=("Data", "A", "B", Path("/tmp/output.xlsx"), 3, 0),
            ) as process_excel_mock,
            patch("tools.term_pair_checker.extract_terms_gui.messagebox.showerror") as showerror_mock,
            patch("tools.term_pair_checker.extract_terms_gui.messagebox.showinfo"),
        ):
            app.run_extraction()

        showerror_mock.assert_not_called()
        process_excel_mock.assert_called_once()
        self.assertIsNone(process_excel_mock.call_args.kwargs["history_tb_file"])
        self.assertIsNone(process_excel_mock.call_args.kwargs["history_sheet"])
        self.assertIsNone(process_excel_mock.call_args.kwargs["history_source_column"])
        self.assertIsNone(process_excel_mock.call_args.kwargs["history_target_column"])
        self.assertEqual(process_excel_mock.call_args.kwargs["history_start_row"], 2)

    def test_term_check_allows_history_tb_without_selected_marks(self) -> None:
        app = ExtractTermsApp.__new__(ExtractTermsApp)
        app.input_file_var = FakeVar("/tmp/input.xlsx")
        app.history_tb_file_var = FakeVar("/tmp/history.xlsx")
        app.history_sheet_var = FakeVar("术语表")
        app.history_source_column_var = FakeVar("A")
        app.history_target_column_var = FakeVar("B")
        app.history_start_row_var = FakeVar("2")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_var = FakeVar("Data")
        app.start_row_var = FakeVar("2")
        app.mark_style_vars = {
            "【】": FakeBoolVar(False),
            "[]": FakeBoolVar(False),
        }

        with (
            patch(
                "tools.term_pair_checker.extract_terms_gui.process_excel",
                return_value=("Data", "A", "B", Path("/tmp/output.xlsx"), 2, 0),
            ) as process_excel_mock,
            patch("tools.term_pair_checker.extract_terms_gui.messagebox.showerror") as showerror_mock,
            patch("tools.term_pair_checker.extract_terms_gui.messagebox.showinfo"),
        ):
            app.run_extraction()

        showerror_mock.assert_not_called()
        process_excel_mock.assert_called_once()
        self.assertEqual(process_excel_mock.call_args.kwargs["mark_styles"], [])
        self.assertEqual(
            process_excel_mock.call_args.kwargs["history_tb_file"],
            "/tmp/history.xlsx",
        )

    def test_workflow_ignores_invalid_history_start_without_history_file(self) -> None:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        app.input_file_var = FakeVar("/tmp/input.xlsx")
        app.output_file_var = FakeVar("/tmp/output.xlsx")
        app.term_history_tb_file_var = FakeVar("")
        app.term_history_sheet_var = FakeVar("术语表")
        app.term_history_source_column_var = FakeVar("C")
        app.term_history_target_column_var = FakeVar("D")
        app.term_history_start_row_var = FakeVar("not-an-int")
        app.source_column_var = FakeVar("A")
        app.target_column_var = FakeVar("B")
        app.sheet_var = FakeVar("Data")
        app.start_row_var = FakeVar("2")
        app.run_term_pair_var = FakeBoolVar(True)
        app.run_tag_check_var = FakeBoolVar(True)
        app.run_line_break_check_var = FakeBoolVar(True)
        app.run_source_consistency_check_var = FakeBoolVar(True)
        app.run_chinese_target_check_var = FakeBoolVar(True)
        app.term_mark_style_vars = {
            "【】": FakeBoolVar(True),
            "[]": FakeBoolVar(True),
        }
        app.angle_var = FakeBoolVar(True)
        app.square_color_var = FakeBoolVar(True)
        app.brace_var = FakeBoolVar(True)
        app.newline_var = FakeBoolVar(True)
        app.tag_mode_var = FakeVar("standard")
        summary = SimpleNamespace(
            output_path=Path("/tmp/output.xlsx"),
            worksheet_title="Data",
            source_column="A",
            target_column="B",
            start_row=2,
            ran_term_pair_check=True,
            ran_tag_check=True,
            ran_line_break_check=True,
            ran_source_consistency_check=True,
            ran_chinese_target_check=True,
            term_count=3,
            term_problem_count=0,
            term_problem_rows=0,
            tag_problem_count=0,
            tag_problem_rows=0,
            line_break_problem_count=0,
            source_consistency_problem_count=0,
            source_consistency_problem_rows=0,
            chinese_target_problem_count=0,
        )

        with (
            patch("tools.workflow.workflow_gui.run_workflow", return_value=summary) as run_workflow_mock,
            patch("tools.workflow.workflow_gui.messagebox.showerror") as showerror_mock,
            patch("tools.workflow.workflow_gui.messagebox.showinfo"),
        ):
            app.run_selected_tasks()

        showerror_mock.assert_not_called()
        run_workflow_mock.assert_called_once()
        self.assertIsNone(run_workflow_mock.call_args.kwargs["term_history_tb_file"])
        self.assertIsNone(run_workflow_mock.call_args.kwargs["term_history_sheet"])
        self.assertIsNone(run_workflow_mock.call_args.kwargs["term_history_source_column"])
        self.assertIsNone(run_workflow_mock.call_args.kwargs["term_history_target_column"])
        self.assertEqual(run_workflow_mock.call_args.kwargs["term_history_start_row"], 2)
        self.assertTrue(run_workflow_mock.call_args.kwargs["run_line_break_check"])
        self.assertTrue(run_workflow_mock.call_args.kwargs["run_source_consistency_check"])
        self.assertTrue(run_workflow_mock.call_args.kwargs["run_chinese_target_check"])

    def test_workflow_defaults_to_all_quality_checks(self) -> None:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        with (
            patch.object(WorkflowRunnerApp, "_build_ui", lambda self: None),
            patch(
                "tools.workflow.workflow_gui.ttk.Frame.__init__",
                lambda self, master=None, padding=None: None,
            ),
            patch("tools.workflow.workflow_gui.tk.StringVar", FakeVar),
            patch("tools.workflow.workflow_gui.tk.BooleanVar", FakeBoolVar),
        ):
            WorkflowRunnerApp.__init__(app, object())

        self.assertTrue(app.run_term_pair_var.get())
        self.assertTrue(app.run_tag_check_var.get())
        self.assertTrue(app.run_line_break_check_var.get())
        self.assertTrue(app.run_source_consistency_check_var.get())
        self.assertTrue(app.run_chinese_target_check_var.get())
        self.assertFalse(app.term_settings_expanded)
        self.assertFalse(app.tag_settings_expanded)

    def test_workflow_select_all_and_clear_all_keep_settings_collapsed(self) -> None:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        app.run_term_pair_var = FakeBoolVar(True)
        app.run_tag_check_var = FakeBoolVar(True)
        app.run_line_break_check_var = FakeBoolVar(True)
        app.run_source_consistency_check_var = FakeBoolVar(True)
        app.run_chinese_target_check_var = FakeBoolVar(True)
        app.term_settings_expanded = True
        app.tag_settings_expanded = True
        app.term_settings_button_text_var = FakeVar("收起设置")
        app.tag_settings_button_text_var = FakeVar("收起设置")
        app.term_settings_button = FakeWidget()
        app.tag_settings_button = FakeWidget()
        app.term_settings_frame = FakeWidget()
        app.tag_settings_frame = FakeWidget()
        app.tag_mode_var = FakeVar("standard")
        app.standard_tag_checkbuttons = []

        app.clear_all_tasks()

        self.assertFalse(any(variable.get() for variable in app.task_vars()))
        self.assertFalse(app.term_settings_expanded)
        self.assertFalse(app.tag_settings_expanded)
        self.assertEqual(app.term_settings_button.state, "disabled")
        self.assertEqual(app.tag_settings_button.state, "disabled")

        app.select_all_tasks()

        self.assertTrue(all(variable.get() for variable in app.task_vars()))
        self.assertFalse(app.term_settings_expanded)
        self.assertFalse(app.tag_settings_expanded)
        self.assertEqual(app.term_settings_button.state, "normal")
        self.assertEqual(app.tag_settings_button.state, "normal")

    def test_workflow_output_preview_uses_automatic_output_name(self) -> None:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        app.input_file_var = FakeVar("D:/project/input.xlsx")
        app.output_preview_var = FakeVar("")

        app.update_output_preview()

        self.assertTrue(
            app.output_preview_var.get().endswith("workflow_check_input.xlsx")
        )

    def test_workflow_only_expands_one_settings_panel_at_a_time(self) -> None:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        app.run_term_pair_var = FakeBoolVar(True)
        app.run_tag_check_var = FakeBoolVar(True)
        app.term_settings_expanded = True
        app.tag_settings_expanded = False
        app.term_settings_button_text_var = FakeVar("收起设置")
        app.tag_settings_button_text_var = FakeVar("展开设置")
        app.term_settings_frame = FakeWidget()
        app.tag_settings_frame = FakeWidget()

        app.toggle_tag_settings()

        self.assertFalse(app.term_settings_expanded)
        self.assertTrue(app.tag_settings_expanded)
        self.assertFalse(app.term_settings_frame.visible)
        self.assertTrue(app.tag_settings_frame.visible)

    def test_term_pair_gui_defaults_to_square_and_book_title_marks(self) -> None:
        app = ExtractTermsApp.__new__(ExtractTermsApp)
        with (
            patch.object(ExtractTermsApp, "_build_ui", lambda self: None),
            patch("tools.term_pair_checker.extract_terms_gui.ttk.Frame.__init__", lambda self, master=None, padding=None: None),
            patch("tools.term_pair_checker.extract_terms_gui.tk.StringVar", FakeVar),
            patch("tools.term_pair_checker.extract_terms_gui.tk.BooleanVar", FakeBoolVar),
        ):
            ExtractTermsApp.__init__(app, object())

        self.assertTrue(app.mark_style_vars["【】"].get())
        self.assertTrue(app.mark_style_vars["[]"].get())
        self.assertNotIn("<>", app.mark_style_vars)
        self.assertFalse(app.history_details_expanded)

    def test_tag_placeholder_gui_defaults_to_standard_tags_not_memoq(self) -> None:
        app = TagPlaceholderCheckerApp.__new__(TagPlaceholderCheckerApp)
        with (
            patch.object(TagPlaceholderCheckerApp, "_build_ui", lambda self: None),
            patch("tools.tag_placeholder_checker.check_tags_and_placeholders_gui.ttk.Frame.__init__", lambda self, master=None, padding=None: None),
            patch("tools.tag_placeholder_checker.check_tags_and_placeholders_gui.tk.StringVar", FakeVar),
            patch("tools.tag_placeholder_checker.check_tags_and_placeholders_gui.tk.BooleanVar", FakeBoolVar),
        ):
            TagPlaceholderCheckerApp.__init__(app, object())

        self.assertTrue(app.angle_var.get())
        self.assertTrue(app.square_color_var.get())
        self.assertTrue(app.brace_var.get())
        self.assertTrue(app.newline_var.get())
        self.assertEqual(app.tag_mode_var.get(), "standard")
        self.assertEqual(
            app.get_selected_token_types(),
            ("angle", "square_color", "brace", "newline"),
        )

    def test_tag_placeholder_gui_makes_memoq_mutually_exclusive_with_standard_tags(self) -> None:
        app = TagPlaceholderCheckerApp.__new__(TagPlaceholderCheckerApp)
        app.angle_var = FakeBoolVar(True)
        app.square_color_var = FakeBoolVar(True)
        app.brace_var = FakeBoolVar(True)
        app.newline_var = FakeBoolVar(True)
        app.tag_mode_var = FakeVar("memoq")
        app.standard_tag_checkbuttons = [FakeWidget(), FakeWidget()]

        app.handle_tag_mode_changed()

        self.assertEqual(app.get_selected_token_types(), ("memoq",))
        self.assertTrue(
            all(widget.state == "disabled" for widget in app.standard_tag_checkbuttons)
        )

        app.tag_mode_var.set("standard")
        app.handle_tag_mode_changed()

        self.assertEqual(
            app.get_selected_token_types(),
            ("angle", "square_color", "brace", "newline"),
        )
        self.assertTrue(
            all(widget.state == "normal" for widget in app.standard_tag_checkbuttons)
        )

    def test_workflow_gui_defaults_to_square_and_book_title_term_marks(self) -> None:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        with (
            patch.object(WorkflowRunnerApp, "_build_ui", lambda self: None),
            patch("tools.workflow.workflow_gui.ttk.Frame.__init__", lambda self, master=None, padding=None: None),
            patch("tools.workflow.workflow_gui.tk.StringVar", FakeVar),
            patch("tools.workflow.workflow_gui.tk.BooleanVar", FakeBoolVar),
        ):
            WorkflowRunnerApp.__init__(app, object())

        self.assertTrue(app.term_mark_style_vars["【】"].get())
        self.assertTrue(app.term_mark_style_vars["[]"].get())
        self.assertNotIn("<>", app.term_mark_style_vars)

    def test_workflow_gui_defaults_to_standard_tags_not_memoq(self) -> None:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        with (
            patch.object(WorkflowRunnerApp, "_build_ui", lambda self: None),
            patch("tools.workflow.workflow_gui.ttk.Frame.__init__", lambda self, master=None, padding=None: None),
            patch("tools.workflow.workflow_gui.tk.StringVar", FakeVar),
            patch("tools.workflow.workflow_gui.tk.BooleanVar", FakeBoolVar),
        ):
            WorkflowRunnerApp.__init__(app, object())

        self.assertTrue(app.angle_var.get())
        self.assertTrue(app.square_color_var.get())
        self.assertTrue(app.brace_var.get())
        self.assertTrue(app.newline_var.get())
        self.assertEqual(app.tag_mode_var.get(), "standard")
        self.assertEqual(
            app.get_selected_tag_token_types(),
            ("angle", "square_color", "brace", "newline"),
        )

    def test_workflow_gui_tag_mode_selects_memoq_or_standard_tags(self) -> None:
        app = WorkflowRunnerApp.__new__(WorkflowRunnerApp)
        app.angle_var = FakeBoolVar(True)
        app.square_color_var = FakeBoolVar(True)
        app.brace_var = FakeBoolVar(True)
        app.newline_var = FakeBoolVar(True)
        app.tag_mode_var = FakeVar("memoq")

        self.assertEqual(app.get_selected_tag_token_types(), ("memoq",))

        app.tag_mode_var.set("standard")
        self.assertEqual(
            app.get_selected_tag_token_types(),
            ("angle", "square_color", "brace", "newline"),
        )


if __name__ == "__main__":
    unittest.main()
