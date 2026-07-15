from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.chinese_target_checker.check_chinese_target import (
    PROBLEM_SHEET_NAME as CHINESE_PROBLEM_SHEET_NAME,
)
from tools.workflow.workflow_runner import (
    WORKFLOW_SUMMARY_SHEET_NAME,
    WORKFLOW_TERM_PROBLEM_SHEET_NAME,
    count_unique_problem_rows,
    run_workflow,
)
from tools.workflow.review_sheet import (
    WORKFLOW_REVIEW_SHEET_NAME,
    collect_review_rows,
    read_review_metadata,
)
from tools.workflow.revision_applier import (
    apply_workflow_revisions,
    build_default_revised_output_path,
)


class WorkflowRunnerTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["A2"] = "第一行 [Alpha] 和 <color=red>{name}"
        worksheet["B2"] = "第一行 [阿尔法] 和 <color=red>{name}"
        worksheet["A3"] = "第二行\n复用 [Alpha] 和 <color=red>{name}"
        worksheet["B3"] = "第二行复用 [错误阿尔法] 和 {name}"
        worksheet["A4"] = "Same source"
        worksheet["B4"] = "译文一"
        worksheet["A5"] = "Same source"
        worksheet["B5"] = "译文二"
        worksheet["C1"] = "note"
        worksheet["C3"] = "keep me"
        workbook.save(path)

    def test_run_workflow_writes_all_quality_check_results_into_same_output_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "workflow_output.xlsx"
            self.create_workbook(input_path)

            summary = run_workflow(
                input_file=input_path,
                output_file=output_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                run_term_pair_check=True,
                term_mark_styles=("[]",),
                run_tag_check=True,
                tag_token_types=("angle", "brace"),
                run_line_break_check=True,
                run_source_consistency_check=True,
                run_chinese_target_check=True,
            )

            self.assertEqual(summary.output_path, output_path.resolve())
            self.assertTrue(summary.ran_term_pair_check)
            self.assertTrue(summary.ran_tag_check)
            self.assertTrue(summary.ran_line_break_check)
            self.assertTrue(summary.ran_source_consistency_check)
            self.assertTrue(summary.ran_chinese_target_check)
            self.assertEqual(summary.term_problem_count, 1)
            self.assertEqual(summary.term_problem_rows, 1)
            self.assertEqual(summary.tag_problem_count, 1)
            self.assertEqual(summary.tag_problem_rows, 1)
            self.assertEqual(summary.line_break_problem_count, 1)
            self.assertEqual(summary.source_consistency_problem_count, 1)
            self.assertEqual(summary.source_consistency_problem_rows, 2)
            self.assertEqual(summary.chinese_target_problem_count, 4)

            workbook = load_workbook(summary.output_path)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Data",
                    "术语表",
                    WORKFLOW_REVIEW_SHEET_NAME,
                    WORKFLOW_SUMMARY_SHEET_NAME,
                ],
            )
            self.assertNotIn("问题列", workbook.sheetnames)
            self.assertNotIn("检查汇总", workbook.sheetnames)
            self.assertNotIn(WORKFLOW_TERM_PROBLEM_SHEET_NAME, workbook.sheetnames)
            self.assertNotIn("标签占位问题", workbook.sheetnames)
            self.assertNotIn("换行数量问题", workbook.sheetnames)
            self.assertNotIn("同源译文不一致", workbook.sheetnames)
            self.assertNotIn(CHINESE_PROBLEM_SHEET_NAME, workbook.sheetnames)
            self.assertEqual(workbook["Data"]["C1"].value, "note")
            self.assertEqual(workbook["Data"]["C3"].value, "keep me")
            summary_sheet = workbook[WORKFLOW_SUMMARY_SHEET_NAME]
            self.assertEqual(
                list(summary_sheet.values),
                [
                    ("检查项", "问题行数"),
                    ("术语检查", 1),
                    ("Tag 检查", 1),
                    ("换行数量检查", 1),
                    ("同源译文一致性", 2),
                    ("Target 中文检查", 4),
                ],
            )
            review_sheet = workbook[WORKFLOW_REVIEW_SHEET_NAME]
            self.assertEqual(
                [review_sheet.cell(1, column).value for column in range(1, 7)],
                [
                    "行号",
                    "source",
                    "target",
                    "修改后target",
                    "问题描述",
                    "检查项",
                ],
            )
            self.assertEqual(
                [review_sheet.cell(row, 1).value for row in range(2, 6)],
                [2, 3, 4, 5],
            )
            merged_check_items = [
                item
                for row in range(2, 6)
                for item in review_sheet.cell(row, 6).value.split("；")
            ]
            self.assertEqual(merged_check_items.count("术语检查"), 1)
            self.assertEqual(merged_check_items.count("Tag 检查"), 1)
            self.assertEqual(merged_check_items.count("换行数量检查"), 1)
            self.assertEqual(merged_check_items.count("同源译文一致性"), 2)
            self.assertEqual(merged_check_items.count("Target 中文检查"), 4)
            self.assertEqual(review_sheet["A3"].value, 3)
            self.assertIsNone(review_sheet["D3"].value)
            self.assertIn("术语检查", review_sheet["F3"].value)
            self.assertIn("Tag 检查", review_sheet["F3"].value)
            self.assertIn("换行数量检查", review_sheet["F3"].value)
            self.assertIn("Target 中文检查", review_sheet["F3"].value)
            row_three_description = review_sheet["E3"].value
            self.assertIn("source术语：Alpha", row_three_description)
            self.assertIn("预期target术语：阿尔法", row_three_description)
            self.assertIn("术语来源：本批次新增", row_three_description)
            self.assertIn("问题类型：尖括号tag不一致", row_three_description)
            self.assertIn("source换行数：1", row_three_description)
            self.assertIn("target换行数：0", row_three_description)
            self.assertIn("数量差：-1", row_three_description)
            self.assertIn("命中字符：", row_three_description)
            row_four_description = review_sheet["E4"].value
            self.assertIn("target版本数：2", row_four_description)
            self.assertIn("同组行号：4、5", row_four_description)
            self.assertEqual(review_sheet["A3"].hyperlink.location, "'Data'!B3")
            self.assertIsNone(review_sheet["A3"].hyperlink.target)
            self.assertTrue(review_sheet.column_dimensions["G"].hidden)
            self.assertTrue(review_sheet.column_dimensions["H"].hidden)
            self.assertEqual(len(review_sheet.data_validations.dataValidation), 0)
            metadata = read_review_metadata(review_sheet)
            self.assertEqual(metadata["data_sheet_name"], "Data")
            self.assertEqual(metadata["source_column"], "A")
            self.assertEqual(metadata["target_column"], "B")
            self.assertEqual(metadata["remove_term_helper"], "0")

    def test_apply_workflow_revisions_writes_targets_and_removes_qa_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            report_path = Path(tmp_dir) / "workflow_output.xlsx"
            revised_path = Path(tmp_dir) / "revised_input.xlsx"
            self.create_workbook(input_path)
            run_workflow(
                input_file=input_path,
                output_file=report_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                term_mark_styles=("[]",),
                tag_token_types=("angle", "brace"),
            )

            report_workbook = load_workbook(report_path)
            review_sheet = report_workbook[WORKFLOW_REVIEW_SHEET_NAME]
            review_sheet["D2"] = "第一行修订"
            review_sheet["D3"] = "第二行修订"
            report_workbook.save(report_path)

            summary = apply_workflow_revisions(report_path, output_file=revised_path)

            self.assertEqual(summary.revised_count, 2)
            self.assertEqual(summary.ignored_count, 2)
            self.assertEqual(summary.unchanged_count, 0)
            self.assertEqual(summary.conflict_rows, ())
            revised_workbook = load_workbook(revised_path)
            self.assertEqual(revised_workbook.sheetnames, ["Data"])
            data_sheet = revised_workbook["Data"]
            self.assertEqual(data_sheet["B2"].value, "第一行修订")
            self.assertEqual(data_sheet["B3"].value, "第二行修订")
            self.assertEqual(data_sheet["B4"].value, "译文一")
            self.assertEqual(data_sheet["B5"].value, "译文二")
            self.assertEqual(data_sheet["C1"].value, "note")
            self.assertEqual(data_sheet["C3"].value, "keep me")

    def test_apply_workflow_revisions_skips_target_conflicts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            report_path = Path(tmp_dir) / "workflow_output.xlsx"
            revised_path = Path(tmp_dir) / "revised_input.xlsx"
            self.create_workbook(input_path)
            run_workflow(
                input_file=input_path,
                output_file=report_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                term_mark_styles=("[]",),
                tag_token_types=("angle", "brace"),
            )

            report_workbook = load_workbook(report_path)
            report_workbook["Data"]["B2"] = "人工直接修改"
            report_workbook[WORKFLOW_REVIEW_SHEET_NAME]["D2"] = "准备回填的修改"
            report_workbook.save(report_path)

            summary = apply_workflow_revisions(report_path, output_file=revised_path)

            self.assertEqual(summary.revised_count, 0)
            self.assertEqual(summary.conflict_rows, (2,))
            revised_workbook = load_workbook(revised_path)
            self.assertEqual(revised_workbook["Data"]["B2"].value, "人工直接修改")

    def test_default_revised_output_path_uses_original_input_name(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            report_path = Path(tmp_dir) / "workflow_output.xlsx"
            self.create_workbook(input_path)
            run_workflow(
                input_file=input_path,
                output_file=report_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                run_term_pair_check=False,
                run_tag_check=False,
                run_line_break_check=False,
                run_source_consistency_check=False,
                run_chinese_target_check=True,
            )

            self.assertEqual(
                build_default_revised_output_path(report_path),
                Path(tmp_dir) / "revised_input.xlsx",
            )

    def test_apply_revisions_preserves_unrelated_sheets_from_disabled_checks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            report_path = Path(tmp_dir) / "workflow_output.xlsx"
            revised_path = Path(tmp_dir) / "revised_input.xlsx"
            self.create_workbook(input_path)
            input_workbook = load_workbook(input_path)
            legitimate_term_sheet = input_workbook.create_sheet("术语表")
            legitimate_term_sheet["A1"] = "用户原有内容"
            input_workbook.save(input_path)
            run_workflow(
                input_file=input_path,
                output_file=report_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                run_term_pair_check=False,
                run_tag_check=False,
                run_line_break_check=False,
                run_source_consistency_check=False,
                run_chinese_target_check=True,
            )

            report_workbook = load_workbook(report_path)
            report_workbook[WORKFLOW_REVIEW_SHEET_NAME]["D2"] = "第一行修订"
            report_workbook.save(report_path)
            apply_workflow_revisions(report_path, output_file=revised_path)

            revised_workbook = load_workbook(revised_path)
            self.assertIn("术语表", revised_workbook.sheetnames)
            self.assertEqual(revised_workbook["术语表"]["A1"].value, "用户原有内容")

    def test_count_unique_problem_rows_deduplicates_source_rows(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["问题行号"])
        worksheet.append([3])
        worksheet.append([3])
        worksheet.append([5])

        self.assertEqual(count_unique_problem_rows(worksheet), 2)

    def test_review_merge_rejects_incompatible_problem_sheet_schema(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "旧问题表"
        worksheet.append(["问题行号", "source", "target", "描述"])
        worksheet.append([2, "Source", "Target", "Problem"])

        with self.assertRaisesRegex(ValueError, "前四列必须为"):
            collect_review_rows(workbook, (("测试检查", "旧问题表"),))

    def test_run_workflow_requires_at_least_one_task(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            with self.assertRaisesRegex(ValueError, "请至少选择一个质量检查项目"):
                run_workflow(
                    input_file=input_path,
                    source_column="A",
                    target_column="B",
                    run_term_pair_check=False,
                    run_tag_check=False,
                    run_line_break_check=False,
                    run_source_consistency_check=False,
                    run_chinese_target_check=False,
                )

    def test_run_workflow_passes_history_tb_to_term_pair_check(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            history_path = Path(tmp_dir) / "history.xlsx"
            output_path = Path(tmp_dir) / "workflow_output.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "本批次 [Alpha]"
            worksheet["B2"] = "本批次 [临时译法]"
            workbook.save(input_path)

            history_workbook = Workbook()
            history_sheet = history_workbook.active
            history_sheet.title = "TB"
            history_sheet["A1"] = "source"
            history_sheet["B1"] = "target"
            history_sheet["A2"] = "Alpha"
            history_sheet["B2"] = "历史译法"
            history_workbook.save(history_path)

            summary = run_workflow(
                input_file=input_path,
                output_file=output_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                run_term_pair_check=True,
                term_mark_styles=("[]",),
                term_history_tb_file=history_path,
                term_history_sheet="TB",
                run_tag_check=False,
                run_line_break_check=False,
                run_source_consistency_check=False,
                run_chinese_target_check=False,
            )

            self.assertEqual(summary.term_count, 1)
            self.assertEqual(summary.term_problem_count, 1)
            self.assertEqual(summary.term_problem_rows, 1)

            workbook = load_workbook(summary.output_path)
            term_sheet = workbook["术语表"]
            self.assertEqual(term_sheet["B2"].value, "历史译法")
            self.assertEqual(term_sheet["E2"].value, "历史TB")

    def test_run_workflow_can_check_history_tb_without_term_marks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            history_path = Path(tmp_dir) / "history.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Alpha appears here"
            worksheet["B2"] = "这里使用错误译法"
            workbook.save(input_path)

            history_workbook = Workbook()
            history_sheet = history_workbook.active
            history_sheet.title = "TB"
            history_sheet["A1"] = "source"
            history_sheet["B1"] = "target"
            history_sheet["A2"] = "Alpha"
            history_sheet["B2"] = "历史译法"
            history_workbook.save(history_path)

            summary = run_workflow(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                run_term_pair_check=True,
                term_mark_styles=(),
                term_history_tb_file=history_path,
                term_history_sheet="TB",
                run_tag_check=False,
                run_line_break_check=False,
                run_source_consistency_check=False,
                run_chinese_target_check=False,
            )

            self.assertEqual(summary.term_count, 1)
            self.assertEqual(summary.term_problem_count, 1)
            self.assertEqual(summary.term_problem_rows, 1)

    def test_term_check_does_not_shift_a_source_column_right_of_target(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["B1"] = "target"
            worksheet["D1"] = "source"
            worksheet["B2"] = "[One]"
            worksheet["D2"] = "[Term]"
            worksheet["B3"] = "Two"
            worksheet["D3"] = "Same"
            worksheet["B4"] = "Three"
            worksheet["D4"] = "Same"
            workbook.save(input_path)

            summary = run_workflow(
                input_file=input_path,
                source_column="D",
                target_column="B",
                sheet="Data",
                term_mark_styles=("[]",),
                run_tag_check=False,
                run_line_break_check=False,
                run_source_consistency_check=True,
                run_chinese_target_check=False,
            )

            self.assertEqual(summary.source_consistency_problem_count, 1)
            self.assertEqual(summary.source_consistency_problem_rows, 2)
            output_workbook = load_workbook(summary.output_path)
            self.assertEqual(output_workbook["Data"]["D1"].value, "source")

if __name__ == "__main__":
    unittest.main()
