from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.excel_line_splitter.split_excel_lines import process_excel, split_cell_lines


class SplitCellLinesTests(unittest.TestCase):
    def test_split_simple_newline(self) -> None:
        self.assertEqual(split_cell_lines("ab\nc"), ["ab", "c"])

    def test_drop_empty_lines(self) -> None:
        self.assertEqual(split_cell_lines("ab\n\nc"), ["ab", "c"])

    def test_trim_spaces(self) -> None:
        self.assertEqual(split_cell_lines(" ab \n c"), ["ab", "c"])

    def test_drop_trailing_empty_line(self) -> None:
        self.assertEqual(split_cell_lines("ab\n"), ["ab"])

    def test_skip_empty_cell(self) -> None:
        self.assertEqual(split_cell_lines(None), [])
        self.assertEqual(split_cell_lines(""), [])


class ProcessExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "result"
        worksheet["A2"] = "ab\nc"
        worksheet["A3"] = " x \n\nz"
        worksheet["B2"] = "old1"
        worksheet["B3"] = "old2"
        worksheet["B4"] = "old3"
        worksheet["B10"] = "stale tail"
        worksheet["Z1000"] = "unrelated tail"
        worksheet["A1001"].number_format = "@"
        workbook.save(path)

    def test_result_column_is_stacked_and_cleared_in_output_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"
            self.create_workbook(input_path)

            worksheet_title, source_col, result_col, saved_path, written_count = process_excel(
                input_file=input_path,
                source_column="A",
                result_column="B",
                start_row=2,
                output_file=output_path,
            )

            self.assertEqual(worksheet_title, "Data")
            self.assertEqual(source_col, "A")
            self.assertEqual(result_col, "B")
            self.assertEqual(saved_path, output_path.resolve())
            self.assertEqual(written_count, 4)

            original_workbook = load_workbook(input_path)
            original_sheet = original_workbook["Data"]
            self.assertEqual(original_sheet["B2"].value, "old1")
            self.assertEqual(original_sheet["B3"].value, "old2")
            self.assertEqual(original_sheet["B4"].value, "old3")

            output_workbook = load_workbook(output_path)
            output_sheet = output_workbook["Data"]
            self.assertEqual(output_sheet["B1"].value, "result")
            self.assertEqual(output_sheet["B2"].value, "ab")
            self.assertEqual(output_sheet["B3"].value, "c")
            self.assertEqual(output_sheet["B4"].value, "x")
            self.assertEqual(output_sheet["B5"].value, "z")
            self.assertIsNone(output_sheet["B10"].value)

    def test_reject_same_source_and_result_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            with self.assertRaisesRegex(ValueError, "源列和结果列不能相同"):
                process_excel(
                    input_file=input_path,
                    source_column="A",
                    result_column="A",
                    start_row=2,
                )


if __name__ == "__main__":
    unittest.main()
