from __future__ import annotations

import io
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from phraseloom import workbook_schema as schema
from phraseloom.string_cluster import SimilarStringCluster
from phraseloom.strings_workflow import (
    export_strings_workbook,
    restore_strings_workbook,
)


def _write_source(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GameStrings"
    ws.append(["source", "target", "context"])
    ws.append(["Pikachu launched an attack", None, "battle"])
    ws.append(["Squirtle launched an attack", None, "battle"])
    ws.append(["Bulbasaur launched an attack", None, "battle"])
    ws.append(["Pikachu launched an attack", None, "tutorial"])
    ws.append(["Already translated", "Déjà traduit", "system"])
    ws.append(["123", None, "value"])
    wb.save(path)
    wb.close()


class StringsWorkflowTests(unittest.TestCase):
    def test_multiline_source_cells_export_segments_and_restore_exact_structure(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "done.xlsx"
            package_path = Path(tmp) / "done_strings.xlsx"
            result_path = Path(tmp) / "done_translated.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append(["三\n四\n五", None])
            worksheet.append(["三\n", None])
            worksheet.append(["四", None])
            worksheet.append(["五", None])
            workbook.save(source_path)
            workbook.close()

            stats = export_strings_workbook(source_path, package_path)

            self.assertEqual(stats["source_row_count"], 4)
            self.assertEqual(stats["source_segment_count"], 6)
            self.assertEqual(stats["multiline_source_row_count"], 2)
            self.assertEqual(stats["pending_row_count"], 4)
            self.assertEqual(stats["pending_segment_count"], 6)
            self.assertEqual(stats["string_count"], 3)
            self.assertEqual(stats["duplicate_row_count"], 3)

            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            source_index = headers.index(schema.SOURCE_COLUMN) + 1
            target_index = headers.index(schema.TARGET_COLUMN) + 1
            self.assertEqual(
                [
                    strings.cell(row=row_number, column=source_index).value
                    for row_number in range(2, strings.max_row + 1)
                ],
                ["三", "四", "五"],
            )
            translations = {"三": "Three", "四": "Four", "五": "Five"}
            for row_number in range(2, strings.max_row + 1):
                source = strings.cell(row=row_number, column=source_index).value
                strings.cell(row=row_number, column=target_index).value = translations[source]

            mapping = package[schema.STRINGS_MAP_SHEET]
            mapping_headers = [cell.value for cell in mapping[1]]
            mapping_records = [
                dict(zip(mapping_headers, values))
                for values in mapping.iter_rows(min_row=2, values_only=True)
            ]
            row_two_segments = sorted(
                (
                    record
                    for record in mapping_records
                    if record[schema.ROW_NUMBER_COLUMN] == 2
                ),
                key=lambda record: record[schema.SEGMENT_INDEX_COLUMN],
            )
            self.assertEqual(
                [record[schema.RAW_SEGMENT_COLUMN] for record in row_two_segments],
                ["三", "四", "五"],
            )
            self.assertEqual(
                [record[schema.SEGMENT_SUFFIX_COLUMN] or "" for record in row_two_segments],
                ["\n", "\n", ""],
            )
            row_three = next(
                record
                for record in mapping_records
                if record[schema.ROW_NUMBER_COLUMN] == 3
            )
            self.assertEqual(row_three[schema.SEGMENT_SUFFIX_COLUMN], "\n")
            package.save(package_path)
            package.close()

            restore_stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(restore_stats["restored_row_count"], 4)
            self.assertEqual(restore_stats["issue_count"], 0)
            restored = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(
                    [
                        row[1]
                        for row in restored.active.iter_rows(
                            min_row=2,
                            values_only=True,
                        )
                    ],
                    ["Three\nFour\nFive", "Three\n", "Four", "Five"],
                )
            finally:
                restored.close()

    def test_multiline_source_cell_can_be_exported_as_one_string(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "whole_cell.xlsx"
            package_path = Path(tmp) / "whole_cell_strings.xlsx"
            result_path = Path(tmp) / "whole_cell_translated.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append(["\n  One\nTwo  \n", None])
            workbook.save(source_path)
            workbook.close()

            stats = export_strings_workbook(
                source_path,
                package_path,
                split_lines=False,
            )

            self.assertFalse(stats["line_splitting_enabled"])
            self.assertEqual(stats["source_row_count"], 1)
            self.assertEqual(stats["source_segment_count"], 1)
            self.assertEqual(stats["pending_segment_count"], 1)
            self.assertEqual(stats["string_count"], 1)

            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            source_index = headers.index(schema.SOURCE_COLUMN) + 1
            target_index = headers.index(schema.TARGET_COLUMN) + 1
            self.assertEqual(
                strings.cell(row=2, column=source_index).value,
                "One\nTwo",
            )
            strings.cell(row=2, column=target_index).value = "Un\nDeux"

            mapping = package[schema.STRINGS_MAP_SHEET]
            mapping_headers = [cell.value for cell in mapping[1]]
            mapping_record = dict(
                zip(
                    mapping_headers,
                    next(mapping.iter_rows(min_row=2, values_only=True)),
                )
            )
            self.assertEqual(mapping_record[schema.SEGMENT_COUNT_COLUMN], 1)
            self.assertEqual(mapping_record[schema.SEGMENT_PREFIX_COLUMN], "\n  ")
            self.assertEqual(mapping_record[schema.SEGMENT_SUFFIX_COLUMN], "  \n")
            package.save(package_path)
            package.close()

            restore_stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(restore_stats["restored_row_count"], 1)
            self.assertEqual(restore_stats["issue_count"], 0)
            restored = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(
                    restored.active.cell(row=2, column=2).value,
                    "\n  Un\nDeux  \n",
                )
            finally:
                restored.close()

    def test_template_cleanup_does_not_drop_sibling_segment_in_same_cell(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "multiline_template.xlsx"
            package_path = Path(tmp) / "multiline_template_strings.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append(["Level 1\nStandalone", None])
            worksheet.append(["Level 2", None])
            workbook.save(source_path)
            workbook.close()

            export_strings_workbook(source_path, package_path)

            package = load_workbook(package_path, data_only=True)
            try:
                strings = package[schema.STRINGS_SHEET]
                headers = [cell.value for cell in strings[1]]
                source_index = headers.index(schema.SOURCE_COLUMN)
                self.assertEqual(
                    [
                        row[source_index]
                        for row in strings.iter_rows(min_row=2, values_only=True)
                    ],
                    ["Level {num1}", "Standalone"],
                )
            finally:
                package.close()

    def test_multiline_restore_does_not_write_a_partially_translated_cell(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "partial.xlsx"
            package_path = Path(tmp) / "partial_strings.xlsx"
            result_path = Path(tmp) / "partial_translated.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append(["One\nTwo", None])
            workbook.save(source_path)
            workbook.close()

            export_strings_workbook(source_path, package_path)
            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            source_index = headers.index(schema.SOURCE_COLUMN) + 1
            target_index = headers.index(schema.TARGET_COLUMN) + 1
            for row_number in range(2, strings.max_row + 1):
                if strings.cell(row=row_number, column=source_index).value == "One":
                    strings.cell(row=row_number, column=target_index).value = "Un"
            package.save(package_path)
            package.close()

            stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(stats["restored_row_count"], 0)
            self.assertEqual(stats["issue_count"], 1)
            restored = load_workbook(result_path, data_only=True)
            try:
                self.assertIsNone(restored.active.cell(row=2, column=2).value)
            finally:
                restored.close()

    def test_cross_line_paired_tags_stay_protected_and_restore_without_issues(self) -> None:
        sources = (
            "<color=#FF0000>Hello\nWorld</color>",
            "[b]Hello\nWorld[/b]",
        )
        for case_number, raw_source in enumerate(sources, start=1):
            with self.subTest(raw_source=raw_source), tempfile.TemporaryDirectory() as tmp:
                source_path = Path(tmp) / f"paired_{case_number}.xlsx"
                package_path = Path(tmp) / f"paired_{case_number}_strings.xlsx"
                result_path = Path(tmp) / f"paired_{case_number}_translated.xlsx"
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.append(["source", "target"])
                worksheet.append([raw_source, None])
                workbook.save(source_path)
                workbook.close()

                stats = export_strings_workbook(source_path, package_path)

                self.assertEqual(stats["source_segment_count"], 2)
                package = load_workbook(package_path)
                strings = package[schema.STRINGS_SHEET]
                headers = [cell.value for cell in strings[1]]
                source_index = headers.index(schema.SOURCE_COLUMN) + 1
                target_index = headers.index(schema.TARGET_COLUMN) + 1
                visible_sources = [
                    strings.cell(row=row_number, column=source_index).value
                    for row_number in range(2, strings.max_row + 1)
                ]
                self.assertEqual(visible_sources, ["{1>Hello", "World<2}"])
                for row_number in range(2, strings.max_row + 1):
                    strings.cell(row=row_number, column=target_index).value = (
                        strings.cell(row=row_number, column=source_index).value
                    )
                package.save(package_path)
                package.close()

                restore_stats = restore_strings_workbook(package_path, result_path)

                self.assertEqual(restore_stats["restored_row_count"], 1)
                self.assertEqual(restore_stats["issue_count"], 0)
                restored = load_workbook(result_path, data_only=True)
                try:
                    self.assertEqual(
                        restored.active.cell(row=2, column=2).value,
                        raw_source,
                    )
                finally:
                    restored.close()

    def test_multiline_atomic_tag_is_not_split_inside_protected_syntax(self) -> None:
        raw_source = '[mq:rxt displaytext="Hello\nWorld" val="value"] suffix'
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "atomic.xlsx"
            package_path = Path(tmp) / "atomic_strings.xlsx"
            result_path = Path(tmp) / "atomic_translated.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append([raw_source, None])
            workbook.save(source_path)
            workbook.close()

            stats = export_strings_workbook(source_path, package_path)

            self.assertEqual(stats["multiline_source_row_count"], 1)
            self.assertEqual(stats["source_segment_count"], 1)
            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            source_index = headers.index(schema.SOURCE_COLUMN) + 1
            target_index = headers.index(schema.TARGET_COLUMN) + 1
            self.assertEqual(
                strings.cell(row=2, column=source_index).value,
                "{1} suffix",
            )
            strings.cell(row=2, column=target_index).value = "{1} suffix"
            package.save(package_path)
            package.close()

            restore_stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(restore_stats["issue_count"], 0)
            restored = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(restored.active.cell(row=2, column=2).value, raw_source)
            finally:
                restored.close()

    def test_multiline_passthrough_stats_distinguish_rows_and_segments(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "passthrough_stats.xlsx"
            package_path = Path(tmp) / "passthrough_stats_strings.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append(["123\n---", None])
            worksheet.append(["Hello\n---", None])
            workbook.save(source_path)
            workbook.close()

            stats = export_strings_workbook(source_path, package_path)

            self.assertEqual(stats["source_row_count"], 2)
            self.assertEqual(stats["source_segment_count"], 4)
            self.assertEqual(stats["auto_completed_row_count"], 1)
            self.assertEqual(stats["auto_completed_segment_count"], 3)
            self.assertEqual(stats["non_translatable_row_count"], 1)
            self.assertEqual(stats["non_translatable_segment_count"], 3)
            self.assertEqual(stats["pending_row_count"], 1)
            self.assertEqual(stats["pending_segment_count"], 1)
            self.assertEqual(stats["duplicate_row_count"], 0)
            self.assertEqual(stats["duplicate_segment_count"], 0)

    def test_restore_accepts_legacy_single_segment_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "legacy.xlsx"
            package_path = Path(tmp) / "legacy_strings.xlsx"
            result_path = Path(tmp) / "legacy_translated.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append(["Hello", None])
            workbook.save(source_path)
            workbook.close()

            export_strings_workbook(source_path, package_path)
            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            string_headers = [cell.value for cell in strings[1]]
            strings.cell(
                row=2,
                column=string_headers.index(schema.TARGET_COLUMN) + 1,
            ).value = "Bonjour"
            mapping = package[schema.STRINGS_MAP_SHEET]
            mapping.delete_cols(
                len(schema.LEGACY_STRINGS_MAP_COLUMNS) + 1,
                len(schema.STRINGS_MAP_COLUMNS)
                - len(schema.LEGACY_STRINGS_MAP_COLUMNS),
            )
            metadata = package[schema.METADATA_SHEET]
            for row in metadata.iter_rows(min_row=2, max_col=2):
                if row[0].value == schema.SCHEMA_VERSION_KEY:
                    row[1].value = "1.0"
                    break
            package.save(package_path)
            package.close()

            stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(stats["restored_row_count"], 1)
            self.assertEqual(stats["issue_count"], 0)
            restored = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(
                    restored.active.cell(row=2, column=2).value,
                    "Bonjour",
                )
            finally:
                restored.close()

    def test_numeric_variants_use_old_cleaning_before_grouping_and_restore(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "equipment.xlsx"
            package_path = Path(tmp) / "equipment_strings.xlsx"
            result_path = Path(tmp) / "equipment_translated.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            for source in (
                "通用补偿器LV1",
                "通用补偿器LV2",
                "通用补偿器LV3",
                "通用补偿器LV4",
                "延长枪管LV2",
                "延长枪管LV3",
                "延长枪管LV4",
            ):
                ws.append([source, None])
            wb.save(source_path)
            wb.close()

            with patch(
                "phraseloom.strings_workflow.cluster_similar_strings",
                return_value=[],
            ) as cluster:
                stats = export_strings_workbook(
                    source_path,
                    package_path,
                    group_similar=True,
                )

            self.assertEqual(stats["pending_row_count"], 7)
            self.assertEqual(stats["string_count"], 2)
            self.assertEqual(stats["duplicate_row_count"], 5)
            self.assertEqual(
                cluster.call_args.args[0],
                ["通用补偿器LV{num1}", "延长枪管LV{num1}"],
            )

            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            source_index = headers.index(schema.SOURCE_COLUMN) + 1
            target_index = headers.index(schema.TARGET_COLUMN) + 1
            translations = {
                "通用补偿器LV{num1}": "Universal Compensator LV{num1}",
                "延长枪管LV{num1}": "Extended Barrel LV{num1}",
            }
            occurrences = {}
            samples = {}
            for row_number in range(2, strings.max_row + 1):
                source = strings.cell(row=row_number, column=source_index).value
                strings.cell(row=row_number, column=target_index).value = translations[source]
                occurrences[source] = strings.cell(
                    row=row_number,
                    column=headers.index(schema.OCCURRENCES_COLUMN) + 1,
                ).value
                samples[source] = strings.cell(
                    row=row_number,
                    column=headers.index(schema.SAMPLE_SOURCES_COLUMN) + 1,
                ).value
            self.assertEqual(
                occurrences,
                {
                    "通用补偿器LV{num1}": 4,
                    "延长枪管LV{num1}": 3,
                },
            )
            self.assertEqual(
                samples["通用补偿器LV{num1}"],
                "通用补偿器LV1",
            )
            package.save(package_path)
            package.close()

            restore_stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(restore_stats["restored_row_count"], 7)
            self.assertEqual(restore_stats["issue_count"], 0)
            result = load_workbook(result_path, data_only=True)
            try:
                targets = [
                    row[1]
                    for row in result.active.iter_rows(min_row=2, values_only=True)
                ]
                self.assertEqual(
                    targets,
                    [
                        "Universal Compensator LV1",
                        "Universal Compensator LV2",
                        "Universal Compensator LV3",
                        "Universal Compensator LV4",
                        "Extended Barrel LV2",
                        "Extended Barrel LV3",
                        "Extended Barrel LV4",
                    ],
                )
            finally:
                result.close()

    def test_mq_rxt_tags_are_atomic_through_export_and_restore(self) -> None:
        def mq_rxt(displaytext: str, value: str) -> str:
            return f'[mq:rxt displaytext="{displaytext}" val="{value}"]'

        def make_source(
            *,
            count: int,
            display_color: str,
            value_color: str,
            display_link_color: str,
            value_link_color: str,
            display_action: int,
            value_action: int,
            references: tuple[int, int, int, int, int, int],
        ) -> tuple[str, str]:
            start = mq_rxt(
                fr"<span color=\&quot;{display_color}\&quot;>",
                fr"<span color=\&quot;{value_color}\&quot;>",
            )
            close = mq_rxt("</>", "</>")
            link = mq_rxt(
                fr"<hyperlink color=\&quot;{display_link_color}\&quot; "
                fr"action=\&quot;{display_action}\&quot;>",
                fr"<hyperlink color=\&quot;{value_link_color}\&quot; "
                fr"action=\&quot;{value_action}\&quot;>",
            )
            reference = mq_rxt(
                fr"\{{{references[0]}}}",
                fr"\{{{references[1]}}}",
            )
            left = mq_rxt(
                fr"\\{{{references[2]}}}",
                fr"\\{{{references[3]}}}",
            )
            right = mq_rxt(
                fr"\\{{{references[4]}}}",
                fr"\\{{{references[5]}}}",
            )
            source = (
                f"在{start}精英难度下的破碎中枢{close}中，使用{count}次"
                f"{link}{reference}{close}({left}/{right})"
            )
            expected_target = (
                f"In {start}the Shattered Nexus on Elite difficulty{close}, "
                f"use {count} times {link}{reference}{close}({left}/{right})"
            )
            return source, expected_target

        source_one, expected_one = make_source(
            count=3,
            display_color="#FF0000",
            value_color="#00FF00",
            display_link_color="#ABCDEF",
            value_link_color="#123456",
            display_action=101,
            value_action=201,
            references=(1, 2, 3, 4, 5, 6),
        )
        source_two, expected_two = make_source(
            count=4,
            display_color="#112233",
            value_color="#445566",
            display_link_color="#778899",
            value_link_color="#AABBCC",
            display_action=102,
            value_action=202,
            references=(7, 8, 9, 10, 11, 12),
        )

        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "mq_tags.xlsx"
            package_path = Path(tmp) / "mq_tags_strings.xlsx"
            result_path = Path(tmp) / "mq_tags_translated.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append([source_one, None])
            worksheet.append([source_two, None])
            workbook.save(source_path)
            workbook.close()

            export_stats = export_strings_workbook(source_path, package_path)

            self.assertEqual(export_stats["pending_row_count"], 2)
            self.assertEqual(export_stats["string_count"], 1)
            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            source_index = headers.index(schema.SOURCE_COLUMN) + 1
            target_index = headers.index(schema.TARGET_COLUMN) + 1
            self.assertEqual(
                strings.cell(row=2, column=source_index).value,
                (
                    "在{1}精英难度下的破碎中枢{2}中，使用{num1}次"
                    "{3}{4}{5}({6}/{7})"
                ),
            )
            strings.cell(row=2, column=target_index).value = (
                "In {1}the Shattered Nexus on Elite difficulty{2}, "
                "use {num1} times {3}{4}{5}({6}/{7})"
            )
            package.save(package_path)
            package.close()

            restore_stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(restore_stats["restored_row_count"], 2)
            self.assertEqual(restore_stats["issue_count"], 0)
            result = load_workbook(result_path, data_only=True)
            try:
                targets = [
                    row[1]
                    for row in result.active.iter_rows(min_row=2, values_only=True)
                ]
                self.assertEqual(targets, [expected_one, expected_two])
            finally:
                result.close()

    def test_raw_span_and_hyperlink_tags_are_protected_and_restored(self) -> None:
        raw_source = (
            '<span color="#BFF8FA" size="36">Hello</> '
            '<hyperlink color="#E98845" action="Key39">details</>'
        )

        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "raw_rich_text_tags.xlsx"
            package_path = Path(tmp) / "raw_rich_text_tags_strings.xlsx"
            result_path = Path(tmp) / "raw_rich_text_tags_translated.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append([raw_source, None])
            workbook.save(source_path)
            workbook.close()

            export_strings_workbook(source_path, package_path)

            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            source_index = headers.index(schema.SOURCE_COLUMN) + 1
            target_index = headers.index(schema.TARGET_COLUMN) + 1
            self.assertEqual(
                strings.cell(row=2, column=source_index).value,
                "{1>Hello<2} {3>details<4}",
            )
            strings.cell(row=2, column=target_index).value = (
                "{1>Bonjour<2} {3>more details<4}"
            )
            package.save(package_path)
            package.close()

            restore_stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(restore_stats["restored_row_count"], 1)
            self.assertEqual(restore_stats["issue_count"], 0)
            result = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(
                    result.active.cell(row=2, column=2).value,
                    (
                        '<span color="#BFF8FA" size="36">Bonjour</> '
                        '<hyperlink color="#E98845" action="Key39">'
                        "more details</>"
                    ),
                )
            finally:
                result.close()

    def test_restore_accepts_mixed_tokens_and_complete_raw_mq_tags(self) -> None:
        hyperlink = (
            r'[mq:rxt displaytext="<hyperlink color=\&quot;#FF8E33\&quot; '
            r'action=\&quot;19040000001\&quot;>" val="<hyperlink color='
            r'\&quot;#FF8E33\&quot; action=\&quot;19040000001\&quot;>"]'
        )
        reference = r'[mq:rxt displaytext="\{2}" val="\{2}"]'
        close = r'[mq:rxt displaytext="</>" val="</>"]'
        left = r'[mq:rxt displaytext="\{0}" val="\{0}"]'
        right = r'[mq:rxt displaytext="\{1}" val="\{1}"]'
        raw_source = (
            f"拾取从该区域回收的{hyperlink}{reference}{close}"
            f"({left}/{right})"
        )
        mixed_target = f"拾取从该区域回收的{{1}}{reference}{{3}}({{4}}/{{5}})"

        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "mq_raw_target.xlsx"
            package_path = Path(tmp) / "mq_raw_target_strings.xlsx"
            result_path = Path(tmp) / "mq_raw_target_restored.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append([raw_source, None])
            workbook.save(source_path)
            workbook.close()

            export_strings_workbook(source_path, package_path)
            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            strings.cell(
                row=2,
                column=headers.index(schema.TARGET_COLUMN) + 1,
            ).value = mixed_target
            package.save(package_path)
            package.close()

            restore_stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(restore_stats["restored_row_count"], 1)
            self.assertEqual(restore_stats["issue_count"], 0)
            result = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(result.active.cell(row=2, column=2).value, raw_source)
            finally:
                result.close()

    def test_export_skips_completed_rows_deduplicates_and_groups(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "source_strings.xlsx"
            _write_source(source_path)

            with patch(
                "phraseloom.strings_workflow.cluster_similar_strings"
            ) as cluster:
                stats = export_strings_workbook(source_path, output_path)

            cluster.assert_not_called()
            self.assertEqual(stats["completed_row_count"], 1)
            self.assertEqual(stats["non_translatable_row_count"], 1)
            self.assertEqual(stats["auto_completed_row_count"], 1)
            self.assertEqual(stats["pending_row_count"], 4)
            self.assertEqual(stats["string_count"], 3)
            self.assertEqual(stats["duplicate_row_count"], 1)
            self.assertEqual(stats["group_count"], 0)
            self.assertEqual(stats["grouped_string_count"], 0)
            self.assertFalse(stats["grouping_enabled"])

            wb = load_workbook(output_path, data_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "GameStrings",
                        schema.STRINGS_SHEET,
                        schema.COMPLETED_STRINGS_SHEET,
                        schema.STRINGS_MAP_SHEET,
                        schema.METADATA_SHEET,
                    ],
                )
                self.assertEqual(wb["GameStrings"].sheet_state, "hidden")
                self.assertEqual(wb[schema.STRINGS_SHEET].sheet_state, "visible")
                self.assertEqual(
                    wb[schema.COMPLETED_STRINGS_SHEET].sheet_state,
                    "visible",
                )
                self.assertEqual(wb[schema.STRINGS_MAP_SHEET].sheet_state, "hidden")
                self.assertTrue(
                    wb[schema.STRINGS_SHEET].column_dimensions["A"].hidden
                )
                rows = list(wb[schema.STRINGS_SHEET].values)
                self.assertEqual(rows[0], tuple(schema.STRINGS_COLUMNS))
                records = [dict(zip(rows[0], row)) for row in rows[1:]]
                self.assertEqual(
                    {row[schema.GROUP_COLUMN] for row in records},
                    {None},
                )
                self.assertEqual(
                    [row[schema.SOURCE_COLUMN] for row in records],
                    [
                        "Pikachu launched an attack",
                        "Squirtle launched an attack",
                        "Bulbasaur launched an attack",
                    ],
                )
                pikachu = next(
                    row
                    for row in records
                    if str(row[schema.SOURCE_COLUMN]).startswith("Pikachu")
                )
                self.assertEqual(pikachu[schema.OCCURRENCES_COLUMN], 2)
                self.assertEqual(
                    pikachu[schema.SAMPLE_SOURCES_COLUMN],
                    "Pikachu launched an attack",
                )
                self.assertEqual(
                    pikachu[schema.CONTEXT_COLUMN],
                    "battle",
                )
                completed_rows = list(
                    wb[schema.COMPLETED_STRINGS_SHEET].values
                )
                self.assertEqual(
                    completed_rows[0],
                    tuple(schema.COMPLETED_STRINGS_COLUMNS),
                )
                completed = [
                    dict(zip(completed_rows[0], row))
                    for row in completed_rows[1:]
                ]
                self.assertEqual(
                    [
                        (
                            row[schema.STATUS_COLUMN],
                            row[schema.SOURCE_COLUMN],
                            row[schema.TARGET_COLUMN],
                        )
                        for row in completed
                    ],
                    [
                        (
                            "existing_target",
                            "Already translated",
                            "Déjà traduit",
                        ),
                        ("auto_passthrough", "123", "123"),
                    ],
                )
            finally:
                wb.close()

    def test_optional_grouping_places_clustered_units_after_unclustered(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "ordering.xlsx"
            output_path = Path(tmp) / "ordering_strings.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            for source in (
                "Unclustered first",
                "Alpha attacks",
                "Unclustered middle",
                "Beta attacks",
                "Gamma attacks",
            ):
                ws.append([source, None])
            wb.save(source_path)
            wb.close()

            cluster_result = [
                SimilarStringCluster(
                    group_id="G001",
                    source_pattern="{variant} attacks",
                    member_indexes=(1, 3, 4),
                )
            ]
            with patch(
                "phraseloom.strings_workflow.cluster_similar_strings",
                return_value=cluster_result,
            ) as cluster:
                stats = export_strings_workbook(
                    source_path,
                    output_path,
                    group_similar=True,
                )

            cluster.assert_called_once_with(
                [
                    "Unclustered first",
                    "Alpha attacks",
                    "Unclustered middle",
                    "Beta attacks",
                    "Gamma attacks",
                ],
                min_group_size=3,
            )
            self.assertTrue(stats["grouping_enabled"])
            self.assertEqual(stats["group_count"], 1)
            workbook = load_workbook(output_path, data_only=True)
            try:
                strings = workbook[schema.STRINGS_SHEET]
                headers = [cell.value for cell in strings[1]]
                records = [
                    dict(zip(headers, row))
                    for row in strings.iter_rows(min_row=2, values_only=True)
                ]
                self.assertEqual(
                    [
                        (
                            row[schema.GROUP_COLUMN],
                            row[schema.SOURCE_COLUMN],
                        )
                        for row in records
                    ],
                    [
                        (None, "Unclustered first"),
                        (None, "Unclustered middle"),
                        ("G001", "Alpha attacks"),
                        ("G001", "Beta attacks"),
                        ("G001", "Gamma attacks"),
                    ],
                )
            finally:
                workbook.close()

    def test_restore_writes_pending_rows_and_preserves_existing_targets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "source.xlsx"
            package_path = Path(tmp) / "source_strings.xlsx"
            result_path = Path(tmp) / "source_translated.xlsx"
            _write_source(source_path)
            export_strings_workbook(source_path, package_path)

            translations = {
                "Pikachu launched an attack": "Pikachu attaque",
                "Squirtle launched an attack": "Squirtle attaque",
                "Bulbasaur launched an attack": "Bulbasaur attaque",
            }
            wb = load_workbook(package_path)
            ws = wb[schema.STRINGS_SHEET]
            headers = [cell.value for cell in ws[1]]
            source_index = headers.index(schema.SOURCE_COLUMN) + 1
            target_index = headers.index(schema.TARGET_COLUMN) + 1
            for row_number in range(2, ws.max_row + 1):
                source = ws.cell(row=row_number, column=source_index).value
                ws.cell(row=row_number, column=target_index).value = translations[source]
            wb.save(package_path)
            wb.close()

            stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(stats["restored_row_count"], 5)
            self.assertEqual(stats["issue_count"], 0)
            self.assertNotIn("audit_output_path", stats)
            result = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(result.sheetnames, ["GameStrings"])
                values = list(result["GameStrings"].values)
                self.assertEqual(values[1][1], "Pikachu attaque")
                self.assertEqual(values[2][1], "Squirtle attaque")
                self.assertEqual(values[3][1], "Bulbasaur attaque")
                self.assertEqual(values[4][1], "Pikachu attaque")
                self.assertEqual(values[5][1], "Déjà traduit")
                self.assertEqual(values[6][1], "123")
            finally:
                result.close()

    def test_restore_creates_review_file_only_when_a_target_is_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "source.xlsx"
            package_path = Path(tmp) / "source_strings.xlsx"
            _write_source(source_path)
            export_strings_workbook(source_path, package_path)

            stats = restore_strings_workbook(package_path)

            self.assertEqual(stats["restored_row_count"], 1)
            self.assertGreater(stats["issue_count"], 0)
            self.assertTrue(Path(str(stats["audit_output_path"])).exists())

    def test_completed_sheet_and_restore_cover_non_translatable_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "passthrough.xlsx"
            package_path = Path(tmp) / "passthrough_strings.xlsx"
            result_path = Path(tmp) / "passthrough_translated.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target", "context"])
            ws.append(["123", None, "number"])
            ws.append(["---", None, "symbol"])
            ws.append(["<br/>", None, "tag"])
            ws.append(["{player_name}", None, "placeholder"])
            wb.save(source_path)
            wb.close()

            stats = export_strings_workbook(source_path, package_path)

            self.assertEqual(stats["string_count"], 0)
            self.assertEqual(stats["auto_completed_row_count"], 4)
            package = load_workbook(package_path, data_only=True)
            try:
                self.assertEqual(package[schema.STRINGS_SHEET].max_row, 1)
                completed = package[schema.COMPLETED_STRINGS_SHEET]
                headers = [cell.value for cell in completed[1]]
                rows = [
                    dict(zip(headers, row))
                    for row in completed.iter_rows(min_row=2, values_only=True)
                ]
                self.assertEqual(
                    [
                        (
                            row[schema.SOURCE_COLUMN],
                            row[schema.TARGET_COLUMN],
                            row[schema.CONTEXT_COLUMN],
                        )
                        for row in rows
                    ],
                    [
                        ("123", "123", "number"),
                        ("---", "---", "symbol"),
                        ("<br/>", "<br/>", "tag"),
                        ("{player_name}", "{player_name}", "placeholder"),
                    ],
                )
            finally:
                package.close()

            restore_stats = restore_strings_workbook(package_path, result_path)

            self.assertEqual(restore_stats["restored_row_count"], 4)
            self.assertEqual(restore_stats["issue_count"], 0)
            result = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(
                    [row[1] for row in result.active.iter_rows(min_row=2, values_only=True)],
                    ["123", "---", "<br/>", "{player_name}"],
                )
            finally:
                result.close()

    def test_missing_target_column_is_created_during_restore(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "source.xlsx"
            package_path = Path(tmp) / "source_strings.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source"])
            ws.append(["Hello"])
            wb.save(source_path)
            wb.close()

            export_strings_workbook(source_path, package_path)
            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            strings.cell(
                row=2,
                column=headers.index(schema.TARGET_COLUMN) + 1,
            ).value = "Bonjour"
            package.save(package_path)
            package.close()

            stats = restore_strings_workbook(package_path)
            result = load_workbook(stats["output_path"], data_only=True)
            try:
                self.assertEqual(
                    list(result.active.values),
                    [("source", "target"), ("Hello", "Bonjour")],
                )
            finally:
                result.close()

    def test_restore_writes_target_and_reports_missing_protected_tags(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "source.xlsx"
            package_path = Path(tmp) / "source_strings.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<color=red>Hello</color>", None])
            wb.save(source_path)
            wb.close()

            export_strings_workbook(source_path, package_path)
            package = load_workbook(package_path)
            strings = package[schema.STRINGS_SHEET]
            headers = [cell.value for cell in strings[1]]
            strings.cell(
                row=2,
                column=headers.index(schema.TARGET_COLUMN) + 1,
            ).value = "Bonjour"
            package.save(package_path)
            package.close()

            stats = restore_strings_workbook(package_path)

            self.assertEqual(stats["restored_row_count"], 1)
            self.assertGreater(stats["issue_count"], 0)
            result = load_workbook(stats["output_path"], data_only=True)
            try:
                self.assertEqual(result.active.cell(row=2, column=2).value, "Bonjour")
            finally:
                result.close()
            audit = load_workbook(stats["audit_output_path"], data_only=True)
            try:
                issue_text = "\n".join(
                    str(row[-1] or "")
                    for row in audit["restore_issues"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                )
                self.assertIn("protected_token_mismatch", issue_text)
            finally:
                audit.close()

    def test_cli_and_interactive_use_the_two_step_strings_language(self) -> None:
        from phraseloom.cli import _dispatch

        help_output = io.StringIO()
        with redirect_stdout(help_output):
            self.assertEqual(_dispatch(["--help"]), 0)
        help_text = help_output.getvalue()
        self.assertIn("Export untranslated Strings", help_text)
        self.assertIn("Restore translated Strings", help_text)
        self.assertNotIn("TM", help_text)
        self.assertNotIn("Entity", help_text)

        menu_output = io.StringIO()
        with patch("builtins.input", side_effect=["q"]), redirect_stdout(menu_output):
            self.assertEqual(_dispatch([]), 0)
        menu_text = menu_output.getvalue()
        self.assertIn("1) Export untranslated Strings", menu_text)
        self.assertIn("2) Restore translated Strings", menu_text)
        self.assertNotIn("TM", menu_text)
        self.assertNotIn("Entity", menu_text)

    def test_cli_can_disable_multiline_source_splitting(self) -> None:
        from phraseloom.cli import _dispatch

        with (
            patch(
                "phraseloom.cli.export_strings_workbook",
                return_value={},
            ) as export_mock,
            patch("phraseloom.cli._print_export_stats"),
        ):
            self.assertEqual(
                _dispatch(["export", "source.xlsx", "--no-split-lines"]),
                0,
            )

        self.assertFalse(export_mock.call_args.kwargs["split_lines"])

    def test_removed_legacy_commands_are_rejected(self) -> None:
        from phraseloom.cli import main

        for command in ("tm-extract", "prepare", "fill", "entity"):
            error_output = io.StringIO()
            with redirect_stderr(error_output):
                self.assertEqual(main([command]), 1)
            self.assertIn("Unknown command", error_output.getvalue())


if __name__ == "__main__":
    unittest.main()
