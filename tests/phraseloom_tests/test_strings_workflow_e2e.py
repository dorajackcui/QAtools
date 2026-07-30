from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from phraseloom import workbook_schema as schema
from phraseloom.strings_workflow import (
    export_strings_workbook,
    restore_strings_workbook,
)


class CompleteStringsWorkflowE2ETests(unittest.TestCase):
    def test_complete_export_clean_passthrough_group_and_restore_pipeline(self) -> None:
        source_rows = [
            (None, None, "blank"),
            ("Already translated", "Déjà traduit", "existing"),
            ("12345", None, "number"),
            ("---", None, "symbol"),
            ("<br/>", None, "tag-only"),
            ("{player_name}", None, "placeholder-only"),
            ("Welcome commander", None, "intro-first"),
            ("Welcome commander", None, "intro-second"),
            ("通用补偿器LV1", None, "equipment-first"),
            ("通用补偿器LV2", None, "equipment-second"),
            ("通用补偿器LV3", None, "equipment-third"),
            ("Clear Story 10-20", None, "stage-first"),
            ("Clear Story 11-20", None, "stage-second"),
            ("Version 1.2.3", None, "sequence-first"),
            ("Version 2.3.4", None, "sequence-second"),
            ("Use color #FF0000", None, "color-first"),
            ("Use color #00FF00", None, "color-second"),
            ("<color=#FF0000>Level 1</color>", None, "angle-tag-first"),
            ("<color=#00FF00>Level 2</color>", None, "angle-tag-second"),
            ("[b]Damage {0}: 10[/b]", None, "bbcode-first"),
            ("[b]Damage {0}: 20[/b]", None, "bbcode-second"),
            ("Pikachu launched an attack", None, "cluster-first"),
            ("Squirtle launched an attack", None, "cluster-second"),
            ("Bulbasaur launched an attack", None, "cluster-third"),
            ("Final standalone", None, "after-cluster-in-source"),
        ]
        translations = {
            "Welcome commander": "Bienvenue commandant",
            "通用补偿器LV{num1}": "Universal Compensator LV{num1}",
            "Clear Story {stage1}": "Terminer l'histoire {stage1}",
            "Version {seq1}": "Version {seq1}",
            "Use color {color1}": "Utiliser la couleur {color1}",
            "{1>Level {num1}<2}": "{1>Niveau {num1}<2}",
            "{1>Damage {2}: {num1}<3}": "{1>Dégâts {2} : {num1}<3}",
            "Pikachu launched an attack": "Pikachu attaque",
            "Squirtle launched an attack": "Squirtle attaque",
            "Bulbasaur launched an attack": "Bulbasaur attaque",
            "Final standalone": "Final autonome",
        }
        default_order = [
            "Welcome commander",
            "通用补偿器LV{num1}",
            "Clear Story {stage1}",
            "Version {seq1}",
            "Use color {color1}",
            "{1>Level {num1}<2}",
            "{1>Damage {2}: {num1}<3}",
            "Pikachu launched an attack",
            "Squirtle launched an attack",
            "Bulbasaur launched an attack",
            "Final standalone",
        ]
        grouped_order = [
            "Welcome commander",
            "通用补偿器LV{num1}",
            "Clear Story {stage1}",
            "Version {seq1}",
            "Use color {color1}",
            "{1>Level {num1}<2}",
            "{1>Damage {2}: {num1}<3}",
            "Final standalone",
            "Pikachu launched an attack",
            "Squirtle launched an attack",
            "Bulbasaur launched an attack",
        ]
        expected_targets = [
            None,
            "Déjà traduit",
            "12345",
            "---",
            "<br/>",
            "{player_name}",
            "Bienvenue commandant",
            "Bienvenue commandant",
            "Universal Compensator LV1",
            "Universal Compensator LV2",
            "Universal Compensator LV3",
            "Terminer l'histoire 10-20",
            "Terminer l'histoire 11-20",
            "Version 1.2.3",
            "Version 2.3.4",
            "Utiliser la couleur #FF0000",
            "Utiliser la couleur #00FF00",
            "<color=#FF0000>Niveau 1</color>",
            "<color=#00FF00>Niveau 2</color>",
            "[b]Dégâts {0} : 10[/b]",
            "[b]Dégâts {0} : 20[/b]",
            "Pikachu attaque",
            "Squirtle attaque",
            "Bulbasaur attaque",
            "Final autonome",
        ]

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_path = root / "complete_source.xlsx"
            default_package_path = root / "complete_default_strings.xlsx"
            grouped_package_path = root / "complete_grouped_strings.xlsx"
            restored_path = root / "complete_translated.xlsx"
            self._write_source_workbook(source_path, source_rows)

            default_stats = export_strings_workbook(
                source_path,
                default_package_path,
            )

            self.assertFalse(default_stats["grouping_enabled"])
            self.assertEqual(default_stats["group_count"], 0)
            self.assertEqual(default_stats["string_count"], len(default_order))
            default_book = load_workbook(default_package_path, data_only=True)
            try:
                default_records = self._sheet_records(
                    default_book[schema.STRINGS_SHEET]
                )
                self.assertEqual(
                    [
                        record[schema.SOURCE_COLUMN]
                        for record in default_records
                    ],
                    default_order,
                )
                self.assertEqual(
                    {
                        record[schema.GROUP_COLUMN]
                        for record in default_records
                    },
                    {None},
                )
            finally:
                default_book.close()

            grouped_stats = export_strings_workbook(
                source_path,
                grouped_package_path,
                group_similar=True,
            )

            self.assertTrue(grouped_stats["grouping_enabled"])
            self.assertEqual(grouped_stats["group_count"], 1)
            self.assertEqual(grouped_stats["grouped_string_count"], 3)
            self.assertEqual(grouped_stats["completed_row_count"], 1)
            self.assertEqual(grouped_stats["auto_completed_row_count"], 4)
            self.assertEqual(grouped_stats["pending_row_count"], 19)
            self.assertEqual(grouped_stats["string_count"], len(grouped_order))

            grouped_book = load_workbook(grouped_package_path)
            strings = grouped_book[schema.STRINGS_SHEET]
            completed = grouped_book[schema.COMPLETED_STRINGS_SHEET]
            try:
                records = self._sheet_records(strings)
                self.assertEqual(
                    [record[schema.SOURCE_COLUMN] for record in records],
                    grouped_order,
                )
                self.assertTrue(
                    all(
                        record[schema.GROUP_COLUMN] is None
                        for record in records[:-3]
                    )
                )
                self.assertEqual(
                    {
                        record[schema.GROUP_COLUMN]
                        for record in records[-3:]
                    },
                    {"G001"},
                )

                by_source = {
                    record[schema.SOURCE_COLUMN]: record
                    for record in records
                }
                self.assertEqual(
                    by_source["Welcome commander"][schema.OCCURRENCES_COLUMN],
                    2,
                )
                self.assertEqual(
                    by_source["Welcome commander"][schema.CONTEXT_COLUMN],
                    "intro-first",
                )
                self.assertEqual(
                    by_source["通用补偿器LV{num1}"][
                        schema.SAMPLE_SOURCES_COLUMN
                    ],
                    "通用补偿器LV1",
                )
                self.assertEqual(
                    by_source["通用补偿器LV{num1}"][schema.CONTEXT_COLUMN],
                    "equipment-first",
                )

                completed_records = self._sheet_records(completed)
                self.assertEqual(
                    [
                        (
                            record[schema.STATUS_COLUMN],
                            record[schema.SOURCE_COLUMN],
                            record[schema.TARGET_COLUMN],
                        )
                        for record in completed_records
                    ],
                    [
                        (
                            "existing_target",
                            "Already translated",
                            "Déjà traduit",
                        ),
                        ("auto_passthrough", "12345", "12345"),
                        ("auto_passthrough", "---", "---"),
                        ("auto_passthrough", "<br/>", "<br/>"),
                        (
                            "auto_passthrough",
                            "{player_name}",
                            "{player_name}",
                        ),
                    ],
                )

                headers = [cell.value for cell in strings[1]]
                source_column = headers.index(schema.SOURCE_COLUMN) + 1
                target_column = headers.index(schema.TARGET_COLUMN) + 1
                for row_number in range(2, strings.max_row + 1):
                    cleaned_source = strings.cell(
                        row=row_number,
                        column=source_column,
                    ).value
                    strings.cell(
                        row=row_number,
                        column=target_column,
                    ).value = translations[cleaned_source]
                grouped_book.save(grouped_package_path)
            finally:
                grouped_book.close()

            restore_stats = restore_strings_workbook(
                grouped_package_path,
                restored_path,
            )

            self.assertEqual(restore_stats["issue_count"], 0)
            self.assertEqual(restore_stats["restored_row_count"], 23)
            self.assertNotIn("audit_output_path", restore_stats)
            restored_book = load_workbook(restored_path, data_only=True)
            try:
                self.assertEqual(restored_book.sheetnames, ["GameStrings"])
                actual_targets = [
                    row[1]
                    for row in restored_book["GameStrings"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                ]
                self.assertEqual(actual_targets, expected_targets)
            finally:
                restored_book.close()

    @staticmethod
    def _write_source_workbook(
        path: Path,
        rows: list[tuple[str | None, str | None, str]],
    ) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "GameStrings"
        worksheet.append(["source", "target", "context"])
        for row in rows:
            worksheet.append(row)
        workbook.save(path)
        workbook.close()

    @staticmethod
    def _sheet_records(worksheet) -> list[dict[str, object]]:
        headers = [cell.value for cell in worksheet[1]]
        return [
            dict(zip(headers, row))
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]


if __name__ == "__main__":
    unittest.main()
