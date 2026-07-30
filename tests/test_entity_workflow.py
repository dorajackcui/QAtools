from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from phraseloom import workbook_schema as schema


def _write_todo_workbook(path: Path, rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = schema.TO_TRANSLATE_SHEET
    ws.append(schema.TO_TRANSLATE_COLUMNS)
    for row in rows:
        ws.append([row.get(column) for column in schema.TO_TRANSLATE_COLUMNS])
    wb.save(path)


def _unit_header_index(headers: list[object], column: str) -> int:
    aliases = {
        schema.SOURCE_UNIT_COLUMN: (schema.SOURCE_UNIT_COLUMN, schema.SOURCE_COLUMN),
        schema.TARGET_UNIT_COLUMN: (schema.TARGET_UNIT_COLUMN, schema.TARGET_COLUMN),
    }
    for candidate in aliases.get(column, (column,)):
        if candidate in headers:
            return headers.index(candidate)
    raise ValueError(f"{column!r} is not in list")


def _with_unit_aliases(row: dict[str, object]) -> dict[str, object]:
    if schema.SOURCE_COLUMN in row and schema.SOURCE_UNIT_COLUMN not in row:
        row[schema.SOURCE_UNIT_COLUMN] = row[schema.SOURCE_COLUMN]
    if schema.TARGET_COLUMN in row and schema.TARGET_UNIT_COLUMN not in row:
        row[schema.TARGET_UNIT_COLUMN] = row[schema.TARGET_COLUMN]
    return row


def _rows_by_header(path: Path, sheet_name: str) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        return [
            _with_unit_aliases(
                {
                    str(headers[index]): value
                    for index, value in enumerate(row)
                    if headers[index] is not None
                }
            )
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        wb.close()


def _headers(path: Path, sheet_name: str) -> list[str]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        return [str(cell.value) for cell in ws[1]]
    finally:
        wb.close()


def _sheet_state(path: Path, sheet_name: str) -> str:
    wb = load_workbook(path, data_only=True)
    try:
        return wb[sheet_name].sheet_state
    finally:
        wb.close()


def _is_column_hidden(path: Path, sheet_name: str, header: str) -> bool:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        column_letter = get_column_letter(headers.index(header) + 1)
        return bool(ws.column_dimensions[column_letter].hidden)
    finally:
        wb.close()


def _set_first_todo_target(path: Path, target: str) -> None:
    wb = load_workbook(path)
    try:
        ws = wb[schema.TO_TRANSLATE_SHEET]
        headers = [cell.value for cell in ws[1]]
        ws.cell(
            row=2,
            column=_unit_header_index(headers, schema.TARGET_UNIT_COLUMN) + 1,
        ).value = target
        wb.save(path)
    finally:
        wb.close()


def _set_first_non_related_target(path: Path, target: str) -> None:
    wb = load_workbook(path)
    try:
        ws = wb[schema.NON_RELATED_UNITS_SHEET]
        headers = [cell.value for cell in ws[1]]
        ws.cell(
            row=2,
            column=_unit_header_index(headers, schema.TARGET_UNIT_COLUMN) + 1,
        ).value = target
        wb.save(path)
    finally:
        wb.close()


def _set_first_related_original_index(path: Path, value: object) -> None:
    wb = load_workbook(path)
    try:
        ws = wb[schema.RELATED_UNITS_SHEET]
        headers = [cell.value for cell in ws[1]]
        ws.cell(
            row=2,
            column=headers.index(schema.ORIGINAL_INDEX_COLUMN) + 1,
        ).value = value
        wb.save(path)
    finally:
        wb.close()


def _set_first_entity_map_original_index(path: Path, value: object) -> None:
    wb = load_workbook(path)
    try:
        ws = wb[schema.ENTITY_MAP_SHEET]
        headers = [cell.value for cell in ws[1]]
        ws.cell(
            row=2,
            column=headers.index(schema.ORIGINAL_INDEX_COLUMN) + 1,
        ).value = value
        wb.save(path)
    finally:
        wb.close()


def _write_entity_tm_workbook(path: Path) -> None:
    wb = Workbook()
    structures = wb.active
    structures.title = "entity_structures"
    structures.append(
        [
            "structure_id",
            "source_structure",
            "target_structure",
            "coverage_count",
            "confidence",
            "risk",
            "status",
            "sample_sources",
            "row_numbers",
            "warning",
        ]
    )
    structures.append(
        [
            "ES9001",
            "{entity1} launched an attack and dealt damage.",
            "{entity1} launched a localized attack and dealt localized damage.",
            2,
            1,
            None,
            "ready",
            None,
            "1,2",
            None,
        ]
    )
    terms = wb.create_sheet("entity_terms")
    terms.append(
        [
            "term_id",
            "source_entity",
            "target_entity",
            "occurrence_count",
            "structure_ids",
            "status",
            "warning",
        ]
    )
    terms.append(["ET9001", "Squirtle", "Carapuce", 1, "ES9001", "ready", None])
    terms.append(["ET9002", "Pikachu", "Pikachu", 1, "ES9001", "ready", None])
    wb.save(path)


def _write_tm_pairs_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = schema.TM_PAIRS_SHEET
    ws.append(schema.TM_PAIR_COLUMNS)
    rows = [
        (
            "TM00001",
            "segment",
            "Squirtle launched an attack and dealt damage.",
            "Carapuce launched a localized attack and dealt localized damage.",
        ),
        (
            "TM00002",
            "segment",
            "Pikachu launched an attack and dealt damage.",
            "Pikachu launched a localized attack and dealt localized damage.",
        ),
        (
            "TM00003",
            "segment",
            "Bulbasaur launched an attack and dealt damage.",
            "Bulbizarre launched a localized attack and dealt localized damage.",
        ),
    ]
    for tm_id, unit_type, source_unit, target_unit in rows:
        ws.append(
            [
                tm_id,
                unit_type,
                source_unit,
                target_unit,
                1,
                1,
                None,
                source_unit,
                target_unit,
                None,
                None,
            ]
        )
    wb.save(path)


def _complete_entity_workbook(
    path: Path,
    *,
    term_targets: dict[str, str],
) -> None:
    wb = load_workbook(path)
    try:
        structures = wb["entity_structures"]
        structure_headers = [cell.value for cell in structures[1]]
        structures.cell(
            row=2,
            column=structure_headers.index("target_structure") + 1,
        ).value = "{entity1} launched a localized attack and dealt localized damage."

        terms = wb["entity_terms"]
        term_headers = [cell.value for cell in terms[1]]
        source_index = term_headers.index("source_entity") + 1
        target_index = term_headers.index("target_entity") + 1
        for row_number in range(2, terms.max_row + 1):
            source_entity = terms.cell(row=row_number, column=source_index).value
            target_entity = term_targets.get(str(source_entity))
            if target_entity:
                terms.cell(row=row_number, column=target_index).value = target_entity
        wb.save(path)
    finally:
        wb.close()


def _complete_entity_tables(
    path: Path,
    *,
    term_targets: dict[str, str],
) -> None:
    wb = load_workbook(path)
    try:
        structures = wb[schema.ENTITY_STRUCTURES_SHEET]
        structure_headers = [cell.value for cell in structures[1]]
        structures.cell(
            row=2,
            column=structure_headers.index(schema.TARGET_STRUCTURE_COLUMN) + 1,
        ).value = "{entity1} launched a localized attack and dealt localized damage."

        terms = wb[schema.ENTITY_TERMS_SHEET]
        term_headers = [cell.value for cell in terms[1]]
        source_index = term_headers.index(schema.SOURCE_ENTITY_COLUMN) + 1
        target_index = term_headers.index(schema.TARGET_ENTITY_COLUMN) + 1
        for row_number in range(2, terms.max_row + 1):
            source_entity = terms.cell(row=row_number, column=source_index).value
            target_entity = term_targets.get(str(source_entity))
            if target_entity:
                terms.cell(row=row_number, column=target_index).value = target_entity
        wb.save(path)
    finally:
        wb.close()


class EntityWorkflowTests(unittest.TestCase):
    def test_split_creates_parallel_entity_and_non_entity_workbooks(self):
        from phraseloom.entity_workflow import split_entity_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "target_todo.xlsx"
            entity_path = tmp_path / "target_entity_related.xlsx"
            non_entity_path = tmp_path / "target_not_entity_related.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                        schema.COVERAGE_COUNT_COLUMN: 1,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                        schema.COVERAGE_COUNT_COLUMN: 1,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Login failed.",
                        schema.TARGET_UNIT_COLUMN: None,
                        schema.COVERAGE_COUNT_COLUMN: 1,
                    },
                ],
            )

            stats = split_entity_workbook(
                todo_path,
                entity_path,
                non_entity_path,
                min_group_size=2,
            )

            self.assertEqual(stats["entity_unit_count"], 2)
            self.assertEqual(stats["non_entity_unit_count"], 1)

            entity_todo_rows = _rows_by_header(entity_path, schema.TO_TRANSLATE_SHEET)
            self.assertEqual(
                [row[schema.UNIT_ID_COLUMN] for row in entity_todo_rows],
                ["U0001", "U0002"],
            )
            self.assertEqual(
                [row["original_index"] for row in entity_todo_rows],
                [1, 2],
            )

            structures = _rows_by_header(entity_path, "entity_structures")
            self.assertEqual(len(structures), 1)
            self.assertEqual(
                structures[0]["source_structure"],
                "{entity1} launched an attack and dealt damage.",
            )

            terms = _rows_by_header(entity_path, "entity_terms")
            self.assertEqual(
                sorted(row["source_entity"] for row in terms),
                ["Pikachu", "Squirtle"],
            )

            source_map = _rows_by_header(entity_path, "entity_source_map")
            self.assertEqual(
                [row[schema.UNIT_ID_COLUMN] for row in source_map],
                ["U0001", "U0002"],
            )
            self.assertIn('"entity1": "Squirtle"', source_map[0]["entities_json"])

            non_entity_rows = _rows_by_header(non_entity_path, schema.TO_TRANSLATE_SHEET)
            self.assertEqual(len(non_entity_rows), 1)
            self.assertEqual(non_entity_rows[0][schema.UNIT_ID_COLUMN], "U0003")
            self.assertEqual(non_entity_rows[0]["original_index"], 3)

    def test_prefill_copies_matching_structure_and_terms(self):
        from phraseloom.entity_workflow import (
            prefill_entity_workbook,
            split_entity_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "target_todo.xlsx"
            entity_path = tmp_path / "target_entity_related.xlsx"
            non_entity_path = tmp_path / "target_not_entity_related.xlsx"
            tm_path = tmp_path / "entity_tm.xlsx"
            prefilled_path = tmp_path / "target_entity_prefilled.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                        schema.COVERAGE_COUNT_COLUMN: 1,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                        schema.COVERAGE_COUNT_COLUMN: 1,
                    },
                ],
            )
            _write_entity_tm_workbook(tm_path)
            split_entity_workbook(
                todo_path,
                entity_path,
                non_entity_path,
                min_group_size=2,
            )

            stats = prefill_entity_workbook(entity_path, tm_path, prefilled_path)

            self.assertEqual(stats["prefilled_structure_count"], 1)
            self.assertEqual(stats["prefilled_term_count"], 2)
            structures = _rows_by_header(prefilled_path, "entity_structures")
            self.assertEqual(
                structures[0]["target_structure"],
                "{entity1} launched a localized attack and dealt localized damage.",
            )
            terms = _rows_by_header(prefilled_path, "entity_terms")
            targets = {
                row["source_entity"]: row["target_entity"]
                for row in terms
            }
            self.assertEqual(
                targets,
                {"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )
            self.assertEqual({row["status"] for row in terms}, {"ready"})

    def test_extract_entity_tm_workbook_reads_tm_pairs(self):
        from phraseloom.entity_workflow import extract_entity_tm_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_pairs_path = tmp_path / "tm_pairs.xlsx"
            entity_tm_path = tmp_path / "entity_tm.xlsx"
            _write_tm_pairs_workbook(tm_pairs_path)

            stats = extract_entity_tm_workbook(
                tm_pairs_path,
                entity_tm_path,
                min_group_size=2,
            )

            self.assertEqual(stats["entity_structure_count"], 1)
            self.assertEqual(stats["entity_term_count"], 3)
            structures = _rows_by_header(entity_tm_path, "entity_structures")
            self.assertEqual(
                structures[0]["source_structure"],
                "{entity1} launched an attack and dealt damage.",
            )
            self.assertEqual(
                structures[0]["target_structure"],
                "{entity1} launched a localized attack and dealt localized damage.",
            )
            terms = _rows_by_header(entity_tm_path, "entity_terms")
            self.assertEqual(
                {
                    row["source_entity"]: row["target_entity"]
                    for row in terms
                },
                {
                    "Bulbasaur": "Bulbizarre",
                    "Pikachu": "Pikachu",
                    "Squirtle": "Carapuce",
                },
            )

    def test_fill_writes_ready_entity_targets(self):
        from phraseloom.entity_workflow import (
            fill_entity_workbook,
            split_entity_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "target_todo.xlsx"
            entity_path = tmp_path / "target_entity_related.xlsx"
            filled_path = tmp_path / "target_entity_filled.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            split_entity_workbook(
                todo_path,
                entity_path,
                tmp_path / "target_not_entity_related.xlsx",
                min_group_size=2,
            )
            _complete_entity_workbook(
                entity_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )

            stats = fill_entity_workbook(entity_path, filled_path)

            self.assertEqual(stats["filled_entity_unit_count"], 2)
            todo_rows = _rows_by_header(filled_path, schema.TO_TRANSLATE_SHEET)
            self.assertEqual(
                [row[schema.TARGET_UNIT_COLUMN] for row in todo_rows],
                [
                    "Carapuce launched a localized attack and dealt localized damage.",
                    "Pikachu launched a localized attack and dealt localized damage.",
                ],
            )
            source_map = _rows_by_header(filled_path, "entity_source_map")
            self.assertEqual(
                [row["fill_status"] for row in source_map],
                ["filled", "filled"],
            )
            self.assertEqual(
                source_map[0]["preview_target"],
                "Carapuce launched a localized attack and dealt localized damage.",
            )

    def test_fill_blocks_when_entity_translation_is_missing(self):
        from phraseloom.entity_workflow import (
            fill_entity_workbook,
            split_entity_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "target_todo.xlsx"
            entity_path = tmp_path / "target_entity_related.xlsx"
            filled_path = tmp_path / "target_entity_filled.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            split_entity_workbook(
                todo_path,
                entity_path,
                tmp_path / "target_not_entity_related.xlsx",
                min_group_size=2,
            )
            _complete_entity_workbook(
                entity_path,
                term_targets={"Squirtle": "Carapuce"},
            )

            stats = fill_entity_workbook(entity_path, filled_path)

            self.assertEqual(stats["filled_entity_unit_count"], 1)
            todo_rows = _rows_by_header(filled_path, schema.TO_TRANSLATE_SHEET)
            self.assertEqual(
                todo_rows[0][schema.TARGET_UNIT_COLUMN],
                "Carapuce launched a localized attack and dealt localized damage.",
            )
            self.assertIsNone(todo_rows[1][schema.TARGET_UNIT_COLUMN])
            source_map = _rows_by_header(filled_path, "entity_source_map")
            self.assertEqual(source_map[1]["fill_status"], "missing_entity_translation")
            self.assertIn("Pikachu", source_map[1]["warning"])

    def test_merge_restores_original_order_and_targets(self):
        from phraseloom.entity_workflow import (
            fill_entity_workbook,
            merge_entity_workbooks,
            split_entity_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "target_todo.xlsx"
            entity_path = tmp_path / "target_entity_related.xlsx"
            non_entity_path = tmp_path / "target_not_entity_related.xlsx"
            filled_entity_path = tmp_path / "target_entity_filled.xlsx"
            merged_path = tmp_path / "target_merged_todo.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Login failed.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            split_entity_workbook(
                todo_path,
                entity_path,
                non_entity_path,
                min_group_size=2,
            )
            _complete_entity_workbook(
                entity_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )
            fill_entity_workbook(entity_path, filled_entity_path)
            _set_first_todo_target(non_entity_path, "Login failed localized.")

            stats = merge_entity_workbooks(
                filled_entity_path,
                non_entity_path,
                merged_path,
            )

            self.assertEqual(stats["merged_unit_count"], 3)
            self.assertNotIn(
                "original_index",
                _headers(merged_path, schema.TO_TRANSLATE_SHEET),
            )
            merged_rows = _rows_by_header(merged_path, schema.TO_TRANSLATE_SHEET)
            self.assertEqual(
                [row[schema.UNIT_ID_COLUMN] for row in merged_rows],
                ["U0001", "U0002", "U0003"],
            )
            self.assertEqual(
                [row[schema.TARGET_UNIT_COLUMN] for row in merged_rows],
                [
                    "Carapuce launched a localized attack and dealt localized damage.",
                    "Login failed localized.",
                    "Pikachu launched a localized attack and dealt localized damage.",
                ],
            )


class EntityPackWorkflowCliTests(unittest.TestCase):
    def test_top_level_help_lists_simplified_entity_commands(self):
        from contextlib import redirect_stdout
        from io import StringIO

        from phraseloom.cli import _dispatch

        stream = StringIO()
        with redirect_stdout(stream):
            self.assertEqual(_dispatch(["--help"]), 0)

        help_text = stream.getvalue()
        self.assertIn("phraseloom entity-tm TM_REUSABLE_UNITS.xlsx", help_text)
        self.assertIn("phraseloom entity-prepare TRANSLATOR_WORKBOOK.xlsx", help_text)
        self.assertIn("phraseloom entity-fill-pack ENTITY_PACK.xlsx", help_text)
        self.assertIn("phraseloom entity-merge-pack FILLED_ENTITY_PACK.xlsx", help_text)
        self.assertIn("Advanced entity commands:", help_text)
        self.assertIn("phraseloom entity-split TRANSLATOR_WORKBOOK.xlsx", help_text)

    def test_entity_interactive_aliases_open_entity_menu(self):
        from contextlib import redirect_stdout
        from io import StringIO
        from unittest.mock import patch

        from phraseloom.cli import _dispatch

        for command in ("entity", "entity-interactive"):
            with self.subTest(command=command):
                stream = StringIO()
                with patch("builtins.input", side_effect=["q"]), redirect_stdout(stream):
                    self.assertEqual(_dispatch([command]), 0)

                menu = stream.getvalue()
                self.assertIn("Entity Workflow", menu)
                self.assertIn("1) Build entity memory from TM reusable units", menu)
                self.assertIn("4) Merge filled entity pack back to translator todo", menu)

    def test_top_level_interactive_option_opens_entity_menu(self):
        from contextlib import redirect_stdout
        from io import StringIO
        from unittest.mock import patch

        from phraseloom.cli import _dispatch

        stream = StringIO()
        with patch("builtins.input", side_effect=["a", "1", "q"]), redirect_stdout(stream):
            self.assertEqual(_dispatch([]), 0)

        menu = stream.getvalue()
        self.assertIn("a) Advanced tools", menu)
        self.assertIn("Advanced Tools", menu)
        self.assertIn("1) Entity workflow", menu)
        self.assertIn("Entity Workflow", menu)

    def test_entity_interactive_step_1_writes_memory_workbook(self):
        from contextlib import redirect_stdout
        from io import StringIO
        from unittest.mock import patch

        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_pairs_path = tmp_path / "tm_pairs.xlsx"
            output_path = tmp_path / "entity_memory.xlsx"
            _write_tm_pairs_workbook(tm_pairs_path)

            answers = iter(["1", str(tm_pairs_path), str(output_path), "2"])
            with (
                patch("builtins.input", side_effect=lambda _="": next(answers)),
                redirect_stdout(StringIO()),
            ):
                self.assertEqual(_dispatch(["entity"]), 0)

            self.assertTrue(output_path.exists())
            self.assertEqual(
                _headers(output_path, schema.ENTITY_STRUCTURES_SHEET)[0],
                schema.STRUCTURE_ID_COLUMN,
            )

    def test_entity_interactive_step_2_writes_pack_with_optional_memory(self):
        from contextlib import redirect_stdout
        from io import StringIO
        from unittest.mock import patch

        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            memory_path = tmp_path / "entity_memory.xlsx"
            output_path = tmp_path / "entity_pack.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Login failed.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            _write_entity_tm_workbook(memory_path)

            answers = iter(["2", str(todo_path), str(memory_path), str(output_path), "2"])
            with (
                patch("builtins.input", side_effect=lambda _="": next(answers)),
                redirect_stdout(StringIO()),
            ):
                self.assertEqual(_dispatch(["entity"]), 0)

            self.assertTrue(output_path.exists())
            structures = _rows_by_header(output_path, schema.ENTITY_STRUCTURES_SHEET)
            self.assertEqual(
                structures[0][schema.TARGET_STRUCTURE_COLUMN],
                "{entity1} launched a localized attack and dealt localized damage.",
            )

    def test_entity_interactive_step_3_fills_entity_pack(self):
        from contextlib import redirect_stdout
        from io import StringIO
        from unittest.mock import patch

        from phraseloom.cli import _dispatch
        from phraseloom.entity_workflow import prepare_entity_pack_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "entity_pack.xlsx"
            output_path = tmp_path / "entity_pack_filled.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(todo_path, pack_path, min_group_size=2)
            _complete_entity_tables(
                pack_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )

            answers = iter(["3", str(pack_path), str(output_path)])
            with (
                patch("builtins.input", side_effect=lambda _="": next(answers)),
                redirect_stdout(StringIO()),
            ):
                self.assertEqual(_dispatch(["entity"]), 0)

            related_rows = _rows_by_header(output_path, schema.RELATED_UNITS_SHEET)
            self.assertEqual(
                [row[schema.TARGET_UNIT_COLUMN] for row in related_rows],
                [
                    "Carapuce launched a localized attack and dealt localized damage.",
                    "Pikachu launched a localized attack and dealt localized damage.",
                ],
            )

    def test_entity_interactive_step_4_merges_filled_entity_pack(self):
        from contextlib import redirect_stdout
        from io import StringIO
        from unittest.mock import patch

        from phraseloom.cli import _dispatch
        from phraseloom.entity_workflow import (
            fill_entity_pack_workbook,
            prepare_entity_pack_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "entity_pack.xlsx"
            filled_path = tmp_path / "entity_pack_filled.xlsx"
            output_path = tmp_path / "merged_todo.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Login failed.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(todo_path, pack_path, min_group_size=2)
            _complete_entity_tables(
                pack_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )
            _set_first_non_related_target(pack_path, "Login failed localized.")
            fill_entity_pack_workbook(pack_path, filled_path)

            answers = iter(["4", str(filled_path), str(output_path)])
            with (
                patch("builtins.input", side_effect=lambda _="": next(answers)),
                redirect_stdout(StringIO()),
            ):
                self.assertEqual(_dispatch(["entity"]), 0)

            merged_rows = _rows_by_header(output_path, schema.TO_TRANSLATE_SHEET)
            self.assertEqual(
                [row[schema.TARGET_UNIT_COLUMN] for row in merged_rows],
                [
                    "Carapuce launched a localized attack and dealt localized damage.",
                    "Login failed localized.",
                    "Pikachu launched a localized attack and dealt localized damage.",
                ],
            )

    def test_entity_tm_cli_writes_memory_workbook_to_l10n(self):
        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_pairs_path = tmp_path / "tm_pairs.xlsx"
            _write_tm_pairs_workbook(tm_pairs_path)

            self.assertEqual(
                _dispatch(
                    [
                        "entity-tm",
                        str(tm_pairs_path),
                        "--min-group-size",
                        "2",
                    ]
                ),
                0,
            )

            entity_memory_path = (
                tmp_path
                / "tm_pairs_l10n"
                / "tm_pairs_entity_memory.xlsx"
            )
            self.assertTrue(entity_memory_path.exists())
            self.assertEqual(
                _headers(entity_memory_path, schema.ENTITY_STRUCTURES_SHEET),
                [
                    schema.STRUCTURE_ID_COLUMN,
                    schema.SOURCE_STRUCTURE_COLUMN,
                    schema.TARGET_STRUCTURE_COLUMN,
                    schema.COVERAGE_COUNT_COLUMN,
                    schema.CONFIDENCE_COLUMN,
                    schema.RISK_COLUMN,
                    schema.STATUS_COLUMN,
                    schema.SAMPLE_SOURCES_COLUMN,
                    schema.ROW_NUMBERS_COLUMN,
                    schema.WARNING_COLUMN,
                ],
            )
            self.assertEqual(
                _headers(entity_memory_path, schema.ENTITY_TERMS_SHEET),
                [
                    schema.TERM_ID_COLUMN,
                    schema.SOURCE_ENTITY_COLUMN,
                    schema.TARGET_ENTITY_COLUMN,
                    schema.OCCURRENCE_COUNT_COLUMN,
                    schema.STRUCTURE_IDS_COLUMN,
                    schema.STATUS_COLUMN,
                    schema.WARNING_COLUMN,
                ],
            )

    def test_entity_prepare_cli_writes_single_pack_with_prefill(self):
        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            entity_memory_path = tmp_path / "tm_entity_memory.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                        schema.CONTEXT_COLUMN: "battle starter context",
                        schema.COVERAGE_COUNT_COLUMN: 1,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Login failed.",
                        schema.TARGET_UNIT_COLUMN: None,
                        schema.CONTEXT_COLUMN: "login menu context",
                        schema.COVERAGE_COUNT_COLUMN: 1,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                        schema.CONTEXT_COLUMN: "electric starter context",
                        schema.COVERAGE_COUNT_COLUMN: 1,
                    },
                ],
            )
            _write_entity_tm_workbook(entity_memory_path)
            wb = load_workbook(todo_path)
            try:
                ws = wb[schema.TO_TRANSLATE_SHEET]
                original_index_column = ws.max_column + 1
                ws.cell(
                    row=1,
                    column=original_index_column,
                ).value = schema.ORIGINAL_INDEX_COLUMN
                for row_number, original_index in enumerate([1, 2, 3], start=2):
                    ws.cell(
                        row=row_number,
                        column=original_index_column,
                    ).value = original_index
                wb.save(todo_path)
            finally:
                wb.close()

            self.assertEqual(
                _dispatch(
                    [
                        "entity-prepare",
                        str(todo_path),
                        "--tm",
                        str(entity_memory_path),
                        "--min-group-size",
                        "2",
                    ]
                ),
                0,
            )

            pack_path = (
                tmp_path
                / "source_translator_todo_l10n"
                / "source_translator_todo_entity_pack.xlsx"
            )
            self.assertTrue(pack_path.exists())
            wb = load_workbook(pack_path, data_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames[:5],
                    [
                        schema.RELATED_UNITS_SHEET,
                        schema.NON_RELATED_UNITS_SHEET,
                        schema.ENTITY_STRUCTURES_SHEET,
                        schema.ENTITY_TERMS_SHEET,
                        schema.ENTITY_MAP_SHEET,
                    ],
                )
            finally:
                wb.close()

            self.assertEqual(_sheet_state(pack_path, schema.ENTITY_MAP_SHEET), "hidden")
            self.assertEqual(
                _sheet_state(pack_path, schema.METADATA_SHEET),
                "hidden",
            )
            related_rows = _rows_by_header(pack_path, schema.RELATED_UNITS_SHEET)
            non_related_rows = _rows_by_header(pack_path, schema.NON_RELATED_UNITS_SHEET)
            self.assertEqual(
                [row[schema.UNIT_ID_COLUMN] for row in related_rows],
                ["U0001", "U0003"],
            )
            self.assertEqual(
                [row[schema.UNIT_ID_COLUMN] for row in non_related_rows],
                ["U0002"],
            )
            structure_headers = _headers(pack_path, schema.ENTITY_STRUCTURES_SHEET)
            self.assertNotIn(schema.STATUS_COLUMN, structure_headers)
            self.assertIn(schema.SAMPLE_CONTEXT_COLUMN, structure_headers)
            self.assertTrue(
                _is_column_hidden(
                    pack_path,
                    schema.ENTITY_STRUCTURES_SHEET,
                    schema.CONFIDENCE_COLUMN,
                )
            )
            self.assertTrue(
                _is_column_hidden(
                    pack_path,
                    schema.ENTITY_STRUCTURES_SHEET,
                    schema.RISK_COLUMN,
                )
            )
            term_headers = _headers(pack_path, schema.ENTITY_TERMS_SHEET)
            self.assertNotIn(schema.STATUS_COLUMN, term_headers)
            self.assertIn(schema.SAMPLE_SOURCES_COLUMN, term_headers)
            self.assertIn(schema.SAMPLE_CONTEXT_COLUMN, term_headers)
            for sheet_name in [
                schema.RELATED_UNITS_SHEET,
                schema.NON_RELATED_UNITS_SHEET,
            ]:
                self.assertIn(schema.ORIGINAL_INDEX_COLUMN, _headers(pack_path, sheet_name))
                self.assertTrue(
                    _is_column_hidden(
                        pack_path,
                        sheet_name,
                        schema.ORIGINAL_INDEX_COLUMN,
                    )
                )
            entity_map = _rows_by_header(pack_path, schema.ENTITY_MAP_SHEET)
            self.assertEqual(
                [
                    (
                        row[schema.ORIGINAL_INDEX_COLUMN],
                        row[schema.UNIT_ID_COLUMN],
                        row[schema.STRUCTURE_ID_COLUMN],
                        row[schema.ENTITIES_JSON_COLUMN],
                    )
                    for row in entity_map
                ],
                [
                    (1, "U0001", "ES0001", '{"entity1": "Squirtle"}'),
                    (3, "U0003", "ES0001", '{"entity1": "Pikachu"}'),
                ],
            )
            metadata = _rows_by_header(pack_path, schema.METADATA_SHEET)
            self.assertIn(
                {
                    schema.KEY_COLUMN: schema.SCHEMA_VERSION_KEY,
                    schema.VALUE_COLUMN: schema.SCHEMA_VERSION,
                },
                metadata,
            )
            structures = _rows_by_header(pack_path, schema.ENTITY_STRUCTURES_SHEET)
            self.assertEqual(
                structures[0][schema.TARGET_STRUCTURE_COLUMN],
                "{entity1} launched a localized attack and dealt localized damage.",
            )
            self.assertIn(
                "battle starter context",
                structures[0][schema.SAMPLE_CONTEXT_COLUMN],
            )
            self.assertIn(
                "electric starter context",
                structures[0][schema.SAMPLE_CONTEXT_COLUMN],
            )
            terms = _rows_by_header(pack_path, schema.ENTITY_TERMS_SHEET)
            self.assertEqual(
                {
                    row[schema.SOURCE_ENTITY_COLUMN]: row[schema.TARGET_ENTITY_COLUMN]
                    for row in terms
                },
                {"Pikachu": "Pikachu", "Squirtle": "Carapuce"},
            )
            terms_by_source = {row[schema.SOURCE_ENTITY_COLUMN]: row for row in terms}
            self.assertEqual(
                terms_by_source["Squirtle"][schema.SAMPLE_SOURCES_COLUMN],
                "Squirtle launched an attack and dealt damage.",
            )
            self.assertEqual(
                terms_by_source["Squirtle"][schema.SAMPLE_CONTEXT_COLUMN],
                "battle starter context",
            )

    def test_entity_fill_pack_cli_writes_targets_to_related_units(self):
        from phraseloom.entity_workflow import prepare_entity_pack_workbook
        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Login failed.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
            )
            _complete_entity_tables(
                pack_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )

            self.assertEqual(_dispatch(["entity-fill-pack", str(pack_path)]), 0)

            filled_pack_path = tmp_path / "source_entity_pack_filled.xlsx"
            self.assertTrue(filled_pack_path.exists())
            related_rows = _rows_by_header(filled_pack_path, schema.RELATED_UNITS_SHEET)
            self.assertEqual(
                [row[schema.TARGET_UNIT_COLUMN] for row in related_rows],
                [
                    "Carapuce launched a localized attack and dealt localized damage.",
                    "Pikachu launched a localized attack and dealt localized damage.",
                ],
            )
            non_related_rows = _rows_by_header(
                filled_pack_path,
                schema.NON_RELATED_UNITS_SHEET,
            )
            self.assertEqual(
                [row[schema.TARGET_UNIT_COLUMN] for row in non_related_rows],
                [None],
            )
            map_rows = _rows_by_header(filled_pack_path, schema.ENTITY_MAP_SHEET)
            self.assertEqual(
                [row[schema.FILL_STATUS_COLUMN] for row in map_rows],
                ["filled", "filled"],
            )

    def test_entity_pack_unit_sheets_use_translator_todo_headers(self):
        from phraseloom.entity_workflow import prepare_entity_pack_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "legacy_source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = schema.TO_TRANSLATE_SHEET
            legacy_headers = [
                schema.UNIT_ID_COLUMN,
                schema.UNIT_TYPE_COLUMN,
                schema.SOURCE_UNIT_COLUMN,
                schema.TARGET_UNIT_COLUMN,
                schema.COVERAGE_COUNT_COLUMN,
                schema.TARGET_UNIT_SOURCE_COLUMN,
            ]
            ws.append(legacy_headers)
            for row in [
                {
                    schema.UNIT_ID_COLUMN: "U0001",
                    schema.UNIT_TYPE_COLUMN: "segment",
                    schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                    schema.TARGET_UNIT_COLUMN: None,
                    schema.COVERAGE_COUNT_COLUMN: 1,
                },
                {
                    schema.UNIT_ID_COLUMN: "U0002",
                    schema.UNIT_TYPE_COLUMN: "segment",
                    schema.SOURCE_UNIT_COLUMN: "Login failed.",
                    schema.TARGET_UNIT_COLUMN: None,
                    schema.COVERAGE_COUNT_COLUMN: 1,
                },
                {
                    schema.UNIT_ID_COLUMN: "U0003",
                    schema.UNIT_TYPE_COLUMN: "segment",
                    schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                    schema.TARGET_UNIT_COLUMN: None,
                    schema.COVERAGE_COUNT_COLUMN: 1,
                },
            ]:
                ws.append([row.get(column) for column in legacy_headers])
            wb.save(todo_path)

            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
            )

            expected_headers = [
                schema.ORIGINAL_INDEX_COLUMN,
                schema.UNIT_ID_COLUMN,
                schema.UNIT_TYPE_COLUMN,
                schema.SOURCE_COLUMN,
                schema.TARGET_COLUMN,
                schema.SAMPLE_SOURCES_COLUMN,
                schema.CONTEXT_COLUMN,
                schema.ROW_NUMBER_COLUMN,
                schema.COVERAGE_COUNT_COLUMN,
                schema.VARIABLES_COLUMN,
                schema.WARNING_COLUMN,
                schema.TRANSLATOR_NOTE_COLUMN,
            ]
            for sheet_name in [
                schema.RELATED_UNITS_SHEET,
                schema.NON_RELATED_UNITS_SHEET,
            ]:
                self.assertEqual(_headers(pack_path, sheet_name), expected_headers)
                self.assertTrue(
                    _is_column_hidden(
                        pack_path,
                        sheet_name,
                        schema.ORIGINAL_INDEX_COLUMN,
                    )
                )

    def test_entity_pack_related_units_groups_same_structure_segments(self):
        from phraseloom.entity_workflow import prepare_entity_pack_workbook
        from phraseloom.models import EntityCluster

        class InterleavedStrategy:
            name = "interleaved"

            def find_clusters(self, rows):
                return [
                    EntityCluster(
                        source_pattern="{entity1} launched an attack.",
                        coverage_count=2,
                        unique_source_count=2,
                        unique_entity_count=2,
                        entity_values=("Pikachu", "Squirtle"),
                        confidence=0.95,
                        risk="",
                        sample_sources=(),
                        sample_targets=(),
                        row_numbers=(2, 4),
                    ),
                    EntityCluster(
                        source_pattern="{entity1} prepares a potion.",
                        coverage_count=2,
                        unique_source_count=2,
                        unique_entity_count=2,
                        entity_values=("Bulbasaur", "Charmander"),
                        confidence=0.95,
                        risk="",
                        sample_sources=(),
                        sample_targets=(),
                        row_numbers=(3, 5),
                    ),
                ]

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Bulbasaur prepares a potion.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0004",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Charmander prepares a potion.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )

            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
                strategy=InterleavedStrategy(),
            )

            related_rows = _rows_by_header(pack_path, schema.RELATED_UNITS_SHEET)
            self.assertEqual(
                [row[schema.UNIT_ID_COLUMN] for row in related_rows],
                ["U0001", "U0003", "U0002", "U0004"],
            )

            structure_by_index = {
                row[schema.ORIGINAL_INDEX_COLUMN]: row[schema.STRUCTURE_ID_COLUMN]
                for row in _rows_by_header(pack_path, schema.ENTITY_MAP_SHEET)
            }
            related_structures = [
                structure_by_index[row[schema.ORIGINAL_INDEX_COLUMN]]
                for row in related_rows
            ]
            self.assertEqual(related_structures, ["ES0001", "ES0001", "ES0002", "ES0002"])

    def test_entity_merge_pack_cli_restores_full_todo_order(self):
        from phraseloom.entity_workflow import (
            fill_entity_pack_workbook,
            prepare_entity_pack_workbook,
        )
        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"
            filled_pack_path = tmp_path / "source_entity_pack_filled.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Login failed.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
            )
            _complete_entity_tables(
                pack_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )
            _set_first_non_related_target(pack_path, "Login failed localized.")
            fill_entity_pack_workbook(pack_path, filled_pack_path)

            self.assertEqual(_dispatch(["entity-merge-pack", str(filled_pack_path)]), 0)

            merged_path = tmp_path / "source_entity_pack_filled_merged_todo.xlsx"
            self.assertTrue(merged_path.exists())
            wb = load_workbook(merged_path, data_only=True)
            try:
                self.assertIn(schema.METADATA_SHEET, wb.sheetnames)
                self.assertNotIn(schema.RELATED_UNITS_SHEET, wb.sheetnames)
                self.assertNotIn(schema.NON_RELATED_UNITS_SHEET, wb.sheetnames)
                self.assertNotIn(schema.ENTITY_STRUCTURES_SHEET, wb.sheetnames)
                self.assertNotIn(schema.ENTITY_TERMS_SHEET, wb.sheetnames)
                self.assertNotIn(schema.ENTITY_MAP_SHEET, wb.sheetnames)
            finally:
                wb.close()
            self.assertNotIn(
                schema.ORIGINAL_INDEX_COLUMN,
                _headers(merged_path, schema.TO_TRANSLATE_SHEET),
            )
            merged_rows = _rows_by_header(merged_path, schema.TO_TRANSLATE_SHEET)
            self.assertEqual(
                [row[schema.UNIT_ID_COLUMN] for row in merged_rows],
                ["U0001", "U0002", "U0003"],
            )
            self.assertEqual(
                [row[schema.TARGET_UNIT_COLUMN] for row in merged_rows],
                [
                    "Carapuce launched a localized attack and dealt localized damage.",
                    "Login failed localized.",
                    "Pikachu launched a localized attack and dealt localized damage.",
                ],
            )

    def test_entity_merge_pack_requires_original_index_values(self):
        from phraseloom.entity_workflow import (
            merge_entity_pack_workbook,
            prepare_entity_pack_workbook,
        )
        from phraseloom.errors import WorkflowError

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"
            merged_path = tmp_path / "source_merged_todo.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
            )
            wb = load_workbook(pack_path)
            try:
                ws = wb[schema.RELATED_UNITS_SHEET]
                headers = [cell.value for cell in ws[1]]
                ws.cell(
                    row=2,
                    column=headers.index(schema.ORIGINAL_INDEX_COLUMN) + 1,
                ).value = None
                wb.save(pack_path)
            finally:
                wb.close()

            with self.assertRaisesRegex(
                WorkflowError,
                "missing required original_index",
            ):
                merge_entity_pack_workbook(pack_path, merged_path)

    def test_entity_merge_pack_reports_invalid_related_original_index(self):
        from phraseloom.entity_workflow import (
            merge_entity_pack_workbook,
            prepare_entity_pack_workbook,
        )
        from phraseloom.errors import WorkflowError

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"
            merged_path = tmp_path / "source_merged_todo.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
            )
            _set_first_related_original_index(pack_path, "oops")

            with self.assertRaisesRegex(
                WorkflowError,
                "invalid original_index",
            ):
                merge_entity_pack_workbook(pack_path, merged_path)

    def test_entity_fill_pack_reports_invalid_related_original_index(self):
        from phraseloom.entity_workflow import (
            fill_entity_pack_workbook,
            prepare_entity_pack_workbook,
        )
        from phraseloom.errors import WorkflowError

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"
            filled_path = tmp_path / "source_entity_pack_filled.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
            )
            _set_first_related_original_index(pack_path, "oops")
            _complete_entity_tables(
                pack_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )

            with self.assertRaisesRegex(
                WorkflowError,
                "invalid original_index",
            ):
                fill_entity_pack_workbook(pack_path, filled_path)

    def test_entity_fill_pack_reports_invalid_entity_map_original_index(self):
        from phraseloom.entity_workflow import (
            fill_entity_pack_workbook,
            prepare_entity_pack_workbook,
        )
        from phraseloom.errors import WorkflowError

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"
            filled_path = tmp_path / "source_entity_pack_filled.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
            )
            _complete_entity_tables(
                pack_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )
            _set_first_entity_map_original_index(pack_path, "oops")

            with self.assertRaisesRegex(
                WorkflowError,
                "invalid original_index",
            ):
                fill_entity_pack_workbook(pack_path, filled_path)

    def test_entity_fill_pack_cli_rejects_output_with_in_place(self):
        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            pack_path = tmp_path / "source_entity_pack.xlsx"
            output_path = tmp_path / "source_entity_pack_filled.xlsx"

            with self.assertRaises(SystemExit) as raised:
                _dispatch(
                    [
                        "entity-fill-pack",
                        str(pack_path),
                        "--in-place",
                        "-o",
                        str(output_path),
                    ]
                )
            self.assertEqual(raised.exception.code, 2)

    def test_entity_fill_pack_cli_in_place_updates_input_pack(self):
        from phraseloom.entity_workflow import prepare_entity_pack_workbook
        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "source_translator_todo.xlsx"
            pack_path = tmp_path / "source_entity_pack.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            prepare_entity_pack_workbook(
                todo_path,
                pack_path,
                min_group_size=2,
            )
            _complete_entity_tables(
                pack_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )

            self.assertEqual(
                _dispatch(["entity-fill-pack", str(pack_path), "--in-place"]),
                0,
            )

            related_rows = _rows_by_header(pack_path, schema.RELATED_UNITS_SHEET)
            self.assertEqual(
                [row[schema.TARGET_UNIT_COLUMN] for row in related_rows],
                [
                    "Carapuce launched a localized attack and dealt localized damage.",
                    "Pikachu launched a localized attack and dealt localized damage.",
                ],
            )

    def test_entity_pack_preserves_self_contained_translation_package(self):
        from phraseloom.entity_workflow import (
            fill_entity_pack_workbook,
            merge_entity_pack_workbook,
            prepare_entity_pack_workbook,
        )
        from phraseloom.workflow import (
            fill_translation_package,
            prepare_translation_package,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_path = tmp_path / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Strings"
            ws.append(["source", "target"])
            ws.append(["Squirtle launched an attack and dealt damage.", None])
            ws.append(["Login failed.", None])
            ws.append(["Pikachu launched an attack and dealt damage.", None])
            ws["A2"].fill = PatternFill("solid", fgColor="00FF00")
            wb.save(source_path)
            wb.close()

            package_stats = prepare_translation_package(source_path)
            package_path = Path(str(package_stats["to_translate_path"]))
            entity_pack = tmp_path / "entity_pack.xlsx"
            filled_pack = tmp_path / "entity_pack_filled.xlsx"
            merged_package = tmp_path / "merged_translator_todo.xlsx"

            prepare_entity_pack_workbook(
                package_path,
                entity_pack,
                min_group_size=2,
            )
            self.assertEqual(_sheet_state(entity_pack, "Strings"), "hidden")
            self.assertEqual(
                _sheet_state(entity_pack, schema.PREFILLED_UNITS_SHEET),
                "visible",
            )

            _complete_entity_tables(
                entity_pack,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )
            _set_first_non_related_target(entity_pack, "Échec de connexion.")
            fill_entity_pack_workbook(entity_pack, filled_pack)
            merge_entity_pack_workbook(filled_pack, merged_package)

            merged = load_workbook(merged_package, data_only=True)
            try:
                self.assertIn(schema.TO_TRANSLATE_SHEET, merged.sheetnames)
                self.assertIn(schema.PREFILLED_UNITS_SHEET, merged.sheetnames)
                self.assertEqual(merged["Strings"].sheet_state, "hidden")
            finally:
                merged.close()

            fill_stats = fill_translation_package(merged_package)
            result = load_workbook(fill_stats["output_path"], data_only=True)
            try:
                self.assertEqual(result.sheetnames, ["Strings"])
                self.assertEqual(result["Strings"]["A2"].fill.fgColor.rgb, "0000FF00")
                self.assertEqual(
                    [row[1] for row in result["Strings"].iter_rows(min_row=2, values_only=True)],
                    [
                        "Carapuce launched a localized attack and dealt localized damage.",
                        "Échec de connexion.",
                        "Pikachu launched a localized attack and dealt localized damage.",
                    ],
                )
            finally:
                result.close()


class EntityWorkflowCliTests(unittest.TestCase):
    def test_entity_extract_tm_cli_reads_tm_pairs(self):
        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_pairs_path = tmp_path / "tm_pairs.xlsx"
            entity_tm_path = tmp_path / "entity_tm.xlsx"
            _write_tm_pairs_workbook(tm_pairs_path)

            self.assertEqual(
                _dispatch(
                    [
                        "entity-extract-tm",
                        str(tm_pairs_path),
                        "-o",
                        str(entity_tm_path),
                        "--min-group-size",
                        "2",
                    ]
                ),
                0,
            )

            structures = _rows_by_header(entity_tm_path, "entity_structures")
            self.assertEqual(
                structures[0]["target_structure"],
                "{entity1} launched a localized attack and dealt localized damage.",
            )

    def test_entity_cli_commands_run_workflow(self):
        from phraseloom.cli import _dispatch

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            todo_path = tmp_path / "target_todo.xlsx"
            entity_path = tmp_path / "target_entity_related.xlsx"
            non_entity_path = tmp_path / "target_not_entity_related.xlsx"
            tm_path = tmp_path / "entity_tm.xlsx"
            prefilled_path = tmp_path / "target_entity_prefilled.xlsx"
            filled_path = tmp_path / "target_entity_filled.xlsx"
            merged_path = tmp_path / "target_merged_todo.xlsx"
            _write_todo_workbook(
                todo_path,
                [
                    {
                        schema.UNIT_ID_COLUMN: "U0001",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Squirtle launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0002",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Login failed.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                    {
                        schema.UNIT_ID_COLUMN: "U0003",
                        schema.UNIT_TYPE_COLUMN: "segment",
                        schema.SOURCE_UNIT_COLUMN: "Pikachu launched an attack and dealt damage.",
                        schema.TARGET_UNIT_COLUMN: None,
                    },
                ],
            )
            _write_entity_tm_workbook(tm_path)

            self.assertEqual(
                _dispatch(
                    [
                        "entity-split",
                        str(todo_path),
                        "--entity-output",
                        str(entity_path),
                        "--non-entity-output",
                        str(non_entity_path),
                        "--min-group-size",
                        "2",
                    ]
                ),
                0,
            )
            self.assertEqual(
                _dispatch(
                    [
                        "entity-prefill",
                        str(entity_path),
                        "--tm",
                        str(tm_path),
                        "-o",
                        str(prefilled_path),
                    ]
                ),
                0,
            )
            _complete_entity_workbook(
                prefilled_path,
                term_targets={"Squirtle": "Carapuce", "Pikachu": "Pikachu"},
            )
            self.assertEqual(
                _dispatch(["entity-fill", str(prefilled_path), "-o", str(filled_path)]),
                0,
            )
            _set_first_todo_target(non_entity_path, "Login failed localized.")
            self.assertEqual(
                _dispatch(
                    [
                        "entity-merge",
                        "--entity",
                        str(filled_path),
                        "--non-entity",
                        str(non_entity_path),
                        "-o",
                        str(merged_path),
                    ]
                ),
                0,
            )

            merged_rows = _rows_by_header(merged_path, schema.TO_TRANSLATE_SHEET)
            self.assertEqual(len(merged_rows), 3)
            self.assertEqual(
                merged_rows[0][schema.TARGET_UNIT_COLUMN],
                "Carapuce launched a localized attack and dealt localized damage.",
            )
            self.assertEqual(
                merged_rows[1][schema.TARGET_UNIT_COLUMN],
                "Login failed localized.",
            )


if __name__ == "__main__":
    unittest.main()
