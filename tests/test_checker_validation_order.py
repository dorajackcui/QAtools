from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from tools.chinese_target_checker.check_chinese_target import (
    process_excel as check_chinese_target,
)
from tools.content_fidelity_checker.check_content_fidelity import (
    process_excel as check_content_fidelity,
)
from tools.line_break_checker.check_line_breaks import (
    process_excel as check_line_breaks,
)
from tools.source_consistency_checker.check_source_consistency import (
    process_excel as check_source_consistency,
)
from tools.tag_placeholder_checker.check_tags_and_placeholders import (
    process_excel as check_tags,
)
from tools.target_consistency_checker.check_target_consistency import (
    process_excel as check_target_consistency,
)
from tools.target_text_checker.check_target_text import (
    process_excel as check_target_text,
)
from tools.term_pair_checker.extract_terms_from_excel import (
    process_excel as check_term_pairs,
)


CHECKERS = (
    (
        "chinese",
        check_chinese_target,
        "tools.chinese_target_checker.check_chinese_target.load_workbook_for_editing",
    ),
    (
        "content",
        check_content_fidelity,
        "tools.content_fidelity_checker.check_content_fidelity.load_workbook_for_editing",
    ),
    (
        "line-break",
        check_line_breaks,
        "tools.line_break_checker.check_line_breaks.load_workbook_for_editing",
    ),
    (
        "source-consistency",
        check_source_consistency,
        "tools.source_consistency_checker.check_source_consistency.load_workbook_for_editing",
    ),
    (
        "tag",
        check_tags,
        "tools.tag_placeholder_checker.check_tags_and_placeholders.load_workbook_for_editing",
    ),
    (
        "target-consistency",
        check_target_consistency,
        "tools.target_consistency_checker.check_target_consistency.load_workbook_for_editing",
    ),
    (
        "target-text",
        check_target_text,
        "tools.target_text_checker.check_target_text.load_workbook_for_editing",
    ),
    (
        "term",
        check_term_pairs,
        "tools.term_pair_checker.extract_terms_from_excel.load_workbook_for_editing",
    ),
)


class CheckerValidationOrderTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.input_path = Path(self.temp_dir.name) / "input.xlsx"
        workbook = Workbook()
        workbook.active.append(["source", "target"])
        workbook.save(self.input_path)
        workbook.close()

    def test_invalid_start_row_is_rejected_before_loading_the_workbook(self) -> None:
        for checker_name, checker, loader_path in CHECKERS:
            with self.subTest(checker=checker_name), patch(
                loader_path,
                side_effect=AssertionError("workbook should not load"),
            ):
                with self.assertRaisesRegex(ValueError, "开始行"):
                    checker(
                        input_file=self.input_path,
                        source_column="A",
                        target_column="B",
                        start_row=0,
                    )

    def test_same_columns_are_rejected_before_loading_the_workbook(self) -> None:
        for checker_name, checker, loader_path in CHECKERS:
            with self.subTest(checker=checker_name), patch(
                loader_path,
                side_effect=AssertionError("workbook should not load"),
            ):
                with self.assertRaisesRegex(ValueError, "不能相同"):
                    checker(
                        input_file=self.input_path,
                        source_column="A",
                        target_column="A",
                    )


if __name__ == "__main__":
    unittest.main()
