from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from tools.term_pair_checker.extract_terms_from_excel import process_excel
from tools.workflow.revision_applier import apply_workflow_revisions
from tools.workflow.workflow_runner import run_workflow


class TermProblemStatusTests(unittest.TestCase):
    def write_input(self, path: Path, rows: tuple[tuple[str, str], ...]) -> None:
        workbook = Workbook()
        try:
            sheet = workbook.active
            sheet.title = "Data"
            sheet.append(["source", "target"])
            for row in rows:
                sheet.append(row)
            workbook.save(path)
        finally:
            workbook.close()

    def test_only_failing_terms_are_marked_without_notes_or_links(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.write_input(input_path, (
                ("[Alpha] [Beta]", "[阿尔法] [贝塔]"),
                ("[Alpha] [Beta]", "[阿尔法]"),
                ("Use Beta", "Wrong"),
            ))
            _, _, _, output, _, problem_count = process_excel(
                input_path, "A", "B", mark_styles=("[]",),
            )
            self.assertEqual(problem_count, 3)
            workbook = load_workbook(output)
            try:
                terms = workbook["术语表"]
                self.assertEqual(terms["F1"].value, "本批次是否有问题")
                self.assertEqual(terms.max_column, 6)
                self.assertEqual(terms.auto_filter.ref, "A1:F3")
                self.assertEqual(terms["F2"].value, "无问题")
                self.assertIsNone(terms["A2"].hyperlink)
                self.assertEqual(terms["F3"].value, "有问题")
                self.assertEqual(terms["F3"].fill.fgColor.rgb, "00FCE4D6")
                for row in terms.iter_rows():
                    for cell in row:
                        self.assertIsNone(cell.comment)
                        self.assertIsNone(cell.hyperlink)
                        self.assertIsNone(cell.font.underline)
                self.assertEqual(workbook["问题列"]["A2"].value, 3)
            finally:
                workbook.close()

    def test_status_handles_separators_case_variants_and_unpaired_terms(self) -> None:
        cases = (
            (
                (("[Alpha、Beta]", "[正确]"), ("[alpha、beta]", "[错误]")),
                "Alpha、Beta",
            ),
            ((("[Orphan]", "No translation"),), "Orphan"),
        )
        for rows, expected_source in cases:
            with self.subTest(source=expected_source), tempfile.TemporaryDirectory() as tmp_dir:
                input_path = Path(tmp_dir) / "input.xlsx"
                self.write_input(input_path, rows)
                _, _, _, output, term_count, _ = process_excel(
                    input_path, "A", "B", mark_styles=("[]",),
                )
                self.assertEqual(term_count, 1)
                workbook = load_workbook(output)
                try:
                    terms = workbook["术语表"]
                    self.assertEqual(terms["C2"].value, expected_source)
                    self.assertEqual(terms["F2"].value, "有问题")
                    self.assertIsNone(terms["F2"].comment)
                    self.assertIsNone(terms["F2"].hyperlink)
                finally:
                    workbook.close()

    def test_clean_or_empty_term_reports_have_no_problem_links(self) -> None:
        for rows in ((("[Alpha]", "[阿尔法]"),), (("Plain text", "Translation"),)):
            with self.subTest(rows=rows), tempfile.TemporaryDirectory() as tmp_dir:
                input_path = Path(tmp_dir) / "input.xlsx"
                self.write_input(input_path, rows)
                _, _, _, output, term_count, problem_count = process_excel(
                    input_path, "A", "B", mark_styles=("[]",),
                )
                self.assertEqual(problem_count, 0)
                workbook = load_workbook(output)
                try:
                    terms = workbook["术语表"]
                    self.assertEqual(terms.max_row, term_count + 1)
                    self.assertEqual(terms["F1"].value, "本批次是否有问题")
                    self.assertIsNone(terms["F1"].comment)
                    for row in range(2, terms.max_row + 1):
                        self.assertEqual(terms.cell(row, 6).value, "无问题")
                        self.assertIsNone(terms.cell(row, 6).comment)
                        self.assertIsNone(terms.cell(row, 6).hyperlink)
                finally:
                    workbook.close()

    def test_workflow_status_preserves_merged_sorted_rows_and_revision_targets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.write_input(input_path, (
                ("[Alpha] [Beta]", "[阿尔法] [贝塔]"),
                ("Use Gamma", "Wrong"),
                ("[Alpha] [Beta]", "[阿尔法]"),
                ("[Gamma]", "[伽马]"),
                ("Other {name}", "残留"),
            ))
            summary = run_workflow(
                input_file=input_path, source_column="A", target_column="B",
                term_mark_styles=("[]",), tag_token_types=("brace",),
            )
            self.assertEqual(summary.term_problem_rows, 2)
            workbook = load_workbook(summary.output_path)
            try:
                self.assertNotIn("问题列", workbook.sheetnames)
                self.assertNotIn("术语问题", workbook.sheetnames)
                review = workbook["问题处理"]
                self.assertEqual([review.cell(row, 1).value for row in range(2, 7)], [2, 4, 3, 5, 6])
                terms = workbook["术语表"]
                self.assertEqual(terms["F2"].value, "无问题")
                self.assertIsNone(terms["A2"].hyperlink)
                for term_row in (3, 4):
                    self.assertEqual(terms.cell(term_row, 6).value, "有问题")
                for row in terms.iter_rows():
                    for cell in row:
                        self.assertIsNone(cell.comment)
                        self.assertIsNone(cell.hyperlink)
                review["D3"] = "[阿尔法] [贝塔]"
                workbook.save(summary.output_path)
            finally:
                workbook.close()

            revision = apply_workflow_revisions(summary.output_path)

            self.assertEqual(revision.revised_count, 1)
            self.assertEqual(revision.conflict_rows, ())
            workbook = load_workbook(revision.output_path)
            try:
                self.assertEqual(workbook.sheetnames, ["Data"])
                self.assertEqual(workbook["Data"]["B4"].value, "[阿尔法] [贝塔]")
                self.assertEqual(workbook["Data"]["B3"].value, "Wrong")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
