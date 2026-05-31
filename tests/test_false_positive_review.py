from __future__ import annotations

import unittest
from pathlib import Path
from subprocess import CompletedProcess
from unittest.mock import patch

from openpyxl import Workbook

from tools.false_positive_review import (
    GLOSSARY_PROBLEM_MAPPING,
    TERM_PAIR_PROBLEM_MAPPING,
    ReviewCluster,
    ReviewDecision,
    ReviewExample,
    apply_false_positive_review_to_sheet,
    build_codex_prompt,
    collect_review_clusters,
    parse_codex_decisions,
    review_cluster_batch_with_codex,
    review_clusters_in_batches,
)


class FalsePositiveReviewTests(unittest.TestCase):
    def test_glossary_review_clusters_include_source_and_target_text(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "术语命中问题"
        worksheet.append(["行号", "问题类型", "source术语", "期望target术语", "source文本", "target文本"])
        worksheet.append([2, "术语未按术语表翻译", "Move", "Capacité", "Move details", "Détails de la capacité"])
        worksheet.append([3, "术语未按术语表翻译", "Move", "Capacité", "Move forward", "Avancer"])

        clusters = collect_review_clusters(worksheet, GLOSSARY_PROBLEM_MAPPING)

        self.assertEqual(len(clusters), 2)
        self.assertEqual([cluster.row_numbers for cluster in clusters], [(2,), (3,)])

    def test_reviews_term_pair_problem_sheet_by_text_specific_cluster(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "问题列"
        worksheet.append(["问题行号", "问题source术语", "预期target术语", "术语来源", "问题简述", "source原文", "target原文"])
        worksheet.append([2, "Lineup", "Équipe", "本批次新增", "target缺少预期术语", "Lineup RCMD:", "Compo conseillée:"])
        worksheet.append([3, "Lineup", "Équipe", "本批次新增", "target缺少预期术语", "No lineup recommendation", "Aucune recommandation de compo"])
        worksheet.append([4, "Advanced", "Avancé", "本批次新增", "target缺少预期术语", "Cannot be advanced", "Ne peut pas passer au rang supérieur"])

        seen_clusters = []

        def reviewer(clusters):
            seen_clusters.extend(clusters)
            return {
                clusters[0].key: ReviewDecision(
                    decision="review",
                    category="需人工确认",
                    confidence="medium",
                    note="项目是否接受 Compo 作为 Équipe 短称需要确认",
                ),
                clusters[1].key: ReviewDecision(
                    decision="review",
                    category="需人工确认",
                    confidence="medium",
                    note="项目是否接受 Compo 作为 Équipe 短称需要确认",
                ),
                clusters[2].key: ReviewDecision(
                    decision="false_positive",
                    category="词性变化/自然改写",
                    confidence="high",
                    note="advanced 在源文中是动词用法，target 使用自然改写",
                ),
            }

        summary = apply_false_positive_review_to_sheet(
            workbook,
            "问题列",
            TERM_PAIR_PROBLEM_MAPPING,
            reviewer=reviewer,
            sample_size=1,
        )

        self.assertEqual(summary.cluster_count, 3)
        self.assertEqual(summary.reviewed_row_count, 3)
        self.assertEqual(len(seen_clusters), 3)
        self.assertEqual(len(seen_clusters[0].examples), 1)
        self.assertEqual(worksheet["H1"].value, "fp_decision")
        self.assertEqual(worksheet["I1"].value, "fp_category")
        self.assertEqual(worksheet["J1"].value, "fp_confidence")
        self.assertEqual(worksheet["K1"].value, "fp_note")
        self.assertEqual(worksheet["L1"].value, "fp_by")
        self.assertEqual(worksheet["H2"].value, "review")
        self.assertEqual(worksheet["H3"].value, "review")
        self.assertEqual(worksheet["H4"].value, "false_positive")
        self.assertEqual(worksheet["L4"].value, "codex")

    def test_reviews_glossary_problem_sheet_with_its_headers(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "术语命中问题"
        worksheet.append(["行号", "问题类型", "source术语", "期望target术语", "source文本", "target文本"])
        worksheet.append([2, "术语未按术语表翻译", "Advanced", "Avancé", "Cannot be advanced", "Ne peut pas passer au rang supérieur"])

        def reviewer(clusters):
            return {
                clusters[0].key: ReviewDecision(
                    decision="false_positive",
                    category="词性变化/自然改写",
                    confidence="high",
                    note="源文是动词用法",
                )
            }

        summary = apply_false_positive_review_to_sheet(
            workbook,
            "术语命中问题",
            GLOSSARY_PROBLEM_MAPPING,
            reviewer=reviewer,
        )

        self.assertEqual(summary.cluster_count, 1)
        self.assertEqual(summary.reviewed_row_count, 1)
        self.assertEqual(worksheet["G1"].value, "fp_decision")
        self.assertEqual(worksheet["G2"].value, "false_positive")
        self.assertEqual(worksheet["H2"].value, "词性变化/自然改写")

    def test_glossary_review_skips_suspected_plural_variant_rows(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "术语命中问题"
        worksheet.append(["行号", "问题类型", "source术语", "期望target术语", "source文本", "target文本"])
        worksheet.append([2, "术语未按术语表翻译", "Advanced", "Avancé", "Cannot be advanced", "Passer au rang supérieur"])
        worksheet.append([3, "术语未按术语表翻译：疑似复数变体", "Reward", "Récompense", "Rewards", "Récompenses"])

        seen_clusters = []

        def reviewer(clusters):
            seen_clusters.extend(clusters)
            return {
                clusters[0].key: ReviewDecision(
                    decision="false_positive",
                    category="词性变化/自然改写",
                    confidence="high",
                    note="源文是动词用法",
                )
            }

        summary = apply_false_positive_review_to_sheet(
            workbook,
            "术语命中问题",
            GLOSSARY_PROBLEM_MAPPING,
            reviewer=reviewer,
        )

        self.assertEqual(summary.cluster_count, 1)
        self.assertEqual(summary.reviewed_row_count, 1)
        self.assertEqual(len(seen_clusters), 1)
        self.assertEqual(seen_clusters[0].issue_type, "术语未按术语表翻译")
        self.assertEqual(worksheet["G2"].value, "false_positive")
        self.assertIsNone(worksheet["G3"].value)

    def test_review_clusters_in_batches_combines_chunk_results(self) -> None:
        clusters = [
            ReviewCluster(
                key=f"cluster-{index}",
                source_term=f"Source {index}",
                expected_target=f"Target {index}",
                issue_type="术语未按术语表翻译",
                source_text="source",
                target_text="target",
                row_numbers=(index + 2,),
                examples=(ReviewExample(source_text="source", target_text="target"),),
            )
            for index in range(5)
        ]
        chunk_sizes = []

        def batch_reviewer(batch):
            chunk_sizes.append(len(batch))
            return {
                cluster.key: ReviewDecision(
                    decision="review",
                    category="需人工确认",
                    confidence="low",
                    note=cluster.source_term,
                )
                for cluster in batch
            }

        decisions = review_clusters_in_batches(clusters, batch_size=2, batch_reviewer=batch_reviewer)

        self.assertEqual(chunk_sizes, [2, 2, 1])
        self.assertEqual(set(decisions), {cluster.key for cluster in clusters})
        self.assertEqual(decisions["cluster-4"].note, "Source 4")

    def test_review_clusters_in_batches_can_retry_missing_cluster_decisions(self) -> None:
        clusters = [
            ReviewCluster(
                key=f"cluster-{index}",
                source_term=f"Source {index}",
                expected_target=f"Target {index}",
                issue_type="术语未按术语表翻译",
                source_text=f"source {index}",
                target_text=f"target {index}",
                row_numbers=(index + 2,),
                examples=(ReviewExample(source_text=f"source {index}", target_text=f"target {index}"),),
            )
            for index in range(3)
        ]
        calls = []

        def batch_reviewer(batch):
            calls.append([cluster.key for cluster in batch])
            if len(batch) > 1:
                return {
                    batch[0].key: ReviewDecision(
                        decision="review",
                        category="需人工确认",
                        confidence="low",
                        note="initial",
                    )
                }
            return {
                batch[0].key: ReviewDecision(
                    decision="false_positive",
                    category="词性变化/自然改写",
                    confidence="high",
                    note="retried",
                )
            }

        decisions = review_clusters_in_batches(
            clusters,
            batch_size=3,
            batch_reviewer=batch_reviewer,
            retry_missing=True,
        )

        self.assertEqual(calls, [["cluster-0", "cluster-1", "cluster-2"], ["cluster-1"], ["cluster-2"]])
        self.assertEqual(set(decisions), {cluster.key for cluster in clusters})
        self.assertEqual(decisions["cluster-0"].note, "initial")
        self.assertEqual(decisions["cluster-1"].note, "retried")

    def test_codex_batch_uses_short_prompt_ids_and_maps_back_to_cluster_keys(self) -> None:
        cluster = ReviewCluster(
            key='["Move","Capacité","术语未按术语表翻译","Move forward","Avancer"]',
            source_term="Move",
            expected_target="Capacité",
            issue_type="术语未按术语表翻译",
            source_text="Move forward",
            target_text="Avancer",
            row_numbers=(2,),
            examples=(ReviewExample(source_text="Move forward", target_text="Avancer"),),
        )

        def fake_run(command, input, **kwargs):
            self.assertIn('"id": "cluster-1"', input)
            self.assertNotIn('"id": "[', input)
            output_path = Path(command[command.index("--output-last-message") + 1])
            output_path.write_text(
                '{"results":[{"id":"cluster-1","decision":"false_positive","category":"词性变化/自然改写","confidence":"high","note":"源文中是普通动词用法"}]}',
                encoding="utf-8",
            )
            return CompletedProcess(command, 0, "", "")

        with patch("tools.codex_runner.subprocess.run", side_effect=fake_run):
            decisions = review_cluster_batch_with_codex([cluster], model="gpt-5.3-codex-spark")

        self.assertEqual(set(decisions), {cluster.key})
        self.assertEqual(decisions[cluster.key].decision, "false_positive")

    def test_codex_prompt_prioritizes_terminology_consistency_over_semantic_equivalence(self) -> None:
        cluster = ReviewCluster(
            key='["Fiery Assault","Assaut Enflammé","术语未按术语表翻译","Fiery Assault deals damage","Assaut flamboyant inflige des dégâts"]',
            source_term="Fiery Assault",
            expected_target="Assaut Enflammé",
            issue_type="术语未按术语表翻译",
            source_text="Fiery Assault deals damage",
            target_text="Assaut flamboyant inflige des dégâts",
            row_numbers=(2,),
            examples=(
                ReviewExample(
                    source_text="Fiery Assault deals damage",
                    target_text="Assaut flamboyant inflige des dégâts",
                ),
            ),
        )

        prompt = build_codex_prompt([cluster], prompt_ids=["cluster-1"])

        self.assertIn("这是术语一致性 QA，不是一般翻译质量评估", prompt)
        self.assertIn("动名词/名词化、形容词化、词族派生或法语句法重组", prompt)
        self.assertIn("词性变化/动名词/自然句法重组", prompt)
        self.assertIn("另一个同义译名、近义译名、自然改写、短称、看似官方的另一定稿或语义等价译名", prompt)
        self.assertIn("source_term=Fiery Assault", prompt)
        self.assertIn("target_text 使用 Assaut flamboyant", prompt)
        self.assertIn("判 true_issue", prompt)
        self.assertIn("同义译名/定稿差异但未按术语表", prompt)
        self.assertNotIn("官方译名/固定译法变体", prompt)

    def test_parse_codex_decisions_tolerates_literal_control_characters_in_note(self) -> None:
        text = (
            '{"results":[{"id":"cluster-1","decision":"review","category":"需人工确认",'
            '"confidence":"low","note":"第一行\n第二行"}]}'
        )

        decisions = parse_codex_decisions(text)

        self.assertEqual(decisions["cluster-1"].note, "第一行\n第二行")


if __name__ == "__main__":
    unittest.main()
