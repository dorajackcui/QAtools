from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.target_text_checker.check_target_text import (
    ABNORMAL_PUNCTUATION_RULE,
    CONSECUTIVE_SPACES_RULE,
    LEADING_TRAILING_SPACES_RULE,
    MIXED_WIDTH_RULE,
    PROBLEM_SHEET_NAME,
    find_text_issues,
    process_excel,
)


class TargetTextRuleTests(unittest.TestCase):
    def test_abnormal_punctuation_allows_valid_ellipsis_and_repeated_exclamations_or_questions(self) -> None:
        self.assertEqual(find_text_issues("Wait... Done… Wow!! Really？？"), ())

        issues = find_text_issues("Wait.. Then.... Finally.....")

        self.assertEqual(
            [issue.rule for issue in issues],
            [ABNORMAL_PUNCTUATION_RULE],
        )
        self.assertEqual(issues[0].matched_content, "..、....、.....")

    def test_abnormal_punctuation_detects_repeated_supported_families(self) -> None:
        issues = find_text_issues(
            "A,, B，， C,， D。。 E：： F:: G;; H；； I、、",
            rules=(ABNORMAL_PUNCTUATION_RULE,),
        )

        self.assertEqual(len(issues), 1)
        self.assertEqual(
            issues[0].matched_content,
            ",,、，，、,，、。。、：：、::、;;、；；、、、",
        )

    def test_legacy_ellipsis_rule_name_selects_upgraded_punctuation_rule(self) -> None:
        issues = find_text_issues("Text,,", rules=("abnormal-ellipsis",))

        self.assertEqual([issue.rule for issue in issues], [ABNORMAL_PUNCTUATION_RULE])

    def test_consecutive_spaces_checks_ascii_space_runs_at_any_position(self) -> None:
        issues = find_text_issues(
            "  Hello  world   again  ",
            rules=(CONSECUTIVE_SPACES_RULE,),
        )

        self.assertEqual([issue.rule for issue in issues], [CONSECUTIVE_SPACES_RULE])
        self.assertEqual(issues[0].matched_content, "2 个空格、3 个空格")
        self.assertEqual(
            find_text_issues("Hello world", rules=(CONSECUTIVE_SPACES_RULE,)),
            (),
        )
        self.assertEqual(
            find_text_issues(
                "Hello  ",
                rules=(CONSECUTIVE_SPACES_RULE,),
            )[0].matched_content,
            "2 个空格",
        )
        self.assertEqual(
            find_text_issues(" Hello ", rules=(CONSECUTIVE_SPACES_RULE,)),
            (),
        )
        self.assertEqual(
            find_text_issues(
                "Hello\u00a0\u00a0world",
                rules=(CONSECUTIVE_SPACES_RULE,),
            ),
            (),
        )

    def test_leading_trailing_spaces_reports_each_edge_and_ignores_nbsp(self) -> None:
        issues = find_text_issues(
            " Hello  ",
            rules=(LEADING_TRAILING_SPACES_RULE,),
        )

        self.assertEqual(
            [issue.rule for issue in issues],
            [LEADING_TRAILING_SPACES_RULE],
        )
        self.assertEqual(issues[0].matched_content, "开头 1 个空格、结尾 2 个空格")
        self.assertEqual(
            find_text_issues(
                "Hello ",
                rules=(LEADING_TRAILING_SPACES_RULE,),
            )[0].matched_content,
            "结尾 1 个空格",
        )
        self.assertEqual(
            find_text_issues(
                "   ",
                rules=(LEADING_TRAILING_SPACES_RULE,),
            )[0].matched_content,
            "首尾 3 个空格",
        )
        self.assertEqual(
            find_text_issues(
                "\u00a0Hello\u00a0",
                rules=(LEADING_TRAILING_SPACES_RULE,),
            ),
            (),
        )

    def test_edge_rule_complements_consecutive_spaces_at_text_boundaries(self) -> None:
        self.assertEqual(
            [issue.rule for issue in find_text_issues("Hello  ")],
            [CONSECUTIVE_SPACES_RULE, LEADING_TRAILING_SPACES_RULE],
        )
        self.assertEqual(
            [issue.rule for issue in find_text_issues("Hello ")],
            [LEADING_TRAILING_SPACES_RULE],
        )

    def test_mixed_width_compares_equivalent_character_families(self) -> None:
        issues = find_text_issues("Hello, world，（test) ABC12１２")

        self.assertEqual([issue.rule for issue in issues], [MIXED_WIDTH_RULE])
        self.assertIn("逗号（半角 , / 全角 ，）", issues[0].matched_content)
        self.assertIn("圆括号（半角 ) / 全角 （）", issues[0].matched_content)
        self.assertIn("数字（半角 12 / 全角 １２）", issues[0].matched_content)

    def test_mixed_width_does_not_flag_unrelated_styles(self) -> None:
        self.assertEqual(find_text_issues("Version 1.2（test）"), ())
        self.assertEqual(find_text_issues("普通中文，with ASCII words"), ())

    def test_rules_can_be_selected_independently(self) -> None:
        value = "Wait..  now,，"

        issues = find_text_issues(value, rules=(CONSECUTIVE_SPACES_RULE,))

        self.assertEqual([issue.rule for issue in issues], [CONSECUTIVE_SPACES_RULE])

    def test_rejects_empty_or_unknown_rule_selection(self) -> None:
        with self.assertRaisesRegex(ValueError, "至少需要选择一项规则"):
            find_text_issues("Text", rules=())
        with self.assertRaisesRegex(ValueError, "不支持"):
            find_text_issues("Text", rules=("unknown",))


class TargetTextExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["source", "target"])
        worksheet.append(["row 2", "Wait..  now,，"])
        worksheet.append(["row 3", "Allowed... text"])
        worksheet.append(["row 4", "（mixed)"])
        workbook.save(path)

    def test_process_excel_writes_one_problem_per_triggered_rule(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                output_file=output_path,
            )

            self.assertEqual(summary.processed_count, 3)
            self.assertEqual(summary.problem_count, 4)
            self.assertEqual(summary.problem_rows, 2)
            workbook = load_workbook(output_path)
            problem_sheet = workbook[PROBLEM_SHEET_NAME]
            self.assertEqual(
                list(problem_sheet.values),
                [
                    ("行号", "source原文", "target原文", "问题描述", "问题类型", "命中内容"),
                    (
                        2,
                        "row 2",
                        "Wait..  now,，",
                        "Target 中存在异常或重复标点符号。",
                        "异常标点符号",
                        "..、,，",
                    ),
                    (2, "row 2", "Wait..  now,，", "Target 中存在连续空格。", "连续空格", "2 个空格"),
                    (
                        2,
                        "row 2",
                        "Wait..  now,，",
                        "Target 中存在同类字符的全半角混用。",
                        "全半角混用",
                        "逗号（半角 , / 全角 ，）",
                    ),
                    (
                        4,
                        "row 4",
                        "（mixed)",
                        "Target 中存在同类字符的全半角混用。",
                        "全半角混用",
                        "圆括号（半角 ) / 全角 （）",
                    ),
                ],
            )
            self.assertEqual(problem_sheet["A2"].hyperlink.location, "'Data'!B2")

    def test_process_excel_honors_selected_rules(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_path,
                source_column="A",
                target_column="B",
                rules=(ABNORMAL_PUNCTUATION_RULE,),
            )

            self.assertEqual(summary.problem_count, 1)
            self.assertEqual(summary.problem_rows, 1)

    def test_process_excel_reports_leading_trailing_spaces(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["source", "target"])
            worksheet.append(["row 2", "Done  "])
            workbook.save(input_path)
            workbook.close()

            summary = process_excel(
                input_path,
                source_column="A",
                target_column="B",
                rules=(LEADING_TRAILING_SPACES_RULE,),
                output_file=output_path,
            )

            self.assertEqual(summary.problem_count, 1)
            output_workbook = load_workbook(output_path)
            problem_sheet = output_workbook[PROBLEM_SHEET_NAME]
            self.assertEqual(problem_sheet["E2"].value, "首尾空格")
            self.assertEqual(problem_sheet["F2"].value, "结尾 2 个空格")
            output_workbook.close()


if __name__ == "__main__":
    unittest.main()
