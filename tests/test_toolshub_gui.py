from __future__ import annotations

import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import Workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtCore import QPoint, Qt  # noqa: E402
from PySide6.QtGui import QCloseEvent  # noqa: E402
from PySide6.QtWidgets import (  # noqa: E402
    QApplication,
    QDialog,
    QFrame,
    QLabel,
    QLineEdit,
    QStackedWidget,
    QToolButton,
)

from toolshub_gui import (  # noqa: E402
    SIDEBAR_WIDTH,
    TOOL_GROUPS,
    ToolshubApp,
    _acquire_gui_instance_lock,
    build_argument_parser,
    calculate_initial_window_size,
    main,
)
from tools.header_aliases import HeaderAliasStore  # noqa: E402
from tools.qt_gui_common import (  # noqa: E402
    ACCENT_COLOR,
    ACCENT_FOREGROUND_COLOR,
    APP_BACKGROUND,
    TEXT_COLOR,
    configure_qt_application,
)
from tools.qt_pages import (  # noqa: E402
    FrenchNbspPage,
    PhraseLoomPage,
    SettingsPage,
    WorkflowPage,
)
from tools.tb_projects import TbProject  # noqa: E402


class ToolshubLayoutTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.qt_app = QApplication.instance() or QApplication([])
        configure_qt_application(cls.qt_app)

    def setUp(self) -> None:
        config_dir = tempfile.TemporaryDirectory()
        self.addCleanup(config_dir.cleanup)
        self.config_dir = Path(config_dir.name)
        self.enterContext(patch(
            "tools.workflow.gui_options.default_header_aliases_path",
            return_value=self.config_dir / "header_aliases.json",
        ))
        self.enterContext(patch(
            "tools.tb_projects.default_tb_projects_path",
            return_value=self.config_dir / "tb_projects.json",
        ))

    def make_app(self) -> ToolshubApp:
        return ToolshubApp(show_window=False)

    def test_initial_window_size_adds_breathing_room(self) -> None:
        self.assertEqual(
            calculate_initial_window_size(
                requested_width=1238,
                requested_height=725,
                screen_width=1920,
                screen_height=1080,
            ),
            (1254, 741),
        )

    def test_initial_window_size_uses_compact_default_floor(self) -> None:
        self.assertEqual(
            calculate_initial_window_size(
                requested_width=960,
                requested_height=500,
                screen_width=1920,
                screen_height=1080,
            ),
            (1000, 660),
        )

    def test_initial_window_size_stays_within_available_screen(self) -> None:
        self.assertEqual(
            calculate_initial_window_size(
                requested_width=1500,
                requested_height=900,
                screen_width=1366,
                screen_height=768,
            ),
            (1286, 688),
        )

    def test_theme_uses_supplied_surface_ink_and_accent_anchors(self) -> None:
        self.assertEqual(APP_BACKGROUND, "#f9f9f7")
        self.assertEqual(TEXT_COLOR, "#2d2d2b")
        self.assertEqual(ACCENT_COLOR, "#cc7d5e")
        self.assertEqual(ACCENT_FOREGROUND_COLOR, "#ffffff")

    def test_smoke_test_builds_complete_app_without_showing_window(self) -> None:
        receiver = Mock()
        receiver.start.return_value = False
        with patch("toolshub_gui.WorkflowFileReceiver", return_value=receiver):
            self.assertEqual(main(["--smoke-test"]), 0)
        receiver.close.assert_called()

    def test_navigation_only_exposes_top_level_workflows_and_utilities(self) -> None:
        grouped_tools = {
            group.title: [tool.title for tool in group.tools]
            for group in TOOL_GROUPS
        }
        self.assertEqual(grouped_tools["常用流程"], ["一键质量检查", "PhraseLoom"])
        self.assertEqual(grouped_tools["文本修复"], ["法语 NBSP 恢复"])
        self.assertEqual(
            grouped_tools["其他"],
            ["Batch 拆分", "合并表格", "Xbench QA 转换"],
        )

    def test_all_qt_pages_are_created_once_and_kept_in_stack(self) -> None:
        window = self.make_app()
        try:
            self.assertIsInstance(window.page_stack, QStackedWidget)
            self.assertEqual(
                set(window.tool_frames),
                {
                    "workflow",
                    "phraseloom",
                    "french_nbsp",
                    "excel_batcher",
                    "excel_merger",
                    "xbench_report",
                    "settings",
                },
            )
            self.assertEqual(window.page_stack.count(), 7)
            self.assertIsInstance(window.tool_frames["workflow"], WorkflowPage)
            self.assertIsInstance(window.tool_frames["phraseloom"], PhraseLoomPage)
            self.assertIsInstance(window.tool_frames["settings"], SettingsPage)
        finally:
            window.close()

    def test_shell_and_workflow_actions_use_fixed_compact_action_bar(self) -> None:
        window = self.make_app()
        try:
            window.resize(1000, 660)
            window.show()
            self.qt_app.processEvents()

            workflow = window.tool_frames["workflow"]
            margins = workflow.content_layout.contentsMargins()

            self.assertEqual(window.sidebar.width(), SIDEBAR_WIDTH)
            self.assertEqual(SIDEBAR_WIDTH, 184)
            self.assertEqual(
                (margins.left(), margins.top(), margins.right(), margins.bottom()),
                (2, 2, 6, 2),
            )
            self.assertEqual(workflow.run_button.width(), 136)
            self.assertEqual(workflow.revision_button.width(), 108)
            self.assertEqual(window.title_label.objectName(), "pageTitle")
            self.assertEqual(
                window.nav_buttons["workflow"].property("navItem"),
                True,
            )
            self.assertIsInstance(workflow.term_settings_dialog, QDialog)
            self.assertEqual(
                workflow.term_settings_dialog.objectName(),
                "settingsDialog",
            )
            self.assertFalse(workflow.term_settings_dialog.isVisible())
            self.assertEqual(workflow.action_bar.objectName(), "pageActionBar")
            self.assertEqual(workflow.action_bar.height(), 46)
            self.assertIs(workflow.run_button.parentWidget(), workflow.action_bar)
            self.assertGreaterEqual(
                workflow.action_bar.y(),
                workflow.height() - workflow.action_bar.height() - 1,
            )
            self.assertGreater(
                workflow.run_button.x(),
                workflow.revision_button.x(),
            )

            action_geometry = workflow.action_bar.geometry().getRect()
            workflow.term_settings_button.click()
            self.qt_app.processEvents()

            self.assertTrue(workflow.term_settings_dialog.isVisible())
            self.assertTrue(workflow.term_settings_dialog.isModal())
            with patch("tools.qt_pages._choose_excel", return_value="") as choose:
                workflow.choose_history_file()
            self.assertIs(
                choose.call_args.args[0],
                workflow.term_settings_dialog,
            )
            self.assertEqual(
                workflow.action_bar.geometry().getRect(),
                action_geometry,
            )
            workflow.term_settings_dialog.reject()
        finally:
            window.close()

    def test_selecting_a_tool_reuses_persistent_page(self) -> None:
        window = self.make_app()
        try:
            page = window.tool_frames["french_nbsp"]
            window.select_tool("french_nbsp")
            window.select_tool("workflow")
            window.select_tool("french_nbsp")
            self.assertEqual(window.current_tool_key, "french_nbsp")
            self.assertEqual(window.title_label.text(), "法语 NBSP 恢复")
            self.assertIs(window.current_tool_frame, page)
            self.assertIs(window.page_stack.currentWidget(), page)
            self.assertTrue(window.nav_buttons["french_nbsp"].isChecked())
        finally:
            window.close()

    def test_settings_gear_is_pinned_to_sidebar_bottom(self) -> None:
        window = self.make_app()
        try:
            window.resize(1000, 660)
            window.show()
            self.qt_app.processEvents()

            settings_button = window.nav_buttons["settings"]
            self.assertEqual(settings_button.text(), "⚙  设置")
            self.assertEqual(settings_button.objectName(), "settingsNavButton")
            self.assertGreater(
                settings_button.y(),
                window.nav_buttons["xbench_report"].y(),
            )

            settings_button.click()
            self.assertEqual(window.current_tool_key, "settings")
            self.assertEqual(window.title_label.text(), "设置")
            self.assertTrue(settings_button.isChecked())
        finally:
            window.close()

    def test_saved_header_aliases_refresh_loaded_workbook_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_path = Path(tmp_dir)
            workbook_path = temp_path / "localized.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet.append(["SourceFile", "Key", "MsgStr", "Translation", "Source", "Target"])
            workbook.create_sheet("Missing").append(["SourceFile", "TargetFile"])
            workbook.create_sheet("Duplicate").append(
                ["MsgStr", "MSGSTR", "Translation", "TRANSLATION", "Source", "Target"]
            )
            workbook.save(workbook_path)
            workbook.close()
            store = HeaderAliasStore(temp_path / "header_aliases.json")
            window = ToolshubApp(show_window=False, header_alias_store=store)
            try:
                workflow = window.tool_frames["workflow"]
                restorer = window.tool_frames["french_nbsp"]
                settings = window.tool_frames["settings"]
                self.assertIsInstance(workflow, WorkflowPage)
                self.assertIsInstance(settings, SettingsPage)
                workflow.load_input_file(str(workbook_path))
                restorer.load_input_file(str(workbook_path))
                self.assertEqual(workflow.source_column.text(), "E")
                self.assertEqual(workflow.target_column.text(), "F")
                self.assertEqual(restorer.target_column.text(), "F")

                settings.source_aliases.setPlainText("MsgStr")
                settings.target_aliases.setPlainText("Translation")
                settings.save_button.click()
                self.qt_app.processEvents()

                self.assertEqual(workflow.source_column.text(), "C")
                self.assertEqual(workflow.target_column.text(), "D")
                self.assertEqual(restorer.target_column.text(), "D")
                self.assertEqual(store.load().source, ("MsgStr",))
                self.assertEqual(store.load().target, ("Translation",))

                for sheet in ("Missing", "Duplicate"):
                    with self.subTest(sheet=sheet):
                        workflow.sheet.setCurrentText(sheet)
                        restorer.sheet.setCurrentText(sheet)
                        self.assertEqual(workflow.source_column.text(), "")
                        self.assertEqual(workflow.target_column.text(), "")
                        self.assertEqual(restorer.target_column.text(), "")
                        workflow.run_in_background = Mock()
                        with patch("tools.qt_pages.show_error"):
                            workflow.run_selected_tasks()
                        workflow.run_in_background.assert_not_called()

                workflow.sheet.setCurrentText("Data")
                self.assertEqual(workflow.source_column.text(), "C")
                self.assertEqual(workflow.target_column.text(), "D")
            finally:
                window.close()

    def test_all_primary_page_actions_live_in_fixed_action_bars(self) -> None:
        window = self.make_app()
        try:
            action_buttons = (
                window.tool_frames["workflow"].run_button,
                window.tool_frames["phraseloom"].export_button,
                window.tool_frames["french_nbsp"].run_button,
                window.tool_frames["excel_merger"].run_button,
                window.tool_frames["xbench_report"].run_button,
                window.tool_frames["excel_batcher"].split_button,
                window.tool_frames["excel_batcher"].restore_button,
            )

            for button in action_buttons:
                with self.subTest(button=button.text()):
                    self.assertEqual(
                        button.parentWidget().objectName(),
                        "pageActionBar",
                    )
        finally:
            window.close()

    def test_workflow_file_path_is_read_only_and_button_driven(self) -> None:
        window = self.make_app()
        try:
            workflow = window.tool_frames["workflow"]
            self.assertTrue(workflow.input_picker.line_edit.isReadOnly())
            self.assertIsInstance(workflow.input_picker.line_edit, QLineEdit)
            workflow.input_picker.set_path(r"C:\very\long\folder\sample.xlsx")
            self.assertEqual(
                workflow.input_picker.path(),
                r"C:\very\long\folder\sample.xlsx",
            )
        finally:
            window.close()

    def test_workflow_optional_tag_config_can_be_cleared(self) -> None:
        window = self.make_app()
        try:
            workflow = window.tool_frames["workflow"]
            workflow.angle_config.set_path(r"C:\configs\tags.json")
            workflow.angle_config.clear()
            self.assertEqual(workflow.angle_config.path(), "")
        finally:
            window.close()

    def test_workflow_groups_quality_checks_and_keeps_reverse_check_optional(self) -> None:
        window = self.make_app()
        try:
            window.resize(1000, 760)
            window.show()
            workflow = window.tool_frames["workflow"]
            self.qt_app.processEvents()

            headings = {
                label.text()
                for label in workflow.findChildren(QLabel)
                if label.objectName() == "sectionTitle"
            }
            labels = {label.text() for label in workflow.findChildren(QLabel)}
            self.assertTrue(
                {
                    "术语与翻译一致性",
                    "内容保真检查",
                    "Target 文本质量",
                }.issubset(headings)
            )
            self.assertEqual(workflow.consistency_check.text(), "同 Source 不同 Target")
            self.assertEqual(
                workflow.target_consistency_check.text(),
                "同 Target 不同 Source",
            )
            self.assertEqual(
                workflow.consistency_check.y(),
                workflow.target_consistency_check.y(),
            )
            self.assertTrue(workflow.consistency_check.isChecked())
            self.assertFalse(workflow.target_consistency_check.isChecked())
            self.assertTrue(workflow.number_check.isChecked())
            self.assertTrue(workflow.url_check.isChecked())
            self.assertIsInstance(workflow.term_settings_button, QToolButton)
            self.assertEqual(workflow.term_settings_button.text(), "")
            self.assertEqual(
                workflow.term_settings_button.arrowType(),
                Qt.ArrowType.DownArrow,
            )
            self.assertEqual(
                workflow.term_settings_button.toolTip(),
                "术语检查设置",
            )
            self.assertEqual(
                workflow.tag_settings_button.toolTip(),
                "Tag / Placeholder 设置",
            )
            self.assertEqual(
                workflow.target_settings_button.toolTip(),
                "Target 文本规范设置",
            )
            for check, button in (
                (workflow.term_check, workflow.term_settings_button),
                (workflow.tag_check, workflow.tag_settings_button),
                (workflow.target_text_check, workflow.target_settings_button),
            ):
                gap = button.mapTo(workflow, QPoint(0, 0)).x() - (
                    check.mapTo(workflow, QPoint(0, 0)).x() + check.width()
                )
                self.assertLessEqual(gap, 12)
            self.assertNotIn("双向文本一致性：", labels)
            self.assertNotIn("常用检查默认开启，可按需调整", labels)
            self.assertNotIn("质量检查项目", labels)
            self.assertNotIn("展开设置", labels)
            self.assertNotIn("收起设置", labels)
            second_column_positions = {
                check.mapTo(workflow, QPoint(0, 0)).x()
                for check in (
                    workflow.target_consistency_check,
                    workflow.line_break_check,
                    workflow.url_check,
                    workflow.target_text_check,
                )
            }
            self.assertEqual(len(second_column_positions), 1)
            separators = [
                frame
                for frame in workflow.findChildren(QFrame)
                if frame.frameShape() == QFrame.Shape.HLine
            ]
            self.assertGreaterEqual(len(separators), 2)

            workflow.set_all_tasks(False)
            self.assertTrue(
                all(
                    not button.isEnabled()
                    for button in (
                        workflow.term_settings_button,
                        workflow.tag_settings_button,
                        workflow.target_settings_button,
                    )
                )
            )
            workflow.set_all_tasks(True)
            self.assertTrue(
                all(
                    button.isEnabled()
                    for button in (
                        workflow.term_settings_button,
                        workflow.tag_settings_button,
                        workflow.target_settings_button,
                    )
                )
            )
        finally:
            window.close()

    def test_tag_modes_are_visually_grouped_and_strictly_exclusive(self) -> None:
        window = self.make_app()
        try:
            window.resize(1000, 660)
            window.show()
            workflow = window.tool_frames["workflow"]
            workflow.tag_settings_button.click()
            self.qt_app.processEvents()

            self.assertTrue(workflow.tag_settings_dialog.isVisible())
            self.assertTrue(workflow.tag_mode_group.exclusive())
            self.assertEqual(workflow.standard_mode.property("segmentedMode"), True)
            self.assertEqual(workflow.memoq_mode.property("segmentedMode"), True)
            self.assertEqual(workflow.tag_mode_group.checkedId(), 0)

            workflow.memoq_mode.setChecked(True)
            self.qt_app.processEvents()

            self.assertFalse(workflow.standard_mode.isChecked())
            self.assertTrue(workflow.memoq_mode.isChecked())
            self.assertEqual(workflow.tag_mode_group.checkedId(), 1)
            self.assertTrue(
                all(not check.isEnabled() for check in workflow.standard_tag_checks)
            )
            self.assertFalse(workflow.angle_config.isEnabled())
            self.assertEqual(
                workflow.content_scroll.horizontalScrollBar().maximum(),
                0,
            )

            workflow.standard_mode.setChecked(True)
            self.qt_app.processEvents()

            self.assertFalse(workflow.memoq_mode.isChecked())
            self.assertTrue(
                all(check.isEnabled() for check in workflow.standard_tag_checks)
            )
            self.assertTrue(workflow.angle_config.isEnabled())
        finally:
            window.close()

    def test_workflow_settings_dialog_cancel_restores_previous_values(self) -> None:
        window = self.make_app()
        try:
            window.resize(1000, 660)
            window.show()
            workflow = window.tool_frames["workflow"]
            workflow.target_settings_button.click()
            self.qt_app.processEvents()

            workflow.abnormal_rule.setChecked(False)
            workflow.edge_spaces_rule.setChecked(False)
            workflow.target_settings_dialog.reject()
            self.qt_app.processEvents()

            self.assertTrue(workflow.abnormal_rule.isChecked())
            self.assertTrue(workflow.edge_spaces_rule.isChecked())

            workflow.target_settings_button.click()
            self.qt_app.processEvents()
            workflow.abnormal_rule.setChecked(False)
            workflow.target_settings_dialog.accept()
            self.qt_app.processEvents()

            self.assertFalse(workflow.abnormal_rule.isChecked())
        finally:
            window.close()

    def test_remember_options_restores_all_choices_without_reading_workbooks(self) -> None:
        for memoq in (True, False):
            with self.subTest(memoq=memoq):
                window = self.make_app()
                workflow = window.tool_frames["workflow"]
                try:
                    workflow.input_picker.set_path(r"C:\input\本批次.xlsx")
                    workflow._restore_combo(workflow.sheet, ("Sheet1", "Data"), "Data")
                    workflow.source_column.setText("C")
                    workflow.target_column.setText("F")
                    workflow.start_row.setValue(7)
                    checks = tuple(bool(index % 2) == memoq for index in range(9))
                    for check, checked in zip(workflow.task_checks, checks, strict=True):
                        check.setChecked(checked)
                    workflow.mark_book.setChecked(not memoq)
                    workflow.mark_square.setChecked(memoq)
                    workflow.tb_store.save_project(TbProject(
                        "项目 TB", r"C:\tb\术语.xlsx", "Terms", "D", "H", 5,
                    ))
                    workflow.refresh_tb_projects("项目 TB")
                    workflow.history_picker.set_path(r"C:\tb\术语.xlsx")
                    workflow._restore_combo(workflow.history_sheet, ("Old", "Terms"), "Terms")
                    workflow.history_source.setText("D")
                    workflow.history_target.setText("H")
                    workflow.history_start_row.setValue(5)
                    workflow.memoq_mode.setChecked(memoq)
                    workflow.standard_mode.setChecked(not memoq)
                    tag_checks = (True, False, True, False)
                    for check, checked in zip(workflow.standard_tag_checks, tag_checks, strict=True):
                        check.setChecked(checked)
                    workflow.angle_config.set_path(r"C:\configs\tags.json")
                    rule_checks = (False, True, False, True)
                    for check, checked in zip(workflow.rule_checks.values(), rule_checks, strict=True):
                        check.setChecked(checked)
                    workflow.remember_options_button.click()
                    self.assertIn("已记住", workflow.status.text())
                    saved_bytes = workflow.options_store.config_path.read_bytes()
                    workflow.set_all_tasks(not memoq)
                    workflow.mark_book.setChecked(memoq)
                    workflow.history_picker.clear()
                    workflow.angle_config.clear()
                    workflow.width_rule.setChecked(False)
                finally:
                    window.close()

                with (
                    patch("tools.qt_pages.list_workbook_sheets") as list_sheets,
                    patch("tools.qt_pages.detect_source_target_columns") as detect_main,
                    patch("tools.qt_pages.detect_history_tb_columns") as detect_history,
                ):
                    restored_window = self.make_app()
                    list_sheets.assert_not_called()
                    detect_main.assert_not_called()
                    detect_history.assert_not_called()
                try:
                    restored = restored_window.tool_frames["workflow"]
                    self.assertEqual(restored.input_picker.path(), r"C:\input\本批次.xlsx")
                    self.assertEqual(restored.sheet.count(), 2)
                    self.assertEqual(restored.sheet.currentText(), "Data")
                    self.assertEqual(restored.source_column.text(), "C")
                    self.assertEqual(restored.target_column.text(), "F")
                    self.assertEqual(restored.start_row.value(), 7)
                    self.assertEqual(tuple(check.isChecked() for check in restored.task_checks), checks)
                    self.assertEqual(restored.mark_book.isChecked(), not memoq)
                    self.assertEqual(restored.mark_square.isChecked(), memoq)
                    self.assertEqual(restored.tb_project.currentText(), "项目 TB")
                    self.assertEqual(restored.history_picker.path(), r"C:\tb\术语.xlsx")
                    self.assertEqual(restored.history_sheet.count(), 2)
                    self.assertEqual(restored.history_sheet.currentText(), "Terms")
                    self.assertEqual(restored.history_source.text(), "D")
                    self.assertEqual(restored.history_target.text(), "H")
                    self.assertEqual(restored.history_start_row.value(), 5)
                    self.assertEqual(restored.memoq_mode.isChecked(), memoq)
                    self.assertEqual(tuple(check.isChecked() for check in restored.standard_tag_checks), tag_checks)
                    self.assertEqual(restored.angle_config.path(), r"C:\configs\tags.json")
                    self.assertEqual(restored.angle_config.isEnabled(), not memoq)
                    self.assertEqual(tuple(check.isChecked() for check in restored.rule_checks.values()), rule_checks)
                    for check, button in (
                        (restored.term_check, restored.term_settings_button),
                        (restored.tag_check, restored.tag_settings_button),
                        (restored.target_text_check, restored.target_settings_button),
                    ):
                        self.assertEqual(button.isEnabled(), check.isChecked())
                    self.assertIn("本批次.xlsx", restored.output_preview.text())
                    self.assertEqual(restored.last_workflow_output_path, "")
                    self.assertEqual(restored.options_store.config_path.read_bytes(), saved_bytes)

                    workbook_path = self.config_dir / "new_input.xlsx"
                    workbook = Workbook()
                    try:
                        workbook.active.append(["SourceFile", "SOURCE", "TARGET"])
                        workbook.save(workbook_path)
                    finally:
                        workbook.close()
                    restored.load_input_file(str(workbook_path))
                    self.assertEqual(restored.source_column.text(), "B")
                    self.assertEqual(restored.target_column.text(), "C")
                finally:
                    restored_window.close()

    def test_invalid_remembered_options_keep_defaults_without_blocking_startup(self) -> None:
        config_path = self.config_dir / "workflow_options.json"
        for payload in (
            "{invalid",
            json.dumps({"version": 1, "options": {"checks": [False]}}),
            json.dumps({"version": 1, "options": {"input": {"start_row": 10**30}}}),
            json.dumps({"version": 1, "options": {"settings": {"term": {"mark_book": "false"}}}}),
        ):
            with self.subTest(payload=payload):
                config_path.write_text(payload, encoding="utf-8")
                window = self.make_app()
                try:
                    workflow = window.tool_frames["workflow"]
                    self.assertEqual(workflow.input_picker.path(), "")
                    self.assertEqual(workflow.start_row.value(), 2)
                    self.assertTrue(workflow.term_check.isChecked())
                    self.assertFalse(workflow.target_consistency_check.isChecked())
                    self.assertTrue(workflow.mark_book.isChecked())
                    self.assertIn("已使用默认设置", workflow.status.text())
                finally:
                    window.close()

    def test_remember_options_write_failure_preserves_previous_saved_choices(self) -> None:
        window = self.make_app()
        try:
            workflow = window.tool_frames["workflow"]
            workflow.remember_options_button.click()
            saved = workflow.options_store.config_path.read_bytes()
            workflow.set_all_tasks(False)
            with (
                patch("tools.workflow.gui_options.os.replace", side_effect=OSError("无法写入")),
                patch("tools.qt_pages.show_error") as error,
            ):
                workflow.remember_options_button.click()
            error.assert_called_once_with(workflow, "记住选项失败", "无法写入")
            self.assertEqual(workflow.options_store.config_path.read_bytes(), saved)
            self.assertFalse(workflow.options_store.config_path.with_suffix(".json.tmp").exists())
        finally:
            window.close()

    def test_workflow_runs_excel_processing_through_background_worker(self) -> None:
        window = self.make_app()
        try:
            workflow = window.tool_frames["workflow"]
            workflow.input_picker.set_path(r"C:\input\sample.xlsx")
            workflow.sheet.addItem("Sheet1")
            workflow.source_column.setText("A")
            workflow.target_column.setText("B")
            workflow.run_in_background = Mock()

            workflow.run_selected_tasks()

            workflow.run_in_background.assert_called_once()
            call_kwargs = workflow.run_in_background.call_args.kwargs
            self.assertEqual(call_kwargs["kwargs"]["input_file"], r"C:\input\sample.xlsx")
            self.assertEqual(call_kwargs["kwargs"]["sheet"], "Sheet1")
            self.assertTrue(call_kwargs["kwargs"]["run_term_pair_check"])
            self.assertTrue(call_kwargs["kwargs"]["run_target_text_check"])
            self.assertTrue(call_kwargs["kwargs"]["run_source_consistency_check"])
            self.assertFalse(call_kwargs["kwargs"]["run_target_consistency_check"])
            self.assertTrue(call_kwargs["kwargs"]["run_number_check"])
            self.assertTrue(call_kwargs["kwargs"]["run_url_check"])
            self.assertEqual(call_kwargs["kwargs"]["term_mark_styles"], ("【】", "[]"))
            self.assertEqual(
                call_kwargs["kwargs"]["tag_token_types"],
                ("angle", "square_color", "brace", "newline"),
            )
        finally:
            window.close()

    def test_batch_worker_uses_values_captured_on_gui_thread(self) -> None:
        window = self.make_app()
        try:
            batcher = window.tool_frames["excel_batcher"]
            batcher.split_input.set_path(r"C:\input\sample.xlsx")
            batcher.split_sheet.addItem("Sheet1")
            batcher.batch_size.setValue(250)
            batcher.header_rows.setValue(2)
            batcher.split_output.set_path(r"C:\output\batches")
            batcher.run_in_background = Mock()

            batcher.run_split()
            task = batcher.run_in_background.call_args.args[0]

            batcher.split_sheet.setCurrentText("")
            batcher.batch_size.setValue(999)
            batcher.header_rows.setValue(0)
            batcher.split_output.clear()
            with patch("tools.qt_pages.split_workbook", return_value=Mock()) as split:
                task()

            split.assert_called_once_with(
                input_file=r"C:\input\sample.xlsx",
                sheet="Sheet1",
                batch_size=250,
                header_rows=2,
                output_dir=r"C:\output\batches",
            )
        finally:
            window.close()

    def test_window_refuses_to_close_while_background_task_is_running(self) -> None:
        window = self.make_app()
        workflow = window.tool_frames["workflow"]
        worker = Mock()
        workflow._workers.add(worker)
        event = QCloseEvent()
        try:
            with patch("toolshub_gui.show_warning") as warning:
                window.closeEvent(event)

            self.assertFalse(event.isAccepted())
            warning.assert_called_once()
            self.assertIn("一键质量检查", warning.call_args.args[2])
        finally:
            workflow._workers.clear()
            window.close()

    def test_qa_workflow_argument_accepts_finder_excel_path(self) -> None:
        args = build_argument_parser().parse_args(["--qa-workflow", "/tmp/QA input.xlsx"])
        self.assertEqual(args.qa_workflow, "/tmp/QA input.xlsx")

    def test_nbsp_restore_argument_accepts_finder_excel_path(self) -> None:
        args = build_argument_parser().parse_args(["--nbsp-restore", "/tmp/French input.xlsx"])
        self.assertEqual(args.nbsp_restore, "/tmp/French input.xlsx")

    def test_gui_instance_lock_rejects_a_second_instance(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            lock_path = str(Path(tmp_dir) / "gui.lock")
            first_lock = _acquire_gui_instance_lock(lock_path)
            try:
                self.assertIsNotNone(first_lock)
                self.assertIsNone(_acquire_gui_instance_lock(lock_path))
            finally:
                if first_lock is not None:
                    first_lock.unlock()

    def test_nbsp_finder_action_loads_safe_defaults_and_schedules_restore(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "French input.xlsx"
            workbook_path.touch()
            window = self.make_app()
            try:
                restorer = window.tool_frames["french_nbsp"]
                self.assertIsInstance(restorer, FrenchNbspPage)
                restorer.load_input_file = Mock()
                restorer.run_restore = Mock()
                window._bring_window_to_front = Mock()
                with patch("toolshub_gui.QTimer.singleShot", side_effect=lambda _delay, callback: callback()):
                    window.open_french_nbsp_restore_file(str(workbook_path))
                restorer.load_input_file.assert_called_once_with(
                    str(workbook_path.absolute()),
                    reset_options=True,
                )
                restorer.run_restore.assert_called_once_with()
                window._bring_window_to_front.assert_called_once_with()
            finally:
                window.close()


if __name__ == "__main__":
    unittest.main()
