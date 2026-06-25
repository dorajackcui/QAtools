from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.xbench_report_transformer.transform_xbench_report import (
    collect_detail_rows,
    find_header_columns,
    format_issue_text,
    group_detail_rows,
    parse_metadata,
    process_excel,
    parse_qa_title,
)


class MetadataParsingTests(unittest.TestCase):
    def test_parse_two_metadata_lines_as_key_and_file_name(self) -> None:
        metadata = parse_metadata("LDLG_Text_ZH_q203101_Line_3\n磐城【配音】.xlsx")
        self.assertEqual(metadata.key, "LDLG_Text_ZH_q203101_Line_3")
        self.assertEqual(metadata.file_name, "磐城【配音】.xlsx")

    def test_parse_single_file_name_line_without_key(self) -> None:
        metadata = parse_metadata("磐城【配音】.xlsx")
        self.assertEqual(metadata.key, "")
        self.assertEqual(metadata.file_name, "磐城【配音】.xlsx")

    def test_parse_single_non_file_line_as_key(self) -> None:
        metadata = parse_metadata("LDLG_Text_ZH_q203101_Line_3")
        self.assertEqual(metadata.key, "LDLG_Text_ZH_q203101_Line_3")
        self.assertEqual(metadata.file_name, "")

    def test_parse_empty_metadata(self) -> None:
        metadata = parse_metadata(None)
        self.assertEqual(metadata.key, "")
        self.assertEqual(metadata.file_name, "")

    def test_parse_metadata_ignores_extra_lines_for_now(self) -> None:
        metadata = parse_metadata("Key_1\nfile.xlsx\nextra")
        self.assertEqual(metadata.key, "Key_1")
        self.assertEqual(metadata.file_name, "file.xlsx")


class QaTitleParsingTests(unittest.TestCase):
    def test_format_key_term_mismatch_title_as_term_pair_issue(self) -> None:
        issue = parse_qa_title("Key Term Mismatch (提示 / Avis)")
        self.assertEqual(issue.issue_type, "Key Term Mismatch")
        self.assertEqual(issue.source_term, "提示")
        self.assertEqual(issue.target_term, "Avis")
        self.assertEqual(format_issue_text(issue), "提示 -> Avis：Key Term Mismatch")

    def test_format_issue_preserves_quoted_terms(self) -> None:
        issue = parse_qa_title('Key Term Mismatch (“斑鸠” / "Colombe")')
        self.assertEqual(format_issue_text(issue), '“斑鸠” -> "Colombe"：Key Term Mismatch')

    def test_format_issue_allows_slashes_inside_terms(self) -> None:
        issue = parse_qa_title("Key Term Mismatch (HP/MP / PV/PM)")
        self.assertEqual(issue.issue_type, "Key Term Mismatch")
        self.assertEqual(issue.source_term, "HP/MP")
        self.assertEqual(issue.target_term, "PV/PM")
        self.assertEqual(format_issue_text(issue), "HP/MP -> PV/PM：Key Term Mismatch")

    def test_title_without_spaced_term_separator_is_unparseable(self) -> None:
        issue = parse_qa_title("Key Term Mismatch (HP/MP)")
        self.assertEqual(issue.issue_type, "Key Term Mismatch (HP/MP)")
        self.assertEqual(issue.source_term, "")
        self.assertEqual(issue.target_term, "")

    def test_unparseable_title_is_used_as_issue_type(self) -> None:
        issue = parse_qa_title("Target same as Source")
        self.assertEqual(issue.issue_type, "Target same as Source")
        self.assertEqual(issue.source_term, "")
        self.assertEqual(issue.target_term, "")
        self.assertEqual(format_issue_text(issue), "Target same as Source")


class RowExtractionAndGroupingTests(unittest.TestCase):
    def build_xbench_workbook(self) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Xbench QA"
        worksheet["A1"] = "Exported QA Report"
        worksheet["C4"] = "Source"
        worksheet["D4"] = "Target"
        worksheet["E4"] = "Comments"
        worksheet["F4"] = "Metadata"
        worksheet["A5"] = "Key Term Mismatch (提示 / Avis)"
        worksheet["C6"] = "好刻意的提示！"
        worksheet["D6"] = "Sans blague !"
        worksheet["F6"] = "Key_1\nUI弹窗文字.xlsx"
        worksheet["A7"] = 'Key Term Mismatch (“斑鸠” / "Colombe")'
        worksheet["C8"] = "“斑鸠”&“诗人”"
        worksheet["D8"] = '"Colombe" & "Poète"'
        worksheet["F8"] = "Key_2\n磐城【配音】.xlsx"
        worksheet["A9"] = 'Key Term Mismatch (“诗人” / "Poète")'
        worksheet["C10"] = "“斑鸠”&“诗人”"
        worksheet["D10"] = '"Colombe" & "Poète"'
        worksheet["F10"] = "Key_2\n磐城【配音】.xlsx"
        return workbook

    def test_find_header_columns_detects_table_header_after_intro_rows(self) -> None:
        workbook = self.build_xbench_workbook()
        worksheet = workbook["Xbench QA"]
        header_row, columns = find_header_columns(worksheet)
        self.assertEqual(header_row, 4)
        self.assertEqual(columns["source"], 3)
        self.assertEqual(columns["target"], 4)
        self.assertEqual(columns["comments"], 5)
        self.assertEqual(columns["metadata"], 6)

    def test_collect_detail_rows_attaches_current_qa_issue(self) -> None:
        workbook = self.build_xbench_workbook()
        worksheet = workbook["Xbench QA"]
        detail_rows = collect_detail_rows(worksheet)
        self.assertEqual(len(detail_rows), 3)
        self.assertEqual(detail_rows[0].file_name, "UI弹窗文字.xlsx")
        self.assertEqual(detail_rows[0].key, "Key_1")
        self.assertEqual(detail_rows[0].source, "好刻意的提示！")
        self.assertEqual(detail_rows[0].target, "Sans blague !")
        self.assertEqual(detail_rows[0].qa_issue, "提示 -> Avis：Key Term Mismatch")
        self.assertEqual(detail_rows[0].group_key, "key:Key_1")

    def test_collect_detail_rows_uses_comments_column_for_qa_issue_when_a_cell_is_empty(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["C1"] = "Source"
        worksheet["D1"] = "Target"
        worksheet["E1"] = "Comments"
        worksheet["F1"] = "Metadata"
        worksheet["E2"] = "Key Term Mismatch (提示 / Avis)"
        worksheet["C3"] = "提示"
        worksheet["D3"] = "Avis"
        detail_rows = collect_detail_rows(worksheet)
        self.assertEqual(len(detail_rows), 1)
        self.assertEqual(detail_rows[0].qa_issue, "提示 -> Avis：Key Term Mismatch")

    def test_collect_detail_rows_prefers_a_column_title_over_comments_value(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["C1"] = "Source"
        worksheet["D1"] = "Target"
        worksheet["E1"] = "Comments"
        worksheet["F1"] = "Metadata"
        worksheet["A2"] = "Key Term Mismatch (提示 / Avis)"
        worksheet["E2"] = "terms.xlsx"
        worksheet["C3"] = "提示"
        worksheet["D3"] = "Avis"
        detail_rows = collect_detail_rows(worksheet)
        self.assertEqual(len(detail_rows), 1)
        self.assertEqual(detail_rows[0].qa_issue, "提示 -> Avis：Key Term Mismatch")

    def test_group_detail_rows_merges_duplicate_key_issues_with_chinese_semicolon(self) -> None:
        workbook = self.build_xbench_workbook()
        worksheet = workbook["Xbench QA"]
        grouped_rows = group_detail_rows(collect_detail_rows(worksheet))
        self.assertEqual(len(grouped_rows), 2)
        self.assertEqual(grouped_rows[1]["文件名"], "磐城【配音】.xlsx")
        self.assertEqual(grouped_rows[1]["key"], "Key_2")
        self.assertEqual(grouped_rows[1]["source"], "“斑鸠”&“诗人”")
        self.assertEqual(
            grouped_rows[1]["QA问题"],
            '“斑鸠” -> "Colombe"：Key Term Mismatch；“诗人” -> "Poète"：Key Term Mismatch',
        )

    def test_group_key_uses_file_name_and_source_when_metadata_has_only_file_name(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["C1"] = "Source"
        worksheet["D1"] = "Target"
        worksheet["E1"] = "Comments"
        worksheet["F1"] = "Metadata"
        worksheet["A2"] = "Key Term Mismatch (提示 / Avis)"
        worksheet["C3"] = "提示"
        worksheet["D3"] = "Avis"
        worksheet["E3"] = "terms.xlsx"
        worksheet["F3"] = "UI弹窗文字.xlsx"
        worksheet["C4"] = "提示"
        worksheet["D4"] = "Avis"
        worksheet["E4"] = "terms.xlsx"
        worksheet["F4"] = "另一个文件.xlsx"
        detail_rows = collect_detail_rows(worksheet)
        grouped_rows = group_detail_rows(detail_rows)
        self.assertEqual(len(grouped_rows), 2)
        self.assertEqual(detail_rows[0].group_key, "file_source:UI弹窗文字.xlsx\x1f提示")
        self.assertEqual(detail_rows[1].group_key, "file_source:另一个文件.xlsx\x1f提示")

    def test_group_key_uses_source_when_metadata_is_empty(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = "Intro"
        worksheet["C2"] = "Source"
        worksheet["D2"] = "Target"
        worksheet["E2"] = "Comments"
        worksheet["F2"] = "Metadata"
        worksheet["A3"] = "Key Term Mismatch (提示 / Avis)"
        worksheet["C4"] = "提示"
        worksheet["D4"] = "Avis"
        detail_rows = collect_detail_rows(worksheet)
        self.assertEqual(len(detail_rows), 1)
        self.assertEqual(detail_rows[0].group_key, "source:提示")


class ProcessExcelTests(unittest.TestCase):
    def create_report(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Xbench QA"
        worksheet["A1"] = "Exported QA Report"
        worksheet["C4"] = "Source"
        worksheet["D4"] = "Target"
        worksheet["E4"] = "Comments"
        worksheet["F4"] = "Metadata"
        worksheet["A5"] = "Key Term Mismatch (提示 / Avis)"
        worksheet["C6"] = "好刻意的提示！"
        worksheet["D6"] = "Sans blague !"
        worksheet["E6"] = "terms.xlsx"
        worksheet["F6"] = "Key_1\nUI弹窗文字.xlsx"
        worksheet["A7"] = 'Key Term Mismatch (“诗人” / "Poète")'
        worksheet["C8"] = "“斑鸠”&“诗人”"
        worksheet["D8"] = '"Colombe" & "Poète"'
        worksheet["E8"] = "terms.xlsx"
        worksheet["F8"] = "Key_2\n磐城【配音】.xlsx"
        workbook.save(path)

    def test_process_excel_writes_flat_output_workbook_and_preserves_input(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "Xbench_QA_Report.xlsx"
            output_path = Path(tmp_dir) / "flat.xlsx"
            self.create_report(input_path)
            summary = process_excel(input_path, output_file=output_path)
            self.assertEqual(summary.worksheet_title, "Xbench QA")
            self.assertEqual(summary.output_path, output_path.resolve())
            self.assertEqual(summary.detail_count, 2)
            self.assertEqual(summary.grouped_count, 2)
            original = load_workbook(input_path)
            self.assertEqual(original.sheetnames, ["Xbench QA"])
            result = load_workbook(output_path)
            self.assertEqual(result.sheetnames, ["Xbench QA整理"])
            sheet = result["Xbench QA整理"]
            self.assertEqual(
                [sheet.cell(1, column).value for column in range(1, 6)],
                ["文件名", "key", "source", "target", "QA问题"],
            )
            self.assertEqual(sheet["A2"].value, "UI弹窗文字.xlsx")
            self.assertEqual(sheet["B2"].value, "Key_1")
            self.assertEqual(sheet["C2"].value, "好刻意的提示！")
            self.assertEqual(sheet["D2"].value, "Sans blague !")
            self.assertEqual(sheet["E2"].value, "提示 -> Avis：Key Term Mismatch")

    def test_process_excel_uses_default_prefixed_output_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "Xbench_QA_Report.xlsx"
            self.create_report(input_path)
            summary = process_excel(input_path)
            self.assertEqual(
                summary.output_path,
                (Path(tmp_dir) / "xbench_transform_Xbench_QA_Report.xlsx").resolve(),
            )
            self.assertTrue(summary.output_path.exists())

    def test_process_excel_rejects_missing_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "Xbench_QA_Report.xlsx"
            self.create_report(input_path)
            with self.assertRaisesRegex(ValueError, "工作表不存在"):
                process_excel(input_path, sheet="Missing")


if __name__ == "__main__":
    unittest.main()
