import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


def _unit_header_index(headers, column):
    aliases = {
        "source_unit": ("source_unit", "source"),
        "target_unit": ("target_unit", "target"),
    }
    for candidate in aliases.get(column, (column,)):
        if candidate in headers:
            return headers.index(candidate)
    raise ValueError(f"{column!r} is not in list")


def _unit_row_value(row, column):
    aliases = {
        "source_unit": ("source_unit", "source"),
        "target_unit": ("target_unit", "target"),
    }
    for candidate in aliases.get(column, (column,)):
        if candidate in row:
            return row[candidate]
    raise KeyError(column)


class TemplateDemoTests(unittest.TestCase):
    def test_infers_target_template_from_one_example_and_applies_to_matching_source(self):
        from phraseloom.template_engine import (
            apply_target_template,
            infer_target_template,
            parse_template,
        )

        example = parse_template("VIP10 Paid Pack")

        self.assertEqual(example.template, "VIP{num1} Paid Pack")
        self.assertEqual(example.values, {"num1": "10"})

        target_template = infer_target_template(example.values, "VIP10pack")

        self.assertEqual(target_template, "VIP{num1}pack")
        self.assertEqual(apply_target_template(target_template, {"num1": "11"}), "VIP11pack")

    def test_template_parser_ignores_raw_brace_placeholders_after_protection(self):
        from phraseloom.tag_engine import extract_tags
        from phraseloom.template_engine import parse_template

        protected = extract_tags("Player reaches level {a}").text
        stage = parse_template("Clear Story 10-20")

        self.assertEqual(protected, "Player reaches level {1}")
        self.assertEqual(parse_template(protected).template, "Player reaches level {1}")
        self.assertEqual(parse_template(protected).values, {})
        self.assertEqual(stage.template, "Clear Story {stage1}")
        self.assertEqual(stage.values, {"stage1": "10-20"})

    def test_template_parser_preserves_protected_tokens_without_values(self):
        from phraseloom.template_engine import parse_template

        match = parse_template("{1>VIP10 Pack<2} {3}")

        self.assertEqual(match.text, "{1>VIP10 Pack<2} {3}")
        self.assertEqual(match.template, "{1>VIP{num1} Pack<2} {3}")
        self.assertEqual(match.values, {"num1": "10"})

    def test_target_template_inference_does_not_replace_protected_token_digits(self):
        from phraseloom.template_engine import infer_target_template, parse_template

        match = parse_template("{1>Level 1<2}")

        self.assertEqual(match.template, "{1>Level {num1}<2}")
        self.assertEqual(match.values, {"num1": "1"})
        self.assertEqual(
            infer_target_template(match.values, "{1>Niveau 1<2}"),
            "{1>Niveau {num1}<2}",
        )

    def test_protected_only_units_autofill_and_template_fill_restores_raw_tags(self):
        from phraseloom.workflow import fill_target_column_workbook, generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            pack_path = Path(tmp) / "pack.xlsx"
            filled_path = Path(tmp) / "filled.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(['<img src="coin.png"/>', ""])
            ws.append(['<color=#fff>VIP10 Pack</color>', ""])
            ws.append(['<color=#fff>VIP20 Pack</color>', ""])
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                pack_path,
                source_col="source",
                target_col="target",
                min_group_size=2,
                use_existing_targets=False,
            )

            self.assertEqual(stats["prefilled_translation_unit_count"], 1)
            self.assertEqual(stats["untranslated_translation_unit_count"], 1)

            pack = load_workbook(pack_path)
            units = pack["translation_units"]
            headers = [cell.value for cell in units[1]]
            source_idx = _unit_header_index(headers, "source_unit") + 1
            target_idx = _unit_header_index(headers, "target_unit") + 1
            variables_idx = headers.index("variables") + 1
            source_to_row = {
                row[source_idx - 1].value: row
                for row in units.iter_rows(min_row=2)
            }
            self.assertEqual(
                source_to_row["{1}"][target_idx - 1].value,
                "{1}",
            )
            self.assertEqual(
                source_to_row["{1>VIP{num1} Pack<2}"][variables_idx - 1].value,
                "{num1}=10,20",
            )
            source_to_row["{1>VIP{num1} Pack<2}"][target_idx - 1].value = (
                "{1>Pack VIP{num1}<2}"
            )
            pack.save(pack_path)

            fill_stats = fill_target_column_workbook(
                input_path,
                filled_path,
                source_col="source",
                target_col="target",
                template_workbook=pack_path,
                min_group_size=2,
            )

            self.assertEqual(fill_stats["autofilled_count"], 3)
            filled = load_workbook(filled_path, data_only=True)
            rows = list(filled.active.iter_rows(values_only=True))
            self.assertEqual(rows[1][1], '<img src="coin.png"/>')
            self.assertEqual(rows[2][1], '<color=#fff>Pack VIP10</color>')
            self.assertEqual(rows[3][1], '<color=#fff>Pack VIP20</color>')

    def test_read_source_rows_serializes_source_and_existing_target_tags(self):
        from phraseloom.excel_io import _read_source_rows

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "tagged.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(
                [
                    '<color=#fff>VIP10</color>',
                    '<color=#fff>VIP10 Pack FR</color>',
                ]
            )
            wb.save(input_path)

            rows = _read_source_rows(input_path, "source", "target")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].raw_source, '<color=#fff>VIP10</color>')
        self.assertEqual(
            rows[0].raw_existing_target, '<color=#fff>VIP10 Pack FR</color>'
        )
        self.assertEqual(rows[0].source, "{1>VIP10<2}")
        self.assertEqual(rows[0].existing_target, "{1>VIP10 Pack FR<2}")
        self.assertEqual(rows[0].match.template, "{1>VIP{num1}<2}")
        self.assertEqual(rows[0].match.values, {"num1": "10"})
        self.assertEqual(rows[0].tag_warnings, ())
        self.assertEqual(rows[0].target_tag_warnings, ())

    def test_custom_tag_config_controls_workflow_extraction(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "source.xlsx"
            output_path = tmp_path / "pack.xlsx"
            config_path = tmp_path / "tag_rules.toml"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<foo>Power {a}</foo>", ""])
            ws.append(["<foo>Power {b}</foo>", ""])
            wb.save(input_path)

            config_path.write_text(
                "\n".join(
                    [
                        "version = 1",
                        "",
                        "[angle_tags]",
                        'mode = "allowlist"',
                        'allowed = ["foo"]',
                        "",
                        "[bbcode_tags]",
                        'mode = "allowlist"',
                        'allowed = ["color"]',
                        "",
                        "[raw_braces]",
                        "protect_all = true",
                    ]
                ),
                encoding="utf-8",
            )

            generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col="target",
                min_group_size=2,
                use_existing_targets=False,
                tag_config=config_path,
            )

            out = load_workbook(output_path, data_only=True)
            units = out["translation_units"]
            headers = [cell.value for cell in units[1]]
            source_idx = _unit_header_index(headers, "source_unit")
            source_units = [
                row[source_idx]
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            out.close()

        self.assertIn("{1>Power {2}<3}", source_units)

    def test_cli_accepts_tag_config_option(self):
        from phraseloom.cli import main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "source.xlsx"
            output_path = tmp_path / "pack.xlsx"
            config_path = tmp_path / "tag_rules.toml"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<foo>Power {a}</foo>", ""])
            ws.append(["<foo>Power {b}</foo>", ""])
            wb.save(input_path)

            config_path.write_text(
                "\n".join(
                    [
                        "version = 1",
                        "",
                        "[angle_tags]",
                        'mode = "allowlist"',
                        'allowed = ["foo"]',
                        "",
                        "[bbcode_tags]",
                        'mode = "allowlist"',
                        'allowed = ["color"]',
                        "",
                        "[raw_braces]",
                        "protect_all = true",
                    ]
                ),
                encoding="utf-8",
            )

            exit_code = main(
                [
                    "extract",
                    str(input_path),
                    "-o",
                    str(output_path),
                    "--source-col",
                    "source",
                    "--target-col",
                    "target",
                    "--tag-config",
                    str(config_path),
                    "--no-existing-targets",
                ]
            )
            self.assertEqual(exit_code, 0)
            self.assertTrue(output_path.exists())

    def test_raw_brace_placeholders_are_translator_facing_protected_tokens(self):
        from phraseloom.workflow import fill_target_column_workbook, generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            pack_path = Path(tmp) / "pack.xlsx"
            filled_path = Path(tmp) / "filled.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Hit deals {0} damage", ""])
            ws.append(["Hit deals {value} damage", ""])
            wb.save(input_path)

            generate_workbook(
                input_path,
                pack_path,
                source_col="source",
                target_col="target",
                min_group_size=2,
                use_existing_targets=False,
            )

            pack = load_workbook(pack_path)
            units = pack["translation_units"]
            headers = [cell.value for cell in units[1]]
            source_idx = _unit_header_index(headers, "source_unit") + 1
            target_idx = _unit_header_index(headers, "target_unit") + 1
            rows = {
                row[source_idx - 1].value: row
                for row in units.iter_rows(min_row=2)
            }

            self.assertIn("Hit deals {1} damage", rows)
            rows["Hit deals {1} damage"][target_idx - 1].value = "Inflige {1} degats"
            pack.save(pack_path)

            fill_target_column_workbook(
                input_path,
                filled_path,
                source_col="source",
                target_col="target",
                template_workbook=pack_path,
                min_group_size=2,
            )

            filled = load_workbook(filled_path, data_only=True)
            output_rows = list(filled.active.iter_rows(values_only=True))
            self.assertEqual(output_rows[1][1], "Inflige {0} degats")
            self.assertEqual(output_rows[2][1], "Inflige {value} degats")

    def test_existing_target_raw_braces_keep_placeholder_order(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["饱食度：{0}/{1}", "Satiété : {0}/{1}"])
            wb.save(input_path)

            generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col="target",
                use_existing_targets=True,
            )

            pack = load_workbook(output_path, data_only=True)
            try:
                units = pack["translation_units"]
                unit_headers = [cell.value for cell in units[1]]
                unit = dict(zip(unit_headers, next(units.iter_rows(min_row=2, values_only=True))))
                self.assertEqual(_unit_row_value(unit, "source_unit"), "饱食度：{1}/{2}")
                self.assertEqual(_unit_row_value(unit, "target_unit"), "Satiété : {1}/{2}")

                filled = pack["filled_workbook"]
                filled_headers = [cell.value for cell in filled[1]]
                filled_row = dict(
                    zip(filled_headers, next(filled.iter_rows(min_row=2, values_only=True)))
                )
                self.assertEqual(filled_row["auto_target"], "Satiété : {0}/{1}")
            finally:
                pack.close()

    def test_existing_target_template_with_adjacent_color_digits_writes_workbook(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Difficulty: [63E7A4]1[-]", "難易度：[63E7A4]1[-]"])
            ws.append(["Difficulty: [63E7A4]2[-]", "難易度：[63E7A4]2[-]"])
            ws.append(["Difficulty: [63E7A4]3[-]", "難易度：[63E7A4]3[-]"])
            wb.save(input_path)

            generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col="target",
                use_existing_targets=True,
            )

            pack = load_workbook(output_path, data_only=True)
            try:
                units = pack["translation_units"]
                unit_headers = [cell.value for cell in units[1]]
                unit = dict(zip(unit_headers, next(units.iter_rows(min_row=2, values_only=True))))
                target_unit = _unit_row_value(unit, "target_unit")

                self.assertNotIn("\x00", target_unit)
                self.assertEqual(
                    target_unit,
                    "難易度：[{num1}E{num2}A{num3}]{num4}[-]",
                )
            finally:
                pack.close()

    def test_row_item_carries_optional_tag_metadata(self):
        from phraseloom.models import RowFillResult, RowItem
        from phraseloom.template_engine import parse_template

        row = RowItem(2, "VIP{num1}", "", parse_template("VIP10"), ("VIP10",))
        result = RowFillResult(row=row, unit=None, auto_target=None, warning="tag warning")

        self.assertEqual(row.raw_source, "")
        self.assertEqual(row.raw_existing_target, "")
        self.assertEqual(row.tag_tokens, ())
        self.assertEqual(row.tag_warnings, ())
        self.assertEqual(row.target_tag_warnings, ())
        self.assertEqual(result.warning, "tag warning")

    def test_generated_workbook_prefills_auto_targets_from_cli_example(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet2"
            ws.append(["中文", "source", "target"])
            ws.append(["VIP0付费礼包", "VIP0 Paid Pack", ""])
            ws.append(["VIP10付费礼包", "VIP10 Paid Pack", ""])
            ws.append(["登录游戏3天", "Sign in for 3 days", ""])
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col="target",
                examples=[("VIP10 Paid Pack", "VIP10pack")],
                min_group_size=2,
            )

            self.assertEqual(stats["template_count"], 1)
            self.assertEqual(stats["clustered_source_segments"], 2)
            self.assertEqual(stats["autofilled_count"], 2)

            out = load_workbook(output_path, data_only=True)
            summary = {
                row[0].value: row[1].value
                for row in out["summary"].iter_rows(min_row=1, max_col=2)
            }
            self.assertEqual(
                summary,
                {
                    "schema_version": "1.0",
                    "total_source_rows": 3,
                    "total_translation_units": 2,
                    "already_filled_units": 1,
                    "already_filled_source_rows": 2,
                    "units_to_translate": 1,
                    "source_rows_to_translate": 1,
                },
            )

            units = out["translation_units"]
            unit_headers = [cell.value for cell in units[1]]
            unit_rows = [
                dict(zip(unit_headers, row))
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            by_source = {_unit_row_value(row, "source_unit"): row for row in unit_rows}
            self.assertEqual(by_source["VIP{num1} Paid Pack"]["unit_type"], "template")
            self.assertEqual(by_source["Sign in for 3 days"]["unit_type"], "segment")

            auto = out["source_map"]
            headers = [cell.value for cell in auto[1]]
            source_idx = headers.index("source") + 1
            auto_idx = headers.index("auto_target") + 1

            actual = {
                row[source_idx - 1].value: row[auto_idx - 1].value
                for row in auto.iter_rows(min_row=2)
            }

            self.assertEqual(actual["VIP0 Paid Pack"], "VIP0pack")
            self.assertEqual(actual["VIP10 Paid Pack"], "VIP10pack")
            self.assertEqual(actual["Sign in for 3 days"], None)

    def test_raw_tagged_cli_examples_prefill_serialized_tagged_rows(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(['<color=#fff>VIP10 Pack</color>', ""])
            ws.append(['<color=#fff>VIP20 Pack</color>', ""])
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col="target",
                examples=[
                    (
                        '<color=#fff>VIP10 Pack</color>',
                        '<color=#fff>Pack VIP10 FR</color>',
                    )
                ],
                min_group_size=2,
                use_existing_targets=False,
            )

            self.assertEqual(stats["autofilled_count"], 2)

            out = load_workbook(output_path, data_only=True)
            filled = out["filled_workbook"]
            headers = [cell.value for cell in filled[1]]
            source_idx = headers.index("source") + 1
            auto_idx = headers.index("auto_target") + 1
            actual = {
                row[source_idx - 1].value: row[auto_idx - 1].value
                for row in filled.iter_rows(min_row=2)
            }

            self.assertEqual(
                actual['<color=#fff>VIP10 Pack</color>'],
                '<color=#fff>Pack VIP10 FR</color>',
            )
            self.assertEqual(
                actual['<color=#fff>VIP20 Pack</color>'],
                '<color=#fff>Pack VIP20 FR</color>',
            )

    def test_minimal_translation_units_deduplicate_segments_and_fill_duplicate_rows(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            pack_path = Path(tmp) / "pack.xlsx"
            filled_path = Path(tmp) / "filled.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["source", "target"])
            ws.append(["Obtain 3 stars in Chapter 1", ""])
            ws.append(["Obtain 3 stars in Chapter 1", ""])
            ws.append(["Obtain 3 stars in Chapter 2", ""])
            ws.append(["Login failed", ""])
            ws.append(["Login failed", ""])
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                pack_path,
                source_col="source",
                target_col="target",
                min_group_size=2,
                use_existing_targets=False,
            )
            self.assertEqual(stats["template_count"], 1)
            self.assertEqual(stats["segment_unit_count"], 1)
            self.assertEqual(stats["translation_unit_count"], 2)
            self.assertEqual(stats["duplicate_source_segments"], 2)

            pack = load_workbook(pack_path)
            units = pack["translation_units"]
            headers = [cell.value for cell in units[1]]
            target_idx = _unit_header_index(headers, "target_unit") + 1
            source_idx = _unit_header_index(headers, "source_unit") + 1
            for row in units.iter_rows(min_row=2):
                source_unit = row[source_idx - 1].value
                if source_unit == "Obtain {num1} stars in Chapter {num2}":
                    row[target_idx - 1].value = "第{num2}章获得{num1}颗星"
                elif source_unit == "Login failed":
                    row[target_idx - 1].value = "登录失败"
            pack.save(pack_path)

            fill_stats = generate_workbook(
                input_path,
                filled_path,
                source_col="source",
                target_col="target",
                template_workbook=pack_path,
                min_group_size=2,
                use_existing_targets=False,
            )
            self.assertEqual(fill_stats["autofilled_count"], 5)

            filled = load_workbook(filled_path, data_only=True)
            source_map = filled["source_map"]
            headers = [cell.value for cell in source_map[1]]
            targets = [
                row[headers.index("auto_target")]
                for row in source_map.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(
                targets,
                [
                    "第1章获得3颗星",
                    "第1章获得3颗星",
                    "第2章获得3颗星",
                    "登录失败",
                    "登录失败",
                ],
            )

    def test_tm_pairs_prefill_translation_units_and_leave_new_units_blank(self):
        from phraseloom.workflow import generate_tm_pairs, generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tm_input = Path(tmp) / "tm.xlsx"
            tm_pairs = Path(tmp) / "tm_pairs.xlsx"
            source_input = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "new_units.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Obtain 3 stars in Chapter 1", "第1章获得3颗星"])
            ws.append(["Obtain 3 stars in Chapter 2", "第2章获得3颗星"])
            ws.append(["Login failed", "登录失败"])
            ws.append(["Login failed", "登录失败"])
            wb.save(tm_input)

            tm_stats = generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
                min_group_size=2,
            )
            self.assertEqual(tm_stats["tm_pair_count"], 2)
            self.assertEqual(tm_stats["template_pair_count"], 1)
            self.assertEqual(tm_stats["segment_pair_count"], 1)

            wb = Workbook()
            ws = wb.active
            ws.append(["source"])
            ws.append(["Obtain 5 stars in Chapter 3"])
            ws.append(["Login failed"])
            ws.append(["Brand new line"])
            wb.save(source_input)

            stats = generate_workbook(
                source_input,
                output_path,
                source_col="source",
                target_col=None,
                tm_workbook=tm_pairs,
                min_group_size=2,
                use_existing_targets=False,
            )
            self.assertEqual(stats["translation_unit_count"], 3)
            self.assertEqual(stats["prefilled_translation_unit_count"], 2)
            self.assertEqual(stats["untranslated_translation_unit_count"], 1)
            self.assertEqual(stats["new_source_segment_count"], 1)

            out = load_workbook(output_path, data_only=True)
            self.assertIn("to_translate", out.sheetnames)
            self.assertIn("filled_workbook", out.sheetnames)

            summary = {
                row[0].value: row[1].value
                for row in out["summary"].iter_rows(min_row=1, max_col=2)
            }
            self.assertEqual(
                summary,
                {
                    "schema_version": "1.0",
                    "total_source_rows": 3,
                    "total_translation_units": 3,
                    "already_filled_units": 2,
                    "already_filled_source_rows": 2,
                    "units_to_translate": 1,
                    "source_rows_to_translate": 1,
                },
            )

            units = out["translation_units"]
            headers = [cell.value for cell in units[1]]
            rows = [
                dict(zip(headers, row))
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            by_source = {_unit_row_value(row, "source_unit"): row for row in rows}

            self.assertEqual(
                _unit_row_value(by_source["Obtain {num1} stars in Chapter {num2}"], "target_unit"),
                "第{num2}章获得{num1}颗星",
            )
            self.assertEqual(
                by_source["Obtain {num1} stars in Chapter {num2}"]["target_unit_source"],
                "tm_pairs",
            )
            self.assertEqual(
                _unit_row_value(by_source["Login failed"], "target_unit"),
                "登录失败",
            )
            self.assertIsNone(
                _unit_row_value(by_source["Brand new line"], "target_unit")
            )

            todo = out["to_translate"]
            todo_headers = [cell.value for cell in todo[1]]
            todo_rows = [
                dict(zip(todo_headers, row))
                for row in todo.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(len(todo_rows), 1)
            self.assertEqual(_unit_row_value(todo_rows[0], "source_unit"), "Brand new line")

            filled = out["filled_workbook"]
            filled_headers = [cell.value for cell in filled[1]]
            self.assertIn("auto_target", filled_headers)
            self.assertIn("fill_status", filled_headers)
            filled_rows = [
                dict(zip(filled_headers, row))
                for row in filled.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(filled_rows[0]["auto_target"], "第3章获得5颗星")
            self.assertEqual(filled_rows[1]["auto_target"], "登录失败")
            self.assertEqual(filled_rows[2]["fill_status"], "missing_target_unit")

            standalone_todo = source_input.parent / "source_l10n" / "source_translator_todo.xlsx"
            self.assertTrue(standalone_todo.exists())
            todo_book = load_workbook(standalone_todo, data_only=True)
            self.assertEqual(
                todo_book.sheetnames,
                ["Sheet", "to_translate", "prefilled_units", "_metadata"],
            )
            self.assertEqual(todo_book["Sheet"].sheet_state, "hidden")
            self.assertEqual(todo_book["to_translate"].sheet_state, "visible")
            self.assertEqual(todo_book["prefilled_units"].sheet_state, "visible")

    def test_non_translatable_numeric_and_symbol_segments_are_autofilled(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source"])
            ws.append(["123"])
            ws.append(["..."])
            ws.append(["10-20"])
            ws.append(["$0.99"])
            ws.append(["New text"])
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col=None,
                min_group_size=2,
                use_existing_targets=False,
            )

            self.assertEqual(stats["autofilled_count"], 4)
            self.assertEqual(stats["new_translation_unit_count"], 1)
            self.assertEqual(stats["new_source_segment_count"], 1)

            out = load_workbook(output_path, data_only=True)
            units = out["translation_units"]
            headers = [cell.value for cell in units[1]]
            rows = [
                dict(zip(headers, row))
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            by_source = {_unit_row_value(row, "source_unit"): row for row in rows}

            for source in ["123", "...", "10-20", "$0.99"]:
                self.assertEqual(_unit_row_value(by_source[source], "target_unit"), source)
                self.assertEqual(by_source[source]["target_unit_source"], "non_translatable")

            todo = out["to_translate"]
            todo_headers = [cell.value for cell in todo[1]]
            todo_rows = [
                dict(zip(todo_headers, row))
                for row in todo.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(
                [_unit_row_value(row, "source_unit") for row in todo_rows],
                ["New text"],
            )

    def test_translator_todo_orders_units_by_first_source_row_and_includes_context(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "context"])
            ws.append(["Intro line", "Opening scene"])
            ws.append(["VIP10 Pack", "Shop scene"])
            ws.append(["Later line", "Battle scene"])
            ws.append(["VIP20 Pack", "Shop scene"])
            wb.save(input_path)

            generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col=None,
                min_group_size=2,
                use_existing_targets=False,
            )

            todo_path = input_path.parent / "source_l10n" / "source_translator_todo.xlsx"
            todo_book = load_workbook(todo_path, data_only=True)
            try:
                todo = todo_book["to_translate"]
                headers = [cell.value for cell in todo[1]]
                rows = [
                    dict(zip(headers, row))
                    for row in todo.iter_rows(min_row=2, values_only=True)
                ]
            finally:
                todo_book.close()

            self.assertEqual(
                headers,
                [
                    "unit_id",
                    "unit_type",
                    "source",
                    "target",
                    "sample_sources",
                    "context",
                    "row_number",
                    "coverage_count",
                    "variables",
                    "warning",
                    "translator_note",
                ],
            )
            self.assertEqual(
                [_unit_row_value(row, "source_unit") for row in rows],
                ["Intro line", "VIP{num1} Pack", "Later line"],
            )
            self.assertEqual([row["row_number"] for row in rows], [2, 3, 4])
            self.assertEqual(
                [row["context"] for row in rows],
                ["Opening scene", "Shop scene", "Battle scene"],
            )
            self.assertEqual(rows[1]["sample_sources"], "VIP10 Pack")

    def test_prepare_accepts_a_custom_context_column(self):
        from phraseloom.workflow import prepare_translation_package

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target", "screen_notes"])
            ws.append(["Start", None, "Login screen"])
            wb.save(input_path)

            stats = prepare_translation_package(
                input_path,
                source_col="source",
                target_col="target",
                context_col="screen_notes",
            )

            package = load_workbook(stats["to_translate_path"], data_only=True)
            try:
                todo = package["to_translate"]
                headers = [cell.value for cell in todo[1]]
                row = dict(zip(headers, next(todo.iter_rows(min_row=2, values_only=True))))
            finally:
                package.close()

            self.assertEqual(row["context"], "Login screen")

    def test_tm_output_includes_custom_context_in_pairs_and_map(self):
        from phraseloom.workflow import generate_tm_pairs

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "tm.xlsx"
            output_path = Path(tmp) / "tm_pairs.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target", "screen_notes"])
            ws.append(["Start", "Démarrer", "Login screen"])
            wb.save(input_path)

            generate_tm_pairs(
                input_path,
                output_path,
                context_col="screen_notes",
            )

            result = load_workbook(output_path, data_only=True)
            try:
                pairs = result["tm_pairs"]
                pair_headers = [cell.value for cell in pairs[1]]
                pair = dict(
                    zip(
                        pair_headers,
                        next(pairs.iter_rows(min_row=2, values_only=True)),
                    )
                )
                tm_map = result["tm_map"]
                map_headers = [cell.value for cell in tm_map[1]]
                mapped = dict(
                    zip(
                        map_headers,
                        next(tm_map.iter_rows(min_row=2, values_only=True)),
                    )
                )
            finally:
                result.close()

            self.assertEqual(pair["context"], "Login screen")
            self.assertEqual(mapped["context"], "Login screen")

    def test_generated_unit_tables_use_source_and_target_headers(self):
        from phraseloom.workflow import (
            fill_target_column_workbook,
            generate_tm_pairs,
            generate_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_input = tmp_path / "tm.xlsx"
            tm_pairs = tmp_path / "tm_pairs.xlsx"
            source_input = tmp_path / "source.xlsx"
            pack_path = tmp_path / "pack.xlsx"
            filled_path = tmp_path / "filled.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Login failed", "登录失败"])
            ws.append(["Login failed", "登录失败"])
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
            )

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "fr"])
            ws.append(["Login failed", ""])
            ws.append(["Brand new line", ""])
            wb.save(source_input)

            generate_workbook(
                source_input,
                pack_path,
                source_col="source",
                target_col="fr",
                tm_workbook=tm_pairs,
                use_existing_targets=False,
            )

            pack_book = load_workbook(pack_path)
            try:
                self.assertIn("source", [cell.value for cell in pack_book["translation_units"][1]])
                self.assertIn("target", [cell.value for cell in pack_book["translation_units"][1]])
                self.assertNotIn("source_unit", [cell.value for cell in pack_book["translation_units"][1]])
                self.assertNotIn("target_unit", [cell.value for cell in pack_book["translation_units"][1]])
            finally:
                pack_book.close()

            tm_book = load_workbook(tm_pairs)
            try:
                tm_headers = [cell.value for cell in tm_book["tm_pairs"][1]]
                self.assertIn("source", tm_headers)
                self.assertIn("target", tm_headers)
                self.assertNotIn("source_unit", tm_headers)
                self.assertNotIn("target_unit", tm_headers)
            finally:
                tm_book.close()

            todo_path = tmp_path / "source_l10n" / "source_translator_todo.xlsx"
            todo_book = load_workbook(todo_path)
            try:
                todo_ws = todo_book["to_translate"]
                todo_headers = [cell.value for cell in todo_ws[1]]
                self.assertIn("source", todo_headers)
                self.assertIn("target", todo_headers)
                self.assertNotIn("source_unit", todo_headers)
                self.assertNotIn("target_unit", todo_headers)
                target_idx = todo_headers.index("target") + 1
                for row in todo_ws.iter_rows(min_row=2):
                    row[target_idx - 1].value = "全新文本"

                prefilled_headers = [
                    cell.value for cell in todo_book["prefilled_units"][1]
                ]
                self.assertEqual(prefilled_headers, todo_headers)
                self.assertIn("source", prefilled_headers)
                self.assertIn("target", prefilled_headers)
                self.assertNotIn("source_unit", prefilled_headers)
                self.assertNotIn("target_unit", prefilled_headers)
                todo_book.save(todo_path)
            finally:
                todo_book.close()

            stats = fill_target_column_workbook(
                source_input,
                filled_path,
                source_col="source",
                target_col="fr",
                template_workbook=todo_path,
            )

        self.assertEqual(stats["autofilled_count"], 2)

    def test_translated_to_translate_can_fill_target_column_in_output_copy(self):
        from phraseloom.workflow import (
            fill_target_column_workbook,
            generate_tm_pairs,
            generate_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tm_input = Path(tmp) / "tm.xlsx"
            tm_pairs = Path(tmp) / "tm_pairs.xlsx"
            source_input = Path(tmp) / "source.xlsx"
            pack_path = Path(tmp) / "source_pack.xlsx"
            target_output = Path(tmp) / "source_filled_target.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Obtain 3 stars in Chapter 1", "第1章获得3颗星"])
            ws.append(["Obtain 3 stars in Chapter 2", "第2章获得3颗星"])
            ws.append(["Login failed", "登录失败"])
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
                min_group_size=2,
            )

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "fr"])
            ws.append(["Obtain 5 stars in Chapter 3", ""])
            ws.append(["Login failed", ""])
            ws.append(["Brand new line", ""])
            wb.save(source_input)

            generate_workbook(
                source_input,
                pack_path,
                source_col="source",
                target_col="fr",
                tm_workbook=tm_pairs,
                min_group_size=2,
                use_existing_targets=False,
            )

            standalone_todo = source_input.parent / "source_l10n" / "source_translator_todo.xlsx"
            todo_book = load_workbook(standalone_todo)
            ws = todo_book["to_translate"]
            headers = [cell.value for cell in ws[1]]
            target_idx = _unit_header_index(headers, "target_unit") + 1
            for row in ws.iter_rows(min_row=2):
                row[target_idx - 1].value = "全新文本"
            todo_book.save(standalone_todo)

            stats = fill_target_column_workbook(
                source_input,
                target_output,
                source_col="source",
                target_col="fr",
                template_workbook=standalone_todo,
                min_group_size=2,
            )
            self.assertEqual(stats["autofilled_count"], 3)

            filled = load_workbook(target_output, data_only=True)
            rows = list(filled.active.iter_rows(values_only=True))
            self.assertEqual(rows[0], ("source", "fr"))
            self.assertEqual(rows[1][1], "第3章获得5颗星")
            self.assertEqual(rows[2][1], "登录失败")
            self.assertEqual(rows[3][1], "全新文本")

    def test_target_column_fill_writes_restore_audit_workbook(self):
        from phraseloom.workflow import fill_target_column_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_input = tmp_path / "source.xlsx"
            todo_path = tmp_path / "todo.xlsx"
            filled_path = tmp_path / "filled.xlsx"
            audit_path = tmp_path / "filled_restore_audit.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Keep {0}", ""])
            ws.append(["Missing line", ""])
            ws.append(["Bad <color>tag", ""])
            wb.save(source_input)

            wb = Workbook()
            ws = wb.active
            ws.title = "to_translate"
            ws.append(["unit_type", "source", "target"])
            ws.append(["segment", "Keep {1}", "Garde {1} {2}"])
            ws.append(["segment", "Bad {1>tag", "Balise {1>"])
            wb.save(todo_path)

            stats = fill_target_column_workbook(
                source_input,
                filled_path,
                source_col="source",
                target_col="target",
                template_workbook=todo_path,
            )

            self.assertIn("audit_output_path", stats)
            self.assertEqual(stats["audit_output_path"], str(audit_path))
            self.assertTrue(audit_path.exists())

            audit = load_workbook(audit_path, data_only=True)
            try:
                self.assertEqual(
                    audit.sheetnames[:2],
                    ["summary", "restore_warnings"],
                )
                self.assertIn("_metadata", audit.sheetnames)

                summary = {
                    row[0]: row[1]
                    for row in audit["summary"].iter_rows(min_row=2, values_only=True)
                }
                self.assertEqual(summary["filled_rows"], 2)
                self.assertEqual(summary["unfilled_rows"], 1)
                self.assertEqual(summary["warning_rows"], 3)
                self.assertEqual(summary["source_warning_rows"], 1)
                self.assertEqual(summary["target_warning_rows"], 1)
                self.assertEqual(summary["restore_warning_rows"], 1)

                warning_headers = [
                    cell.value for cell in audit["restore_warnings"][1]
                ]
                warnings = list(
                    audit["restore_warnings"].iter_rows(min_row=2, values_only=True)
                )
                warning_rows = {
                    row[warning_headers.index("row_number")]: dict(
                        zip(warning_headers, row)
                    )
                    for row in warnings
                }
                self.assertEqual(len(warning_rows), 3)
                self.assertIn(
                    "protected_token_mismatch",
                    warning_rows[2]["target_warning"],
                )
                self.assertIsNone(warning_rows[2]["source_warning"])
                self.assertEqual(
                    warning_rows[3]["restore_warning"],
                    "fill target in to_translate, then rerun fill",
                )
                self.assertIn(
                    "open tag has no close partner",
                    warning_rows[4]["source_warning"],
                )
                self.assertIsNone(warning_rows[4]["target_warning"])
            finally:
                audit.close()

    def test_fill_cli_accepts_restore_audit_output_path(self):
        from phraseloom.cli import main

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_input = tmp_path / "source.xlsx"
            todo_path = tmp_path / "todo.xlsx"
            filled_path = tmp_path / "filled.xlsx"
            audit_path = tmp_path / "audit.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Keep {0}", ""])
            wb.save(source_input)

            wb = Workbook()
            ws = wb.active
            ws.title = "to_translate"
            ws.append(["unit_type", "source", "target"])
            ws.append(["segment", "Keep {1}", "Garde {1}"])
            wb.save(todo_path)

            stdout = StringIO()
            with redirect_stdout(stdout):
                exit_code = main(
                    [
                        "fill",
                        str(source_input),
                        "--templates",
                        str(todo_path),
                        "--mode",
                        "target-column",
                        "-o",
                        str(filled_path),
                        "--audit-output",
                        str(audit_path),
                    ]
                )

            self.assertEqual(exit_code, 0)
            self.assertTrue(filled_path.exists())
            self.assertTrue(audit_path.exists())
            self.assertIn(f"Restore audit workbook: {audit_path}", stdout.getvalue())

    def test_tm_prefill_restores_current_row_color_attributes(self):
        from phraseloom.workflow import (
            fill_target_column_workbook,
            generate_tm_pairs,
            generate_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_input = tmp_path / "tm.xlsx"
            tm_pairs = tmp_path / "tm_pairs.xlsx"
            source_input = tmp_path / "source.xlsx"
            pack_path = tmp_path / "pack.xlsx"
            filled_path = tmp_path / "filled.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<color=#123>Source</>", "<color=#123>Target</>"])
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
            )

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<color=#3333>Source</>", ""])
            wb.save(source_input)

            generate_workbook(
                source_input,
                pack_path,
                source_col="source",
                target_col="target",
                tm_workbook=tm_pairs,
                use_existing_targets=False,
            )
            fill_target_column_workbook(
                source_input,
                filled_path,
                source_col="source",
                target_col="target",
                template_workbook=pack_path,
            )

            filled = load_workbook(filled_path, data_only=True)
            rows = list(filled.active.iter_rows(values_only=True))
            filled.close()

        self.assertEqual(rows[1][1], "<color=#3333>Target</>")

    def test_tm_prefill_keeps_unlisted_angle_labels_translatable(self):
        from phraseloom.workflow import generate_tm_pairs

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_input = tmp_path / "tm.xlsx"
            tm_pairs = tmp_path / "tm_pairs.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(
                [
                    "<Activate> HP increased by {a}%",
                    "<Active> PV augmentes de {a}%",
                ]
            )
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
            )

            out = load_workbook(tm_pairs, data_only=True)
            ws = out["tm_pairs"]
            headers = [cell.value for cell in ws[1]]
            source_idx = _unit_header_index(headers, "source_unit")
            target_idx = _unit_header_index(headers, "target_unit")
            row = next(ws.iter_rows(min_row=2, values_only=True))
            out.close()

        self.assertEqual(row[source_idx], "<Activate> HP increased by {1}%")
        self.assertEqual(row[target_idx], "<Active> PV augmentes de {1}%")

    def test_tm_pair_warning_includes_target_serialization_warning(self):
        from phraseloom.workflow import generate_tm_pairs

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            tm_input = tmp_path / "tm.xlsx"
            tm_pairs = tmp_path / "tm_pairs.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<br/>", "no matching span"])
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
            )

            out = load_workbook(tm_pairs, data_only=True)
            ws = out["tm_pairs"]
            headers = [cell.value for cell in ws[1]]
            warning_idx = headers.index("warning")
            row = next(ws.iter_rows(min_row=2, values_only=True))
            out.close()

        self.assertIn("source_protected_span_not_found: <br/>", row[warning_idx])

    def test_source_map_warning_deduplicates_row_serialization_warning(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "source.xlsx"
            output_path = tmp_path / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<br/>", "no matching span"])
            wb.save(input_path)

            generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col="target",
            )

            out = load_workbook(output_path, data_only=True)
            ws = out["source_map"]
            headers = [cell.value for cell in ws[1]]
            warning_idx = headers.index("warning")
            row = next(ws.iter_rows(min_row=2, values_only=True))
            out.close()

        warning = row[warning_idx]
        self.assertIsNotNone(warning)
        self.assertEqual(warning.count("source_protected_span_not_found: <br/>"), 1)

    def test_source_map_warning_preserves_semicolon_bearing_tag_payload(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "source.xlsx"
            output_path = tmp_path / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<color=a;b/>", "no matching span"])
            wb.save(input_path)

            generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col="target",
                use_existing_targets=True,
            )

            out = load_workbook(output_path, data_only=True)
            ws = out["source_map"]
            headers = [cell.value for cell in ws[1]]
            warning_idx = headers.index("warning")
            row = next(ws.iter_rows(min_row=2, values_only=True))
            out.close()

        warning = row[warning_idx]
        self.assertIsNotNone(warning)
        parts = [part.strip() for part in warning.split("; ")]
        self.assertIn("source_protected_span_not_found: <color=a;b/>", parts)
        self.assertIn("protected_token_mismatch: missing {1}", parts)
        self.assertNotIn("source_protected_span_not_found: <color=a", parts)
        self.assertNotIn("b/>", parts)

    def test_default_output_paths_are_inside_input_work_folder(self):
        from phraseloom.excel_io import (
            _default_extract_output_path,
            _default_fill_output_path,
            _default_tm_output_path,
            _default_to_translate_output_path,
            _default_work_dir,
        )

        input_path = Path("/tmp/project/source.xlsx")

        self.assertEqual(
            _default_work_dir(input_path),
            Path("/tmp/project/source_l10n"),
        )
        self.assertEqual(
            _default_extract_output_path(input_path),
            Path("/tmp/project/source_l10n/source_tm_prefill_pack.xlsx"),
        )
        self.assertEqual(
            _default_to_translate_output_path(input_path),
            Path("/tmp/project/source_l10n/source_translator_todo.xlsx"),
        )
        self.assertEqual(
            _default_fill_output_path(input_path),
            Path("/tmp/project/source_l10n/source_filled_result.xlsx"),
        )
        self.assertEqual(
            _default_tm_output_path(input_path),
            Path("/tmp/project/source_l10n/source_reusable_units.xlsx"),
        )

    def test_missing_source_column_reports_available_headers(self):
        from phraseloom.errors import ColumnNotFoundError
        from phraseloom.excel_io import _read_source_rows

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Login failed", "登录失败"])
            wb.save(input_path)

            with self.assertRaises(ColumnNotFoundError) as raised:
                _read_source_rows(input_path, "missing", "target")

        message = str(raised.exception)
        self.assertIn("Column 'missing' not found", message)
        self.assertIn("Available columns: source, target", message)

    def test_generate_workbook_handles_sparse_header_row(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Login failed", ""])
            ws.cell(row=2, column=184).value = "trailing export data"
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col="target",
                use_existing_targets=False,
            )

            self.assertEqual(stats["row_count"], 1)
            self.assertTrue(output_path.exists())

    def test_malformed_translated_workbook_reports_required_columns(self):
        from phraseloom.errors import TranslationUnitLoadError
        from phraseloom.excel_io import _load_translated_units

        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "bad_to_translate.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "to_translate"
            ws.append(["unit_type", "source_unit"])
            ws.append(["segment", "Login failed"])
            wb.save(workbook_path)

            with self.assertRaises(TranslationUnitLoadError) as raised:
                _load_translated_units(workbook_path)

        message = str(raised.exception)
        self.assertIn("to_translate", message)
        self.assertIn("target_unit", message)
        self.assertIn("Available columns: unit_type, source_unit", message)

    def test_translated_workbook_without_supported_sheet_reports_expected_sheets(self):
        from phraseloom.errors import TranslationUnitLoadError
        from phraseloom.excel_io import _load_translated_units

        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "unsupported.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "notes"
            ws.append(["source", "target"])
            ws.append(["Login failed", "登录失败"])
            wb.save(workbook_path)

            with self.assertRaises(TranslationUnitLoadError) as raised:
                _load_translated_units(workbook_path)

        message = str(raised.exception)
        self.assertIn("does not contain a supported translation sheet", message)
        self.assertIn("tm_pairs", message)
        self.assertIn("translation_units", message)
        self.assertIn("to_translate", message)

    def test_generated_workbooks_include_schema_version_metadata(self):
        from phraseloom.workflow import generate_tm_pairs, generate_workbook
        from phraseloom.workbook_schema import (
            METADATA_SHEET,
            SCHEMA_VERSION,
            SCHEMA_VERSION_KEY,
        )

        with tempfile.TemporaryDirectory() as tmp:
            source_input = Path(tmp) / "source.xlsx"
            pack_output = Path(tmp) / "pack.xlsx"
            tm_input = Path(tmp) / "tm.xlsx"
            tm_output = Path(tmp) / "tm_pairs.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Login failed", ""])
            ws.append(["Login failed", ""])
            wb.save(source_input)

            generate_workbook(
                source_input,
                pack_output,
                source_col="source",
                target_col="target",
                min_group_size=2,
                use_existing_targets=False,
            )

            pack = load_workbook(pack_output, data_only=True)
            pack_summary = {
                row[0].value: row[1].value
                for row in pack["summary"].iter_rows(min_row=1, max_col=2)
            }
            self.assertEqual(pack_summary[SCHEMA_VERSION_KEY], SCHEMA_VERSION)

            standalone_todo = source_input.parent / "source_l10n" / "source_translator_todo.xlsx"
            todo = load_workbook(standalone_todo, data_only=True)
            self.assertIn(METADATA_SHEET, todo.sheetnames)
            self.assertEqual(todo[METADATA_SHEET].sheet_state, "hidden")
            todo_metadata = {
                row[0].value: row[1].value
                for row in todo[METADATA_SHEET].iter_rows(min_row=2, max_col=2)
            }
            self.assertEqual(todo_metadata[SCHEMA_VERSION_KEY], SCHEMA_VERSION)

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Login failed", "登录失败"])
            ws.append(["Login failed", "登录失败"])
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_output,
                source_col="source",
                target_col="target",
                min_group_size=2,
            )

            tm = load_workbook(tm_output, data_only=True)
            tm_summary = {
                row[0].value: row[1].value
                for row in tm["summary"].iter_rows(min_row=1, max_col=2)
            }
            self.assertEqual(tm_summary[SCHEMA_VERSION_KEY], SCHEMA_VERSION)

    def test_generated_workbooks_record_tag_rules_metadata(self):
        from phraseloom.tag_rules import default_tag_rules, normalized_tag_rules_hash
        from phraseloom.workflow import generate_tm_pairs, generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            source_input = tmp_path / "source.xlsx"
            pack_output = tmp_path / "pack.xlsx"
            tm_input = tmp_path / "tm.xlsx"
            tm_output = tmp_path / "tm_pairs.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<color=#123>Source</>", ""])
            wb.save(source_input)

            generate_workbook(
                source_input,
                pack_output,
                source_col="source",
                target_col="target",
                use_existing_targets=False,
            )

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<color=#123>Source</>", "<color=#123>Target</>"])
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_output,
                source_col="source",
                target_col="target",
            )

            expected_hash = normalized_tag_rules_hash(default_tag_rules())
            pack = load_workbook(pack_output, data_only=True)
            tm = load_workbook(tm_output, data_only=True)

            pack_metadata = {
                row[0]: row[1]
                for row in pack["_metadata"].iter_rows(min_row=2, values_only=True)
            }
            tm_metadata = {
                row[0]: row[1]
                for row in tm["_metadata"].iter_rows(min_row=2, values_only=True)
            }
            pack.close()
            tm.close()

        self.assertEqual(pack_metadata["tag_rules_version"], 1)
        self.assertEqual(pack_metadata["tag_rules_hash"], expected_hash)
        self.assertEqual(pack_metadata["tag_rules_source"], "default")
        self.assertEqual(tm_metadata["tag_rules_hash"], expected_hash)

    def test_tag_config_mismatch_reports_user_facing_error(self):
        from phraseloom.errors import ConfigError
        from phraseloom.workflow import generate_tm_pairs, generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            first_config = tmp_path / "first.toml"
            second_config = tmp_path / "second.toml"
            tm_input = tmp_path / "tm.xlsx"
            tm_pairs = tmp_path / "tm_pairs.xlsx"
            source_input = tmp_path / "source.xlsx"
            pack_output = tmp_path / "pack.xlsx"

            first_config.write_text(
                "\n".join(
                    [
                        "version = 1",
                        "",
                        "[angle_tags]",
                        'mode = "allowlist"',
                        'allowed = ["color"]',
                        "",
                        "[bbcode_tags]",
                        'mode = "allowlist"',
                        'allowed = ["color"]',
                        "",
                        "[raw_braces]",
                        "protect_all = true",
                    ]
                ),
                encoding="utf-8",
            )
            second_config.write_text(
                "\n".join(
                    [
                        "version = 1",
                        "",
                        "[angle_tags]",
                        'mode = "allowlist"',
                        'allowed = ["foo"]',
                        "",
                        "[bbcode_tags]",
                        'mode = "allowlist"',
                        'allowed = ["color"]',
                        "",
                        "[raw_braces]",
                        "protect_all = true",
                    ]
                ),
                encoding="utf-8",
            )

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<color=#123>Source</>", "<color=#123>Target</>"])
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
                tag_config=first_config,
            )

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<color=#333>Source</>", ""])
            wb.save(source_input)

            with self.assertRaises(ConfigError) as raised:
                generate_workbook(
                    source_input,
                    pack_output,
                    source_col="source",
                    target_col="target",
                    tm_workbook=tm_pairs,
                    use_existing_targets=False,
                    tag_config=second_config,
                )

        self.assertIn("tag config mismatch", str(raised.exception))

    def test_cli_reports_phraseloom_errors_without_traceback(self):
        from phraseloom.cli import main

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Login failed", ""])
            wb.save(input_path)

            stderr = StringIO()
            with patch("sys.stderr", stderr):
                exit_code = main(
                    [
                        "extract",
                        str(input_path),
                        "-o",
                        str(output_path),
                        "--source-col",
                        "missing",
                        "--target-col",
                        "target",
                    ]
                )

        self.assertEqual(exit_code, 1)
        self.assertIn("Column 'missing' not found", stderr.getvalue())
        self.assertIn("Available columns: source, target", stderr.getvalue())
        self.assertFalse(output_path.exists())

    def test_top_level_help_promotes_package_cli(self):
        from phraseloom.cli import main

        stdout = StringIO()
        with redirect_stdout(stdout):
            exit_code = main(["--help"])

        self.assertEqual(exit_code, 0)
        help_text = stdout.getvalue()
        self.assertIn("phraseloom tm-extract COMPLETED_TM.xlsx", help_text)
        self.assertIn("phraseloom prepare SOURCE.xlsx --tm TM_PAIRS.xlsx", help_text)
        self.assertIn("phraseloom fill TRANSLATOR_WORKBOOK.xlsx", help_text)
        self.assertIn("phraseloom extract SOURCE.xlsx", help_text)
        self.assertIn("python template_demo.py SOURCE.xlsx", help_text)

    def test_interactive_path_input_accepts_copied_shell_quotes(self):
        from phraseloom.interactive import _user_path

        self.assertEqual(
            _user_path("'/tmp/project/source.xlsx'"),
            Path("/tmp/project/source.xlsx"),
        )
        self.assertEqual(
            _user_path('"/tmp/project/source.xlsx"'),
            Path("/tmp/project/source.xlsx"),
        )

    def test_repeated_numeric_source_without_variants_stays_segment(self):
        from phraseloom.workflow import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source"])
            ws.append(["VIP10 Paid Pack"])
            ws.append(["VIP10 Paid Pack"])
            ws.append(["VIP10 Paid Pack"])
            wb.save(input_path)

            generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col=None,
                min_group_size=2,
                use_existing_targets=False,
            )

            out = load_workbook(output_path, data_only=True)
            units = out["translation_units"]
            headers = [cell.value for cell in units[1]]
            rows = [
                dict(zip(headers, row))
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["unit_type"], "segment")
            self.assertEqual(_unit_row_value(rows[0], "source_unit"), "VIP10 Paid Pack")
            self.assertEqual(rows[0]["coverage_count"], 3)

    def test_interactive_menu_shows_three_step_workflow(self):
        from phraseloom.cli import main

        stdout = StringIO()
        with patch("builtins.input", side_effect=["q"]), redirect_stdout(stdout):
            exit_code = main([])

        menu = stdout.getvalue()
        self.assertEqual(exit_code, 0)
        self.assertIn("1) Build TM from completed Excel", menu)
        self.assertIn("2) Prepare translator file for new source", menu)
        self.assertIn("3) Fill source from translated file", menu)
        self.assertIn("a) Advanced tools", menu)
        self.assertNotIn("4) Entity workflow", menu)
        self.assertNotIn("Extract source translation units", menu)

    def test_interactive_extract_creates_self_contained_translator_workbook(self):
        from phraseloom.cli import main

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "input_l10n" / "input_translator_todo.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet2"
            ws.append(["中文", "source", "target"])
            ws.append(["VIP0付费礼包", "VIP0 Paid Pack", ""])
            ws.append(["VIP10付费礼包", "VIP10 Paid Pack", ""])
            wb.save(input_path)

            answers = iter(
                [
                    "2",
                    f"'{input_path}'",
                    "",
                ]
            )
            stdout = StringIO()
            with patch("builtins.input", side_effect=lambda _="": next(answers)), redirect_stdout(stdout):
                exit_code = main([])

            self.assertEqual(exit_code, 0)
            self.assertTrue(output_path.exists())
            self.assertNotIn("Optional: add source=target examples", stdout.getvalue())
            self.assertNotIn("Minimum variants", stdout.getvalue())
            self.assertNotIn("Output process workbook", stdout.getvalue())

            out = load_workbook(output_path, data_only=True)
            self.assertEqual(
                out.sheetnames,
                ["Sheet2", "to_translate", "prefilled_units", "_metadata"],
            )
            self.assertEqual(out["Sheet2"].sheet_state, "hidden")
            self.assertEqual(out["to_translate"].max_row, 2)
            self.assertEqual(out["prefilled_units"].max_row, 1)
            out.close()

    def test_translation_package_prefilled_and_todo_edits_fill_original_copy(self):
        from phraseloom.workflow import (
            fill_translation_package,
            prepare_translation_package,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "source.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Strings"
            ws.append(["source", "target"])
            ws.append(["Hello", "Bonjour old"])
            ws.append(["New text", None])
            wb.save(input_path)
            wb.close()

            prepare_stats = prepare_translation_package(
                input_path,
                use_existing_targets=True,
            )
            package_path = Path(str(prepare_stats["to_translate_path"]))
            package = load_workbook(package_path)
            try:
                self.assertEqual(package["Strings"].sheet_state, "hidden")
                self.assertEqual(package["to_translate"].sheet_state, "visible")
                self.assertEqual(package["prefilled_units"].sheet_state, "visible")
                for sheet_name, target in (
                    ("to_translate", "Nouveau texte"),
                    ("prefilled_units", "Bonjour modifié"),
                ):
                    sheet = package[sheet_name]
                    headers = [cell.value for cell in sheet[1]]
                    sheet.cell(
                        row=2,
                        column=headers.index("target") + 1,
                    ).value = target
                package.save(package_path)
            finally:
                package.close()

            fill_stats = fill_translation_package(package_path)
            output_path = Path(str(fill_stats["output_path"]))
            self.assertNotIn("audit_output_path", fill_stats)

            filled = load_workbook(output_path, data_only=True)
            try:
                self.assertEqual(filled.sheetnames, ["Strings"])
                self.assertEqual(filled["Strings"].sheet_state, "visible")
                self.assertEqual(
                    list(filled["Strings"].values),
                    [
                        ("source", "target"),
                        ("Hello", "Bonjour modifié"),
                        ("New text", "Nouveau texte"),
                    ],
                )
            finally:
                filled.close()

    def test_translation_package_writes_review_workbook_only_for_unfilled_rows(self):
        from phraseloom.workflow import (
            fill_translation_package,
            prepare_translation_package,
        )

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Still missing", None])
            wb.save(input_path)
            wb.close()

            prepare_stats = prepare_translation_package(input_path)
            fill_stats = fill_translation_package(prepare_stats["to_translate_path"])

            self.assertIn("audit_output_path", fill_stats)
            self.assertTrue(Path(str(fill_stats["audit_output_path"])).exists())

    def test_interactive_prepare_asks_for_tm_and_current_target_prefill(self):
        from phraseloom.cli import main

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Hello", "Existing translation"])
            wb.save(input_path)
            wb.close()

            answers = iter(["2", str(input_path), "", "n"])
            prompts = []

            def answer(prompt=""):
                prompts.append(prompt)
                return next(answers)

            with patch("builtins.input", side_effect=answer), redirect_stdout(StringIO()):
                self.assertEqual(main([]), 0)

            prompt_text = "\n".join(prompts)
            self.assertIn("TM workbook path", prompt_text)
            self.assertIn("Use current target values as prefill", prompt_text)
            self.assertNotIn("Minimum variants", prompt_text)
            self.assertNotIn("Output process workbook", prompt_text)

    def test_interactive_fill_needs_only_translator_workbook_path(self):
        from phraseloom.cli import main
        from phraseloom.workflow import prepare_translation_package

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Hello", None])
            wb.save(input_path)
            wb.close()

            prepare_stats = prepare_translation_package(input_path)
            package_path = Path(str(prepare_stats["to_translate_path"]))
            package = load_workbook(package_path)
            todo = package["to_translate"]
            headers = [cell.value for cell in todo[1]]
            todo.cell(row=2, column=headers.index("target") + 1).value = "Bonjour"
            package.save(package_path)
            package.close()

            prompts = []
            answers = iter(["3", str(package_path)])

            def answer(prompt=""):
                prompts.append(prompt)
                return next(answers)

            with patch("builtins.input", side_effect=answer), redirect_stdout(StringIO()):
                self.assertEqual(main([]), 0)

            self.assertEqual(len(prompts), 2)
            self.assertIn("Translated to_translate workbook", prompts[1])
            self.assertTrue(
                (package_path.parent / "source_filled_result.xlsx").exists()
            )

    def test_package_cli_prepare_and_fill_use_one_translator_workbook(self):
        from phraseloom.cli import main

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Hello", None])
            wb.save(input_path)
            wb.close()

            self.assertEqual(main(["prepare", str(input_path)]), 0)
            package_path = input_path.parent / "source_l10n" / "source_translator_todo.xlsx"
            package = load_workbook(package_path)
            todo = package["to_translate"]
            headers = [cell.value for cell in todo[1]]
            todo.cell(row=2, column=headers.index("target") + 1).value = "Bonjour"
            package.save(package_path)
            package.close()

            self.assertEqual(main(["fill", str(package_path)]), 0)
            result_path = package_path.parent / "source_filled_result.xlsx"
            result = load_workbook(result_path, data_only=True)
            try:
                self.assertEqual(result.active.cell(row=2, column=2).value, "Bonjour")
            finally:
                result.close()

    def test_translation_package_embeds_custom_tag_rules_for_fill(self):
        from phraseloom.workflow import (
            fill_translation_package,
            prepare_translation_package,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "source.xlsx"
            config_path = tmp_path / "tag_rules.toml"
            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["<foo>Power</foo>", None])
            wb.save(input_path)
            wb.close()
            config_path.write_text(
                "\n".join(
                    [
                        "version = 1",
                        "[angle_tags]",
                        'mode = "allowlist"',
                        'allowed = ["foo"]',
                        "[bbcode_tags]",
                        'mode = "allowlist"',
                        "allowed = []",
                        "[raw_braces]",
                        "protect_all = true",
                    ]
                ),
                encoding="utf-8",
            )

            prepare_stats = prepare_translation_package(
                input_path,
                tag_config=config_path,
            )
            package_path = Path(str(prepare_stats["to_translate_path"]))
            package = load_workbook(package_path)
            todo = package["to_translate"]
            headers = [cell.value for cell in todo[1]]
            todo.cell(row=2, column=headers.index("target") + 1).value = (
                "{1>Puissance<2}"
            )
            package.save(package_path)
            package.close()
            config_path.unlink()

            fill_stats = fill_translation_package(package_path)
            result = load_workbook(fill_stats["output_path"], data_only=True)
            try:
                self.assertEqual(
                    result.active.cell(row=2, column=2).value,
                    "<foo>Puissance</foo>",
                )
            finally:
                result.close()


class CompatibilityShimTests(unittest.TestCase):
    def test_template_demo_shim_exports_existing_workflow_api(self):
        from template_demo import generate_workbook, main, parse_template

        self.assertTrue(callable(generate_workbook))
        self.assertTrue(callable(main))
        self.assertEqual(parse_template("VIP10 Pack").template, "VIP{num1} Pack")

    def test_entity_cluster_probe_shim_exports_existing_api(self):
        from entity_cluster_probe import find_entity_clusters

        self.assertTrue(callable(find_entity_clusters))


if __name__ == "__main__":
    unittest.main()
