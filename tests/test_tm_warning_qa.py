import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from phraseloom import workbook_schema as schema


class TmWarningQaTests(unittest.TestCase):
    def test_export_uses_raw_workbook_text_and_serialized_tm_map_units(self):
        from scripts.export_tm_warning_qa import export_warning_qa

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            raw_tm = tmp_path / "TM.xlsx"
            reusable_units = tmp_path / "TM_reusable_units.xlsx"
            output = tmp_path / "TM_warning_qa.xlsx"
            _write_raw_tm(raw_tm)
            _write_reusable_units(reusable_units)

            stats = export_warning_qa(raw_tm, reusable_units, output)

            self.assertEqual(
                stats,
                {
                    "warning_pair_count": 1,
                    "warning_raw_row_count": 1,
                    "exported_record_count": 1,
                },
            )

            workbook = load_workbook(output, data_only=True)
            try:
                worksheet = workbook["warning_qa"]
                headers = [cell.value for cell in worksheet[1]]
                row = dict(zip(headers, next(worksheet.iter_rows(min_row=2, values_only=True))))
            finally:
                workbook.close()

        self.assertEqual(row["tm_id"], "TM00001")
        self.assertEqual(row["row_number"], 2)
        self.assertEqual(row["warning"], "open tag has no close partner: {1>")
        self.assertEqual(row["raw_source"], "<color=#fff>Raw source")
        self.assertEqual(row["raw_target"], "<color=#fff>Raw target")
        self.assertEqual(row["row_source_unit"], "{1>Raw source")
        self.assertEqual(row["row_target_unit"], "{1>Raw target")
        self.assertEqual(row["pair_source_unit"], "{1>Raw source")
        self.assertEqual(row["pair_target_unit"], "{1>Raw target")
        self.assertEqual(row["variables"], "{}")


def _write_raw_tm(path: Path) -> None:
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.append(["key", "source", "target"])
        worksheet.append(["k1", "<color=#fff>Raw source", "<color=#fff>Raw target"])
        workbook.save(path)
    finally:
        workbook.close()


def _write_reusable_units(path: Path) -> None:
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = schema.SUMMARY_SHEET
        worksheet.append(["check", "count"])

        pairs = workbook.create_sheet(schema.TM_PAIRS_SHEET)
        pairs.append(schema.TM_PAIR_COLUMNS)
        pairs.append(
            [
                "TM00001",
                "segment",
                "{1>Raw source",
                "{1>Raw target",
                1,
                1,
                "{}",
                "{1>Raw source",
                "{1>Raw target",
                "2",
                "open tag has no close partner: {1>",
            ]
        )

        tm_map = workbook.create_sheet(schema.TM_SOURCE_MAP_SHEET)
        tm_map.append(schema.TM_SOURCE_MAP_COLUMNS)
        tm_map.append(
            [
                2,
                "{1>Raw source",
                "{1>Raw target",
                "segment",
                "{1>Raw source",
                "{1>Raw target",
                "{}",
            ]
        )
        workbook.save(path)
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
