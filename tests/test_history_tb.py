from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from tools.history_tb import detect_history_tb_columns, iter_history_rows


class HistoryTbTests(unittest.TestCase):
    def create_nomark_history_workbook(self, path: Path) -> None:
        workbook = Workbook()
        raw_sheet = workbook.active
        raw_sheet.title = "Raw"
        raw_sheet["A1"] = "source"
        raw_sheet["B1"] = "target"

        term_sheet = workbook.create_sheet("术语表")
        term_sheet["A1"] = "source术语"
        term_sheet["B1"] = "target术语"
        term_sheet["C1"] = "source术语（无mark）"
        term_sheet["D1"] = "target术语（无mark）"
        term_sheet["C2"] = "花艺"
        term_sheet["D2"] = "Art Floral"
        workbook.active = 0
        workbook.save(path)

    def create_offset_history_workbook(self, path: Path) -> None:
        workbook = Workbook()
        term_sheet = workbook.active
        term_sheet.title = "术语表"
        term_sheet["A1"] = "metadata"
        term_sheet["B1"] = "notes"
        term_sheet["C2"] = "source术语（无mark）"
        term_sheet["D2"] = "target术语（无mark）"
        term_sheet["C3"] = "花艺"
        term_sheet["D3"] = "Art Floral"
        workbook.save(path)

    def create_blank_history_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "术语表"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["A2"] = "花艺"
        worksheet["B2"] = "Art Floral"
        worksheet["A3"] = "空译文"
        worksheet["B3"] = ""
        worksheet["A4"] = ""
        worksheet["B4"] = "Target Only"
        worksheet["A5"] = "宝库"
        worksheet["B5"] = "Vault"
        worksheet["Z5000"] = "unrelated tail"
        worksheet["A5001"].number_format = "@"
        workbook.save(path)

    def test_detects_nomark_columns_before_marked_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "history.xlsx"
            self.create_nomark_history_workbook(path)

            columns = detect_history_tb_columns(path)

            self.assertEqual(columns.sheet_title, "术语表")
            self.assertEqual(columns.source_column, "C")
            self.assertEqual(columns.target_column, "D")

    def test_detects_header_row_from_start_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "history.xlsx"
            self.create_offset_history_workbook(path)

            columns = detect_history_tb_columns(path, start_row=3)

            self.assertEqual(columns.sheet_title, "术语表")
            self.assertEqual(columns.source_column, "C")
            self.assertEqual(columns.target_column, "D")

    def test_rejects_same_column_when_source_and_target_arguments_match(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "history.xlsx"
            self.create_nomark_history_workbook(path)

            with self.assertRaisesRegex(ValueError, "不能相同"):
                detect_history_tb_columns(path, source_column="A", target_column="A")

    def test_iter_history_rows_skips_blank_source_or_target(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "history.xlsx"
            self.create_blank_history_workbook(path)

            sheet_title, source_column, target_column, rows = iter_history_rows(path)

            self.assertEqual(sheet_title, "术语表")
            self.assertEqual(source_column, "A")
            self.assertEqual(target_column, "B")
            self.assertEqual([(row.row_index, row.source_text, row.target_text) for row in rows], [
                (2, "花艺", "Art Floral"),
                (5, "宝库", "Vault"),
            ])

    def test_iter_history_rows_stops_after_default_empty_tail_threshold(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "history.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Early"
            worksheet["B2"] = "Précoce"
            worksheet["A1005"] = "Late"
            worksheet["B1005"] = "Tardif"
            workbook.save(path)

            *_metadata, default_rows = iter_history_rows(path)
            *_metadata, unlimited_rows = iter_history_rows(
                path,
                empty_row_stop_threshold=None,
            )

            self.assertEqual([row.source_text for row in default_rows], ["Early"])
            self.assertEqual(
                [row.source_text for row in unlimited_rows],
                ["Early", "Late"],
            )


if __name__ == "__main__":
    unittest.main()
