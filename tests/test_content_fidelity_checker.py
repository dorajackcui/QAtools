from __future__ import annotations

import tempfile
import unittest
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.content_fidelity_checker.check_content_fidelity import (
    NUMBER_PROBLEM_SHEET_NAME,
    NUMBER_RULE,
    URL_PROBLEM_SHEET_NAME,
    extract_numbers,
    extract_urls,
    process_excel,
)


class ContentFidelityExtractionTests(unittest.TestCase):
    def test_extract_numbers_normalizes_width_and_masks_urls_and_tokens(self) -> None:
        text = (
            "Build v１．２．３ costs −５０％; {123} <size=20> "
            r"https://x.test/99 &#x20; \u00A0"
        )

        self.assertEqual(extract_numbers(text), ("1.2.3", "-50%"))

    def test_extract_urls_trims_sentence_punctuation_and_unmatched_brackets(self) -> None:
        text = "See https://example.com/a(1). and “www.test.com/path”。"

        self.assertEqual(
            extract_urls(text),
            ("https://example.com/a(1)", "www.test.com/path"),
        )


class ContentFidelityExcelTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["source", "target"])
        worksheet.append(
            [
                "Download 10 files at https://example.com/a?x=1.",
                "下载 11 个文件：https://example.com/a?x=1。",
            ]
        )
        worksheet.append(
            [
                "See https://a.example/x(1).",
                "查看 https://b.example/x(1)。",
            ]
        )
        worksheet.append(["Version 1.2.3: 50%", "50％：版本 １．２．３"])
        worksheet.append(["Use {1} and <size=20>", "使用 {2} 与 <size=30>"])
        worksheet.append(["Open www.example.com/path).", "打开 www.example.com/path"])
        worksheet.append(["Repeat 10 and 10", "重复 10"])
        workbook.save(path)
        workbook.close()

    def test_process_excel_reports_number_and_url_differences_separately(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(input_path, "A", "B", sheet="Data")

            self.assertEqual(summary.number_problem_rows, 2)
            self.assertEqual(summary.url_problem_rows, 1)
            workbook = load_workbook(summary.output_path)
            try:
                number_sheet = workbook[NUMBER_PROBLEM_SHEET_NAME]
                url_sheet = workbook[URL_PROBLEM_SHEET_NAME]
                self.assertEqual(
                    [number_sheet[f"A{row}"].value for row in range(2, 4)],
                    [2, 7],
                )
                self.assertEqual(url_sheet["A2"].value, 3)
                self.assertIn("10", number_sheet["G2"].value)
                self.assertIn("11", number_sheet["H2"].value)
                self.assertEqual(url_sheet["G2"].value, "https://a.example/x(1)")
                self.assertEqual(url_sheet["H2"].value, "https://b.example/x(1)")
                self.assertEqual(number_sheet["A2"].hyperlink.location, "'Data'!B2")
            finally:
                workbook.close()

    def test_process_excel_can_run_only_number_rule(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(input_path)

            summary = process_excel(
                input_path,
                "A",
                "B",
                sheet="Data",
                rules=(NUMBER_RULE,),
            )

            self.assertEqual(summary.selected_rules, (NUMBER_RULE,))
            workbook = load_workbook(summary.output_path)
            try:
                self.assertIn(NUMBER_PROBLEM_SHEET_NAME, workbook.sheetnames)
                self.assertNotIn(URL_PROBLEM_SHEET_NAME, workbook.sheetnames)
            finally:
                workbook.close()

    def test_number_comparison_uses_a_multiset_not_source_order(self) -> None:
        self.assertEqual(
            Counter(extract_numbers("10, then 20, then 10")),
            Counter(extract_numbers("10 / 10 / 20")),
        )


if __name__ == "__main__":
    unittest.main()
