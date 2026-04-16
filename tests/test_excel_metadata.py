from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from tools.excel_metadata import detect_source_target_columns, list_workbook_sheets


class WorkbookMetadataTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = "Data"
        first_sheet["A1"] = " Source "
        first_sheet["C1"] = " TARGET"

        second_sheet = workbook.create_sheet("Glossary")
        second_sheet["B1"] = "source"
        second_sheet["D1"] = "target"

        duplicate_sheet = workbook.create_sheet("Duplicate")
        duplicate_sheet["A1"] = "source"
        duplicate_sheet["B1"] = "source"
        duplicate_sheet["C1"] = "target"

        source_only_sheet = workbook.create_sheet("SourceOnly")
        source_only_sheet["F1"] = "source"

        workbook.active = 1
        workbook.save(path)

    def test_list_workbook_sheets_returns_sheet_names_and_active_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(workbook_path)

            sheet_choices = list_workbook_sheets(workbook_path)

            self.assertEqual(sheet_choices.sheet_names, ("Data", "Glossary", "Duplicate", "SourceOnly"))
            self.assertEqual(sheet_choices.default_sheet, "Glossary")

    def test_detect_source_target_columns_matches_trimmed_case_insensitive_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(workbook_path)

            detected_columns = detect_source_target_columns(workbook_path, sheet="Data")

            self.assertEqual(detected_columns.detected_source_column, "A")
            self.assertEqual(detected_columns.detected_target_column, "C")

    def test_detect_source_target_columns_returns_none_for_duplicate_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(workbook_path)

            detected_columns = detect_source_target_columns(workbook_path, sheet="Duplicate")

            self.assertIsNone(detected_columns.detected_source_column)
            self.assertEqual(detected_columns.detected_target_column, "C")

    def test_detect_source_target_columns_can_return_single_detected_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(workbook_path)

            detected_columns = detect_source_target_columns(workbook_path, sheet="SourceOnly")

            self.assertEqual(detected_columns.detected_source_column, "F")
            self.assertIsNone(detected_columns.detected_target_column)

    def test_list_workbook_sheets_rejects_missing_file(self) -> None:
        with self.assertRaisesRegex(FileNotFoundError, "输入文件不存在"):
            list_workbook_sheets("missing.xlsx")

    def test_detect_source_target_columns_rejects_missing_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "input.xlsx"
            self.create_workbook(workbook_path)

            with self.assertRaisesRegex(ValueError, "工作表不存在"):
                detect_source_target_columns(workbook_path, sheet="Missing")

    def test_list_workbook_sheets_rejects_invalid_excel_file(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "input.txt"
            workbook_path.write_text("not an excel file", encoding="utf-8")

            with self.assertRaises(InvalidFileException):
                list_workbook_sheets(workbook_path)


if __name__ == "__main__":
    unittest.main()
