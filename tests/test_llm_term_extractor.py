from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from argparse import Namespace
from pathlib import Path
from subprocess import CompletedProcess
from unittest.mock import patch

from tools.llm_term_extractor.codex_term_review import (
    DEFAULT_CODEX_MODEL,
    DEFAULT_REASONING_EFFORT,
    ConflictGroup,
    InputBatchRow,
    build_codex_command,
    load_prompt_template,
    parse_conflict_response,
    parse_extraction_response,
    render_conflict_prompt,
    render_extraction_prompt,
    run_codex_prompt,
)


class CodexTermReviewTests(unittest.TestCase):
    def test_render_extraction_prompt_includes_mode_rows_and_schema(self) -> None:
        template = load_prompt_template(
            Path("tools/llm_term_extractor/prompts/extract_terms_zh_target.md")
        )
        rows = [
            InputBatchRow(
                row_id="row-1",
                source_text="Unlock the Abyssal Vault in Event Clash.",
                target_text="在活动冲突中解锁深渊宝库。",
            )
        ]

        prompt = render_extraction_prompt(template, mode="source_target", rows=rows)

        self.assertIn("source_target", prompt)
        self.assertIn('"row_id": "row-1"', prompt)
        self.assertIn("Abyssal Vault", prompt)
        self.assertIn("source decides whether something is a term", prompt)
        self.assertIn("fixed names/events/gameplay/items/currencies/systems/titles/fixed phrases/cross-context anchors", prompt)
        self.assertIn("do not collect ordinary UI state/action/adjective/full sentence text", prompt)
        self.assertIn("extract existing target expression only; do not recommend or rewrite", prompt)
        self.assertIn('"rows"', prompt)
        self.assertIn('"terms"', prompt)
        self.assertNotIn("{{MODE}}", prompt)
        self.assertNotIn("{{ROWS_JSON}}", prompt)
        self.assertNotIn("{{OUTPUT_SCHEMA}}", prompt)

    def test_parse_extraction_response_reads_plain_and_fenced_json(self) -> None:
        plain = (
            '{"rows":[{"row_id":"row-1","terms":[{"source_term":"Abyssal Vault",'
            '"target_term":"深渊宝库","category":"item","note":"第一行\n第二行"}]}]}'
        )
        fenced = (
            "```json\n"
            '{"rows":[{"row_id":"row-2","terms":[{"source_term":"Event Clash",'
            '"target_term":"活动冲突","category":"event","note":"fixed event"}]}]}\n'
            "```"
        )

        plain_rows = parse_extraction_response(plain)
        fenced_rows = parse_extraction_response(fenced)

        self.assertEqual(plain_rows[0].row_id, "row-1")
        self.assertEqual(plain_rows[0].terms[0].source_term, "Abyssal Vault")
        self.assertEqual(plain_rows[0].terms[0].target_term, "深渊宝库")
        self.assertEqual(plain_rows[0].terms[0].note, "第一行\n第二行")
        self.assertEqual(fenced_rows[0].row_id, "row-2")
        self.assertEqual(fenced_rows[0].terms[0].source_term, "Event Clash")

    def test_parse_conflict_response_reads_decisions(self) -> None:
        template = load_prompt_template(
            Path("tools/llm_term_extractor/prompts/conflict_review_zh_target.md")
        )
        prompt = render_conflict_prompt(
            template,
            groups=[
                ConflictGroup(
                    group_id="group-1",
                    source_term="Abyssal Vault",
                    target_terms=("深渊宝库", "深渊金库"),
                    examples=("Unlock the Abyssal Vault.",),
                )
            ],
        )
        response = (
            '{"decisions":[{"group_id":"group-1","decision":"conflict",'
            '"canonical_target":"深渊宝库","reason":"official-looking variant"}]}'
        )

        decisions = parse_conflict_response(response)

        self.assertIn("ignore case, singular/plural, ordinary grammar, punctuation/mark-only differences", prompt)
        self.assertIn(
            "flag substantial target wording, official-looking variant, conceptual difference, or project-context decisions",
            prompt,
        )
        self.assertIn("return strict JSON only", prompt)
        self.assertIn('"groups"', prompt)
        self.assertIn('"decisions"', prompt)
        self.assertNotIn("{{GROUPS_JSON}}", prompt)
        self.assertNotIn("{{OUTPUT_SCHEMA}}", prompt)
        self.assertEqual(decisions[0].group_id, "group-1")
        self.assertEqual(decisions[0].decision, "conflict")
        self.assertEqual(decisions[0].canonical_target, "深渊宝库")
        self.assertEqual(decisions[0].reason, "official-looking variant")

    def test_build_codex_command_uses_spark_model_and_reasoning_effort(self) -> None:
        command = build_codex_command(
            Path("/tmp/codex-output.txt"),
            DEFAULT_CODEX_MODEL,
            DEFAULT_REASONING_EFFORT,
        )

        self.assertEqual(command[0], "codex")
        self.assertIn("--ask-for-approval", command)
        self.assertIn("never", command)
        self.assertIn('model_reasoning_effort="high"', command)
        self.assertIn("--ephemeral", command)
        self.assertIn("--ignore-rules", command)
        self.assertIn("--skip-git-repo-check", command)
        self.assertIn("read-only", command)
        self.assertEqual(
            command[command.index("--output-last-message") + 1],
            str(Path("/tmp/codex-output.txt")),
        )
        self.assertEqual(command[command.index("--model") + 1], "gpt-5.3-codex-spark")
        self.assertEqual(command[-1], "-")

    def test_run_codex_prompt_writes_prompt_to_stdin_and_reads_output_file(self) -> None:
        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-test-") as tmp_dir:
            output_path = Path(tmp_dir) / "codex-output.txt"

            def fake_run(command, input, **kwargs):
                self.assertEqual(input, "prompt text")
                self.assertEqual(command[command.index("--output-last-message") + 1], str(output_path))
                output_path.write_text('{"rows":[]}', encoding="utf-8")
                return CompletedProcess(command, 0, stdout="stdout fallback", stderr="")

            with patch("tools.codex_runner.subprocess.run", side_effect=fake_run):
                output = run_codex_prompt(
                    "prompt text",
                    output_path,
                    DEFAULT_CODEX_MODEL,
                    DEFAULT_REASONING_EFFORT,
                    timeout_seconds=12,
                )

        self.assertEqual(output, '{"rows":[]}')


class LlmTermCodexOrchestrationTests(unittest.TestCase):
    def test_default_output_path_uses_llm_terms_prefix(self) -> None:
        from tools.llm_term_extractor.extract_llm_terms import build_default_output_path

        output_path = build_default_output_path("/tmp/source-workbook.xlsx")

        self.assertEqual(output_path, Path("/tmp/llm_terms_source-workbook.xlsx"))

    def test_default_batch_extractor_renders_prompt_dumps_raw_output_and_retries_invalid_json(self) -> None:
        from tools.llm_term_extractor.extract_llm_terms import build_codex_batch_extractor

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-codex-") as tmp_dir:
            tmp_path = Path(tmp_dir)
            output_path = tmp_path / "input_llm_terms.xlsx"
            dump_dir = tmp_path / "prompts"
            responses = iter(
                [
                    "not valid json",
                    (
                        '{"rows":[{"row_id":"2","terms":[{"source_term":"Abyssal Vault",'
                        '"target_term":"深渊宝库","category":"item","note":"fixed name"}]}]}'
                    ),
                ]
            )
            prompts: list[str] = []

            def fake_run_codex_prompt(prompt, output_path_arg, model, reasoning_effort, timeout_seconds):
                prompts.append(prompt)
                self.assertEqual(model, DEFAULT_CODEX_MODEL)
                self.assertEqual(reasoning_effort, DEFAULT_REASONING_EFFORT)
                self.assertTrue(str(output_path_arg).endswith(".txt"))
                return next(responses)

            with patch(
                "tools.llm_term_extractor.extract_llm_terms.run_codex_prompt",
                side_effect=fake_run_codex_prompt,
            ):
                extractor = build_codex_batch_extractor(
                    output_path=output_path,
                    dump_prompts_dir=dump_dir,
                    keep_raw_codex_output=True,
                )
                rows = list(
                    extractor(
                        [
                            InputBatchRow(
                                row_id="2",
                                source_text="Unlock the Abyssal Vault.",
                                target_text="解锁深渊宝库。",
                            )
                        ]
                    )
                )

            self.assertEqual(len(prompts), 2)
            self.assertIn("source_target", prompts[0])
            self.assertIn('"row_id": "2"', prompts[0])
            self.assertIn("return strict JSON", prompts[1])
            self.assertEqual(rows[0].row_id, "2")
            self.assertEqual(rows[0].terms[0].source_term, "Abyssal Vault")

            dumped_prompt = dump_dir / "extract-batch-0001.md"
            self.assertTrue(dumped_prompt.exists())
            self.assertIn("Abyssal Vault", dumped_prompt.read_text(encoding="utf-8"))

            raw_output_path = tmp_path / "input_llm_terms_codex_raw.jsonl"
            raw_lines = [
                json.loads(line)
                for line in raw_output_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertEqual([line["kind"] for line in raw_lines], ["extract", "extract"])
            self.assertEqual([line["attempt"] for line in raw_lines], [1, 2])
            self.assertEqual(raw_lines[0]["raw_output"], "not valid json")
            self.assertIn('"rows"', raw_lines[1]["raw_output"])

    def test_conflict_reviewer_renders_prompt_and_retries_invalid_json(self) -> None:
        from tools.llm_term_extractor.extract_llm_terms import build_codex_conflict_reviewer

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-conflict-codex-") as tmp_dir:
            tmp_path = Path(tmp_dir)
            output_path = tmp_path / "input_llm_terms.xlsx"
            dump_dir = tmp_path / "prompts"
            responses = iter(
                [
                    "not valid json",
                    (
                        '{"decisions":[{"group_id":"abyssal vault","decision":"same",'
                        '"canonical_target":"深渊宝库","reason":"same term"}]}'
                    ),
                ]
            )
            prompts: list[str] = []

            def fake_run_codex_prompt(prompt, output_path_arg, model, reasoning_effort, timeout_seconds):
                prompts.append(prompt)
                self.assertEqual(model, DEFAULT_CODEX_MODEL)
                self.assertEqual(reasoning_effort, DEFAULT_REASONING_EFFORT)
                self.assertTrue(str(output_path_arg).endswith(".txt"))
                return next(responses)

            with patch(
                "tools.llm_term_extractor.extract_llm_terms.run_codex_prompt",
                side_effect=fake_run_codex_prompt,
            ):
                reviewer = build_codex_conflict_reviewer(
                    output_path=output_path,
                    dump_prompts_dir=dump_dir,
                    keep_raw_codex_output=True,
                )
                decisions = reviewer(
                    [
                        ConflictGroup(
                            group_id="abyssal vault",
                            source_term="Abyssal Vault",
                            target_terms=("深渊宝库", "深渊金库"),
                            examples=("Unlock the Abyssal Vault.",),
                        )
                    ]
                )

            self.assertEqual(len(prompts), 2)
            self.assertIn('"group_id": "abyssal vault"', prompts[0])
            self.assertIn("return strict JSON", prompts[1])
            self.assertEqual(decisions[0].group_id, "abyssal vault")
            self.assertEqual(decisions[0].decision, "same")

            dumped_prompt = dump_dir / "conflict-review.md"
            self.assertTrue(dumped_prompt.exists())
            self.assertIn("Abyssal Vault", dumped_prompt.read_text(encoding="utf-8"))

            raw_output_path = tmp_path / "input_llm_terms_codex_raw.jsonl"
            raw_lines = [
                json.loads(line)
                for line in raw_output_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertEqual([line["kind"] for line in raw_lines], ["conflict", "conflict"])
            self.assertEqual([line["attempt"] for line in raw_lines], [1, 2])
            self.assertEqual(raw_lines[0]["raw_output"], "not valid json")
            self.assertIn('"decisions"', raw_lines[1]["raw_output"])

    def test_codex_extractors_do_not_leave_last_message_artifacts_when_raw_output_disabled(self) -> None:
        from tools.llm_term_extractor.extract_llm_terms import (
            build_codex_batch_extractor,
            build_codex_conflict_reviewer,
        )

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-codex-clean-") as tmp_dir:
            tmp_path = Path(tmp_dir)
            output_path = tmp_path / "input_llm_terms.xlsx"
            dump_dir = tmp_path / "prompts"

            def fake_run_codex_prompt(prompt, output_path_arg, model, reasoning_effort, timeout_seconds):
                output_path_for_message = Path(output_path_arg)
                if '"groups"' in prompt:
                    raw_output = '{"decisions":[{"group_id":"apple","decision":"same"}]}'
                else:
                    raw_output = (
                        '{"rows":[{"row_id":"2","terms":[{"source_term":"apple",'
                        '"target_term":"苹果"}]}]}'
                    )
                output_path_for_message.write_text(raw_output, encoding="utf-8")
                return raw_output

            with patch(
                "tools.llm_term_extractor.extract_llm_terms.run_codex_prompt",
                side_effect=fake_run_codex_prompt,
            ):
                extractor = build_codex_batch_extractor(
                    output_path=output_path,
                    dump_prompts_dir=dump_dir,
                )
                list(
                    extractor(
                        [
                            InputBatchRow(
                                row_id="2",
                                source_text="Collect apple.",
                                target_text="收集苹果。",
                            )
                        ]
                    )
                )

                reviewer = build_codex_conflict_reviewer(
                    output_path=output_path,
                    dump_prompts_dir=dump_dir,
                )
                reviewer(
                    [
                        ConflictGroup(
                            group_id="apple",
                            source_term="apple",
                            target_terms=("苹果", "历史苹果"),
                        )
                    ]
                )

            self.assertTrue((dump_dir / "extract-batch-0001.md").exists())
            self.assertTrue((dump_dir / "conflict-review.md").exists())
            self.assertFalse((tmp_path / "input_llm_terms_codex_raw.jsonl").exists())
            leftover_text_files = [
                path.name
                for path in tmp_path.iterdir()
                if path.is_file() and path.suffix == ".txt"
            ]
            self.assertEqual(leftover_text_files, [])

    def test_prompt_if_missing_requires_input_and_source_in_noninteractive_mode(self) -> None:
        from tools.llm_term_extractor.extract_llm_terms import prompt_if_missing

        with patch("sys.stdin.isatty", return_value=False):
            with self.assertRaises(SystemExit) as missing_input:
                prompt_if_missing(Namespace(input_file=None, source_column="A"))
            self.assertIn("input_file", str(missing_input.exception))

            with self.assertRaises(SystemExit) as missing_source:
                prompt_if_missing(Namespace(input_file="/tmp/input.xlsx", source_column=None))
            self.assertIn("source_column", str(missing_source.exception))

    def test_cli_help_runs_when_invoked_as_script(self) -> None:
        completed = subprocess.run(
            [sys.executable, "tools/llm_term_extractor/extract_llm_terms.py", "--help"],
            capture_output=True,
            text=True,
            check=False,
        )

        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertIn("--history-tb", completed.stdout)


if __name__ == "__main__":
    unittest.main()


class LlmTermWorkbookTests(unittest.TestCase):
    def test_process_excel_handles_source_only_and_source_target_rows(self) -> None:
        from openpyxl import Workbook, load_workbook

        from tools.llm_term_extractor.codex_term_review import ExtractedLlmTerm, RowExtraction
        from tools.llm_term_extractor.extract_llm_terms import process_excel

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-workbook-") as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Unlock the Abyssal Vault."
            worksheet["B2"] = "解锁深渊宝库。"
            worksheet["A3"] = ""
            worksheet["B3"] = ""
            worksheet["A4"] = "Claim the Heart Flower Gift Box."
            worksheet["B4"] = ""
            workbook.save(input_path)

            def fake_extractor(rows):
                self.assertEqual([row.row_id for row in rows], ["2", "4"])
                return [
                    RowExtraction(
                        row_id="2",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="Abyssal Vault",
                                target_term="深渊宝库",
                                category="item",
                                note="fixed item name",
                            ),
                        ),
                    ),
                    RowExtraction(
                        row_id="4",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="Heart Flower Gift Box",
                                target_term="",
                                category="item",
                                note="source-only item name",
                            ),
                        ),
                    ),
                    RowExtraction(
                        row_id="99",
                        terms=(ExtractedLlmTerm(source_term="Ignored Term", target_term="忽略"),),
                    ),
                ]

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                batch_extractor=fake_extractor,
            )

            self.assertEqual(summary.output_path, input_path.with_name("llm_terms_input.xlsx"))
            self.assertEqual(summary.worksheet_title, "Data")
            self.assertEqual(summary.scanned_row_count, 2)
            self.assertEqual(summary.batch_count, 1)
            self.assertEqual(summary.term_count, 2)
            self.assertEqual(summary.evidence_count, 2)
            self.assertEqual(summary.import_candidate_count, 1)
            self.assertEqual(summary.review_before_import_count, 1)
            self.assertEqual(summary.already_in_history_count, 0)

            result = load_workbook(summary.output_path)
            self.assertEqual(
                result.sheetnames[-7:],
                [
                    "Terms_Source_Dedup",
                    "Extraction_Evidence",
                    "Conflicts_To_Review",
                    "Import_Candidate",
                    "Review_Before_Import",
                    "Already_In_History",
                    "Summary",
                ],
            )

            terms = result["Terms_Source_Dedup"]
            self.assertEqual(terms["A2"].value, "Abyssal Vault")
            self.assertEqual(terms["B2"].value, "深渊宝库")
            self.assertEqual(terms["A3"].value, "Heart Flower Gift Box")

            import_candidates = result["Import_Candidate"]
            self.assertEqual(import_candidates["A2"].value, "Abyssal Vault")
            self.assertEqual(import_candidates["B2"].value, "深渊宝库")

            review = result["Review_Before_Import"]
            self.assertEqual(review["A2"].value, "Heart Flower Gift Box")
            self.assertEqual(review["D2"].value, "target缺失")

    def test_process_excel_routes_history_matches(self) -> None:
        from openpyxl import Workbook, load_workbook

        from tools.llm_term_extractor.codex_term_review import ExtractedLlmTerm, RowExtraction
        from tools.llm_term_extractor.extract_llm_terms import process_excel

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-history-") as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_path = tmp_path / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Unlock the Abyssal Vault."
            worksheet["B2"] = "解锁深渊宝库。"
            workbook.save(input_path)

            history_path = tmp_path / "history.xlsx"
            history = Workbook()
            history_sheet = history.active
            history_sheet.title = "术语表"
            history_sheet["A1"] = "source术语（无mark）"
            history_sheet["B1"] = "target术语(无mark)"
            history_sheet["A2"] = " abyssal   vault "
            history_sheet["B2"] = "深渊宝库"
            history.save(history_path)

            def fake_extractor(rows):
                return [
                    RowExtraction(
                        row_id="2",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="Abyssal Vault",
                                target_term="深渊宝库",
                                category="item",
                                note="already known term",
                            ),
                        ),
                    )
                ]

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                batch_extractor=fake_extractor,
                history_tb_file=history_path,
            )

            self.assertEqual(summary.term_count, 1)
            self.assertEqual(summary.import_candidate_count, 0)
            self.assertEqual(summary.review_before_import_count, 0)
            self.assertEqual(summary.already_in_history_count, 1)

            result = load_workbook(summary.output_path)
            history_output = result["Already_In_History"]
            self.assertEqual(history_output["A2"].value, "Abyssal Vault")
            self.assertEqual(history_output["B2"].value, "深渊宝库")

            import_candidates = result["Import_Candidate"]
            self.assertIsNone(import_candidates["A2"].value)

    def test_load_history_tb_mapping_detects_toolshub_nomark_headers(self) -> None:
        from openpyxl import Workbook

        from tools.llm_term_extractor.extract_llm_terms import load_history_tb_mapping

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-history-headers-") as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "source(无mark)"
            worksheet["B1"] = "target（无mark）"
            worksheet["C1"] = "术语来源"
            worksheet["A2"] = "  Abyssal   Vault "
            worksheet["B2"] = "深渊宝库"
            worksheet["C2"] = "history"
            workbook.save(history_path)

            mapping = load_history_tb_mapping(history_path)

        self.assertEqual(mapping, {"abyssal vault": "深渊宝库"})

    def test_load_history_tb_mapping_prefers_toolshub_nomark_columns(self) -> None:
        from openpyxl import Workbook

        from tools.llm_term_extractor.extract_llm_terms import (
            load_history_tb_mapping,
            normalize_term_key,
        )

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-toolshub-history-") as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "source术语"
            worksheet["B1"] = "target术语"
            worksheet["C1"] = "source术语（无mark）"
            worksheet["D1"] = "target术语（无mark）"
            worksheet["E1"] = "术语来源"
            worksheet["A2"] = "<apple>"
            worksheet["B2"] = "<历史苹果>"
            worksheet["C2"] = "apple"
            worksheet["D2"] = "历史苹果"
            worksheet["E2"] = "历史TB"
            workbook.save(history_path)

            mapping = load_history_tb_mapping(history_path)

        self.assertEqual(mapping, {normalize_term_key("apple"): "历史苹果"})

    def test_blank_history_target_does_not_suppress_import_candidate(self) -> None:
        from openpyxl import Workbook, load_workbook

        from tools.llm_term_extractor.codex_term_review import ExtractedLlmTerm, RowExtraction
        from tools.llm_term_extractor.extract_llm_terms import load_history_tb_mapping, process_excel

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-blank-history-target-") as tmp_dir:
            tmp_path = Path(tmp_dir)
            history_path = tmp_path / "history.xlsx"
            history = Workbook()
            history_sheet = history.active
            history_sheet.title = "术语表"
            history_sheet["A1"] = "source"
            history_sheet["B1"] = "target"
            history_sheet["A2"] = "apple"
            history_sheet["B2"] = "   "
            history.save(history_path)

            self.assertEqual(load_history_tb_mapping(history_path), {})

            input_path = tmp_path / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Collect apple."
            worksheet["B2"] = "收集苹果。"
            workbook.save(input_path)

            def fake_extractor(rows):
                return [
                    RowExtraction(
                        row_id="2",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="apple",
                                target_term="苹果",
                                category="item",
                                note="blank history target should be ignored",
                            ),
                        ),
                    )
                ]

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                batch_extractor=fake_extractor,
                history_tb_file=history_path,
            )

            self.assertEqual(summary.already_in_history_count, 0)
            self.assertEqual(summary.import_candidate_count, 1)
            self.assertEqual(summary.review_before_import_count, 0)

            result = load_workbook(summary.output_path)
            history_output = result["Already_In_History"]
            self.assertIsNone(history_output["A2"].value)
            import_candidates = result["Import_Candidate"]
            self.assertEqual(import_candidates["A2"].value, "apple")
            self.assertEqual(import_candidates["B2"].value, "苹果")

    def test_load_history_tb_mapping_uses_other_fallback_column_when_source_detected(self) -> None:
        from openpyxl import Workbook

        from tools.llm_term_extractor.extract_llm_terms import load_history_tb_mapping

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-history-fallback-source-") as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "custom target header"
            worksheet["B1"] = "source"
            worksheet["A2"] = "深渊宝库"
            worksheet["B2"] = "Abyssal Vault"
            workbook.save(history_path)

            mapping = load_history_tb_mapping(history_path)

        self.assertEqual(mapping, {"abyssal vault": "深渊宝库"})

    def test_load_history_tb_mapping_uses_other_fallback_column_when_target_detected(self) -> None:
        from openpyxl import Workbook

        from tools.llm_term_extractor.extract_llm_terms import load_history_tb_mapping

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-history-fallback-target-") as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "target"
            worksheet["B1"] = "custom source header"
            worksheet["A2"] = "深渊宝库"
            worksheet["B2"] = "Abyssal Vault"
            workbook.save(history_path)

            mapping = load_history_tb_mapping(history_path)

        self.assertEqual(mapping, {"abyssal vault": "深渊宝库"})

    def test_load_history_tb_mapping_rejects_explicit_same_column(self) -> None:
        from openpyxl import Workbook

        from tools.llm_term_extractor.extract_llm_terms import load_history_tb_mapping

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-history-same-column-") as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "source"
            worksheet["A2"] = "Abyssal Vault"
            workbook.save(history_path)

            with self.assertRaisesRegex(ValueError, "不能相同|same"):
                load_history_tb_mapping(
                    history_path,
                    source_column="A",
                    target_column="A",
                )

    def test_load_history_tb_mapping_rejects_same_column_resolved_from_headers(self) -> None:
        from openpyxl import Workbook

        from tools.llm_term_extractor.extract_llm_terms import load_history_tb_mapping

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-history-same-header-") as tmp_dir:
            history_path = Path(tmp_dir) / "history.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "术语表"
            worksheet["A1"] = "shared term header"
            worksheet["A2"] = "Abyssal Vault"
            workbook.save(history_path)

            with self.assertRaisesRegex(ValueError, "不能相同|same"):
                load_history_tb_mapping(
                    history_path,
                    source_column="shared term header",
                    target_column="shared term header",
                )

    def test_process_excel_routes_toolshub_nomark_history_match(self) -> None:
        from openpyxl import Workbook, load_workbook

        from tools.llm_term_extractor.codex_term_review import ExtractedLlmTerm, RowExtraction
        from tools.llm_term_extractor.extract_llm_terms import process_excel

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-toolshub-route-") as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_path = tmp_path / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Collect apple."
            worksheet["B2"] = "收集苹果。"
            workbook.save(input_path)

            history_path = tmp_path / "history.xlsx"
            history = Workbook()
            history_sheet = history.active
            history_sheet.title = "术语表"
            history_sheet["A1"] = "source术语"
            history_sheet["B1"] = "target术语"
            history_sheet["C1"] = "source术语（无mark）"
            history_sheet["D1"] = "target术语（无mark）"
            history_sheet["E1"] = "术语来源"
            history_sheet["A2"] = "<apple>"
            history_sheet["B2"] = "<历史苹果>"
            history_sheet["C2"] = "apple"
            history_sheet["D2"] = "历史苹果"
            history_sheet["E2"] = "历史TB"
            history.save(history_path)

            def fake_extractor(rows):
                return [
                    RowExtraction(
                        row_id="2",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="apple",
                                target_term="苹果",
                                category="item",
                                note="unmarked extracted term",
                            ),
                        ),
                    )
                ]

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                batch_extractor=fake_extractor,
                history_tb_file=history_path,
            )

            self.assertEqual(summary.already_in_history_count, 1)
            self.assertEqual(summary.import_candidate_count, 0)

            result = load_workbook(summary.output_path)
            history_output = result["Already_In_History"]
            self.assertEqual(history_output["A2"].value, "apple")
            self.assertEqual(history_output["B2"].value, "历史苹果")
            import_candidates = result["Import_Candidate"]
            self.assertIsNone(import_candidates["A2"].value)

    def test_process_excel_routes_real_conflicts_to_review_sheets(self) -> None:
        from openpyxl import Workbook, load_workbook

        from tools.llm_term_extractor.codex_term_review import (
            ConflictDecision,
            ExtractedLlmTerm,
            RowExtraction,
        )
        from tools.llm_term_extractor.extract_llm_terms import process_excel

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-conflict-") as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Flower Art gameplay"
            worksheet["B2"] = "Art Floral"
            worksheet["A3"] = "Flower Art piece"
            worksheet["B3"] = "composition florale"
            workbook.save(input_path)

            def fake_extractor(rows):
                return [
                    RowExtraction(
                        row_id="2",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="Flower Art",
                                target_term="Art Floral",
                                category="system_or_concept",
                                note="fixed concept",
                            ),
                        ),
                    ),
                    RowExtraction(
                        row_id="3",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="  flower   art ",
                                target_term="composition florale",
                                category="system_or_concept",
                                note="same source with another target",
                            ),
                        ),
                    ),
                ]

            def fake_conflict_reviewer(groups):
                self.assertEqual(len(groups), 1)
                self.assertEqual(groups[0].source_term, "Flower Art")
                self.assertEqual(groups[0].target_terms, ("Art Floral", "composition florale"))
                return [
                    ConflictDecision(
                        group_id=groups[0].group_id,
                        decision="conflict",
                        reason="substantial target wording difference",
                    )
                ]

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                batch_extractor=fake_extractor,
                conflict_reviewer=fake_conflict_reviewer,
            )

            self.assertEqual(summary.term_count, 1)
            self.assertEqual(summary.evidence_count, 2)
            self.assertEqual(summary.conflict_count, 1)
            self.assertEqual(summary.import_candidate_count, 0)
            self.assertEqual(summary.review_before_import_count, 1)

            result = load_workbook(summary.output_path)
            terms = result["Terms_Source_Dedup"]
            self.assertEqual(terms["A2"].value, "Flower Art")
            self.assertIn("Art Floral", terms["B2"].value)
            self.assertIn("composition florale", terms["B2"].value)

            conflicts = result["Conflicts_To_Review"]
            self.assertEqual(conflicts["A2"].value, "Flower Art")
            self.assertIn("Art Floral", conflicts["B2"].value)
            self.assertIn("composition florale", conflicts["B2"].value)
            self.assertEqual(conflicts["D2"].value, "conflict")
            self.assertEqual(conflicts["F2"].value, "substantial target wording difference")

            review = result["Review_Before_Import"]
            self.assertEqual(review["A2"].value, "Flower Art")
            self.assertEqual(review["D2"].value, "conflict")

    def test_process_excel_routes_unreviewed_multi_target_terms_as_conflicts(self) -> None:
        from openpyxl import Workbook, load_workbook

        from tools.llm_term_extractor.codex_term_review import ExtractedLlmTerm, RowExtraction
        from tools.llm_term_extractor.extract_llm_terms import process_excel

        with tempfile.TemporaryDirectory(prefix="tag-exactor-llm-unreviewed-conflict-") as tmp_dir:
            input_path = Path(tmp_dir) / "input.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "Guild Trial stage"
            worksheet["B2"] = "Epreuve de guilde"
            worksheet["A3"] = "Guild Trial reward"
            worksheet["B3"] = "Defi de guilde"
            workbook.save(input_path)

            def fake_extractor(rows):
                return [
                    RowExtraction(
                        row_id="2",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="Guild Trial",
                                target_term="Epreuve de guilde",
                                category="event",
                                note="event name",
                            ),
                        ),
                    ),
                    RowExtraction(
                        row_id="3",
                        terms=(
                            ExtractedLlmTerm(
                                source_term="Guild Trial",
                                target_term="Defi de guilde",
                                category="event",
                                note="same event different target",
                            ),
                        ),
                    ),
                ]

            summary = process_excel(
                input_file=input_path,
                source_column="A",
                target_column="B",
                sheet="Data",
                start_row=2,
                batch_extractor=fake_extractor,
            )

            self.assertEqual(summary.term_count, 1)
            self.assertEqual(summary.conflict_count, 1)
            self.assertEqual(summary.import_candidate_count, 0)
            self.assertEqual(summary.review_before_import_count, 1)

            result = load_workbook(summary.output_path)
            conflicts = result["Conflicts_To_Review"]
            self.assertEqual(conflicts["A2"].value, "Guild Trial")
            self.assertIn("Epreuve de guilde", conflicts["B2"].value)
            self.assertIn("Defi de guilde", conflicts["B2"].value)
            self.assertEqual(conflicts["D2"].value, "review")
            self.assertEqual(conflicts["F2"].value, "多译法需确认")

            review = result["Review_Before_Import"]
            self.assertEqual(review["A2"].value, "Guild Trial")
            self.assertEqual(review["D2"].value, "多译法需确认")
