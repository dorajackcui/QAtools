from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.term_glossary_checker.check_terms_against_glossary import (
    find_row_terms,
    load_glossary_entries,
    process_excel,
    text_contains_term,
)


class GlossaryLoadingTests(unittest.TestCase):
    def test_conflicting_targets_are_reported_and_duplicates_are_deduped(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            glossary_path = Path(tmp_dir) / "glossary.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Glossary"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "API"
            worksheet["B2"] = "接口"
            worksheet["A3"] = "API"
            worksheet["B3"] = "接口"
            worksheet["A4"] = "Term"
            worksheet["B4"] = "译法一"
            worksheet["A5"] = "Term"
            worksheet["B5"] = "译法二"
            workbook.save(glossary_path)

            sheet_title, entries, conflicts = load_glossary_entries(
                glossary_file=glossary_path,
                source_column="A",
                target_column="B",
                start_row=2,
                case_sensitive=False,
            )

            self.assertEqual(sheet_title, "Glossary")
            self.assertEqual([(entry.source_term, entry.target_term) for entry in entries], [("API", "接口")])
            self.assertEqual(len(conflicts), 1)
            self.assertEqual(conflicts[0].source_term, "Term")
            self.assertEqual(conflicts[0].target_terms, ("译法一", "译法二"))


class MatchingTests(unittest.TestCase):
    def create_entries(self, rows: list[tuple[str, str]]) -> list:
        with tempfile.TemporaryDirectory() as tmp_dir:
            glossary_path = Path(tmp_dir) / "glossary.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            for row_index, (source_term, target_term) in enumerate(rows, start=2):
                worksheet[f"A{row_index}"] = source_term
                worksheet[f"B{row_index}"] = target_term
            workbook.save(glossary_path)

            _, entries, _ = load_glossary_entries(
                glossary_file=glossary_path,
                source_column="A",
                target_column="B",
                start_row=2,
                case_sensitive=False,
            )
            return entries

    def test_longest_overlapping_term_wins(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            glossary_path = Path(tmp_dir) / "glossary.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "API"
            worksheet["B2"] = "接口"
            worksheet["A3"] = "API key"
            worksheet["B3"] = "接口密钥"
            workbook.save(glossary_path)

            _, entries, _ = load_glossary_entries(
                glossary_file=glossary_path,
                source_column="A",
                target_column="B",
                start_row=2,
                case_sensitive=False,
            )

            class FallbackMatcher(list):
                pass

            matches = find_row_terms(
                "Use API key to sign in.",
                FallbackMatcher(entries),
                case_sensitive=False,
                match_mode="hybrid-boundary",
            )
            self.assertEqual([entry.source_term for entry in matches], ["API key"])

    def test_hybrid_boundary_does_not_match_rain_inside_training(self) -> None:
        class FallbackMatcher(list):
            pass

        entries = self.create_entries([("rain", "雨")])
        matches = find_row_terms(
            "training material",
            FallbackMatcher(entries),
            case_sensitive=False,
            match_mode="hybrid-boundary",
        )
        self.assertEqual(matches, [])

    def test_hybrid_boundary_does_not_match_acc_inside_account(self) -> None:
        class FallbackMatcher(list):
            pass

        entries = self.create_entries([("ACC", "ACC")])
        matches = find_row_terms(
            "account setup",
            FallbackMatcher(entries),
            case_sensitive=False,
            match_mode="hybrid-boundary",
        )
        self.assertEqual(matches, [])

    def test_hybrid_boundary_allows_api_in_api_dash_key(self) -> None:
        class FallbackMatcher(list):
            pass

        entries = self.create_entries([("API", "接口")])
        matches = find_row_terms(
            "Use API-key to sign in.",
            FallbackMatcher(entries),
            case_sensitive=False,
            match_mode="hybrid-boundary",
        )
        self.assertEqual([entry.source_term for entry in matches], ["API"])

    def test_hybrid_boundary_does_not_match_acc_inside_acc_001(self) -> None:
        class FallbackMatcher(list):
            pass

        entries = self.create_entries([("ACC", "ACC")])
        matches = find_row_terms(
            "ACC_001 pending",
            FallbackMatcher(entries),
            case_sensitive=False,
            match_mode="hybrid-boundary",
        )
        self.assertEqual(matches, [])

    def test_chinese_term_still_matches_inside_sentence(self) -> None:
        class FallbackMatcher(list):
            pass

        entries = self.create_entries([("接口", "API")])
        matches = find_row_terms(
            "请检查接口配置",
            FallbackMatcher(entries),
            case_sensitive=False,
            match_mode="hybrid-boundary",
        )
        self.assertEqual([entry.source_term for entry in matches], ["接口"])

    def test_substring_mode_keeps_old_behavior(self) -> None:
        class FallbackMatcher(list):
            pass

        entries = self.create_entries([("rain", "雨")])
        matches = find_row_terms(
            "training material",
            FallbackMatcher(entries),
            case_sensitive=False,
            match_mode="substring",
        )
        self.assertEqual([entry.source_term for entry in matches], ["rain"])

    def test_target_contains_term_uses_same_boundary_rules(self) -> None:
        self.assertFalse(text_contains_term("account setup", "acc", match_mode="hybrid-boundary"))
        self.assertTrue(text_contains_term("api-key", "api", match_mode="hybrid-boundary"))
        self.assertTrue(text_contains_term("account setup", "acc", match_mode="substring"))


class ProcessExcelTests(unittest.TestCase):
    def create_glossary_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Glossary"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["A2"] = "API"
        worksheet["B2"] = "接口"
        worksheet["A3"] = "API key"
        worksheet["B3"] = "接口密钥"
        worksheet["A4"] = "UI"
        worksheet["B4"] = "界面"
        worksheet["A5"] = "Button"
        worksheet["B5"] = "按钮"
        worksheet["A6"] = "Name"
        worksheet["B6"] = "名称"
        worksheet["A7"] = "Term"
        worksheet["B7"] = "译法一"
        worksheet["A8"] = "Term"
        worksheet["B8"] = "译法二"
        workbook.save(path)

    def create_data_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "source"
        worksheet["B1"] = "target"
        worksheet["A2"] = "Use API key to sign in."
        worksheet["B2"] = "使用接口登录。"
        worksheet["A3"] = "API API"
        worksheet["B3"] = "接口"
        worksheet["A4"] = "Open UI Button"
        worksheet["B4"] = "打开界面按钮"
        worksheet["A5"] = "No glossary match here"
        worksheet["B5"] = "这里没有术语"
        worksheet["A6"] = "name field"
        worksheet["B6"] = "名称字段"
        workbook.save(path)

    def test_process_excel_writes_result_sheets_and_keeps_source_sheet_unchanged(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            glossary_path = Path(tmp_dir) / "glossary.xlsx"
            data_path = Path(tmp_dir) / "data.xlsx"
            output_path = Path(tmp_dir) / "data_glossary_checked.xlsx"
            self.create_glossary_workbook(glossary_path)
            self.create_data_workbook(data_path)

            summary = process_excel(
                glossary_file=glossary_path,
                data_file=data_path,
                glossary_source_column="A",
                glossary_target_column="B",
                data_source_column="A",
                data_target_column="B",
                start_row=2,
                case_sensitive=False,
                match_mode="hybrid-boundary",
            )

            self.assertEqual(summary.glossary_sheet_title, "Glossary")
            self.assertEqual(summary.data_sheet_title, "Data")
            self.assertEqual(summary.glossary_term_count, 5)
            self.assertEqual(summary.conflict_count, 1)
            self.assertEqual(summary.total_rows_checked, 5)
            self.assertEqual(summary.matched_rows, 4)
            self.assertEqual(summary.problem_rows, 1)
            self.assertEqual(summary.problem_count, 1)
            self.assertEqual(summary.match_mode, "hybrid-boundary")
            self.assertEqual(summary.output_path, output_path.resolve())

            original_workbook = load_workbook(data_path)
            self.assertEqual(original_workbook.sheetnames, ["Data"])
            self.assertEqual(original_workbook["Data"]["A2"].value, "Use API key to sign in.")
            self.assertEqual(original_workbook["Data"]["B2"].value, "使用接口登录。")

            workbook = load_workbook(output_path)
            self.assertIn("术语命中问题", workbook.sheetnames)
            self.assertIn("检查汇总", workbook.sheetnames)

            data_sheet = workbook["Data"]
            self.assertEqual(data_sheet["A2"].value, "Use API key to sign in.")
            self.assertEqual(data_sheet["B2"].value, "使用接口登录。")

            problem_sheet = workbook["术语命中问题"]
            self.assertEqual(problem_sheet["A2"].value, 2)
            self.assertEqual(problem_sheet["B2"].value, "术语未按术语表翻译")
            self.assertEqual(problem_sheet["C2"].value, "API key")
            self.assertEqual(problem_sheet["D2"].value, "接口密钥")

            summary_sheet = workbook["检查汇总"]
            self.assertEqual(summary_sheet["B2"].value, "Glossary")

            summary_values = {
                summary_sheet.cell(row_index, 1).value: summary_sheet.cell(row_index, 2).value
                for row_index in range(2, 17)
            }
            self.assertEqual(summary_values["匹配模式"], "混合边界")
            self.assertEqual(summary_values["总行数"], 5)
            self.assertEqual(summary_values["命中术语行数"], 4)
            self.assertEqual(summary_values["问题行数"], 1)
            self.assertEqual(summary_values["问题条数"], 1)
            self.assertEqual(summary_values["术语表条数"], 5)
            self.assertEqual(summary_values["冲突术语数"], 1)
            self.assertEqual(summary_sheet["D2"].value, "Term")
            self.assertEqual(summary_sheet["E2"].value, "译法一 / 译法二")

            second_summary = process_excel(
                glossary_file=glossary_path,
                data_file=data_path,
                glossary_source_column="A",
                glossary_target_column="B",
                data_source_column="A",
                data_target_column="B",
                start_row=2,
                case_sensitive=False,
                match_mode="hybrid-boundary",
            )
            self.assertEqual(second_summary.problem_count, 1)
            self.assertEqual(second_summary.output_path, output_path.resolve())

            rerun_workbook = load_workbook(output_path)
            rerun_problem_sheet = rerun_workbook["术语命中问题"]
            self.assertEqual(rerun_problem_sheet.max_row, 2)

    def test_case_sensitive_mode_changes_source_hits(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            glossary_path = Path(tmp_dir) / "glossary.xlsx"
            data_path = Path(tmp_dir) / "data.xlsx"
            self.create_glossary_workbook(glossary_path)
            self.create_data_workbook(data_path)

            summary = process_excel(
                glossary_file=glossary_path,
                data_file=data_path,
                glossary_source_column="A",
                glossary_target_column="B",
                data_source_column="A",
                data_target_column="B",
                start_row=2,
                case_sensitive=True,
                match_mode="hybrid-boundary",
            )

            self.assertEqual(summary.matched_rows, 3)
            self.assertEqual(summary.problem_count, 1)

    def test_target_boundary_prevents_false_positive_and_substring_mode_allows_it(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            glossary_path = Path(tmp_dir) / "glossary.xlsx"
            data_path = Path(tmp_dir) / "data.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "ACC"
            worksheet["B2"] = "ACC"
            workbook.save(glossary_path)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "source"
            worksheet["B1"] = "target"
            worksheet["A2"] = "ACC details"
            worksheet["B2"] = "account details"
            workbook.save(data_path)

            boundary_summary = process_excel(
                glossary_file=glossary_path,
                data_file=data_path,
                glossary_source_column="A",
                glossary_target_column="B",
                data_source_column="A",
                data_target_column="B",
                start_row=2,
                match_mode="hybrid-boundary",
            )
            self.assertEqual(boundary_summary.problem_count, 1)

            substring_summary = process_excel(
                glossary_file=glossary_path,
                data_file=data_path,
                glossary_source_column="A",
                glossary_target_column="B",
                data_source_column="A",
                data_target_column="B",
                start_row=2,
                match_mode="substring",
            )
            self.assertEqual(substring_summary.problem_count, 0)


if __name__ == "__main__":
    unittest.main()
