from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from phraseloom import workbook_schema as schema
from phraseloom.excel_io import (
    _default_extract_output_path,
    _default_fill_output_path,
    _default_tm_output_path,
)
from phraseloom.workflow import (
    fill_target_column_workbook,
    generate_tm_pairs,
    generate_workbook,
)


DEFAULT_SOURCE_WORKBOOK = REPO_ROOT / "testfiles" / "for_test.xlsx"
DEFAULT_TM_WORKBOOK = REPO_ROOT / "testfiles" / "TM.xlsx"
PROTECTED_TOKEN_RE = re.compile(r"\{[1-9]\d*>|<[1-9]\d*\}|\{[1-9]\d*\}")
MISSING_TARGET_WARNING = "fill target in translation_units, then rerun fill"


@dataclass(frozen=True)
class StandardWorkflowResult:
    source_workbook: Path
    tm_workbook: Path
    tm_pairs_workbook: Path
    prefill_pack: Path
    translator_todo: Path
    filled_workbook: Path
    tm_stats: dict[str, Any]
    extract_stats: dict[str, Any]
    fill_stats: dict[str, Any]
    rates: dict[str, str]
    sheet_counts: dict[str, int | None]
    qa_report: dict[str, Any]
    source_map: dict[str, Any]
    filled_result: dict[str, int]


def run_standard_workflow(
    *,
    source_workbook: str | Path = DEFAULT_SOURCE_WORKBOOK,
    tm_workbook: str | Path = DEFAULT_TM_WORKBOOK,
    source_col: str = "source",
    target_col: str = "target",
    min_group_size: int = 2,
    tag_config: str | Path | None = None,
) -> StandardWorkflowResult:
    source_workbook = Path(source_workbook)
    tm_workbook = Path(tm_workbook)
    tag_config_path = Path(tag_config) if tag_config else None

    tm_pairs_workbook = _default_tm_output_path(tm_workbook)
    prefill_pack = _default_extract_output_path(source_workbook)
    filled_workbook = _default_fill_output_path(source_workbook)

    tm_stats = generate_tm_pairs(
        tm_workbook,
        tm_pairs_workbook,
        source_col=source_col,
        target_col=target_col,
        min_group_size=min_group_size,
        tag_config=tag_config_path,
    )
    extract_stats = generate_workbook(
        source_workbook,
        prefill_pack,
        source_col=source_col,
        target_col=target_col,
        tm_workbook=tm_pairs_workbook,
        min_group_size=min_group_size,
        use_existing_targets=False,
        tag_config=tag_config_path,
    )
    translator_todo = Path(str(extract_stats["to_translate_path"]))
    fill_stats = fill_target_column_workbook(
        source_workbook,
        filled_workbook,
        source_col=source_col,
        target_col=target_col,
        template_workbook=translator_todo,
        min_group_size=min_group_size,
        tag_config=tag_config_path,
    )

    return StandardWorkflowResult(
        source_workbook=source_workbook,
        tm_workbook=tm_workbook,
        tm_pairs_workbook=tm_pairs_workbook,
        prefill_pack=prefill_pack,
        translator_todo=translator_todo,
        filled_workbook=filled_workbook,
        tm_stats=tm_stats,
        extract_stats=extract_stats,
        fill_stats=fill_stats,
        rates=_coverage_rates(extract_stats),
        sheet_counts=_sheet_counts(
            tm_pairs_workbook, prefill_pack, translator_todo
        ),
        qa_report=_read_qa_report(prefill_pack),
        source_map=_read_source_map_stats(prefill_pack),
        filled_result=_read_filled_result_stats(filled_workbook, target_col),
    )


def format_summary(result: StandardWorkflowResult) -> list[str]:
    lines = [
        "Standard workflow complete.",
        "",
        "Inputs:",
        f"  Source workbook: {result.source_workbook}",
        f"  TM workbook: {result.tm_workbook}",
        "",
        "Outputs:",
        f"  TM reusable units: {result.tm_pairs_workbook}",
        f"  TM prefill pack: {result.prefill_pack}",
        f"  Translator todo: {result.translator_todo}",
        f"  Filled result: {result.filled_workbook}",
        "",
        "TM pairs:",
        f"  TM source rows: {result.tm_stats['row_count']}",
        f"  Unique source rows: {result.tm_stats['unique_source_segments']}",
        f"  Duplicate source rows: {result.tm_stats['duplicate_source_segments']}",
        f"  TM pairs: {result.tm_stats['tm_pair_count']}",
        f"  Template pairs: {result.tm_stats['template_pair_count']}",
        f"  Segment pairs: {result.tm_stats['segment_pair_count']}",
        "",
        "Source coverage:",
        f"  Source rows: {result.extract_stats['row_count']}",
        f"  Translation units: {result.extract_stats['translation_unit_count']}",
        f"  Template units: {result.extract_stats['template_unit_count']}",
        f"  Segment units: {result.extract_stats['segment_unit_count']}",
        f"  Already filled units: {result.extract_stats['prefilled_translation_unit_count']}",
        f"  Units to translate: {result.extract_stats['new_translation_unit_count']}",
        f"  Already filled source rows: {result.extract_stats['autofilled_count']}",
        f"  Source rows to translate: {result.extract_stats['new_source_segment_count']}",
        f"  TM unit hit rate: {result.rates['tm_unit_hit_rate']}",
        f"  TM row hit rate: {result.rates['tm_row_hit_rate']}",
        "",
        "QA:",
        f"  QA report: {_format_mapping(result.qa_report)}",
        f"  Source map status: {_format_mapping(result.source_map['status'])}",
        f"  Source map review warnings: {_format_mapping(result.source_map['review_warnings'])}",
        f"  Filled targets: {result.filled_result['non_empty_targets']}",
        f"  Residual protected-token-like rows: {result.filled_result['residual_token_rows']}",
        "",
        "Sheet counts:",
        f"  {_format_mapping(result.sheet_counts)}",
    ]
    return lines


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    result = run_standard_workflow(
        source_workbook=args.source_workbook,
        tm_workbook=args.tm_workbook,
        source_col=args.source_col,
        target_col=args.target_col,
        min_group_size=args.min_group_size,
        tag_config=args.tag_config,
    )
    for line in format_summary(result):
        _print(line)
    return 0


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run the standard PhraseLoom TM extraction, TM prefill, and fill "
            "workflow, then print pair and coverage metrics."
        )
    )
    parser.add_argument(
        "--source-workbook",
        type=Path,
        default=DEFAULT_SOURCE_WORKBOOK,
        help="Source workbook to prefill.",
    )
    parser.add_argument(
        "--tm-workbook",
        type=Path,
        default=DEFAULT_TM_WORKBOOK,
        help="Completed TM workbook used to build reusable units.",
    )
    parser.add_argument("--source-col", default="source")
    parser.add_argument("--target-col", default="target")
    parser.add_argument("--min-group-size", type=int, default=2)
    parser.add_argument(
        "--tag-config",
        type=Path,
        help="Optional TOML file defining protected tag rules.",
    )
    return parser.parse_args(argv)


def _coverage_rates(stats: dict[str, Any]) -> dict[str, str]:
    row_count = int(stats["row_count"])
    unit_count = int(stats["translation_unit_count"])
    filled_rows = int(stats["autofilled_count"])
    filled_units = int(stats["prefilled_translation_unit_count"])
    rows_to_translate = int(stats["new_source_segment_count"])
    units_to_translate = int(stats["new_translation_unit_count"])
    return {
        "tm_unit_hit_rate": _format_rate(filled_units, unit_count),
        "tm_row_hit_rate": _format_rate(filled_rows, row_count),
        "units_to_translate_rate": _format_rate(units_to_translate, unit_count),
        "source_rows_to_translate_rate": _format_rate(rows_to_translate, row_count),
    }


def _sheet_counts(
    tm_pairs_workbook: Path, prefill_pack: Path, translator_todo: Path
) -> dict[str, int | None]:
    return {
        "tm_pairs": _sheet_row_count(tm_pairs_workbook, schema.TM_PAIRS_SHEET),
        "tm_map": _sheet_row_count(tm_pairs_workbook, schema.TM_SOURCE_MAP_SHEET),
        "translation_units": _sheet_row_count(
            prefill_pack, schema.TRANSLATION_UNITS_SHEET
        ),
        "source_map": _sheet_row_count(prefill_pack, schema.SOURCE_MAP_SHEET),
        "filled_workbook": _sheet_row_count(
            prefill_pack, schema.FILLED_WORKBOOK_SHEET
        ),
        "to_translate": _sheet_row_count(translator_todo, schema.TO_TRANSLATE_SHEET),
        "prefilled_units": _sheet_row_count(
            translator_todo, schema.PREFILLED_UNITS_SHEET
        ),
    }


def _read_qa_report(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if schema.QA_REPORT_SHEET not in workbook.sheetnames:
            return {}
        worksheet = workbook[schema.QA_REPORT_SHEET]
        return {
            str(check): count
            for check, count, *_rest in worksheet.iter_rows(values_only=True)
            if check not in (None, schema.CHECK_COLUMN)
        }
    finally:
        workbook.close()


def _read_source_map_stats(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[schema.SOURCE_MAP_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        headers = list(next(rows))
        fill_status_index = headers.index(schema.FILL_STATUS_COLUMN)
        warning_index = headers.index(schema.WARNING_COLUMN)
        status: Counter[str] = Counter()
        review_warnings: Counter[str] = Counter()
        warning_rows = 0
        review_warning_rows = 0
        for row in rows:
            status[str(row[fill_status_index] or "")] += 1
            warning = str(row[warning_index] or "")
            if not warning:
                continue
            warning_rows += 1
            parts = warning.split("; ")
            if all(part == MISSING_TARGET_WARNING for part in parts):
                continue
            review_warning_rows += 1
            for part in parts:
                if part == MISSING_TARGET_WARNING:
                    continue
                review_warnings[_normalize_warning_part(part)] += 1
        return {
            "status": dict(status),
            "warning_rows": warning_rows,
            "review_warning_rows": review_warning_rows,
            "review_warnings": dict(review_warnings),
        }
    finally:
        workbook.close()


def _read_filled_result_stats(path: Path, target_col: str) -> dict[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(header or "").strip().lower() for header in next(rows)]
        target_index = headers.index(target_col.strip().lower())
        row_count = 0
        non_empty_targets = 0
        residual_token_rows = 0
        for row in rows:
            row_count += 1
            target = row[target_index] if target_index < len(row) else None
            if target in (None, ""):
                continue
            non_empty_targets += 1
            if PROTECTED_TOKEN_RE.search(str(target)):
                residual_token_rows += 1
        return {
            "rows": row_count,
            "non_empty_targets": non_empty_targets,
            "residual_token_rows": residual_token_rows,
        }
    finally:
        workbook.close()


def _sheet_row_count(path: Path, sheet_name: str) -> int | None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return None
        return max(workbook[sheet_name].max_row - 1, 0)
    finally:
        workbook.close()


def _normalize_warning_part(part: str) -> str:
    if part.startswith("open tag has no close partner"):
        return "open tag has no close partner"
    if part.startswith("unpaired close tag"):
        return "unpaired close tag"
    if part.startswith("protected_token_mismatch"):
        return "protected_token_mismatch"
    if part.startswith("source_protected_span_not_found"):
        return "source_protected_span_not_found"
    return part


def _format_rate(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return "0.00%"
    return f"{numerator / denominator * 100:.2f}%"


def _format_mapping(values: dict[str, Any]) -> str:
    if not values:
        return "(none)"
    return ", ".join(f"{key}={value}" for key, value in values.items())


def _print(message: str) -> None:
    encoding = sys.stdout.encoding or "utf-8"
    safe = message.encode(encoding, errors="backslashreplace").decode(encoding)
    print(safe)


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
