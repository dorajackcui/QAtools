from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from phraseloom import workbook_schema as schema


WARNING_QA_SHEET = "warning_qa"
SUMMARY_SHEET = "summary"
WARNING_QA_COLUMNS = [
    "tm_id",
    "row_number",
    "unit_type",
    "warning",
    "raw_source",
    "raw_target",
    "row_source_unit",
    "row_target_unit",
    "pair_source_unit",
    "pair_target_unit",
    "variables",
]


@dataclass(frozen=True)
class RawTmRow:
    row_number: int
    source: Any
    target: Any


@dataclass(frozen=True)
class TmMapRow:
    row_number: int
    source_unit: Any
    target_unit: Any
    variables: Any


@dataclass(frozen=True)
class WarningPair:
    tm_id: Any
    unit_type: Any
    source_unit: Any
    target_unit: Any
    variables: Any
    row_numbers: tuple[int, ...]
    warning: Any


def export_warning_qa(
    tm_workbook: str | Path,
    tm_units: str | Path,
    output_path: str | Path,
    *,
    source_col: str = "source",
    target_col: str = "target",
) -> dict[str, int]:
    tm_workbook = Path(tm_workbook)
    tm_units = Path(tm_units)
    output_path = Path(output_path)

    raw_rows = _read_raw_tm_rows(tm_workbook, source_col, target_col)
    tm_map_rows = _read_tm_map_rows(tm_units)
    warning_pairs = _read_warning_pairs(tm_units)
    records = _build_warning_records(warning_pairs, raw_rows, tm_map_rows)
    _write_warning_qa_workbook(output_path, records, warning_pairs)

    return {
        "warning_pair_count": len(warning_pairs),
        "warning_raw_row_count": len({record["row_number"] for record in records}),
        "exported_record_count": len(records),
    }


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    stats = export_warning_qa(
        args.tm_workbook,
        args.tm_units,
        args.output,
        source_col=args.source_col,
        target_col=args.target_col,
    )
    _print(f"TM workbook: {args.tm_workbook}")
    _print(f"TM reusable units: {args.tm_units}")
    _print(f"Output: {args.output}")
    _print(f"Warning pairs: {stats['warning_pair_count']}")
    _print(f"Warning raw rows: {stats['warning_raw_row_count']}")
    _print(f"Exported records: {stats['exported_record_count']}")
    return 0


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    default_dir = REPO_ROOT / "testfiles" / "tm_self_prefill_verification"
    parser = argparse.ArgumentParser(
        description=(
            "Export TM warning rows with raw source/target and serialized unit context."
        )
    )
    parser.add_argument(
        "--tm-workbook",
        type=Path,
        default=default_dir / "TM.xlsx",
        help="Original TM workbook containing raw source/target columns.",
    )
    parser.add_argument(
        "--tm-units",
        type=Path,
        default=default_dir / "TM_reusable_units.xlsx",
        help="Reusable TM units workbook produced by tm-extract.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=default_dir / "TM_warning_qa.xlsx",
        help="QA workbook to write.",
    )
    parser.add_argument("--source-col", default="source")
    parser.add_argument("--target-col", default="target")
    return parser.parse_args(argv)


def _read_raw_tm_rows(
    path: Path, source_col: str, target_col: str
) -> dict[int, RawTmRow]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        source_index = _header_index(worksheet, source_col)
        target_index = _header_index(worksheet, target_col)
        rows: dict[int, RawTmRow] = {}
        for row_number in range(2, worksheet.max_row + 1):
            rows[row_number] = RawTmRow(
                row_number=row_number,
                source=worksheet.cell(row=row_number, column=source_index).value,
                target=worksheet.cell(row=row_number, column=target_index).value,
            )
        return rows
    finally:
        workbook.close()


def _read_tm_map_rows(path: Path) -> dict[int, TmMapRow]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = _required_sheet(workbook, schema.TM_SOURCE_MAP_SHEET)
        headers = _headers(worksheet)
        row_number_index = _column_index(headers, schema.ROW_NUMBER_COLUMN)
        source_unit_index = _column_index(headers, schema.SOURCE_UNIT_COLUMN)
        target_unit_index = _column_index(headers, schema.TARGET_UNIT_COLUMN)
        variables_index = _column_index(headers, schema.VARIABLE_VALUES_COLUMN)
        rows: dict[int, TmMapRow] = {}
        for row in worksheet.iter_rows(min_row=2):
            row_number = row[row_number_index].value
            if row_number in (None, ""):
                continue
            rows[int(row_number)] = TmMapRow(
                row_number=int(row_number),
                source_unit=row[source_unit_index].value,
                target_unit=row[target_unit_index].value,
                variables=row[variables_index].value,
            )
        return rows
    finally:
        workbook.close()


def _read_warning_pairs(path: Path) -> list[WarningPair]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = _required_sheet(workbook, schema.TM_PAIRS_SHEET)
        headers = _headers(worksheet)
        tm_id_index = _column_index(headers, schema.TM_ID_COLUMN)
        unit_type_index = _column_index(headers, schema.UNIT_TYPE_COLUMN)
        source_unit_index = _column_index(headers, schema.SOURCE_UNIT_COLUMN)
        target_unit_index = _column_index(headers, schema.TARGET_UNIT_COLUMN)
        variables_index = _column_index(headers, schema.VARIABLES_COLUMN)
        row_numbers_index = _column_index(headers, schema.ROW_NUMBERS_COLUMN)
        warning_index = _column_index(headers, schema.WARNING_COLUMN)
        pairs: list[WarningPair] = []
        for row in worksheet.iter_rows(min_row=2):
            warning = row[warning_index].value
            if warning in (None, ""):
                continue
            pairs.append(
                WarningPair(
                    tm_id=row[tm_id_index].value,
                    unit_type=row[unit_type_index].value,
                    source_unit=row[source_unit_index].value,
                    target_unit=row[target_unit_index].value,
                    variables=row[variables_index].value,
                    row_numbers=_parse_row_numbers(row[row_numbers_index].value),
                    warning=warning,
                )
            )
        return pairs
    finally:
        workbook.close()


def _build_warning_records(
    warning_pairs: list[WarningPair],
    raw_rows: dict[int, RawTmRow],
    tm_map_rows: dict[int, TmMapRow],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for pair in warning_pairs:
        for row_number in pair.row_numbers:
            raw_row = raw_rows.get(row_number)
            map_row = tm_map_rows.get(row_number)
            records.append(
                {
                    "tm_id": pair.tm_id,
                    "row_number": row_number,
                    "unit_type": pair.unit_type,
                    "warning": pair.warning,
                    "raw_source": raw_row.source if raw_row else None,
                    "raw_target": raw_row.target if raw_row else None,
                    "row_source_unit": map_row.source_unit if map_row else None,
                    "row_target_unit": map_row.target_unit if map_row else None,
                    "pair_source_unit": pair.source_unit,
                    "pair_target_unit": pair.target_unit,
                    "variables": map_row.variables if map_row else pair.variables,
                }
            )
    records.sort(key=lambda record: (record["row_number"], str(record["tm_id"])))
    return records


def _write_warning_qa_workbook(
    output_path: Path, records: list[dict[str, Any]], warning_pairs: list[WarningPair]
) -> None:
    workbook = Workbook()
    try:
        summary = workbook.active
        summary.title = SUMMARY_SHEET
        summary.append(["check", "count"])
        summary.append(["warning_pairs", len(warning_pairs)])
        summary.append(["warning_raw_rows", len({record["row_number"] for record in records})])
        summary.append(["exported_records", len(records)])

        qa = workbook.create_sheet(WARNING_QA_SHEET)
        qa.append(WARNING_QA_COLUMNS)
        for record in records:
            qa.append([record[column] for column in WARNING_QA_COLUMNS])

        _style_sheet(summary)
        _style_sheet(qa)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()


def _style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        header = column_cells[0].value
        if header in {"raw_source", "raw_target", "row_source_unit", "row_target_unit"}:
            width = 64
        elif header in {"pair_source_unit", "pair_target_unit", "warning"}:
            width = 56
        else:
            width = 18
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _required_sheet(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"required sheet not found: {sheet_name}")
    return workbook[sheet_name]


def _headers(worksheet) -> list[Any]:
    return [cell.value for cell in worksheet[1]]


def _column_index(headers: list[Any], name: str) -> int:
    normalized = [str(header).strip().lower() for header in headers]
    for candidate in schema.UNIT_COLUMN_ALIASES.get(name, (name,)):
        try:
            return normalized.index(candidate.strip().lower())
        except ValueError:
            continue
    raise ValueError(f"required column not found: {name}")


def _header_index(worksheet, name: str) -> int:
    headers = _headers(worksheet)
    return _column_index(headers, name) + 1


def _parse_row_numbers(value: Any) -> tuple[int, ...]:
    return tuple(int(match.group(0)) for match in re.finditer(r"\d+", str(value or "")))


def _print(message: str) -> None:
    encoding = sys.stdout.encoding or "utf-8"
    safe = message.encode(encoding, errors="backslashreplace").decode(encoding)
    print(safe)


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
