from __future__ import annotations

import argparse
import re
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from phraseloom.workflow import (
    fill_target_column_workbook,
    generate_tm_pairs,
    generate_workbook,
)


PROTECTED_TOKEN_RE = re.compile(r"\{[1-9]\d*>|<[1-9]\d*\}|\{[1-9]\d*\}")


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    input_path = args.input or REPO_ROOT / "testfiles" / "TM.xlsx"
    output_dir = args.output_dir or REPO_ROOT / "testfiles" / "tm_self_prefill_verification"
    output_dir.mkdir(parents=True, exist_ok=True)

    tm_copy = output_dir / "TM.xlsx"
    targetless = output_dir / "TM_as_source_without_target.xlsx"
    tm_pairs = output_dir / "TM_reusable_units.xlsx"
    pack = output_dir / "TM_as_source_l10n" / "TM_as_source_tm_prefill_pack.xlsx"
    filled = output_dir / "TM_as_source_l10n" / "TM_as_source_filled_result.xlsx"

    shutil.copyfile(input_path, tm_copy)
    _copy_with_empty_target(tm_copy, targetless, args.target_col)

    tm_stats = generate_tm_pairs(
        tm_copy,
        tm_pairs,
        source_col=args.source_col,
        target_col=args.target_col,
        min_group_size=args.min_group_size,
    )
    extract_stats = generate_workbook(
        targetless,
        pack,
        source_col=args.source_col,
        target_col=None,
        tm_workbook=tm_pairs,
        min_group_size=args.min_group_size,
        use_existing_targets=False,
    )
    todo_path = Path(str(extract_stats["to_translate_path"]))
    fill_stats = fill_target_column_workbook(
        targetless,
        filled,
        source_col=args.source_col,
        target_col=args.target_col,
        template_workbook=todo_path,
        min_group_size=args.min_group_size,
    )

    original_rows = _read_rows(tm_copy, args.source_col, args.target_col)
    filled_rows = _read_rows(filled, args.source_col, args.target_col)
    comparison = _compare_rows(original_rows, filled_rows)
    warning_summary = _read_source_map_warning_summary(pack)
    residual_rows = [
        row_number
        for row_number, _source, target in filled_rows
        if target and PROTECTED_TOKEN_RE.search(str(target))
    ]
    empty_filled_rows = [
        row_number for row_number, _source, target in filled_rows if target in (None, "")
    ]

    _print(f"Input: {input_path}")
    _print(f"Output dir: {output_dir}")
    _print(f"TM pairs: {tm_pairs}")
    _print(f"Prefill pack: {pack}")
    _print(f"Translator todo: {todo_path}")
    _print(f"Filled workbook: {filled}")
    _print("")
    _print(f"TM stats: {tm_stats}")
    _print(f"Extract stats: {extract_stats}")
    _print(f"Fill stats: {fill_stats}")
    _print("")
    _print(f"Rows compared: {comparison['rows_compared']}")
    _print(f"Exact target mismatches: {len(comparison['mismatches'])}")
    _print(f"Empty filled targets: {len(empty_filled_rows)}")
    _print(f"Residual protected-token rows: {len(residual_rows)}")
    _print(f"Protected-token warning rows in source_map: {warning_summary['protected_token_rows']}")
    _print(f"Tag warning rows in source_map: {warning_summary['tag_warning_rows']}")

    if comparison["mismatches"]:
        _print("")
        _print("First target mismatches:")
        for row_number, source, expected, actual in comparison["mismatches"][:10]:
            _print(f"- row {row_number}: source={source!r}")
            _print(f"  expected={expected!r}")
            _print(f"  actual  ={actual!r}")

    if warning_summary["protected_token_samples"]:
        _print("")
        _print("First protected-token warning rows:")
        for sample in warning_summary["protected_token_samples"]:
            _print(f"- row {sample['row_number']}: source={sample['source']!r}")
            _print(f"  auto_target={sample['auto_target']!r}")
            _print(f"  warning={sample['warning']!r}")

    failures: list[str] = []
    if comparison["rows_compared"] != len(original_rows):
        failures.append("filled row count differs from original row count")
    if empty_filled_rows:
        failures.append("filled workbook contains empty target cells")
    if residual_rows:
        failures.append("filled workbook contains residual protected tokens")
    if warning_summary["protected_token_rows"]:
        failures.append("source_map contains protected_token_mismatch warnings")
    if args.strict_targets and comparison["mismatches"]:
        failures.append("strict target comparison found mismatches")

    if failures:
        _print("")
        _print("Verification failed:")
        for failure in failures:
            _print(f"- {failure}")
        return 1

    _print("")
    _print("Structural verification passed.")
    if comparison["mismatches"]:
        _print(
            "Exact target mismatches were reported above; use --strict-targets to fail on them."
        )
    return 0


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run TM.xlsx as both TM and targetless source, then verify fill output."
    )
    parser.add_argument("--input", type=Path, help="Workbook to use as both TM and source.")
    parser.add_argument("--output-dir", type=Path, help="Directory for generated workbooks.")
    parser.add_argument("--source-col", default="source")
    parser.add_argument("--target-col", default="target")
    parser.add_argument("--min-group-size", type=int, default=2)
    parser.add_argument(
        "--strict-targets",
        action="store_true",
        help="Fail if any filled target differs from the original TM target.",
    )
    return parser.parse_args(argv)


def _print(message: str) -> None:
    encoding = sys.stdout.encoding or "utf-8"
    safe = message.encode(encoding, errors="backslashreplace").decode(encoding)
    print(safe)


def _copy_with_empty_target(source: Path, target: Path, target_col: str) -> None:
    workbook = load_workbook(source)
    try:
        worksheet = workbook.worksheets[0]
        target_index = _header_index(worksheet, target_col)
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_number, column=target_index).value = None
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(target)
    finally:
        workbook.close()


def _read_rows(path: Path, source_col: str, target_col: str) -> list[tuple[int, Any, Any]]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        source_index = _header_index(worksheet, source_col)
        target_index = _header_index(worksheet, target_col)
        rows: list[tuple[int, Any, Any]] = []
        for row_number in range(2, worksheet.max_row + 1):
            rows.append(
                (
                    row_number,
                    worksheet.cell(row=row_number, column=source_index).value,
                    worksheet.cell(row=row_number, column=target_index).value,
                )
            )
        return rows
    finally:
        workbook.close()


def _compare_rows(
    original_rows: list[tuple[int, Any, Any]],
    filled_rows: list[tuple[int, Any, Any]],
) -> dict[str, Any]:
    mismatches: list[tuple[int, Any, Any, Any]] = []
    rows_compared = min(len(original_rows), len(filled_rows))
    for original, filled in zip(original_rows, filled_rows):
        row_number, source, expected = original
        _filled_row_number, _filled_source, actual = filled
        if expected != actual:
            mismatches.append((row_number, source, expected, actual))
    return {"rows_compared": rows_compared, "mismatches": mismatches}


def _read_source_map_warning_summary(path: Path) -> dict[str, int]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook["source_map"]
        headers = [cell.value for cell in worksheet[1]]
        warning_index = headers.index("warning") + 1
        row_number_index = headers.index("row_number") + 1
        source_index = headers.index("source") + 1
        auto_target_index = headers.index("auto_target") + 1
        protected_token_rows = 0
        tag_warning_rows = 0
        protected_token_samples: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=2):
            warning = row[warning_index - 1].value or ""
            if "protected_token_mismatch:" in str(warning):
                protected_token_rows += 1
                if len(protected_token_samples) < 10:
                    protected_token_samples.append(
                        {
                            "row_number": row[row_number_index - 1].value,
                            "source": row[source_index - 1].value,
                            "auto_target": row[auto_target_index - 1].value,
                            "warning": warning,
                        }
                    )
            if "tag has no close partner" in str(warning) or "unpaired close tag" in str(warning):
                tag_warning_rows += 1
        return {
            "protected_token_rows": protected_token_rows,
            "tag_warning_rows": tag_warning_rows,
            "protected_token_samples": protected_token_samples,
        }
    finally:
        workbook.close()


def _header_index(worksheet, name: str) -> int:
    for index, cell in enumerate(worksheet[1], start=1):
        if str(cell.value).strip().lower() == name.strip().lower():
            return index
    raise ValueError(f"column not found: {name}")


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
