from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .errors import ColumnNotFoundError
from .models import RowItem
from .tag_engine import extract_tags
from .tag_rules import TagRules, default_tag_rules
from .template_engine import parse_template


def read_source_rows(
    input_path: Path,
    source_column: str | int,
    target_column: str | int | None,
    *,
    tag_rules: TagRules | None = None,
) -> list[RowItem]:
    active_rules = tag_rules or default_tag_rules()
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        source_index = resolve_column(worksheet, source_column)
        target_index = (
            resolve_column(worksheet, target_column)
            if target_column is not None
            else None
        )

        rows: list[RowItem] = []
        seen_source = False
        consecutive_blanks = 0
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            source_value = _cell_value(values, source_index)
            if source_value is None or not str(source_value).strip():
                if seen_source:
                    consecutive_blanks += 1
                    if consecutive_blanks >= 1000:
                        break
                continue

            seen_source = True
            consecutive_blanks = 0
            raw_source = str(source_value).strip()
            protected_source = extract_tags(raw_source, rules=active_rules).text
            target_value = _cell_value(values, target_index)
            raw_target = "" if target_value is None else str(target_value).strip()
            rows.append(
                RowItem(
                    row_number=row_number,
                    source=protected_source,
                    match=parse_template(protected_source),
                    original_values=tuple(values),
                    raw_source=raw_source,
                    raw_existing_target=raw_target,
                )
            )
        return rows
    finally:
        workbook.close()


def read_headers(input_path: Path) -> list[str]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        return header_values(workbook.worksheets[0], fallback=True)
    finally:
        workbook.close()


def resolve_column(worksheet, column: str | int | None) -> int:
    if column is None:
        raise ColumnNotFoundError(column, header_values(worksheet, fallback=True))
    if isinstance(column, int) or str(column).isdigit():
        return int(column)

    wanted = str(column).strip().lower()
    for index, cell in enumerate(worksheet[1], start=1):
        value = "" if cell.value is None else str(cell.value).strip()
        if value.lower() == wanted:
            return getattr(cell, "column", index)
    raise ColumnNotFoundError(column, header_values(worksheet, fallback=True))


def header_values(worksheet, *, fallback: bool = False) -> list[str]:
    headers: list[str] = []
    for index, cell in enumerate(worksheet[1], start=1):
        value = "" if cell.value is None else str(cell.value).strip()
        headers.append(value or (f"column_{index}" if fallback else ""))
    return headers


def workbook_metadata(workbook) -> dict[str, object]:
    from . import workbook_schema as schema

    if schema.METADATA_SHEET not in workbook.sheetnames:
        return {}
    worksheet = workbook[schema.METADATA_SHEET]
    return {
        row[0]: row[1]
        for row in worksheet.iter_rows(min_row=2, max_col=2, values_only=True)
        if row and row[0] is not None
    }


def _cell_value(
    row: tuple[object, ...],
    one_based_index: int | None,
) -> object | None:
    if one_based_index is None:
        return None
    zero_based = one_based_index - 1
    return row[zero_based] if zero_based < len(row) else None


__all__ = [
    "header_values",
    "read_headers",
    "read_source_rows",
    "resolve_column",
    "workbook_metadata",
]
