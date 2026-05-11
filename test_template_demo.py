import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


class TemplateDemoTests(unittest.TestCase):
    def test_infers_target_template_from_one_example_and_applies_to_matching_source(self):
        from template_demo import apply_target_template, infer_target_template, parse_template

        example = parse_template("VIP10 Paid Pack")

        self.assertEqual(example.template, "VIP{num1} Paid Pack")
        self.assertEqual(example.values, {"num1": "10"})

        target_template = infer_target_template(example.values, "VIP10pack")

        self.assertEqual(target_template, "VIP{num1}pack")
        self.assertEqual(apply_target_template(target_template, {"num1": "11"}), "VIP11pack")

    def test_preserves_named_placeholders_and_uses_readable_numeric_names(self):
        from template_demo import parse_template

        named = parse_template("Player reaches level {a}")
        stage = parse_template("Clear Story 10-20")

        self.assertEqual(named.template, "Player reaches level {a}")
        self.assertEqual(named.values, {"a": "{a}"})
        self.assertEqual(stage.template, "Clear Story {stage1}")
        self.assertEqual(stage.values, {"stage1": "10-20"})

    def test_generated_workbook_prefills_auto_targets_from_cli_example(self):
        from template_demo import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet2"
            ws.append(["中文", "英語", "target"])
            ws.append(["VIP0付费礼包", "VIP0 Paid Pack", ""])
            ws.append(["VIP10付费礼包", "VIP10 Paid Pack", ""])
            ws.append(["登录游戏3天", "Sign in for 3 days", ""])
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                output_path,
                source_col="英語",
                target_col="target",
                examples=[("VIP10 Paid Pack", "VIP10pack")],
                min_group_size=2,
            )

            self.assertEqual(stats["template_count"], 1)
            self.assertEqual(stats["clustered_source_segments"], 2)
            self.assertEqual(stats["autofilled_count"], 2)

            out = load_workbook(output_path, data_only=True)
            summary = {
                row[0].value: row[1].value
                for row in out["summary"].iter_rows(min_row=1, max_col=2)
            }
            self.assertEqual(
                summary,
                {
                    "total_source_rows": 3,
                    "total_translation_units": 2,
                    "already_filled_units": 1,
                    "already_filled_source_rows": 2,
                    "units_to_translate": 1,
                    "source_rows_to_translate": 1,
                },
            )

            units = out["translation_units"]
            unit_headers = [cell.value for cell in units[1]]
            unit_rows = [
                dict(zip(unit_headers, row))
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            by_source = {row["source_unit"]: row for row in unit_rows}
            self.assertEqual(by_source["VIP{num1} Paid Pack"]["unit_type"], "template")
            self.assertEqual(by_source["Sign in for 3 days"]["unit_type"], "segment")

            auto = out["source_map"]
            headers = [cell.value for cell in auto[1]]
            source_idx = headers.index("source") + 1
            auto_idx = headers.index("auto_target") + 1

            actual = {
                row[source_idx - 1].value: row[auto_idx - 1].value
                for row in auto.iter_rows(min_row=2)
            }

            self.assertEqual(actual["VIP0 Paid Pack"], "VIP0pack")
            self.assertEqual(actual["VIP10 Paid Pack"], "VIP10pack")
            self.assertEqual(actual["Sign in for 3 days"], None)

    def test_minimal_translation_units_deduplicate_segments_and_fill_duplicate_rows(self):
        from template_demo import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            pack_path = Path(tmp) / "pack.xlsx"
            filled_path = Path(tmp) / "filled.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["英語", "target"])
            ws.append(["Obtain 3 stars in Chapter 1", ""])
            ws.append(["Obtain 3 stars in Chapter 1", ""])
            ws.append(["Obtain 3 stars in Chapter 2", ""])
            ws.append(["Login failed", ""])
            ws.append(["Login failed", ""])
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                pack_path,
                source_col="英語",
                target_col="target",
                min_group_size=2,
                use_existing_targets=False,
            )
            self.assertEqual(stats["template_count"], 1)
            self.assertEqual(stats["segment_unit_count"], 1)
            self.assertEqual(stats["translation_unit_count"], 2)
            self.assertEqual(stats["duplicate_source_segments"], 2)

            pack = load_workbook(pack_path)
            units = pack["translation_units"]
            headers = [cell.value for cell in units[1]]
            target_idx = headers.index("target_unit") + 1
            source_idx = headers.index("source_unit") + 1
            for row in units.iter_rows(min_row=2):
                source_unit = row[source_idx - 1].value
                if source_unit == "Obtain {num1} stars in Chapter {num2}":
                    row[target_idx - 1].value = "第{num2}章获得{num1}颗星"
                elif source_unit == "Login failed":
                    row[target_idx - 1].value = "登录失败"
            pack.save(pack_path)

            fill_stats = generate_workbook(
                input_path,
                filled_path,
                source_col="英語",
                target_col="target",
                template_workbook=pack_path,
                min_group_size=2,
                use_existing_targets=False,
            )
            self.assertEqual(fill_stats["autofilled_count"], 5)

            filled = load_workbook(filled_path, data_only=True)
            source_map = filled["source_map"]
            headers = [cell.value for cell in source_map[1]]
            targets = [
                row[headers.index("auto_target")]
                for row in source_map.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(
                targets,
                [
                    "第1章获得3颗星",
                    "第1章获得3颗星",
                    "第2章获得3颗星",
                    "登录失败",
                    "登录失败",
                ],
            )

    def test_tm_pairs_prefill_translation_units_and_leave_new_units_blank(self):
        from template_demo import generate_tm_pairs, generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            tm_input = Path(tmp) / "tm.xlsx"
            tm_pairs = Path(tmp) / "tm_pairs.xlsx"
            source_input = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "new_units.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Obtain 3 stars in Chapter 1", "第1章获得3颗星"])
            ws.append(["Obtain 3 stars in Chapter 2", "第2章获得3颗星"])
            ws.append(["Login failed", "登录失败"])
            ws.append(["Login failed", "登录失败"])
            wb.save(tm_input)

            tm_stats = generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
                min_group_size=2,
            )
            self.assertEqual(tm_stats["tm_pair_count"], 2)
            self.assertEqual(tm_stats["template_pair_count"], 1)
            self.assertEqual(tm_stats["segment_pair_count"], 1)

            wb = Workbook()
            ws = wb.active
            ws.append(["source"])
            ws.append(["Obtain 5 stars in Chapter 3"])
            ws.append(["Login failed"])
            ws.append(["Brand new line"])
            wb.save(source_input)

            stats = generate_workbook(
                source_input,
                output_path,
                source_col="source",
                target_col=None,
                tm_workbook=tm_pairs,
                min_group_size=2,
                use_existing_targets=False,
            )
            self.assertEqual(stats["translation_unit_count"], 3)
            self.assertEqual(stats["prefilled_translation_unit_count"], 2)
            self.assertEqual(stats["untranslated_translation_unit_count"], 1)
            self.assertEqual(stats["new_source_segment_count"], 1)

            out = load_workbook(output_path, data_only=True)
            self.assertIn("to_translate", out.sheetnames)
            self.assertIn("filled_workbook", out.sheetnames)

            summary = {
                row[0].value: row[1].value
                for row in out["summary"].iter_rows(min_row=1, max_col=2)
            }
            self.assertEqual(
                summary,
                {
                    "total_source_rows": 3,
                    "total_translation_units": 3,
                    "already_filled_units": 2,
                    "already_filled_source_rows": 2,
                    "units_to_translate": 1,
                    "source_rows_to_translate": 1,
                },
            )

            units = out["translation_units"]
            headers = [cell.value for cell in units[1]]
            rows = [
                dict(zip(headers, row))
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            by_source = {row["source_unit"]: row for row in rows}

            self.assertEqual(
                by_source["Obtain {num1} stars in Chapter {num2}"]["target_unit"],
                "第{num2}章获得{num1}颗星",
            )
            self.assertEqual(
                by_source["Obtain {num1} stars in Chapter {num2}"]["target_unit_source"],
                "tm_pairs",
            )
            self.assertEqual(by_source["Login failed"]["target_unit"], "登录失败")
            self.assertIsNone(by_source["Brand new line"]["target_unit"])

            todo = out["to_translate"]
            todo_headers = [cell.value for cell in todo[1]]
            todo_rows = [
                dict(zip(todo_headers, row))
                for row in todo.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(len(todo_rows), 1)
            self.assertEqual(todo_rows[0]["source_unit"], "Brand new line")

            filled = out["filled_workbook"]
            filled_headers = [cell.value for cell in filled[1]]
            self.assertIn("auto_target", filled_headers)
            self.assertIn("fill_status", filled_headers)
            filled_rows = [
                dict(zip(filled_headers, row))
                for row in filled.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(filled_rows[0]["auto_target"], "第3章获得5颗星")
            self.assertEqual(filled_rows[1]["auto_target"], "登录失败")
            self.assertEqual(filled_rows[2]["fill_status"], "missing_target_unit")

            standalone_todo = source_input.parent / "source_l10n" / "source_to_translate.xlsx"
            self.assertTrue(standalone_todo.exists())
            todo_book = load_workbook(standalone_todo, data_only=True)
            self.assertEqual(todo_book.sheetnames, ["to_translate", "prefilled_units"])
            self.assertEqual(todo_book["prefilled_units"].sheet_state, "hidden")

    def test_non_translatable_numeric_and_symbol_segments_are_autofilled(self):
        from template_demo import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source"])
            ws.append(["123"])
            ws.append(["..."])
            ws.append(["10-20"])
            ws.append(["$0.99"])
            ws.append(["New text"])
            wb.save(input_path)

            stats = generate_workbook(
                input_path,
                output_path,
                source_col="source",
                target_col=None,
                min_group_size=2,
                use_existing_targets=False,
            )

            self.assertEqual(stats["autofilled_count"], 4)
            self.assertEqual(stats["new_translation_unit_count"], 1)
            self.assertEqual(stats["new_source_segment_count"], 1)

            out = load_workbook(output_path, data_only=True)
            units = out["translation_units"]
            headers = [cell.value for cell in units[1]]
            rows = [
                dict(zip(headers, row))
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            by_source = {row["source_unit"]: row for row in rows}

            for source in ["123", "...", "10-20", "$0.99"]:
                self.assertEqual(by_source[source]["target_unit"], source)
                self.assertEqual(by_source[source]["target_unit_source"], "non_translatable")

            todo = out["to_translate"]
            todo_headers = [cell.value for cell in todo[1]]
            todo_rows = [
                dict(zip(todo_headers, row))
                for row in todo.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual([row["source_unit"] for row in todo_rows], ["New text"])

    def test_translated_to_translate_can_fill_target_column_in_output_copy(self):
        from template_demo import (
            fill_target_column_workbook,
            generate_tm_pairs,
            generate_workbook,
        )

        with tempfile.TemporaryDirectory() as tmp:
            tm_input = Path(tmp) / "tm.xlsx"
            tm_pairs = Path(tmp) / "tm_pairs.xlsx"
            source_input = Path(tmp) / "source.xlsx"
            pack_path = Path(tmp) / "source_pack.xlsx"
            target_output = Path(tmp) / "source_filled_target.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "target"])
            ws.append(["Obtain 3 stars in Chapter 1", "第1章获得3颗星"])
            ws.append(["Obtain 3 stars in Chapter 2", "第2章获得3颗星"])
            ws.append(["Login failed", "登录失败"])
            wb.save(tm_input)

            generate_tm_pairs(
                tm_input,
                tm_pairs,
                source_col="source",
                target_col="target",
                min_group_size=2,
            )

            wb = Workbook()
            ws = wb.active
            ws.append(["source", "fr"])
            ws.append(["Obtain 5 stars in Chapter 3", ""])
            ws.append(["Login failed", ""])
            ws.append(["Brand new line", ""])
            wb.save(source_input)

            generate_workbook(
                source_input,
                pack_path,
                source_col="source",
                target_col="fr",
                tm_workbook=tm_pairs,
                min_group_size=2,
                use_existing_targets=False,
            )

            standalone_todo = source_input.parent / "source_l10n" / "source_to_translate.xlsx"
            todo_book = load_workbook(standalone_todo)
            ws = todo_book["to_translate"]
            headers = [cell.value for cell in ws[1]]
            target_idx = headers.index("target_unit") + 1
            for row in ws.iter_rows(min_row=2):
                row[target_idx - 1].value = "全新文本"
            todo_book.save(standalone_todo)

            stats = fill_target_column_workbook(
                source_input,
                target_output,
                source_col="source",
                target_col="fr",
                template_workbook=standalone_todo,
                min_group_size=2,
            )
            self.assertEqual(stats["autofilled_count"], 3)

            filled = load_workbook(target_output, data_only=True)
            rows = list(filled.active.iter_rows(values_only=True))
            self.assertEqual(rows[0], ("source", "fr"))
            self.assertEqual(rows[1][1], "第3章获得5颗星")
            self.assertEqual(rows[2][1], "登录失败")
            self.assertEqual(rows[3][1], "全新文本")

    def test_default_output_paths_are_inside_input_work_folder(self):
        from template_demo import (
            _default_extract_output_path,
            _default_fill_output_path,
            _default_tm_output_path,
            _default_to_translate_output_path,
            _default_work_dir,
        )

        input_path = Path("/tmp/project/source.xlsx")

        self.assertEqual(
            _default_work_dir(input_path),
            Path("/tmp/project/source_l10n"),
        )
        self.assertEqual(
            _default_extract_output_path(input_path),
            Path("/tmp/project/source_l10n/source_template_pack.xlsx"),
        )
        self.assertEqual(
            _default_to_translate_output_path(input_path),
            Path("/tmp/project/source_l10n/source_to_translate.xlsx"),
        )
        self.assertEqual(
            _default_fill_output_path(input_path),
            Path("/tmp/project/source_l10n/source_filled.xlsx"),
        )
        self.assertEqual(
            _default_tm_output_path(input_path),
            Path("/tmp/project/source_l10n/source_tm_pairs.xlsx"),
        )

    def test_interactive_path_input_accepts_copied_shell_quotes(self):
        from template_demo import _user_path

        self.assertEqual(
            _user_path("'/tmp/project/source.xlsx'"),
            Path("/tmp/project/source.xlsx"),
        )
        self.assertEqual(
            _user_path('"/tmp/project/source.xlsx"'),
            Path("/tmp/project/source.xlsx"),
        )

    def test_repeated_numeric_source_without_variants_stays_segment(self):
        from template_demo import generate_workbook

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["英語"])
            ws.append(["VIP10 Paid Pack"])
            ws.append(["VIP10 Paid Pack"])
            ws.append(["VIP10 Paid Pack"])
            wb.save(input_path)

            generate_workbook(
                input_path,
                output_path,
                source_col="英語",
                target_col=None,
                min_group_size=2,
                use_existing_targets=False,
            )

            out = load_workbook(output_path, data_only=True)
            units = out["translation_units"]
            headers = [cell.value for cell in units[1]]
            rows = [
                dict(zip(headers, row))
                for row in units.iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["unit_type"], "segment")
            self.assertEqual(rows[0]["source_unit"], "VIP10 Paid Pack")
            self.assertEqual(rows[0]["coverage_count"], 3)

    def test_interactive_menu_shows_three_step_workflow(self):
        from template_demo import main

        stdout = StringIO()
        with patch("builtins.input", side_effect=["q"]), redirect_stdout(stdout):
            exit_code = main([])

        menu = stdout.getvalue()
        self.assertEqual(exit_code, 0)
        self.assertIn("1) Build TM from completed Excel", menu)
        self.assertIn("2) Prepare translator file for new source", menu)
        self.assertIn("3) Fill source from translated file", menu)
        self.assertNotIn("Advanced tools", menu)
        self.assertNotIn("Extract source translation units", menu)

    def test_interactive_extract_creates_template_pack(self):
        from template_demo import main

        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "template_pack.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet2"
            ws.append(["中文", "英語", "target"])
            ws.append(["VIP0付费礼包", "VIP0 Paid Pack", ""])
            ws.append(["VIP10付费礼包", "VIP10 Paid Pack", ""])
            wb.save(input_path)

            answers = iter(
                [
                    "2",
                    f"'{input_path}'",
                    "",
                    "",
                    "-",
                    f'"{output_path}"',
                    "",
                    "n",
                ]
            )
            stdout = StringIO()
            with patch("builtins.input", side_effect=lambda _="": next(answers)), redirect_stdout(stdout):
                exit_code = main([])

            self.assertEqual(exit_code, 0)
            self.assertTrue(output_path.exists())
            self.assertNotIn("Optional: add source=target examples", stdout.getvalue())

            out = load_workbook(output_path, data_only=True)
            summary = {
                row[0].value: row[1].value
                for row in out["summary"].iter_rows(min_row=1, max_col=2)
            }
            self.assertEqual(
                summary,
                {
                    "total_source_rows": 2,
                    "total_translation_units": 1,
                    "already_filled_units": 0,
                    "already_filled_source_rows": 0,
                    "units_to_translate": 1,
                    "source_rows_to_translate": 2,
                },
            )


if __name__ == "__main__":
    unittest.main()
