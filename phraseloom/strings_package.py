from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import workbook_schema as schema
from .errors import ColumnNotFoundError, ConfigError, WorkbookFormatError
from .models import RowItem
from .tag_rules import (
    TagRules,
    default_tag_rules,
    normalized_tag_rules_hash,
    tag_rules_from_payload,
    tag_rules_payload,
)
from .workbook_io import (
    header_values,
    resolve_column,
    workbook_metadata,
)


EXISTING_TARGET_STATUS = "existing_target"
AUTO_PASSTHROUGH_STATUS = "auto_passthrough"
PASSTHROUGH_UNIT_TYPE = "passthrough"


@dataclass(frozen=True)
class StringUnit:
    string_id: str
    unit_type: str
    source: str
    items: tuple[RowItem, ...]
    sample_source: str
    context: object | None
    group_id: str = ""


@dataclass(frozen=True)
class CompletedString:
    item: RowItem
    status: str
    target: str
    context: object | None


def write_strings_package(
    input_path: Path,
    output_path: Path,
    units: list[StringUnit],
    completed_strings: list[CompletedString],
    *,
    tag_rules: TagRules,
    source_column: str | int,
    target_column: str | int,
    context_column: str | int | None,
    group_similar: bool,
    min_group_size: int,
) -> None:
    workbook = load_workbook(input_path)
    try:
        conflicts = sorted(_reserved_sheets() & set(workbook.sheetnames))
        if conflicts:
            raise ConfigError(
                "Source workbook uses reserved PhraseLoom sheet names: "
                + ", ".join(conflicts)
            )
        original_states = {
            worksheet.title: worksheet.sheet_state
            for worksheet in workbook.worksheets
        }
        original_active_sheet = workbook.active.title
        for worksheet in workbook.worksheets:
            worksheet.sheet_state = "hidden"

        strings = workbook.create_sheet(schema.STRINGS_SHEET)
        strings.append(schema.STRINGS_COLUMNS)
        for unit in units:
            strings.append(
                [
                    unit.string_id,
                    unit.group_id or None,
                    unit.source,
                    None,
                    unit.sample_source,
                    unit.context,
                    len(unit.items),
                ]
            )
        strings.column_dimensions["A"].hidden = True
        _style_strings_sheet(strings)

        completed = workbook.create_sheet(schema.COMPLETED_STRINGS_SHEET)
        completed.append(schema.COMPLETED_STRINGS_COLUMNS)
        for item in completed_strings:
            completed.append(
                [
                    item.item.row_number,
                    item.status,
                    item.item.raw_source,
                    item.target,
                    item.context,
                ]
            )
        completed.column_dimensions["A"].hidden = True
        _style_completed_sheet(completed)

        mapping = workbook.create_sheet(schema.STRINGS_MAP_SHEET)
        mapping.append(schema.STRINGS_MAP_COLUMNS)
        for unit in units:
            for item in unit.items:
                mapping.append(
                    [
                        item.row_number,
                        unit.string_id,
                        item.raw_source,
                        unit.unit_type,
                        unit.source,
                        json.dumps(item.match.values, ensure_ascii=False),
                    ]
                )
        passthrough_items = [
            completed.item
            for completed in completed_strings
            if completed.status == AUTO_PASSTHROUGH_STATUS
        ]
        for index, item in enumerate(passthrough_items, start=1):
            mapping.append(
                [
                    item.row_number,
                    f"P{index:04d}",
                    item.raw_source,
                    PASSTHROUGH_UNIT_TYPE,
                    item.source,
                    "{}",
                ]
            )
        mapping.sheet_state = "hidden"

        metadata = workbook.create_sheet(schema.METADATA_SHEET)
        metadata.append(schema.METADATA_COLUMNS)
        metadata_rows = {
            schema.SCHEMA_VERSION_KEY: schema.SCHEMA_VERSION,
            schema.WORKBOOK_KIND_KEY: schema.STRINGS_PACKAGE_KIND,
            schema.ORIGINAL_FILENAME_KEY: input_path.name,
            schema.ORIGINAL_SHEET_STATES_KEY: json.dumps(
                original_states,
                ensure_ascii=False,
            ),
            schema.ORIGINAL_ACTIVE_SHEET_KEY: original_active_sheet,
            schema.SOURCE_COLUMN_KEY: json.dumps(source_column, ensure_ascii=False),
            schema.TARGET_COLUMN_KEY: json.dumps(target_column, ensure_ascii=False),
            schema.CONTEXT_COLUMN_KEY: json.dumps(context_column, ensure_ascii=False),
            schema.GROUP_SIMILAR_KEY: group_similar,
            schema.MIN_GROUP_SIZE_KEY: min_group_size,
            schema.TAG_RULES_VERSION_KEY: tag_rules.version,
            schema.TAG_RULES_HASH_KEY: normalized_tag_rules_hash(tag_rules),
            schema.TAG_RULES_SOURCE_KEY: tag_rules.source,
            schema.TAG_RULES_PAYLOAD_KEY: tag_rules_payload(tag_rules),
        }
        for key, value in metadata_rows.items():
            metadata.append([key, value])
        metadata.sheet_state = "hidden"

        workbook.active = workbook.index(strings)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()


def load_strings_metadata(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        metadata = workbook_metadata(workbook)
        missing_sheets = {
            schema.STRINGS_SHEET,
            schema.COMPLETED_STRINGS_SHEET,
            schema.STRINGS_MAP_SHEET,
        } - set(workbook.sheetnames)
    finally:
        workbook.close()
    if metadata.get(schema.WORKBOOK_KIND_KEY) != schema.STRINGS_PACKAGE_KIND:
        raise WorkbookFormatError(
            f"Workbook {path} is not a PhraseLoom Strings workbook."
        )
    if missing_sheets:
        raise WorkbookFormatError(
            f"Strings workbook is missing sheets: {', '.join(sorted(missing_sheets))}"
        )
    return metadata


def metadata_column(
    metadata: dict[str, object],
    key: str,
) -> str | int | None:
    value = metadata.get(key)
    if value is None:
        raise WorkbookFormatError(f"Strings workbook is missing metadata: {key}")
    try:
        parsed = json.loads(str(value))
    except (TypeError, ValueError) as exc:
        raise WorkbookFormatError(
            f"Strings workbook has invalid metadata: {key}"
        ) from exc
    if parsed is None or isinstance(parsed, (str, int)):
        return parsed
    raise WorkbookFormatError(f"Strings workbook has invalid metadata: {key}")


def embedded_tag_rules(metadata: dict[str, object], path: Path) -> TagRules:
    payload = metadata.get(schema.TAG_RULES_PAYLOAD_KEY)
    if not payload:
        return default_tag_rules()
    return tag_rules_from_payload(str(payload), source=f"embedded:{path.name}")


def read_strings_rows(worksheet) -> dict[str, dict[str, object]]:
    headers = header_values(worksheet)
    required = {
        schema.STRING_ID_COLUMN,
        schema.SOURCE_COLUMN,
        schema.TARGET_COLUMN,
    }
    missing = required - set(headers)
    if missing:
        raise WorkbookFormatError(
            f"Strings sheet is missing columns: {', '.join(sorted(missing))}"
        )

    rows: dict[str, dict[str, object]] = {}
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        row = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        string_id = str(row.get(schema.STRING_ID_COLUMN) or "").strip()
        if not string_id:
            continue
        if string_id in rows:
            raise WorkbookFormatError(
                f"Duplicate string_id in strings sheet: {string_id}"
            )
        if not row.get(schema.SOURCE_COLUMN):
            raise WorkbookFormatError(
                f"Strings row {row_number} has an empty Source."
            )
        rows[string_id] = row
    return rows


def read_mapping_rows(worksheet) -> list[dict[str, object]]:
    headers = header_values(worksheet)
    missing = set(schema.STRINGS_MAP_COLUMNS) - set(headers)
    if missing:
        raise WorkbookFormatError(
            f"Strings mapping is missing columns: {', '.join(sorted(missing))}"
        )
    rows = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
        }
        if row[schema.ROW_NUMBER_COLUMN] is not None:
            rows.append(row)
    return rows


def mapping_variables(map_row: dict[str, object]) -> dict[str, str]:
    raw_values = map_row.get(schema.VARIABLE_VALUES_COLUMN)
    if raw_values in (None, ""):
        return {}
    try:
        values = json.loads(str(raw_values))
    except (TypeError, ValueError) as exc:
        raise WorkbookFormatError(
            "Strings mapping has invalid variable values."
        ) from exc
    if not isinstance(values, dict):
        raise WorkbookFormatError("Strings mapping has invalid variable values.")
    return {str(key): str(value) for key, value in values.items()}


def original_worksheet(workbook):
    for worksheet in workbook.worksheets:
        if worksheet.title not in _reserved_sheets():
            return worksheet
    raise WorkbookFormatError(
        "Strings workbook does not contain the original worksheet."
    )


def restore_original_sheets(
    workbook,
    metadata: dict[str, object],
) -> None:
    try:
        original_states = json.loads(
            str(metadata.get(schema.ORIGINAL_SHEET_STATES_KEY) or "{}")
        )
    except json.JSONDecodeError as exc:
        raise WorkbookFormatError(
            "Strings workbook has invalid sheet-state metadata."
        ) from exc
    for sheet_name in _reserved_sheets():
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
    for worksheet in workbook.worksheets:
        worksheet.sheet_state = str(
            original_states.get(worksheet.title, "visible")
        )
    if not any(
        worksheet.sheet_state == "visible"
        for worksheet in workbook.worksheets
    ):
        workbook.worksheets[0].sheet_state = "visible"
    active_sheet = str(metadata.get(schema.ORIGINAL_ACTIVE_SHEET_KEY) or "")
    if active_sheet in workbook.sheetnames:
        workbook.active = workbook.index(workbook[active_sheet])


def resolve_or_create_column(worksheet, column: str | int) -> int:
    try:
        return resolve_column(worksheet, column)
    except ColumnNotFoundError:
        if isinstance(column, int) or str(column).isdigit():
            raise
        index = worksheet.max_column + 1
        worksheet.cell(row=1, column=index).value = str(column)
        return index


def write_restore_issues(
    output_path: Path,
    issues: list[dict[str, object]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "restore_issues"
    headers = ["string_id", "source", "target", "row_number", "issue"]
    worksheet.append(headers)
    for issue in issues:
        worksheet.append([issue.get(header) for header in headers])
    _style_issue_sheet(worksheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def _reserved_sheets() -> set[str]:
    return {
        schema.STRINGS_SHEET,
        schema.COMPLETED_STRINGS_SHEET,
        schema.STRINGS_MAP_SHEET,
        schema.METADATA_SHEET,
    }


def _style_strings_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="243447")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "B2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for letter, width in {
        "B": 12,
        "C": 48,
        "D": 48,
        "E": 48,
        "F": 36,
        "G": 12,
    }.items():
        worksheet.column_dimensions[letter].width = width

    group_fills = (
        PatternFill("solid", fgColor="EEF4FF"),
        PatternFill("solid", fgColor="F3F8F2"),
    )
    group_colors: dict[str, PatternFill] = {}
    for row_number in range(2, worksheet.max_row + 1):
        group_id = str(worksheet.cell(row=row_number, column=2).value or "")
        if group_id:
            if group_id not in group_colors:
                group_colors[group_id] = group_fills[
                    len(group_colors) % len(group_fills)
                ]
            for column in range(2, worksheet.max_column + 1):
                worksheet.cell(row=row_number, column=column).fill = (
                    group_colors[group_id]
                )
        for column in range(2, worksheet.max_column + 1):
            worksheet.cell(row=row_number, column=column).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def _style_completed_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="3F5F4F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "B2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for letter, width in {"B": 20, "C": 48, "D": 48, "E": 36}.items():
        worksheet.column_dimensions[letter].width = width
    completed_fill = PatternFill("solid", fgColor="F1F7F3")
    for row_number in range(2, worksheet.max_row + 1):
        for column in range(2, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_number, column=column)
            cell.fill = completed_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_issue_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="243447")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        letter = get_column_letter(column[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column
        )
        worksheet.column_dimensions[letter].width = min(
            max(max_length + 2, 12),
            60,
        )
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


__all__ = [
    "AUTO_PASSTHROUGH_STATUS",
    "CompletedString",
    "EXISTING_TARGET_STATUS",
    "PASSTHROUGH_UNIT_TYPE",
    "StringUnit",
    "embedded_tag_rules",
    "load_strings_metadata",
    "mapping_variables",
    "metadata_column",
    "original_worksheet",
    "read_mapping_rows",
    "read_strings_rows",
    "resolve_or_create_column",
    "restore_original_sheets",
    "write_restore_issues",
    "write_strings_package",
]
