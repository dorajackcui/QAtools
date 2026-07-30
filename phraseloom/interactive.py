from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .entity_workflow import (
    default_entity_filled_pack_output_path,
    default_entity_memory_output_path,
    default_entity_merged_todo_output_path,
    default_entity_pack_output_path,
    extract_entity_memory_workbook,
    fill_entity_pack_workbook,
    merge_entity_pack_workbook,
    prepare_entity_pack_workbook,
)
from .errors import ColumnNotFoundError
from .excel_io import (
    _default_tm_output_path,
    _read_headers,
    _resolve_column,
)
from .workflow import (
    fill_translation_package,
    generate_tm_pairs,
    prepare_translation_package,
)


def run_interactive() -> int:
    print("Localization Workflow")
    print()
    print("1) Build TM from completed Excel")
    print("2) Prepare translator file for new source")
    print("3) Fill source from translated file")
    print("a) Advanced tools")
    print("q) Quit")

    action = _prompt_text("Choose step", default="2").lower()
    if action in {"q", "quit", "exit"}:
        print("Bye.")
        return 0
    if action in {"1", "tm", "tm-extract", "extract-tm", "build"}:
        return _interactive_tm_extract()
    if action in {"2", "extract", "prepare", "p"}:
        return _interactive_extract()
    if action in {"3", "fill", "f"}:
        return _interactive_fill()
    if action in {"a", "advanced", "tools"}:
        return run_advanced_interactive()
    # Keep the old entry working for existing operator notes and scripts.
    if action in {"4", "entity", "entity-workflow", "e"}:
        return run_entity_interactive(back_returns_to_main=True)

    print(f"Unknown step: {action}")
    return 2


def run_advanced_interactive() -> int:
    print("Advanced Tools")
    print()
    print("1) Entity workflow")
    print("b) Back")
    print("q) Quit")

    action = _prompt_text("Choose advanced tool", default="1").lower()
    if action in {"q", "quit", "exit"}:
        print("Bye.")
        return 0
    if action in {"b", "back"}:
        return run_interactive()
    if action in {"1", "entity", "e"}:
        return run_entity_interactive(back_returns_to_main=True)
    print(f"Unknown advanced tool: {action}")
    return 2


def _interactive_tm_extract() -> int:
    input_path = _user_path(_prompt_text("Completed Excel path", required=True))
    source_col = _detect_or_prompt_source_column(input_path)
    target_col = _detect_or_prompt_target_column(
        input_path,
        source_col=source_col,
        require_existing=True,
    )
    output_path = _default_tm_output_path(input_path)

    stats = generate_tm_pairs(
        input_path,
        output_path,
        source_col=source_col,
        target_col=target_col,
        min_group_size=2,
    )
    _display_tm_stats(output_path, stats)
    return 0


def _interactive_extract() -> int:
    input_path = _user_path(_prompt_text("New source Excel path", required=True))
    source_col = _detect_or_prompt_source_column(input_path)
    target_col = _detect_or_prompt_target_column(
        input_path,
        source_col=source_col,
        require_existing=False,
    )
    tm_workbook_text = _prompt_text("TM workbook path (blank for none)")
    tm_workbook = _user_path(tm_workbook_text) if tm_workbook_text else None
    use_existing_targets = (
        _prompt_yes_no("Use current target values as prefill", default=True)
        if _column_has_values(input_path, target_col)
        else False
    )

    stats = prepare_translation_package(
        input_path,
        source_col=source_col,
        target_col=target_col,
        tm_workbook=tm_workbook,
        min_group_size=2,
        use_existing_targets=use_existing_targets,
    )
    _display_prepare_stats(stats)
    return 0


def _interactive_fill() -> int:
    package_path = _user_path(
        _prompt_text("Translated to_translate workbook", required=True)
    )
    stats = fill_translation_package(package_path)
    _display_fill_stats(stats)
    return 0


def _detect_or_prompt_source_column(input_path: Path) -> str:
    headers = _read_headers(input_path)
    detected = _case_insensitive_header(headers, "source")
    if detected is not None:
        return detected
    return _prompt_text(
        f"Source column (available: {', '.join(headers)})",
        required=True,
    )


def _detect_or_prompt_target_column(
    input_path: Path,
    *,
    source_col: str,
    require_existing: bool,
) -> str:
    headers = _read_headers(input_path)
    detected = _case_insensitive_header(headers, "target")
    if detected is not None:
        return detected
    candidates = [header for header in headers if header.lower() != source_col.lower()]
    if not require_existing and not candidates:
        return "target"
    default = candidates[-1] if len(candidates) == 1 else None
    return _prompt_text(
        f"Target column (available: {', '.join(headers)})",
        default=default,
        required=default is None,
    )


def _case_insensitive_header(headers: list[str], wanted: str) -> str | None:
    wanted_lower = wanted.lower()
    return next((header for header in headers if header.lower() == wanted_lower), None)


def _column_has_values(input_path: Path, column: str | int) -> bool:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        try:
            column_index = _resolve_column(ws, column)
        except ColumnNotFoundError:
            return False
        return any(
            value is not None and str(value).strip() != ""
            for (value,) in ws.iter_rows(
                min_row=2,
                min_col=column_index,
                max_col=column_index,
                values_only=True,
            )
        )
    finally:
        wb.close()


def run_entity_interactive(*, back_returns_to_main: bool = False) -> int:
    print("Entity Workflow")
    print()
    print("1) Build entity memory from TM reusable units")
    print("2) Prepare source entity pack")
    print("3) Fill completed entity pack")
    print("4) Merge filled entity pack back to translator todo")
    print("b) Back")
    print("q) Quit")

    action = _prompt_text("Choose entity step", default="2").lower()
    if action in {"q", "quit", "exit"}:
        print("Bye.")
        return 0
    if action in {"b", "back"}:
        if back_returns_to_main:
            return run_interactive()
        print("Bye.")
        return 0
    if action in {"1", "tm", "entity-tm", "memory"}:
        return _interactive_entity_tm()
    if action in {"2", "prepare", "pack"}:
        return _interactive_entity_prepare()
    if action in {"3", "fill", "fill-pack"}:
        return _interactive_entity_fill_pack()
    if action in {"4", "merge", "merge-pack"}:
        return _interactive_entity_merge_pack()

    print(f"Unknown entity step: {action}")
    return 2


def _interactive_entity_tm() -> int:
    input_path = _user_path(_prompt_text("TM reusable units path", required=True))
    output_path = _user_path(
        _prompt_text(
            "Output entity memory workbook",
            default=str(default_entity_memory_output_path(input_path)),
        )
    )
    min_group_size = _prompt_int(
        "Minimum variants for a reusable entity structure",
        default=3,
    )

    stats = extract_entity_memory_workbook(
        input_path,
        output_path,
        min_group_size=min_group_size,
    )
    _display_entity_extract_tm_stats(stats)
    return 0


def _interactive_entity_prepare() -> int:
    input_path = _user_path(_prompt_text("Translator todo path", required=True))
    tm_text = _prompt_text("Entity memory path (- for none)", default="-")
    tm_path = _user_path(tm_text) if _normalize_optional_column(tm_text) is not None else None
    output_path = _user_path(
        _prompt_text(
            "Output entity pack workbook",
            default=str(default_entity_pack_output_path(input_path)),
        )
    )
    min_group_size = _prompt_int(
        "Minimum variants for a reusable entity structure",
        default=3,
    )

    stats = prepare_entity_pack_workbook(
        input_path,
        output_path,
        tm_path=tm_path,
        min_group_size=min_group_size,
    )
    _display_entity_prepare_stats(stats)
    return 0


def _interactive_entity_fill_pack() -> int:
    input_path = _user_path(_prompt_text("Source entity pack path", required=True))
    output_path = _user_path(
        _prompt_text(
            "Output filled entity pack workbook",
            default=str(default_entity_filled_pack_output_path(input_path)),
        )
    )

    stats = fill_entity_pack_workbook(input_path, output_path)
    _display_entity_fill_stats(stats)
    return 0


def _interactive_entity_merge_pack() -> int:
    input_path = _user_path(_prompt_text("Filled source entity pack path", required=True))
    output_path = _user_path(
        _prompt_text(
            "Output merged translator todo workbook",
            default=str(default_entity_merged_todo_output_path(input_path)),
        )
    )

    stats = merge_entity_pack_workbook(input_path, output_path)
    _display_entity_merge_stats(stats)
    return 0


def _prompt_text(prompt: str, *, default: str | None = None, required: bool = False) -> str:
    suffix = f" [{default}]" if default is not None else ""
    while True:
        value = input(f"{prompt}{suffix}: ").strip()
        if value:
            return value
        if default is not None:
            return default
        if not required:
            return ""
        print("This value is required.")


def _prompt_int(prompt: str, *, default: int) -> int:
    while True:
        raw = _prompt_text(prompt, default=str(default))
        try:
            value = int(raw)
        except ValueError:
            print("Please enter a number.")
            continue
        if value < 1:
            print("Please enter a number greater than 0.")
            continue
        return value


def _prompt_yes_no(prompt: str, *, default: bool) -> bool:
    default_text = "Y/n" if default else "y/N"
    while True:
        raw = input(f"{prompt} [{default_text}]: ").strip().lower()
        if not raw:
            return default
        if raw in {"y", "yes"}:
            return True
        if raw in {"n", "no"}:
            return False
        print("Please answer y or n.")


def _user_path(value: str | Path) -> Path:
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1].strip()
    return Path(text).expanduser()


def _normalize_optional_column(value: str | int | None) -> str | int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    normalized = value.strip()
    if normalized.lower() in {"", "-", "none", "no", "skip"}:
        return None
    return normalized


def _display_stats(output: Path, stats: dict[str, int]) -> None:
    print(f"Wrote: {output}")
    if "to_translate_path" in stats:
        print(f"To-translate workbook: {stats['to_translate_path']}")
    print(f"Units to translate: {stats['new_translation_unit_count']}")
    print(f"Source rows to translate: {stats['new_source_segment_count']}")
    print(f"Already filled units: {stats['prefilled_translation_unit_count']}")
    print(f"Already filled source rows: {stats['autofilled_count']}")
    print(f"Total translation units: {stats['translation_unit_count']}")
    print(f"Total source rows: {stats['row_count']}")


def _display_prepare_stats(stats: dict[str, int | str]) -> None:
    print(f"Translator workbook: {stats['to_translate_path']}")
    print(f"Units to translate: {stats['new_translation_unit_count']}")
    print(f"Prefilled units: {stats['prefilled_translation_unit_count']}")
    print(f"Source rows covered: {stats['row_count']}")


def _display_fill_stats(stats: dict[str, int | str]) -> None:
    print(f"Filled workbook: {stats['output_path']}")
    print(f"Filled source rows: {stats['autofilled_count']}")
    unfilled = int(stats["row_count"]) - int(stats["autofilled_count"])
    print(f"Unfilled source rows: {unfilled}")
    if "audit_output_path" in stats:
        print(f"Review workbook: {stats['audit_output_path']}")


def _display_tm_stats(output: Path, stats: dict[str, int]) -> None:
    print(f"Wrote: {output}")
    print(f"TM source segments: {stats['row_count']}")
    print(f"Unique source segments: {stats['unique_source_segments']}")
    print(f"Duplicate source segments: {stats['duplicate_source_segments']}")
    print(f"TM pairs: {stats['tm_pair_count']}")
    print(f"Template pairs: {stats['template_pair_count']}")
    print(f"Segment pairs: {stats['segment_pair_count']}")


def _display_entity_prepare_stats(stats: dict[str, int | str]) -> None:
    print(f"Wrote: {stats['output_path']}")
    print(f"Related units: {stats['related_unit_count']}")
    print(f"Non-related units: {stats['non_related_unit_count']}")
    print(f"Entity structures: {stats['entity_structure_count']}")
    print(f"Entity terms: {stats['entity_term_count']}")
    print(f"Prefilled structures: {stats['prefilled_structure_count']}")
    print(f"Prefilled terms: {stats['prefilled_term_count']}")


def _display_entity_extract_tm_stats(stats: dict[str, int | str]) -> None:
    print(f"Wrote: {stats['output_path']}")
    print(f"Entity structures: {stats['entity_structure_count']}")
    print(f"Entity terms: {stats['entity_term_count']}")


def _display_entity_fill_stats(stats: dict[str, int | str]) -> None:
    print(f"Wrote: {stats['output_path']}")
    print(f"Filled entity units: {stats['filled_entity_unit_count']}")


def _display_entity_merge_stats(stats: dict[str, int | str]) -> None:
    print(f"Wrote: {stats['output_path']}")
    print(f"Merged units: {stats['merged_unit_count']}")


__all__ = [
    "run_interactive",
    "run_advanced_interactive",
    "run_entity_interactive",
    "_interactive_tm_extract",
    "_interactive_extract",
    "_interactive_fill",
    "_interactive_entity_tm",
    "_interactive_entity_prepare",
    "_interactive_entity_fill_pack",
    "_interactive_entity_merge_pack",
    "_prompt_text",
    "_prompt_int",
    "_prompt_yes_no",
    "_user_path",
    "_normalize_optional_column",
    "_column_has_values",
    "_detect_or_prompt_source_column",
    "_detect_or_prompt_target_column",
    "_display_fill_stats",
    "_display_prepare_stats",
    "_display_entity_prepare_stats",
    "_display_entity_extract_tm_stats",
    "_display_entity_fill_stats",
    "_display_entity_merge_stats",
]
