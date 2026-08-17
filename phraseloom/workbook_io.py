from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .errors import ColumnNotFoundError
from .models import RowItem
from .tag_engine import extract_tags
from .tag_rules import TagRules, default_tag_rules
from .template_engine import parse_template


def read_source_rows(
    input_path: Path,
    source_column: str | int,
    target_column: str | int | None,
    *,
    tag_rules: TagRules | None = None,
    split_lines: bool = True,
) -> list[RowItem]:
    active_rules = tag_rules or default_tag_rules()
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        source_index = resolve_column(worksheet, source_column)
        target_index = (
            resolve_column(worksheet, target_column)
            if target_column is not None
            else None
        )

        rows: list[RowItem] = []
        seen_source = False
        consecutive_blanks = 0
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            source_value = _cell_value(values, source_index)
            if source_value is None or not str(source_value).strip():
                if seen_source:
                    consecutive_blanks += 1
                    if consecutive_blanks >= 1000:
                        break
                continue

            seen_source = True
            consecutive_blanks = 0
            raw_source = str(source_value)
            target_value = _cell_value(values, target_index)
            raw_target = "" if target_value is None else str(target_value).strip()
            segments = _protected_source_segments(
                raw_source,
                split_lines=split_lines and not raw_target,
                tag_rules=active_rules,
            )
            segment_count = len(segments)
            for segment_index, (
                raw_segment,
                protected_source,
                prefix,
                suffix,
            ) in enumerate(
                segments,
                start=1,
            ):
                rows.append(
                    RowItem(
                        row_number=row_number,
                        source=protected_source,
                        match=parse_template(protected_source),
                        original_values=tuple(values),
                        raw_source=raw_source,
                        raw_segment=raw_segment,
                        raw_existing_target=raw_target,
                        segment_index=segment_index,
                        segment_count=segment_count,
                        segment_prefix=prefix,
                        segment_suffix=suffix,
                    )
                )
        return rows
    finally:
        workbook.close()


def read_headers(input_path: Path) -> list[str]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        return header_values(workbook.worksheets[0], fallback=True)
    finally:
        workbook.close()


def resolve_column(worksheet, column: str | int | None) -> int:
    if column is None:
        raise ColumnNotFoundError(column, header_values(worksheet, fallback=True))
    if isinstance(column, int) or str(column).isdigit():
        return int(column)

    wanted = str(column).strip().lower()
    for index, cell in enumerate(worksheet[1], start=1):
        value = "" if cell.value is None else str(cell.value).strip()
        if value.lower() == wanted:
            return getattr(cell, "column", index)
    raise ColumnNotFoundError(column, header_values(worksheet, fallback=True))


def header_values(worksheet, *, fallback: bool = False) -> list[str]:
    headers: list[str] = []
    for index, cell in enumerate(worksheet[1], start=1):
        value = "" if cell.value is None else str(cell.value).strip()
        headers.append(value or (f"column_{index}" if fallback else ""))
    return headers


def workbook_metadata(workbook) -> dict[str, object]:
    from . import workbook_schema as schema

    if schema.METADATA_SHEET not in workbook.sheetnames:
        return {}
    worksheet = workbook[schema.METADATA_SHEET]
    return {
        row[0]: row[1]
        for row in worksheet.iter_rows(min_row=2, max_col=2, values_only=True)
        if row and row[0] is not None
    }


def _cell_value(
    row: tuple[object, ...],
    one_based_index: int | None,
) -> object | None:
    if one_based_index is None:
        return None
    zero_based = one_based_index - 1
    return row[zero_based] if zero_based < len(row) else None


def _split_source_segments(source: str) -> list[tuple[str, str, str]]:
    """Split a cell into non-empty line segments with reversible whitespace."""
    parts = re.split(r"(\r\n|\r|\n)", source)
    segments: list[list[str]] = []
    pending_prefix = ""
    for position in range(0, len(parts), 2):
        line = parts[position]
        separator = parts[position + 1] if position + 1 < len(parts) else ""
        raw_segment = line.strip()
        if not raw_segment:
            structure = line + separator
            if segments:
                segments[-1][2] += structure
            else:
                pending_prefix += structure
            continue

        core_start = len(line) - len(line.lstrip())
        core_end = len(line.rstrip())
        prefix = pending_prefix + line[:core_start]
        suffix = line[core_end:] + separator
        pending_prefix = ""
        segments.append([raw_segment, prefix, suffix])

    return [(segment, prefix, suffix) for segment, prefix, suffix in segments]


def _protected_source_segments(
    raw_source: str,
    *,
    split_lines: bool,
    tag_rules: TagRules,
) -> list[tuple[str, str, str, str]]:
    extraction = extract_tags(raw_source, rules=tag_rules)
    if not split_lines:
        leading_size = len(raw_source) - len(raw_source.lstrip())
        trailing_start = len(raw_source.rstrip())
        return [
            (
                raw_source.strip(),
                extraction.text.strip(),
                raw_source[:leading_size],
                raw_source[trailing_start:],
            )
        ]

    raw_segments = _split_source_segments(raw_source)
    protected_segments = _split_source_segments(extraction.text)
    if len(raw_segments) == len(protected_segments):
        return [
            (
                raw_segment,
                protected_segment,
                prefix,
                suffix,
            )
            for (
                raw_segment,
                prefix,
                suffix,
            ), (
                protected_segment,
                _protected_prefix,
                _protected_suffix,
            ) in zip(raw_segments, protected_segments, strict=True)
        ]

    # A protected atomic span can itself contain a line break. Splitting inside
    # it would expose partial tag syntax, so keep that uncommon cell as one unit.
    leading_size = len(raw_source) - len(raw_source.lstrip())
    trailing_start = len(raw_source.rstrip())
    return [
        (
            raw_source.strip(),
            extraction.text.strip(),
            raw_source[:leading_size],
            raw_source[trailing_start:],
        )
    ]


__all__ = [
    "header_values",
    "read_headers",
    "read_source_rows",
    "resolve_column",
    "workbook_metadata",
]
