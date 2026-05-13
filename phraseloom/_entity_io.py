from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from . import workbook_schema as schema
from ._entity_types import (
    ENTITY_PACK_STRUCTURE_COLUMNS,
    ENTITY_PACK_TERM_COLUMNS,
    ENTITY_SOURCE_MAP_COLUMNS,
    ENTITY_STRUCTURE_COLUMNS,
    ENTITY_TERM_COLUMNS,
    UnitRow,
)
from .errors import WorkflowError
from .excel_io import _header_values, _style_sheet


def _read_unit_rows(path: Path, sheet_name: str) -> tuple[list[UnitRow], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        headers = _header_values(ws)
        rows = []
        for sequence_index, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=1,
        ):
            values = {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
            original_index = int(
                values.get(schema.ORIGINAL_INDEX_COLUMN) or sequence_index
            )
            if values.get(schema.SOURCE_UNIT_COLUMN):
                rows.append(UnitRow(original_index, values))
        return rows, headers
    finally:
        wb.close()


def _parse_original_index(value: object, *, sheet_name: str, row_number: int) -> int:
    if value in (None, ""):
        raise WorkflowError(
            f"Workbook sheet {sheet_name!r} row {row_number} is missing required "
            f"{schema.ORIGINAL_INDEX_COLUMN}"
        )
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise WorkflowError(
            f"Workbook sheet {sheet_name!r} row {row_number} has invalid "
            f"{schema.ORIGINAL_INDEX_COLUMN}: {value!r}"
        ) from exc


def _read_pack_unit_rows(
    path: Path,
    preferred_sheet_name: str,
    fallback_sheet_name: str | None = None,
) -> tuple[list[UnitRow], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = preferred_sheet_name
        if sheet_name not in wb.sheetnames:
            if fallback_sheet_name is None or fallback_sheet_name not in wb.sheetnames:
                raise WorkflowError(
                    f"Workbook is missing required sheet: {preferred_sheet_name}"
                )
            sheet_name = fallback_sheet_name
        ws = wb[sheet_name]
        headers = _header_values(ws)
        if schema.ORIGINAL_INDEX_COLUMN not in headers:
            raise WorkflowError(
                f"Workbook sheet {sheet_name!r} is missing required column: "
                f"{schema.ORIGINAL_INDEX_COLUMN}"
            )
        rows: list[UnitRow] = []
        for sequence_index, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=1,
        ):
            values = {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
            if values.get(schema.SOURCE_UNIT_COLUMN):
                original_index_value = values.get(schema.ORIGINAL_INDEX_COLUMN)
                original_index = _parse_original_index(
                    original_index_value,
                    sheet_name=sheet_name,
                    row_number=sequence_index + 1,
                )
                rows.append(UnitRow(original_index, values))
        return rows, headers
    finally:
        wb.close()


def _read_tm_pair_rows(path: Path) -> list[UnitRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[schema.TM_PAIRS_SHEET]
        headers = _header_values(ws)
        rows: list[UnitRow] = []
        for sequence_index, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=1,
        ):
            values = {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
            if values.get(schema.SOURCE_UNIT_COLUMN) and values.get(
                schema.TARGET_UNIT_COLUMN
            ):
                values[schema.UNIT_ID_COLUMN] = values.get(schema.TM_ID_COLUMN)
                rows.append(UnitRow(sequence_index, values))
        return rows
    finally:
        wb.close()


def _worksheet_by_name(wb, names: list[str]):
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    expected = " or ".join(names)
    raise WorkflowError(f"Workbook is missing required sheet: {expected}")


def _write_entity_related_workbook(
    output_path: Path,
    input_path: Path,
    headers: list[str],
    rows: list[UnitRow],
    structures: list[dict[str, object]],
    terms: list[dict[str, object]],
    source_map: list[dict[str, object]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = schema.TO_TRANSLATE_SHEET
    _append_unit_rows(ws, headers, rows, include_original_index=True)
    _append_dict_sheet(
        wb.create_sheet(schema.ENTITY_STRUCTURES_SHEET),
        ENTITY_STRUCTURE_COLUMNS,
        structures,
    )
    _append_dict_sheet(
        wb.create_sheet(schema.ENTITY_TERMS_SHEET),
        ENTITY_TERM_COLUMNS,
        terms,
    )
    _append_dict_sheet(
        wb.create_sheet(schema.ENTITY_SOURCE_MAP_SHEET),
        ENTITY_SOURCE_MAP_COLUMNS,
        source_map,
    )
    _copy_support_sheets(input_path, wb, exclude={schema.TO_TRANSLATE_SHEET})
    _save_workbook(wb, output_path)


def _build_entity_pack_workbook(
    input_path: Path,
    headers: list[str],
    entity_rows: list[UnitRow],
    non_entity_rows: list[UnitRow],
    structures: list[dict[str, object]],
    terms: list[dict[str, object]],
    source_map: list[dict[str, object]],
):
    wb = Workbook()
    related_ws = wb.active
    related_ws.title = schema.RELATED_UNITS_SHEET
    _append_unit_rows(related_ws, headers, entity_rows, include_original_index=True)
    _hide_original_index_column(related_ws)
    non_related_ws = wb.create_sheet(schema.NON_RELATED_UNITS_SHEET)
    _append_unit_rows(
        non_related_ws,
        headers,
        non_entity_rows,
        include_original_index=True,
    )
    _hide_original_index_column(non_related_ws)
    structures_ws = wb.create_sheet(schema.ENTITY_STRUCTURES_SHEET)
    _append_dict_sheet(
        structures_ws,
        ENTITY_PACK_STRUCTURE_COLUMNS,
        structures,
    )
    _hide_columns_by_header(
        structures_ws,
        [schema.CONFIDENCE_COLUMN, schema.RISK_COLUMN],
    )
    _append_dict_sheet(
        wb.create_sheet(schema.ENTITY_TERMS_SHEET),
        ENTITY_PACK_TERM_COLUMNS,
        terms,
    )
    entity_map_ws = wb.create_sheet(schema.ENTITY_MAP_SHEET)
    _append_dict_sheet(entity_map_ws, ENTITY_SOURCE_MAP_COLUMNS, source_map)
    entity_map_ws.sheet_state = "hidden"
    _copy_support_sheets(input_path, wb, exclude={schema.TO_TRANSLATE_SHEET})
    if schema.METADATA_SHEET not in wb.sheetnames:
        metadata_ws = wb.create_sheet(schema.METADATA_SHEET)
        metadata_ws.append(schema.METADATA_COLUMNS)
        metadata_ws.append([schema.SCHEMA_VERSION_KEY, schema.SCHEMA_VERSION])
    wb[schema.METADATA_SHEET].sheet_state = "hidden"
    return wb


def _write_non_entity_workbook(
    output_path: Path,
    input_path: Path,
    headers: list[str],
    rows: list[UnitRow],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = schema.TO_TRANSLATE_SHEET
    _append_unit_rows(ws, headers, rows, include_original_index=True)
    _copy_support_sheets(input_path, wb, exclude={schema.TO_TRANSLATE_SHEET})
    _save_workbook(wb, output_path)


def _write_entity_tm_workbook(
    output_path: Path,
    structures: list[dict[str, object]],
    terms: list[dict[str, object]],
) -> None:
    wb = Workbook()
    structures_ws = wb.active
    structures_ws.title = schema.ENTITY_STRUCTURES_SHEET
    _append_dict_sheet(structures_ws, ENTITY_STRUCTURE_COLUMNS, structures)
    _append_dict_sheet(
        wb.create_sheet(schema.ENTITY_TERMS_SHEET),
        ENTITY_TERM_COLUMNS,
        terms,
    )
    _save_workbook(wb, output_path)


def _append_unit_rows(
    ws,
    headers: list[str],
    rows: list[UnitRow],
    *,
    include_original_index: bool,
) -> None:
    output_headers = (
        [schema.ORIGINAL_INDEX_COLUMN] + headers
        if include_original_index and schema.ORIGINAL_INDEX_COLUMN not in headers
        else headers
    )
    ws.append(output_headers)
    for row in rows:
        values = []
        for header in output_headers:
            if header == schema.ORIGINAL_INDEX_COLUMN:
                values.append(row.original_index)
            else:
                values.append(row.values.get(header))
        ws.append(values)


def _hide_original_index_column(ws) -> None:
    headers = _header_values(ws)
    if schema.ORIGINAL_INDEX_COLUMN in headers:
        column_letter = get_column_letter(
            headers.index(schema.ORIGINAL_INDEX_COLUMN) + 1
        )
        ws.column_dimensions[column_letter].hidden = True


def _hide_columns_by_header(ws, headers_to_hide: list[str]) -> None:
    headers = _header_values(ws)
    for header in headers_to_hide:
        if header not in headers:
            continue
        column_letter = get_column_letter(headers.index(header) + 1)
        ws.column_dimensions[column_letter].hidden = True


def _append_dict_sheet(ws, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _copy_support_sheets(input_path: Path, target_wb, *, exclude: set[str]) -> None:
    source_wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        for source_ws in source_wb.worksheets:
            if source_ws.title in exclude:
                continue
            target_ws = target_wb.create_sheet(source_ws.title)
            target_ws.sheet_state = source_ws.sheet_state
            for row in source_ws.iter_rows(values_only=True):
                target_ws.append(row)
    finally:
        source_wb.close()


def _save_workbook(wb, output_path: Path) -> None:
    for ws in wb.worksheets:
        _style_sheet(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _replace_workbook_atomically(wb, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_file = tempfile.NamedTemporaryFile(
        delete=False,
        dir=output_path.parent,
        suffix=output_path.suffix,
    )
    temp_path = Path(temp_file.name)
    temp_file.close()
    try:
        _save_workbook(wb, temp_path)
        temp_path.replace(output_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
