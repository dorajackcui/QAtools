from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from . import workbook_schema as schema
from .cleaning import build_clean_string_units
from .errors import ColumnNotFoundError, ConfigError, WorkbookFormatError
from .models import RowItem
from .string_cluster import cluster_similar_strings
from .strings_package import (
    AUTO_PASSTHROUGH_STATUS,
    EXISTING_TARGET_STATUS,
    PASSTHROUGH_UNIT_TYPE,
    CompletedString,
    StringUnit,
    embedded_tag_rules,
    load_strings_metadata,
    mapping_variables,
    metadata_column,
    original_worksheet,
    read_mapping_rows,
    read_strings_rows,
    resolve_or_create_column,
    restore_original_sheets,
    write_restore_issues,
    write_strings_package,
)
from .tag_engine import (
    extract_tags,
    restore_tags,
    serialize_known_tags,
    validate_tag_placeholders,
)
from .tag_rules import TagRules, load_tag_rules
from .template_engine import (
    PLACEHOLDER_RE,
    apply_target_template,
    is_non_translatable_segment,
)
from .workbook_io import read_headers, read_source_rows, resolve_column


def default_strings_output_path(input_path: str | Path) -> Path:
    input_path = Path(input_path)
    return input_path.with_name(f"{input_path.stem}_strings.xlsx")


def export_strings_workbook(
    input_path: str | Path,
    output_path: str | Path | None = None,
    *,
    source_col: str | int = "source",
    target_col: str | int = "target",
    context_col: str | int | None = None,
    group_similar: bool = False,
    min_group_size: int = 3,
    tag_config: str | Path | None = None,
) -> dict[str, int | str]:
    input_path = Path(input_path)
    output_path = (
        Path(output_path)
        if output_path is not None
        else default_strings_output_path(input_path)
    )
    _require_distinct_paths(input_path, output_path)
    if group_similar and min_group_size < 2:
        raise ConfigError("Similar-string groups need at least 2 members.")

    tag_rules = load_tag_rules(tag_config)
    rows = _read_rows_with_optional_target(
        input_path,
        source_col=source_col,
        target_col=target_col,
        tag_rules=tag_rules,
    )
    context_index = _context_column_index(input_path, context_col)
    completed_strings = [
        CompletedString(
            item=row,
            status=(
                EXISTING_TARGET_STATUS
                if row.raw_existing_target
                else AUTO_PASSTHROUGH_STATUS
            ),
            target=row.raw_existing_target or row.raw_segment,
            context=_row_context(row, context_index),
        )
        for row in rows
        if row.raw_existing_target or is_non_translatable_segment(row.source)
    ]
    pending_rows = [
        row
        for row in rows
        if not row.raw_existing_target
        and not is_non_translatable_segment(row.source)
    ]

    units = [
        StringUnit(
            string_id=unit.unit_id,
            unit_type=unit.unit_type,
            source=unit.source,
            items=unit.items,
            sample_source=_sample_source(unit.items),
            context=_context(unit.items, context_index),
        )
        for unit in build_clean_string_units(
            pending_rows,
            min_template_variants=2,
        )
    ]
    units.sort(key=_source_order)
    grouped_units, group_count, grouped_count = _apply_optional_grouping(
        units,
        enabled=group_similar,
        min_group_size=min_group_size,
    )

    write_strings_package(
        input_path,
        output_path,
        grouped_units,
        completed_strings,
        tag_rules=tag_rules,
        source_column=source_col,
        target_column=target_col,
        context_column=context_col,
        group_similar=group_similar,
        min_group_size=min_group_size,
    )

    source_row_numbers = {row.row_number for row in rows}
    pending_row_numbers = {row.row_number for row in pending_rows}
    existing_target_count = len(
        {row.row_number for row in rows if row.raw_existing_target}
    )
    passthrough_count = sum(
        1
        for row in rows
        if not row.raw_existing_target
        and is_non_translatable_segment(row.source)
    )
    rows_by_number: dict[int, list[RowItem]] = {}
    for row in rows:
        rows_by_number.setdefault(row.row_number, []).append(row)
    auto_completed_row_numbers = {
        row_number
        for row_number, row_segments in rows_by_number.items()
        if all(
            not row.raw_existing_target
            and is_non_translatable_segment(row.source)
            for row in row_segments
        )
    }
    representative_row_numbers = {
        min(
            unit.items,
            key=lambda item: (item.row_number, item.segment_index),
        ).row_number
        for unit in grouped_units
    }
    duplicate_segment_count = len(pending_rows) - len(grouped_units)
    return {
        "output_path": str(output_path),
        "source_row_count": len(source_row_numbers),
        "source_segment_count": len(rows),
        "multiline_source_row_count": len(
            {
                row.row_number
                for row in rows
                if "\n" in row.raw_source or "\r" in row.raw_source
            }
        ),
        "completed_row_count": existing_target_count,
        "non_translatable_row_count": len(auto_completed_row_numbers),
        "non_translatable_segment_count": passthrough_count,
        "auto_completed_row_count": len(auto_completed_row_numbers),
        "auto_completed_segment_count": passthrough_count,
        "pending_row_count": len(pending_row_numbers),
        "pending_segment_count": len(pending_rows),
        "string_count": len(grouped_units),
        "duplicate_row_count": len(
            pending_row_numbers - representative_row_numbers
        ),
        "duplicate_segment_count": duplicate_segment_count,
        "group_count": group_count,
        "grouped_string_count": grouped_count,
        "grouping_enabled": group_similar,
    }


def default_restored_output_path(package_path: str | Path) -> Path:
    package_path = Path(package_path)
    metadata = load_strings_metadata(package_path)
    original_filename = str(
        metadata.get(schema.ORIGINAL_FILENAME_KEY) or "source.xlsx"
    )
    return package_path.with_name(
        f"{Path(original_filename).stem}_translated.xlsx"
    )


def restore_strings_workbook(
    package_path: str | Path,
    output_path: str | Path | None = None,
) -> dict[str, int | str]:
    package_path = Path(package_path)
    output_path = (
        Path(output_path)
        if output_path is not None
        else default_restored_output_path(package_path)
    )
    _require_distinct_paths(package_path, output_path)

    metadata = load_strings_metadata(package_path)
    source_column = metadata_column(metadata, schema.SOURCE_COLUMN_KEY)
    target_column = metadata_column(metadata, schema.TARGET_COLUMN_KEY)
    if target_column is None:
        raise WorkbookFormatError(
            "Strings workbook does not define a Target column."
        )
    tag_rules = embedded_tag_rules(metadata, package_path)

    workbook = load_workbook(package_path)
    issues: list[dict[str, object]] = []
    restored_count = 0
    try:
        string_rows = read_strings_rows(workbook[schema.STRINGS_SHEET])
        mapping_rows = read_mapping_rows(workbook[schema.STRINGS_MAP_SHEET])
        original = original_worksheet(workbook)
        source_index = resolve_column(original, source_column)
        target_index = resolve_or_create_column(original, target_column)

        mapping_rows_by_source_row: dict[int, list[dict[str, object]]] = {}
        for map_row in mapping_rows:
            row_number = int(map_row[schema.ROW_NUMBER_COLUMN])
            mapping_rows_by_source_row.setdefault(row_number, []).append(map_row)

        for source_rows in mapping_rows_by_source_row.values():
            restored = _restore_mapped_cell(
                original,
                source_index=source_index,
                target_index=target_index,
                map_rows=source_rows,
                string_rows=string_rows,
                tag_rules=tag_rules,
                issues=issues,
            )
            restored_count += int(restored)

        restore_original_sheets(workbook, metadata)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()
    stats: dict[str, int | str] = {
        "output_path": str(output_path),
        "restored_row_count": restored_count,
        "issue_count": len(issues),
    }
    if issues:
        audit_path = output_path.with_name(
            f"{output_path.stem}_restore_issues{output_path.suffix}"
        )
        write_restore_issues(audit_path, issues)
        stats["audit_output_path"] = str(audit_path)
    return stats


def _read_rows_with_optional_target(
    input_path: Path,
    *,
    source_col: str | int,
    target_col: str | int,
    tag_rules: TagRules,
) -> list[RowItem]:
    try:
        return read_source_rows(
            input_path,
            source_col,
            target_col,
            tag_rules=tag_rules,
        )
    except ColumnNotFoundError:
        if str(target_col).strip().lower() != schema.TARGET_COLUMN:
            raise
        return read_source_rows(
            input_path,
            source_col,
            None,
            tag_rules=tag_rules,
        )


def _apply_optional_grouping(
    units: list[StringUnit],
    *,
    enabled: bool,
    min_group_size: int,
) -> tuple[list[StringUnit], int, int]:
    if not enabled:
        return units, 0, 0

    clusters = cluster_similar_strings(
        [unit.source for unit in units],
        min_group_size=min_group_size,
    )
    group_by_index = {
        member_index: cluster.group_id
        for cluster in clusters
        for member_index in cluster.member_indexes
    }
    group_source_order = {
        cluster.group_id: min(
            _source_order(units[member_index])
            for member_index in cluster.member_indexes
        )
        for cluster in clusters
    }
    grouped_units = [
        StringUnit(
            string_id=unit.string_id,
            unit_type=unit.unit_type,
            source=unit.source,
            items=unit.items,
            sample_source=unit.sample_source,
            context=unit.context,
            group_id=group_by_index.get(index, ""),
        )
        for index, unit in enumerate(units)
    ]
    grouped_units.sort(
        key=lambda unit: (
            1 if unit.group_id else 0,
            group_source_order.get(unit.group_id, _source_order(unit)),
            _source_order(unit),
        )
    )
    return grouped_units, len(clusters), len(group_by_index)


def _restore_mapped_cell(
    worksheet,
    *,
    source_index: int,
    target_index: int,
    map_rows: list[dict[str, object]],
    string_rows: dict[str, dict[str, object]],
    tag_rules: TagRules,
    issues: list[dict[str, object]],
) -> bool:
    if not map_rows:
        return False

    row_number = int(map_rows[0][schema.ROW_NUMBER_COLUMN])
    raw_source = str(map_rows[0][schema.SOURCE_COLUMN])
    if any(
        int(map_row[schema.ROW_NUMBER_COLUMN]) != row_number
        or str(map_row[schema.SOURCE_COLUMN]) != raw_source
        for map_row in map_rows[1:]
    ):
        raise WorkbookFormatError(
            f"Strings mapping is inconsistent at row {row_number}."
        )

    segmented = all(
        map_row.get(schema.SEGMENT_INDEX_COLUMN) not in (None, "")
        for map_row in map_rows
    )
    if segmented:
        map_rows = sorted(
            map_rows,
            key=lambda map_row: int(map_row[schema.SEGMENT_INDEX_COLUMN]),
        )
        segment_count = int(map_rows[0][schema.SEGMENT_COUNT_COLUMN])
        expected_indexes = list(range(1, segment_count + 1))
        actual_indexes = [
            int(map_row[schema.SEGMENT_INDEX_COLUMN]) for map_row in map_rows
        ]
        if (
            len(map_rows) != segment_count
            or actual_indexes != expected_indexes
            or any(
                int(map_row[schema.SEGMENT_COUNT_COLUMN]) != segment_count
                for map_row in map_rows
            )
        ):
            raise WorkbookFormatError(
                f"Strings mapping has incomplete segments at row {row_number}."
            )
    elif len(map_rows) != 1:
        raise WorkbookFormatError(
            f"Legacy Strings mapping has duplicate rows at row {row_number}."
        )

    actual_source = worksheet.cell(row=row_number, column=source_index).value
    comparable_actual = "" if actual_source is None else str(actual_source)
    if (
        comparable_actual != raw_source
        if segmented
        else comparable_actual.strip() != raw_source
    ):
        raise WorkbookFormatError(
            f"Original Source changed at row {row_number}: {raw_source!r}"
        )

    restored_segments: list[str] = []
    complete = True
    for map_row in map_rows:
        restored_segment = _restore_mapped_segment(
            map_row=map_row,
            string_rows=string_rows,
            tag_rules=tag_rules,
            issues=issues,
            defer_tag_restore=segmented,
        )
        if restored_segment is None:
            complete = False
            continue
        prefix = str(map_row.get(schema.SEGMENT_PREFIX_COLUMN) or "")
        suffix = str(map_row.get(schema.SEGMENT_SUFFIX_COLUMN) or "")
        restored_segments.append(prefix + restored_segment + suffix)

    if not complete:
        return False
    restored_target = "".join(restored_segments)
    if segmented:
        restored_target, warnings = _restore_target_unit(
            raw_source=raw_source,
            source_unit="",
            target_unit=restored_target,
            unit_type="segment",
            variables={},
            tag_rules=tag_rules,
        )
        if warnings:
            issues.append(
                _issue(
                    ",".join(
                        dict.fromkeys(
                            str(map_row[schema.STRING_ID_COLUMN])
                            for map_row in map_rows
                        )
                    ),
                    raw_source,
                    restored_target,
                    row_number,
                    "; ".join(warnings),
                )
            )
    worksheet.cell(row=row_number, column=target_index).value = restored_target
    return True


def _restore_mapped_segment(
    *,
    map_row: dict[str, object],
    string_rows: dict[str, dict[str, object]],
    tag_rules: TagRules,
    issues: list[dict[str, object]],
    defer_tag_restore: bool,
) -> str | None:
    string_id = str(map_row[schema.STRING_ID_COLUMN])
    row_number = int(map_row[schema.ROW_NUMBER_COLUMN])
    raw_segment = str(
        map_row.get(schema.RAW_SEGMENT_COLUMN)
        if map_row.get(schema.RAW_SEGMENT_COLUMN) is not None
        else map_row[schema.SOURCE_COLUMN]
    )
    unit_source = str(map_row[schema.SOURCE_UNIT_COLUMN])
    unit_type = str(map_row[schema.UNIT_TYPE_COLUMN])

    if unit_type == PASSTHROUGH_UNIT_TYPE:
        return raw_segment

    string_row = string_rows.get(string_id)
    if string_row is None:
        issues.append(
            _issue(
                string_id,
                unit_source,
                "",
                row_number,
                "String row is missing.",
            )
        )
        return None
    visible_source = str(string_row[schema.SOURCE_COLUMN] or "").strip()
    if visible_source != unit_source:
        raise WorkbookFormatError(
            f"Source was changed in Strings row {string_id}: {visible_source!r}"
        )
    target_unit = str(string_row[schema.TARGET_COLUMN] or "").strip()
    if not target_unit:
        issues.append(
            _issue(
                string_id,
                unit_source,
                "",
                row_number,
                "Target is empty.",
            )
        )
        return None

    if defer_tag_restore:
        restored_target, warnings = _apply_target_variables(
            source_unit=unit_source,
            target_unit=target_unit,
            unit_type=unit_type,
            variables=mapping_variables(map_row),
        )
    else:
        restored_target, warnings = _restore_target_unit(
            raw_source=raw_segment,
            source_unit=unit_source,
            target_unit=target_unit,
            unit_type=unit_type,
            variables=mapping_variables(map_row),
            tag_rules=tag_rules,
        )
    if warnings:
        issues.append(
            _issue(
                string_id,
                unit_source,
                target_unit,
                row_number,
                "; ".join(warnings),
            )
        )
    return restored_target


def _restore_target_unit(
    *,
    raw_source: str,
    source_unit: str,
    target_unit: str,
    unit_type: str,
    variables: dict[str, str],
    tag_rules: TagRules,
) -> tuple[str, list[str]]:
    candidate, warnings = _apply_target_variables(
        source_unit=source_unit,
        target_unit=target_unit,
        unit_type=unit_type,
        variables=variables,
    )

    source_extraction = extract_tags(raw_source, rules=tag_rules)
    serialized_target = serialize_known_tags(
        candidate,
        source_extraction.tags,
    )
    warnings.extend(source_extraction.warnings)
    warnings.extend(serialized_target.warnings)
    warnings.extend(
        validate_tag_placeholders(
            serialized_target.text,
            source_extraction.tags,
        ).warnings
    )
    restored_target = restore_tags(
        serialized_target.text,
        source_extraction.tags,
    )
    warnings.extend(_target_warnings(raw_source, restored_target, tag_rules))
    return restored_target, list(dict.fromkeys(warnings))


def _apply_target_variables(
    *,
    source_unit: str,
    target_unit: str,
    unit_type: str,
    variables: dict[str, str],
) -> tuple[str, list[str]]:
    if unit_type != "template":
        return target_unit, []
    source_variables = set(PLACEHOLDER_RE.findall(source_unit))
    target_variables = set(PLACEHOLDER_RE.findall(target_unit))
    warnings = [
        f"target_unit is missing source variable: {variable}"
        for variable in sorted(source_variables - target_variables)
    ]
    return apply_target_template(target_unit, variables), warnings


def _target_warnings(
    source: str,
    target: str,
    tag_rules: TagRules,
) -> list[str]:
    source_extraction = extract_tags(source, rules=tag_rules)
    serialized_target = serialize_known_tags(target, source_extraction.tags)
    warnings = list(source_extraction.warnings)
    warnings.extend(serialized_target.warnings)
    warnings.extend(
        validate_tag_placeholders(
            serialized_target.text,
            source_extraction.tags,
        ).warnings
    )
    target_extraction = extract_tags(target, rules=tag_rules)
    warnings.extend(target_extraction.warnings)
    source_tokens = Counter(token.raw for token in source_extraction.tags)
    target_tokens = Counter(token.raw for token in target_extraction.tags)
    for raw in sorted(target_tokens - source_tokens):
        warnings.extend(
            f"protected_token_mismatch: extra {raw}"
            for _ in range(target_tokens[raw] - source_tokens[raw])
        )
    return list(dict.fromkeys(warnings))


def _context_column_index(
    input_path: Path,
    context_col: str | int | None,
) -> int | None:
    headers = read_headers(input_path)
    if context_col is None or not str(context_col).strip():
        return next(
            (
                index
                for index, header in enumerate(headers)
                if header.lower() == schema.CONTEXT_COLUMN
            ),
            None,
        )
    if isinstance(context_col, int) or str(context_col).isdigit():
        index = int(context_col) - 1
        if 0 <= index < len(headers):
            return index
        raise ConfigError(f"Context column is out of range: {context_col}")
    wanted = str(context_col).strip().lower()
    for index, header in enumerate(headers):
        if header.lower() == wanted:
            return index
    raise ConfigError(f"Context column not found: {context_col}")


def _context(
    items: Iterable[RowItem],
    context_index: int | None,
) -> object | None:
    first = _first_item(items)
    if (
        first is None
        or context_index is None
        or context_index >= len(first.original_values)
    ):
        return None
    value = first.original_values[context_index]
    return None if value in (None, "") else value


def _row_context(
    item: RowItem,
    context_index: int | None,
) -> object | None:
    if context_index is None or context_index >= len(item.original_values):
        return None
    value = item.original_values[context_index]
    return None if value in (None, "") else value


def _sample_source(items: Iterable[RowItem]) -> str:
    first = _first_item(items)
    return first.raw_segment if first is not None else ""


def _first_item(items: Iterable[RowItem]) -> RowItem | None:
    return min(
        items,
        key=lambda item: (item.row_number, item.segment_index),
        default=None,
    )


def _source_order(unit: StringUnit) -> tuple[int, int]:
    return min((item.row_number, item.segment_index) for item in unit.items)


def _issue(
    string_id: str,
    source: str,
    target: str,
    row_number: int,
    issue: str,
) -> dict[str, object]:
    return {
        "string_id": string_id,
        "source": source,
        "target": target,
        "row_number": row_number,
        "issue": issue,
    }


def _require_distinct_paths(input_path: Path, output_path: Path) -> None:
    if input_path.resolve() == output_path.resolve():
        raise ConfigError("Input and output paths must be different.")


__all__ = [
    "default_restored_output_path",
    "default_strings_output_path",
    "export_strings_workbook",
    "restore_strings_workbook",
]
