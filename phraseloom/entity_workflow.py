from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

from openpyxl import Workbook, load_workbook

from . import workbook_schema as schema
from ._entity_cluster_probe import find_entity_clusters
from .errors import WorkflowError
from .excel_io import _header_values, _style_sheet
from .models import EntityCluster


ENTITY_STRUCTURE_COLUMNS = [
    schema.STRUCTURE_ID_COLUMN,
    schema.SOURCE_STRUCTURE_COLUMN,
    schema.TARGET_STRUCTURE_COLUMN,
    schema.COVERAGE_COUNT_COLUMN,
    schema.CONFIDENCE_COLUMN,
    schema.RISK_COLUMN,
    schema.STATUS_COLUMN,
    schema.SAMPLE_SOURCES_COLUMN,
    schema.ROW_NUMBERS_COLUMN,
    schema.WARNING_COLUMN,
]
ENTITY_TERM_COLUMNS = [
    schema.TERM_ID_COLUMN,
    schema.SOURCE_ENTITY_COLUMN,
    schema.TARGET_ENTITY_COLUMN,
    schema.OCCURRENCE_COUNT_COLUMN,
    schema.STRUCTURE_IDS_COLUMN,
    schema.STATUS_COLUMN,
    schema.WARNING_COLUMN,
]
ENTITY_SOURCE_MAP_COLUMNS = [
    schema.ORIGINAL_INDEX_COLUMN,
    schema.UNIT_ID_COLUMN,
    schema.UNIT_TYPE_COLUMN,
    schema.SOURCE_UNIT_COLUMN,
    schema.STRUCTURE_ID_COLUMN,
    schema.ENTITIES_JSON_COLUMN,
    schema.PREVIEW_TARGET_COLUMN,
    schema.FILL_STATUS_COLUMN,
    schema.WARNING_COLUMN,
]


@dataclass(frozen=True)
class UnitRow:
    original_index: int
    values: dict[str, object]

    @property
    def unit_id(self) -> str:
        return str(self.values.get(schema.UNIT_ID_COLUMN) or "")

    @property
    def unit_type(self) -> str:
        return str(self.values.get(schema.UNIT_TYPE_COLUMN) or "")

    @property
    def source_unit(self) -> str:
        return str(self.values.get(schema.SOURCE_UNIT_COLUMN) or "")

    @property
    def target_unit(self) -> str:
        return str(self.values.get(schema.TARGET_UNIT_COLUMN) or "")


class EntityExtractionStrategy(Protocol):
    name: str

    def find_clusters(self, rows: list[tuple[str, str]]) -> list[EntityCluster]:
        ...


@dataclass(frozen=True)
class ClusterProbeStrategy:
    min_group_size: int = 3
    max_entity_tokens: int = 4
    min_literal_tokens: int = 3
    top: int = 200
    name: str = "cluster_probe"

    def find_clusters(self, rows: list[tuple[str, str]]) -> list[EntityCluster]:
        return find_entity_clusters(
            rows,
            min_group_size=self.min_group_size,
            max_entity_tokens=self.max_entity_tokens,
            min_literal_tokens=self.min_literal_tokens,
            top=self.top,
        )


def _default_entity_work_dir(input_path: Path) -> Path:
    return input_path.parent / f"{input_path.stem}_l10n"


def default_entity_memory_output_path(input_path: str | Path) -> Path:
    input_path = Path(input_path)
    return _default_entity_work_dir(input_path) / f"{input_path.stem}_entity_memory.xlsx"


def default_entity_pack_output_path(input_path: str | Path) -> Path:
    input_path = Path(input_path)
    return _default_entity_work_dir(input_path) / f"{input_path.stem}_entity_pack.xlsx"


def default_entity_filled_pack_output_path(input_path: str | Path) -> Path:
    input_path = Path(input_path)
    return input_path.with_name(f"{input_path.stem}_filled.xlsx")


def default_entity_merged_todo_output_path(input_path: str | Path) -> Path:
    input_path = Path(input_path)
    return input_path.with_name(f"{input_path.stem}_merged_todo.xlsx")


def extract_entity_memory_workbook(
    input_path: str | Path,
    output_path: str | Path,
    *,
    min_group_size: int = 3,
    strategy: EntityExtractionStrategy | None = None,
) -> dict[str, int | str]:
    return extract_entity_tm_workbook(
        input_path,
        output_path,
        min_group_size=min_group_size,
        strategy=strategy,
    )


def split_entity_workbook(
    input_path: str | Path,
    entity_output_path: str | Path,
    non_entity_output_path: str | Path,
    *,
    min_group_size: int = 3,
    strategy: EntityExtractionStrategy | None = None,
) -> dict[str, int | str]:
    input_path = Path(input_path)
    entity_output_path = Path(entity_output_path)
    non_entity_output_path = Path(non_entity_output_path)
    rows, headers = _read_unit_rows(input_path, schema.TO_TRANSLATE_SHEET)
    active_strategy = strategy or ClusterProbeStrategy(min_group_size=min_group_size)
    clusters = active_strategy.find_clusters(
        [(row.source_unit, row.target_unit) for row in rows]
    )
    entity_rows, structures, terms, source_map = _build_entity_split(rows, clusters)
    entity_indices = {row.original_index for row in entity_rows}
    non_entity_rows = [row for row in rows if row.original_index not in entity_indices]

    _write_entity_related_workbook(
        entity_output_path,
        input_path,
        headers,
        entity_rows,
        structures,
        terms,
        source_map,
    )
    _write_non_entity_workbook(non_entity_output_path, input_path, headers, non_entity_rows)
    return {
        "entity_unit_count": len(entity_rows),
        "non_entity_unit_count": len(non_entity_rows),
        "entity_structure_count": len(structures),
        "entity_term_count": len(terms),
        "entity_output_path": str(entity_output_path),
        "non_entity_output_path": str(non_entity_output_path),
    }


def extract_entity_tm_workbook(
    input_path: str | Path,
    output_path: str | Path,
    *,
    min_group_size: int = 3,
    strategy: EntityExtractionStrategy | None = None,
) -> dict[str, int | str]:
    input_path = Path(input_path)
    output_path = Path(output_path)
    rows = _read_tm_pair_rows(input_path)
    active_strategy = strategy or ClusterProbeStrategy(min_group_size=min_group_size)
    clusters = active_strategy.find_clusters(
        [(row.source_unit, row.target_unit) for row in rows]
    )
    structures, terms = _build_entity_tm(rows, clusters, min_group_size)
    _write_entity_tm_workbook(output_path, structures, terms)
    return {
        "entity_structure_count": len(structures),
        "entity_term_count": len(terms),
        "output_path": str(output_path),
    }


def prefill_entity_workbook(
    entity_input_path: str | Path,
    tm_path: str | Path,
    output_path: str | Path,
) -> dict[str, int | str]:
    entity_input_path = Path(entity_input_path)
    tm_path = Path(tm_path)
    output_path = Path(output_path)
    structure_prefills = _load_unique_prefills(
        tm_path,
        schema.ENTITY_STRUCTURES_SHEET,
        schema.SOURCE_STRUCTURE_COLUMN,
        schema.TARGET_STRUCTURE_COLUMN,
    )
    term_prefills = _load_unique_prefills(
        tm_path,
        schema.ENTITY_TERMS_SHEET,
        schema.SOURCE_ENTITY_COLUMN,
        schema.TARGET_ENTITY_COLUMN,
    )

    wb = load_workbook(entity_input_path)
    try:
        prefilled_structure_count = _apply_prefills(
            wb[schema.ENTITY_STRUCTURES_SHEET],
            source_column=schema.SOURCE_STRUCTURE_COLUMN,
            target_column=schema.TARGET_STRUCTURE_COLUMN,
            prefills=structure_prefills,
            ready_on_prefill=False,
        )
        prefilled_term_count = _apply_prefills(
            wb[schema.ENTITY_TERMS_SHEET],
            source_column=schema.SOURCE_ENTITY_COLUMN,
            target_column=schema.TARGET_ENTITY_COLUMN,
            prefills=term_prefills,
            ready_on_prefill=True,
        )
        _save_workbook(wb, output_path)
    finally:
        wb.close()

    return {
        "prefilled_structure_count": prefilled_structure_count,
        "prefilled_term_count": prefilled_term_count,
        "output_path": str(output_path),
    }


def fill_entity_workbook(
    entity_input_path: str | Path,
    output_path: str | Path,
) -> dict[str, int | str]:
    entity_input_path = Path(entity_input_path)
    output_path = Path(output_path)
    wb = load_workbook(entity_input_path)
    try:
        structures = _load_rows_by_key(
            wb[schema.ENTITY_STRUCTURES_SHEET],
            schema.STRUCTURE_ID_COLUMN,
        )
        terms = _load_rows_by_key(
            wb[schema.ENTITY_TERMS_SHEET],
            schema.SOURCE_ENTITY_COLUMN,
        )
        todo_ws = wb[schema.TO_TRANSLATE_SHEET]
        source_map_ws = wb[schema.ENTITY_SOURCE_MAP_SHEET]
        todo_by_original_index = _todo_rows_by_original_index(todo_ws)
        filled_count = _fill_source_map_rows(
            source_map_ws,
            todo_ws,
            todo_by_original_index,
            structures,
            terms,
        )
        _save_workbook(wb, output_path)
    finally:
        wb.close()

    return {
        "filled_entity_unit_count": filled_count,
        "output_path": str(output_path),
    }


def merge_entity_workbooks(
    entity_path: str | Path,
    non_entity_path: str | Path,
    output_path: str | Path,
) -> dict[str, int | str]:
    entity_path = Path(entity_path)
    non_entity_path = Path(non_entity_path)
    output_path = Path(output_path)
    entity_rows, entity_headers = _read_unit_rows(entity_path, schema.TO_TRANSLATE_SHEET)
    non_entity_rows, non_entity_headers = _read_unit_rows(
        non_entity_path,
        schema.TO_TRANSLATE_SHEET,
    )
    headers = [
        header
        for header in (non_entity_headers or entity_headers)
        if header != schema.ORIGINAL_INDEX_COLUMN
    ]
    merged_rows = _merge_unit_rows(entity_rows + non_entity_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = schema.TO_TRANSLATE_SHEET
    ws.append(headers)
    for row in merged_rows:
        ws.append([row.values.get(header) for header in headers])
    _copy_support_sheets(non_entity_path, wb, exclude={schema.TO_TRANSLATE_SHEET})
    _save_workbook(wb, output_path)
    return {
        "merged_unit_count": len(merged_rows),
        "output_path": str(output_path),
    }


def _read_unit_rows(path: Path, sheet_name: str) -> tuple[list[UnitRow], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        headers = _header_values(ws)
        rows = []
        for sequence_index, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=1,
        ):
            values = {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
            original_index = int(
                values.get(schema.ORIGINAL_INDEX_COLUMN) or sequence_index
            )
            if values.get(schema.SOURCE_UNIT_COLUMN):
                rows.append(UnitRow(original_index, values))
        return rows, headers
    finally:
        wb.close()


def _read_tm_pair_rows(path: Path) -> list[UnitRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[schema.TM_PAIRS_SHEET]
        headers = _header_values(ws)
        rows: list[UnitRow] = []
        for sequence_index, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=1,
        ):
            values = {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
            if values.get(schema.SOURCE_UNIT_COLUMN) and values.get(schema.TARGET_UNIT_COLUMN):
                values[schema.UNIT_ID_COLUMN] = values.get(schema.TM_ID_COLUMN)
                rows.append(UnitRow(sequence_index, values))
        return rows
    finally:
        wb.close()


def _build_entity_tm(
    rows: list[UnitRow],
    clusters: list[EntityCluster],
    min_group_size: int,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    structures: list[dict[str, object]] = []
    term_occurrences: dict[str, Counter[str]] = defaultdict(Counter)
    term_targets: dict[str, set[str]] = defaultdict(set)
    assigned_indexes: set[int] = set()

    for structure_index, cluster in enumerate(clusters, start=1):
        structure_id = f"TMES{structure_index:04d}"
        occurrences: list[tuple[UnitRow, dict[str, str]]] = []
        for cluster_row_number in cluster.row_numbers:
            row_index = cluster_row_number - 2
            if row_index < 0 or row_index >= len(rows):
                continue
            unit_row = rows[row_index]
            if unit_row.original_index in assigned_indexes:
                continue
            source_entities = _extract_entities(
                cluster.source_pattern,
                unit_row.source_unit,
            )
            if source_entities is None:
                continue
            occurrences.append((unit_row, source_entities))

        if not occurrences:
            continue

        target_structure, target_entity_rows = _derive_target_entity_pattern(
            occurrences,
            min_group_size=min(min_group_size, len(occurrences)),
        )
        for unit_row, source_entities in occurrences:
            assigned_indexes.add(unit_row.original_index)
            for source_entity in source_entities.values():
                term_occurrences[source_entity][structure_id] += 1
            target_entities = target_entity_rows.get(unit_row.original_index, {})
            for source_entity, target_entity in _ordered_entity_pairs(
                source_entities,
                target_entities,
            ):
                if target_entity:
                    term_targets[source_entity].add(target_entity)

        structures.append(
            {
                schema.STRUCTURE_ID_COLUMN: structure_id,
                schema.SOURCE_STRUCTURE_COLUMN: cluster.source_pattern,
                schema.TARGET_STRUCTURE_COLUMN: target_structure,
                schema.COVERAGE_COUNT_COLUMN: len(occurrences),
                schema.CONFIDENCE_COLUMN: cluster.confidence,
                schema.RISK_COLUMN: cluster.risk or None,
                schema.STATUS_COLUMN: "ready" if target_structure else "review",
                schema.SAMPLE_SOURCES_COLUMN: "\n".join(
                    dict.fromkeys(row.source_unit for row, _entities in occurrences[:10])
                ),
                schema.ROW_NUMBERS_COLUMN: ",".join(
                    str(row.original_index) for row, _entities in occurrences
                ),
                schema.WARNING_COLUMN: None
                if target_structure
                else "target entity structure not inferred",
            }
        )

    terms = []
    for index, (source_entity, structures_count) in enumerate(
        sorted(term_occurrences.items()),
        start=1,
    ):
        targets = term_targets.get(source_entity, set())
        target_entity = next(iter(targets)) if len(targets) == 1 else None
        warning = "ambiguous_entity_translation" if len(targets) > 1 else None
        if not targets:
            warning = "target entity not inferred"
        terms.append(
            {
                schema.TERM_ID_COLUMN: f"TMET{index:04d}",
                schema.SOURCE_ENTITY_COLUMN: source_entity,
                schema.TARGET_ENTITY_COLUMN: target_entity,
                schema.OCCURRENCE_COUNT_COLUMN: sum(structures_count.values()),
                schema.STRUCTURE_IDS_COLUMN: ",".join(sorted(structures_count)),
                schema.STATUS_COLUMN: "ready" if target_entity else "review",
                schema.WARNING_COLUMN: warning,
            }
        )
    return structures, terms


def _derive_target_entity_pattern(
    occurrences: list[tuple[UnitRow, dict[str, str]]],
    *,
    min_group_size: int,
) -> tuple[str | None, dict[int, dict[str, str]]]:
    if len(occurrences) < min_group_size:
        return None, {}
    target_rows = [(row.target_unit, "") for row, _entities in occurrences]
    target_clusters = find_entity_clusters(
        target_rows,
        min_group_size=min_group_size,
        max_entity_tokens=4,
        min_literal_tokens=3,
        top=1,
    )
    if not target_clusters:
        return None, {}

    target_cluster = target_clusters[0]
    target_entities_by_index: dict[int, dict[str, str]] = {}
    for target_cluster_row_number in target_cluster.row_numbers:
        occurrence_index = target_cluster_row_number - 2
        if occurrence_index < 0 or occurrence_index >= len(occurrences):
            continue
        unit_row, _source_entities = occurrences[occurrence_index]
        target_entities = _extract_entities(
            target_cluster.source_pattern,
            unit_row.target_unit,
        )
        if target_entities is not None:
            target_entities_by_index[unit_row.original_index] = target_entities
    return target_cluster.source_pattern, target_entities_by_index


def _ordered_entity_pairs(
    source_entities: dict[str, str],
    target_entities: dict[str, str],
) -> list[tuple[str, str | None]]:
    source_values = [
        value for key, value in sorted(source_entities.items(), key=lambda item: item[0])
    ]
    target_values = [
        value for key, value in sorted(target_entities.items(), key=lambda item: item[0])
    ]
    return [
        (source_entity, target_values[index] if index < len(target_values) else None)
        for index, source_entity in enumerate(source_values)
    ]


def _load_rows_by_key(ws, key_column: str) -> dict[str, dict[str, object]]:
    headers = _header_values(ws)
    key_index = headers.index(key_column)
    rows: dict[str, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[key_index] if key_index < len(row) else None
        if not key:
            continue
        rows[str(key)] = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
    return rows


def _todo_rows_by_original_index(ws) -> dict[int, int]:
    headers = _header_values(ws)
    original_index_column = headers.index(schema.ORIGINAL_INDEX_COLUMN) + 1
    rows: dict[int, int] = {}
    for row_number in range(2, ws.max_row + 1):
        original_index = ws.cell(row=row_number, column=original_index_column).value
        if original_index is not None:
            rows[int(original_index)] = row_number
    return rows


def _merge_unit_rows(rows: list[UnitRow]) -> list[UnitRow]:
    by_index: dict[int, UnitRow] = {}
    for row in rows:
        if not row.unit_id:
            raise WorkflowError(f"Missing unit_id for original_index {row.original_index}")
        if not row.source_unit:
            raise WorkflowError(
                f"Missing source_unit for original_index {row.original_index}"
            )
        existing = by_index.get(row.original_index)
        if existing is not None:
            if existing.unit_id != row.unit_id or existing.source_unit != row.source_unit:
                raise WorkflowError(
                    f"Conflicting rows for original_index {row.original_index}"
                )
            raise WorkflowError(f"Duplicate original_index {row.original_index}")
        by_index[row.original_index] = row

    if not by_index:
        return []
    expected = set(range(1, max(by_index) + 1))
    actual = set(by_index)
    if expected != actual:
        missing = ",".join(str(index) for index in sorted(expected - actual))
        raise WorkflowError(f"Missing original_index values: {missing}")
    return [by_index[index] for index in sorted(by_index)]


def _fill_source_map_rows(
    source_map_ws,
    todo_ws,
    todo_by_original_index: dict[int, int],
    structures: dict[str, dict[str, object]],
    terms: dict[str, dict[str, object]],
) -> int:
    map_headers = _header_values(source_map_ws)
    todo_headers = _header_values(todo_ws)
    map_columns = _one_based_columns(map_headers)
    todo_columns = _one_based_columns(todo_headers)
    filled_count = 0

    for row_number in range(2, source_map_ws.max_row + 1):
        original_index = int(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.ORIGINAL_INDEX_COLUMN],
            ).value
        )
        unit_id = str(
            source_map_ws.cell(row=row_number, column=map_columns[schema.UNIT_ID_COLUMN]).value
            or ""
        )
        source_unit = str(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.SOURCE_UNIT_COLUMN],
            ).value
            or ""
        )
        structure_id = str(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.STRUCTURE_ID_COLUMN],
            ).value
            or ""
        )
        entities = json.loads(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.ENTITIES_JSON_COLUMN],
            ).value
            or "{}"
        )
        preview_target, fill_status, warning = _build_entity_target(
            original_index,
            unit_id,
            source_unit,
            structure_id,
            entities,
            todo_ws,
            todo_by_original_index,
            todo_columns,
            structures,
            terms,
        )
        source_map_ws.cell(
            row=row_number,
            column=map_columns[schema.PREVIEW_TARGET_COLUMN],
        ).value = preview_target
        source_map_ws.cell(
            row=row_number,
            column=map_columns[schema.FILL_STATUS_COLUMN],
        ).value = fill_status
        source_map_ws.cell(
            row=row_number,
            column=map_columns[schema.WARNING_COLUMN],
        ).value = warning
        if preview_target:
            todo_row_number = todo_by_original_index[original_index]
            todo_ws.cell(
                row=todo_row_number,
                column=todo_columns[schema.TARGET_UNIT_COLUMN],
            ).value = preview_target
            filled_count += 1
    return filled_count


def _build_entity_target(
    original_index: int,
    unit_id: str,
    source_unit: str,
    structure_id: str,
    entities: dict[str, str],
    todo_ws,
    todo_by_original_index: dict[int, int],
    todo_columns: dict[str, int],
    structures: dict[str, dict[str, object]],
    terms: dict[str, dict[str, object]],
) -> tuple[str | None, str, str | None]:
    todo_row_number = todo_by_original_index.get(original_index)
    if todo_row_number is None:
        return None, "unit_not_found", f"original_index not found: {original_index}"
    if (
        str(todo_ws.cell(row=todo_row_number, column=todo_columns[schema.UNIT_ID_COLUMN]).value or "")
        != unit_id
        or str(
            todo_ws.cell(
                row=todo_row_number,
                column=todo_columns[schema.SOURCE_UNIT_COLUMN],
            ).value
            or ""
        )
        != source_unit
    ):
        return None, "unit_mismatch", "unit_id/source_unit changed"

    structure = structures.get(structure_id)
    if structure is None:
        return None, "structure_not_found", f"structure not found: {structure_id}"
    if str(structure.get(schema.STATUS_COLUMN) or "").lower() != "ready":
        return None, "structure_not_ready", f"structure not ready: {structure_id}"
    target_structure = str(structure.get(schema.TARGET_STRUCTURE_COLUMN) or "")
    if not target_structure:
        return None, "missing_structure_translation", f"missing target structure: {structure_id}"

    target = target_structure
    for placeholder_name, source_entity in sorted(entities.items()):
        term = terms.get(str(source_entity))
        if term is None:
            return None, "missing_entity_translation", f"missing entity target: {source_entity}"
        target_entity = str(term.get(schema.TARGET_ENTITY_COLUMN) or "")
        if not target_entity:
            return None, "missing_entity_translation", f"missing entity target: {source_entity}"
        if str(term.get(schema.STATUS_COLUMN) or "").lower() != "ready":
            return None, "entity_not_ready", f"entity not ready: {source_entity}"
        target = target.replace("{" + placeholder_name + "}", target_entity)
    return target, "filled", None


def _one_based_columns(headers: list[str]) -> dict[str, int]:
    return {header: index + 1 for index, header in enumerate(headers) if header}


def _load_unique_prefills(
    path: Path,
    sheet_name: str,
    source_column: str,
    target_column: str,
) -> dict[str, tuple[str | None, str | None]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        headers = _header_values(ws)
        source_index = headers.index(source_column)
        target_index = headers.index(target_column)
        values: dict[str, set[str]] = defaultdict(set)
        for row in ws.iter_rows(min_row=2, values_only=True):
            source = row[source_index] if source_index < len(row) else None
            target = row[target_index] if target_index < len(row) else None
            if source and target:
                values[str(source)].add(str(target))
        result: dict[str, tuple[str | None, str | None]] = {}
        for source, targets in values.items():
            if len(targets) == 1:
                result[source] = (next(iter(targets)), None)
            elif len(targets) > 1:
                result[source] = (None, "ambiguous_entity_prefill")
        return result
    finally:
        wb.close()


def _apply_prefills(
    ws,
    *,
    source_column: str,
    target_column: str,
    prefills: dict[str, tuple[str | None, str | None]],
    ready_on_prefill: bool,
) -> int:
    headers = _header_values(ws)
    source_index = headers.index(source_column) + 1
    target_index = headers.index(target_column) + 1
    status_index = headers.index(schema.STATUS_COLUMN) + 1
    warning_index = headers.index(schema.WARNING_COLUMN) + 1
    prefilled_count = 0
    for row_number in range(2, ws.max_row + 1):
        source = ws.cell(row=row_number, column=source_index).value
        target = ws.cell(row=row_number, column=target_index).value
        if not source or target:
            continue
        prefill_target, warning = prefills.get(str(source), (None, None))
        if prefill_target:
            ws.cell(row=row_number, column=target_index).value = prefill_target
            if ready_on_prefill:
                ws.cell(row=row_number, column=status_index).value = "ready"
            prefilled_count += 1
        elif warning:
            ws.cell(row=row_number, column=warning_index).value = _merge_warnings(
                ws.cell(row=row_number, column=warning_index).value,
                warning,
            )
    return prefilled_count


def _build_entity_split(
    rows: list[UnitRow],
    clusters: list[EntityCluster],
) -> tuple[
    list[UnitRow],
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
]:
    entity_rows: list[UnitRow] = []
    structures: list[dict[str, object]] = []
    source_map: list[dict[str, object]] = []
    term_occurrences: dict[str, Counter[str]] = defaultdict(Counter)
    assigned_indexes: set[int] = set()

    for structure_index, cluster in enumerate(clusters, start=1):
        structure_id = f"ES{structure_index:04d}"
        structure_rows: list[UnitRow] = []
        for cluster_row_number in cluster.row_numbers:
            row_index = cluster_row_number - 2
            if row_index < 0 or row_index >= len(rows):
                continue
            unit_row = rows[row_index]
            if unit_row.original_index in assigned_indexes:
                continue
            entities = _extract_entities(cluster.source_pattern, unit_row.source_unit)
            if entities is None:
                continue
            assigned_indexes.add(unit_row.original_index)
            entity_rows.append(unit_row)
            structure_rows.append(unit_row)
            for source_entity in entities.values():
                term_occurrences[source_entity][structure_id] += 1
            source_map.append(
                {
                    schema.ORIGINAL_INDEX_COLUMN: unit_row.original_index,
                    schema.UNIT_ID_COLUMN: unit_row.unit_id,
                    schema.UNIT_TYPE_COLUMN: unit_row.unit_type,
                    schema.SOURCE_UNIT_COLUMN: unit_row.source_unit,
                    schema.STRUCTURE_ID_COLUMN: structure_id,
                    schema.ENTITIES_JSON_COLUMN: json.dumps(
                        entities,
                        ensure_ascii=False,
                        sort_keys=True,
                    ),
                    schema.PREVIEW_TARGET_COLUMN: None,
                    schema.FILL_STATUS_COLUMN: None,
                    schema.WARNING_COLUMN: None,
                }
            )
        if not structure_rows:
            continue
        structures.append(
            {
                schema.STRUCTURE_ID_COLUMN: structure_id,
                schema.SOURCE_STRUCTURE_COLUMN: cluster.source_pattern,
                schema.TARGET_STRUCTURE_COLUMN: None,
                schema.COVERAGE_COUNT_COLUMN: len(structure_rows),
                schema.CONFIDENCE_COLUMN: cluster.confidence,
                schema.RISK_COLUMN: cluster.risk or None,
                schema.STATUS_COLUMN: _default_structure_status(cluster),
                schema.SAMPLE_SOURCES_COLUMN: "\n".join(
                    dict.fromkeys(row.source_unit for row in structure_rows[:10])
                ),
                schema.ROW_NUMBERS_COLUMN: ",".join(
                    str(row.original_index) for row in structure_rows
                ),
                schema.WARNING_COLUMN: cluster.risk or None,
            }
        )

    terms = [
        {
            schema.TERM_ID_COLUMN: f"ET{index:04d}",
            schema.SOURCE_ENTITY_COLUMN: source_entity,
            schema.TARGET_ENTITY_COLUMN: None,
            schema.OCCURRENCE_COUNT_COLUMN: sum(structures_count.values()),
            schema.STRUCTURE_IDS_COLUMN: ",".join(sorted(structures_count)),
            schema.STATUS_COLUMN: "review",
            schema.WARNING_COLUMN: None,
        }
        for index, (source_entity, structures_count) in enumerate(
            sorted(term_occurrences.items()),
            start=1,
        )
    ]
    entity_rows.sort(key=lambda row: row.original_index)
    source_map.sort(key=lambda row: int(row[schema.ORIGINAL_INDEX_COLUMN]))
    return entity_rows, structures, terms, source_map


def _default_structure_status(cluster: EntityCluster) -> str:
    if cluster.confidence >= 0.85 and not cluster.risk:
        return "ready"
    return "review"


def _extract_entities(source_structure: str, source_unit: str) -> dict[str, str] | None:
    placeholders = re.findall(r"\{entity\d+\}", source_structure)
    if not placeholders:
        return None
    regex_parts: list[str] = ["^"]
    position = 0
    used_names: set[str] = set()
    for found in re.finditer(r"\{entity\d+\}", source_structure):
        regex_parts.append(re.escape(source_structure[position : found.start()]))
        name = found.group(0)[1:-1]
        if name in used_names:
            regex_parts.append(f"(?P={name})")
        else:
            regex_parts.append(f"(?P<{name}>.+?)")
            used_names.add(name)
        position = found.end()
    regex_parts.append(re.escape(source_structure[position:]))
    regex_parts.append("$")
    matched = re.match("".join(regex_parts), source_unit)
    if not matched:
        return None
    return {name: matched.group(name) for name in sorted(used_names)}


def _write_entity_related_workbook(
    output_path: Path,
    input_path: Path,
    headers: list[str],
    rows: list[UnitRow],
    structures: list[dict[str, object]],
    terms: list[dict[str, object]],
    source_map: list[dict[str, object]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = schema.TO_TRANSLATE_SHEET
    _append_unit_rows(ws, headers, rows, include_original_index=True)
    _append_dict_sheet(wb.create_sheet(schema.ENTITY_STRUCTURES_SHEET), ENTITY_STRUCTURE_COLUMNS, structures)
    _append_dict_sheet(wb.create_sheet(schema.ENTITY_TERMS_SHEET), ENTITY_TERM_COLUMNS, terms)
    _append_dict_sheet(wb.create_sheet(schema.ENTITY_SOURCE_MAP_SHEET), ENTITY_SOURCE_MAP_COLUMNS, source_map)
    _copy_support_sheets(input_path, wb, exclude={schema.TO_TRANSLATE_SHEET})
    _save_workbook(wb, output_path)


def _write_non_entity_workbook(
    output_path: Path,
    input_path: Path,
    headers: list[str],
    rows: list[UnitRow],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = schema.TO_TRANSLATE_SHEET
    _append_unit_rows(ws, headers, rows, include_original_index=True)
    _copy_support_sheets(input_path, wb, exclude={schema.TO_TRANSLATE_SHEET})
    _save_workbook(wb, output_path)


def _write_entity_tm_workbook(
    output_path: Path,
    structures: list[dict[str, object]],
    terms: list[dict[str, object]],
) -> None:
    wb = Workbook()
    structures_ws = wb.active
    structures_ws.title = schema.ENTITY_STRUCTURES_SHEET
    _append_dict_sheet(structures_ws, ENTITY_STRUCTURE_COLUMNS, structures)
    _append_dict_sheet(wb.create_sheet(schema.ENTITY_TERMS_SHEET), ENTITY_TERM_COLUMNS, terms)
    _save_workbook(wb, output_path)


def _append_unit_rows(ws, headers: list[str], rows: list[UnitRow], *, include_original_index: bool) -> None:
    output_headers = (
        [schema.ORIGINAL_INDEX_COLUMN] + headers
        if include_original_index and schema.ORIGINAL_INDEX_COLUMN not in headers
        else headers
    )
    ws.append(output_headers)
    for row in rows:
        values = []
        for header in output_headers:
            if header == schema.ORIGINAL_INDEX_COLUMN:
                values.append(row.original_index)
            else:
                values.append(row.values.get(header))
        ws.append(values)


def _append_dict_sheet(ws, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _copy_support_sheets(input_path: Path, target_wb, *, exclude: set[str]) -> None:
    source_wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        for source_ws in source_wb.worksheets:
            if source_ws.title in exclude:
                continue
            target_ws = target_wb.create_sheet(source_ws.title)
            target_ws.sheet_state = source_ws.sheet_state
            for row in source_ws.iter_rows(values_only=True):
                target_ws.append(row)
    finally:
        source_wb.close()


def _save_workbook(wb, output_path: Path) -> None:
    for ws in wb.worksheets:
        _style_sheet(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _merge_warnings(*warnings: object) -> str | None:
    merged = [str(warning) for warning in warnings if warning]
    if not merged:
        return None
    return "; ".join(dict.fromkeys(merged))


__all__ = [
    "ClusterProbeStrategy",
    "EntityExtractionStrategy",
    "default_entity_filled_pack_output_path",
    "default_entity_memory_output_path",
    "default_entity_merged_todo_output_path",
    "default_entity_pack_output_path",
    "extract_entity_memory_workbook",
    "extract_entity_tm_workbook",
    "fill_entity_workbook",
    "merge_entity_workbooks",
    "prefill_entity_workbook",
    "split_entity_workbook",
]
