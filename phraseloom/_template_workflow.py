from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


VAR_RE = re.compile(
    r"\{[^{}]+\}|#[0-9A-Fa-f]{6}|\d+(?:[./:-]\d+)+|\d+(?:\.\d+)?"
)
PLACEHOLDER_RE = re.compile(r"\{[A-Za-z_][A-Za-z0-9_]*\}")
NAMED_PLACEHOLDER_RE = re.compile(r"^\{([A-Za-z_][A-Za-z0-9_]*)\}$")


@dataclass(frozen=True)
class TemplateMatch:
    text: str
    template: str
    values: dict[str, str]


@dataclass(frozen=True)
class RowItem:
    row_number: int
    source: str
    existing_target: str
    match: TemplateMatch
    original_values: tuple[object, ...]


@dataclass(frozen=True)
class TranslationUnit:
    unit_id: str
    unit_type: str
    source_unit: str
    coverage_count: int
    unique_source_count: int
    items: tuple[RowItem, ...]
    target_unit: str
    target_unit_source: str
    suggested_target_unit: str
    warning: str


def parse_template(text: object) -> TemplateMatch:
    source = "" if text is None else str(text)
    chunks: list[str] = []
    values: dict[str, str] = {}
    pos = 0
    counters: dict[str, int] = defaultdict(int)

    for found in VAR_RE.finditer(source):
        chunks.append(source[pos : found.start()])
        value = found.group(0)
        key = _variable_key(value, counters)
        chunks.append("{" + key + "}")
        values[key] = value
        pos = found.end()

    chunks.append(source[pos:])
    return TemplateMatch(source, "".join(chunks), values)


def _variable_key(value: str, counters: dict[str, int]) -> str:
    named = NAMED_PLACEHOLDER_RE.match(value)
    if named:
        return named.group(1)
    if value.startswith("#"):
        prefix = "color"
    elif re.fullmatch(r"\d+-\d+(?:-\d+)*", value):
        prefix = "stage"
    elif re.fullmatch(r"\d+(?:[./:]\d+)+", value):
        prefix = "seq"
    else:
        prefix = "num"
    counters[prefix] += 1
    return f"{prefix}{counters[prefix]}"


def infer_target_template(values: dict[str, str], target_text: object) -> str | None:
    target_template = "" if target_text is None else str(target_text)
    matched = False
    tokens: dict[str, str] = {}

    for index, (key, value) in enumerate(
        sorted(values.items(), key=lambda item: len(item[1]), reverse=True)
    ):
        if not value:
            continue
        token = "\x00" + _letters_token(index) + "\x00"
        if value in target_template:
            target_template = target_template.replace(value, token)
            tokens[token] = "{" + key + "}"
            matched = True

    for token, placeholder in tokens.items():
        target_template = target_template.replace(token, placeholder)

    return target_template if matched else None


def _letters_token(index: int) -> str:
    letters = []
    value = index
    while True:
        letters.append(chr(ord("A") + (value % 26)))
        value = value // 26 - 1
        if value < 0:
            return "".join(reversed(letters))


def apply_target_template(target_template: str, values: dict[str, str]) -> str:
    result = target_template
    for key, value in values.items():
        result = result.replace("{" + key + "}", value)
    return result


def generate_workbook(
    input_path: str | Path,
    output_path: str | Path,
    *,
    source_col: str | int = "英語",
    target_col: str | int | None = "target",
    examples: Iterable[tuple[str, str]] = (),
    template_workbook: str | Path | None = None,
    tm_workbook: str | Path | None = None,
    min_group_size: int = 2,
    use_existing_targets: bool = True,
) -> dict[str, int]:
    input_path = Path(input_path)
    output_path = Path(output_path)

    rows, units, result_rows, autofilled_count = _build_fill_context(
        input_path,
        source_col=source_col,
        target_col=target_col,
        examples=examples,
        template_workbook=template_workbook,
        tm_workbook=tm_workbook,
        min_group_size=min_group_size,
        use_existing_targets=use_existing_targets,
    )

    _write_output_workbook(output_path, input_path, units, result_rows)
    to_translate_path = _default_to_translate_output_path(input_path)
    if not template_workbook or Path(template_workbook).resolve() != to_translate_path.resolve():
        _write_to_translate_workbook(to_translate_path, input_path, units)

    stats = _workbook_stats(rows, units, autofilled_count)
    stats["to_translate_path"] = str(to_translate_path)
    return stats


def fill_target_column_workbook(
    input_path: str | Path,
    output_path: str | Path,
    *,
    source_col: str | int = "英語",
    target_col: str | int = "target",
    template_workbook: str | Path,
    tm_workbook: str | Path | None = None,
    min_group_size: int = 2,
) -> dict[str, int]:
    input_path = Path(input_path)
    output_path = Path(output_path)

    rows, units, result_rows, autofilled_count = _build_fill_context(
        input_path,
        source_col=source_col,
        target_col=target_col,
        template_workbook=template_workbook,
        tm_workbook=tm_workbook,
        min_group_size=min_group_size,
        use_existing_targets=False,
    )
    _write_target_column_workbook(output_path, input_path, target_col, result_rows)
    return _workbook_stats(rows, units, autofilled_count)


def _build_fill_context(
    input_path: Path,
    *,
    source_col: str | int,
    target_col: str | int | None,
    examples: Iterable[tuple[str, str]] = (),
    template_workbook: str | Path | None = None,
    tm_workbook: str | Path | None = None,
    min_group_size: int,
    use_existing_targets: bool,
) -> tuple[
    list[RowItem],
    list[TranslationUnit],
    list[tuple[RowItem, TranslationUnit | None, str | None]],
    int,
]:
    rows = _read_source_rows(input_path, source_col, target_col)
    provided_units, provided_sources = _build_provided_units(
        examples, template_workbook, tm_workbook
    )
    units = _build_translation_units(
        rows,
        min_group_size,
        provided_units,
        provided_sources,
        use_existing_targets,
    )

    unit_by_row_number = {
        item.row_number: unit for unit in units for item in unit.items
    }
    result_rows = []
    autofilled_count = 0
    for row in rows:
        unit = unit_by_row_number.get(row.row_number)
        target_template = unit.target_unit if unit else ""
        auto_target = (
            apply_target_template(target_template, row.match.values)
            if target_template and unit and unit.unit_type == "template"
            else target_template
            if target_template
            else None
        )
        if auto_target:
            autofilled_count += 1
        result_rows.append((row, unit, auto_target))

    return rows, units, result_rows, autofilled_count


def _workbook_stats(
    rows: list[RowItem], units: list[TranslationUnit], autofilled_count: int
) -> dict[str, int]:
    template_units = [unit for unit in units if unit.unit_type == "template"]
    segment_units = [unit for unit in units if unit.unit_type == "segment"]
    template_source_segments = sum(unit.coverage_count for unit in template_units)
    segment_source_segments = sum(unit.coverage_count for unit in segment_units)
    unique_source_segments = len({row.source for row in rows})
    prefilled_translation_unit_count = sum(1 for unit in units if unit.target_unit)
    return {
        "row_count": len(rows),
        "unique_source_segments": unique_source_segments,
        "duplicate_source_segments": len(rows) - unique_source_segments,
        "template_count": len(template_units),
        "template_unit_count": len(template_units),
        "template_source_segments": template_source_segments,
        "template_unique_source_segments": sum(
            unit.unique_source_count for unit in template_units
        ),
        "segment_unit_count": len(segment_units),
        "segment_source_segments": segment_source_segments,
        "translation_unit_count": len(units),
        "prefilled_translation_unit_count": prefilled_translation_unit_count,
        "untranslated_translation_unit_count": len(units)
        - prefilled_translation_unit_count,
        "new_translation_unit_count": len(units) - prefilled_translation_unit_count,
        "new_source_segment_count": sum(
            unit.coverage_count for unit in units if not unit.target_unit
        ),
        "tm_unit_hit_rate": _format_rate(prefilled_translation_unit_count, len(units)),
        "tm_row_hit_rate": _format_rate(autofilled_count, len(rows)),
        "clustered_source_segments": template_source_segments,
        "unclustered_source_segments": segment_source_segments,
        "autofilled_count": autofilled_count,
    }


def generate_tm_pairs(
    input_path: str | Path,
    output_path: str | Path,
    *,
    source_col: str | int = "英語",
    target_col: str | int = "target",
    min_group_size: int = 2,
) -> dict[str, int]:
    input_path = Path(input_path)
    output_path = Path(output_path)
    rows = [
        row
        for row in _read_source_rows(input_path, source_col, target_col)
        if row.existing_target
    ]
    units = _build_translation_units(
        rows,
        min_group_size,
        provided_units={},
        provided_sources={},
        use_existing_targets=True,
    )
    _write_tm_workbook(output_path, input_path, units, rows)

    template_pairs = [unit for unit in units if unit.unit_type == "template"]
    segment_pairs = [unit for unit in units if unit.unit_type == "segment"]
    unique_source_segments = len({row.source for row in rows})
    return {
        "row_count": len(rows),
        "unique_source_segments": unique_source_segments,
        "duplicate_source_segments": len(rows) - unique_source_segments,
        "tm_pair_count": len(units),
        "template_pair_count": len(template_pairs),
        "segment_pair_count": len(segment_pairs),
        "matched_source_segments": sum(unit.coverage_count for unit in units),
    }


def _read_source_rows(
    input_path: Path, source_col: str | int, target_col: str | int | None
) -> list[RowItem]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    source_index = _resolve_column(ws, source_col)
    target_index = _resolve_column(ws, target_col) if target_col is not None else None

    rows: list[RowItem] = []
    seen_source = False
    blank_source_run = 0
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        source_value = _cell_value(row, source_index)
        if source_value is None or str(source_value).strip() == "":
            if seen_source:
                blank_source_run += 1
                if blank_source_run >= 1000:
                    break
            continue
        seen_source = True
        blank_source_run = 0
        source = str(source_value).strip()
        target_value = _cell_value(row, target_index) if target_index else ""
        existing_target = "" if target_value is None else str(target_value).strip()
        rows.append(
            RowItem(row_number, source, existing_target, parse_template(source), tuple(row))
        )

    return rows


def _eligible_groups(
    rows: Iterable[RowItem], min_group_size: int
) -> dict[str, list[RowItem]]:
    grouped: dict[str, list[RowItem]] = defaultdict(list)
    for row in rows:
        if _is_candidate(row.match):
            grouped[row.match.template].append(row)
    return {
        template: items
        for template, items in grouped.items()
        if len({item.source for item in items}) >= min_group_size
    }


def _build_translation_units(
    rows: list[RowItem],
    min_group_size: int,
    provided_units: dict[tuple[str, str], str],
    provided_sources: dict[tuple[str, str], str],
    use_existing_targets: bool,
) -> list[TranslationUnit]:
    template_groups = _eligible_groups(rows, min_group_size)
    assigned_row_numbers: set[int] = set()
    units: list[TranslationUnit] = []

    for row in rows:
        key = ("template", row.match.template)
        if _is_candidate(row.match) and key in provided_units:
            existing = template_groups.setdefault(row.match.template, [])
            if row not in existing:
                existing.append(row)

    for index, (source_unit, items) in enumerate(
        sorted(template_groups.items(), key=lambda item: (-len(item[1]), item[0])),
        start=1,
    ):
        for item in items:
            assigned_row_numbers.add(item.row_number)
        suggested = _suggest_template_target_unit(items) if use_existing_targets else ""
        key = ("template", source_unit)
        target_unit = provided_units.get(key) or suggested
        target_unit_source = (
            provided_sources.get(key, "")
            if key in provided_units
            else "existing_target"
            if suggested
            else ""
        )
        warning = _unit_warning("template", source_unit, target_unit, suggested, items)
        units.append(
            TranslationUnit(
                unit_id=f"T{index:04d}",
                unit_type="template",
                source_unit=source_unit,
                coverage_count=len(items),
                unique_source_count=len({item.source for item in items}),
                items=tuple(items),
                target_unit=target_unit,
                target_unit_source=target_unit_source,
                suggested_target_unit=suggested,
                warning=warning,
            )
        )

    segment_groups: dict[str, list[RowItem]] = defaultdict(list)
    for row in rows:
        if row.row_number not in assigned_row_numbers:
            segment_groups[row.source].append(row)

    for index, (source_unit, items) in enumerate(
        sorted(segment_groups.items(), key=lambda item: (-len(item[1]), item[0])),
        start=1,
    ):
        suggested = _suggest_segment_target_unit(items) if use_existing_targets else ""
        key = ("segment", source_unit)
        if key in provided_units:
            target_unit = provided_units[key]
            target_unit_source = provided_sources.get(key, "")
        elif suggested:
            target_unit = suggested
            target_unit_source = "existing_target"
        elif _is_non_translatable_segment(source_unit):
            target_unit = source_unit
            target_unit_source = "non_translatable"
        else:
            target_unit = ""
            target_unit_source = ""
        warning = _unit_warning("segment", source_unit, target_unit, suggested, items)
        units.append(
            TranslationUnit(
                unit_id=f"S{index:04d}",
                unit_type="segment",
                source_unit=source_unit,
                coverage_count=len(items),
                unique_source_count=1,
                items=tuple(items),
                target_unit=target_unit,
                target_unit_source=target_unit_source,
                suggested_target_unit=suggested,
                warning=warning,
            )
        )

    return units


def _is_non_translatable_segment(source: str) -> bool:
    text = source.strip()
    if not text:
        return True
    return not re.search(
        r"[A-Za-z\u3400-\u9fff\u3040-\u30ff\uac00-\ud7af]", text
    )


def _build_provided_units(
    examples: Iterable[tuple[str, str]],
    template_workbook: str | Path | None,
    tm_workbook: str | Path | None = None,
) -> tuple[dict[tuple[str, str], str], dict[tuple[str, str], str]]:
    provided: dict[tuple[str, str], str] = {}
    sources: dict[tuple[str, str], str] = {}

    if tm_workbook:
        for key, target_unit in _load_translated_units(Path(tm_workbook)).items():
            provided[key] = target_unit
            sources[key] = "tm_pairs"

    if template_workbook:
        for key, target_unit in _load_translated_units(
            Path(template_workbook)
        ).items():
            provided[key] = target_unit
            sources[key] = "translation_units"

    for source, target in examples:
        match = parse_template(source)
        target_template = infer_target_template(match.values, target)
        if target_template:
            key = ("template", match.template)
            provided[key] = target_template
            sources[key] = f"example: {source} => {target}"
        else:
            key = ("segment", source)
            provided[key] = target
            sources[key] = f"example: {source} => {target}"

    return provided, sources


def _load_translated_units(path: Path) -> dict[tuple[str, str], str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "tm_pairs" in wb.sheetnames:
        ws = wb["tm_pairs"]
        headers = [
            str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]
        ]
        try:
            type_index = headers.index("unit_type")
            source_index = headers.index("source_unit")
            target_index = headers.index("target_unit")
        except ValueError:
            return {}

        units: dict[tuple[str, str], str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            unit_type = row[type_index] if type_index < len(row) else None
            source_unit = row[source_index] if source_index < len(row) else None
            target_unit = row[target_index] if target_index < len(row) else None
            if unit_type and source_unit and target_unit:
                units[(str(unit_type), str(source_unit))] = str(target_unit)
        return units

    if "translation_units" in wb.sheetnames:
        ws = wb["translation_units"]
        headers = [
            str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]
        ]
        try:
            type_index = headers.index("unit_type")
            source_index = headers.index("source_unit")
            target_index = headers.index("target_unit")
        except ValueError:
            return {}

        units: dict[tuple[str, str], str] = {}
        status_index = headers.index("status") if "status" in headers else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            unit_type = row[type_index] if type_index < len(row) else None
            source_unit = row[source_index] if source_index < len(row) else None
            target_unit = row[target_index] if target_index < len(row) else None
            status = (
                row[status_index]
                if status_index is not None and status_index < len(row)
                else None
            )
            if (
                unit_type
                and source_unit
                and target_unit
                and str(status or "").strip().lower() != "skip"
            ):
                units[(str(unit_type), str(source_unit))] = str(target_unit)
        return units

    if "to_translate" in wb.sheetnames:
        units = _load_unit_sheet(wb["to_translate"])
        if "prefilled_units" in wb.sheetnames:
            units.update(_load_unit_sheet(wb["prefilled_units"]))
        return units

    if "template_review" not in wb.sheetnames:
        return {}
    ws = wb["template_review"]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    try:
        source_index = headers.index("source_template")
        target_index = headers.index("target_template")
    except ValueError:
        return {}

    templates: dict[tuple[str, str], str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        source_template = row[source_index] if source_index < len(row) else None
        target_template = row[target_index] if target_index < len(row) else None
        if source_template and target_template:
            templates[("template", str(source_template))] = str(target_template)
    return templates


def _load_unit_sheet(ws) -> dict[tuple[str, str], str]:
    headers = [
        str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]
    ]
    try:
        type_index = headers.index("unit_type")
        source_index = headers.index("source_unit")
        target_index = headers.index("target_unit")
    except ValueError:
        return {}

    units: dict[tuple[str, str], str] = {}
    status_index = headers.index("status") if "status" in headers else None
    for row in ws.iter_rows(min_row=2, values_only=True):
        unit_type = row[type_index] if type_index < len(row) else None
        source_unit = row[source_index] if source_index < len(row) else None
        target_unit = row[target_index] if target_index < len(row) else None
        status = (
            row[status_index]
            if status_index is not None and status_index < len(row)
            else None
        )
        if (
            unit_type
            and source_unit
            and target_unit
            and str(status or "").strip().lower() != "skip"
        ):
            units[(str(unit_type), str(source_unit))] = str(target_unit)
    return units


def _suggest_template_target_unit(items: Iterable[RowItem]) -> str:
    suggestions: list[str] = []
    for item in items:
        if not item.existing_target:
            continue
        inferred = infer_target_template(item.match.values, item.existing_target)
        if inferred:
            suggestions.append(inferred)
    if not suggestions:
        return ""
    return Counter(suggestions).most_common(1)[0][0]


def _suggest_segment_target_unit(items: Iterable[RowItem]) -> str:
    suggestions = [item.existing_target for item in items if item.existing_target]
    if not suggestions:
        return ""
    return Counter(suggestions).most_common(1)[0][0]


def _unit_warning(
    unit_type: str,
    source_unit: str,
    target_unit: str,
    suggested_target_unit: str,
    items: Iterable[RowItem],
) -> str:
    warnings: list[str] = []
    source_placeholders = set(PLACEHOLDER_RE.findall(source_unit))
    target_placeholders = set(PLACEHOLDER_RE.findall(target_unit))

    if unit_type == "template" and target_unit and source_placeholders - target_placeholders:
        warnings.append("target_unit is missing source variables")
    if "$" in source_unit:
        warnings.append("price-like text; review manually")
    if re.search(r"\b1\s+(day|time|attempt|task|star|pack)s\b", source_unit):
        warnings.append("plural-sensitive text; review manually")

    inferred = []
    if unit_type == "template":
        for item in items:
            if item.existing_target:
                guess = infer_target_template(item.match.values, item.existing_target)
                if guess:
                    inferred.append(guess)
    else:
        inferred = [item.existing_target for item in items if item.existing_target]
    if suggested_target_unit and len(set(inferred)) > 1:
        warnings.append("multiple existing target patterns found")

    return "; ".join(warnings)


def _is_candidate(match: TemplateMatch) -> bool:
    if not match.values:
        return False
    literal = PLACEHOLDER_RE.sub("", match.template)
    literal = re.sub(r"\s+", "", literal)
    return len(literal) >= 2


def _resolve_column(ws, col: str | int | None) -> int:
    if col is None:
        raise ValueError("Column cannot be None")
    if isinstance(col, int):
        return col
    if str(col).isdigit():
        return int(str(col))

    wanted = str(col).strip()
    wanted_lower = wanted.lower()
    for cell in ws[1]:
        value = "" if cell.value is None else str(cell.value).strip()
        if value == wanted or value.lower() == wanted_lower:
            return cell.column
    raise ValueError(f"Column {col!r} not found in header row")


def _cell_value(row: tuple[object, ...], one_based_index: int | None) -> object | None:
    if one_based_index is None:
        return None
    zero_based = one_based_index - 1
    return row[zero_based] if zero_based < len(row) else None


def _read_headers(input_path: Path) -> list[str]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    return [
        str(cell.value).strip() if cell.value is not None else f"column_{cell.column}"
        for cell in ws[1]
    ]


def _format_rate(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return "0.0%"
    return f"{numerator / denominator * 100:.1f}%"


def _write_output_workbook(
    output_path: Path,
    input_path: Path,
    units: list[TranslationUnit],
    result_rows: list[tuple[RowItem, TranslationUnit | None, str | None]],
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "summary"

    template_units = [unit for unit in units if unit.unit_type == "template"]
    segment_units = [unit for unit in units if unit.unit_type == "segment"]
    total_source_segments = len(result_rows)
    prefilled_units = sum(1 for unit in units if unit.target_unit)
    new_units = sum(1 for unit in units if not unit.target_unit)
    filled_rows = sum(1 for _, _, auto in result_rows if auto)
    new_rows = sum(unit.coverage_count for unit in units if not unit.target_unit)
    summary_rows = [
        ("total_source_rows", total_source_segments),
        ("total_translation_units", len(units)),
        ("already_filled_units", prefilled_units),
        ("already_filled_source_rows", filled_rows),
        ("units_to_translate", new_units),
        ("source_rows_to_translate", new_rows),
    ]
    for row in summary_rows:
        summary.append(row)

    review = wb.create_sheet("translation_units")
    review.append(
        [
            "unit_id",
            "unit_type",
            "source_unit",
            "target_unit",
            "coverage_count",
            "unique_source_count",
            "variables",
            "status",
            "sample_sources",
            "row_numbers",
            "target_unit_source",
            "suggested_target_unit",
            "warning",
            "translator_note",
        ]
    )
    for unit in units:
        review.append(
            [
                unit.unit_id,
                unit.unit_type,
                unit.source_unit,
                unit.target_unit or None,
                unit.coverage_count,
                unit.unique_source_count,
                _variables_summary(unit),
                "ready" if unit.target_unit else None,
                "\n".join(dict.fromkeys(item.source for item in unit.items[:10])),
                ",".join(str(item.row_number) for item in unit.items),
                unit.target_unit_source or None,
                unit.suggested_target_unit or None,
                unit.warning or None,
                None,
            ]
        )

    todo = wb.create_sheet("to_translate")
    todo.append(
        [
            "unit_id",
            "unit_type",
            "source_unit",
            "target_unit",
            "coverage_count",
            "sample_sources",
            "variables",
            "warning",
            "translator_note",
        ]
    )
    for unit in units:
        if unit.target_unit:
            continue
        todo.append(
            [
                unit.unit_id,
                unit.unit_type,
                unit.source_unit,
                None,
                unit.coverage_count,
                "\n".join(dict.fromkeys(item.source for item in unit.items[:5])),
                _variables_summary(unit),
                unit.warning or None,
                None,
            ]
        )

    result = wb.create_sheet("source_map")
    result.append(
        [
            "row_number",
            "source",
            "unit_id",
            "unit_type",
            "source_unit",
            "variable_values",
            "target_unit",
            "auto_target",
            "existing_target",
            "fill_status",
            "warning",
        ]
    )
    for row, unit, auto_target in result_rows:
        if unit is None:
            warning = "no translation unit"
            fill_status = "unit_not_found"
        elif not unit.target_unit:
            warning = "fill target_unit in translation_units, then rerun fill"
            fill_status = "missing_target_unit"
        else:
            warning = unit.warning
            fill_status = "filled"

        result.append(
            [
                row.row_number,
                row.source,
                unit.unit_id if unit else None,
                unit.unit_type if unit else None,
                unit.source_unit if unit else None,
                json.dumps(row.match.values if unit and unit.unit_type == "template" else {}, ensure_ascii=False),
                unit.target_unit if unit else None,
                auto_target,
                row.existing_target or None,
                fill_status,
                warning or None,
            ]
        )

    filled = wb.create_sheet("filled_workbook")
    headers = _read_headers(input_path)
    filled.append(
        headers
        + [
            "auto_target",
            "fill_status",
            "unit_id",
            "unit_type",
            "warning",
        ]
    )
    for row, unit, auto_target in result_rows:
        if unit is None:
            warning = "no translation unit"
            fill_status = "unit_not_found"
        elif not unit.target_unit:
            warning = "fill target_unit in to_translate, then rerun fill"
            fill_status = "missing_target_unit"
        else:
            warning = unit.warning
            fill_status = "filled"
        original = list(row.original_values)
        if len(original) < len(headers):
            original.extend([None] * (len(headers) - len(original)))
        filled.append(
            original[: len(headers)]
            + [
                auto_target,
                fill_status,
                unit.unit_id if unit else None,
                unit.unit_type if unit else None,
                warning or None,
            ]
        )

    qa = wb.create_sheet("qa_report")
    qa.append(["check", "count"])
    qa.append(["missing_target_unit", sum(1 for _, unit, auto in result_rows if unit and not unit.target_unit)])
    qa.append(["filled", sum(1 for _, _, auto in result_rows if auto)])
    qa.append(["warning_units", sum(1 for unit in units if unit.warning)])
    qa.append(["template_units", len(template_units)])
    qa.append(["segment_units", len(segment_units)])

    for ws in wb.worksheets:
        _style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_to_translate_workbook(
    output_path: Path, input_path: Path, units: list[TranslationUnit]
) -> None:
    wb = Workbook()
    todo = wb.active
    todo.title = "to_translate"
    todo.append(
        [
            "unit_id",
            "unit_type",
            "source_unit",
            "target_unit",
            "coverage_count",
            "sample_sources",
            "variables",
            "warning",
            "translator_note",
        ]
    )
    for unit in units:
        if unit.target_unit:
            continue
        todo.append(
            [
                unit.unit_id,
                unit.unit_type,
                unit.source_unit,
                None,
                unit.coverage_count,
                "\n".join(dict.fromkeys(item.source for item in unit.items[:5])),
                _variables_summary(unit),
                unit.warning or None,
                None,
            ]
        )

    prefilled = wb.create_sheet("prefilled_units")
    prefilled.append(
        [
            "unit_id",
            "unit_type",
            "source_unit",
            "target_unit",
            "coverage_count",
            "target_unit_source",
        ]
    )
    for unit in units:
        if not unit.target_unit:
            continue
        prefilled.append(
            [
                unit.unit_id,
                unit.unit_type,
                unit.source_unit,
                unit.target_unit,
                unit.coverage_count,
                unit.target_unit_source or None,
            ]
        )
    prefilled.sheet_state = "hidden"

    for ws in wb.worksheets:
        _style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_target_column_workbook(
    output_path: Path,
    input_path: Path,
    target_col: str | int,
    result_rows: list[tuple[RowItem, TranslationUnit | None, str | None]],
) -> None:
    wb = load_workbook(input_path)
    ws = wb.worksheets[0]
    target_index = _resolve_or_create_column(ws, target_col)

    for row, _, auto_target in result_rows:
        if auto_target:
            ws.cell(row=row.row_number, column=target_index).value = auto_target

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _resolve_or_create_column(ws, col: str | int) -> int:
    try:
        return _resolve_column(ws, col)
    except ValueError:
        if isinstance(col, int) or str(col).isdigit():
            raise
        column = ws.max_column + 1
        ws.cell(row=1, column=column).value = str(col)
        return column


def _write_tm_workbook(
    output_path: Path,
    input_path: Path,
    units: list[TranslationUnit],
    rows: list[RowItem],
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "summary"

    template_units = [unit for unit in units if unit.unit_type == "template"]
    segment_units = [unit for unit in units if unit.unit_type == "segment"]
    unique_source_segments = len({row.source for row in rows})
    summary_rows = [
        ("input_file", str(input_path)),
        ("tm_source_segments", len(rows)),
        ("unique_source_segments", unique_source_segments),
        ("duplicate_source_segments", len(rows) - unique_source_segments),
        ("tm_pair_count", len(units)),
        ("template_pair_count", len(template_units)),
        ("segment_pair_count", len(segment_units)),
        ("matched_source_segments", sum(unit.coverage_count for unit in units)),
        ("next_step", "Use this workbook with extract --tm to prefill new translation_units."),
    ]
    for row in summary_rows:
        summary.append(row)

    pairs = wb.create_sheet("tm_pairs")
    pairs.append(
        [
            "tm_id",
            "unit_type",
            "source_unit",
            "target_unit",
            "coverage_count",
            "unique_source_count",
            "variables",
            "sample_sources",
            "sample_targets",
            "row_numbers",
            "warning",
        ]
    )
    for index, unit in enumerate(units, start=1):
        pairs.append(
            [
                f"TM{index:05d}",
                unit.unit_type,
                unit.source_unit,
                unit.target_unit or None,
                unit.coverage_count,
                unit.unique_source_count,
                _variables_summary(unit),
                "\n".join(dict.fromkeys(item.source for item in unit.items[:10])),
                "\n".join(
                    dict.fromkeys(
                        item.existing_target for item in unit.items[:10] if item.existing_target
                    )
                ),
                ",".join(str(item.row_number) for item in unit.items),
                unit.warning or None,
            ]
        )

    tm_map = wb.create_sheet("tm_map")
    tm_map.append(
        [
            "row_number",
            "source",
            "target",
            "unit_type",
            "source_unit",
            "target_unit",
            "variable_values",
        ]
    )
    unit_by_row_number = {
        item.row_number: unit for unit in units for item in unit.items
    }
    for row in rows:
        unit = unit_by_row_number.get(row.row_number)
        tm_map.append(
            [
                row.row_number,
                row.source,
                row.existing_target,
                unit.unit_type if unit else None,
                unit.source_unit if unit else None,
                unit.target_unit if unit else None,
                json.dumps(
                    row.match.values if unit and unit.unit_type == "template" else {},
                    ensure_ascii=False,
                ),
            ]
        )

    qa = wb.create_sheet("qa_report")
    qa.append(["check", "count"])
    qa.append(["tm_pairs", len(units)])
    qa.append(["template_pairs", len(template_units)])
    qa.append(["segment_pairs", len(segment_units)])
    qa.append(["warning_pairs", sum(1 for unit in units if unit.warning)])

    for ws in wb.worksheets:
        _style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _variables_summary(unit: TranslationUnit) -> str | None:
    if unit.unit_type != "template":
        return None
    placeholders = PLACEHOLDER_RE.findall(unit.source_unit)
    if not placeholders:
        return None
    samples: dict[str, list[str]] = {placeholder: [] for placeholder in placeholders}
    for item in unit.items:
        for key, value in item.match.values.items():
            placeholder = "{" + key + "}"
            if placeholder in samples and value not in samples[placeholder]:
                samples[placeholder].append(value)
    return "; ".join(
        f"{placeholder}={','.join(values[:5])}" for placeholder, values in samples.items()
    )


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions

    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)
        for cell in column[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _parse_examples(raw_examples: Iterable[str]) -> list[tuple[str, str]]:
    examples: list[tuple[str, str]] = []
    for raw in raw_examples:
        if "=" not in raw:
            raise ValueError(f"Example must look like SOURCE=TARGET: {raw!r}")
        source, target = raw.split("=", 1)
        examples.append((source.strip(), target.strip()))
    return examples


def _default_work_dir(input_path: Path) -> Path:
    return input_path.parent / f"{input_path.stem}_l10n"


def _default_extract_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_template_pack.xlsx"


def _default_tm_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_tm_pairs.xlsx"


def _default_to_translate_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_to_translate.xlsx"


def _default_fill_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_filled.xlsx"


def _default_legacy_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_template_demo.xlsx"


def main(argv: list[str] | None = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)

    if argv and argv[0] in {"-h", "--help"}:
        _print_top_level_help()
        return 0
    if not argv or argv[0] in {"interactive", "wizard"}:
        return run_interactive()
    if argv[0] in {"tm-extract", "extract-tm"}:
        return _main_tm_extract(argv[1:])
    if argv[0] == "extract":
        return _main_extract(argv[1:])
    if argv[0] == "fill":
        return _main_fill(argv[1:])
    return _main_legacy(argv)


def run_interactive() -> int:
    print("Localization Workflow")
    print()
    print("1) Build TM from completed Excel")
    print("2) Prepare translator file for new source")
    print("3) Fill source from translated file")
    print("q) Quit")

    action = _prompt_text("Choose step", default="2").lower()
    if action in {"q", "quit", "exit"}:
        print("Bye.")
        return 0
    if action in {"1", "tm", "tm-extract", "extract-tm", "build"}:
        return _interactive_tm_extract()
    if action in {"2", "extract", "prepare", "p"}:
        return _interactive_extract()
    if action in {"3", "fill", "f"}:
        return _interactive_fill()

    print(f"Unknown step: {action}")
    return 2


def _interactive_tm_extract() -> int:
    input_path = _user_path(_prompt_text("Completed Excel path", required=True))
    source_col = _prompt_text("Source column in completed Excel", default="英語")
    target_col = _prompt_text("Target column in completed Excel", default="target")
    output_path = _user_path(
        _prompt_text("Output tm_pairs workbook", default=str(_default_tm_output_path(input_path)))
    )
    min_group_size = _prompt_int("Minimum variants for a reusable template", default=2)

    stats = generate_tm_pairs(
        input_path,
        output_path,
        source_col=source_col,
        target_col=target_col,
        min_group_size=min_group_size,
    )
    _print_tm_stats(output_path, stats)
    return 0


def _interactive_extract() -> int:
    input_path = _user_path(_prompt_text("New source Excel path", required=True))
    source_col = _prompt_text("Source column in new file", default="英語")
    target_col = _normalize_optional_column(
        _prompt_text("Existing target column (- for none)", default="target")
    )
    tm_workbook_text = _prompt_text("Existing tm_pairs path (- for none)", default="-")
    tm_workbook = (
        _user_path(tm_workbook_text)
        if _normalize_optional_column(tm_workbook_text) is not None
        else None
    )
    output_path = _user_path(
        _prompt_text(
            "Output process workbook",
            default=str(_default_extract_output_path(input_path)),
        )
    )
    min_group_size = _prompt_int("Minimum variants for a reusable template", default=2)
    use_existing_targets = (
        _prompt_yes_no("Use existing target column as template suggestions", default=True)
        if target_col is not None
        else False
    )

    stats = generate_workbook(
        input_path,
        output_path,
        source_col=source_col,
        target_col=target_col,
        tm_workbook=tm_workbook,
        min_group_size=min_group_size,
        use_existing_targets=use_existing_targets,
    )
    _print_stats(output_path, stats)
    return 0


def _interactive_fill() -> int:
    input_path = _user_path(_prompt_text("Original source Excel path", required=True))
    template_workbook = _user_path(
        _prompt_text("Translated to_translate file path", required=True)
    )
    source_col = _prompt_text("Source column in original file", default="英語")
    target_col = _normalize_optional_column(
        _prompt_text("Target column to write/check", default="target")
    )
    mode = _prompt_text("Fill mode: report or target-column", default="report")
    output_path = _user_path(
        _prompt_text("Output filled workbook", default=str(_default_fill_output_path(input_path)))
    )
    min_group_size = _prompt_int("Minimum variants for a reusable template", default=2)

    if mode == "target-column":
        if target_col is None:
            raise ValueError("target-column mode needs a target column")
        stats = fill_target_column_workbook(
            input_path,
            output_path,
            source_col=source_col,
            target_col=target_col,
            template_workbook=template_workbook,
            min_group_size=min_group_size,
        )
    else:
        stats = generate_workbook(
            input_path,
            output_path,
            source_col=source_col,
            target_col=target_col,
            template_workbook=template_workbook,
            min_group_size=min_group_size,
            use_existing_targets=False,
        )
    _print_stats(output_path, stats)
    return 0


def _main_tm_extract(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description="Extract reusable TM pairs from completed source/target columns."
    )
    parser.add_argument("input", type=Path, help="Completed TM .xlsx file")
    parser.add_argument("-o", "--output", type=Path, help="Output TM pairs .xlsx file")
    parser.add_argument("--source-col", default="英語", help="Header name or 1-based index")
    parser.add_argument("--target-col", default="target", help="Header name or 1-based index")
    parser.add_argument("--min-group-size", type=int, default=2)

    args = parser.parse_args(argv)
    output = args.output or _default_tm_output_path(args.input)
    stats = generate_tm_pairs(
        args.input,
        output,
        source_col=args.source_col,
        target_col=args.target_col,
        min_group_size=args.min_group_size,
    )
    _print_tm_stats(output, stats)
    return 0


def _main_extract(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description="Extract reusable localization templates from an Excel source column."
    )
    parser.add_argument("input", type=Path, help="Input .xlsx file")
    parser.add_argument("-o", "--output", type=Path, help="Output .xlsx file")
    parser.add_argument("--source-col", default="英語", help="Header name or 1-based index")
    parser.add_argument("--target-col", default="target", help="Header name or 1-based index")
    parser.add_argument(
        "--tm",
        type=Path,
        help="TM pairs workbook used to prefill matching translation units",
    )
    parser.add_argument(
        "--example",
        action="append",
        default=[],
        help='One source=target example, e.g. "VIP10 Paid Pack=VIP10pack"',
    )
    parser.add_argument("--min-group-size", type=int, default=2)
    parser.add_argument(
        "--no-existing-targets",
        action="store_true",
        help="Do not infer target templates from the existing target column",
    )

    args = parser.parse_args(argv)
    output = args.output or _default_extract_output_path(args.input)
    stats = generate_workbook(
        args.input,
        output,
        source_col=args.source_col,
        target_col=_normalize_optional_column(args.target_col),
        tm_workbook=args.tm,
        examples=_parse_examples(args.example),
        min_group_size=args.min_group_size,
        use_existing_targets=not args.no_existing_targets,
    )
    _print_stats(output, stats)
    return 0


def _main_fill(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description="Fill a source workbook from a translated template pack."
    )
    parser.add_argument("input", type=Path, help="Original input .xlsx file")
    parser.add_argument(
        "--templates",
        required=True,
        type=Path,
        help="Translated template pack workbook",
    )
    parser.add_argument("-o", "--output", type=Path, help="Output .xlsx file")
    parser.add_argument("--source-col", default="英語", help="Header name or 1-based index")
    parser.add_argument("--target-col", default="target", help="Header name or 1-based index")
    parser.add_argument("--min-group-size", type=int, default=2)
    parser.add_argument(
        "--mode",
        choices=["report", "target-column"],
        default="report",
        help="report creates analysis sheets; target-column writes auto targets into the target column of an output copy",
    )

    args = parser.parse_args(argv)
    output = args.output or _default_fill_output_path(args.input)
    target_col = _normalize_optional_column(args.target_col)
    if args.mode == "target-column":
        if target_col is None:
            raise ValueError("target-column mode needs a target column")
        stats = fill_target_column_workbook(
            args.input,
            output,
            source_col=args.source_col,
            target_col=target_col,
            template_workbook=args.templates,
            min_group_size=args.min_group_size,
        )
    else:
        stats = generate_workbook(
            args.input,
            output,
            source_col=args.source_col,
            target_col=target_col,
            template_workbook=args.templates,
            min_group_size=args.min_group_size,
            use_existing_targets=False,
        )
    _print_stats(output, stats)
    return 0


def _main_legacy(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description="Extract or fill reusable localization templates from an Excel source column."
    )
    parser.add_argument("input", type=Path, help="Input .xlsx file")
    parser.add_argument("-o", "--output", type=Path, help="Output .xlsx file")
    parser.add_argument("--source-col", default="英語", help="Header name or 1-based index")
    parser.add_argument("--target-col", default="target", help="Header name or 1-based index")
    parser.add_argument(
        "--templates",
        type=Path,
        help="Previous output workbook whose template_review sheet has target_template filled",
    )
    parser.add_argument(
        "--example",
        action="append",
        default=[],
        help='One source=target example, e.g. "VIP10 Paid Pack=VIP10pack"',
    )
    parser.add_argument("--min-group-size", type=int, default=2)
    parser.add_argument(
        "--no-existing-targets",
        action="store_true",
        help="Do not infer target templates from the existing target column",
    )

    args = parser.parse_args(argv)
    output = args.output or _default_legacy_output_path(args.input)
    stats = generate_workbook(
        args.input,
        output,
        source_col=args.source_col,
        target_col=_normalize_optional_column(args.target_col),
        examples=_parse_examples(args.example),
        template_workbook=args.templates,
        min_group_size=args.min_group_size,
        use_existing_targets=not args.no_existing_targets,
    )
    _print_stats(output, stats)
    return 0


def _prompt_text(prompt: str, *, default: str | None = None, required: bool = False) -> str:
    suffix = f" [{default}]" if default is not None else ""
    while True:
        value = input(f"{prompt}{suffix}: ").strip()
        if value:
            return value
        if default is not None:
            return default
        if not required:
            return ""
        print("This value is required.")


def _prompt_int(prompt: str, *, default: int) -> int:
    while True:
        raw = _prompt_text(prompt, default=str(default))
        try:
            value = int(raw)
        except ValueError:
            print("Please enter a number.")
            continue
        if value < 1:
            print("Please enter a number greater than 0.")
            continue
        return value


def _prompt_yes_no(prompt: str, *, default: bool) -> bool:
    default_text = "Y/n" if default else "y/N"
    while True:
        raw = input(f"{prompt} [{default_text}]: ").strip().lower()
        if not raw:
            return default
        if raw in {"y", "yes"}:
            return True
        if raw in {"n", "no"}:
            return False
        print("Please answer y or n.")


def _user_path(value: str | Path) -> Path:
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1].strip()
    return Path(text).expanduser()


def _normalize_optional_column(value: str | int | None) -> str | int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    normalized = value.strip()
    if normalized.lower() in {"", "-", "none", "no", "skip"}:
        return None
    return normalized


def _print_stats(output: Path, stats: dict[str, int]) -> None:
    print(f"Wrote: {output}")
    if "to_translate_path" in stats:
        print(f"To-translate workbook: {stats['to_translate_path']}")
    print(f"Units to translate: {stats['new_translation_unit_count']}")
    print(f"Source rows to translate: {stats['new_source_segment_count']}")
    print(f"Already filled units: {stats['prefilled_translation_unit_count']}")
    print(f"Already filled source rows: {stats['autofilled_count']}")
    print(f"Total translation units: {stats['translation_unit_count']}")
    print(f"Total source rows: {stats['row_count']}")


def _print_tm_stats(output: Path, stats: dict[str, int]) -> None:
    print(f"Wrote: {output}")
    print(f"TM source segments: {stats['row_count']}")
    print(f"Unique source segments: {stats['unique_source_segments']}")
    print(f"Duplicate source segments: {stats['duplicate_source_segments']}")
    print(f"TM pairs: {stats['tm_pair_count']}")
    print(f"Template pairs: {stats['template_pair_count']}")
    print(f"Segment pairs: {stats['segment_pair_count']}")


def _print_top_level_help() -> None:
    print("Localization Workflow")
    print()
    print("Interactive:")
    print("  python template_demo.py")
    print("  python template_demo.py interactive")
    print()
    print("Steps:")
    print("  1) Build TM from completed Excel")
    print("  2) Prepare translator file for new source")
    print("  3) Fill source from translated file")
    print()
    print("Commands:")
    print("  python template_demo.py tm-extract COMPLETED_TM.xlsx [options]")
    print("  python template_demo.py extract SOURCE.xlsx [options]")
    print("  python template_demo.py fill SOURCE.xlsx --templates TEMPLATE_PACK.xlsx [options]")
    print()
    print("Legacy:")
    print("  python template_demo.py SOURCE.xlsx [options]")


if __name__ == "__main__":
    raise SystemExit(main())
