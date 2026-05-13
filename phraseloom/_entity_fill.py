from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from . import workbook_schema as schema
from ._entity_io import (
    _copy_support_sheets,
    _parse_original_index,
    _read_pack_unit_rows,
    _read_unit_rows,
    _replace_workbook_atomically,
    _save_workbook,
    _worksheet_by_name,
)
from ._entity_types import UnitRow
from .errors import WorkflowError
from .excel_io import _header_values


def fill_entity_workbook(
    entity_input_path: str | Path,
    output_path: str | Path,
) -> dict[str, int | str]:
    entity_input_path = Path(entity_input_path)
    output_path = Path(output_path)
    wb = load_workbook(entity_input_path)
    try:
        structures = _load_rows_by_key(
            wb[schema.ENTITY_STRUCTURES_SHEET],
            schema.STRUCTURE_ID_COLUMN,
        )
        terms = _load_rows_by_key(
            wb[schema.ENTITY_TERMS_SHEET],
            schema.SOURCE_ENTITY_COLUMN,
        )
        todo_ws = wb[schema.TO_TRANSLATE_SHEET]
        source_map_ws = wb[schema.ENTITY_SOURCE_MAP_SHEET]
        todo_by_original_index = _todo_rows_by_original_index(todo_ws)
        filled_count = _fill_source_map_rows(
            source_map_ws,
            todo_ws,
            todo_by_original_index,
            structures,
            terms,
        )
        _save_workbook(wb, output_path)
    finally:
        wb.close()

    return {
        "filled_entity_unit_count": filled_count,
        "output_path": str(output_path),
    }


def fill_entity_pack_workbook(
    pack_input_path: str | Path,
    output_path: str | Path,
) -> dict[str, int | str]:
    pack_input_path = Path(pack_input_path)
    output_path = Path(output_path)
    wb = load_workbook(pack_input_path)
    try:
        structures = _load_rows_by_key(
            wb[schema.ENTITY_STRUCTURES_SHEET],
            schema.STRUCTURE_ID_COLUMN,
        )
        terms = _load_rows_by_key(
            wb[schema.ENTITY_TERMS_SHEET],
            schema.SOURCE_ENTITY_COLUMN,
        )
        related_ws = _worksheet_by_name(
            wb,
            [schema.RELATED_UNITS_SHEET, schema.TO_TRANSLATE_SHEET],
        )
        entity_map_ws = _worksheet_by_name(
            wb,
            [schema.ENTITY_MAP_SHEET, schema.ENTITY_SOURCE_MAP_SHEET],
        )
        related_by_original_index = _todo_rows_by_original_index(related_ws)
        filled_count = _fill_source_map_rows(
            entity_map_ws,
            related_ws,
            related_by_original_index,
            structures,
            terms,
        )
        if output_path.resolve() == pack_input_path.resolve():
            _replace_workbook_atomically(wb, output_path)
        else:
            _save_workbook(wb, output_path)
    finally:
        wb.close()

    return {
        "filled_entity_unit_count": filled_count,
        "output_path": str(output_path),
    }


def merge_entity_workbooks(
    entity_path: str | Path,
    non_entity_path: str | Path,
    output_path: str | Path,
) -> dict[str, int | str]:
    entity_path = Path(entity_path)
    non_entity_path = Path(non_entity_path)
    output_path = Path(output_path)
    entity_rows, entity_headers = _read_unit_rows(entity_path, schema.TO_TRANSLATE_SHEET)
    non_entity_rows, non_entity_headers = _read_unit_rows(
        non_entity_path,
        schema.TO_TRANSLATE_SHEET,
    )
    headers = [
        header
        for header in (non_entity_headers or entity_headers)
        if header != schema.ORIGINAL_INDEX_COLUMN
    ]
    merged_rows = _merge_unit_rows(entity_rows + non_entity_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = schema.TO_TRANSLATE_SHEET
    ws.append(headers)
    for row in merged_rows:
        ws.append([row.values.get(header) for header in headers])
    _copy_support_sheets(non_entity_path, wb, exclude={schema.TO_TRANSLATE_SHEET})
    _save_workbook(wb, output_path)
    return {
        "merged_unit_count": len(merged_rows),
        "output_path": str(output_path),
    }


def merge_entity_pack_workbook(
    pack_path: str | Path,
    output_path: str | Path,
) -> dict[str, int | str]:
    pack_path = Path(pack_path)
    output_path = Path(output_path)
    related_rows, related_headers = _read_pack_unit_rows(
        pack_path,
        schema.RELATED_UNITS_SHEET,
        schema.TO_TRANSLATE_SHEET,
    )
    non_related_rows, non_related_headers = _read_pack_unit_rows(
        pack_path,
        schema.NON_RELATED_UNITS_SHEET,
    )
    headers = [
        header
        for header in (related_headers or non_related_headers)
        if header != schema.ORIGINAL_INDEX_COLUMN
    ]
    merged_rows = _merge_unit_rows(related_rows + non_related_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = schema.TO_TRANSLATE_SHEET
    ws.append(headers)
    for row in merged_rows:
        ws.append([row.values.get(header) for header in headers])
    _copy_support_sheets(
        pack_path,
        wb,
        exclude={
            schema.RELATED_UNITS_SHEET,
            schema.NON_RELATED_UNITS_SHEET,
            schema.ENTITY_STRUCTURES_SHEET,
            schema.ENTITY_TERMS_SHEET,
            schema.ENTITY_MAP_SHEET,
            schema.TO_TRANSLATE_SHEET,
            schema.ENTITY_SOURCE_MAP_SHEET,
        },
    )
    _save_workbook(wb, output_path)
    return {
        "merged_unit_count": len(merged_rows),
        "output_path": str(output_path),
    }


def _load_rows_by_key(ws, key_column: str) -> dict[str, dict[str, object]]:
    headers = _header_values(ws)
    key_index = headers.index(key_column)
    rows: dict[str, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[key_index] if key_index < len(row) else None
        if not key:
            continue
        rows[str(key)] = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
    return rows


def _todo_rows_by_original_index(ws) -> dict[int, int]:
    headers = _header_values(ws)
    original_index_column = headers.index(schema.ORIGINAL_INDEX_COLUMN) + 1
    rows: dict[int, int] = {}
    for row_number in range(2, ws.max_row + 1):
        original_index = ws.cell(row=row_number, column=original_index_column).value
        if original_index is not None:
            rows[
                _parse_original_index(
                    original_index,
                    sheet_name=ws.title,
                    row_number=row_number,
                )
            ] = row_number
    return rows


def _merge_unit_rows(rows: list[UnitRow]) -> list[UnitRow]:
    by_index: dict[int, UnitRow] = {}
    for row in rows:
        if not row.unit_id:
            raise WorkflowError(
                f"Missing unit_id for original_index {row.original_index}"
            )
        if not row.source_unit:
            raise WorkflowError(
                f"Missing source_unit for original_index {row.original_index}"
            )
        existing = by_index.get(row.original_index)
        if existing is not None:
            if (
                existing.unit_id != row.unit_id
                or existing.source_unit != row.source_unit
            ):
                raise WorkflowError(
                    f"Conflicting rows for original_index {row.original_index}"
                )
            raise WorkflowError(f"Duplicate original_index {row.original_index}")
        by_index[row.original_index] = row

    if not by_index:
        return []
    expected = set(range(1, max(by_index) + 1))
    actual = set(by_index)
    if expected != actual:
        missing = ",".join(str(index) for index in sorted(expected - actual))
        raise WorkflowError(f"Missing original_index values: {missing}")
    return [by_index[index] for index in sorted(by_index)]


def _fill_source_map_rows(
    source_map_ws,
    todo_ws,
    todo_by_original_index: dict[int, int],
    structures: dict[str, dict[str, object]],
    terms: dict[str, dict[str, object]],
) -> int:
    map_headers = _header_values(source_map_ws)
    todo_headers = _header_values(todo_ws)
    map_columns = _one_based_columns(map_headers)
    todo_columns = _one_based_columns(todo_headers)
    filled_count = 0

    for row_number in range(2, source_map_ws.max_row + 1):
        original_index = _parse_original_index(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.ORIGINAL_INDEX_COLUMN],
            ).value,
            sheet_name=source_map_ws.title,
            row_number=row_number,
        )
        unit_id = str(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.UNIT_ID_COLUMN],
            ).value
            or ""
        )
        source_unit = str(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.SOURCE_UNIT_COLUMN],
            ).value
            or ""
        )
        structure_id = str(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.STRUCTURE_ID_COLUMN],
            ).value
            or ""
        )
        entities = json.loads(
            source_map_ws.cell(
                row=row_number,
                column=map_columns[schema.ENTITIES_JSON_COLUMN],
            ).value
            or "{}"
        )
        preview_target, fill_status, warning = _build_entity_target(
            original_index,
            unit_id,
            source_unit,
            structure_id,
            entities,
            todo_ws,
            todo_by_original_index,
            todo_columns,
            structures,
            terms,
        )
        source_map_ws.cell(
            row=row_number,
            column=map_columns[schema.PREVIEW_TARGET_COLUMN],
        ).value = preview_target
        source_map_ws.cell(
            row=row_number,
            column=map_columns[schema.FILL_STATUS_COLUMN],
        ).value = fill_status
        source_map_ws.cell(
            row=row_number,
            column=map_columns[schema.WARNING_COLUMN],
        ).value = warning
        if preview_target:
            todo_row_number = todo_by_original_index[original_index]
            todo_ws.cell(
                row=todo_row_number,
                column=todo_columns[schema.TARGET_UNIT_COLUMN],
            ).value = preview_target
            filled_count += 1
    return filled_count


def _build_entity_target(
    original_index: int,
    unit_id: str,
    source_unit: str,
    structure_id: str,
    entities: dict[str, str],
    todo_ws,
    todo_by_original_index: dict[int, int],
    todo_columns: dict[str, int],
    structures: dict[str, dict[str, object]],
    terms: dict[str, dict[str, object]],
) -> tuple[str | None, str, str | None]:
    todo_row_number = todo_by_original_index.get(original_index)
    if todo_row_number is None:
        return None, "unit_not_found", f"original_index not found: {original_index}"
    if (
        str(
            todo_ws.cell(
                row=todo_row_number,
                column=todo_columns[schema.UNIT_ID_COLUMN],
            ).value
            or ""
        )
        != unit_id
        or str(
            todo_ws.cell(
                row=todo_row_number,
                column=todo_columns[schema.SOURCE_UNIT_COLUMN],
            ).value
            or ""
        )
        != source_unit
    ):
        return None, "unit_mismatch", "unit_id/source_unit changed"

    structure = structures.get(structure_id)
    if structure is None:
        return None, "structure_not_found", f"structure not found: {structure_id}"
    target_structure = str(structure.get(schema.TARGET_STRUCTURE_COLUMN) or "")
    if not target_structure:
        return (
            None,
            "missing_structure_translation",
            f"missing target structure: {structure_id}",
        )

    target = target_structure
    for placeholder_name, source_entity in sorted(entities.items()):
        term = terms.get(str(source_entity))
        if term is None:
            return (
                None,
                "missing_entity_translation",
                f"missing entity target: {source_entity}",
            )
        target_entity = str(term.get(schema.TARGET_ENTITY_COLUMN) or "")
        if not target_entity:
            return (
                None,
                "missing_entity_translation",
                f"missing entity target: {source_entity}",
            )
        target = target.replace("{" + placeholder_name + "}", target_entity)
    return target, "filled", None


def _one_based_columns(headers: list[str]) -> dict[str, int]:
    return {header: index + 1 for index, header in enumerate(headers) if header}
