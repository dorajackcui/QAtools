from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import workbook_schema as schema
from .errors import (
    ColumnNotFoundError,
    ConfigError,
    TranslationUnitLoadError,
    WorkbookFormatError,
)
from .models import RowFillResult, RowItem, TranslationUnit
from .tag_engine import extract_tags, is_tag_placeholder, serialize_known_tags
from .tag_rules import (
    TagRules,
    default_tag_rules,
    normalized_tag_rules_hash,
    tag_rules_from_payload,
    tag_rules_payload,
)
from .template_engine import PLACEHOLDER_RE, parse_template


def _read_source_rows(
    input_path: Path,
    source_col: str | int,
    target_col: str | int | None,
    *,
    tag_rules: TagRules | None = None,
) -> list[RowItem]:
    active_tag_rules = default_tag_rules() if tag_rules is None else tag_rules
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        source_index = _resolve_column(ws, source_col)
        target_index = _resolve_column(ws, target_col) if target_col is not None else None

        rows: list[RowItem] = []
        seen_source = False
        blank_source_run = 0
        for row_number, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            source_value = _cell_value(row, source_index)
            if source_value is None or str(source_value).strip() == "":
                if seen_source:
                    blank_source_run += 1
                    if blank_source_run >= 1000:
                        break
                continue
            seen_source = True
            blank_source_run = 0
            raw_source = str(source_value).strip()
            source_extraction = extract_tags(raw_source, rules=active_tag_rules)
            target_value = _cell_value(row, target_index) if target_index else ""
            raw_existing_target = "" if target_value is None else str(target_value).strip()
            target_extraction = serialize_known_tags(
                raw_existing_target, source_extraction.tags
            )
            rows.append(
                RowItem(
                    row_number,
                    source_extraction.text,
                    target_extraction.text,
                    parse_template(source_extraction.text),
                    tuple(row),
                    raw_source=raw_source,
                    raw_existing_target=raw_existing_target,
                    tag_tokens=source_extraction.tags,
                    tag_warnings=source_extraction.warnings,
                    target_tag_warnings=target_extraction.warnings,
                )
            )

        return rows
    finally:
        wb.close()


def _load_translated_units(
    path: Path, *, tag_rules: TagRules | None = None
) -> dict[tuple[str, str], str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        _validate_tag_rules_metadata(wb, path, tag_rules)

        if schema.TM_PAIRS_SHEET in wb.sheetnames:
            return _load_unit_sheet(wb[schema.TM_PAIRS_SHEET], schema.TM_PAIRS_SHEET)

        if schema.TRANSLATION_UNITS_SHEET in wb.sheetnames:
            return _load_unit_sheet(
                wb[schema.TRANSLATION_UNITS_SHEET],
                schema.TRANSLATION_UNITS_SHEET,
            )

        if schema.TO_TRANSLATE_SHEET in wb.sheetnames:
            units = _load_unit_sheet(
                wb[schema.TO_TRANSLATE_SHEET], schema.TO_TRANSLATE_SHEET
            )
            if schema.PREFILLED_UNITS_SHEET in wb.sheetnames:
                units.update(
                    _load_unit_sheet(
                        wb[schema.PREFILLED_UNITS_SHEET],
                        schema.PREFILLED_UNITS_SHEET,
                    )
                )
            return units

        if schema.TEMPLATE_REVIEW_SHEET not in wb.sheetnames:
            supported = ", ".join(
                [
                    schema.TM_PAIRS_SHEET,
                    schema.TRANSLATION_UNITS_SHEET,
                    schema.TO_TRANSLATE_SHEET,
                    schema.TEMPLATE_REVIEW_SHEET,
                ]
            )
            raise TranslationUnitLoadError(
                f"Workbook {path} does not contain a supported translation sheet. "
                f"Expected one of: {supported}"
            )

        ws = wb[schema.TEMPLATE_REVIEW_SHEET]
        headers = _header_values(ws)
        indices = _require_columns(
            headers,
            schema.TEMPLATE_REVIEW_SHEET,
            schema.TEMPLATE_REVIEW_REQUIRED_COLUMNS,
        )
        source_index = indices[schema.SOURCE_TEMPLATE_COLUMN]
        target_index = indices[schema.TARGET_TEMPLATE_COLUMN]

        templates: dict[tuple[str, str], str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            source_template = row[source_index] if source_index < len(row) else None
            target_template = row[target_index] if target_index < len(row) else None
            if source_template and target_template:
                templates[("template", str(source_template))] = str(target_template)
        return templates
    finally:
        wb.close()


def _load_unit_sheet(ws, sheet_name: str | None = None) -> dict[tuple[str, str], str]:
    sheet_name = sheet_name or ws.title
    headers = _header_values(ws)
    indices = _require_columns(
        headers,
        sheet_name,
        schema.TRANSLATION_UNIT_REQUIRED_COLUMNS,
    )
    type_index = indices[schema.UNIT_TYPE_COLUMN]
    source_index = indices[schema.SOURCE_UNIT_COLUMN]
    target_index = indices[schema.TARGET_UNIT_COLUMN]

    units: dict[tuple[str, str], str] = {}
    status_index = _optional_column(headers, schema.STATUS_COLUMN)
    for row in ws.iter_rows(min_row=2, values_only=True):
        unit_type = row[type_index] if type_index < len(row) else None
        source_unit = row[source_index] if source_index < len(row) else None
        target_unit = row[target_index] if target_index < len(row) else None
        status = (
            row[status_index]
            if status_index is not None and status_index < len(row)
            else None
        )
        if (
            unit_type
            and source_unit
            and target_unit
            and str(status or "").strip().lower() != "skip"
        ):
            units[(str(unit_type), str(source_unit))] = str(target_unit)
    return units


def _resolve_column(ws, col: str | int | None) -> int:
    if col is None:
        headers = _header_values(ws, fallback=True)
        raise ColumnNotFoundError(col, headers)
    if isinstance(col, int):
        return col
    if str(col).isdigit():
        return int(str(col))

    wanted = str(col).strip()
    wanted_lower = wanted.lower()
    for index, cell in enumerate(ws[1], start=1):
        value = "" if cell.value is None else str(cell.value).strip()
        if value == wanted or value.lower() == wanted_lower:
            return getattr(cell, "column", index)
    headers = _header_values(ws, fallback=True)
    raise ColumnNotFoundError(col, headers)


def _cell_value(row: tuple[object, ...], one_based_index: int | None) -> object | None:
    if one_based_index is None:
        return None
    zero_based = one_based_index - 1
    return row[zero_based] if zero_based < len(row) else None


def _read_headers(input_path: Path) -> list[str]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        return _header_values(ws, fallback=True)
    finally:
        wb.close()


def _header_values(ws, *, fallback: bool = False) -> list[str]:
    headers: list[str] = []
    for index, cell in enumerate(ws[1], start=1):
        if cell.value is None or str(cell.value).strip() == "":
            if fallback:
                headers.append(f"column_{index}")
            else:
                headers.append("")
            continue
        headers.append(str(cell.value).strip())
    return headers


def _optional_column(headers: list[str], column: str) -> int | None:
    return _column_index(headers, column)


def _require_columns(
    headers: list[str], sheet_name: str, required_columns: list[str]
) -> dict[str, int]:
    indices = {
        column: _column_index(headers, column)
        for column in required_columns
    }
    missing = [column for column, index in indices.items() if index is None]
    if missing:
        available = ", ".join(column for column in headers if column) or "(none)"
        raise TranslationUnitLoadError(
            f"Sheet {sheet_name!r} is missing required columns: "
            f"{', '.join(missing)}.\nAvailable columns: {available}"
        )
    return {column: index for column, index in indices.items() if index is not None}


def _column_index(headers: list[str], column: str) -> int | None:
    for candidate in schema.UNIT_COLUMN_ALIASES.get(column, (column,)):
        if candidate in headers:
            return headers.index(candidate)
    return None


def _display_unit_headers(headers: list[str]) -> list[str]:
    return [schema.UNIT_DISPLAY_COLUMN_NAMES.get(header, header) for header in headers]


def _append_schema_version(summary) -> None:
    summary.append([schema.SCHEMA_VERSION_KEY, schema.SCHEMA_VERSION])


def _add_metadata_sheet(
    wb,
    tag_rules: TagRules | None = None,
    *,
    extra_metadata: dict[str, object] | None = None,
) -> None:
    metadata = wb.create_sheet(schema.METADATA_SHEET)
    metadata.append(schema.METADATA_COLUMNS)
    _append_metadata_rows(metadata, tag_rules, extra_metadata=extra_metadata)
    metadata.sheet_state = "hidden"


def _append_metadata_rows(
    ws,
    tag_rules: TagRules | None = None,
    *,
    extra_metadata: dict[str, object] | None = None,
) -> None:
    ws.append([schema.SCHEMA_VERSION_KEY, schema.SCHEMA_VERSION])
    active_rules = default_tag_rules() if tag_rules is None else tag_rules
    ws.append([schema.TAG_RULES_VERSION_KEY, active_rules.version])
    ws.append([schema.TAG_RULES_HASH_KEY, normalized_tag_rules_hash(active_rules)])
    ws.append([schema.TAG_RULES_SOURCE_KEY, active_rules.source])
    ws.append([schema.TAG_RULES_PAYLOAD_KEY, tag_rules_payload(active_rules)])
    for key, value in (extra_metadata or {}).items():
        ws.append([key, value])


def _workbook_metadata(wb) -> dict[str, object]:
    if schema.METADATA_SHEET not in wb.sheetnames:
        return {}
    ws = wb[schema.METADATA_SHEET]
    return {
        row[0]: row[1]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True)
        if row and row[0] is not None
    }


def _translation_package_metadata(path: Path) -> dict[str, object]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        metadata = _workbook_metadata(wb)
    finally:
        wb.close()
    if metadata.get(schema.WORKBOOK_KIND_KEY) != schema.TRANSLATION_PACKAGE_KIND:
        raise WorkbookFormatError(
            f"Workbook {path} is not a self-contained translation package. "
            "Prepare it again with the current PhraseLoom workflow."
        )
    return metadata


def _load_package_tag_rules(path: Path) -> TagRules:
    metadata = _translation_package_metadata(path)
    payload = metadata.get(schema.TAG_RULES_PAYLOAD_KEY)
    if not payload:
        return default_tag_rules()
    return tag_rules_from_payload(str(payload), source=f"embedded:{path.name}")


def _validate_tag_rules_metadata(
    wb, path: Path, tag_rules: TagRules | None
) -> None:
    if tag_rules is None:
        return
    metadata = _workbook_metadata(wb)
    actual_hash = metadata.get(schema.TAG_RULES_HASH_KEY)
    if actual_hash is None:
        return
    expected_hash = normalized_tag_rules_hash(tag_rules)
    if str(actual_hash) != expected_hash:
        raise ConfigError(
            f"Workbook {path} tag config mismatch. "
            f"Expected {expected_hash}, found {actual_hash}."
        )


def _default_work_dir(input_path: Path) -> Path:
    return input_path.parent / f"{input_path.stem}_l10n"


def _default_extract_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_tm_prefill_pack.xlsx"


def _default_tm_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_reusable_units.xlsx"


def _default_to_translate_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_translator_todo.xlsx"


def _default_fill_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_filled_result.xlsx"


def _default_package_fill_output_path(package_path: Path) -> Path:
    metadata = _translation_package_metadata(package_path)
    original_filename = str(metadata.get(schema.ORIGINAL_FILENAME_KEY) or "source.xlsx")
    return package_path.parent / f"{Path(original_filename).stem}_filled_result.xlsx"


def _default_restore_audit_output_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}_restore_audit{output_path.suffix}")


def _default_legacy_output_path(input_path: Path) -> Path:
    return _default_work_dir(input_path) / f"{input_path.stem}_phraseloom_result.xlsx"


def _write_output_workbook(
    output_path: Path,
    input_path: Path,
    units: list[TranslationUnit],
    result_rows: list[RowFillResult],
    *,
    tag_rules: TagRules | None = None,
    context_col: str | int | None = None,
) -> None:
    wb = Workbook()
    source_headers = _read_headers(input_path)
    context_index = _context_column_index(source_headers, context_col)
    summary = wb.active
    summary.title = schema.SUMMARY_SHEET

    template_units = [unit for unit in units if unit.unit_type == "template"]
    segment_units = [unit for unit in units if unit.unit_type == "segment"]
    total_source_segments = len(result_rows)
    prefilled_units = sum(1 for unit in units if unit.target_unit)
    new_units = sum(1 for unit in units if not unit.target_unit)
    filled_rows = sum(1 for result_row in result_rows if result_row.auto_target)
    new_rows = sum(unit.coverage_count for unit in units if not unit.target_unit)
    summary_rows = [
        ("total_source_rows", total_source_segments),
        ("total_translation_units", len(units)),
        ("already_filled_units", prefilled_units),
        ("already_filled_source_rows", filled_rows),
        ("units_to_translate", new_units),
        ("source_rows_to_translate", new_rows),
    ]
    for row in summary_rows:
        summary.append(row)
    _append_schema_version(summary)

    review = wb.create_sheet(schema.TRANSLATION_UNITS_SHEET)
    review.append(_display_unit_headers(schema.TRANSLATION_UNIT_COLUMNS))
    for unit in units:
        review.append(
            [
                unit.unit_id,
                unit.unit_type,
                unit.source_unit,
                unit.target_unit or None,
                unit.coverage_count,
                unit.unique_source_count,
                _variables_summary(unit),
                "ready" if unit.target_unit else None,
                _sample_sources(unit, 10),
                ",".join(str(item.row_number) for item in unit.items),
                unit.target_unit_source or None,
                unit.suggested_target_unit or None,
                unit.warning or None,
                None,
            ]
        )

    todo = wb.create_sheet(schema.TO_TRANSLATE_SHEET)
    todo.append(_display_unit_headers(schema.TO_TRANSLATE_COLUMNS))
    for unit in _units_in_source_order(units):
        if unit.target_unit:
            continue
        todo.append(
            [
                unit.unit_id,
                unit.unit_type,
                unit.source_unit,
                None,
                _sample_sources(unit, 1),
                _context_value(unit, context_index),
                _first_row_number(unit),
                unit.coverage_count,
                _variables_summary(unit),
                unit.warning or None,
                None,
            ]
        )

    result = wb.create_sheet(schema.SOURCE_MAP_SHEET)
    result.append(schema.SOURCE_MAP_COLUMNS)
    for result_row in result_rows:
        row = result_row.row
        unit = result_row.unit
        auto_target = result_row.auto_target
        if unit is None:
            warning = "no translation unit"
            fill_status = "unit_not_found"
        elif not unit.target_unit:
            warning = "fill target in translation_units, then rerun fill"
            fill_status = "missing_target_unit"
        else:
            warning = unit.warning
            fill_status = "filled"
        warning = _merge_warnings(warning, result_row.warning)

        result.append(
            [
                row.row_number,
                row.source,
                unit.unit_id if unit else None,
                unit.unit_type if unit else None,
                unit.source_unit if unit else None,
                json.dumps(
                    row.match.values if unit and unit.unit_type == "template" else {},
                    ensure_ascii=False,
                ),
                unit.target_unit if unit else None,
                auto_target,
                row.existing_target or None,
                fill_status,
                warning or None,
            ]
        )

    filled = wb.create_sheet(schema.FILLED_WORKBOOK_SHEET)
    filled.append(source_headers + schema.FILLED_WORKBOOK_EXTRA_COLUMNS)
    for result_row in result_rows:
        row = result_row.row
        unit = result_row.unit
        auto_target = result_row.auto_target
        if unit is None:
            warning = "no translation unit"
            fill_status = "unit_not_found"
        elif not unit.target_unit:
            warning = "fill target in to_translate, then rerun fill"
            fill_status = "missing_target_unit"
        else:
            warning = unit.warning
            fill_status = "filled"
        warning = _merge_warnings(warning, result_row.warning)
        original = list(row.original_values)
        if len(original) < len(source_headers):
            original.extend([None] * (len(source_headers) - len(original)))
        filled.append(
            original[: len(source_headers)]
            + [
                auto_target,
                fill_status,
                unit.unit_id if unit else None,
                unit.unit_type if unit else None,
                warning or None,
            ]
        )

    qa = wb.create_sheet(schema.QA_REPORT_SHEET)
    qa.append(schema.QA_REPORT_COLUMNS)
    qa.append(
        [
            "missing_target_unit",
            sum(
                1
                for result_row in result_rows
                if result_row.unit and not result_row.unit.target_unit
            ),
        ]
    )
    qa.append(["filled", sum(1 for result_row in result_rows if result_row.auto_target)])
    qa.append(["warning_units", sum(1 for unit in units if unit.warning)])
    qa.append(
        [
            "protected_token_mismatch_rows",
            sum(
                1
                for result_row in result_rows
                if "protected_token_mismatch:" in result_row.warning
            ),
        ]
    )
    qa.append(
        [
            "protected_token_warning_rows",
            sum(
                1
                for result_row in result_rows
                if result_row.row.tag_warnings or result_row.row.target_tag_warnings
            ),
        ]
    )
    qa.append(
        [
            "protected_only_units",
            sum(1 for unit in units if unit.target_unit_source == "tag_only"),
        ]
    )
    qa.append(["template_units", len(template_units)])
    qa.append(["segment_units", len(segment_units)])

    _add_metadata_sheet(wb, tag_rules)

    for ws in wb.worksheets:
        _style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_to_translate_workbook(
    output_path: Path,
    input_path: Path,
    units: list[TranslationUnit],
    *,
    tag_rules: TagRules | None = None,
    source_col: str | int = "source",
    target_col: str | int | None = "target",
    context_col: str | int | None = None,
    min_group_size: int = 2,
    use_existing_targets: bool = False,
) -> None:
    reserved = {
        schema.METADATA_SHEET,
        schema.PREFILLED_UNITS_SHEET,
        schema.TM_PAIRS_SHEET,
        schema.TO_TRANSLATE_SHEET,
        schema.TRANSLATION_UNITS_SHEET,
        schema.TEMPLATE_REVIEW_SHEET,
    }
    wb = load_workbook(input_path)
    try:
        conflicts = sorted(reserved.intersection(wb.sheetnames))
        if conflicts:
            raise ConfigError(
                "Source workbook uses reserved PhraseLoom sheet names: "
                + ", ".join(conflicts)
            )

        source_headers = _header_values(wb.worksheets[0], fallback=True)
        context_index = _context_column_index(source_headers, context_col)
        original_sheet_states = {ws.title: ws.sheet_state for ws in wb.worksheets}
        original_active_sheet = wb.active.title
        for ws in wb.worksheets:
            ws.sheet_state = "hidden"

        todo = wb.create_sheet(schema.TO_TRANSLATE_SHEET)
        todo.append(_display_unit_headers(schema.TO_TRANSLATE_COLUMNS))
        for unit in _units_in_source_order(units):
            if unit.target_unit:
                continue
            todo.append(_to_translate_row(unit, context_index, target_unit=None))

        prefilled = wb.create_sheet(schema.PREFILLED_UNITS_SHEET)
        prefilled.append(_display_unit_headers(schema.TO_TRANSLATE_COLUMNS))
        for unit in units:
            if not unit.target_unit:
                continue
            prefilled.append(
                _to_translate_row(unit, context_index, target_unit=unit.target_unit)
            )

        _add_metadata_sheet(
            wb,
            tag_rules,
            extra_metadata={
                schema.WORKBOOK_KIND_KEY: schema.TRANSLATION_PACKAGE_KIND,
                schema.ORIGINAL_FILENAME_KEY: input_path.name,
                schema.ORIGINAL_SHEET_STATES_KEY: json.dumps(
                    original_sheet_states, ensure_ascii=False
                ),
                schema.ORIGINAL_ACTIVE_SHEET_KEY: original_active_sheet,
                schema.SOURCE_COLUMN_KEY: json.dumps(source_col, ensure_ascii=False),
                schema.TARGET_COLUMN_KEY: json.dumps(target_col, ensure_ascii=False),
                schema.CONTEXT_COLUMN_KEY: json.dumps(context_col, ensure_ascii=False),
                schema.MIN_GROUP_SIZE_KEY: min_group_size,
                schema.USE_EXISTING_TARGETS_KEY: use_existing_targets,
            },
        )

        _style_sheet(todo)
        _style_sheet(prefilled)
        wb.active = wb.index(todo)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    finally:
        wb.close()


def _to_translate_row(
    unit: TranslationUnit,
    context_index: int | None,
    *,
    target_unit: str | None,
) -> list[object | None]:
    return [
        unit.unit_id,
        unit.unit_type,
        unit.source_unit,
        target_unit,
        _sample_sources(unit, 1),
        _context_value(unit, context_index),
        _first_row_number(unit),
        unit.coverage_count,
        _variables_summary(unit),
        unit.warning or None,
        None,
    ]


def _write_target_column_workbook(
    output_path: Path,
    input_path: Path,
    target_col: str | int,
    result_rows: list[RowFillResult],
) -> None:
    wb = load_workbook(input_path)
    try:
        ws = wb.worksheets[0]
        target_index = _resolve_or_create_column(ws, target_col)

        for result_row in result_rows:
            if result_row.auto_target:
                ws.cell(
                    row=result_row.row.row_number,
                    column=target_index,
                ).value = result_row.auto_target

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    finally:
        wb.close()


def _write_translation_package_target_workbook(
    output_path: Path,
    package_path: Path,
    target_col: str | int,
    result_rows: list[RowFillResult],
) -> None:
    wb = load_workbook(package_path)
    try:
        metadata = _workbook_metadata(wb)
        if metadata.get(schema.WORKBOOK_KIND_KEY) != schema.TRANSLATION_PACKAGE_KIND:
            raise WorkbookFormatError(
                f"Workbook {package_path} is not a self-contained translation package."
            )
        ws = wb.worksheets[0]
        target_index = _resolve_or_create_column(ws, target_col)
        for result_row in result_rows:
            if result_row.auto_target:
                ws.cell(
                    row=result_row.row.row_number,
                    column=target_index,
                ).value = result_row.auto_target

        try:
            original_states = json.loads(
                str(metadata.get(schema.ORIGINAL_SHEET_STATES_KEY) or "{}")
            )
        except json.JSONDecodeError as exc:
            raise WorkbookFormatError(
                f"Workbook {package_path} has invalid original sheet metadata."
            ) from exc

        for sheet_name in (
            schema.TO_TRANSLATE_SHEET,
            schema.PREFILLED_UNITS_SHEET,
            schema.METADATA_SHEET,
        ):
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])

        for original_ws in wb.worksheets:
            original_ws.sheet_state = str(original_states.get(original_ws.title, "visible"))
        if not any(sheet.sheet_state == "visible" for sheet in wb.worksheets):
            wb.worksheets[0].sheet_state = "visible"
        active_sheet = str(metadata.get(schema.ORIGINAL_ACTIVE_SHEET_KEY) or "")
        if active_sheet in wb.sheetnames:
            wb.active = wb.index(wb[active_sheet])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    finally:
        wb.close()


def _write_restore_audit_workbook(
    output_path: Path,
    source_path: Path,
    filled_output_path: Path,
    result_rows: list[RowFillResult],
    *,
    tag_rules: TagRules | None = None,
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    summary.append(["metric", "value"])

    audit_rows = [
        _restore_audit_row(result_row)
        for result_row in result_rows
    ]
    filled_rows = [
        row for row in audit_rows if row[schema.FILL_STATUS_COLUMN] == "filled"
    ]
    unfilled_rows = [
        row for row in audit_rows if row[schema.FILL_STATUS_COLUMN] != "filled"
    ]
    warning_rows = [
        row for row in audit_rows if row[schema.WARNING_COLUMN]
    ]
    summary_rows = [
        ("source_workbook", str(source_path)),
        ("filled_workbook", str(filled_output_path)),
        ("total_source_rows", len(audit_rows)),
        ("filled_rows", len(filled_rows)),
        ("unfilled_rows", len(unfilled_rows)),
        ("warning_rows", len(warning_rows)),
        (
            "source_warning_rows",
            sum(1 for row in audit_rows if row["source_warning"]),
        ),
        (
            "target_warning_rows",
            sum(1 for row in audit_rows if row["target_warning"]),
        ),
        (
            "restore_warning_rows",
            sum(1 for row in audit_rows if row["restore_warning"]),
        ),
        (
            "protected_token_mismatch_rows",
            sum(
                1
                for row in audit_rows
                if "protected_token_mismatch:" in str(row["target_warning"] or "")
            ),
        ),
        (
            "missing_target_unit_rows",
            sum(
                1
                for row in audit_rows
                if row[schema.FILL_STATUS_COLUMN] == "missing_target_unit"
            ),
        ),
        (
            "unit_not_found_rows",
            sum(
                1
                for row in audit_rows
                if row[schema.FILL_STATUS_COLUMN] == "unit_not_found"
            ),
        ),
    ]
    for row in summary_rows:
        summary.append(row)
    _append_schema_version(summary)

    columns = [
        schema.ROW_NUMBER_COLUMN,
        schema.SOURCE_COLUMN,
        schema.AUTO_TARGET_COLUMN,
        schema.FILL_STATUS_COLUMN,
        "source_warning",
        "target_warning",
        "restore_warning",
        schema.WARNING_COLUMN,
    ]
    ws = wb.create_sheet("restore_warnings")
    ws.append(columns)
    for row in warning_rows:
        ws.append([row[column] for column in columns])

    _add_metadata_sheet(wb, tag_rules)
    wb[schema.METADATA_SHEET].sheet_state = "hidden"
    for ws in wb.worksheets:
        _style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _restore_audit_row(result_row: RowFillResult) -> dict[str, object]:
    row = result_row.row
    unit = result_row.unit
    source_warning = _restore_source_warning(result_row)
    target_warning = _restore_target_warning(result_row)
    if unit is None:
        restore_warning = "no translation unit"
        fill_status = "unit_not_found"
    elif not unit.target_unit:
        restore_warning = "fill target in to_translate, then rerun fill"
        fill_status = "missing_target_unit"
    else:
        restore_warning = ""
        fill_status = "filled"
    warning = _merge_warnings(source_warning, target_warning, restore_warning)
    return {
        schema.ROW_NUMBER_COLUMN: row.row_number,
        schema.SOURCE_COLUMN: row.raw_source or row.source,
        schema.AUTO_TARGET_COLUMN: result_row.auto_target,
        schema.FILL_STATUS_COLUMN: fill_status,
        "source_warning": source_warning or None,
        "target_warning": target_warning or None,
        "restore_warning": restore_warning or None,
        schema.WARNING_COLUMN: warning or None,
    }


def _restore_source_warning(result_row: RowFillResult) -> str:
    row = result_row.row
    unit = result_row.unit
    warnings = list(row.tag_warnings)
    source_unit = unit.source_unit if unit else row.source
    if "$" in source_unit:
        warnings.append("price-like text; review manually")
    if re.search(r"\b1\s+(day|time|attempt|task|star|pack)s\b", source_unit):
        warnings.append("plural-sensitive text; review manually")
    return _merge_warnings(*warnings)


def _restore_target_warning(result_row: RowFillResult) -> str:
    row = result_row.row
    unit = result_row.unit
    warnings = list(row.target_tag_warnings)
    if unit and unit.unit_type == "template" and unit.target_unit:
        source_placeholders = {
            placeholder
            for placeholder in PLACEHOLDER_RE.findall(unit.source_unit)
            if not is_tag_placeholder(placeholder)
        }
        target_placeholders = {
            placeholder
            for placeholder in PLACEHOLDER_RE.findall(unit.target_unit)
            if not is_tag_placeholder(placeholder)
        }
        if source_placeholders - target_placeholders:
            warnings.append("target_unit is missing source variables")
    if result_row.warning:
        warnings.append(result_row.warning)
    return _merge_warnings(*warnings)


def _resolve_or_create_column(ws, col: str | int) -> int:
    try:
        return _resolve_column(ws, col)
    except ColumnNotFoundError:
        if isinstance(col, int) or str(col).isdigit():
            raise
        column = ws.max_column + 1
        ws.cell(row=1, column=column).value = str(col)
        return column


def _write_tm_workbook(
    output_path: Path,
    input_path: Path,
    units: list[TranslationUnit],
    rows: list[RowItem],
    *,
    tag_rules: TagRules | None = None,
    context_col: str | int | None = None,
) -> None:
    wb = Workbook()
    context_index = _context_column_index(_read_headers(input_path), context_col)
    summary = wb.active
    summary.title = schema.SUMMARY_SHEET

    template_units = [unit for unit in units if unit.unit_type == "template"]
    segment_units = [unit for unit in units if unit.unit_type == "segment"]
    unique_source_segments = len({row.source for row in rows})
    summary_rows = [
        ("input_file", str(input_path)),
        ("tm_source_segments", len(rows)),
        ("unique_source_segments", unique_source_segments),
        ("duplicate_source_segments", len(rows) - unique_source_segments),
        ("tm_pair_count", len(units)),
        ("template_pair_count", len(template_units)),
        ("segment_pair_count", len(segment_units)),
        ("matched_source_segments", sum(unit.coverage_count for unit in units)),
        ("next_step", "Use this workbook with extract --tm to prefill new translation_units."),
    ]
    for row in summary_rows:
        summary.append(row)
    _append_schema_version(summary)

    pairs = wb.create_sheet(schema.TM_PAIRS_SHEET)
    pairs.append(_display_unit_headers(schema.TM_PAIR_COLUMNS))
    for index, unit in enumerate(units, start=1):
        pairs.append(
            [
                f"TM{index:05d}",
                unit.unit_type,
                unit.source_unit,
                unit.target_unit or None,
                unit.coverage_count,
                unit.unique_source_count,
                _variables_summary(unit),
                "\n".join(dict.fromkeys(item.source for item in unit.items[:10])),
                "\n".join(
                    dict.fromkeys(
                        item.existing_target for item in unit.items[:10] if item.existing_target
                    )
                ),
                ",".join(str(item.row_number) for item in unit.items),
                unit.warning or None,
                _sample_contexts(unit, context_index),
            ]
        )

    tm_map = wb.create_sheet(schema.TM_SOURCE_MAP_SHEET)
    tm_map.append(schema.TM_SOURCE_MAP_COLUMNS)
    unit_by_row_number = {
        item.row_number: unit for unit in units for item in unit.items
    }
    for row in rows:
        unit = unit_by_row_number.get(row.row_number)
        tm_map.append(
            [
                row.row_number,
                row.source,
                row.existing_target,
                unit.unit_type if unit else None,
                unit.source_unit if unit else None,
                unit.target_unit if unit else None,
                json.dumps(
                    row.match.values if unit and unit.unit_type == "template" else {},
                    ensure_ascii=False,
                ),
                _original_context_value(row.original_values, context_index),
            ]
        )

    qa = wb.create_sheet(schema.QA_REPORT_SHEET)
    qa.append(schema.QA_REPORT_COLUMNS)
    qa.append(["tm_pairs", len(units)])
    qa.append(["template_pairs", len(template_units)])
    qa.append(["segment_pairs", len(segment_units)])
    qa.append(["warning_pairs", sum(1 for unit in units if unit.warning)])

    _add_metadata_sheet(wb, tag_rules)

    for ws in wb.worksheets:
        _style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _variables_summary(unit: TranslationUnit) -> str | None:
    if unit.unit_type != "template":
        return None
    placeholders = [
        placeholder
        for placeholder in PLACEHOLDER_RE.findall(unit.source_unit)
        if not is_tag_placeholder(placeholder)
    ]
    if not placeholders:
        return None
    samples: dict[str, list[str]] = {placeholder: [] for placeholder in placeholders}
    for item in unit.items:
        for key, value in item.match.values.items():
            placeholder = "{" + key + "}"
            if placeholder in samples and value not in samples[placeholder]:
                samples[placeholder].append(value)
    return "; ".join(
        f"{placeholder}={','.join(values[:5])}" for placeholder, values in samples.items()
    )


def _units_in_source_order(units: Iterable[TranslationUnit]) -> list[TranslationUnit]:
    return sorted(units, key=_source_order_key)


def _source_order_key(unit: TranslationUnit) -> tuple[int, str, str]:
    first_row_number = _first_row_number(unit)
    return (
        first_row_number if first_row_number is not None else 0,
        unit.unit_type,
        unit.source_unit,
    )


def _first_row_item(unit: TranslationUnit) -> RowItem | None:
    return min(unit.items, key=lambda item: item.row_number) if unit.items else None


def _first_row_number(unit: TranslationUnit) -> int | None:
    first = _first_row_item(unit)
    return first.row_number if first else None


def _context_column_index(
    headers: list[str], context_col: str | int | None = None
) -> int | None:
    auto_detect = context_col is None or str(context_col).strip() == ""
    wanted: str | int = schema.CONTEXT_COLUMN if auto_detect else context_col
    if isinstance(wanted, int) or str(wanted).isdigit():
        index = int(wanted)
        if 1 <= index <= len(headers):
            return index - 1
        if auto_detect:
            return None
        raise ColumnNotFoundError(wanted, headers)

    wanted_text = str(wanted).strip().lower()
    for index, header in enumerate(headers):
        if header.strip().lower() == wanted_text:
            return index
    if auto_detect:
        return None
    raise ColumnNotFoundError(wanted, headers)


def _context_value(unit: TranslationUnit, context_index: int | None) -> object | None:
    first = _first_row_item(unit)
    if first is None or context_index is None or context_index >= len(first.original_values):
        return None
    return first.original_values[context_index]


def _sample_contexts(
    unit: TranslationUnit, context_index: int | None, limit: int = 10
) -> str | None:
    if context_index is None:
        return None
    values = [
        str(value)
        for item in unit.items
        if (value := _original_context_value(item.original_values, context_index))
        not in (None, "")
    ]
    samples = list(dict.fromkeys(values))[:limit]
    return "\n".join(samples) or None


def _original_context_value(
    values: tuple[object, ...], context_index: int | None
) -> object | None:
    if context_index is None or context_index >= len(values):
        return None
    return values[context_index]


def _sample_sources(unit: TranslationUnit, limit: int) -> str:
    return "\n".join(dict.fromkeys(item.source for item in unit.items[:limit]))


def _merge_warnings(*warnings: str) -> str:
    return "; ".join(dict.fromkeys(warning for warning in warnings if warning))


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions

    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)
        for cell in column[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


__all__ = [
    "_cell_value",
    "_default_extract_output_path",
    "_default_fill_output_path",
    "_default_legacy_output_path",
    "_default_package_fill_output_path",
    "_default_restore_audit_output_path",
    "_default_tm_output_path",
    "_default_to_translate_output_path",
    "_default_work_dir",
    "_load_package_tag_rules",
    "_load_translated_units",
    "_load_unit_sheet",
    "_read_headers",
    "_read_source_rows",
    "_resolve_column",
    "_translation_package_metadata",
    "_write_output_workbook",
    "_write_restore_audit_workbook",
    "_write_target_column_workbook",
    "_write_translation_package_target_workbook",
    "_write_tm_workbook",
    "_write_to_translate_workbook",
]
