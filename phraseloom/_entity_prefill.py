from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from . import workbook_schema as schema
from ._entity_io import _save_workbook
from .excel_io import _header_values


def _prefill_entity_sheets(wb, tm_path: Path) -> tuple[int, int]:
    structure_prefills = _load_unique_prefills(
        tm_path,
        schema.ENTITY_STRUCTURES_SHEET,
        schema.SOURCE_STRUCTURE_COLUMN,
        schema.TARGET_STRUCTURE_COLUMN,
    )
    term_prefills = _load_unique_prefills(
        tm_path,
        schema.ENTITY_TERMS_SHEET,
        schema.SOURCE_ENTITY_COLUMN,
        schema.TARGET_ENTITY_COLUMN,
    )
    prefilled_structure_count = _apply_prefills(
        wb[schema.ENTITY_STRUCTURES_SHEET],
        source_column=schema.SOURCE_STRUCTURE_COLUMN,
        target_column=schema.TARGET_STRUCTURE_COLUMN,
        prefills=structure_prefills,
        ready_on_prefill=False,
    )
    prefilled_term_count = _apply_prefills(
        wb[schema.ENTITY_TERMS_SHEET],
        source_column=schema.SOURCE_ENTITY_COLUMN,
        target_column=schema.TARGET_ENTITY_COLUMN,
        prefills=term_prefills,
        ready_on_prefill=True,
    )
    return prefilled_structure_count, prefilled_term_count


def prefill_entity_workbook(
    entity_input_path: str | Path,
    tm_path: str | Path,
    output_path: str | Path,
) -> dict[str, int | str]:
    entity_input_path = Path(entity_input_path)
    tm_path = Path(tm_path)
    output_path = Path(output_path)

    wb = load_workbook(entity_input_path)
    try:
        prefilled_structure_count, prefilled_term_count = _prefill_entity_sheets(
            wb,
            tm_path,
        )
        _save_workbook(wb, output_path)
    finally:
        wb.close()

    return {
        "prefilled_structure_count": prefilled_structure_count,
        "prefilled_term_count": prefilled_term_count,
        "output_path": str(output_path),
    }


def _load_unique_prefills(
    path: Path,
    sheet_name: str,
    source_column: str,
    target_column: str,
) -> dict[str, tuple[str | None, str | None]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        headers = _header_values(ws)
        source_index = headers.index(source_column)
        target_index = headers.index(target_column)
        values: dict[str, set[str]] = defaultdict(set)
        for row in ws.iter_rows(min_row=2, values_only=True):
            source = row[source_index] if source_index < len(row) else None
            target = row[target_index] if target_index < len(row) else None
            if source and target:
                values[str(source)].add(str(target))
        result: dict[str, tuple[str | None, str | None]] = {}
        for source, targets in values.items():
            if len(targets) == 1:
                result[source] = (next(iter(targets)), None)
            elif len(targets) > 1:
                result[source] = (None, "ambiguous_entity_prefill")
        return result
    finally:
        wb.close()


def _apply_prefills(
    ws,
    *,
    source_column: str,
    target_column: str,
    prefills: dict[str, tuple[str | None, str | None]],
    ready_on_prefill: bool,
) -> int:
    headers = _header_values(ws)
    source_index = headers.index(source_column) + 1
    target_index = headers.index(target_column) + 1
    status_index = (
        headers.index(schema.STATUS_COLUMN) + 1
        if schema.STATUS_COLUMN in headers
        else None
    )
    warning_index = headers.index(schema.WARNING_COLUMN) + 1
    prefilled_count = 0
    for row_number in range(2, ws.max_row + 1):
        source = ws.cell(row=row_number, column=source_index).value
        target = ws.cell(row=row_number, column=target_index).value
        if not source or target:
            continue
        prefill_target, warning = prefills.get(str(source), (None, None))
        if prefill_target:
            ws.cell(row=row_number, column=target_index).value = prefill_target
            if ready_on_prefill and status_index is not None:
                ws.cell(row=row_number, column=status_index).value = "ready"
            prefilled_count += 1
        elif warning:
            ws.cell(row=row_number, column=warning_index).value = _merge_warnings(
                ws.cell(row=row_number, column=warning_index).value,
                warning,
            )
    return prefilled_count


def _merge_warnings(*warnings: object) -> str | None:
    merged = [str(warning) for warning in warnings if warning]
    if not merged:
        return None
    return "; ".join(dict.fromkeys(merged))
