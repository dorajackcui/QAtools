#!/usr/bin/env python3
"""Shared helpers for generated Excel output files."""

from __future__ import annotations

from collections.abc import Iterable, Mapping
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


ROW_PROBLEM_COLUMN_HEADER = "术语QA问题"
ROW_PROBLEM_SEPARATOR = "；"
PROBLEM_BASE_HEADERS = ("行号", "source原文", "target原文", "问题描述")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


def load_workbook_for_editing(input_file: str | Path):
    """Load an editable workbook without dropping VBA from .xlsm files."""
    input_path = Path(input_file)
    return load_workbook(
        input_path,
        keep_vba=input_path.suffix.casefold() == ".xlsm",
    )


def validate_distinct_source_target_columns(
    source_column: str,
    target_column: str,
) -> None:
    """Reject a configuration that would compare one column with itself."""
    if source_column.strip().upper() == target_column.strip().upper():
        raise ValueError("source 列和 target 列不能相同。")


def find_last_value_row(
    worksheet,
    columns: Iterable[str],
    *,
    start_row: int = 1,
) -> int:
    """Return the last row containing a value in the selected columns."""
    if start_row < 1:
        raise ValueError("开始行必须大于等于 1。")

    column_indexes = {
        column_index_from_string(column.strip().upper()) for column in columns
    }
    if not column_indexes:
        return start_row - 1

    # Editable openpyxl worksheets keep instantiated cells in a sparse mapping.
    # Reading that mapping avoids walking to max_row when only an unrelated column
    # or a styled blank cell extends the worksheet's reported dimensions.
    return max(
        (
            row_index
            for (row_index, column_index), cell in worksheet._cells.items()
            if row_index >= start_row
            and column_index in column_indexes
            and cell.value is not None
        ),
        default=start_row - 1,
    )


def rebuild_output_sheet(workbook, current_sheet_name: str, sheet_name: str):
    if current_sheet_name == sheet_name:
        raise ValueError(f"数据工作表名称不能为 {sheet_name}")
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    return workbook.create_sheet(title=sheet_name)


def join_unique_text(values, separator: str = ROW_PROBLEM_SEPARATOR) -> str:
    """Join non-empty text values once, preserving their first-seen order."""
    unique_values: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value).strip() if value is not None else ""
        if text and text not in seen:
            seen.add(text)
            unique_values.append(text)
    return separator.join(unique_values)


def write_output_table(
    workbook,
    *,
    current_sheet_name: str,
    sheet_name: str,
    headers,
    rows,
    row_link_target_column: str | None = None,
):
    """Write a consistently formatted output table and return its worksheet."""
    worksheet = rebuild_output_sheet(workbook, current_sheet_name, sheet_name)
    normalized_headers = tuple(headers)
    for column_index, header in enumerate(normalized_headers, start=1):
        cell = worksheet.cell(1, column_index, header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    max_lengths = [len(str(header)) for header in normalized_headers]
    for row_index, row in enumerate(rows, start=2):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if value is not None:
                max_lengths[column_index - 1] = max(
                    max_lengths[column_index - 1],
                    max(len(line) for line in str(value).splitlines() or [""]),
                )

    preferred_widths = {1: 10, 2: 52, 3: 52, 4: 48}
    for column_index, max_length in enumerate(max_lengths, start=1):
        width = preferred_widths.get(column_index, min(max(max_length + 2, 12), 30))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A2"
    if normalized_headers:
        last_column = get_column_letter(len(normalized_headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{max(worksheet.max_row, 1)}"
    if row_link_target_column:
        add_row_number_hyperlinks(
            worksheet,
            target_sheet_name=current_sheet_name,
            target_column=row_link_target_column,
        )
    return worksheet


def add_row_number_hyperlinks(
    worksheet,
    *,
    target_sheet_name: str,
    target_column: str,
) -> None:
    """Link first-column source row numbers to cells in the original data sheet."""
    normalized_target_column = target_column.strip().upper()
    column_index_from_string(normalized_target_column)
    escaped_sheet_name = target_sheet_name.replace("'", "''")
    for output_row in range(2, worksheet.max_row + 1):
        row_cell = worksheet.cell(output_row, 1)
        raw_source_row = row_cell.value
        if isinstance(raw_source_row, bool):
            continue
        try:
            source_row = int(raw_source_row)
        except (TypeError, ValueError):
            continue
        if source_row < 1:
            continue
        row_cell.hyperlink = Hyperlink(
            ref=row_cell.coordinate,
            location=f"'{escaped_sheet_name}'!{normalized_target_column}{source_row}",
            display=str(raw_source_row),
        )
        row_cell.style = "Hyperlink"
        row_cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_prefixed_output_path(input_file: str | Path, prefix: str) -> Path:
    input_path = Path(input_file).expanduser()
    return input_path.with_name(f"{prefix}{input_path.name}")


def format_row_problem_text(source_term: str, expected_target_term: str, description: str) -> str:
    source = source_term.strip()
    expected_target = expected_target_term.strip()
    description = description.strip()

    if source and expected_target:
        problem_subject = f"{source} -> {expected_target}"
    elif source:
        problem_subject = source
    elif expected_target:
        problem_subject = f"-> {expected_target}"
    else:
        return description

    if not description:
        return problem_subject
    return f"{problem_subject}：{description}"


def insert_row_problem_column(
    worksheet,
    target_column: str,
    row_problem_texts: Mapping[int, str],
) -> None:
    target_column_index = column_index_from_string(target_column.strip().upper())
    problem_column_index = target_column_index + 1

    worksheet.insert_cols(problem_column_index)
    worksheet.cell(1, problem_column_index, ROW_PROBLEM_COLUMN_HEADER)
    for row_index, problem_text in sorted(row_problem_texts.items()):
        if problem_text:
            worksheet.cell(row_index, problem_column_index, problem_text)
