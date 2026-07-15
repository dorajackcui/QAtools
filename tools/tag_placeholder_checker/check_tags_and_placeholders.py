#!/usr/bin/env python3
"""Check whether source/target rows keep tag and placeholder tokens aligned."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl.utils import column_index_from_string

from tools.excel_output import (
    PROBLEM_BASE_HEADERS,
    build_prefixed_output_path,
    find_last_value_row,
    join_unique_text,
    load_workbook_for_editing,
    rebuild_output_sheet,
    validate_distinct_source_target_columns,
    write_output_table,
)


PROBLEM_SHEET_NAME = "标签占位问题"
SUMMARY_SHEET_NAME = "检查汇总"
CANONICAL_TOKEN_TYPES = ("angle", "square_color", "brace", "newline", "memoq")
LEGACY_TOKEN_TYPE_ALIASES = {"numeric": "memoq"}
SUPPORTED_TOKEN_TYPES = CANONICAL_TOKEN_TYPES + tuple(LEGACY_TOKEN_TYPE_ALIASES)
DEFAULT_TOKEN_TYPES = CANONICAL_TOKEN_TYPES
TOKEN_LABELS = {
    "angle": "尖括号tag",
    "square_color": "方括号color tag",
    "brace": "花括号placeholder",
    "newline": r"\n mark",
    "memoq": "memoQ tag",
    "numeric": "memoQ tag",
}
TOKEN_PATTERNS: dict[str, tuple[re.Pattern[str], ...]] = {
    "angle": (),
    "square_color": (re.compile(r"\[/color\]|\[color\s*=\s*[^\[\]]+\]", re.IGNORECASE),),
    "brace": (),
    "newline": (re.compile(r"\\n"),),
    "memoq": (
        re.compile(r"(?<!\{)\{\d+>|<\d+\}(?!\})|(?<!\{)\{\d+\}(?!\})"),
    ),
}
BRACE_TOKEN_PATTERN = re.compile(
    r"(?<!\{)(?:\{\{(?P<double>[^{}]+)\}\}|\{(?P<single>[^{}]+)\})(?!\})"
)
ANGLE_TAG_NAME_PATTERN = re.compile(r"^[^\s/=<>]+")
NUMERIC_BRACE_PATTERN = re.compile(r"^\d+$")
NUMERIC_TAG_ENVELOPE_PATTERN = re.compile(r"^\d+>.*<\d+$", re.DOTALL)


@dataclass(frozen=True)
class ExtractedToken:
    display_text: str
    inner_text: str
    token_type: str
    start: int
    end: int


@dataclass(frozen=True)
class CheckSummary:
    worksheet_title: str
    output_path: Path
    total_rows_checked: int
    rows_with_selected_tokens: int
    angle_rows: int
    square_color_rows: int
    brace_rows: int
    newline_rows: int
    memoq_rows: int
    problem_rows: int
    problem_count: int
    selected_token_types: tuple[str, ...]

    @property
    def numeric_rows(self) -> int:
        return self.memoq_rows


def normalize_column(column_name: str) -> str:
    normalized = column_name.strip().upper()
    column_index_from_string(normalized)
    return normalized


def build_default_output_path(input_path: Path) -> Path:
    return build_prefixed_output_path(input_path, "tag_check_")


def normalize_token_types(token_types: tuple[str, ...] | list[str] | None) -> tuple[str, ...]:
    raw_types = list(DEFAULT_TOKEN_TYPES if token_types is None else token_types)
    invalid_types = [token_type for token_type in raw_types if token_type not in SUPPORTED_TOKEN_TYPES]
    if invalid_types:
        raise ValueError(f"不支持的检查类型: {'、'.join(invalid_types)}")

    canonical_raw_types = {
        LEGACY_TOKEN_TYPE_ALIASES.get(token_type, token_type)
        for token_type in raw_types
    }
    normalized_types = tuple(
        token_type for token_type in CANONICAL_TOKEN_TYPES if token_type in canonical_raw_types
    )
    if not normalized_types:
        raise ValueError("请至少选择一种检查类型。")
    return normalized_types


def normalize_angle_patterns(angle_patterns: tuple[str, ...] | list[str] | None) -> tuple[str, ...]:
    if angle_patterns is None:
        return ()
    return tuple(pattern.strip() for pattern in angle_patterns if pattern and pattern.strip())


def load_angle_patterns_from_file(config_file: str | Path | None = None) -> tuple[str, ...]:
    if config_file is None:
        raise ValueError("请提供尖括号tag过滤配置文件路径。")
    config_path = Path(config_file).expanduser().resolve()
    if not config_path.exists():
        raise FileNotFoundError(f"尖括号tag配置文件不存在: {config_path}")

    try:
        config_data = json.loads(config_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"尖括号tag配置文件不是有效 JSON: {config_path} ({exc})") from exc

    if not isinstance(config_data, dict):
        raise ValueError(
            f"尖括号tag配置格式错误: {config_path}，需要 JSON 对象中的 patterns 字符串数组。"
        )
    patterns = config_data.get("patterns")
    if not isinstance(patterns, list) or any(not isinstance(pattern, str) for pattern in patterns):
        raise ValueError(f"尖括号tag配置格式错误: {config_path}，需要 JSON 对象中的 patterns 字符串数组。")
    return normalize_angle_patterns(patterns)


def resolve_angle_patterns(
    angle_patterns: tuple[str, ...] | list[str] | None = None,
    angle_config_file: str | Path | None = None,
) -> tuple[str, ...]:
    if angle_patterns is not None:
        return normalize_angle_patterns(angle_patterns)
    if angle_config_file is not None:
        return load_angle_patterns_from_file(angle_config_file)
    return ()


def compile_angle_patterns(
    angle_patterns: tuple[str, ...] | list[str] | None = None,
    angle_config_file: str | Path | None = None,
) -> tuple[re.Pattern[str], ...]:
    compiled_patterns: list[re.Pattern[str]] = []
    for pattern in resolve_angle_patterns(angle_patterns, angle_config_file):
        try:
            compiled_patterns.append(re.compile(pattern, re.IGNORECASE))
        except re.error as exc:
            raise ValueError(f"尖括号tag正则无效: {pattern} ({exc})") from exc
    return tuple(compiled_patterns)


def should_keep_angle_token(
    display_text: str,
    inner_text: str,
    angle_regexes: tuple[re.Pattern[str], ...],
) -> bool:
    if not angle_regexes:
        return True
    return any(regex.search(inner_text) or regex.search(display_text) for regex in angle_regexes)


def should_keep_brace_token(inner_text: str) -> bool:
    return not (
        NUMERIC_BRACE_PATTERN.fullmatch(inner_text)
        or NUMERIC_TAG_ENVELOPE_PATTERN.fullmatch(inner_text)
    )


def _character_is_escaped(text: str, index: int) -> bool:
    backslash_count = 0
    cursor = index - 1
    while cursor >= 0 and text[cursor] == "\\":
        backslash_count += 1
        cursor -= 1
    return backslash_count % 2 == 1


def _iter_angle_token_candidates(text: str) -> Iterator[tuple[int, int, str, str]]:
    search_index = 0
    while True:
        start = text.find("<", search_index)
        if start < 0:
            return

        cursor = start + 1
        quote = ""
        end = -1
        while cursor < len(text):
            character = text[cursor]
            if quote:
                if character == quote and not _character_is_escaped(text, cursor):
                    quote = ""
            elif character in {'"', "'"}:
                quote = character
            elif character == "<":
                break
            elif character == ">":
                end = cursor + 1
                break
            cursor += 1

        if end < 0:
            search_index = start + 1
            continue

        display_text = text[start:end]
        raw_inner_text = text[start + 1 : end - 1]
        if raw_inner_text and raw_inner_text == raw_inner_text.strip():
            yield start, end, display_text, raw_inner_text
        search_index = end


def _iter_brace_token_candidates(text: str) -> Iterator[tuple[int, int, str, str]]:
    for match in BRACE_TOKEN_PATTERN.finditer(text):
        inner_text = match.group("double") or match.group("single") or ""
        yield match.start(), match.end(), match.group(0), inner_text.strip()


def _iter_token_candidates(
    text: str,
    token_type: str,
) -> Iterator[tuple[int, int, str, str]]:
    if token_type == "angle":
        yield from _iter_angle_token_candidates(text)
        return
    if token_type == "brace":
        yield from _iter_brace_token_candidates(text)
        return

    for pattern in TOKEN_PATTERNS[token_type]:
        for match in pattern.finditer(text):
            display_text = match.group(0)
            inner_text = match.group(1).strip() if match.lastindex else display_text
            yield match.start(), match.end(), display_text, inner_text


def extract_token_details(
    text: object,
    token_types: tuple[str, ...] | list[str] | None = None,
    angle_patterns: tuple[str, ...] | list[str] | None = None,
    angle_config_file: str | Path | None = None,
) -> list[ExtractedToken]:
    if text is None:
        return []

    normalized_token_types = normalize_token_types(token_types)
    angle_regexes = (
        compile_angle_patterns(angle_patterns, angle_config_file)
        if "angle" in normalized_token_types
        else ()
    )
    return _extract_token_details(text, normalized_token_types, angle_regexes)


def _extract_token_details(
    text: object,
    normalized_token_types: tuple[str, ...],
    angle_regexes: tuple[re.Pattern[str], ...],
) -> list[ExtractedToken]:
    if text is None:
        return []

    text_value = str(text)
    matches: list[ExtractedToken] = []

    for token_type in normalized_token_types:
        for start, end, display_text, inner_text in _iter_token_candidates(
            text_value,
            token_type,
        ):
            if token_type == "angle" and not should_keep_angle_token(
                display_text,
                inner_text,
                angle_regexes,
            ):
                continue
            if (
                token_type == "brace"
                and not display_text.startswith("{{")
                and not should_keep_brace_token(inner_text)
            ):
                continue
            matches.append(
                ExtractedToken(
                    display_text=display_text,
                    inner_text=inner_text,
                    token_type=token_type,
                    start=start,
                    end=end,
                )
            )

    matches.sort(key=lambda item: (item.start, item.end))
    return matches


def _parse_angle_structure_token(token: ExtractedToken) -> tuple[str, str] | None:
    inner_text = token.inner_text
    if not inner_text or inner_text.startswith(("!", "?")):
        return None

    is_closing = inner_text.startswith("/")
    tag_body = inner_text[1:].lstrip() if is_closing else inner_text
    match = ANGLE_TAG_NAME_PATTERN.match(tag_body)
    if match is None:
        return None

    tag_name = match.group(0)
    if is_closing:
        return "close", tag_name
    if inner_text.rstrip().endswith("/"):
        return "self", tag_name
    return "open", tag_name


def _angle_structure_signature(
    tokens: list[ExtractedToken],
    paired_tag_names: set[str],
) -> tuple[bool, Counter[tuple[str, tuple[str, ...]]]]:
    is_valid = True
    stack: list[tuple[str, str]] = []
    relationships: Counter[tuple[str, tuple[str, ...]]] = Counter()

    for token in tokens:
        if token.token_type != "angle":
            continue
        parsed_token = _parse_angle_structure_token(token)
        if parsed_token is None:
            continue
        token_kind, tag_name = parsed_token
        if tag_name not in paired_tag_names or token_kind == "self":
            continue

        if token_kind == "open":
            relationships[(token.display_text, tuple(item[1] for item in stack))] += 1
            stack.append((tag_name, token.display_text))
            continue

        if stack and stack[-1][0] == tag_name:
            stack.pop()
            continue

        is_valid = False
        matching_index = next(
            (
                index
                for index in range(len(stack) - 1, -1, -1)
                if stack[index][0] == tag_name
            ),
            None,
        )
        if matching_index is not None:
            del stack[matching_index:]

    if stack:
        is_valid = False
    return is_valid, relationships


def _angle_tag_structures_match(
    source_tokens: list[ExtractedToken],
    target_tokens: list[ExtractedToken],
) -> bool:
    paired_tag_names: set[str] = set()
    for token in source_tokens + target_tokens:
        if token.token_type != "angle":
            continue
        parsed_token = _parse_angle_structure_token(token)
        if parsed_token is not None and parsed_token[0] == "close":
            paired_tag_names.add(parsed_token[1])
    if not paired_tag_names:
        return True
    return _angle_structure_signature(
        source_tokens,
        paired_tag_names,
    ) == _angle_structure_signature(
        target_tokens,
        paired_tag_names,
    )


def extract_tokens(
    text: object,
    token_types: tuple[str, ...] | list[str] | None = None,
    angle_patterns: tuple[str, ...] | list[str] | None = None,
    angle_config_file: str | Path | None = None,
) -> list[str]:
    return [
        token.display_text
        for token in extract_token_details(
            text,
            token_types=token_types,
            angle_patterns=angle_patterns,
            angle_config_file=angle_config_file,
        )
    ]


def format_counter(counter: Counter[str]) -> str:
    if not counter:
        return "无"

    parts: list[str] = []
    for token_text, count in sorted(counter.items()):
        parts.append(f"{token_text} x{count}" if count > 1 else token_text)
    return "、".join(parts)


def build_problem_description(
    *,
    token_type: str,
    source_counter: Counter[str],
    target_counter: Counter[str],
) -> str:
    missing_counter = source_counter - target_counter
    extra_counter = target_counter - source_counter
    label = TOKEN_LABELS[token_type]

    description_lines = [
        f"{label}不一致。",
        f"source={format_counter(source_counter)}",
        f"target={format_counter(target_counter)}",
    ]
    if missing_counter:
        description_lines.append(f"target缺少={format_counter(missing_counter)}")
    if extra_counter:
        description_lines.append(f"target多出={format_counter(extra_counter)}")
    return "；".join(description_lines)


def write_problem_sheet(
    workbook,
    worksheet_title: str,
    target_column: str,
    problem_entries: list[tuple[int, str, str, str, str]],
) -> None:
    entries_by_row: dict[int, list[tuple[int, str, str, str, str]]] = {}
    for entry in problem_entries:
        entries_by_row.setdefault(entry[0], []).append(entry)

    rows = []
    for source_row, row_entries in entries_by_row.items():
        first_entry = row_entries[0]
        rows.append(
            (
                source_row,
                first_entry[3],
                first_entry[4],
                join_unique_text(
                    f"{entry[1]}：{entry[2]}" for entry in row_entries
                ),
                join_unique_text(entry[1] for entry in row_entries),
            )
        )

    write_output_table(
        workbook,
        current_sheet_name=worksheet_title,
        sheet_name=PROBLEM_SHEET_NAME,
        headers=PROBLEM_BASE_HEADERS + ("问题类型",),
        rows=rows,
        row_link_target_column=target_column,
    )


def write_summary_sheet(
    workbook,
    worksheet_title: str,
    summary: CheckSummary,
    source_column: str,
    target_column: str,
    start_row: int,
) -> None:
    summary_sheet = rebuild_output_sheet(workbook, worksheet_title, SUMMARY_SHEET_NAME)
    summary_sheet["A1"] = "统计项"
    summary_sheet["B1"] = "值"

    summary_rows = [
        ("检查工作表", summary.worksheet_title),
        ("source列", source_column),
        ("target列", target_column),
        ("开始行", start_row),
        (
            "检查类型",
            "、".join(TOKEN_LABELS[token_type] for token_type in summary.selected_token_types),
        ),
        ("总行数", summary.total_rows_checked),
        ("命中检查类型行数", summary.rows_with_selected_tokens),
        ("含尖括号tag行数", summary.angle_rows),
        ("含方括号color tag行数", summary.square_color_rows),
        ("含花括号placeholder行数", summary.brace_rows),
        (r"含\n mark行数", summary.newline_rows),
        ("含memoQ tag行数", summary.memoq_rows),
        ("问题行数", summary.problem_rows),
        ("问题条数", summary.problem_count),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=2):
        summary_sheet.cell(row_index, 1, label)
        summary_sheet.cell(row_index, 2, value)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="检查双语 Excel 中 source / target 的 tag、placeholder 和 mark 是否一致。"
    )
    parser.add_argument("input_file", nargs="?", help="输入 Excel 文件路径，例如 input.xlsx")
    parser.add_argument("-s", "--sheet", help="工作表名称，不填则使用活动工作表")
    parser.add_argument("-c", "--source-column", help="source 列，例如 A")
    parser.add_argument("-t", "--target-column", help="target 列，例如 B")
    parser.add_argument("--start-row", type=int, default=None, help="开始处理的行号，默认 2")
    parser.add_argument(
        "--token-type",
        action="append",
        choices=SUPPORTED_TOKEN_TYPES,
        default=None,
        help="检查类型，可重复传入，例如 --token-type angle --token-type memoq",
    )
    parser.add_argument(
        "--angle-config",
        help=(
            "可选的尖括号tag过滤配置文件路径；不传时检查所有 <...>。"
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        help="输出 Excel 文件路径，默认生成 tag_check_<原文件名>",
    )
    return parser.parse_args()


def prompt_if_missing(args: argparse.Namespace) -> argparse.Namespace:
    interactive_mode = sys.stdin.isatty()

    if not args.input_file and not interactive_mode:
        raise ValueError("缺少输入文件路径，请传入 input_file 参数。")
    if not args.input_file:
        args.input_file = input("请输入 Excel 文件路径: ").strip()
    if not args.sheet and interactive_mode and len(sys.argv) == 1:
        args.sheet = input("请输入工作表名称（直接回车使用当前活动工作表）: ").strip() or None
    if not args.source_column and not interactive_mode:
        raise ValueError("缺少 source 列，请使用 -c 或 --source-column 指定。")
    if not args.source_column:
        args.source_column = input("请输入 source 列（例如 A）: ").strip().upper()
    if not args.target_column and not interactive_mode:
        raise ValueError("缺少 target 列，请使用 -t 或 --target-column 指定。")
    if not args.target_column:
        args.target_column = input("请输入 target 列（例如 B）: ").strip().upper()
    if args.start_row is None:
        if interactive_mode and len(sys.argv) == 1:
            start_row_text = input("请输入开始处理的行号（默认 2）: ").strip()
            args.start_row = int(start_row_text) if start_row_text else 2
        else:
            args.start_row = 2
    args.token_type = normalize_token_types(args.token_type)
    return args


def process_excel(
    input_file: str | Path,
    source_column: str,
    target_column: str,
    sheet: str | None = None,
    start_row: int = 2,
    token_types: tuple[str, ...] | list[str] | None = None,
    angle_patterns: tuple[str, ...] | list[str] | None = None,
    angle_config_file: str | Path | None = None,
    output_file: str | Path | None = None,
) -> CheckSummary:
    if start_row < 1:
        raise ValueError("开始行必须大于等于 1。")

    normalized_token_types = normalize_token_types(token_types)
    source_column = normalize_column(source_column)
    target_column = normalize_column(target_column)
    validate_distinct_source_target_columns(source_column, target_column)
    angle_regexes = (
        compile_angle_patterns(angle_patterns, angle_config_file)
        if "angle" in normalized_token_types
        else ()
    )

    input_path = Path(input_file).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_output_path(input_path)
    )

    workbook = load_workbook_for_editing(input_path)
    worksheet = workbook[sheet] if sheet else workbook.active

    last_row = find_last_value_row(
        worksheet,
        (source_column, target_column),
        start_row=start_row,
    )
    total_rows_checked = max(0, last_row - start_row + 1)
    rows_with_selected_tokens = 0
    angle_rows = 0
    square_color_rows = 0
    brace_rows = 0
    newline_rows = 0
    memoq_rows = 0
    problem_rows_set: set[int] = set()
    problem_entries: list[tuple[int, str, str, str, str]] = []

    for row_index in range(start_row, last_row + 1):
        source_text = worksheet[f"{source_column}{row_index}"].value
        target_text = worksheet[f"{target_column}{row_index}"].value
        source_tokens = _extract_token_details(
            source_text,
            normalized_token_types,
            angle_regexes,
        )
        target_tokens = _extract_token_details(
            target_text,
            normalized_token_types,
            angle_regexes,
        )

        row_source_snapshot = "" if source_text is None else str(source_text)
        row_target_snapshot = "" if target_text is None else str(target_text)

        combined_tokens = source_tokens + target_tokens
        if combined_tokens:
            rows_with_selected_tokens += 1

        if any(token.token_type == "angle" for token in combined_tokens):
            angle_rows += 1
        if any(token.token_type == "square_color" for token in combined_tokens):
            square_color_rows += 1
        if any(token.token_type == "brace" for token in combined_tokens):
            brace_rows += 1
        if any(token.token_type == "newline" for token in combined_tokens):
            newline_rows += 1
        if any(token.token_type == "memoq" for token in combined_tokens):
            memoq_rows += 1

        for token_type in normalized_token_types:
            source_counter = Counter(
                token.display_text for token in source_tokens if token.token_type == token_type
            )
            target_counter = Counter(
                token.display_text for token in target_tokens if token.token_type == token_type
            )
            counters_match = source_counter == target_counter
            structures_match = (
                token_type != "angle"
                or not counters_match
                or _angle_tag_structures_match(source_tokens, target_tokens)
            )
            if counters_match and structures_match:
                continue

            problem_rows_set.add(row_index)
            if counters_match:
                problem_type = "尖括号tag结构不一致"
                problem_description = (
                    "尖括号tag结构不一致。source/target 的 tag 文本和数量相同，"
                    "但嵌套或闭合结构不同。"
                )
            else:
                problem_type = f"{TOKEN_LABELS[token_type]}不一致"
                problem_description = build_problem_description(
                    token_type=token_type,
                    source_counter=source_counter,
                    target_counter=target_counter,
                )
            problem_entries.append(
                (
                    row_index,
                    problem_type,
                    problem_description,
                    row_source_snapshot,
                    row_target_snapshot,
                )
            )

    summary = CheckSummary(
        worksheet_title=worksheet.title,
        output_path=output_path,
        total_rows_checked=total_rows_checked,
        rows_with_selected_tokens=rows_with_selected_tokens,
        angle_rows=angle_rows,
        square_color_rows=square_color_rows,
        brace_rows=brace_rows,
        newline_rows=newline_rows,
        memoq_rows=memoq_rows,
        problem_rows=len(problem_rows_set),
        problem_count=len(problem_entries),
        selected_token_types=normalized_token_types,
    )

    write_problem_sheet(workbook, worksheet.title, target_column, problem_entries)
    write_summary_sheet(
        workbook=workbook,
        worksheet_title=worksheet.title,
        summary=summary,
        source_column=source_column,
        target_column=target_column,
        start_row=start_row,
    )
    workbook.save(output_path)
    return summary


def main() -> None:
    args = prompt_if_missing(parse_args())
    summary = process_excel(
        input_file=args.input_file,
        sheet=args.sheet,
        source_column=args.source_column,
        target_column=args.target_column,
        start_row=args.start_row,
        token_types=args.token_type,
        angle_config_file=args.angle_config,
        output_file=args.output,
    )

    print("处理完成。")
    print(f"检查工作表: {summary.worksheet_title}")
    print(f"检查类型: {'、'.join(TOKEN_LABELS[token_type] for token_type in summary.selected_token_types)}")
    print(f"总行数: {summary.total_rows_checked}")
    print(f"命中检查类型行数: {summary.rows_with_selected_tokens}")
    print(f"含尖括号tag行数: {summary.angle_rows}")
    print(f"含方括号color tag行数: {summary.square_color_rows}")
    print(f"含花括号placeholder行数: {summary.brace_rows}")
    print(rf"含\n mark行数: {summary.newline_rows}")
    print(f"含memoQ tag行数: {summary.memoq_rows}")
    print(f"问题行数: {summary.problem_rows}")
    print(f"问题条数: {summary.problem_count}")
    print(f"输出文件: {summary.output_path}")


if __name__ == "__main__":
    main()
