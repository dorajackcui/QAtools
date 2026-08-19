#!/usr/bin/env python3
"""Split one Excel worksheet into batches and restore edited batches."""

from __future__ import annotations

import argparse
from collections.abc import Callable
from copy import copy, deepcopy
from dataclasses import dataclass
import hashlib
from io import BytesIO
import json
import os
from pathlib import Path
import posixpath
import shutil
import sys
import tempfile
from typing import Any, Sequence
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator, TranslatorError
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.utils.indexed_list import IndexedList

from tools.excel_output import load_workbook_for_editing


MANIFEST_FORMAT = "qatools.excel-batches.v1"
MANIFEST_FILE_NAME = "batch_manifest.json"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
ProgressCallback = Callable[[int, int], None]
RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
XML_SPACE_ATTRIBUTE = "{http://www.w3.org/XML/1998/namespace}space"
STYLE_REPOSITORY_ATTRIBUTES = (
    "_fonts",
    "_fills",
    "_borders",
    "_alignments",
    "_number_formats",
    "_protections",
)


@dataclass(frozen=True)
class BatchInfo:
    index: int
    file_name: str
    source_start_row: int
    source_end_row: int
    row_count: int


@dataclass(frozen=True)
class SplitSummary:
    output_dir: Path
    manifest_path: Path
    worksheet_title: str
    batch_size: int
    batch_count: int
    data_row_count: int
    batch_files: tuple[Path, ...]


@dataclass(frozen=True)
class RestoreSummary:
    output_path: Path
    worksheet_title: str
    batch_count: int
    restored_row_count: int


@dataclass(frozen=True)
class _Manifest:
    path: Path
    original_file_name: str
    source_snapshot_name: str
    source_sha256: str
    worksheet_title: str
    header_rows: int
    batch_size: int
    data_row_count: int
    original_max_column: int
    batches: tuple[BatchInfo, ...]


@dataclass(frozen=True)
class _CellSnapshot:
    value: object
    style_array: object | None


def build_default_output_dir(input_path: str | Path) -> Path:
    path = Path(input_path).expanduser()
    return path.with_name(f"{path.stem}_batches")


def _resolve_excel_input(input_file: str | Path) -> Path:
    input_path = Path(input_file).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    if not input_path.is_file():
        raise ValueError(f"输入路径不是文件: {input_path}")
    if input_path.suffix.casefold() not in SUPPORTED_EXTENSIONS:
        raise ValueError("仅支持 .xlsx 和 .xlsm 文件。")
    return input_path


def _last_value_row(worksheet) -> int:
    return max(
        (
            cell.row
            for cell in worksheet._cells.values()
            if not isinstance(cell, MergedCell) and cell.value is not None
        ),
        default=0,
    )


def _snapshot_read_only_row(row) -> tuple[_CellSnapshot, ...]:
    snapshots = []
    for cell in row:
        style_array = (
            copy(cell.style_array)
            if getattr(cell, "has_style", False)
            else None
        )
        if style_array is not None:
            style_array.xfId = 0
        snapshots.append(
            _CellSnapshot(
                value=getattr(cell, "value", None),
                style_array=style_array,
            )
        )
    return tuple(snapshots)


def _scan_read_only_worksheet(
    worksheet,
    *,
    header_rows: int,
    max_column: int,
) -> tuple[int, tuple[tuple[_CellSnapshot, ...], ...]]:
    last_value_row = 0
    header_snapshots = []
    for row_index, row in enumerate(
        worksheet.iter_rows(max_col=max_column),
        start=1,
    ):
        if row_index <= header_rows:
            header_snapshots.append(_snapshot_read_only_row(row))
        if any(getattr(cell, "value", None) is not None for cell in row):
            last_value_row = row_index
    return last_value_row, tuple(header_snapshots)


def _copy_workbook_style_repositories(source_workbook, destination_workbook) -> None:
    for attribute in STYLE_REPOSITORY_ATTRIBUTES:
        if hasattr(source_workbook, attribute):
            setattr(
                destination_workbook,
                attribute,
                IndexedList(getattr(source_workbook, attribute)),
            )
    destination_workbook.epoch = source_workbook.epoch
    destination_workbook.iso_dates = source_workbook.iso_dates


def _write_snapshot_row(worksheet, row_index: int, snapshots) -> None:
    for column_index, snapshot in enumerate(snapshots, start=1):
        if snapshot.value is None and snapshot.style_array is None:
            continue
        cell = worksheet.cell(row_index, column_index)
        cell.value = snapshot.value
        if snapshot.style_array is not None:
            cell._style = copy(snapshot.style_array)


def _write_read_only_row(worksheet, row_index: int, source_row) -> None:
    for column_index, source_cell in enumerate(source_row, start=1):
        value = getattr(source_cell, "value", None)
        has_style = getattr(source_cell, "has_style", False)
        if value is None and not has_style:
            continue
        cell = worksheet.cell(row_index, column_index)
        cell.value = value
        if has_style:
            style_array = copy(source_cell.style_array)
            style_array.xfId = 0
            cell._style = style_array


def _new_slim_batch_workbook(
    source_workbook,
    *,
    worksheet_title: str,
    header_snapshots,
) -> tuple[Workbook, object]:
    batch_workbook = Workbook()
    _copy_workbook_style_repositories(source_workbook, batch_workbook)
    batch_worksheet = batch_workbook.active
    batch_worksheet.title = worksheet_title
    for row_index, snapshots in enumerate(header_snapshots, start=1):
        _write_snapshot_row(batch_worksheet, row_index, snapshots)
    if header_snapshots:
        batch_worksheet.freeze_panes = f"A{len(header_snapshots) + 1}"
    return batch_workbook, batch_worksheet


def _write_streamed_batches(
    source_workbook,
    source_worksheet,
    *,
    staging_dir: Path,
    batches: list[BatchInfo],
    header_rows: int,
    header_snapshots,
    max_column: int,
    progress_callback: ProgressCallback | None,
) -> None:
    source_rows = source_worksheet.iter_rows(
        min_row=header_rows + 1,
        max_row=batches[-1].source_end_row,
        max_col=max_column,
    )
    if progress_callback:
        progress_callback(0, len(batches))

    for completed_count, batch in enumerate(batches, start=1):
        batch_workbook, batch_worksheet = _new_slim_batch_workbook(
            source_workbook,
            worksheet_title=source_worksheet.title,
            header_snapshots=header_snapshots,
        )
        try:
            for local_row in range(header_rows + 1, header_rows + batch.row_count + 1):
                try:
                    source_row = next(source_rows)
                except StopIteration as exc:
                    raise ValueError("读取工作表时数据行提前结束。") from exc
                _write_read_only_row(batch_worksheet, local_row, source_row)
            batch_workbook.save(staging_dir / batch.file_name)
        finally:
            batch_workbook.close()
        if progress_callback:
            progress_callback(completed_count, len(batches))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for block in iter(lambda: file_handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _map_included_row(
    row_index: int,
    *,
    header_rows: int,
    source_start_row: int,
    source_end_row: int,
) -> int | None:
    if row_index <= header_rows:
        return row_index
    if source_start_row <= row_index <= source_end_row:
        return header_rows + 1 + row_index - source_start_row
    return None


def _translate_range(
    cell_range: str | None,
    *,
    header_rows: int,
    source_start_row: int,
    source_end_row: int,
) -> str | None:
    if not cell_range or " " in cell_range or "," in cell_range:
        return None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except ValueError:
        return None

    included_rows = [
        row
        for row in range(min_row, max_row + 1)
        if row <= header_rows or source_start_row <= row <= source_end_row
    ]
    if not included_rows:
        return None

    mapped_min_row = _map_included_row(
        included_rows[0],
        header_rows=header_rows,
        source_start_row=source_start_row,
        source_end_row=source_end_row,
    )
    mapped_max_row = _map_included_row(
        included_rows[-1],
        header_rows=header_rows,
        source_start_row=source_start_row,
        source_end_row=source_end_row,
    )
    if mapped_min_row is None or mapped_max_row is None:
        return None
    return (
        f"{get_column_letter(min_col)}{mapped_min_row}:"
        f"{get_column_letter(max_col)}{mapped_max_row}"
    )


def _trim_worksheet_to_batch(
    worksheet,
    *,
    header_rows: int,
    source_start_row: int,
    source_end_row: int,
) -> None:
    original_max_row = worksheet.max_row
    merged_ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
    for cell_range in merged_ranges:
        worksheet.unmerge_cells(cell_range)

    row_dimensions = {}
    for source_row, dimension in worksheet.row_dimensions.items():
        mapped_row = _map_included_row(
            source_row,
            header_rows=header_rows,
            source_start_row=source_start_row,
            source_end_row=source_end_row,
        )
        if mapped_row is not None:
            copied_dimension = copy(dimension)
            copied_dimension.index = mapped_row
            row_dimensions[mapped_row] = copied_dimension

    # TableList.items() intentionally exposes ``(name, ref)`` pairs rather
    # than the Table objects returned by indexed lookup.
    table_refs = dict(worksheet.tables.items())
    auto_filter_ref = worksheet.auto_filter.ref

    if original_max_row > source_end_row:
        worksheet.delete_rows(
            source_end_row + 1,
            original_max_row - source_end_row,
        )
    rows_before_batch = source_start_row - header_rows - 1
    if rows_before_batch > 0:
        worksheet.delete_rows(header_rows + 1, rows_before_batch)

    worksheet.row_dimensions.clear()
    for mapped_row, dimension in row_dimensions.items():
        dimension.worksheet = worksheet
        worksheet.row_dimensions[mapped_row] = dimension

    for cell_range in merged_ranges:
        translated = _translate_range(
            cell_range,
            header_rows=header_rows,
            source_start_row=source_start_row,
            source_end_row=source_end_row,
        )
        if translated:
            worksheet.merge_cells(translated)

    translated_filter = _translate_range(
        auto_filter_ref,
        header_rows=header_rows,
        source_start_row=source_start_row,
        source_end_row=source_end_row,
    )
    worksheet.auto_filter.ref = translated_filter

    for table_name, original_ref in table_refs.items():
        translated_ref = _translate_range(
            original_ref,
            header_rows=header_rows,
            source_start_row=source_start_row,
            source_end_row=source_end_row,
        )
        if translated_ref:
            table = worksheet.tables[table_name]
            table.ref = translated_ref
            if table.autoFilter is not None:
                table.autoFilter.ref = translated_ref
        else:
            del worksheet.tables[table_name]


def _prepare_output_directory(output_dir: Path) -> tuple[Path, bool]:
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    existed_empty = False
    if output_dir.exists():
        if not output_dir.is_dir():
            raise ValueError(f"输出路径不是目录: {output_dir}")
        if any(output_dir.iterdir()):
            raise ValueError(f"输出目录必须为空: {output_dir}")
        existed_empty = True
    staging_dir = Path(
        tempfile.mkdtemp(
            prefix=f".{output_dir.name}.tmp-",
            dir=output_dir.parent,
        )
    )
    return staging_dir, existed_empty


def split_workbook(
    input_file: str | Path,
    *,
    batch_size: int = 1000,
    sheet: str | None = None,
    header_rows: int = 1,
    output_dir: str | Path | None = None,
    progress_callback: ProgressCallback | None = None,
) -> SplitSummary:
    """Split one worksheet, retaining header rows in every batch workbook."""
    if batch_size < 1:
        raise ValueError("每个 batch 的行数必须大于 0。")
    if header_rows < 0:
        raise ValueError("表头行数不能小于 0。")

    input_path = _resolve_excel_input(input_file)
    destination = (
        Path(output_dir).expanduser().resolve()
        if output_dir
        else build_default_output_dir(input_path).resolve()
    )

    use_streaming_batches = input_path.suffix.casefold() == ".xlsx"
    workbook = (
        load_workbook(input_path, read_only=True, data_only=False)
        if use_streaming_batches
        else load_workbook_for_editing(input_path)
    )
    staging_dir: Path | None = None
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {sheet}")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.active
        worksheet_title = worksheet.title
        original_max_column = worksheet.max_column
        if use_streaming_batches:
            last_data_row, header_snapshots = _scan_read_only_worksheet(
                worksheet,
                header_rows=header_rows,
                max_column=original_max_column,
            )
        else:
            last_data_row = _last_value_row(worksheet)
            header_snapshots = ()

        if last_data_row <= header_rows:
            raise ValueError("指定工作表在表头之后没有可拆分的数据行。")

        data_row_count = last_data_row - header_rows
        batch_count = (data_row_count + batch_size - 1) // batch_size
        batch_number_width = max(3, len(str(batch_count)))
        snapshot_name = f"_qatools_restore_source_{input_path.name}"
        batches: list[BatchInfo] = []
        for index in range(1, batch_count + 1):
            source_start = header_rows + 1 + (index - 1) * batch_size
            source_end = min(source_start + batch_size - 1, last_data_row)
            file_name = (
                f"{input_path.stem}_batch_"
                f"{index:0{batch_number_width}d}_of_"
                f"{batch_count:0{batch_number_width}d}{input_path.suffix}"
            )
            batches.append(
                BatchInfo(
                    index=index,
                    file_name=file_name,
                    source_start_row=source_start,
                    source_end_row=source_end,
                    row_count=source_end - source_start + 1,
                )
            )

        staging_dir, destination_existed_empty = _prepare_output_directory(
            destination
        )
        snapshot_path = staging_dir / snapshot_name
        shutil.copy2(input_path, snapshot_path)
        source_hash = _sha256(snapshot_path)

        if use_streaming_batches:
            _write_streamed_batches(
                workbook,
                worksheet,
                staging_dir=staging_dir,
                batches=batches,
                header_rows=header_rows,
                header_snapshots=header_snapshots,
                max_column=original_max_column,
                progress_callback=progress_callback,
            )
        else:
            if progress_callback:
                progress_callback(0, len(batches))
            for completed_count, batch in enumerate(batches, start=1):
                batch_path = staging_dir / batch.file_name
                shutil.copy2(snapshot_path, batch_path)
                batch_workbook = load_workbook_for_editing(batch_path)
                try:
                    batch_worksheet = batch_workbook[worksheet_title]
                    _trim_worksheet_to_batch(
                        batch_worksheet,
                        header_rows=header_rows,
                        source_start_row=batch.source_start_row,
                        source_end_row=batch.source_end_row,
                    )
                    batch_workbook.save(batch_path)
                finally:
                    batch_workbook.close()
                if progress_callback:
                    progress_callback(completed_count, len(batches))

        manifest_data = {
            "format": MANIFEST_FORMAT,
            "original_file_name": input_path.name,
            "source_snapshot": snapshot_name,
            "source_sha256": source_hash,
            "worksheet": worksheet_title,
            "header_rows": header_rows,
            "batch_size": batch_size,
            "data_row_count": data_row_count,
            "original_max_column": original_max_column,
            "batches": [
                {
                    "index": batch.index,
                    "file_name": batch.file_name,
                    "source_start_row": batch.source_start_row,
                    "source_end_row": batch.source_end_row,
                    "row_count": batch.row_count,
                }
                for batch in batches
            ],
        }
        manifest_path = staging_dir / MANIFEST_FILE_NAME
        manifest_path.write_text(
            json.dumps(manifest_data, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        if destination_existed_empty:
            destination.rmdir()
        staging_dir.replace(destination)
        staging_dir = None

        return SplitSummary(
            output_dir=destination,
            manifest_path=destination / MANIFEST_FILE_NAME,
            worksheet_title=worksheet_title,
            batch_size=batch_size,
            batch_count=batch_count,
            data_row_count=data_row_count,
            batch_files=tuple(destination / batch.file_name for batch in batches),
        )
    finally:
        workbook.close()
        if staging_dir is not None:
            shutil.rmtree(staging_dir, ignore_errors=True)


def _require_manifest_int(data: dict[str, Any], key: str, minimum: int) -> int:
    value = data.get(key)
    if isinstance(value, bool) or not isinstance(value, int) or value < minimum:
        raise ValueError(f"batch manifest 字段无效: {key}")
    return value


def _require_manifest_text(data: dict[str, Any], key: str) -> str:
    value = data.get(key)
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"batch manifest 字段无效: {key}")
    return value


def _safe_member_name(value: str, *, field_name: str) -> str:
    if Path(value).name != value or value in {".", ".."}:
        raise ValueError(f"batch manifest 中的 {field_name} 不是安全文件名。")
    return value


def _load_manifest(manifest_or_directory: str | Path) -> _Manifest:
    source_path = Path(manifest_or_directory).expanduser().resolve()
    manifest_path = source_path / MANIFEST_FILE_NAME if source_path.is_dir() else source_path
    if not manifest_path.exists():
        raise FileNotFoundError(f"batch manifest 不存在: {manifest_path}")
    try:
        data = json.loads(manifest_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"batch manifest 不是有效 JSON: {manifest_path}") from exc
    if not isinstance(data, dict) or data.get("format") != MANIFEST_FORMAT:
        raise ValueError("不支持的 batch manifest 格式。")

    batches_data = data.get("batches")
    if not isinstance(batches_data, list) or not batches_data:
        raise ValueError("batch manifest 没有批次记录。")

    batches: list[BatchInfo] = []
    for batch_data in batches_data:
        if not isinstance(batch_data, dict):
            raise ValueError("batch manifest 的批次记录无效。")
        batches.append(
            BatchInfo(
                index=_require_manifest_int(batch_data, "index", 1),
                file_name=_safe_member_name(
                    _require_manifest_text(batch_data, "file_name"),
                    field_name="file_name",
                ),
                source_start_row=_require_manifest_int(
                    batch_data,
                    "source_start_row",
                    1,
                ),
                source_end_row=_require_manifest_int(
                    batch_data,
                    "source_end_row",
                    1,
                ),
                row_count=_require_manifest_int(batch_data, "row_count", 1),
            )
        )

    expected_indexes = list(range(1, len(batches) + 1))
    if [batch.index for batch in batches] != expected_indexes:
        raise ValueError("batch manifest 的批次顺序不连续。")
    if len({batch.file_name for batch in batches}) != len(batches):
        raise ValueError("batch manifest 包含重复文件名。")

    data_row_count = _require_manifest_int(data, "data_row_count", 1)
    if sum(batch.row_count for batch in batches) != data_row_count:
        raise ValueError("batch manifest 的总行数与批次记录不一致。")
    header_rows = _require_manifest_int(data, "header_rows", 0)
    batch_size = _require_manifest_int(data, "batch_size", 1)
    expected_start_row = header_rows + 1
    for batch in batches:
        if batch.source_end_row - batch.source_start_row + 1 != batch.row_count:
            raise ValueError(f"batch {batch.index} 的原始行范围无效。")
        if batch.source_start_row != expected_start_row:
            raise ValueError("batch manifest 的原始行范围不连续。")
        if batch.row_count > batch_size:
            raise ValueError(f"batch {batch.index} 超过 manifest 中的 batch 行数。")
        expected_start_row = batch.source_end_row + 1

    return _Manifest(
        path=manifest_path,
        original_file_name=_safe_member_name(
            _require_manifest_text(data, "original_file_name"),
            field_name="original_file_name",
        ),
        source_snapshot_name=_safe_member_name(
            _require_manifest_text(data, "source_snapshot"),
            field_name="source_snapshot",
        ),
        source_sha256=_require_manifest_text(data, "source_sha256"),
        worksheet_title=_require_manifest_text(data, "worksheet"),
        header_rows=header_rows,
        batch_size=batch_size,
        data_row_count=data_row_count,
        original_max_column=_require_manifest_int(data, "original_max_column", 1),
        batches=tuple(batches),
    )


def build_default_restore_path(manifest_or_directory: str | Path) -> Path:
    manifest = _load_manifest(manifest_or_directory)
    original_path = Path(manifest.original_file_name)
    return (
        manifest.path.parent.parent
        / f"{original_path.stem}_restored{original_path.suffix}"
    ).resolve()


def _batch_path(manifest: _Manifest, file_name: str) -> Path:
    path = (manifest.path.parent / file_name).resolve()
    if path.parent != manifest.path.parent.resolve():
        raise ValueError("batch 文件路径超出 manifest 所在目录。")
    if not path.exists():
        raise FileNotFoundError(f"batch 文件不存在: {path}")
    return path


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _worksheet_archive_path(archive: ZipFile, worksheet_title: str) -> str:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    relationship_id = None
    for node in workbook_root.iter():
        if (
            _xml_local_name(node.tag) == "sheet"
            and node.attrib.get("name") == worksheet_title
        ):
            relationship_id = node.attrib.get(
                f"{{{RELATIONSHIP_NAMESPACE}}}id"
            )
            break
    if not relationship_id:
        raise ValueError(f"复原模板缺少工作表: {worksheet_title}")

    relationships_root = ElementTree.fromstring(
        archive.read("xl/_rels/workbook.xml.rels")
    )
    target = None
    for node in relationships_root.iter():
        if (
            _xml_local_name(node.tag) == "Relationship"
            and node.attrib.get("Id") == relationship_id
        ):
            target = node.attrib.get("Target")
            break
    if not target:
        raise ValueError(f"无法定位工作表文件: {worksheet_title}")
    if target.startswith("/"):
        worksheet_path = target.lstrip("/")
    else:
        worksheet_path = posixpath.normpath(posixpath.join("xl", target))
    if worksheet_path not in archive.namelist():
        raise ValueError(f"复原模板缺少工作表数据: {worksheet_path}")
    return worksheet_path


def _cell_column_index(cell_element) -> int:
    coordinate = cell_element.attrib.get("r", "")
    column_text = "".join(character for character in coordinate if character.isalpha())
    column_index = 0
    for character in column_text.upper():
        column_index = column_index * 26 + ord(character) - ord("A") + 1
    return column_index


def _worksheet_namespace(root) -> str:
    if root.tag.startswith("{"):
        return root.tag[1:].split("}", 1)[0]
    return ""


def _parse_xml_with_namespace_prefixes(
    xml_data: bytes,
) -> tuple[object, tuple[tuple[str, str], ...]]:
    namespace_prefixes: list[tuple[str, str]] = []
    parser = ElementTree.iterparse(
        BytesIO(xml_data),
        events=("start-ns",),
    )
    for event, value in parser:
        if event == "start-ns" and value not in namespace_prefixes:
            namespace_prefixes.append(value)
    return parser.root, tuple(namespace_prefixes)


def _xml_namespace_uri(name: object) -> str | None:
    if isinstance(name, str) and name.startswith("{"):
        return name[1:].split("}", 1)[0]
    return None


def _preserve_namespace_prefixes(
    root,
    namespace_prefixes: tuple[tuple[str, str], ...],
) -> None:
    preferred_prefixes: dict[str, str] = {}
    for prefix, uri in namespace_prefixes:
        preferred_prefixes.setdefault(uri, prefix)

    for uri, prefix in preferred_prefixes.items():
        try:
            ElementTree.register_namespace(prefix, uri)
        except ValueError:
            # ElementTree reserves generated prefixes such as ns0. Those
            # prefixes cannot occur in mc:Ignorable/Requires in valid OOXML.
            continue

    used_namespace_uris: set[str] = set()
    for node in root.iter():
        tag_namespace = _xml_namespace_uri(node.tag)
        if tag_namespace:
            used_namespace_uris.add(tag_namespace)
        for attribute_name in node.attrib:
            attribute_namespace = _xml_namespace_uri(attribute_name)
            if attribute_namespace:
                used_namespace_uris.add(attribute_namespace)

    # ElementTree drops declarations that are referenced only from lexical
    # QName values such as mc:Ignorable="x14ac". Add those declarations back
    # explicitly so Excel can still resolve the original prefix.
    for prefix, uri in namespace_prefixes:
        if not prefix or prefix == "xml":
            continue
        if (
            preferred_prefixes.get(uri) != prefix
            or uri not in used_namespace_uris
        ):
            root.attrib.setdefault(f"xmlns:{prefix}", uri)


def _qualified_tag(namespace: str, local_name: str) -> str:
    return f"{{{namespace}}}{local_name}" if namespace else local_name


def _clear_cell_xml_value(cell_element) -> None:
    for child in list(cell_element):
        cell_element.remove(child)
    cell_element.attrib.pop("t", None)


def _read_shared_strings(archive: ZipFile) -> tuple[str, ...]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return ()
    root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in root:
        if _xml_local_name(item.tag) != "si":
            continue
        strings.append(
            "".join(
                text_node.text or ""
                for text_node in item.iter()
                if _xml_local_name(text_node.tag) == "t"
            )
        )
    return tuple(strings)


def _expand_shared_formulas(root, *, batch_index: int) -> None:
    shared_formula_masters: dict[str, tuple[str, str]] = {}
    formula_cells: list[tuple[object, object, str]] = []
    for cell in root.iter():
        if _xml_local_name(cell.tag) != "c":
            continue
        coordinate = cell.attrib.get("r", "")
        formula = next(
            (
                child
                for child in cell
                if _xml_local_name(child.tag) == "f"
            ),
            None,
        )
        if formula is None:
            continue
        formula_type = formula.attrib.get("t", "normal")
        if formula_type in {"array", "dataTable"}:
            raise ValueError(
                f"batch {batch_index} 包含数组或数据表公式，"
                "请先转换为普通公式或值后再复原。"
            )
        if formula_type != "shared":
            continue
        shared_index = formula.attrib.get("si")
        if shared_index is None:
            raise ValueError(f"batch {batch_index} 的共享公式缺少 si。")
        formula_cells.append((cell, formula, shared_index))
        if formula.text:
            shared_formula_masters[shared_index] = (coordinate, formula.text)

    for cell, formula, shared_index in formula_cells:
        coordinate = cell.attrib.get("r", "")
        if formula.text:
            expanded_formula = formula.text
        else:
            master = shared_formula_masters.get(shared_index)
            if master is None:
                raise ValueError(
                    f"batch {batch_index} 的共享公式组 {shared_index} 缺少主公式。"
                )
            master_coordinate, master_formula = master
            try:
                expanded_formula = Translator(
                    f"={master_formula}",
                    origin=master_coordinate,
                ).translate_formula(coordinate)[1:]
            except (TranslatorError, ValueError) as exc:
                raise ValueError(
                    f"batch {batch_index} 的共享公式无法安全展开。"
                ) from exc
        formula.text = expanded_formula
        for attribute in ("t", "si", "ref"):
            formula.attrib.pop(attribute, None)


def _set_inline_string_xml(
    cell_element,
    *,
    value: str,
    namespace: str,
) -> None:
    qualified = lambda name: _qualified_tag(namespace, name)
    cell_element.attrib["t"] = "inlineStr"
    inline_string = ElementTree.SubElement(cell_element, qualified("is"))
    text_element = ElementTree.SubElement(inline_string, qualified("t"))
    if value != value.strip() or "\n" in value or "\r" in value or "\t" in value:
        text_element.attrib[XML_SPACE_ATTRIBUTE] = "preserve"
    text_element.text = value


def _copy_batch_cell_xml_value(
    source_cell,
    destination_cell,
    *,
    namespace: str,
    shared_strings: tuple[str, ...],
) -> None:
    _clear_cell_xml_value(destination_cell)
    if source_cell is None:
        return
    source_type = source_cell.attrib.get("t")
    if source_type == "s":
        value_node = next(
            (
                child
                for child in source_cell
                if _xml_local_name(child.tag) == "v"
            ),
            None,
        )
        if value_node is None or value_node.text is None:
            return
        try:
            string_value = shared_strings[int(value_node.text)]
        except (IndexError, ValueError) as exc:
            raise ValueError("batch 的共享字符串索引无效。") from exc
        _set_inline_string_xml(
            destination_cell,
            value=string_value,
            namespace=namespace,
        )
        return
    if source_type:
        destination_cell.attrib["t"] = source_type
    for child in source_cell:
        destination_cell.append(deepcopy(child))


def _get_or_create_row_element(
    sheet_data,
    rows_by_index: dict[int, object],
    *,
    row_index: int,
    namespace: str,
):
    row_element = rows_by_index.get(row_index)
    if row_element is None:
        row_element = ElementTree.Element(
            _qualified_tag(namespace, "row"),
            {"r": str(row_index)},
        )
        sheet_data.append(row_element)
        rows_by_index[row_index] = row_element
    return row_element


def _apply_batch_xml_row(
    sheet_data,
    rows_by_index: dict[int, object],
    *,
    source_row,
    destination_row_index: int,
    min_column: int,
    max_column: int,
    namespace: str,
    shared_strings: tuple[str, ...],
) -> None:
    row_element = _get_or_create_row_element(
        sheet_data,
        rows_by_index,
        row_index=destination_row_index,
        namespace=namespace,
    )
    destination_cells = {
        _cell_column_index(cell): cell
        for cell in row_element
        if _xml_local_name(cell.tag) == "c"
    }
    source_cells = (
        {
            _cell_column_index(cell): cell
            for cell in source_row
            if _xml_local_name(cell.tag) == "c"
        }
        if source_row is not None
        else {}
    )
    for column_index in range(min_column, max_column + 1):
        source_cell = source_cells.get(column_index)
        destination_cell = destination_cells.get(column_index)
        if destination_cell is None:
            if source_cell is None:
                continue
            destination_cell = ElementTree.Element(
                _qualified_tag(namespace, "c"),
                {
                    "r": (
                        f"{get_column_letter(column_index)}"
                        f"{destination_row_index}"
                    )
                },
            )
            row_element.append(destination_cell)
            destination_cells[column_index] = destination_cell
        _copy_batch_cell_xml_value(
            source_cell,
            destination_cell,
            namespace=namespace,
            shared_strings=shared_strings,
        )
    row_element[:] = sorted(
        row_element,
        key=lambda cell: _cell_column_index(cell)
        if _xml_local_name(cell.tag) == "c"
        else 10**9,
    )


def _cell_xml_has_value(cell) -> bool:
    return any(
        _xml_local_name(child.tag) in {"f", "v", "is"}
        for child in cell
    )


def _sheet_xml_dimensions(root) -> tuple[int, int]:
    namespace = _worksheet_namespace(root)
    dimension = root.find(_qualified_tag(namespace, "dimension"))
    if dimension is not None and dimension.attrib.get("ref"):
        try:
            _, _, max_column, max_row = range_boundaries(
                dimension.attrib["ref"]
            )
            return max_row, max_column
        except ValueError:
            pass
    max_row = 0
    max_column = 1
    for node in root.iter():
        if _xml_local_name(node.tag) == "row":
            try:
                max_row = max(max_row, int(node.attrib.get("r", "0")))
            except ValueError:
                pass
        elif _xml_local_name(node.tag) == "c":
            max_column = max(max_column, _cell_column_index(node))
    return max_row, max_column


def _copy_zip_with_replaced_part(
    source_path: Path,
    output_path: Path,
    *,
    part_name: str,
    replacement: bytes,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.name}.",
        suffix=".tmp",
        dir=output_path.parent,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        with ZipFile(source_path) as source_archive, ZipFile(
            temporary_path,
            mode="w",
            compression=ZIP_DEFLATED,
            allowZip64=True,
        ) as output_archive:
            for item in source_archive.infolist():
                data = replacement if item.filename == part_name else source_archive.read(item)
                output_archive.writestr(item, data)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def _restore_batches_via_xml(
    manifest: _Manifest,
    *,
    snapshot_path: Path,
    batch_paths: list[Path],
    output_path: Path,
    progress_callback: ProgressCallback | None,
) -> None:
    with ZipFile(snapshot_path) as snapshot_archive:
        worksheet_path = _worksheet_archive_path(
            snapshot_archive,
            manifest.worksheet_title,
        )
        worksheet_xml = snapshot_archive.read(worksheet_path)

    root, namespace_prefixes = _parse_xml_with_namespace_prefixes(
        worksheet_xml
    )
    namespace = _worksheet_namespace(root)
    _preserve_namespace_prefixes(root, namespace_prefixes)
    qualified = lambda name: _qualified_tag(namespace, name)
    sheet_data = root.find(qualified("sheetData"))
    if sheet_data is None:
        raise ValueError("复原模板中的工作表缺少 sheetData。")
    rows_by_index = {
        int(row.attrib["r"]): row
        for row in sheet_data
        if _xml_local_name(row.tag) == "row" and row.attrib.get("r", "").isdigit()
    }

    max_column = manifest.original_max_column
    if progress_callback:
        progress_callback(0, len(manifest.batches))
    for completed_count, (batch, batch_path) in enumerate(
        zip(manifest.batches, batch_paths, strict=True),
        start=1,
    ):
        with ZipFile(batch_path) as batch_archive:
            batch_worksheet_path = _worksheet_archive_path(
                batch_archive,
                manifest.worksheet_title,
            )
            batch_root = ElementTree.fromstring(
                batch_archive.read(batch_worksheet_path)
            )
            shared_strings = _read_shared_strings(batch_archive)
        _expand_shared_formulas(batch_root, batch_index=batch.index)
        batch_namespace = _worksheet_namespace(batch_root)
        batch_sheet_data = batch_root.find(
            _qualified_tag(batch_namespace, "sheetData")
        )
        if batch_sheet_data is None:
            raise ValueError(f"batch {batch.index} 缺少 sheetData。")
        batch_rows = {
            int(row.attrib["r"]): row
            for row in batch_sheet_data
            if _xml_local_name(row.tag) == "row"
            and row.attrib.get("r", "").isdigit()
        }
        _, detected_max_column = _sheet_xml_dimensions(batch_root)
        expected_last_row = manifest.header_rows + batch.row_count
        for local_row_index, source_row in batch_rows.items():
            if local_row_index <= expected_last_row:
                continue
            if any(
                _xml_local_name(cell.tag) == "c" and _cell_xml_has_value(cell)
                for cell in source_row
            ):
                raise ValueError(
                    f"batch {batch.index} 在预期范围之后包含额外数据行。"
                )
        batch_max_column = max(
            manifest.original_max_column,
            detected_max_column,
        )
        if batch.index == 1:
            header_min_column = 1
        else:
            header_min_column = max_column + 1
        if header_min_column <= batch_max_column:
            for header_row in range(1, manifest.header_rows + 1):
                _apply_batch_xml_row(
                    sheet_data,
                    rows_by_index,
                    source_row=batch_rows.get(header_row),
                    destination_row_index=header_row,
                    min_column=header_min_column,
                    max_column=batch_max_column,
                    namespace=namespace,
                    shared_strings=shared_strings,
                )
        for local_row_index in range(
            manifest.header_rows + 1,
            expected_last_row + 1,
        ):
            destination_row = (
                batch.source_start_row
                + local_row_index
                - manifest.header_rows
                - 1
            )
            _apply_batch_xml_row(
                sheet_data,
                rows_by_index,
                source_row=batch_rows.get(local_row_index),
                destination_row_index=destination_row,
                min_column=1,
                max_column=batch_max_column,
                namespace=namespace,
                shared_strings=shared_strings,
            )
        max_column = max(max_column, batch_max_column)
        if progress_callback:
            progress_callback(completed_count, len(manifest.batches))

    sheet_data[:] = sorted(
        sheet_data,
        key=lambda row: int(row.attrib.get("r", "0")),
    )
    dimension = root.find(qualified("dimension"))
    if dimension is not None:
        dimension.attrib["ref"] = (
            f"A1:{get_column_letter(max_column)}"
            f"{manifest.header_rows + manifest.data_row_count}"
        )
    restored_xml = ElementTree.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    )
    _copy_zip_with_replaced_part(
        snapshot_path,
        output_path,
        part_name=worksheet_path,
        replacement=restored_xml,
    )


def restore_batches(
    manifest_or_directory: str | Path,
    *,
    output_file: str | Path | None = None,
    progress_callback: ProgressCallback | None = None,
) -> RestoreSummary:
    """Restore all batch rows into the original workbook snapshot."""
    manifest = _load_manifest(manifest_or_directory)
    snapshot_path = _batch_path(manifest, manifest.source_snapshot_name)
    if _sha256(snapshot_path) != manifest.source_sha256:
        raise ValueError("复原模板已被修改，无法安全复原 batch。")

    batch_paths = [
        _batch_path(manifest, batch.file_name)
        for batch in manifest.batches
    ]
    original_suffix = Path(manifest.original_file_name).suffix
    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_restore_path(manifest.path)
    )
    if output_path.suffix.casefold() != original_suffix.casefold():
        raise ValueError(f"复原文件必须使用原扩展名 {original_suffix}。")
    forbidden_paths = {
        manifest.path.resolve(),
        snapshot_path.resolve(),
        *(path.resolve() for path in batch_paths),
    }
    if output_path in forbidden_paths:
        raise ValueError("复原输出不能覆盖 manifest、模板或 batch 文件。")

    _restore_batches_via_xml(
        manifest,
        snapshot_path=snapshot_path,
        batch_paths=batch_paths,
        output_path=output_path,
        progress_callback=progress_callback,
    )

    return RestoreSummary(
        output_path=output_path,
        worksheet_title=manifest.worksheet_title,
        batch_count=len(manifest.batches),
        restored_row_count=manifest.data_row_count,
    )


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="把 Excel 工作表按指定行数拆成 batch，或按 manifest 复原。"
    )
    subparsers = parser.add_subparsers(dest="action", required=True)

    split_parser = subparsers.add_parser("split", help="拆分 Excel 工作表")
    split_parser.add_argument("input_file", help="输入 .xlsx 或 .xlsm 文件")
    split_parser.add_argument("-s", "--sheet", help="要拆分的工作表，默认当前工作表")
    split_parser.add_argument(
        "--batch-size",
        type=int,
        default=1000,
        help="每个 batch 的数据行数，默认 1000",
    )
    split_parser.add_argument(
        "--header-rows",
        type=int,
        default=1,
        help="每个 batch 重复保留的表头行数，默认 1",
    )
    split_parser.add_argument(
        "-o",
        "--output-dir",
        help="batch 输出目录，默认 <原文件名>_batches",
    )

    restore_parser = subparsers.add_parser("restore", help="复原 batch")
    restore_parser.add_argument(
        "batch_source",
        help=f"batch 目录或 {MANIFEST_FILE_NAME} 路径",
    )
    restore_parser.add_argument("-o", "--output", help="复原后的 Excel 路径")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_argument_parser().parse_args(argv)
    if args.action == "split":
        summary = split_workbook(
            args.input_file,
            batch_size=args.batch_size,
            sheet=args.sheet,
            header_rows=args.header_rows,
            output_dir=args.output_dir,
        )
        print(f"工作表: {summary.worksheet_title}")
        print(f"数据行数: {summary.data_row_count}")
        print(f"batch 数: {summary.batch_count}")
        print(f"输出目录: {summary.output_dir}")
        print(f"复原 manifest: {summary.manifest_path}")
        return 0

    summary = restore_batches(args.batch_source, output_file=args.output)
    print(f"工作表: {summary.worksheet_title}")
    print(f"复原 batch 数: {summary.batch_count}")
    print(f"复原数据行数: {summary.restored_row_count}")
    print(f"输出文件: {summary.output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
