from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import Workbook, load_workbook

from tools.excel_output import build_prefixed_output_path

FILE_NAME_EXTENSIONS = (
    ".xlsx",
    ".xls",
    ".xlsm",
    ".csv",
    ".txt",
    ".po",
    ".pot",
    ".json",
    ".xliff",
    ".xlf",
    ".sdlxliff",
    ".xml",
    ".resx",
    ".strings",
    ".properties",
    ".yml",
    ".yaml",
)
OUTPUT_HEADERS = ("文件名", "key", "source", "target", "QA问题")
OUTPUT_SHEET_NAME = "Xbench QA整理"
HEADER_SOURCE = "source"
HEADER_TARGET = "target"
HEADER_COMMENTS = "comments"
HEADER_METADATA = "metadata"
GROUP_KEY_SEPARATOR = "\x1f"
REQUIRED_HEADERS = (HEADER_COMMENTS, HEADER_METADATA, HEADER_SOURCE, HEADER_TARGET)
QA_TITLE_PATTERN = re.compile(r"^\s*(?P<issue_type>.*?)\s*\((?P<terms>.*)\)\s*$")


@dataclass(frozen=True)
class ParsedMetadata:
    key: str
    file_name: str


@dataclass(frozen=True)
class XbenchIssue:
    issue_type: str
    source_term: str
    target_term: str


@dataclass(frozen=True)
class XbenchDetailRow:
    file_name: str
    key: str
    source: str
    target: str
    qa_issue: str
    group_key: str


@dataclass(frozen=True)
class TransformSummary:
    worksheet_title: str
    output_path: Path
    detail_count: int
    grouped_count: int


def value_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: object) -> str:
    return value_to_text(value).casefold()


def looks_like_file_name(value: str) -> bool:
    return value.casefold().endswith(FILE_NAME_EXTENSIONS)


def parse_metadata(value: object) -> ParsedMetadata:
    lines = [line.strip() for line in value_to_text(value).splitlines() if line.strip()]
    if len(lines) >= 2:
        return ParsedMetadata(key=lines[0], file_name=lines[1])
    if not lines:
        return ParsedMetadata(key="", file_name="")
    line = lines[0]
    if looks_like_file_name(line):
        return ParsedMetadata(key="", file_name=line)
    return ParsedMetadata(key=line, file_name="")


def parse_qa_title(title: object) -> XbenchIssue:
    text = value_to_text(title)
    match = QA_TITLE_PATTERN.match(text)
    if match is None:
        return XbenchIssue(issue_type=text, source_term="", target_term="")
    source_term, separator, target_term = match.group("terms").partition(" / ")
    if not separator:
        return XbenchIssue(issue_type=text, source_term="", target_term="")
    return XbenchIssue(
        issue_type=match.group("issue_type").strip(),
        source_term=source_term.strip(),
        target_term=target_term.strip(),
    )


def format_issue_text(issue: XbenchIssue) -> str:
    if issue.source_term and issue.target_term:
        return f"{issue.source_term} -> {issue.target_term}：{issue.issue_type}"
    if issue.source_term:
        return f"{issue.source_term}：{issue.issue_type}"
    if issue.target_term:
        return f"-> {issue.target_term}：{issue.issue_type}"
    return issue.issue_type


def find_header_columns(worksheet) -> tuple[int, dict[str, int]]:
    for row in worksheet.iter_rows():
        columns: dict[str, int] = {}
        for cell in row:
            header = normalize_header(cell.value)
            if header in REQUIRED_HEADERS:
                columns[header] = cell.column
        if all(header in columns for header in REQUIRED_HEADERS):
            return row[0].row, columns
    raise ValueError("未找到 Xbench 明细表头，预期包含: comments, metadata, source, target")


def build_group_key(metadata: ParsedMetadata, source: str) -> str:
    if metadata.key:
        return f"key:{metadata.key}"
    if metadata.file_name:
        return f"file_source:{metadata.file_name}{GROUP_KEY_SEPARATOR}{source}"
    return f"source:{source}"


def choose_qa_group_title(first_cell: object, comments: str, source: str, target: str, metadata: str) -> str:
    first_cell_text = value_to_text(first_cell)
    if first_cell_text and not source and not target and not metadata:
        return first_cell_text
    if comments and not first_cell_text and not source and not target and not metadata:
        return comments
    return ""


def collect_detail_rows(worksheet) -> list[XbenchDetailRow]:
    header_row, columns = find_header_columns(worksheet)
    detail_rows: list[XbenchDetailRow] = []
    current_issue = XbenchIssue(issue_type="", source_term="", target_term="")

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        row_values = [value_to_text(cell.value) for cell in worksheet[row_index]]
        if not any(row_values):
            continue

        first_cell = worksheet.cell(row=row_index, column=1).value
        source = value_to_text(worksheet.cell(row=row_index, column=columns[HEADER_SOURCE]).value)
        target = value_to_text(worksheet.cell(row=row_index, column=columns[HEADER_TARGET]).value)
        comments = value_to_text(worksheet.cell(row=row_index, column=columns[HEADER_COMMENTS]).value)
        metadata_text = value_to_text(worksheet.cell(row=row_index, column=columns[HEADER_METADATA]).value)

        group_title = choose_qa_group_title(first_cell, comments, source, target, metadata_text)
        if group_title:
            current_issue = parse_qa_title(group_title)
            continue

        if not source and not target and not metadata_text:
            continue

        metadata = parse_metadata(metadata_text)
        detail_rows.append(
            XbenchDetailRow(
                file_name=metadata.file_name,
                key=metadata.key,
                source=source,
                target=target,
                qa_issue=format_issue_text(current_issue),
                group_key=build_group_key(metadata, source),
            )
        )

    return detail_rows


def first_non_empty(existing: str, candidate: str) -> str:
    return existing if existing else candidate


def group_detail_rows(detail_rows: Iterable[XbenchDetailRow]) -> list[dict[str, str]]:
    grouped: dict[str, dict[str, str]] = {}
    grouped_issues: dict[str, list[str]] = {}

    for detail_row in detail_rows:
        if detail_row.group_key not in grouped:
            grouped[detail_row.group_key] = dict.fromkeys(OUTPUT_HEADERS, "")
            grouped_issues[detail_row.group_key] = []

        output_row = grouped[detail_row.group_key]
        output_row["文件名"] = first_non_empty(output_row["文件名"], detail_row.file_name)
        output_row["key"] = first_non_empty(output_row["key"], detail_row.key)
        output_row["source"] = first_non_empty(output_row["source"], detail_row.source)
        output_row["target"] = first_non_empty(output_row["target"], detail_row.target)

        issues = grouped_issues[detail_row.group_key]
        if detail_row.qa_issue and detail_row.qa_issue not in issues:
            issues.append(detail_row.qa_issue)
            output_row["QA问题"] = "；".join(issues)

    return list(grouped.values())


def build_default_output_path(input_path: Path) -> Path:
    return build_prefixed_output_path(input_path, "xbench_transform_")


def write_output_workbook(output_path: Path, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = OUTPUT_SHEET_NAME
        worksheet.append(list(OUTPUT_HEADERS))

        for row in rows:
            worksheet.append([row[header] for header in OUTPUT_HEADERS])

        workbook.save(output_path)
    finally:
        workbook.close()


def process_excel(
    input_file: str | Path,
    sheet: str | None = None,
    output_file: str | Path | None = None,
) -> TransformSummary:
    input_path = Path(input_file).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_output_path(input_path).resolve()
    )
    if output_path == input_path:
        raise ValueError("输出文件不能与输入文件相同")

    workbook = load_workbook(input_path)
    try:
        if sheet:
            if sheet not in workbook.sheetnames:
                raise ValueError(f"工作表不存在: {sheet}")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.active

        worksheet_title = worksheet.title
        detail_rows = collect_detail_rows(worksheet)
        grouped_rows = group_detail_rows(detail_rows)
    finally:
        workbook.close()

    write_output_workbook(output_path, grouped_rows)
    return TransformSummary(
        worksheet_title=worksheet_title,
        output_path=output_path,
        detail_count=len(detail_rows),
        grouped_count=len(grouped_rows),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="将 Xbench QA Report 整理为按 key/source 合并的扁平 Excel 工作簿。"
    )
    parser.add_argument("input_file", nargs="?", help="输入 Xbench QA Report Excel 文件路径")
    parser.add_argument(
        "-s",
        "--sheet",
        help="工作表名称，不填则默认处理当前活动工作表",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="输出 Excel 文件路径，默认生成 xbench_transform_<原文件名>",
    )
    return parser.parse_args()


def prompt_if_missing(args: argparse.Namespace) -> argparse.Namespace:
    interactive_mode = sys.stdin.isatty()

    if not args.input_file and not interactive_mode:
        raise ValueError("缺少输入文件路径，请传入 input_file 参数。")
    if not args.input_file:
        args.input_file = input("请输入 Xbench QA Report Excel 文件路径: ").strip()
    if not args.sheet and interactive_mode and len(sys.argv) == 1:
        args.sheet = input("请输入工作表名称（直接回车使用当前活动工作表）: ").strip() or None
    if not args.output and interactive_mode and len(sys.argv) == 1:
        args.output = input("请输入输出文件路径（直接回车使用默认路径）: ").strip() or None
    return args


def main() -> None:
    args = prompt_if_missing(parse_args())
    summary = process_excel(
        input_file=args.input_file,
        sheet=args.sheet,
        output_file=args.output,
    )

    print("处理完成。")
    print(f"工作表: {summary.worksheet_title}")
    print(f"读取明细数: {summary.detail_count}")
    print(f"输出行数: {summary.grouped_count}")
    print(f"输出文件: {summary.output_path}")


if __name__ == "__main__":
    main()
