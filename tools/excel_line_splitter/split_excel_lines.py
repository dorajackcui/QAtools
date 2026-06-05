#!/usr/bin/env python3
"""Split multi-line Excel cell content into a stacked result column."""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from tools.excel_output import build_prefixed_output_path


LINE_BREAK_PATTERN = re.compile(r"\r\n|\n|\r")


def normalize_column(column_name: str) -> str:
    normalized = column_name.strip().upper()
    column_index_from_string(normalized)
    return normalized


def split_cell_lines(value: object) -> list[str]:
    if value is None:
        return []

    text = str(value)
    if text == "":
        return []

    return [line.strip() for line in LINE_BREAK_PATTERN.split(text) if line.strip()]


def build_default_output_path(input_path: Path) -> Path:
    return build_prefixed_output_path(input_path, "split_lines_")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="将 Excel 指定列中的多行文本按回车拆开，并连续写入结果列。"
    )
    parser.add_argument("input_file", nargs="?", help="输入 Excel 文件路径，例如 input.xlsx")
    parser.add_argument(
        "-s",
        "--sheet",
        help="工作表名称，不填则默认处理当前活动工作表",
    )
    parser.add_argument(
        "-c",
        "--source-column",
        help="源列，例如 A",
    )
    parser.add_argument(
        "-r",
        "--result-column",
        help="结果列，例如 B",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=None,
        help="开始处理的行号，默认 2",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="输出 Excel 文件路径，默认生成 split_lines_<原文件名>",
    )
    return parser.parse_args()


def prompt_if_missing(args: argparse.Namespace) -> argparse.Namespace:
    interactive_mode = sys.stdin.isatty()

    if not args.input_file and not interactive_mode:
        raise ValueError("缺少输入文件路径，请传入 input_file 参数。")
    if not args.input_file:
        args.input_file = input("请输入 Excel 文件路径: ").strip()
    if not args.sheet and interactive_mode and len(sys.argv) == 1:
        args.sheet = input("请输入工作表名称（直接回车使用当前活动工作表）: ").strip() or None
    if not args.source_column and not interactive_mode:
        raise ValueError("缺少源列，请使用 -c 或 --source-column 指定。")
    if not args.source_column:
        args.source_column = input("请输入源列（例如 A）: ").strip().upper()
    if not args.result_column and not interactive_mode:
        raise ValueError("缺少结果列，请使用 -r 或 --result-column 指定。")
    if not args.result_column:
        args.result_column = input("请输入结果列（例如 B）: ").strip().upper()
    if args.start_row is None:
        if interactive_mode and len(sys.argv) == 1:
            start_row_text = input("请输入开始处理的行号（默认 2）: ").strip()
            args.start_row = int(start_row_text) if start_row_text else 2
        else:
            args.start_row = 2
    return args


def process_excel(
    input_file: str | Path,
    source_column: str,
    result_column: str,
    sheet: str | None = None,
    start_row: int = 2,
    output_file: str | Path | None = None,
) -> tuple[str, str, str, Path, int]:
    input_path = Path(input_file).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    if start_row < 1:
        raise ValueError("开始行必须大于等于 1。")

    source_column = normalize_column(source_column)
    result_column = normalize_column(result_column)
    if source_column == result_column:
        raise ValueError("源列和结果列不能相同。")

    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_output_path(input_path)
    )

    workbook = load_workbook(input_path)
    worksheet = workbook[sheet] if sheet else workbook.active

    stacked_values: list[str] = []
    for row_index in range(start_row, worksheet.max_row + 1):
        cell_value = worksheet[f"{source_column}{row_index}"].value
        stacked_values.extend(split_cell_lines(cell_value))

    for row_index in range(start_row, worksheet.max_row + 1):
        worksheet[f"{result_column}{row_index}"] = None

    for row_index, value in enumerate(stacked_values, start=start_row):
        worksheet[f"{result_column}{row_index}"] = value

    workbook.save(output_path)

    return (
        worksheet.title,
        source_column,
        result_column,
        output_path,
        len(stacked_values),
    )


def main() -> None:
    args = prompt_if_missing(parse_args())

    worksheet_title, source_column, result_column, output_path, written_count = process_excel(
        input_file=args.input_file,
        source_column=args.source_column,
        result_column=args.result_column,
        sheet=args.sheet,
        start_row=args.start_row,
        output_file=args.output,
    )

    print("处理完成。")
    print(f"工作表: {worksheet_title}")
    print(f"源列: {source_column}")
    print(f"结果列: {result_column}")
    print(f"开始行: {args.start_row}")
    print(f"写入条目数: {written_count}")
    print(f"输出文件: {output_path}")


if __name__ == "__main__":
    main()
