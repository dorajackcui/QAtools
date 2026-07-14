#!/usr/bin/env python3
"""Workbook aggregation for LLM-assisted term extraction."""

from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
from collections.abc import Callable, Iterable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from tools.llm_term_extractor.codex_term_review import (
    ConflictDecision,
    ConflictGroup,
    DEFAULT_CODEX_MODEL,
    DEFAULT_REASONING_EFFORT,
    DEFAULT_TIMEOUT_SECONDS,
    InputBatchRow,
    RowExtraction,
    load_prompt_template,
    parse_conflict_response,
    parse_extraction_response,
    render_conflict_prompt,
    render_extraction_prompt,
    run_codex_prompt,
)
from tools.history_tb import (
    detect_history_tb_columns as shared_detect_history_tb_columns,
    iter_history_rows,
)
from tools.excel_output import build_prefixed_output_path


TERMS_SHEET_NAME = "本批次术语汇总表"
CONFLICTS_SHEET_NAME = "冲突汇总"
DEFAULT_EXTRACTION_PROMPT = "extract_terms_zh_target.md"
DEFAULT_CONFLICT_PROMPT = "conflict_review_zh_target.md"
STRICT_JSON_RETRY_REMINDER = (
    "\n\nYour previous response could not be parsed. "
    "Return strict JSON only, with no markdown fences, no prose, and the same schema."
)
BatchExtractor = Callable[[list[InputBatchRow]], Iterable[RowExtraction]]
ConflictReviewer = Callable[[list[ConflictGroup]], Any]


@dataclass(frozen=True)
class SourceRow:
    row_index: int
    source_text: str
    target_text: str

    @property
    def row_id(self) -> str:
        return str(self.row_index)


@dataclass(frozen=True)
class ExtractionObservation:
    row_index: int
    source_text: str
    target_text: str
    source_term: str
    target_term: str
    term_type: str
    confidence: str
    note: str


@dataclass
class AggregatedTerm:
    source_term: str
    source_key: str
    target_terms: list[str] = field(default_factory=list)
    row_indexes: list[int] = field(default_factory=list)
    source_examples: list[str] = field(default_factory=list)
    target_examples: list[str] = field(default_factory=list)
    term_types: list[str] = field(default_factory=list)
    confidences: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    conflict_decision: Any | None = None


@dataclass(frozen=True)
class LlmTermExtractionSummary:
    output_path: Path
    worksheet_title: str
    source_column: str
    target_column: str
    start_row: int
    scanned_row_count: int
    batch_count: int
    term_count: int
    evidence_count: int
    conflict_count: int
    import_candidate_count: int
    review_before_import_count: int
    already_in_history_count: int


def normalize_column(column_name: str) -> str:
    normalized = column_name.strip().upper()
    column_index_from_string(normalized)
    return normalized


def normalize_term_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).casefold()


def build_default_output_path(input_file: str | Path) -> Path:
    return build_prefixed_output_path(input_file, "llm_terms_")


def default_prompt_path(name: str) -> Path:
    return Path(__file__).resolve().parent / "prompts" / name


def _batch_mode(rows: list[InputBatchRow]) -> str:
    has_target = [bool(row.target_text.strip()) for row in rows]
    if has_target and all(has_target):
        return "source_target"
    if any(has_target):
        return "mixed"
    return "source_only"


def _codex_raw_output_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}_codex_raw.jsonl")


def _resolve_codex_output_path(
    output_path: str | Path | None,
    output_stem: str | Path | None,
) -> Path:
    if output_path is None and output_stem is None:
        raise TypeError("output_path or output_stem is required.")
    return Path(output_path or output_stem).expanduser().absolute()


def _append_raw_codex_output(
    output_path: Path,
    *,
    keep_raw_codex_output: bool,
    kind: str,
    item_index: int,
    attempt: int,
    raw_output: str,
) -> None:
    if not keep_raw_codex_output:
        return
    raw_output_path = _codex_raw_output_path(output_path)
    with raw_output_path.open("a", encoding="utf-8") as raw_file:
        raw_file.write(
            json.dumps(
                {
                    "kind": kind,
                    "item_index": item_index,
                    "attempt": attempt,
                    "raw_output": raw_output,
                },
                ensure_ascii=False,
            )
        )
        raw_file.write("\n")


def _dump_prompt(dump_prompts_dir: Path | None, file_name: str, prompt: str) -> None:
    if dump_prompts_dir is None:
        return
    dump_prompts_dir.mkdir(parents=True, exist_ok=True)
    (dump_prompts_dir / file_name).write_text(prompt, encoding="utf-8")


def _run_codex_with_json_retry(
    *,
    prompt: str,
    output_path: Path,
    kind: str,
    item_index: int,
    parser: Callable[[str], Any],
    model: str,
    reasoning_effort: str,
    timeout_seconds: int,
    keep_raw_codex_output: bool,
) -> Any:
    for attempt in (1, 2):
        attempt_prompt = prompt if attempt == 1 else f"{prompt}{STRICT_JSON_RETRY_REMINDER}"
        with tempfile.TemporaryDirectory(prefix="llm-term-extract-") as tmp_dir:
            codex_output_path = (
                Path(tmp_dir) / f"{kind}-{item_index:04d}-attempt-{attempt}-output.txt"
            )
            raw_output = run_codex_prompt(
                attempt_prompt,
                codex_output_path,
                model,
                reasoning_effort,
                timeout_seconds,
            )
            _append_raw_codex_output(
                output_path,
                keep_raw_codex_output=keep_raw_codex_output,
                kind=kind,
                item_index=item_index,
                attempt=attempt,
                raw_output=raw_output,
            )
            try:
                return parser(raw_output)
            except ValueError:
                if attempt == 2:
                    raise
    raise RuntimeError("unreachable Codex retry state")


def build_codex_batch_extractor(
    *,
    output_path: str | Path | None = None,
    output_stem: str | Path | None = None,
    prompt_file: str | Path | None = None,
    dump_prompts_dir: str | Path | None = None,
    keep_raw_codex_output: bool = False,
    codex_model: str = DEFAULT_CODEX_MODEL,
    codex_reasoning_effort: str = DEFAULT_REASONING_EFFORT,
    model: str | None = None,
    reasoning_effort: str | None = None,
    timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS,
) -> BatchExtractor:
    resolved_output_path = _resolve_codex_output_path(output_path, output_stem)
    resolved_dump_dir = (
        Path(dump_prompts_dir).expanduser().absolute() if dump_prompts_dir else None
    )
    template = load_prompt_template(prompt_file or default_prompt_path(DEFAULT_EXTRACTION_PROMPT))
    resolved_model = model or codex_model
    resolved_reasoning_effort = reasoning_effort or codex_reasoning_effort
    batch_index = 0

    def extract_batch(rows: list[InputBatchRow]) -> Iterable[RowExtraction]:
        nonlocal batch_index
        batch_index += 1
        prompt = render_extraction_prompt(
            template,
            mode=_batch_mode(rows),
            rows=rows,
        )
        _dump_prompt(resolved_dump_dir, f"extract-batch-{batch_index:04d}.md", prompt)
        return _run_codex_with_json_retry(
            prompt=prompt,
            output_path=resolved_output_path,
            kind="extract",
            item_index=batch_index,
            parser=parse_extraction_response,
            model=resolved_model,
            reasoning_effort=resolved_reasoning_effort,
            timeout_seconds=timeout_seconds,
            keep_raw_codex_output=keep_raw_codex_output,
        )

    return extract_batch


def build_codex_conflict_reviewer(
    *,
    output_path: str | Path | None = None,
    output_stem: str | Path | None = None,
    prompt_file: str | Path | None = None,
    dump_prompts_dir: str | Path | None = None,
    keep_raw_codex_output: bool = False,
    codex_model: str = DEFAULT_CODEX_MODEL,
    codex_reasoning_effort: str = DEFAULT_REASONING_EFFORT,
    model: str | None = None,
    reasoning_effort: str | None = None,
    timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS,
) -> ConflictReviewer:
    resolved_output_path = _resolve_codex_output_path(output_path, output_stem)
    resolved_dump_dir = (
        Path(dump_prompts_dir).expanduser().absolute() if dump_prompts_dir else None
    )
    template = load_prompt_template(prompt_file or default_prompt_path(DEFAULT_CONFLICT_PROMPT))
    resolved_model = model or codex_model
    resolved_reasoning_effort = reasoning_effort or codex_reasoning_effort

    def review_conflicts(groups: list[ConflictGroup]) -> list[ConflictDecision]:
        if not groups:
            return []
        prompt = render_conflict_prompt(template, groups=groups)
        _dump_prompt(resolved_dump_dir, "conflict-review.md", prompt)
        return _run_codex_with_json_retry(
            prompt=prompt,
            output_path=resolved_output_path,
            kind="conflict",
            item_index=1,
            parser=parse_conflict_response,
            model=resolved_model,
            reasoning_effort=resolved_reasoning_effort,
            timeout_seconds=timeout_seconds,
            keep_raw_codex_output=keep_raw_codex_output,
        )

    return review_conflicts


def unique_join(values: Iterable[object], separator: str = "、") -> str:
    seen: set[str] = set()
    unique_values: list[str] = []
    for value in values:
        text = str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        unique_values.append(text)
    return separator.join(unique_values)


def iter_batches(rows: list[SourceRow], batch_size: int) -> Iterable[list[SourceRow]]:
    if batch_size < 1:
        raise ValueError("batch_size must be at least 1.")
    for start_index in range(0, len(rows), batch_size):
        yield rows[start_index : start_index + batch_size]


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _read_source_rows(
    worksheet: Any,
    source_column: str,
    target_column: str,
    start_row: int,
) -> list[SourceRow]:
    rows: list[SourceRow] = []
    for row_index in range(start_row, worksheet.max_row + 1):
        source_text = _cell_text(worksheet[f"{source_column}{row_index}"].value)
        target_text = (
            _cell_text(worksheet[f"{target_column}{row_index}"].value)
            if target_column
            else ""
        )
        if not source_text and not target_text:
            continue
        rows.append(
            SourceRow(
                row_index=row_index,
                source_text=source_text,
                target_text=target_text,
            )
        )
    return rows


def detect_history_tb_columns(
    history_tb_file: str | Path,
    *,
    sheet: str | None = None,
    source_column: str | None = None,
    target_column: str | None = None,
    start_row: int = 2,
) -> tuple[str, str, str]:
    """Return history TB worksheet title and detected source/target columns."""
    columns = shared_detect_history_tb_columns(
        history_tb_file,
        sheet=sheet,
        source_column=source_column,
        target_column=target_column,
        start_row=start_row,
        prefer_no_mark=True,
    )
    assert columns.source_column is not None
    assert columns.target_column is not None
    return columns.sheet_title, columns.source_column, columns.target_column


def load_history_tb_mapping(
    history_tb_file: str | Path,
    *,
    sheet: str | None = None,
    source_column: str | None = None,
    target_column: str | None = None,
    start_row: int = 2,
) -> dict[str, str]:
    mapping: dict[str, str] = {}
    _sheet_title, _source_column, _target_column, rows = iter_history_rows(
        history_tb_file,
        sheet=sheet,
        source_column=source_column,
        target_column=target_column,
        start_row=start_row,
        prefer_no_mark=True,
    )
    for row in rows:
        source_text = row.source_text
        target_text = row.target_text
        source_key = normalize_term_key(source_text)
        if not source_key:
            continue
        if source_key not in mapping or (not mapping[source_key] and target_text):
            mapping[source_key] = target_text
    return mapping


def _to_input_batch(rows: list[SourceRow]) -> list[InputBatchRow]:
    return [
        InputBatchRow(
            row_id=row.row_id,
            source_text=row.source_text,
            target_text=row.target_text,
        )
        for row in rows
    ]


def _value_from(value: object, *names: str) -> str:
    if isinstance(value, dict):
        for name in names:
            text = _cell_text(value.get(name))
            if text:
                return text
        return ""
    for name in names:
        text = _cell_text(getattr(value, name, ""))
        if text:
            return text
    return ""


def _row_extraction_id(row_extraction: object) -> str:
    row_id = _value_from(row_extraction, "row_id", "id")
    if row_id:
        return row_id
    row_index = _value_from(row_extraction, "row_index")
    return row_index


def collect_observations(
    rows_by_id: dict[str, SourceRow],
    extractions: Iterable[RowExtraction],
) -> list[ExtractionObservation]:
    observations: list[ExtractionObservation] = []
    for row_extraction in extractions:
        source_row = rows_by_id.get(_row_extraction_id(row_extraction))
        if source_row is None:
            continue
        terms = getattr(row_extraction, "terms", ())
        if isinstance(row_extraction, dict):
            terms = row_extraction.get("terms", ())
        for term in terms:
            source_term = _value_from(term, "source_term", "source", "term")
            if not source_term:
                continue
            observations.append(
                ExtractionObservation(
                    row_index=source_row.row_index,
                    source_text=source_row.source_text,
                    target_text=source_row.target_text,
                    source_term=source_term,
                    target_term=_value_from(term, "target_term", "target", "target_expression"),
                    term_type=_value_from(term, "term_type", "category", "type"),
                    confidence=_value_from(term, "confidence"),
                    note=_value_from(term, "note", "reason"),
                )
            )
    return observations


def _append_unique(values: list[str], value: str) -> None:
    if value and value not in values:
        values.append(value)


def aggregate_observations(observations: list[ExtractionObservation]) -> dict[str, AggregatedTerm]:
    aggregated: dict[str, AggregatedTerm] = {}
    for observation in observations:
        source_key = normalize_term_key(observation.source_term)
        if not source_key:
            continue
        term = aggregated.setdefault(
            source_key,
            AggregatedTerm(source_term=observation.source_term, source_key=source_key),
        )
        _append_unique(term.target_terms, observation.target_term)
        if observation.row_index not in term.row_indexes:
            term.row_indexes.append(observation.row_index)
        _append_unique(term.source_examples, observation.source_text)
        _append_unique(term.target_examples, observation.target_text)
        _append_unique(term.term_types, observation.term_type)
        _append_unique(term.confidences, observation.confidence)
        _append_unique(term.notes, observation.note)
    return aggregated


def _ordered_terms(aggregated_terms: dict[str, AggregatedTerm]) -> list[AggregatedTerm]:
    return sorted(
        aggregated_terms.values(),
        key=lambda term: (
            min(term.row_indexes) if term.row_indexes else 0,
            normalize_term_key(term.source_term),
        ),
    )


def _build_conflict_groups(aggregated_terms: dict[str, AggregatedTerm]) -> list[ConflictGroup]:
    return [
        ConflictGroup(
            group_id=term.source_key,
            source_term=term.source_term,
            target_terms=tuple(term.target_terms),
            examples=tuple(term.source_examples),
        )
        for term in _ordered_terms(aggregated_terms)
        if len(term.target_terms) > 1
    ]


def _iter_decision_items(decisions: object) -> Iterable[tuple[str, object]]:
    if decisions is None:
        return []
    if isinstance(decisions, dict):
        return [(str(key), value) for key, value in decisions.items()]
    return [("", decision) for decision in decisions]


def _apply_conflict_decisions(
    aggregated_terms: dict[str, AggregatedTerm],
    groups: list[ConflictGroup],
    conflict_reviewer: ConflictReviewer | None,
) -> None:
    if not groups or conflict_reviewer is None:
        return

    decisions = conflict_reviewer(groups)
    group_to_key = {group.group_id: group.group_id for group in groups}
    group_to_key.update({group.source_term: group.group_id for group in groups})

    for fallback_key, decision in _iter_decision_items(decisions):
        decision_group_id = _value_from(decision, "group_id", "source_term", "source")
        lookup_key = decision_group_id or fallback_key
        source_key = group_to_key.get(lookup_key) or normalize_term_key(lookup_key)
        term = aggregated_terms.get(source_key)
        if term is not None:
            term.conflict_decision = decision


def _decision_value(decision: object | None, *names: str) -> str:
    if decision is None:
        return ""
    return _value_from(decision, *names)


def _decision_status(decision: object | None) -> str:
    return _decision_value(decision, "decision").casefold()


def _effective_target(term: AggregatedTerm) -> str:
    canonical_target = _decision_value(term.conflict_decision, "canonical_target", "preferred_target")
    if canonical_target:
        return canonical_target
    if len(term.target_terms) == 1:
        return term.target_terms[0]
    return ""


def _write_output_sheets(
    workbook: Any,
    observations: list[ExtractionObservation],
    aggregated_terms: dict[str, AggregatedTerm],
    history_mapping: dict[str, str] | None = None,
) -> tuple[int, int, int, int]:
    history_mapping = history_mapping or {}
    terms_sheet = workbook.active
    terms_sheet.title = TERMS_SHEET_NAME
    terms_sheet.append(
        [
            "source术语",
            "target术语",
            "本批次target观察值",
            "术语来源",
            "出现行数",
            "原始行号",
            "实例原文",
            "实例译文",
            "类型",
            "备注",
            "多译法判断",
            "判断理由",
        ]
    )

    conflicts_sheet = workbook.create_sheet(title=CONFLICTS_SHEET_NAME)
    conflicts_sheet.append(
        [
            "source术语",
            "本批次target观察值",
            "原始行号",
            "决策",
            "建议统一target",
            "冲突/复核原因",
            "实例原文",
            "实例译文",
            "备注",
        ]
    )

    conflict_count = 0
    import_candidate_count = 0
    review_before_import_count = 0
    already_in_history_count = 0

    for term in _ordered_terms(aggregated_terms):
        target_terms = unique_join(term.target_terms)
        rows = unique_join(term.row_indexes)
        source_examples = unique_join(term.source_examples, separator=" | ")
        target_examples = unique_join(term.target_examples, separator=" | ")
        notes = unique_join(term.notes, separator=" | ")
        decision = _decision_status(term.conflict_decision)
        decision_reason = _decision_value(term.conflict_decision, "reason", "note")
        canonical_target = _decision_value(
            term.conflict_decision,
            "canonical_target",
            "preferred_target",
        )
        is_history_term = term.source_key in history_mapping
        target_term = history_mapping[term.source_key] if is_history_term else _effective_target(term)

        terms_sheet.append(
            [
                term.source_term,
                target_term,
                target_terms,
                "历史术语" if is_history_term else "本批次新增",
                len(term.row_indexes),
                rows,
                source_examples,
                target_examples,
                unique_join(term.term_types),
                notes,
                decision,
                decision_reason,
            ]
        )

        if is_history_term:
            already_in_history_count += 1
            continue

        if decision in {"conflict", "review"}:
            conflict_count += 1
            review_before_import_count += 1
            conflicts_sheet.append(
                [
                    term.source_term,
                    target_terms,
                    rows,
                    decision,
                    canonical_target,
                    decision_reason,
                    source_examples,
                    target_examples,
                    notes,
                ]
            )
            continue

        if len(term.target_terms) > 1 and term.conflict_decision is None:
            conflict_count += 1
            review_before_import_count += 1
            conflicts_sheet.append(
                [
                    term.source_term,
                    target_terms,
                    rows,
                    "review",
                    "",
                    "多译法需确认",
                    source_examples,
                    target_examples,
                    notes,
                ]
            )
            continue

        effective_target = _effective_target(term)
        if effective_target and (decision != "same" or len(term.target_terms) == 1 or canonical_target):
            import_candidate_count += 1
            continue

        review_before_import_count += 1

    return (
        conflict_count,
        import_candidate_count,
        review_before_import_count,
        already_in_history_count,
    )


def process_excel(
    input_file: str | Path,
    source_column: str,
    target_column: str | None = None,
    sheet: str | None = None,
    start_row: int = 2,
    batch_size: int = 50,
    output_file: str | Path | None = None,
    batch_extractor: BatchExtractor | None = None,
    conflict_reviewer: ConflictReviewer | None = None,
    history_tb_file: str | Path | None = None,
    history_sheet: str | None = None,
    history_source_column: str | None = None,
    history_target_column: str | None = None,
    history_start_row: int = 2,
    codex_model: str = DEFAULT_CODEX_MODEL,
    codex_reasoning_effort: str = DEFAULT_REASONING_EFFORT,
    extract_prompt_file: str | Path | None = None,
    conflict_prompt_file: str | Path | None = None,
    dump_prompts_dir: str | Path | None = None,
    keep_raw_codex_output: bool = False,
) -> LlmTermExtractionSummary:
    if start_row < 1:
        raise ValueError("start_row must be at least 1.")
    if history_start_row < 1:
        raise ValueError("history_start_row must be at least 1.")

    input_path = Path(input_file).expanduser().absolute()
    if not input_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {input_path}")

    source_column = normalize_column(source_column)
    normalized_target_column = normalize_column(target_column) if target_column else ""
    output_path = (
        Path(output_file).expanduser().absolute()
        if output_file
        else build_default_output_path(input_path)
    )
    if output_path == input_path:
        raise ValueError("输出文件不能与输入文件相同，请指定不同的输出路径。")
    using_default_extractor = batch_extractor is None
    if batch_extractor is None:
        batch_extractor = build_codex_batch_extractor(
            output_path=output_path,
            prompt_file=extract_prompt_file,
            dump_prompts_dir=dump_prompts_dir,
            keep_raw_codex_output=keep_raw_codex_output,
            codex_model=codex_model,
            codex_reasoning_effort=codex_reasoning_effort,
        )
    if conflict_reviewer is None and using_default_extractor:
        conflict_reviewer = build_codex_conflict_reviewer(
            output_path=output_path,
            prompt_file=conflict_prompt_file,
            dump_prompts_dir=dump_prompts_dir,
            keep_raw_codex_output=keep_raw_codex_output,
            codex_model=codex_model,
            codex_reasoning_effort=codex_reasoning_effort,
        )
    history_mapping = (
        load_history_tb_mapping(
            history_tb_file,
            sheet=history_sheet,
            source_column=history_source_column,
            target_column=history_target_column,
            start_row=history_start_row,
        )
        if history_tb_file
        else {}
    )

    workbook = load_workbook(input_path)
    worksheet = workbook[sheet] if sheet else workbook.active
    source_rows = _read_source_rows(
        worksheet=worksheet,
        source_column=source_column,
        target_column=normalized_target_column,
        start_row=start_row,
    )

    observations: list[ExtractionObservation] = []
    batch_count = 0
    for batch in iter_batches(source_rows, batch_size):
        batch_count += 1
        rows_by_id = {row.row_id: row for row in batch}
        extractions = batch_extractor(_to_input_batch(batch))
        observations.extend(collect_observations(rows_by_id, extractions or []))

    aggregated_terms = aggregate_observations(observations)
    conflict_groups = _build_conflict_groups(aggregated_terms)
    _apply_conflict_decisions(aggregated_terms, conflict_groups, conflict_reviewer)

    output_workbook = Workbook()
    (
        conflict_count,
        import_candidate_count,
        review_before_import_count,
        already_in_history_count,
    ) = _write_output_sheets(
        workbook=output_workbook,
        observations=observations,
        aggregated_terms=aggregated_terms,
        history_mapping=history_mapping,
    )
    output_workbook.save(output_path)

    return LlmTermExtractionSummary(
        output_path=output_path,
        worksheet_title=worksheet.title,
        source_column=source_column,
        target_column=normalized_target_column,
        start_row=start_row,
        scanned_row_count=len(source_rows),
        batch_count=batch_count,
        term_count=len(aggregated_terms),
        evidence_count=len(observations),
        conflict_count=conflict_count,
        import_candidate_count=import_candidate_count,
        review_before_import_count=review_before_import_count,
        already_in_history_count=already_in_history_count,
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract LLM-reviewed terms from an Excel workbook.")
    parser.add_argument("input_file", nargs="?", help="Input Excel workbook.")
    parser.add_argument("-s", "--sheet", help="Worksheet name to scan.")
    parser.add_argument("-c", "--source-column", help="Source text column, for example A.")
    parser.add_argument("-t", "--target-column", help="Target text column, for example B.")
    parser.add_argument("--start-row", type=int, default=2, help="First data row to scan.")
    parser.add_argument("--batch-size", type=int, default=50, help="Rows per Codex extraction batch.")
    parser.add_argument("--codex-model", default=DEFAULT_CODEX_MODEL, help="Codex model name.")
    parser.add_argument(
        "--codex-reasoning-effort",
        default=DEFAULT_REASONING_EFFORT,
        choices=("low", "medium", "high", "xhigh"),
        help="Codex model reasoning effort.",
    )
    parser.add_argument("--extract-prompt-file", help="Custom extraction prompt template.")
    parser.add_argument("--conflict-prompt-file", help="Custom conflict-review prompt template.")
    parser.add_argument("--dump-prompts-dir", help="Directory for rendered prompt dumps.")
    parser.add_argument(
        "--keep-raw-codex-output",
        action="store_true",
        help="Append raw Codex responses to a JSONL file next to the output workbook.",
    )
    parser.add_argument("--history-tb", help="Existing Toolshub history TB workbook.")
    parser.add_argument("--history-sheet", help="History TB worksheet name.")
    parser.add_argument("--history-source-column", help="History TB source column or header.")
    parser.add_argument("--history-target-column", help="History TB target column or header.")
    parser.add_argument(
        "--history-start-row",
        type=int,
        default=2,
        help="First history TB data row; the header is the previous row.",
    )
    parser.add_argument("-o", "--output", help="Output workbook path.")
    return parser.parse_args(argv)


def _prompt_required(label: str) -> str:
    while True:
        value = input(f"{label}: ").strip()
        if value:
            return value


def prompt_if_missing(args: argparse.Namespace) -> argparse.Namespace:
    if not sys.stdin.isatty():
        if not getattr(args, "input_file", None):
            raise SystemExit("input_file is required in non-interactive mode.")
        if not getattr(args, "source_column", None):
            raise SystemExit("source_column is required in non-interactive mode.")
        return args

    if not args.input_file:
        args.input_file = _prompt_required("Input file")
    if not args.source_column:
        args.source_column = _prompt_required("Source column")
    if getattr(args, "target_column", None) == "":
        args.target_column = None
    return args


def _print_completion_summary(summary: LlmTermExtractionSummary) -> None:
    print("LLM term extraction completed.")
    print(f"worksheet: {summary.worksheet_title}")
    print(f"source column: {summary.source_column}")
    print(f"target column: {summary.target_column or '未指定'}")
    print(f"scanned rows: {summary.scanned_row_count}")
    print(f"batch count: {summary.batch_count}")
    print(f"term count: {summary.term_count}")
    print(f"conflict count: {summary.conflict_count}")
    print(f"output file: {summary.output_path}")


def main(argv: list[str] | None = None) -> int:
    args = prompt_if_missing(parse_args(argv))
    summary = process_excel(
        input_file=args.input_file,
        source_column=args.source_column,
        target_column=args.target_column,
        sheet=args.sheet,
        start_row=args.start_row,
        batch_size=args.batch_size,
        output_file=args.output,
        history_tb_file=args.history_tb,
        history_sheet=args.history_sheet,
        history_source_column=args.history_source_column,
        history_target_column=args.history_target_column,
        history_start_row=args.history_start_row,
        codex_model=args.codex_model,
        codex_reasoning_effort=args.codex_reasoning_effort,
        extract_prompt_file=args.extract_prompt_file,
        conflict_prompt_file=args.conflict_prompt_file,
        dump_prompts_dir=args.dump_prompts_dir,
        keep_raw_codex_output=args.keep_raw_codex_output,
    )
    _print_completion_summary(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
