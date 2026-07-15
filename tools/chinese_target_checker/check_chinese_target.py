#!/usr/bin/env python3
"""Check whether Excel target cells contain Chinese characters."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl.utils import column_index_from_string

from tools.excel_output import (
    PROBLEM_BASE_HEADERS,
    build_prefixed_output_path,
    find_last_value_row,
    load_workbook_for_editing,
    validate_distinct_source_target_columns,
    write_output_table,
)


PROBLEM_SHEET_NAME = "Target中文问题"
LEGACY_PROBLEM_SHEET_NAMES = ("中文检查问题",)
CHINESE_PATTERN = re.compile(
    r"[\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF"
    r"\U00020000-\U0002FA1F\U00030000-\U000323AF"
    r"\u3000-\u303F\uFF01-\uFF0F\uFF1A-\uFF20\uFF3B-\uFF40\uFF5B-\uFF65"
    r"\u00B7\u2018\u2019\u201C\u201D]"
)


@dataclass(frozen=True)
class CheckSummary:
    output_path: Path
    worksheet_title: str
    source_column: str
    target_column: str
    start_row: int
    processed_count: int
    matched_count: int


def normalize_column(column_name: str) -> str:
    normalized = column_name.strip().upper()
    column_index_from_string(normalized)
    return normalized


def cell_text(value: object) -> str:
    return "" if value is None else str(value)


def build_default_output_path(input_file: str | Path) -> Path:
    return build_prefixed_output_path(input_file, "target_chinese_check_")


def contains_chinese(value: object) -> bool:
    return isinstance(value, str) and bool(CHINESE_PATTERN.search(value))


def extract_chinese_characters(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return "".join(CHINESE_PATTERN.findall(value))


def process_excel(
    input_file: str | Path,
    source_column: str,
    target_column: str,
    sheet: str | None = None,
    start_row: int = 2,
    output_file: str | Path | None = None,
) -> CheckSummary:
    input_path = Path(input_file).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    if start_row < 1:
        raise ValueError("开始行必须大于等于 1。")

    source_column = normalize_column(source_column)
    target_column = normalize_column(target_column)
    validate_distinct_source_target_columns(source_column, target_column)
    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_output_path(input_path)
    )

    workbook = load_workbook_for_editing(input_path)
    worksheet = workbook[sheet] if sheet else workbook.active
    for legacy_sheet_name in LEGACY_PROBLEM_SHEET_NAMES:
        if (
            legacy_sheet_name in workbook.sheetnames
            and worksheet.title != legacy_sheet_name
        ):
            del workbook[legacy_sheet_name]

    processed_count = 0
    problem_entries: list[tuple[int, str, str, str, str]] = []
    last_row = find_last_value_row(
        worksheet,
        (source_column, target_column),
        start_row=start_row,
    )
    for row_index in range(start_row, last_row + 1):
        source_value = worksheet[f"{source_column}{row_index}"].value
        target_value = worksheet[f"{target_column}{row_index}"].value
        chinese_characters = extract_chinese_characters(target_value)
        processed_count += 1
        if not chinese_characters:
            continue
        problem_entries.append(
            (
                row_index,
                cell_text(source_value),
                cell_text(target_value),
                f"Target 中包含中文或全角标点：{chinese_characters}",
                chinese_characters,
            )
        )

    write_output_table(
        workbook,
        current_sheet_name=worksheet.title,
        sheet_name=PROBLEM_SHEET_NAME,
        headers=PROBLEM_BASE_HEADERS + ("命中字符",),
        rows=problem_entries,
        row_link_target_column=target_column,
    )
    workbook.save(output_path)
    return CheckSummary(
        output_path=output_path,
        worksheet_title=worksheet.title,
        source_column=source_column,
        target_column=target_column,
        start_row=start_row,
        processed_count=processed_count,
        matched_count=len(problem_entries),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="检查 Excel target 列是否包含中文字符或中文/全角标点。"
    )
    parser.add_argument("input_file", nargs="?", help="输入 Excel 文件路径，例如 input.xlsx")
    parser.add_argument("-s", "--sheet", help="工作表名称，不填则默认处理当前活动工作表")
    parser.add_argument("-c", "--source-column", help="source 列，例如 A")
    parser.add_argument("-t", "--target-column", help="需要检查的 target 列，例如 B")
    parser.add_argument("--start-row", type=int, default=None, help="开始处理的行号，默认 2")
    parser.add_argument(
        "-o",
        "--output",
        help="输出 Excel 文件路径，默认生成 target_chinese_check_<原文件名>",
    )
    return parser.parse_args()


def prompt_if_missing(args: argparse.Namespace) -> argparse.Namespace:
    interactive_mode = sys.stdin.isatty()
    if not args.input_file and not interactive_mode:
        raise ValueError("缺少输入文件路径，请传入 input_file 参数。")
    if not args.input_file:
        args.input_file = input("请输入 Excel 文件路径: ").strip()
    if not args.sheet and interactive_mode and len(sys.argv) == 1:
        args.sheet = input("请输入工作表名称（直接回车使用当前活动工作表）: ").strip() or None
    if not args.source_column and not interactive_mode:
        raise ValueError("缺少 source 列，请使用 -c 或 --source-column 指定。")
    if not args.source_column:
        args.source_column = input("请输入 source 列（例如 A）: ").strip().upper()
    if not args.target_column and not interactive_mode:
        raise ValueError("缺少 target 列，请使用 -t 或 --target-column 指定。")
    if not args.target_column:
        args.target_column = input("请输入 target 列（例如 B）: ").strip().upper()
    if args.start_row is None:
        if interactive_mode and len(sys.argv) == 1:
            start_row_text = input("请输入开始处理的行号（默认 2）: ").strip()
            args.start_row = int(start_row_text) if start_row_text else 2
        else:
            args.start_row = 2
    return args


def main() -> None:
    args = prompt_if_missing(parse_args())
    summary = process_excel(
        input_file=args.input_file,
        source_column=args.source_column,
        target_column=args.target_column,
        sheet=args.sheet,
        start_row=args.start_row,
        output_file=args.output,
    )
    print("处理完成。")
    print(f"工作表: {summary.worksheet_title}")
    print(f"source / target 列: {summary.source_column} / {summary.target_column}")
    print(f"开始行: {summary.start_row}")
    print(f"处理行数: {summary.processed_count}")
    print(f"含中文行数: {summary.matched_count}")
    print(f"输出文件: {summary.output_path}")


if __name__ == "__main__":
    main()
