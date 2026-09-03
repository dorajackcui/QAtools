#!/usr/bin/env python3
"""Shared helpers for reading Excel workbook metadata in GUI flows."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from tools.header_aliases import SOURCE_HEADER, TARGET_HEADER, HeaderAliases, HeaderAliasStore


@dataclass(frozen=True)
class WorkbookSheetChoices:
    sheet_names: tuple[str, ...]
    default_sheet: str | None


@dataclass(frozen=True)
class SourceTargetColumns:
    detected_source_column: str | None
    detected_target_column: str | None


def resolve_workbook_path(input_file: str | Path) -> Path:
    workbook_path = Path(input_file).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {workbook_path}")
    return workbook_path


def normalize_header(value: object) -> str:
    return "" if value is None else str(value).strip().casefold()


def list_workbook_sheets(input_file: str | Path) -> WorkbookSheetChoices:
    workbook_path = resolve_workbook_path(input_file)
    try:
        with ZipFile(workbook_path) as archive:
            workbook_xml = archive.read("xl/workbook.xml")
    except (BadZipFile, KeyError) as exc:
        raise InvalidFileException(
            f"无法读取 Excel 工作簿结构: {workbook_path}"
        ) from exc

    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    root = ElementTree.fromstring(workbook_xml)
    sheet_nodes = root.findall(f"{namespace}sheets/{namespace}sheet")
    sheet_names = tuple(node.attrib.get("name", "") for node in sheet_nodes)
    visible_indexes = [
        index
        for index, node in enumerate(sheet_nodes)
        if node.attrib.get("state", "visible") == "visible"
    ]
    view = root.find(f"{namespace}bookViews/{namespace}workbookView")
    try:
        active_index = int(view.attrib.get("activeTab", "0")) if view is not None else 0
    except ValueError:
        active_index = 0
    if active_index not in visible_indexes:
        active_index = visible_indexes[0] if visible_indexes else 0
    default_sheet = (
        sheet_names[active_index]
        if sheet_names and 0 <= active_index < len(sheet_names)
        else None
    )
    return WorkbookSheetChoices(
        sheet_names=sheet_names,
        default_sheet=default_sheet,
    )


def _preferred_header_column(
    headers: tuple[str, ...], custom_aliases: tuple[str, ...], builtin_header: str,
) -> str | None:
    for candidates in ({alias.casefold() for alias in custom_aliases}, {builtin_header}):
        matches = [
            get_column_letter(index)
            for index, header in enumerate(headers, start=1)
            if header in candidates
        ]
        if matches:
            # Ambiguous custom matches must not fall back to a built-in header.
            return matches[0] if len(matches) == 1 else None
    return None


def detect_source_target_columns(
    input_file: str | Path,
    sheet: str | None = None,
    *,
    header_aliases: HeaderAliases | None = None,
) -> SourceTargetColumns:
    configured_aliases = (
        header_aliases if header_aliases is not None else HeaderAliasStore().load()
    )
    aliases = HeaderAliases.create(
        source=configured_aliases.source,
        target=configured_aliases.target,
    )
    workbook_path = resolve_workbook_path(input_file)
    workbook = load_workbook(workbook_path, read_only=True)

    try:
        if sheet and sheet not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet}")

        worksheet = workbook[sheet] if sheet else workbook.active
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )

        headers = tuple(normalize_header(value) for value in header_row)
        return SourceTargetColumns(
            detected_source_column=_preferred_header_column(headers, aliases.source, SOURCE_HEADER),
            detected_target_column=_preferred_header_column(headers, aliases.target, TARGET_HEADER),
        )
    finally:
        workbook.close()
