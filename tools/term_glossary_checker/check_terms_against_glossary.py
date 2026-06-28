#!/usr/bin/env python3
"""Check whether bilingual Excel rows follow a glossary."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from tools.term_matching import (
    SUPPORTED_MATCH_MODES,
    TermMappingEntry as GlossaryEntry,
    build_matcher,
    find_row_terms,
    normalize_text,
    term_has_expected_target,
    term_has_simple_s_plural_variant,
    text_contains_term,
)
from tools.false_positive_review import (
    GLOSSARY_PROBLEM_MAPPING,
    Reviewer,
    apply_false_positive_review_to_sheet,
    review_clusters_with_codex,
)
from tools.excel_output import insert_row_problem_column
from tools.term_glossary_checker.workbook_output import (
    PROBLEM_SHEET_NAME,
    SUMMARY_SHEET_NAME,
    GlossaryProblemEntry,
    build_default_output_path,
    build_row_problem_summaries,
    write_problem_sheet,
    write_summary_sheet,
)


GLOSSARY_EMPTY_ROW_STOP_THRESHOLD = 1000
PLACEHOLDER_TARGET_VALUES = {"#N/A", "N/A", "NA", "#NA", "NAN"}


@dataclass(frozen=True)
class GlossaryConflict:
    source_term: str
    target_terms: tuple[str, ...]


@dataclass(frozen=True)
class CheckSummary:
    glossary_sheet_title: str
    data_sheet_title: str
    output_path: Path
    glossary_term_count: int
    conflict_count: int
    total_rows_checked: int
    matched_rows: int
    problem_rows: int
    problem_count: int
    case_sensitive: bool
    match_mode: str


def normalize_column(column_name: str) -> str:
    normalized = column_name.strip().upper()
    column_index_from_string(normalized)
    return normalized


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="按术语表检查双语 Excel 中的术语是否按 target 进行了翻译。"
    )
    parser.add_argument("glossary_file", nargs="?", help="术语表 Excel 文件路径，例如 glossary.xlsx")
    parser.add_argument("data_file", nargs="?", help="检查文本 Excel 文件路径，例如 data.xlsx")
    parser.add_argument("--glossary-sheet", help="术语表工作表名称，不填则使用活动工作表")
    parser.add_argument("--glossary-source-column", help="术语表 source 列，例如 A")
    parser.add_argument("--glossary-target-column", help="术语表 target 列，例如 B")
    parser.add_argument("--data-sheet", help="检查文本工作表名称，不填则使用活动工作表")
    parser.add_argument("--data-source-column", help="检查文本 source 列，例如 A")
    parser.add_argument("--data-target-column", help="检查文本 target 列，例如 B")
    parser.add_argument("--start-row", type=int, default=None, help="开始处理的行号，默认 2")
    parser.add_argument(
        "--case-sensitive",
        action="store_true",
        help="按大小写敏感方式匹配 source 术语和 target 译法",
    )
    parser.add_argument(
        "--match-mode",
        choices=SUPPORTED_MATCH_MODES,
        default="hybrid-boundary",
        help="匹配模式：hybrid-boundary 为混合边界，substring 为纯包含兼容模式",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="输出 Excel 文件路径，默认生成 glossary_check_<原文件名>",
    )
    parser.add_argument(
        "--codex-fp-review",
        action="store_true",
        help="检查完成后调用 Codex 对术语问题做假阳性筛查，并写回 fp_* 辅助列。",
    )
    parser.add_argument(
        "--codex-fp-sample-size",
        type=int,
        default=5,
        help="每个术语问题 cluster 发送给 Codex 的样本条数，默认 5。",
    )
    parser.add_argument(
        "--codex-model",
        help="Codex 假阳性筛查使用的模型；不填则使用 Codex 默认模型。",
    )
    parser.add_argument(
        "--codex-reasoning-effort",
        default="high",
        choices=("low", "medium", "high"),
        help="Codex 假阳性筛查使用的 reasoning effort，默认 high。",
    )
    return parser.parse_args()


def prompt_if_missing(args: argparse.Namespace) -> argparse.Namespace:
    interactive_mode = sys.stdin.isatty()

    if not args.glossary_file and not interactive_mode:
        raise ValueError("缺少术语表文件路径，请传入 glossary_file 参数。")
    if not args.data_file and not interactive_mode:
        raise ValueError("缺少检查文本文件路径，请传入 data_file 参数。")
    if not args.glossary_source_column and not interactive_mode:
        raise ValueError("缺少术语表 source 列，请使用 --glossary-source-column 指定。")
    if not args.glossary_target_column and not interactive_mode:
        raise ValueError("缺少术语表 target 列，请使用 --glossary-target-column 指定。")
    if not args.data_source_column and not interactive_mode:
        raise ValueError("缺少检查文本 source 列，请使用 --data-source-column 指定。")
    if not args.data_target_column and not interactive_mode:
        raise ValueError("缺少检查文本 target 列，请使用 --data-target-column 指定。")

    if not args.glossary_file:
        args.glossary_file = input("请输入术语表 Excel 文件路径: ").strip()
    if not args.data_file:
        args.data_file = input("请输入检查文本 Excel 文件路径: ").strip()
    if not args.glossary_sheet and interactive_mode and len(sys.argv) == 1:
        args.glossary_sheet = input("请输入术语表工作表名称（直接回车使用当前活动工作表）: ").strip() or None
    if not args.glossary_source_column:
        args.glossary_source_column = input("请输入术语表 source 列（例如 A）: ").strip().upper()
    if not args.glossary_target_column:
        args.glossary_target_column = input("请输入术语表 target 列（例如 B）: ").strip().upper()
    if not args.data_sheet and interactive_mode and len(sys.argv) == 1:
        args.data_sheet = input("请输入检查文本工作表名称（直接回车使用当前活动工作表）: ").strip() or None
    if not args.data_source_column:
        args.data_source_column = input("请输入检查文本 source 列（例如 A）: ").strip().upper()
    if not args.data_target_column:
        args.data_target_column = input("请输入检查文本 target 列（例如 B）: ").strip().upper()
    if args.start_row is None:
        if interactive_mode and len(sys.argv) == 1:
            start_row_text = input("请输入开始处理的行号（默认 2）: ").strip()
            args.start_row = int(start_row_text) if start_row_text else 2
        else:
            args.start_row = 2
    return args


def row_value(row: tuple[object, ...], column_index: int) -> object:
    return row[column_index - 1] if len(row) >= column_index else None


def row_values_are_empty(*values: object) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def is_placeholder_glossary_entry(target_term: str) -> bool:
    return target_term.strip().upper() in PLACEHOLDER_TARGET_VALUES


def load_glossary_entries(
    glossary_file: str | Path,
    source_column: str,
    target_column: str,
    sheet: str | None = None,
    start_row: int = 2,
    case_sensitive: bool = False,
) -> tuple[str, list[GlossaryEntry], list[GlossaryConflict]]:
    glossary_path = Path(glossary_file).expanduser().resolve()
    if not glossary_path.exists():
        raise FileNotFoundError(f"术语表文件不存在: {glossary_path}")
    if start_row < 1:
        raise ValueError("开始行必须大于等于 1。")

    source_column = normalize_column(source_column)
    target_column = normalize_column(target_column)
    source_column_index = column_index_from_string(source_column)
    target_column_index = column_index_from_string(target_column)
    max_column_index = max(source_column_index, target_column_index)

    workbook = load_workbook(glossary_path, read_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.active

        source_to_targets: dict[str, dict[str, tuple[str, str]]] = {}
        consecutive_empty_rows = 0

        for row in worksheet.iter_rows(
            min_row=start_row,
            max_col=max_column_index,
            values_only=True,
        ):
            raw_source = row_value(row, source_column_index)
            raw_target = row_value(row, target_column_index)
            if row_values_are_empty(raw_source, raw_target):
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= GLOSSARY_EMPTY_ROW_STOP_THRESHOLD:
                    break
                continue

            consecutive_empty_rows = 0
            source_term = "" if raw_source is None else str(raw_source).strip()
            target_term = "" if raw_target is None else str(raw_target).strip()

            if not source_term or not target_term:
                continue
            if is_placeholder_glossary_entry(target_term):
                continue

            normalized_source = normalize_text(source_term, case_sensitive=case_sensitive)
            normalized_target = normalize_text(target_term, case_sensitive=case_sensitive)
            if not normalized_source or not normalized_target:
                continue

            targets = source_to_targets.setdefault(normalized_source, {})
            targets.setdefault(normalized_target, (source_term, target_term))

        worksheet_title = worksheet.title
    finally:
        workbook.close()

    entries: list[GlossaryEntry] = []
    conflicts: list[GlossaryConflict] = []
    for normalized_source, targets in source_to_targets.items():
        if len(targets) == 1:
            source_term, target_term = next(iter(targets.values()))
            entries.append(
                GlossaryEntry(
                    source_term=source_term,
                    target_term=target_term,
                    normalized_source=normalized_source,
                    normalized_target=normalize_text(target_term, case_sensitive=case_sensitive),
                )
            )
            continue

        source_term = next(iter(targets.values()))[0]
        conflict_targets = tuple(sorted(target_term for _, target_term in targets.values()))
        conflicts.append(GlossaryConflict(source_term=source_term, target_terms=conflict_targets))

    entries.sort(key=lambda entry: (len(entry.normalized_source), entry.normalized_source), reverse=True)
    conflicts.sort(key=lambda item: item.source_term)
    return worksheet_title, entries, conflicts


def process_excel(
    glossary_file: str | Path,
    data_file: str | Path,
    glossary_source_column: str,
    glossary_target_column: str,
    data_source_column: str,
    data_target_column: str,
    glossary_sheet: str | None = None,
    data_sheet: str | None = None,
    start_row: int = 2,
    case_sensitive: bool = False,
    match_mode: str = "hybrid-boundary",
    output_file: str | Path | None = None,
    false_positive_reviewer: Reviewer | None = None,
    false_positive_sample_size: int = 5,
) -> CheckSummary:
    if start_row < 1:
        raise ValueError("开始行必须大于等于 1。")
    if match_mode not in SUPPORTED_MATCH_MODES:
        raise ValueError(f"不支持的匹配模式: {match_mode}")

    glossary_source_column = normalize_column(glossary_source_column)
    glossary_target_column = normalize_column(glossary_target_column)
    data_source_column = normalize_column(data_source_column)
    data_target_column = normalize_column(data_target_column)

    glossary_sheet_title, entries, conflicts = load_glossary_entries(
        glossary_file=glossary_file,
        source_column=glossary_source_column,
        target_column=glossary_target_column,
        sheet=glossary_sheet,
        start_row=start_row,
        case_sensitive=case_sensitive,
    )

    data_path = Path(data_file).expanduser().resolve()
    if not data_path.exists():
        raise FileNotFoundError(f"检查文本文件不存在: {data_path}")
    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_output_path(data_path)
    )

    workbook = load_workbook(data_path)
    worksheet = workbook[data_sheet] if data_sheet else workbook.active
    matcher = build_matcher(entries)

    total_rows_checked = max(0, worksheet.max_row - start_row + 1)
    matched_rows = 0
    problem_row_set: set[int] = set()
    problem_entries: list[GlossaryProblemEntry] = []

    for row_index in range(start_row, worksheet.max_row + 1):
        source_text = worksheet[f"{data_source_column}{row_index}"].value
        target_text = worksheet[f"{data_target_column}{row_index}"].value
        matched_entries = find_row_terms(
            source_text,
            matcher,
            case_sensitive=case_sensitive,
            match_mode=match_mode,
        )
        if not matched_entries:
            continue

        matched_rows += 1
        normalized_source_text = normalize_text(source_text, case_sensitive=case_sensitive)
        normalized_target_text = normalize_text(target_text, case_sensitive=case_sensitive)
        source_snapshot = "" if source_text is None else str(source_text)
        target_snapshot = "" if target_text is None else str(target_text)

        for entry in matched_entries:
            if term_has_expected_target(
                normalized_source_text,
                normalized_target_text,
                entry,
                match_mode=match_mode,
                allow_target_plural_variants=True,
            ):
                continue

            if term_has_simple_s_plural_variant(
                normalized_source_text,
                normalized_target_text,
                entry,
                match_mode=match_mode,
            ):
                continue

            problem_row_set.add(row_index)
            problem_entries.append(
                (
                    row_index,
                    "术语未按术语表翻译",
                    entry.source_term,
                    entry.target_term,
                    source_snapshot,
                    target_snapshot,
                )
            )

    summary = CheckSummary(
        glossary_sheet_title=glossary_sheet_title,
        data_sheet_title=worksheet.title,
        output_path=output_path,
        glossary_term_count=len(entries),
        conflict_count=len(conflicts),
        total_rows_checked=total_rows_checked,
        matched_rows=matched_rows,
        problem_rows=len(problem_row_set),
        problem_count=len(problem_entries),
        case_sensitive=case_sensitive,
        match_mode=match_mode,
    )

    insert_row_problem_column(
        worksheet,
        data_target_column,
        build_row_problem_summaries(problem_entries),
    )
    write_problem_sheet(workbook, worksheet.title, problem_entries)
    write_summary_sheet(
        workbook=workbook,
        worksheet_title=worksheet.title,
        summary=summary,
        conflicts=conflicts,
        glossary_source_column=glossary_source_column,
        glossary_target_column=glossary_target_column,
        data_source_column=data_source_column,
        data_target_column=data_target_column,
        start_row=start_row,
        match_mode=match_mode,
    )
    if false_positive_reviewer is not None:
        apply_false_positive_review_to_sheet(
            workbook,
            PROBLEM_SHEET_NAME,
            GLOSSARY_PROBLEM_MAPPING,
            reviewer=false_positive_reviewer,
            sample_size=false_positive_sample_size,
        )
    workbook.save(output_path)
    return summary


def main() -> None:
    args = prompt_if_missing(parse_args())
    false_positive_reviewer = None
    if args.codex_fp_review:
        false_positive_reviewer = lambda clusters: review_clusters_with_codex(  # noqa: E731
            clusters,
            model=args.codex_model,
            reasoning_effort=args.codex_reasoning_effort,
        )
    summary = process_excel(
        glossary_file=args.glossary_file,
        data_file=args.data_file,
        glossary_sheet=args.glossary_sheet,
        glossary_source_column=args.glossary_source_column,
        glossary_target_column=args.glossary_target_column,
        data_sheet=args.data_sheet,
        data_source_column=args.data_source_column,
        data_target_column=args.data_target_column,
        start_row=args.start_row,
        case_sensitive=args.case_sensitive,
        match_mode=args.match_mode,
        output_file=args.output,
        false_positive_reviewer=false_positive_reviewer,
        false_positive_sample_size=args.codex_fp_sample_size,
    )

    print("处理完成。")
    print(f"术语表工作表: {summary.glossary_sheet_title}")
    print(f"检查工作表: {summary.data_sheet_title}")
    print(f"大小写模式: {'严格区分' if summary.case_sensitive else '忽略大小写'}")
    print(f"匹配模式: {'混合边界' if summary.match_mode == 'hybrid-boundary' else '纯包含'}")
    print(f"术语表条数: {summary.glossary_term_count}")
    print(f"冲突术语数: {summary.conflict_count}")
    print(f"总行数: {summary.total_rows_checked}")
    print(f"命中术语行数: {summary.matched_rows}")
    print(f"问题行数: {summary.problem_rows}")
    print(f"问题条数: {summary.problem_count}")
    if args.codex_fp_review:
        print("Codex 假阳性筛查: 已写入 fp_* 辅助列")
    print(f"输出文件: {summary.output_path}")


if __name__ == "__main__":
    main()
