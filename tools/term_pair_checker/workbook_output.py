"""Workbook output helpers for term pair checking."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl.styles import Font, PatternFill

from tools.excel_output import (
    PROBLEM_BASE_HEADERS,
    ROW_PROBLEM_SEPARATOR,
    build_prefixed_output_path,
    format_row_problem_text,
    join_unique_text,
    write_output_table,
)
from tools.term_matching import normalize_text

if TYPE_CHECKING:
    from tools.term_pair_checker.extract_terms_from_excel import (
        ProblemEntry,
        RecordedTermPair,
    )


TERM_SHEET_NAME = "术语表"
PROBLEM_SHEET_NAME = "问题列"
LEGACY_TERM_SHEET_NAMES = ("术语表（无mark）", "术语汇总")
TERM_STATUS_HEADER = "本批次是否有问题"


def build_default_output_path(input_path: Path) -> Path:
    return build_prefixed_output_path(input_path, "term_pair_check_")


def build_row_problem_summaries(problem_entries: Iterable["ProblemEntry"]) -> dict[int, str]:
    summaries_by_row: dict[int, list[str]] = {}
    for problem_entry in problem_entries:
        summary = format_row_problem_text(
            problem_entry.problem_source_term,
            problem_entry.expected_target_term,
            problem_entry.description,
        )
        if summary:
            summaries_by_row.setdefault(problem_entry.row_index, []).append(summary)
    return {
        row_index: ROW_PROBLEM_SEPARATOR.join(summaries)
        for row_index, summaries in summaries_by_row.items()
    }


def write_term_sheet(
    workbook,
    worksheet_title: str,
    term_pairs: Iterable["RecordedTermPair"],
    *,
    problem_entries: Iterable["ProblemEntry"] = (),
) -> None:
    problem_term_keys: set[str] = set()
    for entry in problem_entries:
        # Use structured term identities: term text can itself contain separators.
        source_terms = entry.affected_source_terms
        if source_terms is None:
            source_terms = (entry.problem_source_term,) if entry.problem_source_term else ()
        for source_term in source_terms:
            key = normalize_text(source_term, case_sensitive=False)
            problem_term_keys.add(key)

    rows = []
    for term_pair in term_pairs:
        key = normalize_text(term_pair.source_plain_text, case_sensitive=False)
        rows.append((
            term_pair.source_display_text or None,
            term_pair.target_display_text or None,
            term_pair.source_plain_text or None,
            term_pair.target_plain_text or None,
            term_pair.term_source or None,
            "有问题" if key in problem_term_keys else "无问题",
        ))
    term_sheet = write_output_table(
        workbook,
        current_sheet_name=worksheet_title,
        sheet_name=TERM_SHEET_NAME,
        headers=(
            "source术语", "target术语", "source术语（无mark）",
            "target术语（无mark）", "术语来源", TERM_STATUS_HEADER,
        ),
        rows=rows,
    )
    for column in ("A", "B", "C", "D"):
        term_sheet.column_dimensions[column].width = 36
    term_sheet.column_dimensions["E"].width = 16
    term_sheet.column_dimensions["F"].width = 24
    problem_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    problem_font = Font(color="9C0006", bold=True)
    for row_index, row in enumerate(rows, start=2):
        if row[-1] != "有问题":
            continue
        status_cell = term_sheet.cell(row_index, 6)
        status_cell.fill = problem_fill
        status_cell.font = problem_font


def write_problem_sheet(
    workbook,
    worksheet_title: str,
    target_column: str,
    problem_entries: Iterable["ProblemEntry"],
    *,
    format_output: bool = True,
) -> None:
    entries_by_row: dict[int, list["ProblemEntry"]] = {}
    for problem_entry in problem_entries:
        entries_by_row.setdefault(problem_entry.row_index, []).append(problem_entry)

    rows = []
    for source_row, row_entries in entries_by_row.items():
        first_entry = row_entries[0]
        descriptions = (
            format_row_problem_text(
                entry.problem_source_term,
                entry.expected_target_term,
                entry.description,
            )
            for entry in row_entries
        )
        rows.append(
            (
                source_row,
                first_entry.source_snapshot,
                first_entry.target_snapshot,
                join_unique_text(descriptions),
                join_unique_text(entry.problem_source_term for entry in row_entries),
                join_unique_text(entry.expected_target_term for entry in row_entries),
                join_unique_text(entry.term_source for entry in row_entries),
            )
        )

    write_output_table(
        workbook,
        current_sheet_name=worksheet_title,
        sheet_name=PROBLEM_SHEET_NAME,
        headers=PROBLEM_BASE_HEADERS
        + ("source术语", "预期target术语", "术语来源"),
        rows=rows,
        row_link_target_column=target_column,
        format_output=format_output,
    )


def delete_legacy_term_sheets(workbook) -> None:
    for sheet_name in LEGACY_TERM_SHEET_NAMES:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
