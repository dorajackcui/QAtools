#!/usr/bin/env python3
"""Shared helpers for reading Toolshub history TB workbooks."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


DEFAULT_HISTORY_SHEET_NAME = "术语表"
HISTORY_SOURCE_EXACT_HEADERS = {"source"}
HISTORY_TARGET_EXACT_HEADERS = {"target"}
HISTORY_SOURCE_NO_MARK_HEADERS = {
    "source(无mark)",
    "source术语(无mark)",
}
HISTORY_TARGET_NO_MARK_HEADERS = {
    "target(无mark)",
    "target术语(无mark)",
}
HISTORY_SOURCE_MARKED_HEADERS = {"source术语"}
HISTORY_TARGET_MARKED_HEADERS = {"target术语"}


@dataclass(frozen=True)
class HistoryTbColumns:
    sheet_title: str
    source_column: str | None
    target_column: str | None


@dataclass(frozen=True)
class HistoryTbRow:
    row_index: int
    source_text: str
    target_text: str


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_history_header(value: object) -> str:
    return re.sub(
        r"\s+",
        "",
        cell_text(value)
        .casefold()
        .replace("（", "(")
        .replace("）", ")"),
    )


def resolve_history_path(history_tb_file: str | Path) -> Path:
    history_path = Path(history_tb_file).expanduser().resolve()
    if not history_path.exists():
        raise FileNotFoundError(f"历史 TB 文件不存在: {history_path}")
    return history_path


def choose_history_worksheet(
    workbook: Any,
    sheet: str | None,
    *,
    preferred_sheet: str = DEFAULT_HISTORY_SHEET_NAME,
) -> Any:
    if sheet:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"历史 TB 工作表不存在: {sheet}")
        return workbook[sheet]
    if preferred_sheet in workbook.sheetnames:
        return workbook[preferred_sheet]
    return workbook.active


def _source_candidate_groups(prefer_no_mark: bool) -> tuple[set[str], ...]:
    if prefer_no_mark:
        return (
            HISTORY_SOURCE_EXACT_HEADERS,
            HISTORY_SOURCE_NO_MARK_HEADERS,
            HISTORY_SOURCE_MARKED_HEADERS,
        )
    return (
        HISTORY_SOURCE_EXACT_HEADERS,
        HISTORY_SOURCE_MARKED_HEADERS,
        HISTORY_SOURCE_NO_MARK_HEADERS,
    )


def _target_candidate_groups(prefer_no_mark: bool) -> tuple[set[str], ...]:
    if prefer_no_mark:
        return (
            HISTORY_TARGET_EXACT_HEADERS,
            HISTORY_TARGET_NO_MARK_HEADERS,
            HISTORY_TARGET_MARKED_HEADERS,
        )
    return (
        HISTORY_TARGET_EXACT_HEADERS,
        HISTORY_TARGET_MARKED_HEADERS,
        HISTORY_TARGET_NO_MARK_HEADERS,
    )


def _column_from_argument(worksheet: Any, column: str | None, header_row: int) -> str | None:
    if not column:
        return None
    try:
        column_index_from_string(column.strip().upper())
        return column.strip().upper()
    except ValueError:
        expected_header = normalize_history_header(column)
        for column_index in range(1, worksheet.max_column + 1):
            if normalize_history_header(worksheet.cell(header_row, column_index).value) == expected_header:
                return get_column_letter(column_index)
        raise


def _find_header_column(
    headers: list[tuple[str, str]],
    candidate_groups: tuple[set[str], ...],
    *,
    require_unique_matches: bool = False,
) -> str | None:
    for candidates in candidate_groups:
        matches = [
            column_letter
            for column_letter, normalized_header in headers
            if normalized_header in candidates
        ]
        if require_unique_matches:
            if len(matches) == 1:
                return matches[0]
            continue
        if matches:
            return matches[0]
    return None


def _non_empty_headers(worksheet: Any, header_row: int) -> list[tuple[str, str]]:
    headers: list[tuple[str, str]] = []
    for column_index in range(1, worksheet.max_column + 1):
        normalized_header = normalize_history_header(worksheet.cell(header_row, column_index).value)
        if normalized_header:
            headers.append((get_column_letter(column_index), normalized_header))
    return headers


def detect_history_columns_from_header_row(
    worksheet: Any,
    *,
    header_row: int,
    source_column: str | None = None,
    target_column: str | None = None,
    prefer_no_mark: bool = True,
    allow_partial: bool = False,
    require_unique_header_matches: bool = False,
) -> tuple[str | None, str | None]:
    if header_row < 1:
        raise ValueError("历史 TB 表头行必须大于等于 1。")

    detected_source_column = _column_from_argument(worksheet, source_column, header_row)
    detected_target_column = _column_from_argument(worksheet, target_column, header_row)
    headers = _non_empty_headers(worksheet, header_row)

    if not detected_source_column:
        detected_source_column = _find_header_column(
            headers,
            _source_candidate_groups(prefer_no_mark),
            require_unique_matches=require_unique_header_matches,
        )
    if not detected_target_column:
        detected_target_column = _find_header_column(
            headers,
            _target_candidate_groups(prefer_no_mark),
            require_unique_matches=require_unique_header_matches,
        )

    if (not detected_source_column or not detected_target_column) and len(headers) == 2:
        fallback_columns = [column for column, _header in headers]
        if detected_source_column and not detected_target_column:
            detected_target_column = next(column for column in fallback_columns if column != detected_source_column)
        elif detected_target_column and not detected_source_column:
            detected_source_column = next(column for column in fallback_columns if column != detected_target_column)
        else:
            detected_source_column, detected_target_column = fallback_columns

    if detected_source_column and detected_target_column and detected_source_column == detected_target_column:
        raise ValueError("历史 TB source/target 列不能相同。")
    if not allow_partial and (not detected_source_column or not detected_target_column):
        raise ValueError(
            "历史 TB 缺少 source/target 列，请指定 source/target 列或使用可识别表头。"
        )
    return detected_source_column, detected_target_column


def detect_history_tb_columns(
    history_tb_file: str | Path,
    *,
    sheet: str | None = None,
    source_column: str | None = None,
    target_column: str | None = None,
    start_row: int = 2,
    header_row: int | None = None,
    preferred_sheet: str = DEFAULT_HISTORY_SHEET_NAME,
    prefer_no_mark: bool = True,
    allow_partial: bool = False,
    require_unique_header_matches: bool = False,
) -> HistoryTbColumns:
    if start_row < 1:
        raise ValueError("历史 TB 开始行必须大于等于 1。")
    if header_row is not None and header_row < 1:
        raise ValueError("历史 TB 表头行必须大于等于 1。")

    workbook = load_workbook(resolve_history_path(history_tb_file), read_only=True, data_only=True)
    try:
        worksheet = choose_history_worksheet(workbook, sheet, preferred_sheet=preferred_sheet)
        source_column, target_column = detect_history_columns_from_header_row(
            worksheet,
            header_row=header_row or max(1, start_row - 1),
            source_column=source_column,
            target_column=target_column,
            prefer_no_mark=prefer_no_mark,
            allow_partial=allow_partial,
            require_unique_header_matches=require_unique_header_matches,
        )
        return HistoryTbColumns(
            sheet_title=worksheet.title,
            source_column=source_column,
            target_column=target_column,
        )
    finally:
        workbook.close()


def iter_history_rows(
    history_tb_file: str | Path,
    *,
    sheet: str | None = None,
    source_column: str | None = None,
    target_column: str | None = None,
    start_row: int = 2,
    header_row: int | None = None,
    preferred_sheet: str = DEFAULT_HISTORY_SHEET_NAME,
    prefer_no_mark: bool = True,
    empty_row_stop_threshold: int | None = None,
    require_unique_header_matches: bool = False,
) -> tuple[str, str, str, tuple[HistoryTbRow, ...]]:
    if start_row < 1:
        raise ValueError("历史 TB 开始行必须大于等于 1。")
    if header_row is not None and header_row < 1:
        raise ValueError("历史 TB 表头行必须大于等于 1。")

    workbook = load_workbook(resolve_history_path(history_tb_file), read_only=True, data_only=True)
    try:
        worksheet = choose_history_worksheet(workbook, sheet, preferred_sheet=preferred_sheet)
        detected_source_column, detected_target_column = detect_history_columns_from_header_row(
            worksheet,
            header_row=header_row or max(1, start_row - 1),
            source_column=source_column,
            target_column=target_column,
            prefer_no_mark=prefer_no_mark,
            require_unique_header_matches=require_unique_header_matches,
        )
        assert detected_source_column is not None
        assert detected_target_column is not None

        rows: list[HistoryTbRow] = []
        consecutive_empty_rows = 0
        for row_index in range(start_row, worksheet.max_row + 1):
            source_text = cell_text(worksheet[f"{detected_source_column}{row_index}"].value)
            target_text = cell_text(worksheet[f"{detected_target_column}{row_index}"].value)
            if not source_text and not target_text:
                consecutive_empty_rows += 1
                if (
                    empty_row_stop_threshold is not None
                    and consecutive_empty_rows >= empty_row_stop_threshold
                ):
                    break
                continue
            consecutive_empty_rows = 0
            if not source_text or not target_text:
                continue
            rows.append(
                HistoryTbRow(
                    row_index=row_index,
                    source_text=source_text,
                    target_text=target_text,
                )
            )
        return worksheet.title, detected_source_column, detected_target_column, tuple(rows)
    finally:
        workbook.close()
