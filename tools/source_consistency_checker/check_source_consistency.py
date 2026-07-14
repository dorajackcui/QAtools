#!/usr/bin/env python3
"""Find identical source text associated with different target text."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from tools.excel_output import (
    PROBLEM_BASE_HEADERS,
    build_prefixed_output_path,
    write_output_table,
)


PROBLEM_SHEET_NAME = "同源译文不一致"


@dataclass(frozen=True)
class SourceOccurrence:
    row_index: int
    target_text: str


@dataclass(frozen=True)
class CheckSummary:
    output_path: Path
    worksheet_title: str
    source_column: str
    target_column: str
    start_row: int
    total_rows_checked: int
    non_empty_source_rows: int
    repeated_source_count: int
    inconsistent_source_count: int
    problem_rows: int


def normalize_column(column_name: str) -> str:
    normalized = column_name.strip().upper()
    column_index_from_string(normalized)
    return normalized


def cell_text(value: object) -> str:
    return "" if value is None else str(value)


def build_default_output_path(input_file: str | Path) -> Path:
    return build_prefixed_output_path(input_file, "source_consistency_check_")


def write_problem_sheet(
    workbook,
    worksheet_title: str,
    target_column: str,
    problem_entries: list[tuple[int, str, str, str, int, str]],
) -> None:
    write_output_table(
        workbook,
        current_sheet_name=worksheet_title,
        sheet_name=PROBLEM_SHEET_NAME,
        headers=PROBLEM_BASE_HEADERS + ("target版本数", "同组行号"),
        rows=problem_entries,
        row_link_target_column=target_column,
    )


def process_excel(
    input_file: str | Path,
    source_column: str,
    target_column: str,
    sheet: str | None = None,
    start_row: int = 2,
    output_file: str | Path | None = None,
) -> CheckSummary:
    if start_row < 1:
        raise ValueError("开始行必须大于等于 1。")

    input_path = Path(input_file).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    source_column = normalize_column(source_column)
    target_column = normalize_column(target_column)
    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_output_path(input_path)
    )

    workbook = load_workbook(input_path)
    worksheet = workbook[sheet] if sheet else workbook.active
    occurrences_by_source: dict[str, list[SourceOccurrence]] = {}

    for row_index in range(start_row, worksheet.max_row + 1):
        source_text = cell_text(worksheet[f"{source_column}{row_index}"].value)
        if not source_text.strip():
            continue
        target_text = cell_text(worksheet[f"{target_column}{row_index}"].value)
        occurrences_by_source.setdefault(source_text, []).append(
            SourceOccurrence(row_index=row_index, target_text=target_text)
        )

    repeated_source_count = 0
    inconsistent_source_count = 0
    problem_entries: list[tuple[int, str, str, str, int, str]] = []
    for source_text, occurrences in occurrences_by_source.items():
        if len(occurrences) < 2:
            continue
        repeated_source_count += 1
        target_variants = {occurrence.target_text for occurrence in occurrences}
        if len(target_variants) < 2:
            continue

        inconsistent_source_count += 1
        grouped_rows = "、".join(str(occurrence.row_index) for occurrence in occurrences)
        for occurrence in occurrences:
            problem_entries.append(
                (
                    occurrence.row_index,
                    source_text,
                    occurrence.target_text,
                    f"同一 source 对应 {len(target_variants)} 个不同 target",
                    len(target_variants),
                    grouped_rows,
                )
            )

    write_problem_sheet(workbook, worksheet.title, target_column, problem_entries)
    workbook.save(output_path)
    return CheckSummary(
        output_path=output_path,
        worksheet_title=worksheet.title,
        source_column=source_column,
        target_column=target_column,
        start_row=start_row,
        total_rows_checked=max(0, worksheet.max_row - start_row + 1),
        non_empty_source_rows=sum(len(items) for items in occurrences_by_source.values()),
        repeated_source_count=repeated_source_count,
        inconsistent_source_count=inconsistent_source_count,
        problem_rows=len(problem_entries),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="检查 Excel 中是否存在完全相同的 source 对应不同 target。"
    )
    parser.add_argument("input_file", nargs="?", help="输入 Excel 文件路径，例如 input.xlsx")
    parser.add_argument("-s", "--sheet", help="工作表名称，不填则使用活动工作表")
    parser.add_argument("-c", "--source-column", help="source 列，例如 A")
    parser.add_argument("-t", "--target-column", help="target 列，例如 B")
    parser.add_argument("--start-row", type=int, default=None, help="开始处理的行号，默认 2")
    parser.add_argument(
        "-o",
        "--output",
        help="输出 Excel 文件路径，默认生成 source_consistency_check_<原文件名>",
    )
    return parser.parse_args()


def prompt_if_missing(args: argparse.Namespace) -> argparse.Namespace:
    interactive_mode = sys.stdin.isatty()
    if not args.input_file and not interactive_mode:
        raise ValueError("缺少输入文件路径，请传入 input_file 参数。")
    if not args.input_file:
        args.input_file = input("请输入 Excel 文件路径: ").strip()
    if not args.sheet and interactive_mode and len(sys.argv) == 1:
        args.sheet = input("请输入工作表名称（直接回车使用当前活动工作表）: ").strip() or None
    if not args.source_column and not interactive_mode:
        raise ValueError("缺少 source 列，请使用 -c 或 --source-column 指定。")
    if not args.source_column:
        args.source_column = input("请输入 source 列（例如 A）: ").strip().upper()
    if not args.target_column and not interactive_mode:
        raise ValueError("缺少 target 列，请使用 -t 或 --target-column 指定。")
    if not args.target_column:
        args.target_column = input("请输入 target 列（例如 B）: ").strip().upper()
    if args.start_row is None:
        if interactive_mode and len(sys.argv) == 1:
            raw_start_row = input("请输入开始处理的行号（默认 2）: ").strip()
            args.start_row = int(raw_start_row) if raw_start_row else 2
        else:
            args.start_row = 2
    return args


def main() -> None:
    args = prompt_if_missing(parse_args())
    summary = process_excel(
        input_file=args.input_file,
        source_column=args.source_column,
        target_column=args.target_column,
        sheet=args.sheet,
        start_row=args.start_row,
        output_file=args.output,
    )
    print("处理完成。")
    print(f"检查工作表: {summary.worksheet_title}")
    print(f"source / target 列: {summary.source_column} / {summary.target_column}")
    print(f"总行数: {summary.total_rows_checked}")
    print(f"重复 source 数: {summary.repeated_source_count}")
    print(f"不一致 source 数: {summary.inconsistent_source_count}")
    print(f"问题行数: {summary.problem_rows}")
    print(f"输出文件: {summary.output_path}")


if __name__ == "__main__":
    main()
