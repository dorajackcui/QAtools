#!/usr/bin/env python3
"""Run multiple Excel checks in sequence and keep results in one workbook."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl.utils import column_index_from_string

from tools.chinese_target_checker.check_chinese_target import (
    PROBLEM_SHEET_NAME as CHINESE_PROBLEM_SHEET_NAME,
    process_excel as run_chinese_target_check_excel,
)
from tools.content_fidelity_checker.check_content_fidelity import (
    NUMBER_PROBLEM_SHEET_NAME,
    NUMBER_RULE,
    URL_PROBLEM_SHEET_NAME,
    URL_RULE,
    process_excel as run_content_fidelity_check_excel,
)
from tools.excel_output import (
    ROW_PROBLEM_COLUMN_HEADER,
    build_prefixed_output_path,
    load_workbook_for_editing,
    validate_distinct_source_target_columns,
)
from tools.line_break_checker.check_line_breaks import (
    PROBLEM_SHEET_NAME as LINE_BREAK_PROBLEM_SHEET_NAME,
    process_excel as run_line_break_check_excel,
)
from tools.source_consistency_checker.check_source_consistency import (
    PROBLEM_SHEET_NAME as SOURCE_CONSISTENCY_PROBLEM_SHEET_NAME,
    process_excel as run_source_consistency_check_excel,
)
from tools.target_consistency_checker.check_target_consistency import (
    PROBLEM_SHEET_NAME as TARGET_CONSISTENCY_PROBLEM_SHEET_NAME,
    process_excel as run_target_consistency_check_excel,
)
from tools.tag_placeholder_checker.check_tags_and_placeholders import (
    PROBLEM_SHEET_NAME as TAG_PROBLEM_SHEET_NAME,
    SUMMARY_SHEET_NAME as TAG_SUMMARY_SHEET_NAME,
    process_excel as run_tag_check_excel,
)
from tools.target_text_checker.check_target_text import (
    PROBLEM_SHEET_NAME as TARGET_TEXT_PROBLEM_SHEET_NAME,
    process_excel as run_target_text_check_excel,
)
from tools.term_pair_checker.extract_terms_from_excel import (
    process_excel as run_term_pair_check_excel,
)
from tools.term_pair_checker.workbook_output import (
    PROBLEM_SHEET_NAME as TERM_PROBLEM_SHEET_NAME,
    TERM_SHEET_NAME,
)
from tools.workflow.review_sheet import WORKFLOW_REVIEW_SHEET_NAME, write_review_sheet


WORKFLOW_TERM_PROBLEM_SHEET_NAME = "术语问题"
WORKFLOW_SUMMARY_SHEET_NAME = "质量检查汇总"


@dataclass(frozen=True)
class WorkflowSummary:
    output_path: Path
    worksheet_title: str
    source_column: str
    target_column: str
    start_row: int
    ran_term_pair_check: bool
    ran_tag_check: bool
    ran_line_break_check: bool
    ran_source_consistency_check: bool
    ran_chinese_target_check: bool
    ran_target_text_check: bool
    term_count: int
    term_problem_count: int
    term_problem_rows: int
    tag_problem_count: int
    tag_problem_rows: int
    line_break_problem_count: int
    source_consistency_problem_count: int
    source_consistency_problem_rows: int
    chinese_target_problem_count: int
    target_text_problem_count: int
    target_text_problem_rows: int
    ran_target_consistency_check: bool = False
    ran_number_check: bool = False
    ran_url_check: bool = False
    target_consistency_problem_count: int = 0
    target_consistency_problem_rows: int = 0
    number_problem_rows: int = 0
    url_problem_rows: int = 0


def build_default_output_path(input_file: str | Path) -> Path:
    return build_prefixed_output_path(input_file, "workflow_check_")


def count_unique_problem_rows(worksheet) -> int:
    """Count unique source row numbers recorded in an output worksheet's first column."""
    return len(
        {
            row_number
            for (row_number,) in worksheet.iter_rows(
                min_row=2,
                min_col=1,
                max_col=1,
                values_only=True,
            )
            if row_number is not None
        }
    )


def finalize_workflow_output(
    *,
    output_path: Path,
    worksheet_title: str,
    run_term_pair_check: bool,
    run_tag_check: bool,
    run_line_break_check: bool,
    run_source_consistency_check: bool,
    run_target_consistency_check: bool,
    run_number_check: bool,
    run_url_check: bool,
    run_chinese_target_check: bool,
    run_target_text_check: bool,
    tag_problem_rows: int,
    line_break_problem_rows: int,
    source_consistency_problem_rows: int,
    target_consistency_problem_rows: int,
    number_problem_rows: int,
    url_problem_rows: int,
    chinese_target_problem_rows: int,
    target_text_problem_rows: int,
    input_file: Path,
    source_column: str,
    target_column: str,
    start_row: int,
) -> int:
    """Normalize workflow-only sheets and write the compact quality summary."""
    workbook = load_workbook_for_editing(output_path)
    term_problem_rows = 0

    if run_term_pair_check and TERM_PROBLEM_SHEET_NAME in workbook.sheetnames:
        if worksheet_title == WORKFLOW_TERM_PROBLEM_SHEET_NAME:
            raise ValueError(
                f"数据工作表名称不能为 {WORKFLOW_TERM_PROBLEM_SHEET_NAME}"
            )
        term_problem_sheet = workbook[TERM_PROBLEM_SHEET_NAME]
        term_problem_rows = count_unique_problem_rows(term_problem_sheet)
        if WORKFLOW_TERM_PROBLEM_SHEET_NAME in workbook.sheetnames:
            del workbook[WORKFLOW_TERM_PROBLEM_SHEET_NAME]
        term_problem_sheet.title = WORKFLOW_TERM_PROBLEM_SHEET_NAME

    if worksheet_title == WORKFLOW_SUMMARY_SHEET_NAME:
        raise ValueError(f"数据工作表名称不能为 {WORKFLOW_SUMMARY_SHEET_NAME}")

    if TAG_SUMMARY_SHEET_NAME in workbook.sheetnames:
        del workbook[TAG_SUMMARY_SHEET_NAME]
    if WORKFLOW_SUMMARY_SHEET_NAME in workbook.sheetnames:
        del workbook[WORKFLOW_SUMMARY_SHEET_NAME]

    problem_sheets = []
    if run_term_pair_check:
        problem_sheets.append(("术语检查", WORKFLOW_TERM_PROBLEM_SHEET_NAME))
    if run_source_consistency_check:
        problem_sheets.append(
            ("同 Source 不同 Target", SOURCE_CONSISTENCY_PROBLEM_SHEET_NAME)
        )
    if run_target_consistency_check:
        problem_sheets.append(
            ("同 Target 不同 Source", TARGET_CONSISTENCY_PROBLEM_SHEET_NAME)
        )
    if run_tag_check:
        problem_sheets.append(("Tag 检查", TAG_PROBLEM_SHEET_NAME))
    if run_line_break_check:
        problem_sheets.append(("换行数量检查", LINE_BREAK_PROBLEM_SHEET_NAME))
    if run_number_check:
        problem_sheets.append(("数字一致性", NUMBER_PROBLEM_SHEET_NAME))
    if run_url_check:
        problem_sheets.append(("URL 一致性", URL_PROBLEM_SHEET_NAME))
    if run_chinese_target_check:
        problem_sheets.append(("Target 中文检查", CHINESE_PROBLEM_SHEET_NAME))
    if run_target_text_check:
        problem_sheets.append(("Target 文本规范检查", TARGET_TEXT_PROBLEM_SHEET_NAME))
    generated_sheet_names = []
    if run_term_pair_check:
        generated_sheet_names.append(TERM_SHEET_NAME)
    generated_sheet_names.extend(
        (WORKFLOW_REVIEW_SHEET_NAME, WORKFLOW_SUMMARY_SHEET_NAME)
    )
    write_review_sheet(
        workbook,
        current_sheet_name=worksheet_title,
        input_file=input_file,
        source_column=source_column,
        target_column=target_column,
        start_row=start_row,
        problem_sheets=problem_sheets,
        generated_sheet_names=generated_sheet_names,
        remove_term_helper=False,
    )

    for _, problem_sheet_name in problem_sheets:
        if problem_sheet_name in workbook.sheetnames:
            del workbook[problem_sheet_name]

    if run_term_pair_check:
        data_sheet = workbook[worksheet_title]
        helper_column_index = column_index_from_string(target_column) + 1
        if data_sheet.cell(1, helper_column_index).value == ROW_PROBLEM_COLUMN_HEADER:
            data_sheet.delete_cols(helper_column_index)

    summary_sheet = workbook.create_sheet(WORKFLOW_SUMMARY_SHEET_NAME)
    summary_sheet.append(["检查项", "问题行数"])
    summary_rows = []
    if run_term_pair_check:
        summary_rows.append(("术语检查", term_problem_rows))
    if run_source_consistency_check:
        summary_rows.append(("同 Source 不同 Target", source_consistency_problem_rows))
    if run_target_consistency_check:
        summary_rows.append(("同 Target 不同 Source", target_consistency_problem_rows))
    if run_tag_check:
        summary_rows.append(("Tag 检查", tag_problem_rows))
    if run_line_break_check:
        summary_rows.append(("换行数量检查", line_break_problem_rows))
    if run_number_check:
        summary_rows.append(("数字一致性", number_problem_rows))
    if run_url_check:
        summary_rows.append(("URL 一致性", url_problem_rows))
    if run_chinese_target_check:
        summary_rows.append(("Target 中文检查", chinese_target_problem_rows))
    if run_target_text_check:
        summary_rows.append(("Target 文本规范检查", target_text_problem_rows))
    for summary_row in summary_rows:
        summary_sheet.append(summary_row)

    summary_sheet.column_dimensions["A"].width = 22
    summary_sheet.column_dimensions["B"].width = 12
    workbook.save(output_path)
    workbook.close()
    return term_problem_rows


def run_workflow(
    *,
    input_file: str | Path,
    source_column: str,
    target_column: str,
    output_file: str | Path | None = None,
    sheet: str | None = None,
    start_row: int = 2,
    run_term_pair_check: bool = True,
    term_mark_styles: Iterable[str] | None = None,
    term_history_tb_file: str | Path | None = None,
    term_history_sheet: str | None = None,
    term_history_source_column: str | None = None,
    term_history_target_column: str | None = None,
    term_history_start_row: int = 2,
    run_tag_check: bool = True,
    tag_token_types: tuple[str, ...] | list[str] | None = None,
    tag_angle_config_file: str | Path | None = None,
    run_line_break_check: bool = True,
    run_source_consistency_check: bool = True,
    run_target_consistency_check: bool = False,
    run_number_check: bool = True,
    run_url_check: bool = True,
    run_chinese_target_check: bool = True,
    run_target_text_check: bool = True,
    target_text_rules: Iterable[str] | None = None,
) -> WorkflowSummary:
    if not any(
        (
            run_term_pair_check,
            run_tag_check,
            run_line_break_check,
            run_source_consistency_check,
            run_target_consistency_check,
            run_number_check,
            run_url_check,
            run_chinese_target_check,
            run_target_text_check,
        )
    ):
        raise ValueError("请至少选择一个质量检查项目。")

    input_path = Path(input_file).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    if start_row < 1:
        raise ValueError("开始行必须大于等于 1。")

    normalized_source_column = source_column.strip().upper()
    normalized_target_column = target_column.strip().upper()
    column_index_from_string(normalized_source_column)
    column_index_from_string(normalized_target_column)
    validate_distinct_source_target_columns(
        normalized_source_column,
        normalized_target_column,
    )
    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_output_path(input_path)
    )

    current_input_path = input_path
    worksheet_title = sheet or ""
    term_count = 0
    term_problem_count = 0
    term_problem_rows = 0
    tag_problem_count = 0
    tag_problem_rows = 0
    line_break_problem_count = 0
    source_consistency_problem_count = 0
    source_consistency_problem_rows = 0
    target_consistency_problem_count = 0
    target_consistency_problem_rows = 0
    number_problem_rows = 0
    url_problem_rows = 0
    chinese_target_problem_count = 0
    target_text_problem_count = 0
    target_text_problem_rows = 0

    if run_term_pair_check:
        (
            worksheet_title,
            normalized_source_column,
            normalized_target_column,
            saved_path,
            term_count,
            term_problem_count,
        ) = run_term_pair_check_excel(
            input_file=current_input_path,
            source_column=normalized_source_column,
            target_column=normalized_target_column,
            sheet=sheet,
            start_row=start_row,
            mark_styles=term_mark_styles,
            history_tb_file=term_history_tb_file,
            history_sheet=term_history_sheet,
            history_source_column=term_history_source_column,
            history_target_column=term_history_target_column,
            history_start_row=term_history_start_row,
            output_file=output_path,
            include_row_problem_column=False,
        )
        current_input_path = saved_path

    if run_tag_check:
        tag_summary = run_tag_check_excel(
            input_file=current_input_path,
            source_column=normalized_source_column,
            target_column=normalized_target_column,
            sheet=sheet,
            start_row=start_row,
            token_types=tag_token_types,
            angle_config_file=tag_angle_config_file,
            output_file=output_path,
        )
        worksheet_title = tag_summary.worksheet_title
        tag_problem_count = tag_summary.problem_count
        tag_problem_rows = tag_summary.problem_rows
        current_input_path = tag_summary.output_path

    if run_line_break_check:
        line_break_summary = run_line_break_check_excel(
            input_file=current_input_path,
            source_column=normalized_source_column,
            target_column=normalized_target_column,
            sheet=sheet,
            start_row=start_row,
            output_file=output_path,
        )
        worksheet_title = line_break_summary.worksheet_title
        line_break_problem_count = line_break_summary.problem_rows
        current_input_path = line_break_summary.output_path

    if run_source_consistency_check:
        source_consistency_summary = run_source_consistency_check_excel(
            input_file=current_input_path,
            source_column=normalized_source_column,
            target_column=normalized_target_column,
            sheet=sheet,
            start_row=start_row,
            output_file=output_path,
        )
        worksheet_title = source_consistency_summary.worksheet_title
        source_consistency_problem_count = (
            source_consistency_summary.inconsistent_source_count
        )
        source_consistency_problem_rows = source_consistency_summary.problem_rows
        current_input_path = source_consistency_summary.output_path

    if run_target_consistency_check:
        target_consistency_summary = run_target_consistency_check_excel(
            input_file=current_input_path,
            source_column=normalized_source_column,
            target_column=normalized_target_column,
            sheet=sheet,
            start_row=start_row,
            output_file=output_path,
        )
        worksheet_title = target_consistency_summary.worksheet_title
        target_consistency_problem_count = (
            target_consistency_summary.inconsistent_target_count
        )
        target_consistency_problem_rows = target_consistency_summary.problem_rows
        current_input_path = target_consistency_summary.output_path

    if run_number_check or run_url_check:
        content_fidelity_summary = run_content_fidelity_check_excel(
            input_file=current_input_path,
            source_column=normalized_source_column,
            target_column=normalized_target_column,
            sheet=sheet,
            start_row=start_row,
            rules=tuple(
                rule
                for rule, enabled in (
                    (NUMBER_RULE, run_number_check),
                    (URL_RULE, run_url_check),
                )
                if enabled
            ),
            output_file=output_path,
        )
        worksheet_title = content_fidelity_summary.worksheet_title
        number_problem_rows = content_fidelity_summary.number_problem_rows
        url_problem_rows = content_fidelity_summary.url_problem_rows
        current_input_path = content_fidelity_summary.output_path

    if run_chinese_target_check:
        chinese_target_summary = run_chinese_target_check_excel(
            input_file=current_input_path,
            source_column=normalized_source_column,
            target_column=normalized_target_column,
            sheet=sheet,
            start_row=start_row,
            output_file=output_path,
        )
        worksheet_title = chinese_target_summary.worksheet_title
        chinese_target_problem_count = chinese_target_summary.matched_count
        current_input_path = chinese_target_summary.output_path

    if run_target_text_check:
        target_text_summary = run_target_text_check_excel(
            input_file=current_input_path,
            source_column=normalized_source_column,
            target_column=normalized_target_column,
            sheet=sheet,
            start_row=start_row,
            rules=target_text_rules,
            output_file=output_path,
        )
        worksheet_title = target_text_summary.worksheet_title
        target_text_problem_count = target_text_summary.problem_count
        target_text_problem_rows = target_text_summary.problem_rows

    term_problem_rows = finalize_workflow_output(
        output_path=output_path,
        worksheet_title=worksheet_title,
        run_term_pair_check=run_term_pair_check,
        run_tag_check=run_tag_check,
        run_line_break_check=run_line_break_check,
        run_source_consistency_check=run_source_consistency_check,
        run_target_consistency_check=run_target_consistency_check,
        run_number_check=run_number_check,
        run_url_check=run_url_check,
        run_chinese_target_check=run_chinese_target_check,
        run_target_text_check=run_target_text_check,
        tag_problem_rows=tag_problem_rows,
        line_break_problem_rows=line_break_problem_count,
        source_consistency_problem_rows=source_consistency_problem_rows,
        target_consistency_problem_rows=target_consistency_problem_rows,
        number_problem_rows=number_problem_rows,
        url_problem_rows=url_problem_rows,
        chinese_target_problem_rows=chinese_target_problem_count,
        target_text_problem_rows=target_text_problem_rows,
        input_file=input_path,
        source_column=normalized_source_column,
        target_column=normalized_target_column,
        start_row=start_row,
    )

    return WorkflowSummary(
        output_path=output_path,
        worksheet_title=worksheet_title,
        source_column=normalized_source_column,
        target_column=normalized_target_column,
        start_row=start_row,
        ran_term_pair_check=run_term_pair_check,
        ran_tag_check=run_tag_check,
        ran_line_break_check=run_line_break_check,
        ran_source_consistency_check=run_source_consistency_check,
        ran_target_consistency_check=run_target_consistency_check,
        ran_number_check=run_number_check,
        ran_url_check=run_url_check,
        ran_chinese_target_check=run_chinese_target_check,
        ran_target_text_check=run_target_text_check,
        term_count=term_count,
        term_problem_count=term_problem_count,
        term_problem_rows=term_problem_rows,
        tag_problem_count=tag_problem_count,
        tag_problem_rows=tag_problem_rows,
        line_break_problem_count=line_break_problem_count,
        source_consistency_problem_count=source_consistency_problem_count,
        source_consistency_problem_rows=source_consistency_problem_rows,
        target_consistency_problem_count=target_consistency_problem_count,
        target_consistency_problem_rows=target_consistency_problem_rows,
        number_problem_rows=number_problem_rows,
        url_problem_rows=url_problem_rows,
        chinese_target_problem_count=chinese_target_problem_count,
        target_text_problem_count=target_text_problem_count,
        target_text_problem_rows=target_text_problem_rows,
    )
