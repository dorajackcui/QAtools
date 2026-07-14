"""Apply edits from a workflow report's review sheet to a clean workbook copy."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from tools.excel_output import ROW_PROBLEM_COLUMN_HEADER
from tools.workflow.review_sheet import (
    WORKFLOW_REVIEW_SHEET_NAME,
    WORKFLOW_SCHEMA_VERSION,
    cell_text,
    read_review_metadata,
)


@dataclass(frozen=True)
class RevisionSummary:
    output_path: Path
    worksheet_title: str
    revised_count: int
    ignored_count: int
    unchanged_count: int
    conflict_rows: tuple[int, ...]


def build_default_revised_output_path(report_file: str | Path) -> Path:
    report_path = Path(report_file).expanduser().resolve()
    original_name = report_path.name
    try:
        workbook = load_workbook(report_path, read_only=True)
        if WORKFLOW_REVIEW_SHEET_NAME in workbook.sheetnames:
            metadata = read_review_metadata(workbook[WORKFLOW_REVIEW_SHEET_NAME])
            input_file = metadata.get("input_file")
            if input_file:
                original_name = Path(str(input_file)).name
        workbook.close()
    except Exception:
        pass
    return report_path.with_name(f"revised_{original_name}")


def _required_metadata(review_sheet) -> dict[str, object]:
    metadata = read_review_metadata(review_sheet)
    if str(metadata.get("schema_version", "")) != WORKFLOW_SCHEMA_VERSION:
        raise ValueError("问题处理工作表版本不受支持，请重新运行一键质量检查。")
    for key in ("data_sheet_name", "source_column", "target_column"):
        if not metadata.get(key):
            raise ValueError("问题处理工作表缺少回填信息，请重新运行一键质量检查。")
    return metadata


def _remove_workflow_artifacts(
    workbook,
    data_sheet_name: str,
    target_column: str,
    generated_sheet_names: set[str],
    remove_term_helper: bool,
) -> None:
    for sheet_name in generated_sheet_names:
        if sheet_name in workbook.sheetnames and sheet_name != data_sheet_name:
            del workbook[sheet_name]

    if remove_term_helper:
        data_sheet = workbook[data_sheet_name]
        helper_column_index = column_index_from_string(target_column) + 1
        if data_sheet.cell(1, helper_column_index).value == ROW_PROBLEM_COLUMN_HEADER:
            data_sheet.delete_cols(helper_column_index)


def apply_workflow_revisions(
    report_file: str | Path,
    output_file: str | Path | None = None,
) -> RevisionSummary:
    report_path = Path(report_file).expanduser().resolve()
    if not report_path.exists():
        raise FileNotFoundError(f"检查报告不存在: {report_path}")
    output_path = (
        Path(output_file).expanduser().resolve()
        if output_file
        else build_default_revised_output_path(report_path)
    )
    if output_path == report_path:
        raise ValueError("修订稿输出路径不能与检查报告相同。")

    workbook = load_workbook(report_path)
    if WORKFLOW_REVIEW_SHEET_NAME not in workbook.sheetnames:
        raise ValueError("未找到“问题处理”工作表，请选择 workflow 检查报告。")
    review_sheet = workbook[WORKFLOW_REVIEW_SHEET_NAME]
    metadata = _required_metadata(review_sheet)
    data_sheet_name = str(metadata["data_sheet_name"])
    target_column = str(metadata["target_column"]).strip().upper()
    if data_sheet_name not in workbook.sheetnames:
        raise ValueError(f"未找到原数据工作表: {data_sheet_name}")
    data_sheet = workbook[data_sheet_name]

    revised_count = 0
    ignored_count = 0
    unchanged_count = 0
    conflict_rows: list[int] = []

    for review_row in range(2, review_sheet.max_row + 1):
        raw_row_number = review_sheet.cell(review_row, 1).value
        if raw_row_number is None:
            continue
        try:
            source_row = int(raw_row_number)
        except (TypeError, ValueError):
            continue
        original_target = review_sheet.cell(review_row, 3).value
        revised_target = review_sheet.cell(review_row, 4).value

        if revised_target is None:
            ignored_count += 1
            continue

        target_cell = data_sheet[f"{target_column}{source_row}"]
        if cell_text(target_cell.value) != cell_text(original_target):
            conflict_rows.append(source_row)
            continue

        if cell_text(target_cell.value) == cell_text(revised_target):
            unchanged_count += 1
            continue
        target_cell.value = revised_target
        revised_count += 1

    generated_sheet_names = {
        name
        for name in cell_text(metadata.get("generated_sheet_names")).splitlines()
        if name
    }
    generated_sheet_names.add(WORKFLOW_REVIEW_SHEET_NAME)
    _remove_workflow_artifacts(
        workbook,
        data_sheet_name,
        target_column,
        generated_sheet_names,
        remove_term_helper=cell_text(metadata.get("remove_term_helper")) == "1",
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return RevisionSummary(
        output_path=output_path,
        worksheet_title=data_sheet_name,
        revised_count=revised_count,
        ignored_count=ignored_count,
        unchanged_count=unchanged_count,
        conflict_rows=tuple(dict.fromkeys(conflict_rows)),
    )
