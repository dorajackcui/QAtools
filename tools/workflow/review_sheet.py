"""Build the editable review sheet included in workflow reports."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from tools.excel_output import (
    PROBLEM_BASE_HEADERS,
    join_unique_text,
    write_output_table,
)


WORKFLOW_REVIEW_SHEET_NAME = "问题处理"
WORKFLOW_REVIEW_HEADERS = (
    "行号",
    "source",
    "target",
    "修改后target",
    "问题描述",
    "检查项",
)
METADATA_KEY_COLUMN = 7
METADATA_VALUE_COLUMN = 8
WORKFLOW_SCHEMA_VERSION = "2"


@dataclass
class _ReviewEntry:
    source_text: str = ""
    target_text: str = ""
    check_items: list[str] = field(default_factory=list)
    descriptions: list[str] = field(default_factory=list)


def cell_text(value: object) -> str:
    return "" if value is None else str(value)


def _normalized_row_number(value: object) -> int | None:
    if isinstance(value, bool):
        return None
    try:
        row_number = int(value)
    except (TypeError, ValueError):
        return None
    return row_number if row_number >= 1 else None


def collect_review_rows(
    workbook,
    problem_sheets: Iterable[tuple[str, str]],
) -> list[tuple[object, ...]]:
    """Aggregate all enabled problem sheets into one row per source row."""
    entries_by_row: dict[int, _ReviewEntry] = {}
    for check_item, sheet_name in problem_sheets:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{check_item}缺少问题工作表“{sheet_name}”，无法完整合并。")
        worksheet = workbook[sheet_name]
        headers = tuple(
            cell_text(worksheet.cell(1, column).value).strip()
            for column in range(1, worksheet.max_column + 1)
        )
        if headers[:4] != PROBLEM_BASE_HEADERS:
            raise ValueError(
                f"{check_item}问题表格式不兼容：前四列必须为 "
                + " / ".join(PROBLEM_BASE_HEADERS)
            )
        for row_values in worksheet.iter_rows(
            min_row=2,
            min_col=1,
            max_col=max(4, worksheet.max_column),
            values_only=True,
        ):
            row_number = _normalized_row_number(row_values[0])
            if row_number is None:
                continue
            source_text = cell_text(row_values[1])
            target_text = cell_text(row_values[2])
            description = cell_text(row_values[3]).strip()
            entry = entries_by_row.setdefault(row_number, _ReviewEntry())
            if entry.source_text and source_text and entry.source_text != source_text:
                raise ValueError(f"第 {row_number} 行的 source 原文在检查结果中不一致。")
            if entry.target_text and target_text and entry.target_text != target_text:
                raise ValueError(f"第 {row_number} 行的 target 原文在检查结果中不一致。")
            if not entry.source_text and source_text:
                entry.source_text = source_text
            if not entry.target_text and target_text:
                entry.target_text = target_text
            if check_item not in entry.check_items:
                entry.check_items.append(check_item)
            detail_parts = []
            for column_index in range(5, len(row_values) + 1):
                detail_header = headers[column_index - 1]
                detail_value = cell_text(row_values[column_index - 1]).strip()
                if detail_header and detail_value:
                    detail_parts.append(f"{detail_header}：{detail_value}")
            detail_text = "；".join(detail_parts)
            combined_description = description
            if detail_text:
                combined_description = (
                    f"{description}（{detail_text}）" if description else detail_text
                )
            formatted_description = (
                f"【{check_item}】{combined_description}"
                if combined_description
                else f"【{check_item}】"
            )
            if formatted_description not in entry.descriptions:
                entry.descriptions.append(formatted_description)

    return [
        (
            row_number,
            entry.source_text,
            entry.target_text,
            None,
            join_unique_text(entry.descriptions),
            join_unique_text(entry.check_items),
        )
        for row_number, entry in sorted(entries_by_row.items())
    ]


def write_review_metadata(
    worksheet,
    *,
    input_file: str | Path,
    data_sheet_name: str,
    source_column: str,
    target_column: str,
    start_row: int,
    generated_sheet_names: Iterable[str],
    remove_term_helper: bool,
) -> None:
    metadata = (
        ("schema_version", WORKFLOW_SCHEMA_VERSION),
        ("input_file", str(Path(input_file).expanduser().resolve())),
        ("data_sheet_name", data_sheet_name),
        ("source_column", source_column),
        ("target_column", target_column),
        ("start_row", start_row),
        ("generated_sheet_names", "\n".join(generated_sheet_names)),
        ("remove_term_helper", "1" if remove_term_helper else "0"),
    )
    for row_index, (key, value) in enumerate(metadata, start=1):
        worksheet.cell(row_index, METADATA_KEY_COLUMN, key)
        worksheet.cell(row_index, METADATA_VALUE_COLUMN, value)
    worksheet.column_dimensions["G"].hidden = True
    worksheet.column_dimensions["H"].hidden = True


def read_review_metadata(worksheet) -> dict[str, object]:
    metadata: dict[str, object] = {}
    for row_index in range(1, 20):
        key = worksheet.cell(row_index, METADATA_KEY_COLUMN).value
        if key is None:
            continue
        metadata[str(key)] = worksheet.cell(row_index, METADATA_VALUE_COLUMN).value
    return metadata


def write_review_sheet(
    workbook,
    *,
    current_sheet_name: str,
    input_file: str | Path,
    source_column: str,
    target_column: str,
    start_row: int,
    problem_sheets: Iterable[tuple[str, str]],
    generated_sheet_names: Iterable[str],
    remove_term_helper: bool,
) -> int:
    rows = collect_review_rows(workbook, problem_sheets)
    worksheet = write_output_table(
        workbook,
        current_sheet_name=current_sheet_name,
        sheet_name=WORKFLOW_REVIEW_SHEET_NAME,
        headers=WORKFLOW_REVIEW_HEADERS,
        rows=rows,
        row_link_target_column=target_column,
    )
    custom_widths = {
        "A": 10,
        "B": 52,
        "C": 52,
        "D": 52,
        "E": 64,
        "F": 28,
    }
    for column, width in custom_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet["D1"].comment = Comment(
        "在这里填写最终 target。填写后会回填；留空表示忽略该行。",
        "QAtools",
    )

    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for worksheet_row in range(2, len(rows) + 2):
        worksheet.cell(worksheet_row, 4).fill = editable_fill

    write_review_metadata(
        worksheet,
        input_file=input_file,
        data_sheet_name=current_sheet_name,
        source_column=source_column,
        target_column=target_column,
        start_row=start_row,
        generated_sheet_names=generated_sheet_names,
        remove_term_helper=remove_term_helper,
    )
    return len(rows)
